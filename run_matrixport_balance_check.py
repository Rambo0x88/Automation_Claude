#!/usr/bin/env python3
"""
Matrixport Balance vs DAM Portfolio Comparison

Steps:
  1. Call Matrixport Balance API  → available_balance per currency
  2. Call Matrixport Balance+ API → flexi-saving balances
  3. Login to DAM and fetch CEX balances detail for a given portfolio
  4. Compare API totals against DAM portfolio amounts
  5. Export a 4-sheet Excel report

Configure the variables in the CONFIG section, then run:
    python3 run_matrixport_balance_check.py
"""

import hmac
import hashlib
import time
import json
import smtplib
import os
import requests
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────

MATRIXPORT_ACCESS_KEY = "ak-9c60c430-a9d5-483f-8d90-32494ab20022"
MATRIXPORT_SECRET_KEY = "qGtXrDOMsHHVca5dg4Rimcy0gqjz0mBCW3cs3sV1iRzYAmv6GQ1Pjdw1mgEu0vEU"

DAM_BASE_URL   = "https://dam-sit.mqbc21.com"
DAM_EMAIL      = "roninx688@gmail.com"
DAM_PASSWORD   = "0987654321a@A"
DAM_CAPTCHA    = "1x0000000000000000000000000000000AA"  # Cloudflare test-mode bypass
PORTFOLIO_ID   = "6cc8ad6a-8249-4a6e-be62-d7c968fe8dac"

MATRIXPORT_BASE = "https://mapi.matrixport.com"
OUTPUT_FILE     = "matrixport_balance.xlsx"

# ── EMAIL CONFIG ──────────────────────────────────────────────────────────────
# Generate a Gmail App Password at: https://myaccount.google.com/apppasswords
GMAIL_SENDER       = "merqbcqa2@gmail.com"
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")  # set env var or paste here
EMAIL_RECIPIENT    = "roninx688@gmail.com"

# ── MATRIXPORT AUTH ───────────────────────────────────────────────────────────

def matrixport_headers(path: str) -> dict:
    """Generate signed headers for Matrixport API v2."""
    ts = str(int(time.time() * 1000))
    msg = ts + "GET" + path + "&"
    sig = hmac.new(MATRIXPORT_SECRET_KEY.encode(), msg.encode(), hashlib.sha256).hexdigest()
    return {
        "X-MatrixPort-Access-Key": MATRIXPORT_ACCESS_KEY,
        "X-Signature":             sig,
        "X-Timestamp":             ts,
        "X-Auth-Version":          "v2",
        "Content-Type":            "application/json",
    }

# ── STEP 1: Balance API ───────────────────────────────────────────────────────

def fetch_balance_api() -> list[dict]:
    """Return non-zero wallet balances from Matrixport Balance API."""
    path = "/mapi/v1/wallet/balance"
    resp = requests.get(MATRIXPORT_BASE + path, headers=matrixport_headers(path), timeout=15)
    resp.raise_for_status()
    items = resp.json()["data"]["items"]
    non_zero = [i for i in items if float(i["available_balance"]) > 0]
    print(f"[Step 1] Balance API → {len(non_zero)} non-zero currencies")
    for i in non_zero:
        print(f"         {i['currency']:10s}  {i['available_balance']}")
    return non_zero

# ── STEP 2: Balance+ API ──────────────────────────────────────────────────────

def fetch_balance_plus_api() -> list[dict]:
    """Return flexi-saving balances from Matrixport Balance+ API."""
    path = "/flexible/api/v2/user/asset/summary"
    resp = requests.get(MATRIXPORT_BASE + path, headers=matrixport_headers(path), timeout=15)
    resp.raise_for_status()
    currencies = resp.json()["data"].get("currencies", [])
    print(f"[Step 2] Balance+ API → {len(currencies)} currencies")
    for c in currencies:
        print(f"         {c['currency']:10s}  {c['balance']}  ({c['product_type']})")
    return currencies

# ── DAM AUTH ──────────────────────────────────────────────────────────────────

def dam_login() -> str:
    """Login to DAM and return the access token."""
    url  = f"{DAM_BASE_URL}/api/v1/user/signin"
    body = {"email": DAM_EMAIL, "password": DAM_PASSWORD, "captcha_token": DAM_CAPTCHA}
    resp = requests.post(url, json=body, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") != 201:
        raise RuntimeError(f"DAM login failed: {data.get('message')}")
    token = data["data"]["access_token"]
    print(f"[DAM]    Login successful (user_id: {data['data']['user_id']})")
    return token

# ── STEP 3: DAM Portfolio CEX Balances ───────────────────────────────────────

def fetch_dam_portfolio(token: str) -> list[dict]:
    """Return CEX token rows for the configured portfolio."""
    url  = f"{DAM_BASE_URL}/api/v1/portfolio/{PORTFOLIO_ID}/cex/balances/details"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=15)
    resp.raise_for_status()
    accounts = resp.json().get("data") or []
    rows = []
    for account in accounts:
        label = account["account_label"]
        for token_data in account["tokens"]:
            rows.append({
                "account":      label,
                "token":        token_data["symbol"],
                "price":        token_data["price"],
                "price_24h":    f"{float(token_data['price_change_percent_24h']):+.3f}%",
                "total_amount": token_data["total_amount"],
                "value":        token_data["total_value"],
            })
    print(f"[Step 3] DAM portfolio → {len(rows)} token rows")
    for r in rows:
        print(f"         [{r['account']}] {r['token']:10s}  amount={r['total_amount']}  value={r['value']}")
    return rows

# ── STEP 4: Comparison ────────────────────────────────────────────────────────

def build_comparison(balance_api: list, balance_plus: list, dam_rows: list) -> list[dict]:
    """Match API balances against DAM totals per token."""
    balance_map      = {r["currency"]: float(r["available_balance"]) for r in balance_api}
    balance_plus_map = {r["currency"]: float(r["balance"])           for r in balance_plus}

    results = []
    for row in dam_rows:
        token     = row["token"]
        api_bal   = balance_map.get(token, 0.0)
        plus_bal  = balance_plus_map.get(token, 0.0)
        total_api = api_bal + plus_bal
        dam_amt   = float(row["total_amount"])
        diff      = total_api - dam_amt
        match     = "MATCH" if abs(diff) < 0.01 else "MISMATCH"

        results.append({
            "Account":           row["account"],
            "Token":             token,
            "Balance API":       round(api_bal, 8),
            "Balance+ API":      round(plus_bal, 8),
            "Total API":         round(total_api, 8),
            "DAM Total Amount":  round(dam_amt, 8),
            "Difference":        round(diff, 8),
            "Status":            match,
            "DAM Price":         row["price"],
            "DAM Price (24H)":   row["price_24h"],
            "DAM Value (USD)":   round(float(row["value"]), 6),
        })
    return results

# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────

BLUE  = "1F4E79"
GREEN = "1E6B3C"
MATCH_FILL    = "C6EFCE"
MISMATCH_FILL = "FFC7CE"
THIN = Side(style="thin")

def _header(ws, row: int, ncols: int, color: str = BLUE):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _auto_width(ws):
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 36)

def _border_row(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = Border(
            left=THIN, right=THIN, top=THIN, bottom=THIN
        )

def export_excel(balance_api, balance_plus, dam_rows, comparison, path: str, overall: str = "", reason: str = ""):
    wb = Workbook()

    # Sheet 1 — Balance API
    ws1 = wb.active
    ws1.title = "1. Balance API"
    h1 = ["Currency", "Available Balance", "Balance", "Frozen Balance"]
    ws1.append(h1)
    _header(ws1, 1, len(h1))
    for r in balance_api:
        ws1.append([r["currency"], float(r["available_balance"]),
                    float(r["balance"]), float(r["frozen_balance"])])
    _auto_width(ws1)

    # Sheet 2 — Balance+ API
    ws2 = wb.create_sheet("2. Balance+ API")
    h2 = ["Currency", "Product Type", "Balance", "Interest", "Balance (USD)", "Interest (USD)"]
    ws2.append(h2)
    _header(ws2, 1, len(h2), color=GREEN)
    for r in balance_plus:
        ws2.append([r["currency"], r["product_type"],
                    float(r["balance"]), float(r.get("interest", 0)),
                    float(r.get("balance_usd", 0)), float(r.get("interest_usd", 0))])
    _auto_width(ws2)

    # Sheet 3 — DAM Portfolio
    ws3 = wb.create_sheet("3. DAM Portfolio")
    h3 = ["Account", "Token", "Price", "Price (24H)", "Total Amount", "Value (USD)"]
    ws3.append(h3)
    _header(ws3, 1, len(h3), color="7B2D8B")
    for r in dam_rows:
        ws3.append([r["account"], r["token"], float(r["price"]),
                    r["price_24h"], float(r["total_amount"]), float(r["value"])])
    _auto_width(ws3)

    # Sheet 4 — Comparison
    ws4 = wb.create_sheet("4. Comparison Report")

    # Verdict row
    verdict_color = "375623" if overall == "PASSED" else "9C0006"
    ws4.append([f"RESULT: {overall}", reason, f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    for col in range(1, 4):
        c = ws4.cell(row=1, column=col)
        c.font      = Font(bold=True, color="FFFFFF", size=12)
        c.fill      = PatternFill("solid", fgColor=verdict_color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws4.row_dimensions[1].height = 20

    h4  = list(comparison[0].keys()) if comparison else []
    ws4.append(h4)
    _header(ws4, 2, len(h4))
    for row_data in comparison:
        ws4.append(list(row_data.values()))
        row_idx  = ws4.max_row
        fill_hex = MATCH_FILL if row_data["Status"] == "MATCH" else MISMATCH_FILL
        for col in range(1, len(h4) + 1):
            c = ws4.cell(row=row_idx, column=col)
            c.fill   = PatternFill("solid", fgColor=fill_hex)
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    _auto_width(ws4)

    wb.save(path)
    print(f"\n[Export] Saved → {path}")

# ── EMAIL ─────────────────────────────────────────────────────────────────────

def send_email(overall: str, reason: str, attachment_path: str):
    if not GMAIL_APP_PASSWORD:
        print("[Email] Skipped — GMAIL_APP_PASSWORD not set.")
        return

    subject = f"Matrixport Balance Report – {datetime.now().strftime('%Y-%m-%d')}  [{overall}]"
    body    = (
        f"Matrixport Balance Check\n"
        f"Run time : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"Result   : {overall}\n"
        f"Detail   : {reason}\n\n"
        f"See the attached Excel report for the full breakdown."
    )

    msg = MIMEMultipart()
    msg["From"]    = GMAIL_SENDER
    msg["To"]      = EMAIL_RECIPIENT
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(attachment_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(attachment_path)}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_SENDER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_SENDER, EMAIL_RECIPIENT, msg.as_string())
    print(f"[Email] Report sent to {EMAIL_RECIPIENT}")

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print(f"Matrixport Balance Check  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    balance_api  = fetch_balance_api()
    balance_plus = fetch_balance_plus_api()

    token    = dam_login()
    dam_rows = fetch_dam_portfolio(token)

    comparison = build_comparison(balance_api, balance_plus, dam_rows)

    print("\n" + "=" * 60)
    print("COMPARISON SUMMARY")
    print("=" * 60)
    print(f"{'Token':<8} {'API':>14} {'Balance+':>12} {'Total API':>14} {'DAM Amount':>14} {'Diff':>12}  Status")
    print("-" * 80)
    for r in comparison:
        icon = "✅" if r["Status"] == "MATCH" else "❌"
        print(f"{r['Token']:<8} {r['Balance API']:>14.8f} {r['Balance+ API']:>12.8f} "
              f"{r['Total API']:>14.8f} {r['DAM Total Amount']:>14.8f} "
              f"{r['Difference']:>12.8f}  {icon} {r['Status']}")

    if not comparison:
        overall = "FAILED"
        reason  = "No DAM portfolio data to compare"
    elif all(r["Status"] == "MATCH" for r in comparison):
        overall = "PASSED"
        reason  = f"All {len(comparison)} token(s) matched"
    else:
        mismatches = [r["Token"] for r in comparison if r["Status"] != "MATCH"]
        overall = "FAILED"
        reason  = f"{len(mismatches)} mismatch(es): {', '.join(mismatches)}"

    print("\n" + "=" * 60)
    verdict_icon = "✅" if overall == "PASSED" else "❌"
    print(f"OVERALL RESULT:  {verdict_icon} {overall}  —  {reason}")
    print("=" * 60)

    export_excel(balance_api, balance_plus, dam_rows, comparison, OUTPUT_FILE, overall, reason)
    send_email(overall, reason, OUTPUT_FILE)

if __name__ == "__main__":
    main()
