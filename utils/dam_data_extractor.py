#!/usr/bin/env python3
"""
DAM Data Extractor - Extract portfolio data from Overview and Holdings pages
"""
import pandas as pd
from datetime import datetime
from playwright.sync_api import Page
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class DAMDataExtractor:
    """Extract data from DAM portfolio pages and export to Excel"""

    def __init__(self, page: Page):
        self.page = page

    def extract_price_24h_change(self, cell):
        """
        Extract 24h price change from a cell containing the rounded pill element

        Args:
            cell: The Playwright cell element

        Returns:
            String with 24h price change (e.g., "↑ 0.79%") or empty string if not found
        """
        try:
            # Look for the price change pill element with specific classes
            price_change_selectors = [
                '.typography-body.font-normal.rounded-full',
                '[class*="rounded-full"][class*="typography-body"]',
                '.rounded-full'
            ]

            for selector in price_change_selectors:
                price_change_elem = cell.locator(selector).first
                if price_change_elem.count() > 0:
                    try:
                        if price_change_elem.is_visible(timeout=100):
                            # Get the text content
                            price_text = price_change_elem.inner_text().strip()

                            # Get the class attribute to determine if it's positive or negative
                            class_attr = price_change_elem.get_attribute('class')

                            # Determine arrow based on color/class
                            # Green = positive (up), Red = negative (down)
                            arrow = ""
                            if class_attr:
                                # Check for success/green classes (positive change)
                                if 'text-success' in class_attr or 'text-green' in class_attr or 'bg-success' in class_attr or 'bg-green' in class_attr:
                                    arrow = "↑ "
                                # Check for error/red classes (negative change)
                                elif 'text-error' in class_attr or 'text-red' in class_attr or 'bg-error' in class_attr or 'bg-red' in class_attr:
                                    arrow = "↓ "

                            # If no arrow detected from class, check if text starts with + or -
                            if not arrow:
                                if price_text.startswith('+') or price_text.startswith('↑'):
                                    arrow = "↑ "
                                    # Remove the + if present
                                    price_text = price_text.lstrip('+').lstrip('↑').strip()
                                elif price_text.startswith('-') or price_text.startswith('↓'):
                                    arrow = "↓ "
                                    # Remove the - if present
                                    price_text = price_text.lstrip('-').lstrip('↓').strip()

                            return arrow + price_text
                    except:
                        continue
            return ""
        except:
            return ""

    def extract_table_headers(self) -> list:
        """
        Extract table headers from current page

        Returns:
            List of header names
        """
        try:
            # Look for table headers
            headers = []

            # Try thead > tr > th
            header_cells = self.page.locator("thead th, thead [role='columnheader']").all()

            if len(header_cells) == 0:
                # Try alternative: first row with th elements
                header_cells = self.page.locator("tr th, [role='row'] [role='columnheader']").all()

            for cell in header_cells:
                header_text = cell.text_content().strip()
                if header_text:
                    headers.append(header_text)

            if len(headers) > 0:
                print(f"   📋 Found {len(headers)} column headers: {headers}")

            return headers

        except Exception as e:
            print(f"   ⚠️  Could not extract headers: {e}")
            return []

    def extract_overview_data(self) -> list:
        """
        Extract token breakdown table data from Overview page

        Returns:
            List of dictionaries with token breakdown data
        """
        print("📊 Extracting Overview page - Token Breakdown table data...")

        overview_data = []

        try:
            # Wait for page to load
            self.page.wait_for_timeout(2000)

            # === SCROLL TO LOAD ALL LAZY-LOADED TOKENS ===
            print(f"   🔄 Scrolling to load all tokens (handling lazy-loading)...")
            viewport_height = self.page.viewport_size['height']
            total_height = self.page.evaluate("document.body.scrollHeight")
            current_scroll = 0
            last_row_count = 0
            stable_count = 0

            while current_scroll < total_height and stable_count < 3:
                current_scroll += viewport_height // 2
                self.page.evaluate(f"window.scrollTo(0, {current_scroll})")
                self.page.wait_for_timeout(400)

                current_rows = self.page.locator("tbody tr").count()
                if current_rows == last_row_count:
                    stable_count += 1
                else:
                    stable_count = 0
                    last_row_count = current_rows

                total_height = self.page.evaluate("document.body.scrollHeight")

            self.page.evaluate("window.scrollTo(0, 0)")
            self.page.wait_for_timeout(500)
            print(f"   ✅ Scrolled to load all tokens")
            # === END SCROLL LOGIC ===

            # Look for table element on Overview page
            # Try multiple selectors to find the token breakdown table
            table_selectors = [
                "table",
                "[role='table']",
                "tbody",
                "[data-testid*='table']"
            ]

            table_found = False
            table_rows = []

            for selector in table_selectors:
                try:
                    print(f"   🔍 Trying table selector: {selector}")
                    table_element = self.page.locator(selector).first

                    if table_element.is_visible(timeout=2000):
                        # Found table, now get rows
                        table_rows = self.page.locator("tbody tr, [role='row']").all()

                        # Filter out header rows
                        table_rows = [row for row in table_rows if row.locator("td, [role='cell']").count() > 0]

                        if len(table_rows) > 0:
                            print(f"   ✅ Found table with {len(table_rows)} data rows")
                            table_found = True
                            break
                except Exception as e:
                    print(f"   ⚠️  Selector failed: {e}")
                    continue

            if not table_found:
                print("   ⚠️  No table found, trying alternative extraction...")
                # Try to extract any visible data
                all_rows = self.page.locator("tr").all()
                table_rows = [row for row in all_rows if row.locator("td").count() > 0]
                print(f"   Found {len(table_rows)} rows with td elements")

            # Extract table headers
            headers = self.extract_table_headers()

            # If no headers found, use correct column names for Overview Token Breakdown table
            # Based on actual HTML structure: Chain, Name, Price, Price(24h), Share, Amount, Value
            if len(headers) == 0:
                headers = ["Chain", "Name", "Price", "Price_24h", "Share", "Amount", "Value"]
                print(f"   Using default headers: {headers}")

            # Extract data from each row
            for idx, row in enumerate(table_rows, 1):
                try:
                    # Get all cells in the row
                    cells = row.locator("td, [role='cell']").all()

                    if len(cells) == 0:
                        continue

                    # Extract text from each cell - get only the primary value (first line/element)
                    # Also extract 24h price change and tooltip amounts separately
                    cell_values = []
                    price_24h_values = []
                    tooltip_amounts = []

                    for cell_idx, cell in enumerate(cells):
                        try:
                            # Get inner text and split by newlines to get first value only
                            text = cell.inner_text().strip()
                            # Split by newline and take first non-empty line (the main value)
                            lines = [line.strip() for line in text.split('\n') if line.strip()]
                            text = lines[0] if lines else text
                            cell_values.append(text)

                            # Extract 24h price change from this cell (if it exists)
                            price_24h = self.extract_price_24h_change(cell)
                            price_24h_values.append(price_24h)

                            # Extract tooltip amount if it exists
                            # Tooltips show the full unformatted amount (e.g., "0.00332" instead of "0.003...")
                            tooltip_amount = ""
                            try:
                                # Check if this cell contains elements with amount tooltip
                                tooltip_elems = cell.locator('[data-tooltip-id*="amount"]').all()

                                if len(tooltip_elems) > 0:
                                    for tooltip_elem in tooltip_elems:
                                        try:
                                            # Get the tooltip ID to find the corresponding tooltip content
                                            tooltip_id = tooltip_elem.get_attribute("data-tooltip-id")

                                            if tooltip_id and "amount" in tooltip_id.lower():
                                                # Hover over the element to trigger tooltip
                                                tooltip_elem.hover(timeout=500)
                                                self.page.wait_for_timeout(200)

                                                # Find the actual tooltip popup by ID
                                                # The tooltip appears as a separate div with matching id or data attribute
                                                tooltip_selector = f'[id="{tooltip_id}"], [data-tooltip-id="{tooltip_id}"][role="tooltip"]'
                                                tooltip_popup = self.page.locator(tooltip_selector).first

                                                if tooltip_popup.is_visible(timeout=300):
                                                    # Extract the text from tooltip
                                                    tooltip_text = tooltip_popup.inner_text().strip()
                                                    if tooltip_text and any(c.isdigit() for c in tooltip_text):
                                                        tooltip_amount = tooltip_text
                                                        # Move mouse away to hide tooltip
                                                        self.page.mouse.move(0, 0)
                                                        self.page.wait_for_timeout(100)
                                                        break

                                                # Alternative: The tooltip might be in a global container
                                                # Try finding any visible tooltip with role="tooltip"
                                                if not tooltip_amount:
                                                    visible_tooltips = self.page.locator('[role="tooltip"]:visible').all()
                                                    for tt in visible_tooltips:
                                                        try:
                                                            tt_text = tt.inner_text().strip()
                                                            # Check if this tooltip contains numeric data
                                                            if tt_text and any(c.isdigit() for c in tt_text):
                                                                tooltip_amount = tt_text
                                                                break
                                                        except:
                                                            continue

                                                # Move mouse away to hide tooltip
                                                self.page.mouse.move(0, 0)
                                                self.page.wait_for_timeout(50)

                                                if tooltip_amount:
                                                    break

                                        except Exception as te:
                                            continue

                            except Exception as e:
                                pass

                            tooltip_amounts.append(tooltip_amount)

                        except:
                            # Fallback to text_content if inner_text fails
                            text = cell.text_content().strip()
                            lines = [line.strip() for line in text.split('\n') if line.strip()]
                            text = lines[0] if lines else text
                            cell_values.append(text)
                            price_24h_values.append("")
                            tooltip_amounts.append("")

                    # Create row data in correct column order: Row_Number, Chain, Name, Price, Price_24h, Share, Amount, Amount_Tooltip, Value
                    # Column structure in HTML: 0=Chain, 1=Name, 2=Price, 3=Price(24h) pill, 4=Share, 5=Amount, 6=Value
                    row_data = {}
                    row_data["Row_Number"] = idx

                    # Map columns in order
                    if len(cell_values) >= 1:
                        row_data["Chain"] = cell_values[0]
                    if len(cell_values) >= 2:
                        row_data["Name"] = cell_values[1]
                    if len(cell_values) >= 3:
                        row_data["Price"] = cell_values[2]

                    # Add Price_24h from the rounded pill element (from column 3 or 2)
                    price_24h = ""
                    if len(price_24h_values) > 3 and price_24h_values[3]:
                        price_24h = price_24h_values[3]
                    elif len(price_24h_values) > 2 and price_24h_values[2]:
                        price_24h = price_24h_values[2]
                    row_data["Price_24h"] = price_24h

                    if len(cell_values) >= 4:
                        row_data["Share"] = cell_values[3]
                    if len(cell_values) >= 5:
                        row_data["Amount"] = cell_values[4]

                    # Add Amount_Tooltip from tooltip (index 4 or 5 = Amount column)
                    # Debug: Check all tooltip_amounts to find which index has the data
                    amount_tooltip_value = ""
                    if len(tooltip_amounts) > 0:
                        # Try to find the first non-empty tooltip amount
                        for i, tt in enumerate(tooltip_amounts):
                            if tt and str(tt).strip():
                                # Prefer tooltip from Amount column (index 4 or 5)
                                if i in [4, 5]:
                                    amount_tooltip_value = tt
                                    break
                        # If no tooltip found in expected columns, use the first non-empty one
                        if not amount_tooltip_value:
                            for tt in tooltip_amounts:
                                if tt and str(tt).strip():
                                    amount_tooltip_value = tt
                                    break

                    row_data["Amount_Tooltip"] = amount_tooltip_value

                    if len(cell_values) >= 6:
                        row_data["Value"] = cell_values[5]

                    # Add any extra columns
                    for i in range(6, len(cell_values)):
                        row_data[f"Column_{i+1}"] = cell_values[i]

                    overview_data.append(row_data)

                except Exception as e:
                    print(f"   ⚠️  Error extracting row {idx}: {e}")
                    continue

            print(f"✅ Extracted {len(overview_data)} rows from Overview page token breakdown table")

        except Exception as e:
            print(f"❌ Error extracting Overview data: {e}")
            import traceback
            traceback.print_exc()

        return overview_data

    def extract_holdings_token_table(self, portfolio_name: str) -> list:
        """
        Extract holdings data from Holdings page - Token view

        In Token view, there are two types of rows:
        1. Main token rows with: Token, Portfolio %, Amount, Value
        2. Expandable chain detail rows per token with: Chain, Amount, Price, Value

        Args:
            portfolio_name: Name of the portfolio

        Returns:
            List of dictionaries with both main token rows and expandable chain detail rows
        """
        print(f"📊 Extracting Holdings - Token table for '{portfolio_name}'...")

        holdings_data = []

        try:
            # Wait for page to load
            self.page.wait_for_timeout(2000)

            # Wait for table to be attached (not necessarily visible due to scrolling)
            self.page.wait_for_selector("tbody tr", state="attached", timeout=5000)
            self.page.wait_for_timeout(1000)  # Additional wait for dynamic content

            # === SCROLL TO LOAD ALL LAZY-LOADED TOKENS ===
            # The table uses virtualization/lazy loading, so we must scroll to load all rows
            print(f"   🔄 Scrolling to load all tokens (handling lazy-loading)...")

            # Try to find the scrollable table container
            # Look for tbody, table container, or the page itself
            scroll_container = None
            scroll_selectors = [
                "div[class*='overflow-y-auto']",
                "div[class*='overflow-auto']",
                "div[class*='scroll']",
                "tbody",
                "table"
            ]

            for selector in scroll_selectors:
                container = self.page.locator(selector).first
                if container.count() > 0:
                    try:
                        scroll_info = container.evaluate("(el) => ({ scrollHeight: el.scrollHeight, clientHeight: el.clientHeight })")
                        if scroll_info['scrollHeight'] > scroll_info['clientHeight']:
                            scroll_container = container
                            print(f"   📍 Found scrollable container: {selector}")
                            break
                    except:
                        continue

            # Scroll within the container or use page scroll
            if scroll_container:
                # Scroll within the table container
                max_scroll = scroll_container.evaluate("(el) => el.scrollHeight - el.clientHeight")
                current_scroll = 0
                scroll_step = 500  # Pixels per scroll
                last_row_count = 0
                stable_count = 0

                while current_scroll < max_scroll and stable_count < 3:
                    current_scroll += scroll_step
                    if current_scroll > max_scroll:
                        current_scroll = max_scroll
                    scroll_container.evaluate(f"(el) => el.scrollTop = {current_scroll}")
                    self.page.wait_for_timeout(300)  # Wait for lazy load

                    # Check if new rows were loaded
                    current_rows = self.page.locator("tbody tr").count()
                    if current_rows == last_row_count:
                        stable_count += 1
                    else:
                        stable_count = 0
                        last_row_count = current_rows

                    # Update max_scroll in case content grew
                    max_scroll = scroll_container.evaluate("(el) => el.scrollHeight - el.clientHeight")

                # Scroll back to top
                scroll_container.evaluate("(el) => el.scrollTop = 0")
                self.page.wait_for_timeout(500)
                print(f"   ✅ Scrolled container to load all tokens")
            else:
                # Fallback: Use page-level scrolling
                viewport_height = self.page.viewport_size['height']
                total_height = self.page.evaluate("document.body.scrollHeight")
                current_scroll = 0
                last_row_count = 0
                stable_count = 0

                while current_scroll < total_height and stable_count < 3:
                    current_scroll += viewport_height // 2  # Scroll half viewport at a time
                    self.page.evaluate(f"window.scrollTo(0, {current_scroll})")
                    self.page.wait_for_timeout(500)  # Wait for lazy load

                    # Check if new rows were loaded
                    current_rows = self.page.locator("tbody tr").count()
                    if current_rows == last_row_count:
                        stable_count += 1
                    else:
                        stable_count = 0
                        last_row_count = current_rows

                    # Update total_height in case content grew
                    total_height = self.page.evaluate("document.body.scrollHeight")

                # Scroll back to top
                self.page.evaluate("window.scrollTo(0, 0)")
                self.page.wait_for_timeout(500)
                print(f"   ✅ Scrolled page to load all tokens")

            # === END SCROLL LOGIC ===

            # Look for ALL table rows including nested/expanded ones
            # Use a more comprehensive selector to catch expanded chain rows
            table_rows = self.page.locator("tbody tr[data-slot], tbody tr[role='row'], tbody tr").all()

            # Filter out rows without td elements
            table_rows = [row for row in table_rows if row.locator("td, [role='cell']").count() > 0]

            print(f"   Found {len(table_rows)} total rows (main + expanded)")

            # Debug: Print first 5 row types to understand structure
            print(f"   DEBUG: Checking first 5 rows...")
            for i in range(min(5, len(table_rows))):
                row = table_rows[i]
                cells = row.locator("td, [role='cell']").all()
                if len(cells) > 0:
                    first_cell = cells[0].inner_text().strip().split('\n')[0]
                    has_btn = row.locator("td:first-child button, td:first-child svg").count() > 0
                    print(f"   DEBUG Row {i+1}: cells={len(cells)}, has_button={has_btn}, first_cell='{first_cell[:20]}'...")

            # Based on screenshot: Extract only Chain_Detail rows (expanded chain details)
            # Each row should have: Portfolio, Row_Type, Parent_Token, Chain, Share, Amount, Price, Price_24h, Value
            # Token Main rows structure: [button] | Token | Portfolio% | Amount | Value
            # Chain Detail rows structure: Chain | Amount | Price | Value
            chain_detail_headers = ["Chain", "Amount", "Price", "Value"]

            current_token = None  # Track the current parent token
            current_share = None  # Track the current parent token's portfolio share

            for idx, row in enumerate(table_rows, 1):
                try:
                    # Extract data from each row
                    cells = row.locator("td, [role='cell']").all()

                    if len(cells) == 0:
                        continue

                    # Extract text from each cell - get only the primary value (first line/element)
                    # Also extract 24h price change and tooltip amounts separately
                    cell_values = []
                    price_24h_values = []
                    tooltip_amounts = []

                    for cell_idx, cell in enumerate(cells):
                        try:
                            # Get inner text and split by newlines to get first value only
                            text = cell.inner_text().strip()
                            # Split by newline and take first non-empty line (the main value)
                            lines = [line.strip() for line in text.split('\n') if line.strip()]
                            text = lines[0] if lines else text
                            cell_values.append(text)

                            # Extract 24h price change from this cell (if it exists)
                            price_24h = self.extract_price_24h_change(cell)
                            price_24h_values.append(price_24h)

                            # Extract tooltip amount if it exists (same logic as Overview)
                            tooltip_amount = ""
                            try:
                                # Only extract tooltip for Amount column
                                tooltip_elems = cell.locator('[data-tooltip-id*="amount"]').all()

                                for tooltip_elem in tooltip_elems:
                                    try:
                                        tooltip_id = tooltip_elem.get_attribute("data-tooltip-id")

                                        if tooltip_id and "amount" in tooltip_id.lower():
                                            # Hover over the element to trigger tooltip
                                            tooltip_elem.hover(timeout=300)
                                            self.page.wait_for_timeout(50)

                                            # Find the actual tooltip popup
                                            tooltip_selector = f'[id="{tooltip_id}"], [data-tooltip-id="{tooltip_id}"][role="tooltip"]'
                                            tooltip_popup = self.page.locator(tooltip_selector).first

                                            if tooltip_popup.is_visible(timeout=100):
                                                tooltip_text = tooltip_popup.inner_text().strip()
                                                if tooltip_text and any(c.isdigit() for c in tooltip_text):
                                                    tooltip_amount = tooltip_text
                                                    break

                                            # Alternative: Try finding any visible tooltip
                                            if not tooltip_amount:
                                                visible_tooltips = self.page.locator('[role="tooltip"]:visible, .core-styles-module_tooltip__3vRRp:visible').all()
                                                for tt in visible_tooltips:
                                                    try:
                                                        tt_text = tt.inner_text().strip()
                                                        if tt_text and any(c.isdigit() for c in tt_text):
                                                            tooltip_amount = tt_text
                                                            break
                                                    except:
                                                        continue

                                            # Move mouse away to hide tooltip
                                            self.page.mouse.move(0, 0)
                                            break

                                    except:
                                        continue

                            except:
                                pass

                            tooltip_amounts.append(tooltip_amount)

                        except:
                            # Fallback to text_content if inner_text fails
                            text = cell.text_content().strip()
                            lines = [line.strip() for line in text.split('\n') if line.strip()]
                            text = lines[0] if lines else text
                            cell_values.append(text)
                            price_24h_values.append("")
                            tooltip_amounts.append("")

                    # Check for button/icon in first cell
                    first_cell_has_button = row.locator("td:first-child button.cursor-pointer.rounded").count() > 0

                    if first_cell_has_button:
                        # Token Main row - extract token name and share% for use in child chain rows
                        # Cells: [button] | Token | Share% | Amount | Value
                        data_cells = cell_values[1:5] if len(cell_values) >= 5 else cell_values[1:]

                        if len(data_cells) >= 2:
                            current_token = data_cells[0]  # Token name
                            current_share = data_cells[1]  # Portfolio share %

                        # Skip Token_Main rows - we only want chain details

                    else:
                        # Chain Detail row - this is what we want to extract
                        # Cells: Chain | Amount | Price | Value
                        if current_token:  # Only add if we have a parent token
                            # Skip header rows (where first cell is "Chain", "Amount", "Price", etc.)
                            if len(cell_values) > 0 and cell_values[0].lower() in ["chain", "amount", "price", "value"]:
                                continue

                            row_data = {
                                "Portfolio": portfolio_name,
                                "Row_Type": "Token_Main",  # User wants all rows labeled as Token_Main
                                "Parent_Token": current_token,
                                "Row_Number": idx
                            }

                            # Map cells: Chain | Amount | Price | Value
                            data_cells = cell_values[:4] if len(cell_values) >= 4 else cell_values

                            for i in range(min(len(data_cells), len(chain_detail_headers))):
                                column_name = chain_detail_headers[i]
                                row_data[column_name] = data_cells[i]

                            # Add Share from parent token
                            row_data["Share"] = current_share if current_share else ""

                            # Extract Price_24h from Price cell (index 2 in chain detail row)
                            if len(price_24h_values) > 2:
                                price_24h = price_24h_values[2]  # Price column is at index 2
                                if price_24h:
                                    row_data["Price_24h"] = price_24h
                                else:
                                    row_data["Price_24h"] = ""
                            else:
                                row_data["Price_24h"] = ""

                            # Add Amount_Tooltip from tooltip (index 1 = Amount column in chain detail row)
                            if len(tooltip_amounts) > 1 and tooltip_amounts[1]:
                                row_data["Amount_Tooltip"] = tooltip_amounts[1]
                            else:
                                row_data["Amount_Tooltip"] = ""

                            holdings_data.append(row_data)

                except Exception as e:
                    print(f"   ⚠️  Error extracting row {idx}: {e}")
                    continue

            # Filter out rows without Parent_Token
            filtered_data = []
            orphan_count = 0
            for row in holdings_data:
                if row.get("Parent_Token"):
                    filtered_data.append(row)
                else:
                    orphan_count += 1

            if orphan_count > 0:
                print(f"   🚫 Filtered out {orphan_count} rows without Parent_Token")

            holdings_data = filtered_data

            # Remove duplicate rows based on unique combination of fields
            seen = set()
            deduped_data = []
            for row in holdings_data:
                # Create a unique key based on all main data fields
                row_key = (
                    row.get("Parent_Token"),
                    row.get("Chain"),
                    row.get("Share"),
                    row.get("Amount"),
                    row.get("Price"),
                    row.get("Value")
                )

                if row_key not in seen:
                    seen.add(row_key)
                    deduped_data.append(row)

            holdings_data = deduped_data

            # Count rows
            print(f"✅ Extracted {len(holdings_data)} chain detail rows (after filtering & deduplication)")

        except Exception as e:
            print(f"❌ Error extracting Holdings Token data: {e}")
            import traceback
            traceback.print_exc()

        return holdings_data

    def extract_holdings_chain_table(self, portfolio_name: str, holdings_token_data: list = None) -> list:
        """
        Extract holdings data from Holdings page - Chain view

        IMPORTANT: The "Chain" view actually groups data by TOKEN (not by Chain!)
        Structure:
        1. Address row (no chevron): Address, blank, Share%, Token count, Total value
        2. Token row (with chevron): blank, Token name, Share%, Amount, Value
        3. Chain detail rows (no chevron, nested under token): Chain name, Amount, Price, Value

        Expected output columns:
        Portfolio, Row_Type, Parent_Chain, Chain, Share, Token, Amount, Price, Price (24h), Value

        For the output, we transform this to show:
        - Chain_Main rows: One row per unique chain with aggregated data
        - Token_Detail rows: Tokens grouped under their respective chains

        Args:
            portfolio_name: Name of the portfolio
            holdings_token_data: Optional list of holdings token data to fill in missing chain info

        Returns:
            List of dictionaries with Chain_Main and Token_Detail rows
        """
        print(f"📊 Extracting Holdings - Chain table for '{portfolio_name}'...")

        holdings_chain_data = []

        try:
            # Wait for page to load
            self.page.wait_for_timeout(2000)

            # Look for table rows in Holdings page
            table_rows = self.page.locator("tbody tr").all()

            # Filter out rows without td elements
            table_rows = [row for row in table_rows if row.locator("td, [role='cell']").count() > 0]

            print(f"   Found {len(table_rows)} total rows (address + tokens + chain details)")

            # First pass: Extract all token-chain relationships
            token_chain_map = {}  # {token_name: [(chain, amount, price, price_24h, value), ...]}
            current_token = None
            current_token_share = None

            # Debug: Print first 10 rows to understand structure
            print(f"   DEBUG: Analyzing first 10 rows to understand table structure...")
            for i in range(min(10, len(table_rows))):
                row = table_rows[i]
                cells = row.locator("td, [role='cell']").all()
                cell_texts = [c.inner_text().strip().replace('\n', ' ')[:40] for c in cells[:5]]
                has_chevron = row.locator("td:first-child button.cursor-pointer.rounded").count() > 0
                print(f"   DEBUG Row {i+1}: has_chevron={has_chevron}, cells={len(cells)}, first_5_cells={cell_texts}")

            # First pass: Extract raw data and build token-chain map
            for idx, row in enumerate(table_rows, 1):
                try:
                    cells = row.locator("td, [role='cell']").all()
                    if len(cells) == 0:
                        continue

                    # Extract cell values, price changes, and tooltips
                    cell_values = []
                    price_24h_values = []
                    tooltip_amounts = []

                    for cell in cells:
                        try:
                            text = cell.inner_text().strip()
                            lines = [line.strip() for line in text.split('\n') if line.strip()]
                            text = lines[0] if lines else text
                            cell_values.append(text)

                            price_24h = self.extract_price_24h_change(cell)
                            price_24h_values.append(price_24h)

                            # Extract tooltip amount if it exists
                            tooltip_amount = ""
                            try:
                                tooltip_elems = cell.locator('[data-tooltip-id*="amount"]').all()
                                for tooltip_elem in tooltip_elems:
                                    tooltip_id = tooltip_elem.get_attribute("data-tooltip-id")
                                    if tooltip_id and "amount" in tooltip_id.lower():
                                        tooltip_elem.hover(timeout=300)
                                        self.page.wait_for_timeout(50)

                                        tooltip_selector = f'[id="{tooltip_id}"], [data-tooltip-id="{tooltip_id}"][role="tooltip"]'
                                        tooltip_popup = self.page.locator(tooltip_selector).first
                                        if tooltip_popup.is_visible(timeout=100):
                                            tooltip_text = tooltip_popup.inner_text().strip()
                                            if tooltip_text and any(c.isdigit() for c in tooltip_text):
                                                tooltip_amount = tooltip_text
                                                self.page.mouse.move(0, 0)
                                                break
                            except:
                                pass

                            tooltip_amounts.append(tooltip_amount)

                        except:
                            text = cell.text_content().strip()
                            lines = [line.strip() for line in text.split('\n') if line.strip()]
                            text = lines[0] if lines else text
                            cell_values.append(text)
                            price_24h_values.append("")
                            tooltip_amounts.append("")

                    has_chevron = row.locator("td:first-child button.cursor-pointer.rounded").count() > 0

                    if has_chevron:
                        # Token row: [chevron] | Token name | Share% | Amount | Value | [toggle]
                        # Cells: ['', 'GTUSDCP', '99.99%', '1,734,833.3491', '$1,871,234.57', toggle]
                        token_name = cell_values[1] if len(cell_values) > 1 else ""
                        token_share = cell_values[2] if len(cell_values) > 2 else ""

                        current_token = token_name
                        current_token_share = token_share

                        # Initialize list for this token's chains if not exists
                        if token_name not in token_chain_map:
                            token_chain_map[token_name] = {
                                "share": token_share,
                                "chains": []
                            }

                    elif current_token:
                        # Chain detail row (nested under token): Chain | Amount | Price | Value
                        # Skip header rows - they contain "Chain" and have more than 4 cells or unusual structure
                        if len(cell_values) > 0:
                            # Header rows have "Chain" in first cell and unusual cell count (9 cells) or contain tab characters
                            if "Chain" in cell_values[0] and ("Amount" in cell_values[0] or len(cells) > 6):
                                continue
                            # Also skip if first cell starts with "Chain" (exact match)
                            if cell_values[0] == "Chain":
                                continue

                        # Regular chain detail row: Chain | Amount | Price | Value
                        chain_name = cell_values[0] if len(cell_values) > 0 else ""
                        amount = cell_values[1] if len(cell_values) > 1 else ""
                        price = cell_values[2] if len(cell_values) > 2 else ""
                        value = cell_values[3] if len(cell_values) > 3 else ""

                        # Skip empty rows
                        if not chain_name or not amount:
                            continue

                        # Extract price 24h from price cell (index 2)
                        price_24h = price_24h_values[2] if len(price_24h_values) > 2 else ""

                        # Extract amount tooltip from amount cell (index 1)
                        amount_tooltip = tooltip_amounts[1] if len(tooltip_amounts) > 1 else ""

                        # Add to token's chain list
                        if chain_name and current_token in token_chain_map:
                            token_chain_map[current_token]["chains"].append({
                                "chain": chain_name,
                                "amount": amount,
                                "amount_tooltip": amount_tooltip,
                                "price": price,
                                "price_24h": price_24h,
                                "value": value
                            })

                except Exception as e:
                    print(f"   ⚠️  Error extracting row {idx}: {e}")
                    continue

            # Handle tokens without chain details by using holdings_token_data as fallback
            if holdings_token_data:
                for token_name, token_info in token_chain_map.items():
                    if len(token_info["chains"]) == 0:
                        # This token has no chain details in Holdings > Chain view
                        # Find it in holdings_token_data to get chain info
                        for token_record in holdings_token_data:
                            if token_record.get("Parent_Token") == token_name:
                                # Found the token, use its data
                                chain = token_record.get("Chain", "")
                                amount = token_record.get("Amount", "")
                                price = token_record.get("Price", "")
                                value = token_record.get("Value", "")

                                # Try to get Price_24h and Amount_Tooltip from token record
                                price_24h = token_record.get("Price_24h", "")
                                amount_tooltip = token_record.get("Amount_Tooltip", "")

                                token_info["chains"].append({
                                    "chain": chain,
                                    "amount": amount,
                                    "amount_tooltip": amount_tooltip,
                                    "price": price,
                                    "price_24h": price_24h,
                                    "value": value
                                })

                print(f"   ℹ️  Filled {sum(1 for t in token_chain_map.values() if len(t['chains']) > 0)} tokens with chain data")

            # Second pass: Transform to chain-grouped structure
            # Build chain aggregation: {chain_name: {tokens: [...], total_value: ..., share: ...}}
            chain_aggregation = {}

            for token_name, token_data in token_chain_map.items():
                for chain_info in token_data["chains"]:
                    chain = chain_info["chain"]

                    if chain not in chain_aggregation:
                        chain_aggregation[chain] = {
                            "tokens": [],
                            "token_count": 0
                        }

                    chain_aggregation[chain]["tokens"].append({
                        "token": token_name,
                        "share": token_data["share"],
                        "amount": chain_info["amount"],
                        "amount_tooltip": chain_info["amount_tooltip"],
                        "price": chain_info["price"],
                        "price_24h": chain_info["price_24h"],
                        "value": chain_info["value"]
                    })
                    chain_aggregation[chain]["token_count"] += 1

            # Generate output rows: Chain_Main + Token_Detail rows
            for chain_name, chain_data in sorted(chain_aggregation.items()):
                # Calculate chain totals
                total_value = 0
                total_share = 0

                for token_info in chain_data["tokens"]:
                    # Try to parse value for aggregation
                    try:
                        val_str = token_info["value"].replace("$", "").replace(",", "").strip()
                        if val_str and val_str != "":
                            total_value += float(val_str)
                    except:
                        pass

                    # Try to parse share
                    try:
                        share_str = token_info["share"].replace("%", "").strip()
                        if share_str and share_str != "" and share_str != "< 0.00001":
                            total_share += float(share_str)
                    except:
                        pass

                # Add Chain_Main row
                chain_main_row = {
                    "Portfolio": portfolio_name,
                    "Row_Type": "Chain_Main",
                    "Parent_Chain": chain_name,
                    "Chain": chain_name,
                    "Share": f"{total_share:.2f}%" if total_share > 0 else "",
                    "Token": str(chain_data["token_count"]),
                    "Amount": "",
                    "Amount_Tooltip": "",
                    "Price": "",
                    "Price_24h": "",
                    "Value": f"${total_value:,.2f}" if total_value > 0 else ""
                }
                holdings_chain_data.append(chain_main_row)

                # Add Token_Detail rows
                for token_info in chain_data["tokens"]:
                    token_detail_row = {
                        "Portfolio": portfolio_name,
                        "Row_Type": "Token_Detail",
                        "Parent_Chain": chain_name,
                        "Chain": chain_name,
                        "Share": token_info["share"],
                        "Token": token_info["token"],
                        "Amount": token_info["amount"],
                        "Amount_Tooltip": token_info["amount_tooltip"],
                        "Price": token_info["price"],
                        "Price_24h": token_info["price_24h"],
                        "Value": token_info["value"]
                    }
                    holdings_chain_data.append(token_detail_row)

            # Count main and detail rows
            main_rows = sum(1 for row in holdings_chain_data if row.get("Row_Type") == "Chain_Main")
            detail_rows = sum(1 for row in holdings_chain_data if row.get("Row_Type") == "Token_Detail")
            print(f"✅ Extracted {main_rows} chains and {detail_rows} token details (transformed from token-grouped view)")

        except Exception as e:
            print(f"❌ Error extracting Holdings Chain data: {e}")
            import traceback
            traceback.print_exc()

        return holdings_chain_data

    def extract_wallet_data(self, portfolio_name: str) -> list:
        """
        Extract wallet/address data from Holdings page

        Args:
            portfolio_name: Name of the portfolio

        Returns:
            List of dictionaries with wallet data
        """
        print(f"📊 Extracting wallet data for '{portfolio_name}'...")

        wallet_data = []

        try:
            # Wait for page to load
            self.page.wait_for_timeout(2000)

            # Look for wallet information
            # This might be in a different section or table
            wallet_elements = self.page.locator("[data-testid*='wallet'], [class*='wallet']").all()

            print(f"   Found {len(wallet_elements)} wallet elements")

            for idx, element in enumerate(wallet_elements, 1):
                try:
                    text = element.text_content().strip()

                    wallet = {
                        "Portfolio": portfolio_name,
                        "Wallet/Chain": text,
                        "Type": "Unknown"
                    }

                    wallet_data.append(wallet)

                except Exception as e:
                    print(f"   ⚠️  Error extracting wallet {idx}: {e}")
                    continue

            print(f"✅ Extracted {len(wallet_data)} wallets")

        except Exception as e:
            print(f"❌ Error extracting wallet data: {e}")

        return wallet_data


def remove_currency_symbols(value_str: str) -> str:
    """
    Remove only $ and % symbols from value, keep commas and numbers as-is

    - Removes ONLY: $, %
    - Keeps: commas, decimal points, numbers, k/m/b abbreviations
    - Examples:
      - "$1,500.50" → "1,500.50"
      - "45.23%" → "45.23"
      - "$2.5k" → "2.5k"
      - "$1.2m" → "1.2m"

    Args:
        value_str: String value that may contain currency symbols, percentages

    Returns:
        Clean string with only $ and % removed
    """
    if not value_str or value_str == "":
        return ""

    try:
        # Remove only $ and % symbols, keep everything else
        cleaned = value_str.replace("$", "").replace("%", "").strip()
        return cleaned
    except:
        return value_str  # Return original if anything fails


def parse_value_to_full_amount(value_str: str) -> str:
    """
    Parse abbreviated values and clean special characters

    - Removes: $, %,  commas
    - Converts: k (×1,000), m (×1,000,000), b (×1,000,000,000)
    - Examples:
      - "$1,500.50" → "1500.50"
      - "45.23%" → "45.23"
      - "2.5k" → "2500"
      - "1.2m" → "1200000"
      - "3.5b" → "3500000000"

    Args:
        value_str: String value that may contain currency symbols, percentages, abbreviations

    Returns:
        Clean numeric string without special characters
    """
    if not value_str or value_str == "":
        return ""

    try:
        # Store original for error handling
        original = value_str

        # Remove whitespace
        cleaned = value_str.strip()

        # Handle special cases
        if cleaned in ["-", "N/A", "n/a", "N/A"]:
            return ""

        # Remove currency symbols and commas (but keep percentage sign for now)
        cleaned = cleaned.replace("$", "").replace(",", "").strip()

        # Check for k, m, b multipliers (case insensitive)
        multiplier = 1
        if cleaned.lower().endswith('k'):
            multiplier = 1000
            cleaned = cleaned[:-1].strip()
        elif cleaned.lower().endswith('m'):
            multiplier = 1000000
            cleaned = cleaned[:-1].strip()
        elif cleaned.lower().endswith('b'):
            multiplier = 1000000000
            cleaned = cleaned[:-1].strip()

        # Remove percentage sign
        cleaned = cleaned.replace("%", "").strip()

        # Try to parse as float and apply multiplier
        float_val = float(cleaned) * multiplier

        # Format to reasonable precision (remove unnecessary decimals)
        if float_val == int(float_val):
            # If it's a whole number, return as integer
            return str(int(float_val))
        else:
            # Otherwise, keep up to 8 decimal places and remove trailing zeros
            return f"{float_val:.8f}".rstrip('0').rstrip('.')

    except (ValueError, AttributeError, TypeError):
        # If parsing fails, try to at least remove $ and % symbols
        try:
            return value_str.replace("$", "").replace("%", "").replace(",", "").strip()
        except:
            return value_str  # Return original if all else fails


def export_to_excel(overview_data: list, holdings_token_data: list, holdings_chain_data: list,
                    chain_breakdown_data: list = None, chain_activity_data: list = None,
                    address_breakdown_data: list = None, output_path: str = None, screenshot_paths: dict = None):
    """
    Export Overview, Holdings, Chain, and Address data to Excel file

    Args:
        overview_data: List of overview data dictionaries
        holdings_token_data: List of holdings token view data dictionaries
        holdings_chain_data: List of holdings chain view data dictionaries
        chain_breakdown_data: List of chain breakdown data dictionaries (optional)
        chain_activity_data: List of chain activity data dictionaries (optional)
        address_breakdown_data: List of address breakdown data dictionaries (optional)
        output_path: Path to save Excel file (optional)
        screenshot_paths: Dictionary mapping sheet names to screenshot file paths (optional)
                         Example: {'Overview - Token': '/path/to/screenshot.png'}

    Returns:
        Path to saved Excel file
    """
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"test-results/excel-reports/dam_data_export_{timestamp}.xlsx"

    # Ensure directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print(f"\n📊 Exporting data to Excel: {output_path}")

    try:
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            # Export Overview data (Token Breakdown table)
            if overview_data:
                df_overview = pd.DataFrame(overview_data)

                # Add full amount columns for Price, Amount, and Value
                # ORIGINAL data in main columns with only $ and % removed
                if 'Price' in df_overview.columns:
                    df_overview['Price (Full Amount)'] = df_overview['Price'].apply(parse_value_to_full_amount)
                    df_overview['Price'] = df_overview['Price'].apply(remove_currency_symbols)
                if 'Amount' in df_overview.columns:
                    df_overview['Amount (Full Amount)'] = df_overview['Amount'].apply(parse_value_to_full_amount)
                    df_overview['Amount'] = df_overview['Amount'].apply(remove_currency_symbols)
                if 'Amount_Tooltip' in df_overview.columns:
                    # Clean the tooltip amount as well
                    df_overview['Amount_Tooltip'] = df_overview['Amount_Tooltip'].apply(remove_currency_symbols)
                if 'Value' in df_overview.columns:
                    df_overview['Value (Full Amount)'] = df_overview['Value'].apply(parse_value_to_full_amount)
                    df_overview['Value'] = df_overview['Value'].apply(remove_currency_symbols)
                if 'Share' in df_overview.columns:
                    df_overview['Share'] = df_overview['Share'].apply(remove_currency_symbols)
                if 'Price_24h' in df_overview.columns:
                    df_overview['Price_24h'] = df_overview['Price_24h'].apply(remove_currency_symbols)
                if 'Price_Change_24h' in df_overview.columns:
                    df_overview['Price_Change_24h'] = df_overview['Price_Change_24h'].apply(remove_currency_symbols)

                # Add validation columns (empty for now - user will fill manually)
                df_overview['Price Validation'] = ''
                df_overview['Price Diff'] = ''
                df_overview['Price Change Validation'] = ''
                df_overview['Price Change Diff'] = ''
                df_overview['Amount Validation'] = ''
                df_overview['Amount Diff'] = ''
                df_overview['Share Validation'] = ''
                df_overview['Value Validation'] = ''
                df_overview['Value Diff'] = ''
                df_overview['Remark'] = ''

                # Reorganize columns to match user's expected format with validation columns interspersed
                # Expected order: Row_Number, Name, Price, Price (Full Amount), Price Validation, Price Diff,
                #                 Price_24h, Price Change Validation, Price Change Diff,
                #                 Amount, Amount (Full Amount), Amount_Tooltip, Amount Validation, Amount Diff,
                #                 Share, Share Validation, Value, Value (Full Amount), Value Validation, Value Diff, Remark
                # Note: Price, Amount, Value now cleaned (no $, %), Full Amount columns kept for compatibility
                overview_column_order = [
                    'Row_Number', 'Name',
                    'Price', 'Price (Full Amount)', 'Price Validation', 'Price Diff',
                    'Price_24h', 'Price Change Validation', 'Price Change Diff',
                    'Amount', 'Amount (Full Amount)', 'Amount_Tooltip', 'Amount Validation', 'Amount Diff',
                    'Share', 'Share Validation',
                    'Value', 'Value (Full Amount)', 'Value Validation', 'Value Diff',
                    'Remark'
                ]
                # Only include columns that exist in the DataFrame
                existing_cols = [col for col in overview_column_order if col in df_overview.columns]
                df_overview = df_overview[existing_cols]

                df_overview.to_excel(writer, sheet_name='DAM - Overview - Token', index=False)
                print(f"   ✅ Exported {len(overview_data)} rows to 'DAM - Overview - Token' sheet")

            # Combine Chain Breakdown and Chain Activity into one sheet with separate tables
            if chain_breakdown_data or chain_activity_data:
                combined_chain_data = []

                # Add Chain Breakdown section with title row
                if chain_breakdown_data:
                    # Add "Chain Breakdown" title row (no column headers)
                    title_row = {"Table_Title": "Chain Breakdown"}
                    combined_chain_data.append(title_row)

                    # Add Chain Breakdown data rows
                    for row in chain_breakdown_data:
                        data_row = {"Table_Title": "", **row}
                        combined_chain_data.append(data_row)

                    # Add 3 blank separator rows
                    for _ in range(3):
                        separator = {"Table_Title": ""}
                        if chain_breakdown_data:
                            for key in chain_breakdown_data[0].keys():
                                separator[key] = ""
                        combined_chain_data.append(separator)

                # Add Chain Activity section with title row
                if chain_activity_data:
                    # Add "Chain Activity" title row (no column headers)
                    title_row = {"Table_Title": "Chain Activity"}
                    combined_chain_data.append(title_row)

                    # Add Chain Activity data rows
                    for row in chain_activity_data:
                        data_row = {"Table_Title": "", **row}
                        combined_chain_data.append(data_row)

                # Export combined data
                if combined_chain_data:
                    df_combined_chain = pd.DataFrame(combined_chain_data)

                    # Remove $ and % symbols from Value columns if they exist
                    if 'Value' in df_combined_chain.columns:
                        df_combined_chain['Value'] = df_combined_chain['Value'].apply(remove_currency_symbols)
                    if 'Share' in df_combined_chain.columns:
                        df_combined_chain['Share'] = df_combined_chain['Share'].apply(remove_currency_symbols)

                    # Add validation columns (empty for now - user will fill manually)
                    df_combined_chain['Share Validation'] = ''
                    df_combined_chain['Token Validation'] = ''
                    df_combined_chain['Value Validation'] = ''
                    df_combined_chain['Value Diff'] = ''
                    df_combined_chain['Current Value Validation'] = ''

                    # Rearrange columns to match user's expected order
                    desired_column_order = [
                        'Table_Title', 'Row_Number', 'Name', 'Share', 'Share Validation',
                        'Token', 'Token Validation', 'Protocol',
                        'Value', 'Value Validation', 'Value Diff',
                        'Chain', 'Txns', 'Gas_Spent', 'Current_Value', 'Current Value Validation'
                    ]
                    # Only include columns that exist in the DataFrame
                    existing_cols = [col for col in desired_column_order if col in df_combined_chain.columns]
                    df_combined_chain = df_combined_chain[existing_cols]

                    try:
                        df_combined_chain.to_excel(writer, sheet_name='DAM - Overview - Chain', index=False)
                        print(f"   ✅ Exported {len(combined_chain_data)} rows to 'DAM - Overview - Chain' sheet")
                        print(f"      - Chain Breakdown: {len(chain_breakdown_data)} rows")
                        print(f"      - Chain Activity: {len(chain_activity_data)} rows")
                    except Exception as e:
                        print(f"   ⚠️  Error exporting DAM - Overview - Chain: {e}")
                        # Export without validation columns as fallback
                        df_combined_chain_backup = df_combined_chain.drop(columns=['Share Validation', 'Token Validation', 'Value Validation', 'Value Diff', 'Current Value Validation'], errors='ignore')
                        df_combined_chain_backup.to_excel(writer, sheet_name='DAM - Overview - Chain', index=False)
                        print(f"   ✅ Exported {len(combined_chain_data)} rows to 'DAM - Overview - Chain' sheet (without validation columns)")

            # Export Address Breakdown data (from Overview page Address button)
            if address_breakdown_data:
                df_address_breakdown = pd.DataFrame(address_breakdown_data)

                # Remove $ and % symbols from Value columns if they exist
                if 'Value' in df_address_breakdown.columns:
                    df_address_breakdown['Value'] = df_address_breakdown['Value'].apply(remove_currency_symbols)
                if 'Share' in df_address_breakdown.columns:
                    df_address_breakdown['Share'] = df_address_breakdown['Share'].apply(remove_currency_symbols)

                # Add validation columns (empty for now - user will fill manually)
                df_address_breakdown['Token Validation'] = ''
                df_address_breakdown['Value Validation'] = ''

                # Rearrange columns to match user's expected order
                desired_column_order = [
                    'Row_Number', 'Address', 'Chain', 'Share', 'Token', 'Token Validation',
                    'Protocol', 'Value', 'Value Validation'
                ]
                # Only include columns that exist in the DataFrame
                existing_cols = [col for col in desired_column_order if col in df_address_breakdown.columns]
                df_address_breakdown = df_address_breakdown[existing_cols]

                df_address_breakdown.to_excel(writer, sheet_name='DAM - Overview - Address', index=False)
                print(f"   ✅ Exported {len(address_breakdown_data)} rows to 'DAM - Overview - Address' sheet")

            # Export Holdings - Token data
            if holdings_token_data:
                df_holdings_token = pd.DataFrame(holdings_token_data)

                # Add full amount columns - ORIGINAL data with only $ and % removed
                if 'Price' in df_holdings_token.columns:
                    df_holdings_token['Price (Full Amount)'] = df_holdings_token['Price'].apply(parse_value_to_full_amount)
                    df_holdings_token['Price'] = df_holdings_token['Price'].apply(remove_currency_symbols)
                if 'Amount' in df_holdings_token.columns:
                    df_holdings_token['Amount (Full Amount)'] = df_holdings_token['Amount'].apply(parse_value_to_full_amount)
                    df_holdings_token['Amount'] = df_holdings_token['Amount'].apply(remove_currency_symbols)
                if 'Value' in df_holdings_token.columns:
                    df_holdings_token['Value (Full Amount)'] = df_holdings_token['Value'].apply(parse_value_to_full_amount)
                    df_holdings_token['Value'] = df_holdings_token['Value'].apply(remove_currency_symbols)
                if 'Portfolio_%' in df_holdings_token.columns:
                    df_holdings_token['Portfolio_%'] = df_holdings_token['Portfolio_%'].apply(remove_currency_symbols)
                if 'Price_Change_24h' in df_holdings_token.columns:
                    df_holdings_token['Price_Change_24h'] = df_holdings_token['Price_Change_24h'].apply(remove_currency_symbols)

                # Add validation columns (empty for now - user will fill manually)
                df_holdings_token['Portfolio Share Validation'] = ''
                df_holdings_token['Amount Validation'] = ''
                df_holdings_token['Amount Diff'] = ''
                df_holdings_token['Price Validation'] = ''
                df_holdings_token['Price Diff'] = ''
                df_holdings_token['Price Change Validation'] = ''
                df_holdings_token['Price Change Diff'] = ''
                df_holdings_token['Value Validation'] = ''
                df_holdings_token['Value Diff'] = ''
                df_holdings_token['Remark'] = ''

                # Reorganize columns to match user's expected format with validation columns interspersed
                # Expected order: Portfolio, Row_Type, Row_Number, Parent_Token, Chain, Token, Portfolio_%, Portfolio Share Validation,
                #                 Amount, Amount (Full Amount), Amount Validation, Amount Diff,
                #                 Price, Price (Full Amount), Price Validation, Price Diff,
                #                 Price_Change_24h, Price Change Validation, Price Change Diff,
                #                 Value, Value (Full Amount), Value Validation, Value Diff, Remark
                # Note: Price, Amount, Value now cleaned (no $, %), Full Amount columns kept for compatibility
                token_column_order = [
                    'Portfolio', 'Row_Type', 'Row_Number', 'Parent_Token', 'Chain', 'Token', 'Portfolio_%', 'Portfolio Share Validation',
                    'Amount', 'Amount (Full Amount)', 'Amount Validation', 'Amount Diff',
                    'Price', 'Price (Full Amount)', 'Price Validation', 'Price Diff',
                    'Price_Change_24h', 'Price Change Validation', 'Price Change Diff',
                    'Value', 'Value (Full Amount)', 'Value Validation', 'Value Diff',
                    'Remark'
                ]
                # Only include columns that exist in the DataFrame
                existing_cols = [col for col in token_column_order if col in df_holdings_token.columns]
                df_holdings_token = df_holdings_token[existing_cols]

                df_holdings_token.to_excel(writer, sheet_name='DAM - Holdings - Token', index=False)
                print(f"   ✅ Exported {len(holdings_token_data)} rows to 'DAM - Holdings - Token' sheet")

            # Export Holdings - Chain data
            if holdings_chain_data:
                df_holdings_chain = pd.DataFrame(holdings_chain_data)

                # Add full amount columns - ORIGINAL data with only $ and % removed
                if 'Price' in df_holdings_chain.columns:
                    df_holdings_chain['Price (Full Amount)'] = df_holdings_chain['Price'].apply(parse_value_to_full_amount)
                    df_holdings_chain['Price'] = df_holdings_chain['Price'].apply(remove_currency_symbols)
                if 'Amount' in df_holdings_chain.columns:
                    df_holdings_chain['Amount (Full Amount)'] = df_holdings_chain['Amount'].apply(parse_value_to_full_amount)
                    df_holdings_chain['Amount'] = df_holdings_chain['Amount'].apply(remove_currency_symbols)
                if 'Value' in df_holdings_chain.columns:
                    df_holdings_chain['Value (Full Amount)'] = df_holdings_chain['Value'].apply(parse_value_to_full_amount)
                    df_holdings_chain['Value'] = df_holdings_chain['Value'].apply(remove_currency_symbols)
                if 'Share' in df_holdings_chain.columns:
                    df_holdings_chain['Share'] = df_holdings_chain['Share'].apply(remove_currency_symbols)
                if 'Price_Change_24h' in df_holdings_chain.columns:
                    df_holdings_chain['Price_Change_24h'] = df_holdings_chain['Price_Change_24h'].apply(remove_currency_symbols)

                # Add validation columns (empty for now - user will fill manually)
                df_holdings_chain['Share Validation'] = ''
                df_holdings_chain['Amount Validation'] = ''
                df_holdings_chain['Amount Diff'] = ''
                df_holdings_chain['Price Validation'] = ''
                df_holdings_chain['Price Diff'] = ''
                df_holdings_chain['Price Change Validation'] = ''
                df_holdings_chain['Price Change Diff'] = ''
                df_holdings_chain['Value Validation'] = ''
                df_holdings_chain['Value Diff'] = ''
                df_holdings_chain['Remark'] = ''

                # Reorganize columns to match user's expected format with validation columns interspersed
                # Expected order: Portfolio, Row_Type, Parent_Chain, Chain, Share, Share Validation, Token,
                #                 Amount, Amount (Full Amount), Amount Validation, Amount Diff,
                #                 Price, Price (Full Amount), Price Validation, Price Diff,
                #                 Price_24h, Price Change Validation, Price Change Diff,
                #                 Value, Value (Full Amount), Value Validation, Value Diff, Remark
                # Note: Price, Amount, Value now cleaned (no $, %), Full Amount columns kept for compatibility
                chain_column_order = [
                    'Portfolio', 'Row_Type', 'Parent_Chain', 'Chain', 'Share', 'Share Validation', 'Token',
                    'Amount', 'Amount (Full Amount)', 'Amount Validation', 'Amount Diff',
                    'Price', 'Price (Full Amount)', 'Price Validation', 'Price Diff',
                    'Price_24h', 'Price Change Validation', 'Price Change Diff',
                    'Value', 'Value (Full Amount)', 'Value Validation', 'Value Diff',
                    'Remark'
                ]
                # Only include columns that exist in the DataFrame
                existing_cols = [col for col in chain_column_order if col in df_holdings_chain.columns]
                df_holdings_chain = df_holdings_chain[existing_cols]

                df_holdings_chain.to_excel(writer, sheet_name='DAM - Holdings - Chain', index=False)
                print(f"   ✅ Exported {len(holdings_chain_data)} rows to 'DAM - Holdings - Chain' sheet")

            # Debank and CoinGecko export removed as per updated requirements
            # Export Google Drive Summary sheet
            summary_data = []

            # Header rows
            summary_data.append({'Section': 'Google Drive', 'Data': '', 'Data Source': '', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': '', 'Data Source': '', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': '', 'Data Source': '', 'Acceptance criteria': ''})
            summary_data.append({'Section': 'E...', 'Data': 'Export Excel File', 'Data Source': '', 'Acceptance criteria': ''})
            summary_data.append({'Section': 'DAM', 'Data': 'Data shown on DAM', 'Data Source': '', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': '', 'Data Source': '', 'Acceptance criteria': ''})

            # DAM - Overview - Token section
            summary_data.append({'Section': '', 'Data': 'Data', 'Data Source': 'Data Source', 'Acceptance criteria': 'Acceptance criteria'})
            summary_data.append({'Section': '', 'Data': 'Price', 'Data Source': '1. Coingecko\n2. Debank', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Price change', 'Data Source': '1. Coingecko\n2. show 0 for those token which is not listed on Coingecko', 'Acceptance criteria': ''})
            summary_data.append({'Section': 'DAM - Overview - Token', 'Data': 'Amount', 'Data Source': '1. Debank/ Protocol - Balance inside Contract Address\n2. I will double check Etherscan when amount has diff > 2%', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Share', 'Data Source': 'Formula calculation', 'Acceptance criteria': 'share diff <0.01'})
            summary_data.append({'Section': '', 'Data': 'Value', 'Data Source': 'Formula calculation\n1. Debank', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Remark - 2 spam tokens are displayed', 'Data Source': '', 'Acceptance criteria': ''})

            # Excel - Overview - Token section
            summary_data.append({'Section': '', 'Data': 'Price', 'Data Source': '1. Coingecko\n2. Debank', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Price change', 'Data Source': '1. Coingecko\n2. show 0 for those token which is not listed on Coingecko', 'Acceptance criteria': ''})
            summary_data.append({'Section': 'Excel - Overview - Token', 'Data': 'Amount', 'Data Source': '1. Protocol - Balance inside Contract Address\n2. Etherscan when amount has diff > 2%\n3. Debank', 'Acceptance criteria': 'I expect the amount to match exactly with Etherscan, but the value in the Excel file shows more decimal places.'})
            summary_data.append({'Section': '', 'Data': 'Share', 'Data Source': 'Formula calculation', 'Acceptance criteria': 'TBC'})
            summary_data.append({'Section': '', 'Data': 'Value', 'Data Source': 'Formula calculation\n1. Debank', 'Acceptance criteria': 'within 2% diff'})

            # DAM - Overview - Chain - Chain Breakdown section
            summary_data.append({'Section': '', 'Data': 'Share', 'Data Source': 'Formula calculation', 'Acceptance criteria': 'share diff <0.01'})
            summary_data.append({'Section': '', 'Data': 'Token', 'Data Source': '1. Debank', 'Acceptance criteria': ''})
            summary_data.append({'Section': 'DAM - Overview - Chain - Chain Breakdown', 'Data': 'Value', 'Data Source': '1. Debank', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Protocol', 'Data Source': '', 'Acceptance criteria': ''})

            # DAM - Overview - Chain - Chain Activity section
            summary_data.append({'Section': 'DAM - Overview - Chain - Chain Activity', 'Data': 'Txns', 'Data Source': 'TBC', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': 'Gas Spent', 'Data Source': '', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': 'Current Value', 'Data Source': '1. Debank', 'Acceptance criteria': 'within 2% diff'})

            # Excel - Overview - Chain section
            summary_data.append({'Section': '', 'Data': 'Share', 'Data Source': 'Formula calculation', 'Acceptance criteria': 'share diff <0.01'})
            summary_data.append({'Section': '', 'Data': 'Token', 'Data Source': '1. Debank', 'Acceptance criteria': ''})
            summary_data.append({'Section': 'Excel - Overview - Chain', 'Data': 'Value', 'Data Source': '', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Protocol', 'Data Source': 'TBC', 'Acceptance criteria': ''})

            # DAM - Overview - Address section
            summary_data.append({'Section': '', 'Data': 'Token', 'Data Source': '1. Debank', 'Acceptance criteria': ''})
            summary_data.append({'Section': 'DAM - Overview - Address', 'Data': 'Protocol', 'Data Source': '', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': 'Value', 'Data Source': '1. Debank', 'Acceptance criteria': 'within 2% diff'})

            # Excel - Overview - Address section
            summary_data.append({'Section': '', 'Data': 'Token', 'Data Source': '1. Debank', 'Acceptance criteria': ''})
            summary_data.append({'Section': 'Excel - Overview - Address', 'Data': 'Protocol', 'Data Source': 'TBC', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': 'Value', 'Data Source': '1. Debank', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Portfolio Share', 'Data Source': 'Formula calculation', 'Acceptance criteria': ''})

            # DAM - Holdings - Token section
            summary_data.append({'Section': '', 'Data': 'Amount', 'Data Source': '1. Debank/ Protocol - Balance inside Contract Address\n2. I will double check Etherscan when amount has diff > 2%', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': 'DAM - Holdings - Token', 'Data': 'Price', 'Data Source': '1. Coingecko\n2. Debank', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Price change', 'Data Source': '1. Coingecko\n2. show 0 for those token which is not listed on Coingecko', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': 'Value', 'Data Source': '1. Debank', 'Acceptance criteria': 'within 2% diff'})

            # DAM - Holdings - Chain section
            summary_data.append({'Section': '', 'Data': 'Portfolio Share', 'Data Source': 'Formula calculation', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': 'Amount', 'Data Source': '1. Debank/ Protocol - Balance inside Contract Address\n2. I will double check Etherscan when amount has diff > 2%', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': 'DAM - Holdings - Chain', 'Data': 'Price', 'Data Source': '1. Coingecko\n2. Debank', 'Acceptance criteria': 'within 2% diff'})
            summary_data.append({'Section': '', 'Data': 'Price change', 'Data Source': '1. Coingecko\n2. show 0 for those token which is not listed on Coingecko', 'Acceptance criteria': ''})
            summary_data.append({'Section': '', 'Data': 'Value', 'Data Source': '1. Debank', 'Acceptance criteria': 'within 2% diff'})

            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Google Drive n Summary', index=False)
            print(f"   ✅ Exported Google Drive n Summary sheet")

            # Auto-adjust column widths
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Post-processing: Add screenshots and format table titles
        print(f"\n🎨 Post-processing Excel file...")
        wb = load_workbook(output_path)

        # Format table titles and headers as bold in sheets with Table_Title column
        sheets_with_titles = ['DAM - Overview - Chain']
        for sheet_name in sheets_with_titles:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Find rows where first column (Table_Title) is not empty (table titles)
                # Also find header rows (where first column is empty but second column has "Address", etc.)
                for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
                    first_cell = row[0]

                    # Table title row: First cell has table name (e.g., "Wallet", "Morpho")
                    if first_cell.value and first_cell.value != "":
                        # Make the first cell bold (table title)
                        first_cell.font = Font(bold=True, size=12)
                        first_cell.alignment = Alignment(horizontal='left', vertical='center')

                    # Header row: First cell is empty but other cells have column names
                    # Check if this looks like a header row (has "Address", "Token", "Price", etc.)
                    elif len(row) > 1 and first_cell.value == "":
                        # Check if second cell contains column header text
                        second_cell = row[1] if len(row) > 1 else None
                        if second_cell and second_cell.value in ["Address", "Protocol Name", "Token", "Price", "Amount"]:
                            # This is a header row - make all cells bold
                            for cell in row:
                                if cell.value:  # Only format non-empty cells
                                    cell.font = Font(bold=True, size=11)
                                    cell.alignment = Alignment(horizontal='center', vertical='center')

                print(f"   ✅ Formatted table titles and headers in '{sheet_name}' as bold")

        # Highlight validation columns in yellow for Overview - Token and Holdings - Token
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # DAM - Overview - Token validation columns
        if 'DAM - Overview - Token' in wb.sheetnames:
            ws = wb['DAM - Overview - Token']
            # Find validation column indices
            validation_columns = ['Price Validation', 'Price Diff', 'Price Change Validation', 'Price Change Diff',
                                'Amount Validation', 'Amount Diff', 'Share Validation', 'Value Validation', 'Value Diff', 'Remark']

            # Get header row (row 1)
            header_row = ws[1]
            validation_col_indices = []

            for idx, cell in enumerate(header_row, start=1):
                if cell.value in validation_columns:
                    validation_col_indices.append(idx)

            # Apply yellow fill to validation columns (entire column including header)
            for col_idx in validation_col_indices:
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row_idx in range(1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

            if validation_col_indices:
                print(f"   ✅ Highlighted {len(validation_col_indices)} validation columns in yellow for 'DAM - Overview - Token'")

        # DAM - Holdings - Token validation columns
        if 'DAM - Holdings - Token' in wb.sheetnames:
            ws = wb['DAM - Holdings - Token']
            # Find validation column indices
            validation_columns = ['Portfolio Share Validation', 'Amount Validation', 'Amount Diff', 'Price Validation', 'Price Diff',
                                'Price Change Validation', 'Price Change Diff', 'Value Validation', 'Value Diff', 'Remark']

            # Get header row (row 1)
            header_row = ws[1]
            validation_col_indices = []

            for idx, cell in enumerate(header_row, start=1):
                if cell.value in validation_columns:
                    validation_col_indices.append(idx)

            # Apply yellow fill to validation columns (entire column including header)
            for col_idx in validation_col_indices:
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row_idx in range(1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

            if validation_col_indices:
                print(f"   ✅ Highlighted {len(validation_col_indices)} validation columns in yellow for 'DAM - Holdings - Token'")

        # DAM - Holdings - Chain validation columns
        if 'DAM - Holdings - Chain' in wb.sheetnames:
            ws = wb['DAM - Holdings - Chain']
            # Find validation column indices
            validation_columns = ['Share Validation', 'Amount Validation', 'Amount Diff', 'Price Validation', 'Price Diff',
                                'Price Change Validation', 'Price Change Diff', 'Value Validation', 'Value Diff', 'Remark']

            # Get header row (row 1)
            header_row = ws[1]
            validation_col_indices = []

            for idx, cell in enumerate(header_row, start=1):
                if cell.value in validation_columns:
                    validation_col_indices.append(idx)

            # Apply yellow fill to validation columns (entire column including header)
            for col_idx in validation_col_indices:
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row_idx in range(1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

            if validation_col_indices:
                print(f"   ✅ Highlighted {len(validation_col_indices)} validation columns in yellow for 'DAM - Holdings - Chain'")

        # DAM - Overview - Address validation columns
        if 'DAM - Overview - Address' in wb.sheetnames:
            ws = wb['DAM - Overview - Address']
            # Find validation column indices
            validation_columns = ['Token Validation', 'Value Validation']

            # Get header row (row 1)
            header_row = ws[1]
            validation_col_indices = []

            for idx, cell in enumerate(header_row, start=1):
                if cell.value in validation_columns:
                    validation_col_indices.append(idx)

            # Apply yellow fill to validation columns (entire column including header)
            for col_idx in validation_col_indices:
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row_idx in range(1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

            if validation_col_indices:
                print(f"   ✅ Highlighted {len(validation_col_indices)} validation columns in yellow for 'DAM - Overview - Address'")

        # DAM - Overview - Chain validation columns
        if 'DAM - Overview - Chain' in wb.sheetnames:
            ws = wb['DAM - Overview - Chain']
            # Find validation column indices
            validation_columns = ['Share Validation', 'Token Validation', 'Value Validation', 'Value Diff', 'Current Value Validation']

            # Get header row (row 1)
            header_row = ws[1]
            validation_col_indices = []

            for idx, cell in enumerate(header_row, start=1):
                if cell.value in validation_columns:
                    validation_col_indices.append(idx)

            # Apply yellow fill to validation columns (entire column including header)
            for col_idx in validation_col_indices:
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row_idx in range(1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

            if validation_col_indices:
                print(f"   ✅ Highlighted {len(validation_col_indices)} validation columns in yellow for 'DAM - Overview - Chain'")

        # Format Google Drive n Summary sheet
        if 'Google Drive n Summary' in wb.sheetnames:
            ws = wb['Google Drive n Summary']

            # Make header rows bold (rows 1, 5, 7)
            for row_idx in [1, 5, 7]:
                for cell in ws[row_idx]:
                    if cell.value:
                        cell.font = Font(bold=True)

            # Highlight "TBC" cells in yellow
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.value == 'TBC':
                        cell.fill = yellow_fill

            # Highlight the long explanation text in yellow (Excel - Overview - Token, Amount row)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                acceptance_criteria_cell = row[3]  # Column D (Acceptance criteria)
                if acceptance_criteria_cell.value and isinstance(acceptance_criteria_cell.value, str) and 'Etherscan' in acceptance_criteria_cell.value:
                    acceptance_criteria_cell.fill = yellow_fill

            print(f"   ✅ Formatted Google Drive n Summary sheet")

        # Add screenshots to respective tabs if provided
        if screenshot_paths:
            print(f"\n📸 Adding screenshots to Excel tabs...")
            for sheet_name, screenshot_path in screenshot_paths.items():
                if sheet_name in wb.sheetnames and os.path.exists(screenshot_path):
                    ws = wb[sheet_name]
                    # Insert image starting from a column far to the right (e.g., column P = 16)
                    img = OpenpyxlImage(screenshot_path)
                    # Scale image to fit better (adjust scale as needed)
                    img.width = img.width * 0.5  # 50% of original width
                    img.height = img.height * 0.5  # 50% of original height
                    # Anchor image to cell P1 (column 16, row 1)
                    ws.add_image(img, 'P1')
                    print(f"   ✅ Added screenshot to '{sheet_name}' tab")
                elif sheet_name not in wb.sheetnames:
                    print(f"   ⚠️  Sheet '{sheet_name}' not found in workbook")
                elif not os.path.exists(screenshot_path):
                    print(f"   ⚠️  Screenshot file not found: {screenshot_path}")

        # Save the modified workbook
        try:
            wb.save(output_path)
            print(f"✅ Excel file saved with formatting and screenshots: {output_path}")

            # Verify file was created successfully
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                print(f"   📄 File size: {file_size:,} bytes")

                # Note about Google Sheets compatibility
                print(f"\n💡 Note: For Google Sheets compatibility:")
                print(f"   1. Upload the .xlsx file to Google Drive")
                print(f"   2. Right-click and select 'Open with' → 'Google Sheets'")
                print(f"   3. Yellow highlighting may not appear (Google Sheets limitation)")
                print(f"   4. Images/screenshots may not appear (Google Sheets limitation)")

                wb.close()
                return output_path
            else:
                raise Exception(f"File was not created at {output_path}")
        except Exception as save_error:
            print(f"❌ Error saving workbook: {save_error}")
            # Try to close workbook even on error
            try:
                wb.close()
            except:
                pass
            raise

    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        raise
