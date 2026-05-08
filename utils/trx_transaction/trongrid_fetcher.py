#!/usr/bin/env python3
"""
Step 3: TronGrid API — fetch transactions, parse, and produce Excel.

Standalone usage (from automationv2/ root):
  python3 -m utils.trx_transaction.trongrid_fetcher <TRX_ADDRESS> <DDMMYYYY>
  python3 -m utils.trx_transaction.trongrid_fetcher <TRX_ADDRESS> <FROM_DDMMYYYY> <TO_DDMMYYYY>

When imported by the orchestrator, call fetch_and_parse() / build_trongrid_excel().
"""

import os, sys, time, requests
from datetime import datetime, timezone
import openpyxl
from openpyxl.utils import get_column_letter

import utils.trx_transaction.shared as shared
from utils.trx_transaction.shared import (
    _is_trx_address, _parse_date, _date_to_ts_ms, ts_to_utc,
    decode_addr, header_style, data_style,
)


# ── TronGrid API fetch ───────────────────────────────────────────────────────
def fetch_trongrid_transactions():
    """Fetch all transactions with pagination."""
    all_txs = []
    url = (f"https://api.trongrid.io/v1/accounts/{shared.ADDRESS}/transactions"
           f"?min_timestamp={shared.TS_FROM}&max_timestamp={shared.TS_TO}"
           f"&limit=200&order_by=block_timestamp,asc")
    headers = {"Accept": "application/json"}
    page = 0
    while url:
        print(f"   Fetching TronGrid txs page {page+1}... URL: {url[:100]}")
        try:
            r = requests.get(url, headers=headers, timeout=30)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"   ⚠️ Fetch error: {e}")
            break
        batch = data.get("data", [])
        all_txs.extend(batch)
        meta = data.get("meta", {})
        fp   = meta.get("fingerprint")
        page += 1
        if fp and len(batch) > 0:
            url = (f"https://api.trongrid.io/v1/accounts/{shared.ADDRESS}/transactions"
                   f"?min_timestamp={shared.TS_FROM}&max_timestamp={shared.TS_TO}"
                   f"&limit=200&order_by=block_timestamp,asc&fingerprint={fp}")
        else:
            url = None
        time.sleep(0.3)
    print(f"   ✅ Total TronGrid transactions: {len(all_txs)}")
    return all_txs


def fetch_trongrid_trc20():
    """Fetch all TRC20 transfers with pagination."""
    all_transfers = []
    url = (f"https://api.trongrid.io/v1/accounts/{shared.ADDRESS}/transactions/trc20"
           f"?min_timestamp={shared.TS_FROM}&max_timestamp={shared.TS_TO}"
           f"&limit=200&order_by=block_timestamp,asc")
    headers = {"Accept": "application/json"}
    page = 0
    while url:
        print(f"   Fetching TRC20 page {page+1}...")
        try:
            r = requests.get(url, headers=headers, timeout=30)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"   ⚠️ TRC20 fetch error: {e}")
            break
        batch = data.get("data", [])
        all_transfers.extend(batch)
        meta = data.get("meta", {})
        fp   = meta.get("fingerprint")
        page += 1
        if fp and len(batch) > 0:
            url = (f"https://api.trongrid.io/v1/accounts/{shared.ADDRESS}/transactions/trc20"
                   f"?min_timestamp={shared.TS_FROM}&max_timestamp={shared.TS_TO}"
                   f"&limit=200&order_by=block_timestamp,asc&fingerprint={fp}")
        else:
            url = None
        time.sleep(0.3)
    print(f"   ✅ Total TRC20 transfers: {len(all_transfers)}")
    return all_transfers


# ── Parse ────────────────────────────────────────────────────────────────────
def parse_tx(tx, trc20_map):
    """Parse a raw TronGrid transaction into a structured dict."""
    txid   = tx.get("txID", "")
    ts     = tx.get("block_timestamp", 0)
    dt     = ts_to_utc(ts)
    ret    = tx.get("ret", [{}])[0]
    fee    = ret.get("fee", 0)
    status = ret.get("contractRet", "SUCCESS")

    contracts = tx.get("raw_data", {}).get("contract", [])
    if not contracts:
        return {
            "trx_hash": txid, "date_time": dt, "tx_type": "Unknown",
            "from_addr": "", "to_addr": "", "amount": "",
            "resources_fee": f"Fee: {fee/1_000_000:.6f} TRX | Status: {status}",
            "token_transfer": "", "net_transfer": "",
        }

    c     = contracts[0]
    ctype = c.get("type", "")
    val   = c.get("parameter", {}).get("value", {})

    owner    = decode_addr(val.get("owner_address", ""))
    to       = decode_addr(val.get("to_address", ""))
    recv     = decode_addr(val.get("receiver_address", ""))
    caddr    = decode_addr(val.get("contract_address", ""))
    amount   = val.get("amount", val.get("frozen_balance",
               val.get("unfreeze_balance", val.get("balance", 0))))
    resource = val.get("resource", "BANDWIDTH")

    fee_str = f"{fee/1_000_000:.6f} TRX ({fee:,} SUN)"
    if status != "SUCCESS":
        fee_str += f" | Status: {status}"

    if ctype == "TransferContract":
        trx_amount = amount / 1_000_000
        is_out = owner == shared.ADDRESS
        tx_type = "TRX Transfer (Send)" if is_out else "TRX Transfer (Receive)"
        net = f"{'-' if is_out else '+'}{trx_amount:,.6f} TRX"
        return {
            "trx_hash": txid, "date_time": dt, "tx_type": tx_type,
            "from_addr": owner, "to_addr": to, "amount": f"{trx_amount:,.6f} TRX",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": net,
        }
    elif ctype == "TriggerSmartContract":
        t20 = trc20_map.get(txid)
        if t20:
            sym     = t20.get("token_info", {}).get("symbol", "?")
            name    = t20.get("token_info", {}).get("name", "?")
            dec     = t20.get("token_info", {}).get("decimals", 6)
            t20from = t20.get("from", "")
            t20to   = t20.get("to", "")
            raw_val = int(t20.get("value", 0))
            adj_val = raw_val / (10 ** dec)
            is_out  = t20from == shared.ADDRESS
            net     = f"{'-' if is_out else '+'}{adj_val:,.{min(dec,8)}f} {sym}"
            token_str = (f"{sym} ({name}): {adj_val:,.{min(dec,8)}f}\n"
                         f"From: {t20from}\nTo: {t20to}\n"
                         f"Contract: {t20.get('token_info',{}).get('address','')}")
            return {
                "trx_hash": txid, "date_time": dt,
                "tx_type": f"TRC20 Transfer ({sym})",
                "from_addr": t20from, "to_addr": t20to,
                "amount": "0 TRX (token transfer)",
                "resources_fee": f"Energy+Bandwidth Fee: {fee_str}",
                "token_transfer": token_str, "net_transfer": net,
            }
        else:
            call_val = val.get("call_value", 0)
            trx_c = call_val / 1_000_000 if call_val else 0
            label = "Smart Contract Call" if status == "SUCCESS" else f"Smart Contract ({status})"
            return {
                "trx_hash": txid, "date_time": dt, "tx_type": label,
                "from_addr": owner, "to_addr": caddr,
                "amount": f"{trx_c:.6f} TRX" if trx_c else "0 TRX",
                "resources_fee": f"Energy+Bandwidth Fee: {fee_str}",
                "token_transfer": "-", "net_transfer": f"Fee: -{fee/1_000_000:.6f} TRX",
            }
    elif ctype == "FreezeBalanceV2Contract":
        amt = amount / 1_000_000
        return {
            "trx_hash": txid, "date_time": dt,
            "tx_type": f"Stake/Freeze ({resource})",
            "from_addr": owner, "to_addr": owner,
            "amount": f"Freeze: {amt:,.6f} TRX ({resource})",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": f"Staked: {amt:,.6f} TRX",
        }
    elif ctype == "UnfreezeBalanceV2Contract":
        amt = amount / 1_000_000
        return {
            "trx_hash": txid, "date_time": dt,
            "tx_type": f"Unstake/Unfreeze ({resource})",
            "from_addr": owner, "to_addr": owner,
            "amount": f"Unfreeze: {amt:,.6f} TRX ({resource})",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": f"+{amt:,.6f} TRX (unfrozen)",
        }
    elif ctype == "DelegateResourceContract":
        bal = val.get("balance", 0) / 1_000_000
        return {
            "trx_hash": txid, "date_time": dt,
            "tx_type": f"Delegate Resource ({resource})",
            "from_addr": owner, "to_addr": recv,
            "amount": f"Delegate: {bal:,.6f} TRX ({resource})",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": f"Delegated {bal:,.6f} TRX {resource}",
        }
    elif ctype == "UnDelegateResourceContract":
        bal = val.get("balance", 0) / 1_000_000
        return {
            "trx_hash": txid, "date_time": dt,
            "tx_type": f"Undelegate Resource ({resource})",
            "from_addr": owner, "to_addr": recv,
            "amount": f"Undelegate: {bal:,.6f} TRX ({resource})",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": f"Undelegated {bal:,.6f} TRX {resource}",
        }
    elif ctype == "VoteWitnessContract":
        votes = val.get("votes", [])
        vote_str = ", ".join(f"{v.get('vote_address','?')}:{v.get('vote_count',0)}"
                             for v in votes[:3])
        return {
            "trx_hash": txid, "date_time": dt, "tx_type": "Vote Witness",
            "from_addr": owner, "to_addr": vote_str[:80],
            "amount": f"{sum(v.get('vote_count',0) for v in votes)} votes",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": "Vote cast",
        }
    elif ctype == "WithdrawExpireUnfreezeContract":
        return {
            "trx_hash": txid, "date_time": dt, "tx_type": "Withdraw Unstaked TRX",
            "from_addr": owner, "to_addr": owner, "amount": "Unstaked TRX withdrawal",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": "+TRX withdrawn from stake",
        }
    elif ctype == "WithdrawBalanceContract":
        return {
            "trx_hash": txid, "date_time": dt, "tx_type": "Claim Voting Reward",
            "from_addr": owner, "to_addr": owner, "amount": "Voting reward claim",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": "+TRX reward claimed",
        }
    else:
        return {
            "trx_hash": txid, "date_time": dt, "tx_type": ctype,
            "from_addr": owner, "to_addr": to or recv or caddr,
            "amount": f"{amount/1_000_000:,.6f} TRX" if amount else "-",
            "resources_fee": f"Fee: {fee_str}",
            "token_transfer": "-", "net_transfer": "-",
        }


# ── Convenience: fetch + parse in one call ───────────────────────────────────
def fetch_and_parse():
    """Fetch TronGrid data and return (parsed_txs, raw_txs, trc20_list)."""
    raw_txs    = fetch_trongrid_transactions()
    trc20_list = fetch_trongrid_trc20()
    trc20_map  = {t["transaction_id"]: t for t in trc20_list}
    parsed_txs = [parse_tx(tx, trc20_map) for tx in raw_txs]
    return parsed_txs, raw_txs, trc20_list


# ── Build standalone Step 3 Excel ────────────────────────────────────────────
def build_trongrid_excel(parsed_txs, raw_txs, trc20_list):
    """Build Step 3 Excel with TronGrid Transactions + TRC20 sheets."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "TronGrid Transactions"
    hdrs = ["#", "Trx Hash", "Date/Time (UTC)", "Transaction Type",
            "From", "To", "Amount", "Resources Consumed & Fee",
            "Token Transfer", "Net Transfer"]
    ws.append(hdrs)
    header_style(ws, 1, "1F4E79")
    ws.freeze_panes = "A2"

    for i, tx in enumerate(parsed_txs, 1):
        ws.append([i, tx["trx_hash"], tx["date_time"], tx["tx_type"],
                   tx["from_addr"], tx["to_addr"], tx["amount"],
                   tx["resources_fee"], tx["token_transfer"], tx["net_transfer"]])
        data_style(ws, ws.max_row, i % 2 == 0)

    for i, w in enumerate([4, 68, 26, 32, 40, 40, 28, 55, 60, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("TRC20 Transfers")
    ws2.append(["#", "Trx Hash", "Date/Time (UTC)", "Token Symbol", "Token Name",
                "From", "To", "Value (Raw)", "Value (Adjusted)", "Contract"])
    header_style(ws2, 1, "375623")
    ws2.freeze_panes = "A2"
    for i, t in enumerate(trc20_list, 1):
        sym  = t.get("token_info", {}).get("symbol", "?")
        name = t.get("token_info", {}).get("name", "?")
        dec  = t.get("token_info", {}).get("decimals", 6)
        raw  = int(t.get("value", 0))
        adj  = raw / (10 ** dec)
        ws2.append([i, t.get("transaction_id", ""), ts_to_utc(t.get("block_timestamp", 0)),
                    sym, name, t.get("from", ""), t.get("to", ""),
                    str(raw), f"{adj:,.{min(dec,8)}f} {sym}",
                    t.get("token_info", {}).get("address", "")])
        data_style(ws2, ws2.max_row, i % 2 == 0)
    for i, w in enumerate([4, 68, 26, 12, 25, 40, 40, 30, 30, 40], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(shared.EXCEL_DIR, exist_ok=True)
    path = (f"{shared.EXCEL_DIR}/Step3_TronGrid_{shared.ADDRESS[-8:]}"
            f"_{shared.DATE_FROM}_to_{shared.DATE_TO}_{ts}.xlsx")
    wb.save(path)
    print(f"   ✅ Step 3 Excel saved: {path}")
    return path


# ── Standalone CLI ───────────────────────────────────────────────────────────
def _cli_main():
    if len(sys.argv) < 3:
        print("Usage:")
        print("  python3 -m utils.trx_transaction.trongrid_fetcher <TRX_ADDRESS> <DDMMYYYY>")
        print("  python3 -m utils.trx_transaction.trongrid_fetcher <TRX_ADDRESS> <FROM> <TO>")
        sys.exit(1)

    shared.ADDRESS = sys.argv[1]
    if not _is_trx_address(shared.ADDRESS):
        print(f"❌ '{shared.ADDRESS}' is not a valid TRX address (must start with T, 34 chars)")
        sys.exit(1)

    date1_str = sys.argv[2]
    date2_str = sys.argv[3] if len(sys.argv) > 3 else None

    _, _, _, dt_from = _parse_date(date1_str)
    if date2_str:
        _, _, _, dt_to = _parse_date(date2_str)
    else:
        dt_to = dt_from.replace(hour=23, minute=59, second=59)

    shared.DATE_FROM = dt_from.strftime("%Y-%m-%d")
    shared.DATE_TO   = dt_to.strftime("%Y-%m-%d")
    shared.TS_FROM   = _date_to_ts_ms(dt_from)
    shared.TS_TO     = _date_to_ts_ms(dt_to.replace(hour=23, minute=59, second=59))

    local_tz = datetime.now().astimezone().tzinfo
    utc_from = datetime.fromtimestamp(shared.TS_FROM / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    utc_to   = datetime.fromtimestamp(shared.TS_TO / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    print("=" * 70)
    print("Step 3: TronGrid Transaction Extraction")
    print(f"Address : {shared.ADDRESS}")
    print(f"Range   : {shared.DATE_FROM} to {shared.DATE_TO}")
    print(f"Timezone: {local_tz} (detected from system)")
    print(f"UTC range: {utc_from} → {utc_to}")
    print("=" * 70)

    parsed_txs, raw_txs, trc20_list = fetch_and_parse()
    print(f"\n[Step 3] Parsed {len(parsed_txs)} transactions")
    path = build_trongrid_excel(parsed_txs, raw_txs, trc20_list)

    print("\n" + "=" * 70)
    print(f"DONE — {path}")
    print("=" * 70)


if __name__ == "__main__":
    _cli_main()
