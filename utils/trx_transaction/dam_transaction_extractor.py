#!/usr/bin/env python3
"""
DAM Transaction Extraction Script v3 (Merged)
- Supports portfolio name lookup from DAM dropdown (from v2)
- Supports direct portfolio ID navigation (from v2)
- Correct DAM date picker handling: Today → Between → D/M/YYYY (from jan21)
- Credential file fallback (from v2)
- Full CLI argument support (from v2)
- Detailed transaction parsing with DAM column mapping (from jan21)
- Styled Excel output with openpyxl (from jan21)

Usage:
  python3 dam_transaction_extractor.py                            # Run with default portfolio ID
  python3 dam_transaction_extractor.py trx2_Mkx                   # Search by portfolio name
  python3 dam_transaction_extractor.py -p "My Portfolio"           # Search by portfolio name
  python3 dam_transaction_extractor.py --id 8724c50c-...           # Direct portfolio ID
  python3 dam_transaction_extractor.py --date 2026-03-15           # Specify target date
  python3 dam_transaction_extractor.py -p "trx2_Mkx" --date 2026-01-21 --xlsx
"""

import os
import re
import json
import sys
import argparse
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Defaults ──────────────────────────────────────────────────────────────────
TARGET_ADDRESS = "TWd4WrZ9wn84f5x1hZhL4DHvk738ns5jwb"
TARGET_DATE = "2026-01-21"
EMAIL = "roninx688@gmail.com"
PASSWORD = "787193@PyBt7871"
BASE_URL = "https://dam-sit.mqbc21.com"
PORTFOLIO_ID = "8724c50c-d46e-415b-8fd8-42e2fc6b2334"
PORTFOLIO_NAME = None  # Set via CLI argument

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SS_DIR = os.path.join(_SCRIPT_DIR, "test-results", "screenshots")
RESULT_DIR = os.path.join(_SCRIPT_DIR, "test-results")

os.makedirs(SS_DIR, exist_ok=True)
os.makedirs(RESULT_DIR, exist_ok=True)


# ── Helpers ───────────────────────────────────────────────────────────────────
def ss(page, name):
    """Take a full-page screenshot."""
    path = os.path.join(SS_DIR, f"{name}.png")
    page.screenshot(path=path, full_page=True)
    print(f"   📸 {name}.png")
    return path


def save_html(page, name):
    """Save current page HTML for debugging."""
    path = os.path.join(RESULT_DIR, f"{name}.html")
    with open(path, "w") as f:
        f.write(page.content())
    print(f"   💾 HTML: {name}.html")
    return path


# ── Main extraction logic ────────────────────────────────────────────────────
def run():
    """
    Run the full DAM transaction extraction flow.
    Returns: (all_tx_data, row_texts, row_count)
    """
    all_tx_data = []
    row_texts = []

    # Load credentials from tc1_account.json if available (v2 feature)
    # Project root = automationv2/
    _project_root = os.path.normpath(os.path.join(_SCRIPT_DIR, "..", ".."))
    _tc1_path = os.path.join(_project_root, "test_data", "tc1_account.json")
    if os.path.exists(_tc1_path):
        with open(_tc1_path) as _f:
            _acc = json.load(_f)
        email = _acc["email"]
        password = _acc["password"]
        print(f"   📂 Loaded credentials from tc1_account.json")
    else:
        email = EMAIL
        password = PASSWORD

    with sync_playwright() as p:
        HEADLESS = os.environ.get('HEADLESS', 'true').lower() == 'true'
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=400)
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        page = context.new_page()
        page.set_default_timeout(30000)

        # ── STEP 1: Sign in ──────────────────────────────────────────────────
        print(f"\n[1] Signing in to DAM as {email}...")
        page.goto(f"{BASE_URL}/sign-in")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(2000)

        page.fill('input[data-testid="input-email"]', email)
        page.fill('input[data-testid="input-password"]', password)
        ss(page, "01_before_signin")
        page.click('button[data-testid="sign-in-btn"]')

        try:
            page.wait_for_url("**/portfolio**", timeout=25000)
            print("   ✅ Signed in, redirected to portfolio")
        except PlaywrightTimeoutError:
            ss(page, "01_signin_error")
            print("   ❌ Sign-in failed")
            browser.close()
            return [], [], 0

        page.wait_for_timeout(2000)
        ss(page, "02_after_signin")

        # ── STEP 2: Navigate to portfolio (v2 name-lookup + ID) ──────────────
        portfolio_id = PORTFOLIO_ID

        if PORTFOLIO_NAME:
            portfolio_id = _navigate_by_portfolio_name(page)
            if portfolio_id is None:
                browser.close()
                return [], [], 0

        if not portfolio_id:
            print("\n[2] ❌ No portfolio name or ID specified")
            browser.close()
            return [], [], 0

        tx_url = f"{BASE_URL}/portfolio?portfolioId={portfolio_id}&tab=transactions"
        print(f"\n[2] Navigating to transactions: {tx_url}")
        page.goto(tx_url)
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(3000)

        # ── Wait for loading overlay to disappear ────────────────────────────
        loading_overlay = page.locator("div.absolute.top-0.left-0.w-full.h-full.z-50")
        try:
            loading_overlay.wait_for(state="hidden", timeout=60000)
            print("   ✅ Loading overlay dismissed")
        except Exception:
            # Fallback: broader selector for any backdrop-blur overlay
            try:
                page.locator("[class*='backdrop-blur']").wait_for(state="hidden", timeout=30000)
                print("   ✅ Loading overlay dismissed (fallback)")
            except Exception:
                print("   ⚠️ Loading overlay may still be present, continuing anyway")
        page.wait_for_timeout(1000)

        ss(page, "03_transactions_tab")
        save_html(page, "dam_transactions_before_filter")

        # ── STEP 3: Apply date filter (jan21 date-picker logic) ──────────────
        _apply_date_filter(page)

        # ── STEP 4: Read transaction rows ────────────────────────────────────
        print(f"\n[6] Reading transactions...")

        best_rows, best_count, best_sel = _find_transaction_rows(page)
        print(f"   Best selector: '{best_sel}' → {best_count} rows")

        if best_rows and best_count > 0:
            for i in range(best_count):
                row = best_rows.nth(i)
                try:
                    txt = row.text_content() or ""
                    row_texts.append(txt.strip())
                    print(f"   Row {i+1}: {txt[:150].strip()}")
                except Exception:
                    pass

        # Check if target date is present
        _check_target_date(row_texts)

        # ── STEP 5: Get detailed info by clicking each row ───────────────────
        print(f"\n[7] Getting transaction details...")

        if best_rows and best_count > 0:
            for i in range(best_count):
                tx_info = _extract_row_detail(page, best_rows, i)
                all_tx_data.append(tx_info)

        ss(page, "09_final_state")
        browser.close()

    return all_tx_data, row_texts, best_count


# ── Portfolio name lookup (from v2) ──────────────────────────────────────────
def _navigate_by_portfolio_name(page):
    """Search for portfolio by name in the DAM dropdown. Returns portfolio ID or None."""
    print(f"\n[2] Searching for portfolio: '{PORTFOLIO_NAME}'...")

    # Open portfolio dropdown
    portfolio_dropdown = page.locator('button:has-text("Portfolio")').first
    if portfolio_dropdown.count() > 0:
        portfolio_dropdown.click()
        page.wait_for_timeout(2000)
    else:
        page.mouse.click(395, 141)
        page.wait_for_timeout(2000)

    # Scroll dropdown to load all items
    page.mouse.move(490, 450)
    for _ in range(15):
        page.mouse.wheel(0, 300)
        page.wait_for_timeout(100)
    page.wait_for_timeout(500)

    # Find portfolio by name
    found = False
    try:
        match = page.get_by_text(PORTFOLIO_NAME, exact=True)
        if match.count() > 0:
            for i in range(match.count()):
                elem = match.nth(i)
                if elem.is_visible():
                    elem.click()
                    found = True
                    print(f"   ✅ Found and clicked portfolio: '{PORTFOLIO_NAME}'")
                    break
    except Exception:
        pass

    if not found:
        # Fallback: search all visible divs
        try:
            all_divs = page.locator('div').all()
            for div in all_divs:
                try:
                    if not div.is_visible():
                        continue
                    div_text = div.text_content().strip()
                    if not div_text:
                        continue
                    first_line = div_text.split('\n')[0].strip()
                    if first_line.lower() == PORTFOLIO_NAME.lower():
                        if 'Addresses' in div_text or 'Exchange' in div_text or 'Wallet' in div_text:
                            div.click()
                            found = True
                            print(f"   ✅ Found portfolio (fallback): '{first_line}'")
                            break
                except Exception:
                    pass
        except Exception:
            pass

    if not found:
        print(f"   ❌ Portfolio '{PORTFOLIO_NAME}' not found in dropdown")
        ss(page, "02_portfolio_not_found")
        return None

    page.wait_for_timeout(3000)

    # Extract portfolio ID from URL
    current_url = page.url
    id_match = re.search(r'portfolioId=([a-f0-9\-]+)', current_url)
    if id_match:
        portfolio_id = id_match.group(1)
        print(f"   📋 Portfolio ID: {portfolio_id}")
        return portfolio_id
    else:
        print(f"   ⚠️  Could not extract portfolio ID from URL: {current_url}")
        return None


# ── Date filter (jan21 correct date-picker logic) ────────────────────────────
def _apply_date_filter(page):
    """
    Apply the DAM date filter using the correct UI flow:
    1. Click the date popover trigger (Today/Week/Month/Custom button)
    2. Select "Between" option
    3. Fill D / M / YYYY inputs for both From and To
    4. Click "Set" to apply
    """
    # Parse target date
    dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    date_values = {"D": str(dt.day), "M": str(dt.month), "YYYY": str(dt.year)}

    # ── Wait for any loading overlay to clear before interacting ─────────
    try:
        overlay = page.locator("div.absolute.top-0.left-0.w-full.h-full.z-50")
        if overlay.count() > 0 and overlay.first.is_visible(timeout=1000):
            print("   ⏳ Waiting for loading overlay to clear...")
            overlay.first.wait_for(state="hidden", timeout=60000)
            print("   ✅ Overlay cleared")
    except Exception:
        pass  # No overlay or already gone

    # ── Step 3a: Open the date picker popover ─────────────────────────────
    print(f"\n[3] Opening date picker...")
    today_btn = page.locator(
        "button[aria-haspopup='dialog']:has-text('Today'), "
        "button[aria-haspopup='dialog']:has-text('Week'), "
        "button[aria-haspopup='dialog']:has-text('Month'), "
        "button[aria-haspopup='dialog']:has-text('Custom')"
    )
    try:
        today_btn.first.wait_for(state="visible", timeout=8000)
        today_btn.first.click()
        page.wait_for_timeout(1500)
        ss(page, "03_date_picker_opened")
        print("   ✅ Date picker opened")
    except Exception as e:
        print(f"   ⚠️ Date popover trigger not found: {e}")
        for sel in ["button:has-text('Today')", "button:has-text('Date')", "[class*='date']"]:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=2000):
                    btn.click()
                    page.wait_for_timeout(1500)
                    break
            except Exception:
                pass
        ss(page, "03_date_picker_attempt")

    # ── Step 3b: Click "Between" option ───────────────────────────────────
    print(f"\n[4] Selecting 'Between' (custom) date range for {TARGET_DATE}...")
    page.wait_for_timeout(500)

    between_clicked = False
    for sel in [
        "button:has-text('Between')",
        "[role='option']:has-text('Between')",
        "div[class*='cursor']:has-text('Between')",
        "li:has-text('Between')",
    ]:
        try:
            el = page.locator(sel).last
            if el.is_visible(timeout=2000):
                el.click()
                page.wait_for_timeout(1000)
                between_clicked = True
                print(f"   ✅ Clicked 'Between' with {sel}")
                break
        except Exception:
            pass

    if not between_clicked:
        print("   ⚠️ 'Between' option not found, may already be selected")

    ss(page, "04_between_selected")

    # ── Step 3c: Fill D / M / YYYY inputs ─────────────────────────────────
    print(f"\n[5] Filling date inputs (D/M/YYYY format)...")
    page.wait_for_timeout(500)

    all_inputs = page.locator("input").all()
    visible_date_inputs = []
    for inp in all_inputs:
        try:
            if inp.is_visible(timeout=300):
                ph = inp.get_attribute("placeholder") or ""
                if ph in ["D", "M", "YYYY"]:
                    visible_date_inputs.append((ph, inp))
        except Exception:
            pass

    print(f"   Found {len(visible_date_inputs)} D/M/YYYY inputs: {[ph for ph, _ in visible_date_inputs]}")

    # Fill both From and To with the same date (expect 6 inputs: D M YYYY D M YYYY)
    if len(visible_date_inputs) >= 6:
        for i, (ph, inp) in enumerate(visible_date_inputs):
            val = date_values.get(ph, "")
            if val:
                inp.click(click_count=3)
                inp.fill(val)
                page.wait_for_timeout(200)
                print(f"   Input {i} ({ph}) = {val}")
    elif len(visible_date_inputs) >= 3:
        # Only "From" inputs visible
        for ph, inp in visible_date_inputs[:3]:
            val = date_values.get(ph, "")
            if val:
                inp.click(click_count=3)
                inp.fill(val)
                page.wait_for_timeout(200)

    ss(page, "05_dates_filled")

    # ── Step 3d: Click "Set" / "Apply" button ─────────────────────────────
    set_clicked = False
    for apply_text in ["Set", "Apply", "Confirm", "Search", "OK"]:
        try:
            btn = page.locator(f"button:has-text('{apply_text}')").last
            if btn.is_visible(timeout=1000):
                btn.click()
                set_clicked = True
                print(f"   ✅ Clicked '{apply_text}'")
                page.wait_for_timeout(3000)
                break
        except Exception:
            pass

    if not set_clicked:
        print("   ⚠️ No Set/Apply button found")

    page.wait_for_timeout(3000)
    ss(page, "06_filter_applied")
    save_html(page, "dam_transactions_after_filter")
    print(f"   Current URL: {page.url}")


# ── Find transaction rows ────────────────────────────────────────────────────
def _find_transaction_rows(page):
    """Locate the best set of transaction rows on the page."""
    best_rows = None
    best_count = 0
    best_sel = None
    for sel in [
        "table tbody tr",
        "[class*='transaction-row']",
        "[class*='tx-row']",
        "[class*='transaction-item']",
        "tr[class*='hover']",
        "div[role='row']",
    ]:
        try:
            rows = page.locator(sel)
            count = rows.count()
            if count > best_count:
                best_count = count
                best_rows = rows
                best_sel = sel
        except Exception:
            pass
    return best_rows, best_count, best_sel


# ── Check target date presence ───────────────────────────────────────────────
def _check_target_date(row_texts):
    """Log whether the target date appears in the extracted row texts."""
    dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    # Build multiple date format patterns to check
    patterns = [
        dt.strftime("%m/%d/%Y"),   # 01/21/2026
        TARGET_DATE,                # 2026-01-21
        dt.strftime("%b %d"),       # Jan 21
    ]
    has_target = any(p in t for t in row_texts for p in patterns)
    print(f"   Target date ({TARGET_DATE}) in rows: {has_target}")
    if not has_target and row_texts:
        print(f"   ⚠️ Date filter may not be applied. First row: {row_texts[0][:100]}")
    return has_target


# ── Extract detail from a single row (jan21 detailed parsing) ────────────────
def _extract_row_detail(page, best_rows, i):
    """Click a transaction row, capture the detail panel, and parse it."""
    tx_info = {
        "row_index": i + 1,
        "trx_hash": "",
        "date_time": "",
        "transaction_type": "",
        "from_addr": "",
        "to_addr": "",
        "amount": "",
        "token": "",
        "usd_value": "",
        "resources_fee": "",
        "token_transfer": "",
        "net_transfer": "",
        "raw_cells": [],
        "detail_text": "",
    }

    try:
        row = best_rows.nth(i)

        # Get cell data from row
        cells = row.locator("td")
        cell_count = cells.count()
        for j in range(cell_count):
            try:
                tx_info["raw_cells"].append(cells.nth(j).text_content().strip())
            except Exception:
                tx_info["raw_cells"].append("")

        # Parse row text for hash
        row_text = row.text_content() or ""
        hash_match = re.search(r'[0-9a-fA-F]{64}', row_text)
        if hash_match:
            tx_info["trx_hash"] = hash_match.group(0)
        else:
            trunc_match = re.search(r'([0-9a-fA-F]{6,})\.\.\.([0-9a-fA-F]{4,})', row_text)
            if trunc_match:
                tx_info["trx_hash"] = f"{trunc_match.group(1)}...{trunc_match.group(2)}"

        # Click row to open detail panel
        row.click()
        page.wait_for_timeout(2000)
        ss(page, f"07_row_{i+1}_detail")

        # Try to capture the detail panel
        detail_text = ""
        for d_sel in [
            "[role='dialog']", "[class*='detail']", "[class*='modal']",
            "[class*='drawer']", "[class*='panel'][class*='open']",
            "[class*='sheet']", "[data-state='open']",
        ]:
            try:
                detail_el = page.locator(d_sel).first
                if detail_el.is_visible(timeout=2000):
                    detail_text = detail_el.text_content().strip()
                    print(f"   Row {i+1} detail ({d_sel}): {detail_text[:200]}")
                    break
            except Exception:
                pass

        if not detail_text:
            detail_text = row_text

        tx_info["detail_text"] = detail_text

        # Parse detail for structured data (jan21 parser)
        parse_detail(tx_info, detail_text, row_text)

        # Close detail
        page.keyboard.press("Escape")
        page.wait_for_timeout(800)
        for close_sel in ["button[aria-label*='close' i]", "button:has-text('×')", "[data-testid*='close']"]:
            try:
                close_btn = page.locator(close_sel).first
                if close_btn.is_visible(timeout=500):
                    close_btn.click()
                    page.wait_for_timeout(500)
                    break
            except Exception:
                pass

    except Exception as e:
        print(f"   Row {i+1} error: {e}")

    return tx_info


# ── Transaction detail parser (from jan21 — most complete) ───────────────────
def parse_detail(tx_info, detail_text, row_text):
    """Parse transaction detail text into structured fields."""
    combined = detail_text or row_text

    # Extract full 64-char hash
    hash_match = re.search(r'\b([0-9a-fA-F]{64})\b', combined)
    if hash_match:
        tx_info["trx_hash"] = hash_match.group(1)

    # Extract TRON addresses (start with T, 34 chars)
    tron_addrs = re.findall(r'\bT[a-zA-Z0-9]{33}\b', combined)
    if len(tron_addrs) >= 1:
        tx_info["from_addr"] = tron_addrs[0]
    if len(tron_addrs) >= 2:
        tx_info["to_addr"] = tron_addrs[1]

    # Parse line by line
    lines = [l.strip() for l in combined.replace('\t', '\n').split('\n') if l.strip()]
    for line in lines:
        ll = line.lower()
        if any(x in ll for x in ["txid", "tx id", "hash", "transaction id"]):
            h = re.search(r'[0-9a-fA-F]{64}', line)
            if h:
                tx_info["trx_hash"] = h.group(0)
        elif "type" in ll and not tx_info["transaction_type"]:
            tx_info["transaction_type"] = line
        elif "from" in ll and not tx_info["from_addr"]:
            addr = re.search(r'T[a-zA-Z0-9]{33}', line)
            if addr:
                tx_info["from_addr"] = addr.group(0)
        elif "to" in ll and "token" not in ll and not tx_info["to_addr"]:
            addr = re.search(r'T[a-zA-Z0-9]{33}', line)
            if addr:
                tx_info["to_addr"] = addr.group(0)
        elif "amount" in ll and not tx_info["amount"]:
            tx_info["amount"] = line
        elif "fee" in ll or "resource" in ll or "bandwidth" in ll or "energy" in ll:
            if tx_info["resources_fee"]:
                tx_info["resources_fee"] += " | " + line
            else:
                tx_info["resources_fee"] = line
        elif "token" in ll and not tx_info["token_transfer"]:
            tx_info["token_transfer"] = line
        elif re.search(r'\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2}', line):
            if not tx_info["date_time"]:
                tx_info["date_time"] = line
        elif re.search(r'(transfer|operational|receive|send|unfreeze|stake|approve)', ll):
            if not tx_info["transaction_type"]:
                tx_info["transaction_type"] = line

    # Fallback: use raw_cells data
    # DAM columns: Chain | Hash | Method | Time | From | To | Amount | Token | Value | Details
    cells = tx_info.get("raw_cells", [])
    if cells:
        if len(cells) > 1 and not tx_info["trx_hash"]:
            tx_info["trx_hash"] = cells[1]
        if len(cells) > 2 and not tx_info["transaction_type"]:
            tx_info["transaction_type"] = cells[2].split('\n')[0] if '\n' in cells[2] else cells[2]
        if len(cells) > 3 and not tx_info["date_time"]:
            time_match = re.search(r'\d{2}/\d{2}/\d{4}', cells[2] + cells[3] if len(cells) > 3 else cells[2])
            if time_match:
                tx_info["date_time"] = time_match.group(0)
        if len(cells) > 4 and not tx_info["from_addr"]:
            tx_info["from_addr"] = cells[4]
        if len(cells) > 5 and not tx_info["to_addr"]:
            tx_info["to_addr"] = cells[5]
        if len(cells) > 6 and not tx_info["amount"]:
            tx_info["amount"] = cells[6]
        if len(cells) > 7 and not tx_info["token"]:
            tx_info["token"] = cells[7]
        if len(cells) > 8:
            tx_info["usd_value"] = cells[8]


# ── Excel output (from jan21) ────────────────────────────────────────────────
def build_dam_excel(all_tx_data, row_texts, target_date=None):
    """Generate a styled DAM Excel file."""
    if not HAS_OPENPYXL:
        print("   ⚠️ openpyxl not installed — skipping Excel output. Install with: pip install openpyxl")
        return None

    date_label = target_date or TARGET_DATE
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"DAM Transactions {date_label}"

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Trx Hash", "Transaction Type", "From", "To",
        "Amount", "Resources Consumed & Fee", "Token Transfer", "Net Transfer",
        "Date/Time", "USD Value", "Raw Cells",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    # Row fill colors
    fill_light = PatternFill("solid", fgColor="DDEEFF")
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    for idx, tx in enumerate(all_tx_data):
        row = [
            tx.get("trx_hash", ""),
            tx.get("transaction_type", ""),
            tx.get("from_addr", ""),
            tx.get("to_addr", ""),
            tx.get("amount", ""),
            tx.get("resources_fee", ""),
            tx.get("token_transfer", ""),
            tx.get("net_transfer", ""),
            tx.get("date_time", ""),
            tx.get("usd_value", ""),
            " | ".join(tx.get("raw_cells", [])),
        ]
        ws.append(row)
        fill = fill_light if idx % 2 == 0 else fill_white
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Add raw rows sheet
    ws2 = wb.create_sheet("Raw Rows (Unfiltered)")
    ws2.append(["Row #", "Raw Text"])
    for i, txt in enumerate(row_texts, 1):
        ws2.append([i, txt])

    # Column widths
    col_widths = [70, 25, 40, 40, 20, 50, 50, 40, 20, 15, 80]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = os.path.join(RESULT_DIR, f"DAM_Transactions_{date_label}_{ts}.xlsx")
    wb.save(out_path)
    print(f"\n   ✅ DAM Excel saved: {out_path}")
    return out_path


# ── CLI entry point (from v2) ────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="DAM Transaction Extraction v3 (Merged)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 dam_transaction_extractor.py                                # Default portfolio ID
  python3 dam_transaction_extractor.py trx2_Mkx                       # Search by portfolio name
  python3 dam_transaction_extractor.py -p "My Portfolio"               # Search by portfolio name
  python3 dam_transaction_extractor.py --id 8724c50c-...               # Direct portfolio ID
  python3 dam_transaction_extractor.py --date 2026-03-15               # Specify target date
  python3 dam_transaction_extractor.py -p "trx2_Mkx" --date 2026-01-21 --xlsx
        """
    )
    parser.add_argument('name', nargs='?', default=None,
                        help='Portfolio name to search in DAM dropdown')
    parser.add_argument('-p', '--portfolio', type=str, default=None,
                        help='Portfolio name to search in DAM dropdown')
    parser.add_argument('--id', type=str, default=None,
                        help='Direct portfolio ID (skip dropdown search)')
    parser.add_argument('--date', type=str, default=None,
                        help='Target date filter (YYYY-MM-DD)')
    parser.add_argument('--xlsx', action='store_true', default=False,
                        help='Generate styled Excel output (requires openpyxl)')

    args = parser.parse_args()

    # Resolve portfolio name or ID
    if args.portfolio:
        PORTFOLIO_NAME = args.portfolio
        PORTFOLIO_ID = None
    elif args.name:
        PORTFOLIO_NAME = args.name
        PORTFOLIO_ID = None
    elif args.id:
        PORTFOLIO_ID = args.id

    if args.date:
        TARGET_DATE = args.date

    label = PORTFOLIO_NAME or f"ID:{PORTFOLIO_ID}"
    print("=" * 70)
    print(f"DAM Transaction Extraction v3 (Merged)")
    print(f"Portfolio: {label}")
    print(f"Date:      {TARGET_DATE}")
    print("=" * 70)

    all_tx_data, row_texts, count = run()

    print(f"\nCaptured {len(all_tx_data)} transactions (total rows in table: {count})")

    # Save raw JSON
    raw_path = os.path.join(RESULT_DIR, "dam_transactions_detailed.json")
    with open(raw_path, "w") as f:
        json.dump({"transactions": all_tx_data, "row_texts": row_texts}, f, indent=2, default=str)
    print(f"✅ Raw data saved: {raw_path}")

    # Build Excel if requested
    if args.xlsx:
        build_dam_excel(all_tx_data, row_texts, target_date=TARGET_DATE)
    else:
        print("   (Add --xlsx flag to generate Excel output)")

    for i, tx in enumerate(all_tx_data, 1):
        print(f"\n  [{i}] hash={tx.get('trx_hash','')[:20]}... type={tx.get('transaction_type','')} "
              f"from={tx.get('from_addr','')[:10]}... to={tx.get('to_addr','')[:10]}...")

    print("\nDone.")