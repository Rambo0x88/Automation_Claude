#!/usr/bin/env python3
"""
DAM Transaction Extraction - 2026-01-21
Address: TWd4WrZ9wn84f5x1hZhL4DHvk738ns5jwb
"""

import os
import re
import json
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TARGET_ADDRESS = "TWd4WrZ9wn84f5x1hZhL4DHvk738ns5jwb"
TARGET_DATE    = "2026-01-21"
EMAIL          = "roninx688@gmail.com"
PASSWORD       = "787193@PyBt7871"
BASE_URL       = "https://dam-sit.mqbc21.com"
PORTFOLIO_ID   = "8724c50c-d46e-415b-8fd8-42e2fc6b2334"
OUT_DIR        = os.path.dirname(os.path.abspath(__file__))
SS_DIR         = os.path.join(OUT_DIR, "test-results", "screenshots", "jan21")

os.makedirs(SS_DIR, exist_ok=True)


def ss(page, name):
    path = f"{SS_DIR}/{name}.png"
    page.screenshot(path=path, full_page=True)
    print(f"   📸 {name}.png")
    return path


def run():
    transactions = []

    with sync_playwright() as p:
        HEADLESS = os.environ.get('HEADLESS', 'false').lower() == 'true'
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=500)
        context = browser.new_context(viewport={"width": 1440, "height": 900})
        page = context.new_page()
        page.set_default_timeout(30000)

        # ── STEP 1: Sign in ──────────────────────────────────────────────────
        print("\n[1] Signing in...")
        page.goto(f"{BASE_URL}/sign-in")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(2000)

        page.fill('input[data-testid="input-email"]', EMAIL)
        page.fill('input[data-testid="input-password"]', PASSWORD)
        page.click('button[data-testid="sign-in-btn"]')

        try:
            page.wait_for_url("**/portfolio**", timeout=25000)
            print("   ✅ Signed in")
        except PlaywrightTimeoutError:
            ss(page, "01_signin_error")
            print("   ❌ Sign-in failed")
            browser.close()
            return []

        page.wait_for_timeout(2000)
        ss(page, "01_after_signin")

        # ── STEP 2: Navigate to Transactions tab ─────────────────────────────
        tx_url = f"{BASE_URL}/portfolio?portfolioId={PORTFOLIO_ID}&tab=transactions"
        print(f"\n[2] Going to transactions: {tx_url}")
        page.goto(tx_url)
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(3000)
        ss(page, "02_transactions_loaded")

        # ── STEP 3: Open the date picker (Today button) ───────────────────────
        print(f"\n[3] Opening date picker...")

        # Click the "Today" button (popover trigger)
        today_btn = page.locator("button[aria-haspopup='dialog']:has-text('Today'), button[aria-haspopup='dialog']:has-text('Week'), button[aria-haspopup='dialog']:has-text('Month'), button[aria-haspopup='dialog']:has-text('Custom')")
        try:
            today_btn.first.wait_for(state="visible", timeout=8000)
            today_btn.first.click()
            page.wait_for_timeout(1500)
            ss(page, "03_date_picker_opened")
            print("   ✅ Date picker opened")
        except Exception as e:
            print(f"   ⚠️ Today button not found: {e}")
            # Try clicking any date-related button
            for sel in ["button:has-text('Today')", "button:has-text('Date')", "[class*='date']"]:
                try:
                    btn = page.locator(sel).first
                    if btn.is_visible(timeout=2000):
                        btn.click()
                        page.wait_for_timeout(1500)
                        break
                except:
                    pass
            ss(page, "03_date_picker_attempt")

        # ── STEP 4: Click "Between" option ───────────────────────────────────
        print(f"\n[4] Selecting 'Between' (custom) date range for {TARGET_DATE}...")
        page.wait_for_timeout(500)

        # The date picker shows: Today/Yesterday/Last 7 days/.../Between
        # "Between" opens the date range inputs
        between_clicked = False
        for sel in [
            "button:has-text('Between')",
            "[role='option']:has-text('Between')",
            "div[class*='cursor']:has-text('Between')",
            "li:has-text('Between')",
        ]:
            try:
                el = page.locator(sel).last  # click the last "Between" (sidebar option, not calendar label)
                if el.is_visible(timeout=2000):
                    el.click()
                    page.wait_for_timeout(1000)
                    between_clicked = True
                    print(f"   ✅ Clicked 'Between' with {sel}")
                    break
            except:
                pass

        if not between_clicked:
            print("   ⚠️ 'Between' option not found, may already be selected")

        ss(page, "04_between_selected")

        # ── STEP 5: Fill date inputs (D / M / YYYY format) ───────────────────
        print(f"\n[5] Filling date inputs (D/M/YYYY format)...")
        page.wait_for_timeout(500)

        # Find all visible inputs - expect 6: D M YYYY D M YYYY
        all_inputs = page.locator("input").all()
        visible_date_inputs = []
        for inp in all_inputs:
            try:
                if inp.is_visible(timeout=300):
                    ph = inp.get_attribute("placeholder") or ""
                    if ph in ["D", "M", "YYYY"]:
                        visible_date_inputs.append((ph, inp))
            except:
                pass

        print(f"   Found {len(visible_date_inputs)} D/M/YYYY inputs: {[ph for ph, _ in visible_date_inputs]}")

        # Fill: [D=21, M=1, YYYY=2026] x2
        # From date
        date_values = {"D": "21", "M": "1", "YYYY": "2026"}
        if len(visible_date_inputs) >= 6:
            for i, (ph, inp) in enumerate(visible_date_inputs):
                val = date_values.get(ph, "")
                if val:
                    inp.click(click_count=3)
                    inp.fill(val)
                    page.wait_for_timeout(200)
                    print(f"   Input {i} ({ph}) = {val}")
        elif len(visible_date_inputs) >= 3:
            # Only "From" inputs visible
            for ph, inp in visible_date_inputs[:3]:
                val = date_values.get(ph, "")
                if val:
                    inp.click(click_count=3)
                    inp.fill(val)
                    page.wait_for_timeout(200)

        ss(page, "05_dates_filled")

        # Click "Set" button to apply the date range
        set_clicked = False
        for apply_text in ["Set", "Apply", "Confirm", "Search", "OK"]:
            try:
                btn = page.locator(f"button:has-text('{apply_text}')").last
                if btn.is_visible(timeout=1000):
                    btn.click()
                    set_clicked = True
                    print(f"   ✅ Clicked '{apply_text}'")
                    page.wait_for_timeout(3000)
                    break
            except:
                pass

        if not set_clicked:
            print("   ⚠️ No Set/Apply button found")

        page.wait_for_timeout(3000)
        ss(page, "06_filter_applied")
        print(f"   Current URL: {page.url}")

        # ── STEP 6: Check the filter state & read rows ────────────────────────
        print(f"\n[6] Reading transactions...")

        # Check visible text
        body_text = page.locator("body").text_content() or ""
        print(f"   Visible transactions area: {body_text[body_text.find('Transactions'):body_text.find('Transactions')+500] if 'Transactions' in body_text else 'not found'}")

        # Look for table rows - prefer tbody rows to avoid calendar/header rows
        best_rows = None
        best_count = 0
        best_sel = None
        for sel in ["table tbody tr", "[class*='transaction-row']", "tr[class*='hover']", "div[role='row']"]:
            try:
                rows = page.locator(sel)
                count = rows.count()
                if count > best_count:
                    best_count = count
                    best_rows = rows
                    best_sel = sel
            except:
                pass

        print(f"   Best selector: '{best_sel}' → {best_count} rows")

        row_texts = []
        if best_rows and best_count > 0:
            for i in range(best_count):
                row = best_rows.nth(i)
                try:
                    txt = row.text_content() or ""
                    row_texts.append(txt.strip())
                    print(f"   Row {i+1}: {txt[:150].strip()}")
                except:
                    pass

        # Check if we have target date rows
        has_target = any("01/21/2026" in t or "2026-01-21" in t or "Jan 21" in t for t in row_texts)
        print(f"   Target date (Jan 21) in rows: {has_target}")

        if not has_target and best_count > 0:
            print("   ⚠️ Date filter may not be applied. Current data date:", row_texts[0][:100] if row_texts else "N/A")

        # ── STEP 7: Get detailed info by clicking each row ────────────────────
        print(f"\n[7] Getting transaction details...")
        all_tx_data = []

        if best_rows and best_count > 0:
            for i in range(best_count):
                tx_info = {
                    "row_index": i + 1,
                    "trx_hash": "",
                    "date_time": "",
                    "transaction_type": "",
                    "from_addr": "",
                    "to_addr": "",
                    "amount": "",
                    "token": "",
                    "usd_value": "",
                    "resources_fee": "",
                    "token_transfer": "",
                    "net_transfer": "",
                    "raw_cells": [],
                    "detail_text": "",
                }

                try:
                    row = best_rows.nth(i)
                    # Get cell data from row
                    cells = row.locator("td")
                    cell_count = cells.count()
                    for j in range(cell_count):
                        try:
                            tx_info["raw_cells"].append(cells.nth(j).text_content().strip())
                        except:
                            tx_info["raw_cells"].append("")

                    # Parse row text for addresses and hash
                    row_text = row.text_content() or ""

                    # Try to find hash (full or truncated)
                    hash_match = re.search(r'[0-9a-fA-F]{64}', row_text)
                    if hash_match:
                        tx_info["trx_hash"] = hash_match.group(0)
                    else:
                        # Truncated hash like "2a3809...c56742"
                        trunc_match = re.search(r'([0-9a-fA-F]{6,})\.\.\.([0-9a-fA-F]{4,})', row_text)
                        if trunc_match:
                            tx_info["trx_hash"] = f"{trunc_match.group(1)}...{trunc_match.group(2)}"

                    # Click row to open detail panel
                    row.click()
                    page.wait_for_timeout(2000)
                    ss(page, f"07_row_{i+1}_detail")

                    # Try to capture the detail panel
                    detail_text = ""
                    for d_sel in ["[role='dialog']", "[class*='detail']", "[class*='modal']",
                                  "[class*='drawer']", "[class*='panel'][class*='open']",
                                  "[class*='sheet']", "[data-state='open']"]:
                        try:
                            detail_el = page.locator(d_sel).first
                            if detail_el.is_visible(timeout=2000):
                                detail_text = detail_el.text_content().strip()
                                detail_html = detail_el.inner_html()
                                print(f"   Row {i+1} detail ({d_sel}): {detail_text[:200]}")
                                break
                        except:
                            pass

                    if not detail_text:
                        # Try to get the full page text diff
                        detail_text = row_text

                    tx_info["detail_text"] = detail_text

                    # Parse detail for structured data
                    parse_detail(tx_info, detail_text, row_text)

                    all_tx_data.append(tx_info)

                    # Close detail
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(800)
                    for close_sel in ["button[aria-label*='close' i]", "button:has-text('×')", "[data-testid*='close']"]:
                        try:
                            close_btn = page.locator(close_sel).first
                            if close_btn.is_visible(timeout=500):
                                close_btn.click()
                                page.wait_for_timeout(500)
                                break
                        except:
                            pass

                except Exception as e:
                    print(f"   Row {i+1} error: {e}")
                    all_tx_data.append(tx_info)

        # ── STEP 8: If no data for target date, try scrolling through pages ─────
        if not has_target:
            print(f"\n[8] Attempting to navigate to date {TARGET_DATE} via URL or page nav...")
            # Try different period options
            ss(page, "08_no_target_date")

        ss(page, "09_final_state")
        browser.close()

    return all_tx_data, row_texts, best_count


def parse_detail(tx_info, detail_text, row_text):
    """Parse transaction detail text into structured fields."""
    combined = detail_text or row_text

    # Extract full 64-char hash
    hash_match = re.search(r'\b([0-9a-fA-F]{64})\b', combined)
    if hash_match:
        tx_info["trx_hash"] = hash_match.group(1)

    # Extract TRON addresses (start with T, 34 chars)
    tron_addrs = re.findall(r'\bT[a-zA-Z0-9]{33}\b', combined)
    if len(tron_addrs) >= 1:
        tx_info["from_addr"] = tron_addrs[0]
    if len(tron_addrs) >= 2:
        tx_info["to_addr"] = tron_addrs[1]

    # Parse line by line
    lines = [l.strip() for l in combined.replace('\t', '\n').split('\n') if l.strip()]
    for line in lines:
        ll = line.lower()
        if any(x in ll for x in ["txid", "tx id", "hash", "transaction id"]):
            h = re.search(r'[0-9a-fA-F]{64}', line)
            if h:
                tx_info["trx_hash"] = h.group(0)
        elif "type" in ll and not tx_info["transaction_type"]:
            tx_info["transaction_type"] = line
        elif "from" in ll and not tx_info["from_addr"]:
            addr = re.search(r'T[a-zA-Z0-9]{33}', line)
            if addr:
                tx_info["from_addr"] = addr.group(0)
        elif "to" in ll and "token" not in ll and not tx_info["to_addr"]:
            addr = re.search(r'T[a-zA-Z0-9]{33}', line)
            if addr:
                tx_info["to_addr"] = addr.group(0)
        elif "amount" in ll and not tx_info["amount"]:
            tx_info["amount"] = line
        elif "fee" in ll or "resource" in ll or "bandwidth" in ll or "energy" in ll:
            if tx_info["resources_fee"]:
                tx_info["resources_fee"] += " | " + line
            else:
                tx_info["resources_fee"] = line
        elif "token" in ll and not tx_info["token_transfer"]:
            tx_info["token_transfer"] = line
        elif re.search(r'\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2}', line):
            if not tx_info["date_time"]:
                tx_info["date_time"] = line
        elif re.search(r'(transfer|operational|receive|send|unfreeze|stake|approve)', ll):
            if not tx_info["transaction_type"]:
                tx_info["transaction_type"] = line

    # Fallback: use raw_cells data
    cells = tx_info.get("raw_cells", [])
    if cells:
        # DAM columns from HTML: Chain | Hash | Method | Time | From | To | Amount | Token | Value | Details
        if len(cells) > 1 and not tx_info["trx_hash"]:
            tx_info["trx_hash"] = cells[1]
        if len(cells) > 2 and not tx_info["transaction_type"]:
            tx_info["transaction_type"] = cells[2].split('\n')[0] if '\n' in cells[2] else cells[2]
        if len(cells) > 3 and not tx_info["date_time"]:
            # Time might be combined with type in cell[2], or separate in cell[3]
            time_match = re.search(r'\d{2}/\d{2}/\d{4}', cells[2] + cells[3] if len(cells) > 3 else cells[2])
            if time_match:
                tx_info["date_time"] = time_match.group(0)
        if len(cells) > 4 and not tx_info["from_addr"]:
            tx_info["from_addr"] = cells[4]
        if len(cells) > 5 and not tx_info["to_addr"]:
            tx_info["to_addr"] = cells[5]
        if len(cells) > 6 and not tx_info["amount"]:
            tx_info["amount"] = cells[6]
        if len(cells) > 7 and not tx_info["token"]:
            tx_info["token"] = cells[7]
        if len(cells) > 8:
            tx_info["usd_value"] = cells[8]


def build_dam_excel(all_tx_data, row_texts):
    """Generate the DAM Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAM Transactions 2026-01-21"

    # Header style
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Trx Hash", "Transaction Type", "From", "To",
        "Amount", "Resources Consumed & Fee", "Token Transfer", "Net Transfer",
        "Date/Time", "USD Value", "Raw Cells"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    # Row fill colors
    fill_light = PatternFill("solid", fgColor="DDEEFF")
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    for idx, tx in enumerate(all_tx_data):
        row = [
            tx.get("trx_hash", ""),
            tx.get("transaction_type", ""),
            tx.get("from_addr", ""),
            tx.get("to_addr", ""),
            tx.get("amount", ""),
            tx.get("resources_fee", ""),
            tx.get("token_transfer", ""),
            tx.get("net_transfer", ""),
            tx.get("date_time", ""),
            tx.get("usd_value", ""),
            " | ".join(tx.get("raw_cells", [])),
        ]
        ws.append(row)
        fill = fill_light if idx % 2 == 0 else fill_white
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Add raw rows sheet
    ws2 = wb.create_sheet("Raw Rows (Unfiltered)")
    ws2.append(["Row #", "Raw Text"])
    for i, txt in enumerate(row_texts, 1):
        ws2.append([i, txt])

    # Column widths
    col_widths = [70, 25, 40, 40, 20, 50, 50, 40, 20, 15, 80]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = f"{OUT_DIR}/DAM_Transactions_2026-01-21_{ts}.xlsx"
    wb.save(out_path)
    print(f"\n   ✅ DAM Excel saved: {out_path}")
    return out_path


if __name__ == "__main__":
    print("=" * 70)
    print(f"DAM Transaction Extraction for {TARGET_DATE}")
    print(f"Address: {TARGET_ADDRESS}")
    print("=" * 70)

    result = run()
    if isinstance(result, tuple):
        all_tx_data, row_texts, count = result
    else:
        all_tx_data, row_texts, count = result, [], 0

    print(f"\nCaptured {len(all_tx_data)} transactions (total rows in table: {count})")

    # Save raw data
    raw_path = f"{OUT_DIR}/test-results/dam_jan21_raw.json"
    with open(raw_path, "w") as f:
        json.dump({"transactions": all_tx_data, "row_texts": row_texts}, f, indent=2)
    print(f"Raw data: {raw_path}")

    # Build Excel
    excel_path = build_dam_excel(all_tx_data, row_texts)

    print("\nDone.")
