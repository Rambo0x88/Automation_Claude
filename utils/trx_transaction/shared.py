#!/usr/bin/env python3
"""
Shared config, helpers, and Excel styling for TRX transaction pipeline.
"""

import os, json
from datetime import datetime, timezone, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Config (defaults — overridden by CLI args in orchestrator) ────────────────
_SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT  = os.path.normpath(os.path.join(_SCRIPT_DIR, "..", ".."))

ADDRESS        = ""
PORTFOLIO_ID   = None
PORTFOLIO_NAME = None
BASE_URL       = "https://dam-sit.mqbc21.com"
DATE_FROM      = ""
DATE_TO        = ""
TS_FROM        = 0
TS_TO          = 0
OUT_DIR        = _SCRIPT_DIR
EXCEL_DIR      = os.path.join(_PROJECT_ROOT, "test-results")
SS_DIR         = ""

# Credentials — loaded from tc1_account.json or fallback
EMAIL    = "roninx688@gmail.com"
PASSWORD = "787193@PyBt7871"
_tc1_path = os.path.join(_PROJECT_ROOT, "test_data", "tc1_account.json")
if os.path.exists(_tc1_path):
    with open(_tc1_path) as _f:
        _acc = json.load(_f)
    EMAIL    = _acc["email"]
    PASSWORD = _acc["password"]


# ── Helpers ──────────────────────────────────────────────────────────────────
def _is_trx_address(s):
    return s.startswith("T") and len(s) == 34 and s.isalnum()


def _parse_date(s):
    """Parse DDMMYYYY into (day, month, year) strings and a datetime object.
    Uses the system's local timezone so that date ranges match the user's wall clock.
    e.g. in Singapore (UTC+8), '16042026' → 2026-04-16 00:00:00+08:00
    """
    s = s.strip()
    if len(s) == 8 and s.isdigit():
        d, m, y = s[:2], s[2:4], s[4:]
        local_tz = datetime.now().astimezone().tzinfo
        dt = datetime(int(y), int(m), int(d), tzinfo=local_tz)
        return d.lstrip("0"), m.lstrip("0"), y, dt
    raise ValueError(f"Invalid date format '{s}'. Expected DDMMYYYY (e.g. 16042026)")


def _date_to_ts_ms(dt):
    return int(dt.timestamp() * 1000)


def ts_to_utc(ms):
    if not ms:
        return ""
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def hex_to_base58(hex_addr):
    """Convert 41... hex address to TRON base58 address."""
    try:
        import hashlib, base58
        addr_bytes = bytes.fromhex(hex_addr)
        checksum = hashlib.sha256(hashlib.sha256(addr_bytes).digest()).digest()[:4]
        return base58.b58encode(addr_bytes + checksum).decode()
    except Exception:
        return hex_addr


def decode_addr(hex_addr):
    """Try to decode hex TRON address to base58."""
    if not hex_addr:
        return ""
    if hex_addr.startswith("T") and len(hex_addr) == 34:
        return hex_addr
    if hex_addr.startswith("41") and len(hex_addr) == 42:
        try:
            return hex_to_base58(hex_addr)
        except Exception:
            pass
    return hex_addr


# ── Excel styling helpers ────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def header_style(ws, row, fill_hex, font_color="FFFFFF"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_color, size=9, name="Calibri")
    aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row]:
        cell.font  = font
        cell.fill  = fill
        cell.alignment = aln
        cell.border = thin_border()
    ws.row_dimensions[row].height = 22


def data_style(ws, row_num, alt=False):
    fill = PatternFill("solid", fgColor="EEF4FB" if alt else "FFFFFF")
    font = Font(size=8, name="Calibri")
    aln  = Alignment(vertical="top", wrap_text=True)
    for cell in ws[row_num]:
        cell.fill  = fill
        cell.font  = font
        cell.alignment = aln
