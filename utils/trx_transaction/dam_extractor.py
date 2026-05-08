#!/usr/bin/env python3
"""
Steps 4-7: DAM UI — login, find portfolio, filter by date, scrape transactions.

Standalone usage:
  python3 -m utils.trx_transaction.dam_extractor <PORTFOLIO_NAME> <DDMMYYYY>
  python3 -m utils.trx_transaction.dam_extractor <TRX_ADDRESS> <DDMMYYYY>

When imported by the orchestrator, call extract_dam_transactions() / build_dam_excel().
"""

import os, re, sys
from datetime import datetime, timezone
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.utils import get_column_letter

import utils.trx_transaction.shared as shared
from utils.trx_transaction.shared import (
    _is_trx_address, _parse_date, _date_to_ts_ms,
    header_style, data_style,
)


# ── Screenshot helper ────────────────────────────────────────────────────────
def ss(page, name):
    os.makedirs(shared.SS_DIR, exist_ok=True)
    p = f"{shared.SS_DIR}/{name}.png"
    page.screenshot(path=p, full_page=True)
    print(f"   📸 {name}.png → {p}")


# ── DAM Playwright extraction ────────────────────────────────────────────────
def extract_dam_transactions():
    """Login to DAM, find/create portfolio, filter by date range, capture all transactions.
    Returns (dam_rows, dam_raw) or (None, None) on failure.
    Updates shared.PORTFOLIO_ID and shared.ADDRESS as side effects.
    """
    dam_rows = []
    dam_raw  = []

    HEADLESS = os.environ.get('HEADLESS', 'true').lower() == 'true'
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=400)
        ctx     = browser.new_context(viewport={"width": 1920, "height": 1080})
        page    = ctx.new_page()
        page.set_default_timeout(30000)

        # ── Sign in ──────────────────────────────────────────────────────────
        print(f"\n   [DAM] Signing in as {shared.EMAIL}...")
        page.goto(f"{shared.BASE_URL}/sign-in")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(2000)
        page.locator('input[data-testid="input-email"]').wait_for(state="visible", timeout=15000)
        page.fill('input[data-testid="input-email"]', shared.EMAIL)
        page.wait_for_timeout(300)
        page.locator('input[data-testid="input-password"]').wait_for(state="visible", timeout=10000)
        page.fill('input[data-testid="input-password"]', shared.PASSWORD)
        page.wait_for_timeout(500)
        sign_in_btn = page.locator('button[data-testid="sign-in-btn"]')
        sign_in_btn.wait_for(state="visible", timeout=10000)
        for _ in range(20):
            try:
                if sign_in_btn.is_enabled():
                    break
            except: pass
            page.wait_for_timeout(200)
        sign_in_btn.click()

        signed_in = False
        try:
            page.wait_for_url("**/portfolio**", timeout=25000)
            signed_in = True
            print("   [DAM] ✅ Signed in")
        except PWTimeout:
            print("   [DAM] ⚠️ First sign-in attempt failed, retrying...")
            page.goto(f"{shared.BASE_URL}/sign-in")
            page.wait_for_timeout(2000)
            page.fill('input[data-testid="input-email"]', shared.EMAIL)
            page.fill('input[data-testid="input-password"]', shared.PASSWORD)
            page.click('button[data-testid="sign-in-btn"]')
            try:
                page.wait_for_url("**/portfolio**", timeout=25000)
                signed_in = True
                print("   [DAM] ✅ Signed in (retry)")
            except PWTimeout:
                pass

        if not signed_in:
            ss(page, "dam_signin_error")
            print("   [DAM] ❌ Sign-in FAILED after 2 attempts.")
            print(f"   [DAM]    Email: {shared.EMAIL}")
            print(f"   [DAM]    URL: {shared.BASE_URL}/sign-in")
            browser.close()
            return None, None
        page.wait_for_timeout(2000)

        # ── Find or create portfolio ─────────────────────────────────────────
        if shared.PORTFOLIO_ID:
            print(f"   [DAM] Using portfolio ID: {shared.PORTFOLIO_ID}")
        elif shared.PORTFOLIO_NAME or shared.ADDRESS:
            search_term = shared.PORTFOLIO_NAME or shared.ADDRESS
            is_address = _is_trx_address(search_term)
            print(f"   [DAM] Searching for {'address' if is_address else 'portfolio'}: {search_term}")

            dropdown = page.locator('button:has-text("Portfolio")').first
            if dropdown.count() > 0:
                dropdown.click()
            else:
                page.mouse.click(395, 141)
            page.wait_for_timeout(2000)

            page.mouse.move(490, 450)
            for _ in range(15):
                page.mouse.wheel(0, 300)
                page.wait_for_timeout(100)
            page.wait_for_timeout(500)

            found = False
            matched_elem = None

            if is_address:
                search_names = [search_term, search_term[-8:], f"Custom_{search_term[:8]}"]
            else:
                search_names = [search_term]

            for name in search_names:
                try:
                    match = page.get_by_text(name, exact=True)
                    if match.count() > 0:
                        for i in range(match.count()):
                            elem = match.nth(i)
                            if elem.is_visible():
                                matched_elem = elem
                                found = True
                                print(f"   [DAM] ✅ Found portfolio: '{name}'")
                                break
                    if found:
                        break
                except:
                    pass

            if not found:
                for name in search_names:
                    try:
                        for div in page.locator('div').all():
                            try:
                                if not div.is_visible():
                                    continue
                                text = div.text_content().strip()
                                first_line = text.split('\n')[0].strip()
                                if first_line.lower() == name.lower():
                                    if any(kw in text for kw in ['Addresses', 'Exchange', 'Wallet', 'Address']):
                                        matched_elem = div
                                        found = True
                                        print(f"   [DAM] ✅ Found portfolio (fallback): '{first_line}'")
                                        break
                            except:
                                pass
                        if found:
                            break
                    except:
                        pass

            if found and matched_elem:
                matched_elem.click()
                page.wait_for_timeout(3000)
                url = page.url
                id_match = re.search(r'portfolioId=([a-f0-9\-]+)', url)
                if id_match:
                    shared.PORTFOLIO_ID = id_match.group(1)
                    print(f"   [DAM] Portfolio ID: {shared.PORTFOLIO_ID}")
                else:
                    print(f"   [DAM] ⚠️ Could not extract portfolio ID from: {url}")

            elif is_address:
                print(f"   [DAM] Portfolio not found for address, creating new...")
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)

                dropdown = page.locator('button:has-text("Portfolio")').first
                if dropdown.count() > 0:
                    try:
                        dropdown.click(timeout=5000)
                    except:
                        page.mouse.click(395, 141)
                else:
                    page.mouse.click(395, 141)
                page.wait_for_timeout(2000)
                page.mouse.move(490, 450)
                for _ in range(15):
                    page.mouse.wheel(0, 300)
                    page.wait_for_timeout(100)
                page.wait_for_timeout(500)

                create_clicked = False
                for sel_fn in [
                    lambda: page.get_by_role("menuitem", name="Create portfolio", exact=True),
                    lambda: page.get_by_text("Create portfolio", exact=True),
                    lambda: page.get_by_text("Create portfolio"),
                    lambda: page.get_by_role("menuitem").filter(has_text="Create"),
                ]:
                    try:
                        btn = sel_fn()
                        if btn.count() > 0 and btn.first.is_visible():
                            btn.first.click()
                            create_clicked = True
                            break
                    except:
                        pass

                if not create_clicked:
                    print("   [DAM] ❌ Could not find Create portfolio button")
                    ss(page, "dam_create_failed")
                    browser.close()
                    return None, None

                page.wait_for_timeout(2000)

                portfolio_name = f"Custom_{shared.ADDRESS[:8]}"
                name_input = None
                for ph in ["Enter portfolio name", "Portfolio name", "Name"]:
                    candidate = page.get_by_placeholder(ph)
                    if candidate.count() > 0:
                        name_input = candidate
                        break
                if not name_input:
                    for sel in ['dialog input[type="text"]', 'form input[type="text"]', 'input[type="text"]:visible']:
                        candidate = page.locator(sel)
                        if candidate.count() > 0 and candidate.first.is_visible(timeout=2000):
                            name_input = candidate.first
                            break
                if name_input:
                    name_input.click()
                    name_input.fill(portfolio_name)
                    page.wait_for_timeout(500)
                    print(f"   [DAM] Portfolio name: {portfolio_name}")

                addr_field = None
                for sel in [
                    'input[name="wallet.0.address"]',
                    '[placeholder*="wallet address"]:not([disabled])',
                    '[placeholder*="Paste your wallet"]:not([disabled])',
                    'textarea:not([disabled])',
                ]:
                    candidate = page.locator(sel).first
                    if candidate.count() > 0 and candidate.is_visible(timeout=2000):
                        addr_field = candidate
                        break
                if addr_field:
                    addr_field.click()
                    addr_field.fill(shared.ADDRESS)
                    page.wait_for_timeout(500)
                    page.keyboard.press("Enter")
                    page.wait_for_timeout(2000)
                    print(f"   [DAM] Address added: {shared.ADDRESS[:12]}...")

                save_btn = page.locator('button:has-text("Save")').first
                try:
                    save_btn.wait_for(state="visible", timeout=10000)
                    page.wait_for_timeout(3000)
                    save_btn.click(timeout=15000)
                    page.wait_for_timeout(5000)
                    print("   [DAM] ✅ Portfolio created!")
                except:
                    try:
                        page.locator('button[type="submit"]').first.click(timeout=10000)
                        page.wait_for_timeout(5000)
                    except:
                        page.evaluate("document.querySelector('button[type=\"submit\"]')?.click()")
                        page.wait_for_timeout(5000)

                page.goto(f"{shared.BASE_URL}/portfolio")
                page.wait_for_timeout(3000)
                try:
                    link = page.get_by_text(portfolio_name, exact=True)
                    if link.count() > 0:
                        link.first.click()
                        page.wait_for_timeout(3000)
                except:
                    pass

                url = page.url
                id_match = re.search(r'portfolioId=([a-f0-9\-]+)', url)
                if id_match:
                    shared.PORTFOLIO_ID = id_match.group(1)
                    print(f"   [DAM] New portfolio ID: {shared.PORTFOLIO_ID}")
                else:
                    print(f"   [DAM] ⚠️ Could not get portfolio ID after creation")
                    ss(page, "dam_no_portfolio_id")
                    browser.close()
                    return None, None
            else:
                print(f"   [DAM] ❌ Portfolio '{search_term}' not found")
                ss(page, "dam_portfolio_not_found")
                browser.close()
                return None, None

        # ── Extract address from Combined Net Worth (for portfolio name mode) ─
        if not shared.ADDRESS and shared.PORTFOLIO_ID:
            overview_url = f"{shared.BASE_URL}/portfolio?portfolioId={shared.PORTFOLIO_ID}"
            print(f"   [DAM] Navigating to Overview to extract address...")
            page.goto(overview_url)
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(5000)

            try:
                found_addrs = page.evaluate('''() => {
                    const results = [];
                    const els = document.querySelectorAll('[data-tooltip-id^="address-display-tooltip-"]');
                    for (const el of els) {
                        let addr = "";
                        const hlEl = el.querySelector("[data-highlight-target]");
                        if (hlEl) addr = hlEl.getAttribute("data-highlight-target") || "";
                        if (!addr) {
                            const tid = el.getAttribute("data-tooltip-id") || "";
                            addr = tid.replace("address-display-tooltip-", "");
                        }
                        if (addr.trim()) results.push(addr.trim());
                    }
                    return results;
                }''')
                for addr in found_addrs:
                    if addr.startswith("T") and len(addr) == 34:
                        shared.ADDRESS = addr
                        print(f"   [DAM] ✅ Extracted TRX address from Combined Net Worth: {shared.ADDRESS}")
                        break
                if not shared.ADDRESS:
                    page_html = page.content()
                    tron_matches = re.findall(r'(T[A-Za-z0-9]{33})', page_html)
                    if tron_matches:
                        shared.ADDRESS = tron_matches[0]
                        print(f"   [DAM] ✅ Extracted TRX address from page HTML: {shared.ADDRESS}")
                if not shared.ADDRESS:
                    print(f"   [DAM] ⚠️ Could not extract TRX address from Combined Net Worth")
                    ss(page, "dam_no_address_found")
            except Exception as e:
                print(f"   [DAM] ⚠️ Error extracting address: {e}")

        # ── Navigate to Transactions tab ─────────────────────────────────────
        tx_url = f"{shared.BASE_URL}/portfolio?portfolioId={shared.PORTFOLIO_ID}&tab=transactions"
        print(f"   [DAM] Navigating to: {tx_url}")
        page.goto(tx_url)
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(3000)
        ss(page, "dam_01_transactions_loaded")

        # ── Open date picker ─────────────────────────────────────────────────
        print("   [DAM] Opening date picker...")

        _overlay_sel = "div.absolute.z-50.backdrop-blur-sm, div[class*='backdrop-blur']"
        try:
            _overlay = page.locator(_overlay_sel).first
            if _overlay.is_visible(timeout=2000):
                print("   [DAM] ⏳ Waiting for loading overlay to disappear...")
                _overlay.wait_for(state="hidden", timeout=30000)
                print("   [DAM] ✅ Loading overlay gone")
                page.wait_for_timeout(1000)
        except Exception:
            pass

        today_btn = page.locator(
            "button[aria-haspopup='dialog']:has-text('Today'),"
            "button[aria-haspopup='dialog']:has-text('Yesterday'),"
            "button[aria-haspopup='dialog']:has-text('Week'),"
            "button[aria-haspopup='dialog']:has-text('Month'),"
            "button[aria-haspopup='dialog']:has-text('Jan'),"
            "button[aria-haspopup='dialog']:has-text('Feb'),"
            "button[aria-haspopup='dialog']:has-text('Mar'),"
            "button[aria-haspopup='dialog']:has-text('Custom'),"
            "button[aria-haspopup='dialog']:has-text('Between')"
        )
        try:
            today_btn.first.wait_for(state="visible", timeout=8000)
            today_btn.first.click()
            page.wait_for_timeout(1500)
            print("   [DAM] ✅ Date picker opened")
        except Exception as e:
            print(f"   [DAM] ⚠️ Date picker click failed, retrying with force: {e}")
            try:
                today_btn.first.click(force=True)
                page.wait_for_timeout(1500)
                print("   [DAM] ✅ Date picker opened (force click)")
            except Exception as e2:
                print(f"   [DAM] ⚠️ Date picker open failed: {e2}")
        ss(page, "dam_02_date_picker")

        # ── Set date range from CLI args ────────────────────────────────────
        from_parts = shared.DATE_FROM.split("-")  # YYYY-MM-DD
        to_parts   = shared.DATE_TO.split("-")
        from_day, from_month, from_year = from_parts[2].lstrip("0"), from_parts[1].lstrip("0"), from_parts[0]
        to_day,   to_month,   to_year   = to_parts[2].lstrip("0"),   to_parts[1].lstrip("0"),   to_parts[0]

        def fill_date_inputs(from_d, from_m, from_y, to_d, to_m, to_y):
            all_inp = page.locator("input").all()
            dmyyyy  = []
            for inp in all_inp:
                try:
                    if inp.is_visible(timeout=200):
                        ph = inp.get_attribute("placeholder") or ""
                        if ph in ["D", "M", "YYYY"]:
                            dmyyyy.append((ph, inp))
                except: pass

            print(f"   [DAM] Found {len(dmyyyy)} D/M/YYYY inputs")
            vals_from = {"D": from_d, "M": from_m, "YYYY": from_y}
            vals_to   = {"D": to_d,   "M": to_m,   "YYYY": to_y}

            if len(dmyyyy) >= 6:
                for i, (ph, inp) in enumerate(dmyyyy):
                    vals = vals_from if i < 3 else vals_to
                    v = vals.get(ph, "")
                    if v:
                        inp.click(click_count=3)
                        page.wait_for_timeout(80)
                        inp.fill(v)
                        page.wait_for_timeout(150)
                        print(f"     Input {i} ({ph}) = {v}")
                return True
            elif len(dmyyyy) >= 3:
                for ph, inp in dmyyyy[:3]:
                    v = vals_from.get(ph, "")
                    if v:
                        inp.click(click_count=3)
                        page.wait_for_timeout(80)
                        inp.fill(v)
                        page.wait_for_timeout(150)
                return True
            return False

        filled = fill_date_inputs(from_day, from_month, from_year,
                                  to_day,   to_month,   to_year)
        if not filled:
            print("   [DAM] ⚠️ No D/M/YYYY inputs found — trying 'Between'")
            for sel in ["button:has-text('Between')", "[role='option']:has-text('Between')"]:
                try:
                    el = page.locator(sel).last
                    if el.is_visible(timeout=2000):
                        el.click(); page.wait_for_timeout(1000)
                        break
                except: pass
            filled = fill_date_inputs(from_day, from_month, from_year,
                                      to_day,   to_month,   to_year)

        if not filled:
            print("   [DAM] ⚠️ Still no date inputs — waiting for overlay and retrying...")
            try:
                _overlay = page.locator(_overlay_sel).first
                if _overlay.is_visible(timeout=1000):
                    _overlay.wait_for(state="hidden", timeout=60000)
                    page.wait_for_timeout(1000)
            except: pass
            try:
                today_btn.first.click(force=True)
                page.wait_for_timeout(1500)
                for sel in ["button:has-text('Between')", "[role='option']:has-text('Between')"]:
                    try:
                        el = page.locator(sel).last
                        if el.is_visible(timeout=2000):
                            el.click(); page.wait_for_timeout(1000)
                            break
                    except: pass
                filled = fill_date_inputs(from_day, from_month, from_year,
                                          to_day,   to_month,   to_year)
                if filled:
                    print("   [DAM] ✅ Date inputs filled on retry")
                else:
                    print("   [DAM] ❌ Could not fill date inputs after retry — results may show wrong date")
            except Exception as e:
                print(f"   [DAM] ❌ Date retry failed: {e}")

        ss(page, "dam_03_dates_filled")

        # ── Click Set ────────────────────────────────────────────────────────
        for apply_text in ["Set", "Apply", "Confirm", "OK"]:
            try:
                btn = page.locator(f"button:has-text('{apply_text}')").last
                if btn.is_visible(timeout=1000):
                    btn.click()
                    print(f"   [DAM] ✅ Clicked '{apply_text}'")
                    page.wait_for_timeout(3000)
                    break
            except: pass

        ss(page, "dam_04_filter_applied")
        print(f"   [DAM] URL: {page.url}")

        # ── Check for error and retry if needed ──────────────────────────────
        body_check = page.locator("body").text_content() or ""
        if "Something went wrong" in body_check:
            print("   [DAM] ⚠️ Error loading transactions — clicking Refresh...")
            try:
                page.locator("button:has-text('Refresh')").first.click()
                page.wait_for_timeout(3000)
            except: pass
            ss(page, "dam_04b_after_refresh")
            body_check = page.locator("body").text_content() or ""
            if "Something went wrong" in body_check:
                print("   [DAM] Still erroring — retrying date inputs...")
                try:
                    today_btn.first.click()
                    page.wait_for_timeout(1500)
                    fill_date_inputs(from_day, from_month, from_year,
                                     to_day, to_month, to_year)
                    ss(page, "dam_04c_retry_dates")
                    for apply_text in ["Set", "Apply", "Confirm", "OK"]:
                        try:
                            btn = page.locator(f"button:has-text('{apply_text}')").last
                            if btn.is_visible(timeout=800):
                                btn.click(); page.wait_for_timeout(3000); break
                        except: pass
                    ss(page, "dam_04d_retry_applied")
                except Exception as e:
                    print(f"   [DAM] Retry error: {e}")

        # ── Read all pages of transactions ───────────────────────────────────
        print("   [DAM] Reading transaction rows...")
        page_num      = 1
        all_dam_rows  = []
        seen_hashes   = set()
        max_pages     = 20

        while page_num <= max_pages:
            page.wait_for_timeout(1500)
            body_text = page.locator("body").text_content() or ""

            if "No transactions found" in body_text and page_num == 1:
                print("   [DAM] ⚠️ No transactions found with current filter")
                ss(page, f"dam_05_page{page_num}_no_data")
                break

            rows  = page.locator("tr[class*='hover']")
            count = rows.count()
            print(f"   [DAM] Page {page_num}: {count} rows")

            if count == 0:
                ss(page, f"dam_05_page{page_num}_empty")
                break

            page_hashes    = set()
            new_rows_added = 0
            for i in range(count):
                try:
                    row       = rows.nth(i)
                    cells     = row.locator("td")
                    cell_data = [cells.nth(j).text_content().strip()
                                 for j in range(cells.count())]
                    row_text  = row.text_content().strip()
                    row_key   = cell_data[1] if len(cell_data) > 1 else row_text[:30]
                    page_hashes.add(row_key)
                    if row_key not in seen_hashes:
                        all_dam_rows.append({"cells": cell_data, "text": row_text})
                        seen_hashes.add(row_key)
                        new_rows_added += 1
                    if i < 3 or i == count - 1:
                        print(f"     Row {i+1}: {row_text[:120]}")
                except Exception as e:
                    print(f"     Row {i+1} error: {e}")

            if page_num % 3 == 0 or new_rows_added == 0:
                ss(page, f"dam_05_page{page_num}_data")

            if new_rows_added == 0:
                print(f"   [DAM] No new rows on page {page_num} — end of data reached")
                break

            print(f"   [DAM] Page {page_num}: {new_rows_added} new unique rows (total so far: {len(all_dam_rows)})")

            total_pages = None
            try:
                m = re.search(r'(\d+)\s*/\s*(\d+)', body_text)
                if m:
                    cur_p  = int(m.group(1))
                    tot_p  = int(m.group(2))
                    total_pages = tot_p
                    if cur_p >= tot_p:
                        print(f"   [DAM] On last page ({cur_p}/{tot_p}) — stopping")
                        break
            except: pass

            navigated = False
            try:
                pg_input = page.locator("input#pagination-input, input[id*='pagination']").first
                if pg_input.is_visible(timeout=500):
                    current_val = int(pg_input.input_value() or "1")
                    if total_pages and current_val >= total_pages:
                        print(f"   [DAM] Pagination input at {current_val}/{total_pages} — done")
                        break
                    next_val = current_val + 1
                    pg_input.click(click_count=3)
                    pg_input.fill(str(next_val))
                    pg_input.press("Enter")
                    page.wait_for_timeout(2000)
                    page_num += 1
                    navigated = True
            except: pass

            if not navigated:
                for sel in [
                    "button[aria-label='Go to next page']",
                    "button[aria-label*='next' i]",
                    "nav button:last-child",
                    "[class*='pagination'] button:last-child",
                ]:
                    try:
                        btn = page.locator(sel).last
                        if btn.is_visible(timeout=400):
                            if (btn.get_attribute("disabled") is None and
                                    btn.get_attribute("aria-disabled") != "true"):
                                btn.click()
                                page.wait_for_timeout(2000)
                                page_num += 1
                                navigated = True
                                break
                    except: pass

            if not navigated:
                print(f"   [DAM] Pagination ended at page {page_num}")
                break

        # ── Parse rows into structured dicts ─────────────────────────────────
        print(f"\n   [DAM] Processing {len(all_dam_rows)} transaction rows...")
        dam_tx_list = []

        for idx, row_info in enumerate(all_dam_rows):
            cells = row_info.get("cells", [])
            text  = row_info.get("text", "")

            tx = {
                "row_index": idx + 1,
                "trx_hash": "", "date_time": "", "tx_type": "",
                "from_addr": "", "to_addr": "", "amount": "",
                "resources_fee": "", "token_transfer": "", "net_transfer": "",
                "raw_cells": cells, "detail_text": "",
            }

            non_empty = [c for c in cells if c.strip()]
            if len(non_empty) < 3:
                continue

            hash_idx = -1
            for ci, c in enumerate(cells):
                if re.search(r'[0-9a-fA-F]{6}', c) and ('...' in c or len(c) >= 20):
                    hash_idx = ci
                    break

            if hash_idx >= 0:
                tx["trx_hash"] = cells[hash_idx]

                type_time_idx = hash_idx + 1
                if type_time_idx < len(cells):
                    raw = cells[type_time_idx]
                    dt_match = re.search(r'(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})', raw)
                    if dt_match:
                        tx["date_time"] = dt_match.group(1)
                        tx["tx_type"] = raw[:dt_match.start()].strip()
                    else:
                        parts = raw.split('\n')
                        tx["tx_type"] = parts[0].strip()
                        tx["date_time"] = parts[-1].strip() if len(parts) > 1 else ""

                remaining_start = type_time_idx + 1
                remaining = cells[remaining_start:] if remaining_start < len(cells) else []
                if len(remaining) >= 1: tx["from_addr"] = remaining[0]
                if len(remaining) >= 2: tx["to_addr"] = remaining[1]
                if len(remaining) >= 3: tx["amount"] = remaining[2]
                if len(remaining) >= 4: tx["token_transfer"] = remaining[3]
                if len(remaining) >= 5: tx["net_transfer"] = remaining[4]
            else:
                if len(cells) >= 1: tx["trx_hash"] = cells[0]
                if len(cells) >= 2: tx["tx_type"] = cells[1]
                if len(cells) >= 3: tx["from_addr"] = cells[2]
                if len(cells) >= 4: tx["to_addr"] = cells[3]
                if len(cells) >= 5: tx["amount"] = cells[4]

            h = re.search(r'\b([0-9a-fA-F]{64})\b', text)
            if h:
                tx["trx_hash"] = h.group(1)

            dam_tx_list.append(tx)

        dam_rows = dam_tx_list
        dam_raw  = all_dam_rows

        ss(page, "dam_06_final")
        browser.close()

    return dam_rows, dam_raw


# ── Build standalone Step 7 Excel ────────────────────────────────────────────
def build_dam_excel(dam_rows, dam_raw):
    """Build Step 7 Excel with DAM Transactions sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAM Transactions"
    hdrs = ["#", "Trx Hash", "Date/Time (DAM)", "Transaction Type",
            "From", "To", "Amount", "Resources Consumed & Fee",
            "Token Transfer", "Net Transfer", "Raw Cell Data"]
    ws.append(hdrs)
    header_style(ws, 1, "375623")
    ws.freeze_panes = "A2"

    for i, tx in enumerate(dam_rows, 1):
        ws.append([
            i,
            tx.get("trx_hash", ""),
            tx.get("date_time", ""),
            tx.get("tx_type", ""),
            tx.get("from_addr", ""),
            tx.get("to_addr", ""),
            tx.get("amount", ""),
            tx.get("resources_fee", ""),
            tx.get("token_transfer", ""),
            tx.get("net_transfer", ""),
            " | ".join(tx.get("raw_cells", [])),
        ])
        data_style(ws, ws.max_row, i % 2 == 0)

    for i, w in enumerate([4, 68, 26, 28, 40, 40, 28, 40, 45, 35, 70], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(shared.EXCEL_DIR, exist_ok=True)
    path = f"{shared.EXCEL_DIR}/Step7_DAM_{shared.ADDRESS[-8:]}_{shared.DATE_FROM}_to_{shared.DATE_TO}_{ts}.xlsx"
    wb.save(path)
    print(f"   ✅ Step 7 Excel saved: {path}")
    return path


# ── Standalone CLI ───────────────────────────────────────────────────────────
def _cli_main():
    if len(sys.argv) < 3:
        print("Usage:")
        print("  python3 -m utils.trx_transaction.dam_extractor <PORTFOLIO_NAME_OR_ADDRESS> <DDMMYYYY>")
        print("  python3 -m utils.trx_transaction.dam_extractor <PORTFOLIO_NAME_OR_ADDRESS> <FROM> <TO>")
        sys.exit(1)

    input_target = sys.argv[1]
    date1_str    = sys.argv[2]
    date2_str    = sys.argv[3] if len(sys.argv) > 3 else None

    if _is_trx_address(input_target):
        shared.ADDRESS = input_target
        shared.PORTFOLIO_NAME = None
    else:
        shared.PORTFOLIO_NAME = input_target
        shared.ADDRESS = ""

    _, _, _, dt_from = _parse_date(date1_str)
    if date2_str:
        _, _, _, dt_to = _parse_date(date2_str)
    else:
        dt_to = dt_from.replace(hour=23, minute=59, second=59)

    shared.DATE_FROM = dt_from.strftime("%Y-%m-%d")
    shared.DATE_TO   = dt_to.strftime("%Y-%m-%d")
    shared.TS_FROM   = _date_to_ts_ms(dt_from)
    shared.TS_TO     = _date_to_ts_ms(dt_to.replace(hour=23, minute=59, second=59))

    label = shared.ADDRESS[-8:] if shared.ADDRESS else shared.PORTFOLIO_NAME
    shared.SS_DIR = f"{shared.OUT_DIR}/test-results/screenshots/{label}"
    os.makedirs(shared.SS_DIR, exist_ok=True)

    local_tz = datetime.now().astimezone().tzinfo
    print("=" * 70)
    print("Step 7: DAM Transaction Extraction")
    if shared.ADDRESS:
        print(f"Address  : {shared.ADDRESS}")
    if shared.PORTFOLIO_NAME:
        print(f"Portfolio: {shared.PORTFOLIO_NAME}")
    print(f"Range    : {shared.DATE_FROM} to {shared.DATE_TO}")
    print(f"Timezone : {local_tz} (detected from system)")
    print("=" * 70)

    print(f"\n[Steps 4-7] Extracting DAM transactions via Playwright...")
    dam_rows, dam_raw = extract_dam_transactions()

    if dam_rows is None:
        print("\n❌ ABORTED — DAM sign-in or portfolio lookup failed.")
        sys.exit(1)

    print(f"   DAM rows captured: {len(dam_rows)}")
    path = build_dam_excel(dam_rows, dam_raw)

    print("\n" + "=" * 70)
    print(f"DONE — {path}")
    print("=" * 70)


if __name__ == "__main__":
    _cli_main()
