#!/usr/bin/env python3
"""
DAM Transaction Extractor v2 — Pure extraction functions only.

This module provides reusable functions for extracting transaction data
from the DAM Transactions tab. It does NOT handle sign-in or browser setup.

The caller is responsible for:
  1. Launching Playwright browser
  2. Signing in to DAM
  3. Navigating to the portfolio
  4. Passing the `page` object to these functions

Functions:
  - navigate_to_transactions(page, portfolio_id, base_url) → navigates to Transactions tab
  - apply_date_filter(page, target_date) → opens date picker, fills D/M/YYYY, clicks Set
  - extract_transaction_rows(page) → reads all visible transaction rows
  - extract_row_detail(page, rows, index) → clicks a row, captures detail panel
  - parse_transaction_detail(detail_text, row_text, cells) → parses detail into structured dict
  - extract_all_transactions(page) → full extraction: rows + details for each
  - build_transaction_excel(transactions, row_texts, output_path) → exports to Excel

Usage:
    from utils.trx_transaction.dam_transaction_extractor_v2 import (
        navigate_to_transactions,
        apply_date_filter,
        extract_all_transactions,
        build_transaction_excel,
    )

    # After sign-in and portfolio navigation:
    navigate_to_transactions(page, portfolio_id)
    apply_date_filter(page, "2026-01-21")
    transactions, row_texts, count = extract_all_transactions(page)
    build_transaction_excel(transactions, row_texts, "output.xlsx")
"""

import os
import re
import json
from datetime import datetime
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ────────────────────────────────────────────────────────────────
# Import shared auth for default BASE_URL
try:
    from config.config import Config
    BASE_URL = Config.BASE_URL
except ImportError:
    BASE_URL = "https://dam-sit.mqbc21.com"

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SS_DIR = os.path.join(_SCRIPT_DIR, "..", "..", "test-results", "screenshots")
RESULT_DIR = os.path.join(_SCRIPT_DIR, "..", "..", "test-results")

os.makedirs(SS_DIR, exist_ok=True)
os.makedirs(RESULT_DIR, exist_ok=True)


# ── Helper functions ─────────────────────────────────────────────────────────

def _screenshot(page: Page, name: str) -> str:
    """Take a screenshot and return the path."""
    path = os.path.join(SS_DIR, f"{name}.png")
    page.screenshot(path=path, full_page=True)
    print(f"   📸 Screenshot: {name}.png")
    return path


# ── Navigation ───────────────────────────────────────────────────────────────

def navigate_to_transactions(page: Page, portfolio_id: str, base_url: str = BASE_URL):
    """
    Navigate to the Transactions tab for a given portfolio.

    Args:
        page: Playwright page (must already be signed in)
        portfolio_id: DAM portfolio UUID
        base_url: DAM base URL
    """
    tx_url = f"{base_url}/portfolio?portfolioId={portfolio_id}&tab=transactions"
    print(f"   Navigating to transactions: {tx_url}")
    page.goto(tx_url)
    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(3000)
    _screenshot(page, "trx_01_transactions_loaded")


# ── Date Filter ──────────────────────────────────────────────────────────────

def apply_date_filter(page: Page, target_date: str):
    """
    Apply a date filter on the DAM Transactions tab.

    Args:
        page: Playwright page (must be on Transactions tab)
        target_date: Date string in YYYY-MM-DD format (e.g. "2026-01-21")
    """
    print(f"\n   Applying date filter for {target_date}...")

    # Parse date
    parts = target_date.split("-")
    if len(parts) != 3:
        print(f"   ⚠️ Invalid date format: {target_date}. Expected YYYY-MM-DD")
        return

    year, month, day = parts[0], parts[1].lstrip("0"), parts[2].lstrip("0")

    # Open date picker
    date_btn = page.locator(
        "button[aria-haspopup='dialog']:has-text('Today'),"
        "button[aria-haspopup='dialog']:has-text('Yesterday'),"
        "button[aria-haspopup='dialog']:has-text('Week'),"
        "button[aria-haspopup='dialog']:has-text('Month'),"
        "button[aria-haspopup='dialog']:has-text('Custom'),"
        "button[aria-haspopup='dialog']:has-text('Between')"
    )
    try:
        date_btn.first.wait_for(state="visible", timeout=8000)
        date_btn.first.click()
        page.wait_for_timeout(1500)
        print("   ✅ Date picker opened")
    except Exception as e:
        print(f"   ⚠️ Date picker open failed: {e}")

    # Click "Between" for custom date range
    for sel in ["button:has-text('Between')", "[role='option']:has-text('Between')"]:
        try:
            el = page.locator(sel).last
            if el.is_visible(timeout=2000):
                el.click()
                page.wait_for_timeout(1000)
                print("   ✅ Selected 'Between'")
                break
        except Exception:
            pass

    # Fill D/M/YYYY inputs
    all_inputs = page.locator("input").all()
    date_inputs = []
    for inp in all_inputs:
        try:
            if inp.is_visible(timeout=300):
                ph = inp.get_attribute("placeholder") or ""
                if ph in ["D", "M", "YYYY"]:
                    date_inputs.append((ph, inp))
        except Exception:
            pass

    print(f"   Found {len(date_inputs)} D/M/YYYY inputs")

    date_values = {"D": day, "M": month, "YYYY": year}
    if len(date_inputs) >= 6:
        # From + To date inputs
        for i, (ph, inp) in enumerate(date_inputs):
            val = date_values.get(ph, "")
            if val:
                inp.click(click_count=3)
                page.wait_for_timeout(80)
                inp.fill(val)
                page.wait_for_timeout(150)
    elif len(date_inputs) >= 3:
        # From date only
        for ph, inp in date_inputs[:3]:
            val = date_values.get(ph, "")
            if val:
                inp.click(click_count=3)
                page.wait_for_timeout(80)
                inp.fill(val)
                page.wait_for_timeout(150)

    _screenshot(page, "trx_02_dates_filled")

    # Click Set/Apply
    for apply_text in ["Set", "Apply", "Confirm", "OK"]:
        try:
            btn = page.locator(f"button:has-text('{apply_text}')").last
            if btn.is_visible(timeout=1000):
                btn.click()
                print(f"   ✅ Clicked '{apply_text}'")
                page.wait_for_timeout(3000)
                break
        except Exception:
            pass

    _screenshot(page, "trx_03_filter_applied")


# ── Transaction Row Extraction ───────────────────────────────────────────────

def extract_transaction_rows(page: Page):
    """
    Find and read all transaction rows from the current page.

    Returns:
        tuple: (rows_locator, row_count, selector_used, row_texts)
    """
    row_selectors = [
        "table tbody tr",
        "[class*='transaction-row']",
        "[class*='tx-row']",
        "[class*='transaction-item']",
        "tr[class*='hover']",
        "div[role='row']",
    ]

    best_rows = None
    best_count = 0
    best_sel = None

    for sel in row_selectors:
        try:
            rows = page.locator(sel)
            count = rows.count()
            if count > best_count:
                best_count = count
                best_rows = rows
                best_sel = sel
        except Exception:
            pass

    row_texts = []
    if best_rows and best_count > 0:
        for i in range(best_count):
            try:
                txt = best_rows.nth(i).text_content() or ""
                row_texts.append(txt.strip())
            except Exception:
                pass

    return best_rows, best_count, best_sel, row_texts


def extract_row_detail(page: Page, rows, index: int) -> dict:
    """
    Click a transaction row and extract detail from the modal/panel.

    Args:
        page: Playwright page
        rows: Locator for transaction rows
        index: Row index (0-based)

    Returns:
        dict with transaction fields
    """
    tx_info = {
        "row_index": index + 1,
        "trx_hash": "",
        "date_time": "",
        "transaction_type": "",
        "from_addr": "",
        "to_addr": "",
        "amount": "",
        "resources_fee": "",
        "token_transfer": "",
        "net_transfer": "",
        "raw_cells": [],
        "detail_text": "",
    }

    try:
        row = rows.nth(index)

        # Get cell data
        cells = row.locator("td")
        for j in range(cells.count()):
            try:
                tx_info["raw_cells"].append(cells.nth(j).text_content().strip())
            except Exception:
                tx_info["raw_cells"].append("")

        # Click row to open detail
        row.click()
        page.wait_for_timeout(2000)

        # Capture detail panel
        detail_text = ""
        for d_sel in ["[role='dialog']", "[class*='detail']", "[class*='modal']",
                       "[class*='drawer']", "[class*='panel']", "[class*='sheet']"]:
            try:
                detail = page.locator(d_sel).first
                if detail.is_visible(timeout=2000):
                    detail_text = detail.text_content().strip()
                    break
            except Exception:
                pass

        tx_info["detail_text"] = detail_text

        # Parse detail into structured fields
        parse_transaction_detail(tx_info, detail_text, row.text_content() or "")

        # Close detail
        page.keyboard.press("Escape")
        page.wait_for_timeout(500)
        for close_sel in ["button[aria-label*='close' i]", "button:has-text('×')",
                           "button:has-text('Close')", "[data-testid*='close']"]:
            try:
                close_btn = page.locator(close_sel).first
                if close_btn.is_visible(timeout=500):
                    close_btn.click()
                    page.wait_for_timeout(500)
                    break
            except Exception:
                pass

    except Exception as e:
        print(f"   Row {index + 1} error: {e}")

    return tx_info


def parse_transaction_detail(tx_info: dict, detail_text: str, row_text: str):
    """
    Parse transaction detail text into structured fields.

    Args:
        tx_info: Dict to populate (modified in place)
        detail_text: Text from detail panel/modal
        row_text: Text from the table row
    """
    combined = detail_text or row_text

    # Extract full 64-char hash
    hash_match = re.search(r'\b([0-9a-fA-F]{64})\b', combined)
    if hash_match:
        tx_info["trx_hash"] = hash_match.group(1)

    # Extract TRON addresses (T + 33 chars)
    tron_addrs = re.findall(r'\bT[a-zA-Z0-9]{33}\b', combined)
    if len(tron_addrs) >= 1:
        tx_info["from_addr"] = tron_addrs[0]
    if len(tron_addrs) >= 2:
        tx_info["to_addr"] = tron_addrs[1]

    # Parse line by line
    lines = [l.strip() for l in combined.replace('\t', '\n').split('\n') if l.strip()]
    for line in lines:
        ll = line.lower()
        if any(x in ll for x in ["txid", "tx id", "hash", "transaction id"]):
            h = re.search(r'[0-9a-fA-F]{64}', line)
            if h:
                tx_info["trx_hash"] = h.group(0)
        elif "type" in ll and not tx_info["transaction_type"]:
            tx_info["transaction_type"] = line
        elif "from" in ll and not tx_info["from_addr"]:
            addr = re.search(r'T[a-zA-Z0-9]{33}', line)
            if addr:
                tx_info["from_addr"] = addr.group(0)
        elif "to" in ll and "token" not in ll and not tx_info["to_addr"]:
            addr = re.search(r'T[a-zA-Z0-9]{33}', line)
            if addr:
                tx_info["to_addr"] = addr.group(0)
        elif "amount" in ll and not tx_info["amount"]:
            tx_info["amount"] = line
        elif "fee" in ll or "resource" in ll or "bandwidth" in ll or "energy" in ll:
            if tx_info["resources_fee"]:
                tx_info["resources_fee"] += " | " + line
            else:
                tx_info["resources_fee"] = line
        elif "token" in ll and not tx_info["token_transfer"]:
            tx_info["token_transfer"] = line
        elif re.search(r'\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2}', line):
            if not tx_info["date_time"]:
                tx_info["date_time"] = line

    # Fallback: use raw_cells
    cells = tx_info.get("raw_cells", [])
    if cells:
        if len(cells) > 1 and not tx_info["trx_hash"]:
            tx_info["trx_hash"] = cells[1]
        if len(cells) > 2 and not tx_info["transaction_type"]:
            tx_info["transaction_type"] = cells[2].split('\n')[0] if '\n' in cells[2] else cells[2]
        if len(cells) > 3 and not tx_info["date_time"]:
            time_match = re.search(r'\d{2}/\d{2}/\d{4}', cells[2] + (cells[3] if len(cells) > 3 else ""))
            if time_match:
                tx_info["date_time"] = time_match.group(0)
        if len(cells) > 4 and not tx_info["from_addr"]:
            tx_info["from_addr"] = cells[4]
        if len(cells) > 5 and not tx_info["to_addr"]:
            tx_info["to_addr"] = cells[5]
        if len(cells) > 6 and not tx_info["amount"]:
            tx_info["amount"] = cells[6]


# ── Full Extraction ──────────────────────────────────────────────────────────

def extract_all_transactions(page: Page):
    """
    Extract all transaction rows and their details from the current page.

    Args:
        page: Playwright page (must be on Transactions tab with filter applied)

    Returns:
        tuple: (all_tx_data, row_texts, row_count)
    """
    rows, count, sel, row_texts = extract_transaction_rows(page)
    print(f"   Found {count} rows with selector: {sel}")

    all_tx_data = []
    if rows and count > 0:
        for i in range(count):
            print(f"   Extracting row {i + 1}/{count}...")
            tx_info = extract_row_detail(page, rows, i)
            all_tx_data.append(tx_info)

    return all_tx_data, row_texts, count


# ── Excel Export ─────────────────────────────────────────────────────────────

def build_transaction_excel(all_tx_data: list, row_texts: list, output_path: str):
    """
    Export transaction data to Excel.

    Args:
        all_tx_data: List of transaction dicts
        row_texts: List of raw row text strings
        output_path: Full path for the output Excel file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAM Transactions"

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="375623")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "#", "Trx Hash", "Date/Time", "Transaction Type",
        "From", "To", "Amount", "Resources Consumed & Fee",
        "Token Transfer", "Net Transfer", "Raw Cells"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
    ws.freeze_panes = "A2"

    fill_light = PatternFill("solid", fgColor="EEF4FB")
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    for idx, tx in enumerate(all_tx_data):
        row = [
            idx + 1,
            tx.get("trx_hash", ""),
            tx.get("date_time", ""),
            tx.get("transaction_type", ""),
            tx.get("from_addr", ""),
            tx.get("to_addr", ""),
            tx.get("amount", ""),
            tx.get("resources_fee", ""),
            tx.get("token_transfer", ""),
            tx.get("net_transfer", ""),
            " | ".join(tx.get("raw_cells", [])),
        ]
        ws.append(row)
        fill = fill_light if idx % 2 == 0 else fill_white
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    widths = [4, 68, 26, 28, 40, 40, 28, 45, 50, 35, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Raw rows sheet
    ws2 = wb.create_sheet("Raw Rows")
    ws2.append(["Row #", "Raw Text"])
    for i, txt in enumerate(row_texts, 1):
        ws2.append([i, txt])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"   ✅ Excel saved: {output_path}")
    return output_path
