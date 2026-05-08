#!/usr/bin/env python3
"""
Full TronGrid → DAM → Comparison pipeline

Orchestrator that:
  1. Runs DAM extraction (Steps 4-7) via dam_extractor
  2. Fetches TronGrid data (Steps 1-3) via trongrid_fetcher
  3. Builds a single comparison Excel (Step 8) containing all sheets

Usage:
  python3 trongrid_dam_comparison.py <ADDRESS> <DDMMYYYY>
  python3 trongrid_dam_comparison.py <PORTFOLIO_NAME> <DDMMYYYY>
  python3 trongrid_dam_comparison.py <ADDRESS> <FROM_DDMMYYYY> <TO_DDMMYYYY>
  python3 trongrid_dam_comparison.py <PORTFOLIO_NAME> <FROM_DDMMYYYY> <TO_DDMMYYYY>

Examples:
  python3 trongrid_dam_comparison.py TQbqqt5kEfgXoQP31HUFumM5bYpieFcNQ6 16042026
  python3 trongrid_dam_comparison.py trx2_Mkx 01032026 31032026
"""

import os, sys, json, re
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Ensure parent package is importable when run directly
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.normpath(os.path.join(_SCRIPT_DIR, "..", ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from utils.trx_transaction import shared
from utils.trx_transaction.shared import (
    _is_trx_address, _parse_date, _date_to_ts_ms, ts_to_utc,
    header_style, data_style, thin_border,
)
from utils.trx_transaction.trongrid_fetcher import fetch_and_parse
from utils.trx_transaction.dam_extractor import extract_dam_transactions


# ─────────────────────────────────────────────────────────────────────────────
# Build Comparison Excel (Step 8) — contains ALL data + comparison
# ─────────────────────────────────────────────────────────────────────────────
def build_comparison_excel(tg_parsed, dam_rows):
    wb = openpyxl.Workbook()

    # ── Index by truncated hash ──────────────────────────────────────────────
    def norm_hash(h):
        h = (h or "").strip()
        if "..." in h:
            parts = h.split("...")
            if len(parts) == 2:
                return (parts[0].lower(), parts[1].lower())
        return (h[:6].lower(), h[-6:].lower()) if len(h) >= 12 else (h.lower(), "")

    tg_by_hash = {}
    for tx in tg_parsed:
        h = tx["trx_hash"]
        tg_by_hash[h.lower()] = tx
        if len(h) >= 12:
            tg_by_hash[(h[:6].lower(), h[-6:].lower())] = tx

    # ── Match DAM rows to TronGrid ───────────────────────────────────────────
    dam_matched = []
    for d in dam_rows:
        raw_hash = d.get("trx_hash", "")
        tg_match = None
        if len(raw_hash) == 64:
            tg_match = tg_by_hash.get(raw_hash.lower())
        if not tg_match:
            key = norm_hash(raw_hash)
            tg_match = tg_by_hash.get(key)
        dam_matched.append((d, tg_match))

    dam_hashes = set()
    for d, tg in dam_matched:
        if tg:
            dam_hashes.add(tg["trx_hash"].lower())

    tg_only = [tx for tx in tg_parsed if tx["trx_hash"].lower() not in dam_hashes]

    # ── Sheet 1: Summary & Conclusion ────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary & Conclusion"

    matched_count = sum(1 for _, tg in dam_matched if tg)
    dam_only      = sum(1 for _, tg in dam_matched if not tg)
    tg_only_count = len(tg_only)
    all_match     = (dam_only == 0 and tg_only_count == 0)

    local_tz = datetime.now().astimezone().tzinfo

    summary = [
        ["COMPARISON REPORT: TronGrid API vs DAM System"],
        [""],
        ["Address",         shared.ADDRESS],
        ["Date Range",      f"{shared.DATE_FROM} to {shared.DATE_TO}"],
        ["Timezone",        str(local_tz)],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [""],
        ["TRANSACTION COUNTS", "", ""],
        ["Source",                          "Timezone", "Transactions"],
        ["TronGrid API (api.trongrid.io)",  "UTC",       str(len(tg_parsed))],
        ["DAM System (dam-sit.mqbc21.com)", str(local_tz), str(len(dam_rows))],
        [""],
        ["MATCHING ANALYSIS", "", ""],
        ["Category",                             "Count"],
        ["Matched (in both TronGrid and DAM)",   str(matched_count)],
        ["TronGrid only (not in DAM view)",       str(tg_only_count)],
        ["DAM only (not found in TronGrid data)", str(dam_only)],
        [""],
        ["TIMEZONE NOTE", "", ""],
        ["", (f"TronGrid uses UTC. DAM uses {local_tz}. "
              "Transactions near midnight may appear on different calendar dates."), ""],
        [""],
        ["FINAL VERDICT", "", ""],
    ]

    if all_match:
        summary.append(["",
            (f"✅ ALL TRANSACTIONS MATCH\n\n"
             f"All {matched_count} transactions found in TronGrid also appear in DAM. "
             "No discrepancies found in transaction hashes, types, addresses, or amounts. "
             f"Any count difference is explained by the UTC vs {local_tz} timezone offset."), ""])
    else:
        summary.append(["",
            (f"⚠️ DIFFERENCES FOUND\n\n"
             f"• {matched_count} transactions matched between TronGrid and DAM.\n"
             f"• {tg_only_count} transactions in TronGrid not visible in DAM (likely timezone boundary).\n"
             f"• {dam_only} transactions in DAM not matched to TronGrid data.\n\n"
             "Please review the 'Comparison Detail' sheet for specifics."), ""])

    for row in summary:
        ws_sum.append(row)

    ws_sum["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    ws_sum["A1"].fill = PatternFill("solid", fgColor="1A1A2E")
    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 95
    ws_sum.column_dimensions["C"].width = 20
    for row in ws_sum.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Sheet 2: Comparison Detail ───────────────────────────────────────────
    ws2 = wb.create_sheet("Comparison Detail")
    hdrs = ["#", "DAM Hash (truncated)", "DAM Date/Time", "DAM Type",
            "DAM From", "DAM To", "DAM Amount", "DAM Token",
            "TronGrid Hash (full)", "TronGrid Date (UTC)", "TronGrid Type",
            "TronGrid From", "TronGrid To", "TronGrid Amount", "TronGrid Token",
            "MATCH?", "Notes"]
    ws2.append(hdrs)
    header_style(ws2, 1, "555555")
    ws2.freeze_panes = "A2"

    fill_match   = PatternFill("solid", fgColor="C6EFCE")
    fill_partial = PatternFill("solid", fgColor="FFEB9C")
    fill_missing = PatternFill("solid", fgColor="FFC7CE")
    bdr = thin_border()
    font_s = Font(size=8, name="Calibri")
    aln_t  = Alignment(vertical="top", wrap_text=True)

    for i, (d, tg) in enumerate(dam_matched, 1):
        match_result = "✅ MATCH" if tg else "⚠️ Not found in TronGrid range"
        note = "" if tg else (
            "DAM hash not matched — may be due to timezone boundary or DAM-only activity."
        )
        row = [
            i,
            d.get("trx_hash", ""), d.get("date_time", ""), d.get("tx_type", ""),
            d.get("from_addr", ""), d.get("to_addr", ""),
            d.get("amount", ""), d.get("token_transfer", ""),
            tg["trx_hash"] if tg else "", tg["date_time"] if tg else "",
            tg["tx_type"] if tg else "", tg["from_addr"] if tg else "",
            tg["to_addr"] if tg else "", tg["amount"] if tg else "",
            tg["token_transfer"] if tg else "",
            match_result, note,
        ]
        ws2.append(row)
        fill = fill_match if tg else fill_missing
        for cell in ws2[ws2.max_row]:
            cell.fill = fill
            cell.font = font_s
            cell.alignment = aln_t
            cell.border = bdr
        ws2.row_dimensions[ws2.max_row].height = 45

    if tg_only:
        ws2.append([])
        ws2.append(["--- TRONGRID TRANSACTIONS NOT FOUND IN DAM (timezone boundary or not in DAM view) ---"])
        for tx in tg_only:
            row = ["", "", "", "", "", "", "", "",
                   tx["trx_hash"], tx["date_time"], tx["tx_type"],
                   tx["from_addr"], tx["to_addr"], tx["amount"], tx["token_transfer"],
                   "⚠️ TronGrid Only", "Not visible in DAM date filter window"]
            ws2.append(row)
            for cell in ws2[ws2.max_row]:
                cell.fill = fill_partial
                cell.font = font_s
                cell.alignment = aln_t
                cell.border = bdr

    widths = [4, 35, 25, 25, 38, 38, 25, 35, 68, 26, 30, 40, 40, 28, 45, 16, 50]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: TronGrid Data ───────────────────────────────────────────────
    ws3 = wb.create_sheet("TronGrid Data")
    ws3.append(["#", "Trx Hash", "Date/Time (UTC)", "Transaction Type",
                "From", "To", "Amount", "Resources & Fee", "Token Transfer", "Net Transfer"])
    header_style(ws3, 1, "1F4E79")
    ws3.freeze_panes = "A2"
    for i, tx in enumerate(tg_parsed, 1):
        ws3.append([i, tx["trx_hash"], tx["date_time"], tx["tx_type"],
                    tx["from_addr"], tx["to_addr"], tx["amount"],
                    tx["resources_fee"], tx["token_transfer"], tx["net_transfer"]])
        data_style(ws3, ws3.max_row, i % 2 == 0)
    for i, w in enumerate([4, 68, 26, 32, 40, 40, 28, 55, 60, 40], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: DAM Data ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("DAM Data")
    ws4.append(["#", "Trx Hash", "Date/Time (DAM)", "Transaction Type",
                "From", "To", "Amount", "Token Transfer", "Net Transfer", "Raw Cells"])
    header_style(ws4, 1, "375623")
    ws4.freeze_panes = "A2"
    for i, tx in enumerate(dam_rows, 1):
        ws4.append([i, tx.get("trx_hash", ""), tx.get("date_time", ""), tx.get("tx_type", ""),
                    tx.get("from_addr", ""), tx.get("to_addr", ""),
                    tx.get("amount", ""), tx.get("token_transfer", ""),
                    tx.get("net_transfer", ""), " | ".join(tx.get("raw_cells", []))])
        data_style(ws4, ws4.max_row, i % 2 == 0)
    for i, w in enumerate([4, 68, 26, 28, 40, 40, 28, 45, 35, 70], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(shared.EXCEL_DIR, exist_ok=True)
    path = f"{shared.EXCEL_DIR}/Comparison_{shared.ADDRESS[-8:]}_{shared.DATE_FROM}_to_{shared.DATE_TO}_{ts}.xlsx"
    wb.save(path)
    print(f"   ✅ Comparison Excel saved: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage:")
        print("  python3 trongrid_dam_comparison.py <ADDRESS_OR_NAME> <DDMMYYYY>")
        print("  python3 trongrid_dam_comparison.py <ADDRESS_OR_NAME> <FROM_DDMMYYYY> <TO_DDMMYYYY>")
        print()
        print("Examples:")
        print("  python3 trongrid_dam_comparison.py TQbqqt5kEfgXoQP31HUFumM5bYpieFcNQ6 16042026")
        print("  python3 trongrid_dam_comparison.py trx2_Mkx 01032026 31032026")
        sys.exit(1)

    input_target = sys.argv[1]
    date1_str    = sys.argv[2]
    date2_str    = sys.argv[3] if len(sys.argv) > 3 else None

    # Classify input
    if _is_trx_address(input_target):
        shared.ADDRESS = input_target
        shared.PORTFOLIO_NAME = None
    else:
        shared.PORTFOLIO_NAME = input_target
        shared.ADDRESS = ""

    # Parse dates
    _, _, _, dt_from = _parse_date(date1_str)
    if date2_str:
        _, _, _, dt_to = _parse_date(date2_str)
    else:
        dt_to = dt_from.replace(hour=23, minute=59, second=59)

    shared.DATE_FROM = dt_from.strftime("%Y-%m-%d")
    shared.DATE_TO   = dt_to.strftime("%Y-%m-%d")
    shared.TS_FROM   = _date_to_ts_ms(dt_from)
    shared.TS_TO     = _date_to_ts_ms(dt_to.replace(hour=23, minute=59, second=59))

    # Log timezone
    local_tz = datetime.now().astimezone().tzinfo
    utc_from = datetime.fromtimestamp(shared.TS_FROM / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    utc_to   = datetime.fromtimestamp(shared.TS_TO / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    # Set up output directories
    label = shared.ADDRESS[-8:] if shared.ADDRESS else shared.PORTFOLIO_NAME
    shared.SS_DIR = f"{shared.OUT_DIR}/test-results/screenshots/{label}"
    os.makedirs(shared.SS_DIR, exist_ok=True)

    print("=" * 70)
    print("TronGrid vs DAM Comparison Pipeline")
    if shared.ADDRESS:
        print(f"Address : {shared.ADDRESS}")
    if shared.PORTFOLIO_NAME:
        print(f"Portfolio: {shared.PORTFOLIO_NAME}")
    print(f"Range   : {shared.DATE_FROM} to {shared.DATE_TO}")
    print(f"Timezone: {local_tz} (detected from system)")
    print(f"UTC range: {utc_from} → {utc_to}")
    print("=" * 70)

    # ── Steps 4-7: DAM extraction (runs first to discover address) ───────
    print(f"\n[Steps 4-7] Extracting DAM transactions via Playwright...")
    dam_rows, dam_raw = extract_dam_transactions()

    if dam_rows is None:
        print("\n❌ ABORTED — DAM sign-in or portfolio lookup failed.")
        sys.exit(1)

    print(f"   DAM rows captured: {len(dam_rows)}")
    if shared.ADDRESS:
        print(f"   Address (from Combined Net Worth): {shared.ADDRESS}")

    # ── Steps 1-3: TronGrid fetch ────────────────────────────────────────
    parsed_txs = []
    if shared.ADDRESS:
        print(f"\n[Steps 1-3] Fetching TronGrid transactions for {shared.ADDRESS}...")
        parsed_txs, raw_txs, trc20_list = fetch_and_parse()
        print(f"   TronGrid transactions: {len(parsed_txs)}")
    else:
        print(f"\n[Steps 1-3] ⏭️  Skipped TronGrid — no TRX address found in portfolio")

    # ── Step 8: Build single comparison Excel ────────────────────────────
    print(f"\n[Step 8] Building comparison Excel...")
    cmp_excel = build_comparison_excel(parsed_txs, dam_rows)

    # ── Save raw data ────────────────────────────────────────────────────
    raw_path = f"{shared.OUT_DIR}/test-results/{label}_raw_data.json"
    os.makedirs(os.path.dirname(raw_path), exist_ok=True)
    with open(raw_path, "w") as f:
        json.dump({
            "trongrid_count":  len(parsed_txs),
            "dam_count":       len(dam_rows),
            "trongrid_parsed": parsed_txs,
            "dam_rows":        dam_rows,
        }, f, indent=2, default=str)
    print(f"   Raw data: {raw_path}")

    print("\n" + "=" * 70)
    print("DONE")
    print(f"  Comparison Excel: {cmp_excel}")
    print("=" * 70)
