#!/usr/bin/env python3
"""
Portfolio Manager — search existing or create new portfolios in DAM.

Responsibilities:
  1. Excel lookup  — check if address/portfolio exists in DAM addresses.xlsx
  2. DAM dropdown  — search the portfolio dropdown for a given name
  3. Create        — open "Create portfolio" dialog, fill name + addresses, save
  4. Navigate      — click an existing portfolio and land on its page

Usage from run_overview.py:
    from utils.portfolio.portfolio_manager import (
        lookup_portfolio_in_excel,
        lookup_address_in_excel,
        classify_address,
        lookup_portfolio_in_dam,
        search_portfolio_in_dropdown,
        create_portfolio_in_dam,
        navigate_to_portfolio,
    )
"""

import os
import re
import json

DAM_EXCEL_PATH = "test_data/DAM addresses.xlsx"


# ============================================================================
# 1. EXCEL LOOKUPS
# ============================================================================

def lookup_portfolio_in_excel(portfolio_name):
    """
    Look up a portfolio name in DAM addresses.xlsx and return ALL associated addresses.

    Returns:
        (addresses_list, portfolio_name) or ([], None) if not found.
        Addresses are collected from columns B onward.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(DAM_EXCEL_PATH, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2:
                excel_portfolio = str(row[0]).strip() if row[0] else ""
                if excel_portfolio.lower() == portfolio_name.lower():
                    excel_addresses = []
                    for col_val in row[1:]:
                        addr = str(col_val).strip() if col_val else ""
                        if addr and addr.startswith('T') and len(addr) == 34:
                            excel_addresses.append(addr)
                    wb.close()
                    return excel_addresses, excel_portfolio

        wb.close()
        return [], None
    except Exception as e:
        print(f"⚠️  Error reading DAM addresses Excel: {e}")
        return [], None


def lookup_address_in_excel(address):
    """
    Look up an ADDRESS in DAM addresses.xlsx and return the portfolio that contains it.

    Searches all address columns (B, C, D, …).

    Returns:
        (portfolio_name, all_addresses_in_portfolio) or (None, []) if not found.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(DAM_EXCEL_PATH, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2:
                portfolio_name = str(row[0]).strip() if row[0] else ""
                row_addresses = []
                for col_val in row[1:]:
                    addr = str(col_val).strip() if col_val else ""
                    if addr and addr.startswith('T') and len(addr) == 34:
                        row_addresses.append(addr)
                if address in row_addresses:
                    wb.close()
                    return portfolio_name, row_addresses

        wb.close()
        return None, []
    except Exception as e:
        print(f"⚠️  Error reading DAM addresses Excel: {e}")
        return None, []


def update_excel_with_portfolio(portfolio_name, addresses):
    """
    Add or update a portfolio row in DAM addresses.xlsx.

    Args:
        portfolio_name: Name to write in column A.
        addresses: List of addresses to write in columns B, C, D, …
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(DAM_EXCEL_PATH)
        ws = wb.active
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=portfolio_name)
        for idx, addr in enumerate(addresses):
            ws.cell(row=next_row, column=2 + idx, value=addr)
        wb.save(DAM_EXCEL_PATH)
        wb.close()
        print(f"✅ Excel updated — row {next_row}: {portfolio_name} | {len(addresses)} address(es)")
    except Exception as e:
        print(f"⚠️  Could not update Excel: {e}")


# ============================================================================
# 2. ADDRESS CLASSIFICATION
# ============================================================================

def classify_address(addr):
    """
    Classify an address as 'tron', 'evm', or 'exchange'.

    - Tron:     starts with T, 34 characters
    - EVM:      starts with 0x, 42 characters
    - Exchange: everything else (names like "Binance", "moontest")
    """
    addr = addr.strip()
    if (addr.startswith('T') or addr.startswith('t')) and len(addr) == 34:
        return 'tron'
    elif addr.startswith('0x') and len(addr) == 42:
        return 'evm'
    else:
        return 'exchange'


# ============================================================================
# 3. DAM DROPDOWN — SEARCH
# ============================================================================

def search_portfolio_in_dropdown(page, portfolio_name):
    """
    Search the open portfolio dropdown for *portfolio_name*.

    Assumes the dropdown is already open.

    Returns:
        (found: bool, element: Locator | None, display_name: str | None)
    """
    found = False
    element = None
    display_name = None

    # --- Method 0: CSS class selectors for portfolio name elements ---
    _selectors = [
        'div.text-mono-900.break-all',
        'div[class*="text-mono-900"][class*="break-all"]',
        'div.text-mono-900.typography-body.font-normal.break-all.text-left.w-full',
    ]
    for sel in _selectors:
        try:
            for elem_candidate in page.locator(sel).all():
                if elem_candidate.is_visible():
                    txt = elem_candidate.text_content().strip()
                    if txt.lower() == portfolio_name.lower():
                        parent_btn = elem_candidate.locator('xpath=ancestor::button').first
                        if parent_btn.count() > 0:
                            element = parent_btn
                        else:
                            fallback = elem_candidate.locator(
                                'xpath=ancestor::div[contains(@class,"cursor-pointer") '
                                'or @role="option" or @role="menuitem"]'
                            ).first
                            element = fallback if fallback.count() > 0 else elem_candidate
                        found = True
                        display_name = txt
                        return found, element, display_name
        except Exception:
            pass

    # --- Method 1: getByText partial match ---
    try:
        partial = page.get_by_text(portfolio_name)
        if partial.count() > 0:
            for i in range(partial.count()):
                el = partial.nth(i)
                if el.is_visible():
                    first_line = el.text_content().strip().split('\n')[0].strip()
                    if first_line.lower() == portfolio_name.lower():
                        element = el
                        found = True
                        display_name = first_line
                        return found, element, display_name
    except Exception:
        pass

    # --- Method 2: role="menuitem" substring ---
    try:
        items = page.get_by_role("menuitem").all()
        for item in items:
            if not item.is_visible():
                continue
            item_text = item.text_content().strip()
            if portfolio_name.lower() in item_text.lower() and 'create portfolio' not in item_text.lower():
                element = item
                found = True
                display_name = item_text.split('\n')[0].strip()
                return found, element, display_name
    except Exception:
        pass

    # --- Method 3: brute-force div scan ---
    try:
        for div in page.locator('div').all():
            if not div.is_visible():
                continue
            div_text = div.text_content().strip()
            if not div_text:
                continue
            if portfolio_name.lower() in div_text.lower():
                if ('Addresses' in div_text or 'Exchange' in div_text) and 'create portfolio' not in div_text.lower():
                    element = div
                    found = True
                    display_name = div_text.split('\n')[0].strip()
                    return found, element, display_name
    except Exception:
        pass

    return found, element, display_name


def open_portfolio_dropdown(page):
    """
    Open the portfolio dropdown. Returns True if opened successfully.
    """
    dropdown_opened = False

    # Try text-based button first
    btn = page.locator('button:has-text("Portfolio")').first
    if btn.count() > 0:
        try:
            btn.click(timeout=5000)
            page.wait_for_timeout(2000)
            dropdown_opened = True
        except Exception:
            pass

    # Fallback: coordinate click on chevron area
    if not dropdown_opened:
        try:
            page.mouse.click(395, 141)
            page.wait_for_timeout(2000)
            dropdown_opened = True
        except Exception:
            pass

    # Scroll inside dropdown to load all items
    if dropdown_opened:
        page.mouse.move(490, 450)
        for _ in range(15):
            page.mouse.wheel(0, 300)
            page.wait_for_timeout(100)
        page.wait_for_timeout(500)

    return dropdown_opened


# ============================================================================
# 4. DAM — CREATE PORTFOLIO
# ============================================================================

def create_portfolio_in_dam(page, portfolio_name, addresses, screenshot_folder=None):
    """
    Create a new portfolio in DAM via the UI.

    Steps:
      1. Open dropdown → scroll to bottom → click "+ Create portfolio"
      2. Fill portfolio name
      3. Add each address (EVM + TRX)
      4. Click Save

    Args:
        page: Playwright page object (must be signed in).
        portfolio_name: Name for the new portfolio.
        addresses: List of wallet addresses to add (EVM 0x… and/or TRX T…).
        screenshot_folder: Optional folder for debug screenshots.

    Returns:
        True if portfolio was created successfully, False otherwise.
    """
    print("STEP 3: Create Portfolio")
    print("=" * 80)

    # Close any open dropdown, then reopen cleanly
    page.keyboard.press("Escape")
    page.wait_for_timeout(500)

    if not open_portfolio_dropdown(page):
        print("❌ Could not open portfolio dropdown")
        return False

    # Debug screenshot
    if screenshot_folder:
        try:
            page.screenshot(path=f"{screenshot_folder}/debug_dropdown.png")
        except Exception:
            pass

    # --- Click "+ Create portfolio" ---
    print("➕ Clicking Create portfolio...")
    page.wait_for_timeout(1000)

    create_clicked = False
    _create_strategies = [
        lambda: page.get_by_role("menuitem", name="Create portfolio", exact=True).first,
        lambda: page.get_by_text("Create portfolio", exact=True),
        lambda: page.get_by_text("Create portfolio").first,
        lambda: page.get_by_role("menuitem").filter(has_text="Create").first,
    ]
    for strategy in _create_strategies:
        try:
            btn = strategy()
            if btn.count() > 0 and btn.is_visible():
                btn.click()
                create_clicked = True
                break
        except Exception:
            continue

    if not create_clicked:
        print("❌ Could not find Create portfolio button")
        if screenshot_folder:
            try:
                page.screenshot(path=f"{screenshot_folder}/debug_dropdown_failed.png")
            except Exception:
                pass
        return False

    page.wait_for_timeout(2000)
    print("✅ Create Portfolio dialog opened")

    # --- Fill portfolio name ---
    print(f"📝 Portfolio: {portfolio_name}")
    name_input = None
    for placeholder in ["Enter portfolio name", "Portfolio name", "Name", "Enter name"]:
        candidate = page.get_by_placeholder(placeholder)
        if candidate.count() > 0:
            name_input = candidate
            break
    if name_input is None:
        for sel in ['dialog input[type="text"]', 'dialog input', '[role="dialog"] input',
                     'form input[type="text"]', 'input[type="text"]:visible']:
            candidate = page.locator(sel)
            if candidate.count() > 0 and candidate.first.is_visible(timeout=2000):
                name_input = candidate.first
                break
    if not name_input:
        print("❌ Could not find portfolio name input field")
        return False

    name_input.click()
    name_input.fill(portfolio_name)
    page.wait_for_timeout(500)
    print("✅ Portfolio name entered")

    # --- Add addresses ---
    print(f"📍 Adding {len(addresses)} address(es)...")
    for addr_idx, addr in enumerate(addresses):
        print(f"   Adding: {addr}")
        addr_field = None

        # Try specific input by index
        specific = page.locator(f'input[name="wallet.{addr_idx}.address"]').first
        if specific.count() > 0:
            try:
                specific.wait_for(state="visible", timeout=5000)
                if specific.is_enabled(timeout=2000):
                    addr_field = specific
            except Exception:
                pass

        # Fallback: any enabled placeholder input
        if not addr_field:
            for sel in [
                '[placeholder*="wallet address"]:not([disabled])',
                '[placeholder*="Paste your wallet"]:not([disabled])',
                'textarea:not([disabled])',
            ]:
                try:
                    candidate = page.locator(sel).first
                    if candidate.count() > 0 and candidate.is_visible(timeout=2000) and candidate.is_enabled(timeout=1000):
                        addr_field = candidate
                        break
                except Exception:
                    pass

        if addr_field:
            addr_field.click()
            addr_field.fill(addr)
            page.wait_for_timeout(500)
            page.keyboard.press("Enter")
            page.wait_for_timeout(2000)
            print(f"   ✅ Added: {addr[:12]}...")
        else:
            print(f"   ⚠️  Could not find address input for: {addr}")

    # --- Save ---
    print("💾 Saving portfolio...")
    save_button = page.locator('button:has-text("Save")').first
    saved = False
    try:
        save_button.wait_for(state="visible", timeout=10000)
        page.wait_for_timeout(3000)
        save_button.click(timeout=15000)
        page.wait_for_timeout(5000)
        saved = True
        print("✅ Portfolio saved!")
    except Exception as e:
        print(f"⚠️  Save button issue: {e}")
        # Fallback: submit button
        try:
            page.locator('button[type="submit"]').first.click(timeout=10000)
            page.wait_for_timeout(5000)
            saved = True
            print("✅ Portfolio saved (submit fallback)!")
        except Exception:
            # Last resort: JS click
            try:
                page.evaluate(
                    'document.querySelector(\'button:has-text("Save"), button[type="submit"]\')?.click()'
                )
                page.wait_for_timeout(5000)
                saved = True
                print("✅ Portfolio saved (JS click)!")
            except Exception as e2:
                print(f"❌ Could not save portfolio: {e2}")

    if saved:
        print("✅ Portfolio created")
    return saved


# ============================================================================
# 5. NAVIGATE TO EXISTING PORTFOLIO
# ============================================================================

def navigate_to_portfolio(page, portfolio_name, config):
    """
    Navigate to an existing portfolio by name via the dropdown.

    Args:
        page: Playwright page object (must be signed in).
        portfolio_name: Name of the portfolio to navigate to.
        config: Config object with PORTFOLIO_URL, BASE_URL.

    Returns:
        (success: bool, actual_portfolio_name: str | None)
    """
    if not open_portfolio_dropdown(page):
        print("❌ Could not open portfolio dropdown")
        return False, None

    found, element, display_name = search_portfolio_in_dropdown(page, portfolio_name)

    if found and element:
        print(f"   Navigating to existing portfolio: {display_name}")
        element.click()
        page.wait_for_timeout(3000)

        # Extract actual portfolio name from UI
        actual_name = display_name
        try:
            name_elem = page.locator(
                'div.text-mono-900.typography-body.font-normal.text-left.break-all.w-full'
            ).first
            if name_elem.count() > 0:
                actual_name = name_elem.text_content().strip()
                print(f"   📝 Portfolio name from DAM UI: {actual_name}")
        except Exception:
            pass

        return True, actual_name

    print(f"⚠️  Portfolio '{portfolio_name}' NOT found in DAM dropdown")
    return False, None


# ============================================================================
# 6. FULL DAM LOOKUP (sign in → search → extract addresses)
# ============================================================================

def lookup_portfolio_in_dam(portfolio_name):
    """
    Sign in to DAM, search for *portfolio_name* in the dropdown,
    click it, and extract all addresses from the Combined Net Worth section.

    Returns:
        (trx_addresses, evm_addresses, exchanges, portfolio_display_name)
        or ([], [], [], None) if not found.
    """
    from playwright.sync_api import sync_playwright

    print(f"\n🌐 Checking DAM for portfolio: '{portfolio_name}'")

    # Load credentials
    from config.config import Config
    _tc1_path = os.path.join(Config.PROJECT_ROOT, "test_data", "tc1_account.json")
    if os.path.exists(_tc1_path):
        with open(_tc1_path) as f:
            acc = json.load(f)
        test_email = acc["email"]
        test_password = acc["password"]
    else:
        test_email = Config.TEST_EMAIL
        test_password = Config.TEST_PASSWORD

    found_trx = []
    found_evm = []
    found_exchanges = []
    found_portfolio = None

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True, slow_mo=300, channel="chrome",
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        try:
            from playwright_stealth import Stealth
            Stealth(navigator_platform_override="MacIntel").apply_stealth_sync(context)
        except ImportError:
            pass

        page = context.new_page()

        try:
            from config.config import Config

            # Sign in
            print(f"   🔐 Signing in with: {test_email}")
            page.goto(Config.SIGN_IN_URL)
            page.wait_for_timeout(2000)
            page.fill('input[data-testid="input-email"]', test_email)
            page.fill('input[data-testid="input-password"]', test_password)
            page.click('button[data-testid="sign-in-btn"]')
            page.wait_for_timeout(8000)

            # Dismiss popup
            try:
                for sel in ['button:has-text("×")', '[aria-label="close"]']:
                    if page.locator(sel).is_visible(timeout=1000):
                        page.locator(sel).first.click()
                        page.wait_for_timeout(500)
                        break
            except Exception:
                pass

            # If redirected to a specific portfolio, go to portfolio list
            if "portfolioId=" in page.url:
                page.goto(f"{Config.BASE_URL}/portfolio")
                page.wait_for_timeout(3000)

            print("   ✅ Signed in")

            # Open dropdown and search
            if not open_portfolio_dropdown(page):
                print("   ⚠️  Could not open dropdown")
                browser.close()
                return [], [], [], None

            found, element, display_name = search_portfolio_in_dropdown(page, portfolio_name)

            # If not found in current view, scroll through dropdown
            if not found:
                print(f"   🔄 Scrolling through dropdown...")
                for _ in range(30):
                    page.keyboard.press('ArrowDown')
                    page.wait_for_timeout(200)
                    found, element, display_name = search_portfolio_in_dropdown(page, portfolio_name)
                    if found:
                        break

            if not found or not element:
                print(f"   ❌ Portfolio '{portfolio_name}' not found in DAM dropdown")
                browser.close()
                return [], [], [], None

            # Click portfolio
            element.click()
            page.wait_for_timeout(5000)

            # Get display name
            try:
                name_elem = page.locator(
                    'div.text-mono-900.typography-body.font-normal.text-left.break-all.w-full'
                ).first
                if name_elem.count() > 0:
                    found_portfolio = name_elem.text_content().strip()
            except Exception:
                found_portfolio = display_name or portfolio_name

            print(f"   📝 Portfolio: {found_portfolio}")

            # --- Extract addresses from Combined Net Worth ---
            collected_tron = set()
            collected_evm = set()
            collected_exchanges = set()

            # Strategy 1: tooltip IDs
            try:
                entries = page.evaluate('''() => {
                    const results = [];
                    const els = document.querySelectorAll('[data-tooltip-id^="address-display-tooltip-"]');
                    for (const el of els) {
                        let addr = "";
                        const hl = el.querySelector("[data-highlight-target]");
                        if (hl) addr = hl.getAttribute("data-highlight-target") || "";
                        if (!addr) {
                            const tid = el.getAttribute("data-tooltip-id") || "";
                            addr = tid.replace("address-display-tooltip-", "");
                        }
                        if (!addr) {
                            const ne = el.querySelector('div[class*="typography-body"]');
                            addr = ne ? ne.textContent.trim() : el.textContent.trim();
                        }
                        if (addr.trim()) results.push(addr.trim());
                    }
                    return results;
                }''')
                for addr in entries:
                    addr = addr.strip()
                    t = classify_address(addr)
                    if t == 'tron':
                        collected_tron.add(addr)
                    elif t == 'evm':
                        collected_evm.add(addr)
            except Exception:
                pass

            # Strategy 2: hover truncated addresses for full text
            if not collected_tron and not collected_evm:
                try:
                    for pattern_sel in [
                        'text=/[Tt][A-Za-z0-9]{6,8}[\\.…]{2,3}[A-Za-z0-9]{4,8}/',
                        'text=/0x[A-Fa-f0-9]{4,8}[\\.…]{2,3}[A-Fa-f0-9]{4,8}/',
                    ]:
                        for el in page.locator(pattern_sel).all():
                            try:
                                el.hover()
                                page.wait_for_timeout(1500)
                                tooltip = page.locator('[role="tooltip"]').first
                                if tooltip.count() > 0 and tooltip.is_visible():
                                    tt = tooltip.text_content().strip()
                                    for m in re.findall(r'([Tt][A-Za-z0-9]{33})', tt):
                                        collected_tron.add(m)
                                    for m in re.findall(r'(0x[A-Fa-f0-9]{40})', tt):
                                        collected_evm.add(m)
                            except Exception:
                                pass
                except Exception:
                    pass

            # Strategy 3: page source scan
            if not collected_tron and not collected_evm:
                try:
                    html = page.content()
                    for m in re.findall(r'address-display-tooltip-([Tt][A-Za-z0-9]{33})', html):
                        collected_tron.add(m)
                    for m in re.findall(r'address-display-tooltip-(0x[A-Fa-f0-9]{40})', html):
                        collected_evm.add(m)
                except Exception:
                    pass

            found_trx = sorted(collected_tron, key=str.lower)
            found_evm = sorted(collected_evm, key=str.lower)
            found_exchanges = sorted(collected_exchanges)

            print(f"   📋 Tron: {len(found_trx)}, EVM: {len(found_evm)}, Exchanges: {len(found_exchanges)}")

            # Update Excel
            all_addrs = found_trx + found_evm
            if all_addrs:
                update_excel_with_portfolio(found_portfolio, all_addrs)

        except Exception as e:
            print(f"   ❌ Error: {e}")
        finally:
            browser.close()

    return found_trx, found_evm, found_exchanges, found_portfolio
