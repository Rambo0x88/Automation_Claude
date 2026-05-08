"""
Rabby API utility for fetching and exporting DeFi protocol and app positions.

Endpoints:
  - complex_protocol_list: Aave, Morpho, Compound (and other EVM DeFi protocols)
  - complex_app_list: Hyperliquid (and other off-chain/hybrid apps)

Both endpoints require an EVM address (0x + 40 hex chars).
"""

import json
import os
import re
import time
import requests
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from decimal import Decimal, getcontext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

getcontext().prec = 50

_PROTOCOL_URL = "https://api.rabby.io/v1/user/complex_protocol_list"
_APP_URL = "https://api.rabby.io/v1/user/complex_app_list"

_EVM_RE = re.compile(r'^0x[0-9a-fA-F]{40}$')

# ---------------------------------------------------------------------------
# Rate limiter — max 5 calls per 60 seconds across all Rabby API requests
# ---------------------------------------------------------------------------

class _RateLimiter:
    def __init__(self, max_calls: int = 5, period: int = 60):
        self.max_calls = max_calls
        self.period = period
        self._timestamps: deque = deque()

    def wait(self):
        now = time.time()
        # Drop timestamps older than the window
        while self._timestamps and now - self._timestamps[0] >= self.period:
            self._timestamps.popleft()

        if len(self._timestamps) >= self.max_calls:
            wait_for = self.period - (now - self._timestamps[0])
            if wait_for > 0:
                print(f"   ⏳ Rate limit reached (5/min). Waiting {wait_for:.1f}s...", flush=True)
                time.sleep(wait_for)
            # Refresh after sleeping
            now = time.time()
            while self._timestamps and now - self._timestamps[0] >= self.period:
                self._timestamps.popleft()

        self._timestamps.append(time.time())


_rate_limiter = _RateLimiter(max_calls=5, period=60)

_HEADERS = {
    "User-Agent": "PostmanRuntime/7.43.0",
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}


def _is_evm(address: str) -> bool:
    return bool(_EVM_RE.match(address))


def _fetch_with_retry(url: str, params: dict, proxy: str, rate_limiter: "_RateLimiter",
                      label: str, max_retries: int = 3) -> object:
    """
    GET a Rabby API endpoint with rate-limit enforcement and 429 retry.
    Without proxies: retries up to max_retries times with exponential back-off.
    With proxies: each batch is on a separate IP so 429 is rare; still retries.
    """
    proxies = {"http": proxy, "https": proxy} if proxy else None
    for attempt in range(max_retries):
        rate_limiter.wait()
        try:
            resp = requests.get(url, params=params, headers=_HEADERS, proxies=proxies, timeout=30)
            if resp.status_code == 200:
                return resp.json()
            if resp.status_code == 429:
                wait_s = 30 * (attempt + 1)  # 30s, 60s, 90s
                print(f"   ⚠️  429 rate limit for {label} (attempt {attempt+1}/{max_retries}), waiting {wait_s}s...")
                time.sleep(wait_s)
                continue
            resp.raise_for_status()
        except requests.HTTPError:
            raise
        except Exception as e:
            print(f"   ❌ Request error for {label}: {e}")
            return None
    print(f"   ❌ All {max_retries} retries exhausted for {label}")
    return None


def fetch_protocol_list(address: str, proxy: str = None, rate_limiter: "_RateLimiter" = None):
    """GET complex_protocol_list for one EVM address. Returns parsed JSON or None."""
    if not _is_evm(address):
        print(f"   ⚠️  Skipping non-EVM address: {address}")
        return None
    rl = rate_limiter or _rate_limiter
    return _fetch_with_retry(_PROTOCOL_URL, {"id": address}, proxy, rl, f"protocol/{address[-8:]}")


def fetch_app_list(address: str, proxy: str = None, rate_limiter: "_RateLimiter" = None):
    """GET complex_app_list for one EVM address. Returns parsed JSON or None."""
    if not _is_evm(address):
        print(f"   ⚠️  Skipping non-EVM address: {address}")
        return None
    rl = rate_limiter or _rate_limiter
    return _fetch_with_retry(_APP_URL, {"id": address}, proxy, rl, f"app/{address[-8:]}")


def fetch_all_batched(
    addresses: list,
    proxies: list = None,
    batch_size: int = 5,
) -> tuple:
    """
    Divide addresses into batches of `batch_size`, run each batch in parallel.
    Each batch routes through a different proxy IP so Rabby sees independent servers.
    Each batch has its own rate limiter (5 calls/min per IP).

    Returns: (protocol_rows, app_rows, proto_raw_entries, app_raw_entries)
    """
    if not addresses:
        return [], [], [], []

    # Split into batches
    batches = [addresses[i:i + batch_size] for i in range(0, len(addresses), batch_size)]
    proxy_list = proxies or []

    def _process_batch(batch_addrs: list, proxy: str, batch_idx: int):
        label = f"[Batch {batch_idx + 1}]"
        if proxy:
            print(f"   {label} Using proxy: {proxy.split('@')[-1] if '@' in proxy else proxy}")
        else:
            print(f"   {label} No proxy (direct connection)")

        # Each batch gets its own rate limiter — independent IP = independent quota
        rl = _RateLimiter(max_calls=5, period=60)

        batch_proto_rows, batch_app_rows = [], []
        batch_proto_raw, batch_app_raw = [], []

        for addr in batch_addrs:
            print(f"   {label} Fetching protocols for {addr[-8:]}...", flush=True)
            proto_data = fetch_protocol_list(addr, proxy=proxy, rate_limiter=rl)
            if proto_data is not None:
                batch_proto_rows.extend(parse_protocol_data(addr, proto_data))
            batch_proto_raw.append((addr, proto_data if proto_data is not None else []))

            print(f"   {label} Fetching apps     for {addr[-8:]}...", flush=True)
            app_data = fetch_app_list(addr, proxy=proxy, rate_limiter=rl)
            if app_data is not None:
                batch_app_rows.extend(parse_app_data(addr, app_data))
            batch_app_raw.append((addr, app_data if app_data is not None else []))

        return batch_proto_rows, batch_app_rows, batch_proto_raw, batch_app_raw

    all_proto_rows, all_app_rows, all_proto_raw, all_app_raw = [], [], [], []

    print(f"\n   📦 Batched fetch: {len(addresses)} addresses → {len(batches)} batch(es) of {batch_size}")

    with ThreadPoolExecutor(max_workers=len(batches)) as executor:
        futures = {
            executor.submit(
                _process_batch,
                batch,
                proxy_list[i] if i < len(proxy_list) else None,
                i,
            ): i
            for i, batch in enumerate(batches)
        }
        for future in as_completed(futures):
            try:
                p_rows, a_rows, p_raw, a_raw = future.result()
                all_proto_rows.extend(p_rows)
                all_app_rows.extend(a_rows)
                all_proto_raw.extend(p_raw)
                all_app_raw.extend(a_raw)
            except Exception as e:
                print(f"   ❌ Batch error: {e}")

    print(f"   ✅ Batched fetch done: {len(all_proto_rows)} protocol rows, {len(all_app_rows)} app rows")
    return all_proto_rows, all_app_rows, all_proto_raw, all_app_raw


def _calc_value(price, amount) -> str:
    """Return Price × Amount as a full-precision string with no trailing zeros."""
    try:
        result = Decimal(str(price)) * Decimal(str(amount))
        s = str(result).rstrip('0').rstrip('.')
        return s
    except Exception:
        return ""


def _extract_tokens_from_item(item: dict) -> list:
    """
    Return a list of (token_type, token_dict) pairs from a portfolio_item detail.
    token_type is 'supply', 'reward', or 'borrow'.
    """
    detail = item.get("detail", {}) or {}
    tokens = []
    for key in ("supply_token_list", "reward_token_list", "borrow_token_list", "asset_token_list"):
        kind = key.split("_")[0]  # supply / reward / borrow
        for tok in detail.get(key) or []:
            tokens.append((kind, tok))
    return tokens


def parse_protocol_data(address: str, response_data) -> list:
    """
    Parse complex_protocol_list response into a list of row dicts.

    Each row: {Address, Name, ID, Chain, Symbol, Price, Amount, Calculated_Value}
    """
    rows = []
    if not response_data:
        return rows

    # Response is a list of protocol objects
    protocols = response_data if isinstance(response_data, list) else []

    for protocol in protocols:
        name = protocol.get("name", "")
        chain = protocol.get("chain", "")

        for item in protocol.get("portfolio_item_list") or []:
            item_id = item.get("id", "")
            item_desc = ((item.get("detail") or {}).get("description") or
                         item.get("description") or "")
            item_pool_name = item.get("name", "")
            for _kind, tok in _extract_tokens_from_item(item):
                symbol = tok.get("symbol", "")
                price = tok.get("price", 0) or 0
                amount = tok.get("amount", 0) or 0
                tok_chain = tok.get("chain", chain) or chain
                calc = _calc_value(price, amount)
                rows.append({
                    "Address": address,
                    "Name": name,
                    "ID": item_id,
                    "Chain": tok_chain,
                    "Pool_Name": item_pool_name,
                    "Description": item_desc or tok.get("description", ""),
                    "Symbol": symbol,
                    "Price": price,
                    "Amount": amount,
                    "Calculated_Value": calc,
                })

    return rows


def parse_app_data(address: str, response_data) -> list:
    """
    Parse complex_app_list response into a list of row dicts.

    Each row: {Address, Name, ID, Chain, Symbol, Price, Amount, Calculated_Value}
    """
    rows = []
    if not response_data:
        return rows

    # complex_app_list returns {"apps": [...], "error_apps": [...]}
    if isinstance(response_data, dict):
        apps = response_data.get("apps") or []
    else:
        apps = response_data if isinstance(response_data, list) else []

    for app in apps:
        name = app.get("name", "")
        chain = app.get("chain", "")

        for item in app.get("portfolio_item_list") or []:
            item_id = item.get("id", "")
            item_desc = ((item.get("detail") or {}).get("description") or
                         item.get("description") or "")
            item_pool_name = item.get("name", "")

            if item_pool_name == "Perpetuals":
                detail = item.get("detail") or {}
                base_token = detail.get("base_token") or {}
                quote_token = detail.get("quote_token") or {}
                margin_token = detail.get("margin_token") or {}
                currency_pair = f"{base_token.get('symbol', '')}/{quote_token.get('symbol', '')}"
                price = margin_token.get("price", 0) or 0
                amount = margin_token.get("amount", 0) or 0
                calc = _calc_value(price, amount)
                leverage_raw = detail.get("leverage")
                leverage_str = f"{int(leverage_raw)}x" if leverage_raw is not None else ""
                rows.append({
                    "Address": address,
                    "Name": name,
                    "ID": item_id,
                    "Chain": chain,
                    "Pool_Name": item_pool_name,
                    "Description": item_desc,
                    "Side": detail.get("side", ""),
                    "Symbol": currency_pair,
                    "Leverage": leverage_str,
                    "PnL_USD": detail.get("pnl_usd_value", ""),
                    "Price": price,
                    "Amount": amount,
                    "Calculated_Value": calc,
                })
            else:
                for _kind, tok in _extract_tokens_from_item(item):
                    symbol = tok.get("symbol", "")
                    price = tok.get("price", 0) or 0
                    amount = tok.get("amount", 0) or 0
                    tok_chain = tok.get("chain", chain) or chain
                    calc = _calc_value(price, amount)
                    rows.append({
                        "Address": address,
                        "Name": name,
                        "ID": item_id,
                        "Chain": tok_chain,
                        "Pool_Name": item_pool_name,
                        "Description": item_desc or tok.get("description", ""),
                        "Symbol": symbol,
                        "Price": price,
                        "Amount": amount,
                        "Calculated_Value": calc,
                    })

    return rows


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


_URL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_URL_FONT = Font(bold=True, color="1F4E79")


def _write_raw_tab(ws, raw_entries: list):
    """Write 'Rabby Raw' tab: Address | JSON columns."""
    ws.append(["Address", "JSON Response"])
    ws["A1"].font = _HEADER_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["B1"].font = _HEADER_FONT
    ws["B1"].fill = _HEADER_FILL

    for address, data in raw_entries:
        ws.append([address, json.dumps(data, ensure_ascii=False)])

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 80
    ws["B2"].alignment = Alignment(wrap_text=False)


def _write_combined_raw_tab(ws, proto_entries: list, hl_entries: list):
    """
    Write a single Raw tab containing both Protocol and Hyperliquid responses.
    Before each JSON row, inserts a URL row showing the exact GET request sent.
    Structure per entry:
      Row: "GET" | full URL
      Row: address | JSON response
    """
    ws.append(["Address", "JSON Response"])
    ws["A1"].font = _HEADER_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["B1"].font = _HEADER_FONT
    ws["B1"].fill = _HEADER_FILL

    def append_entries(entries, base_url):
        for address, data in entries:
            url = f"{base_url}?id={address}"
            url_row = ws.max_row + 1
            ws.append(["GET", url])
            for col in (1, 2):
                cell = ws.cell(row=url_row, column=col)
                cell.font = _URL_FONT
                cell.fill = _URL_FILL
            ws.append([address, json.dumps(data, ensure_ascii=False)])

    append_entries(proto_entries, _PROTOCOL_URL)
    append_entries(hl_entries, _APP_URL)

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 100


def _write_parsed_tab(ws, rows: list, tab_type: str):
    """Write parsed data tab with 9 columns (Description inserted before Symbol)."""
    headers = ["Address", "Name", "ID", "Chain", "Pool Name", "Description", "Side", "Symbol/Currency Pair", "Leverage", "PnL (USD)", "Price", "Amount", "Calculated Value"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for row in rows:
        ws.append([
            row.get("Address", ""),
            row.get("Name", ""),
            row.get("ID", ""),
            row.get("Chain", ""),
            row.get("Pool_Name", ""),
            row.get("Description", ""),
            row.get("Side", ""),
            row.get("Symbol", ""),
            row.get("Leverage", ""),
            str(row["PnL_USD"]) if row.get("PnL_USD") is not None and row.get("PnL_USD") != "" else "",
            str(row["Price"]) if row.get("Price") is not None else "",
            str(row["Amount"]) if row.get("Amount") is not None else "",
            row.get("Calculated_Value", ""),
        ])

    # Auto-width
    col_widths = [46, 20, 20, 12, 18, 18, 10, 22, 12, 16, 16, 20, 24]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width


def _sanitize_sheet_name(name: str) -> str:
    """Limit sheet name to 31 chars and strip illegal characters."""
    illegal = r'[\\/*?\[\]:]'
    name = re.sub(illegal, '', name)
    return name[:31]


def export_protocol_excel(
    addresses: list,
    portfolio_name: str = "portfolio",
    output_dir: str = "test-results/excel-exports",
) -> str:
    """
    Call complex_protocol_list for each EVM address and export to Excel.

    Returns path to the created Excel file.
    """
    print(f"\n{'='*70}", flush=True)
    print(f"🔗 Rabby Protocol API — complex_protocol_list", flush=True)
    print(f"{'='*70}", flush=True)

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = re.sub(r'[^A-Za-z0-9_\-]', '_', portfolio_name)
    filename = f"Protocol_{safe_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)

    ws_raw = wb.create_sheet(_sanitize_sheet_name("Rabby Raw"))
    ws_parsed = wb.create_sheet(_sanitize_sheet_name("Rabby - Protocol Amount"))

    raw_entries = []
    all_rows = []

    for addr in addresses:
        if not _is_evm(addr):
            print(f"   ⏭️  Skipping non-EVM address: {addr}", flush=True)
            continue
        print(f"\n📡 Fetching protocol list for {addr}...", flush=True)
        data = fetch_protocol_list(addr)
        raw_entries.append((addr, data if data is not None else []))
        if data is not None:
            rows = parse_protocol_data(addr, data)
            all_rows.extend(rows)
            print(f"   ✅ Got {len(data) if isinstance(data, list) else 0} protocol(s), "
                  f"{len(rows)} token row(s)", flush=True)
        else:
            print(f"   ⚠️  No data returned for {addr}", flush=True)

    _write_raw_tab(ws_raw, raw_entries)
    _write_parsed_tab(ws_parsed, all_rows, "protocol")

    wb.save(filepath)
    print(f"\n✅ Protocol Excel saved: {filepath}", flush=True)
    print(f"   Total rows: {len(all_rows)}", flush=True)
    return filepath


def export_hyperliquid_excel(
    addresses: list,
    portfolio_name: str = "portfolio",
    output_dir: str = "test-results/excel-exports",
) -> str:
    """
    Call complex_app_list for each EVM address and export to Excel.

    Returns path to the created Excel file.
    """
    print(f"\n{'='*70}", flush=True)
    print(f"🔗 Rabby App API — complex_app_list (Hyperliquid)", flush=True)
    print(f"{'='*70}", flush=True)

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = re.sub(r'[^A-Za-z0-9_\-]', '_', portfolio_name)
    filename = f"Hyperliquid_{safe_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)

    ws_raw = wb.create_sheet(_sanitize_sheet_name("Hyperliquid Raw"))
    ws_parsed = wb.create_sheet(_sanitize_sheet_name("Hyperliquid - App Amount"))

    raw_entries = []
    all_rows = []

    for addr in addresses:
        if not _is_evm(addr):
            print(f"   ⏭️  Skipping non-EVM address: {addr}", flush=True)
            continue
        print(f"\n📡 Fetching app list for {addr}...", flush=True)
        data = fetch_app_list(addr)
        raw_entries.append((addr, data if data is not None else []))
        if data is not None:
            rows = parse_app_data(addr, data)
            all_rows.extend(rows)
            print(f"   ✅ Got {len(data) if isinstance(data, list) else 0} app(s), "
                  f"{len(rows)} token row(s)", flush=True)
        else:
            print(f"   ⚠️  No data returned for {addr}", flush=True)

    _write_raw_tab(ws_raw, raw_entries)
    _write_parsed_tab(ws_parsed, all_rows, "app")

    wb.save(filepath)
    print(f"\n✅ Hyperliquid Excel saved: {filepath}", flush=True)
    print(f"   Total rows: {len(all_rows)}", flush=True)
    return filepath


def export_combined_excel(
    addresses: list,
    portfolio_name: str = "portfolio",
    output_dir: str = "test-results/excel-exports",
) -> str:
    """
    Call both complex_protocol_list and complex_app_list for each EVM address
    and export all data into a single Excel file with 4 tabs:
      - Rabby Raw
      - Rabby - Protocol Amount
      - Hyperliquid Raw
      - Hyperliquid - App Amount

    Returns path to the created Excel file.
    """
    print(f"\n{'='*70}", flush=True)
    print(f"🔗 Rabby Combined Export (Protocol + Hyperliquid)", flush=True)
    print(f"{'='*70}", flush=True)

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = re.sub(r'[^A-Za-z0-9_\-]', '_', portfolio_name)
    filename = f"Protocol_{safe_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)

    ws_raw = wb.create_sheet(_sanitize_sheet_name("Rabby Raw"))
    ws_amount = wb.create_sheet(_sanitize_sheet_name("Rabby Api Data"))

    proto_raw_entries, proto_rows = [], []
    hl_raw_entries, hl_rows = [], []

    for addr in addresses:
        if not _is_evm(addr):
            print(f"   ⏭️  Skipping non-EVM address: {addr}", flush=True)
            continue

        print(f"\n📡 Fetching protocol list for {addr}...", flush=True)
        proto_data = fetch_protocol_list(addr)
        proto_raw_entries.append((addr, proto_data if proto_data is not None else []))
        if proto_data is not None:
            rows = parse_protocol_data(addr, proto_data)
            proto_rows.extend(rows)
            print(f"   ✅ Got {len(proto_data) if isinstance(proto_data, list) else 0} protocol(s), "
                  f"{len(rows)} token row(s)", flush=True)
        else:
            print(f"   ⚠️  No protocol data for {addr}", flush=True)

        print(f"\n📡 Fetching app list for {addr}...", flush=True)
        hl_data = fetch_app_list(addr)
        hl_raw_entries.append((addr, hl_data if hl_data is not None else []))
        if hl_data is not None:
            rows = parse_app_data(addr, hl_data)
            hl_rows.extend(rows)
            app_count = len(hl_data.get("apps", [])) if isinstance(hl_data, dict) else len(hl_data)
            print(f"   ✅ Got {app_count} app(s), {len(rows)} token row(s)", flush=True)
        else:
            print(f"   ⚠️  No app data for {addr}", flush=True)

    _write_combined_raw_tab(ws_raw, proto_raw_entries, hl_raw_entries)
    _write_parsed_tab(ws_amount, proto_rows + hl_rows, "combined")

    wb.save(filepath)
    print(f"\n✅ Combined Excel saved: {filepath}", flush=True)
    print(f"   Protocol rows: {len(proto_rows)} | Hyperliquid rows: {len(hl_rows)}", flush=True)
    return filepath
