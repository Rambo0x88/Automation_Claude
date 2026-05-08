"""
Fetch Coingecko Coin ID List and export to Excel.

Excel file: Coingecko Coin ID List.xlsx
  Tab 1: Raw Data - Full JSON response
  Tab 2: Coin ID List - id, symbol, name, platform name, platform address
"""
import requests
import json
from openpyxl import Workbook
from datetime import datetime


def fetch_coingecko_coin_list():
    """Fetch coin list from Coingecko API with platform info."""
    url = "https://api.coingecko.com/api/v3/coins/list?include_platform=true"
    headers = {
        "x-cg-demo-api-key": "CG-F3KENg4b1mcvyeg6eo6LGDQU"
    }

    print("Fetching Coingecko Coin ID List...")
    response = requests.get(url, headers=headers, timeout=60)
    response.raise_for_status()
    data = response.json()
    print(f"Fetched {len(data)} coins")
    return data


def export_to_excel(data):
    """Export coin list data to Excel with 2 tabs."""
    wb = Workbook()

    # Tab 1: Raw Data
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    ws_raw.cell(row=1, column=1, value="Raw API Response")
    ws_raw.cell(row=2, column=1, value=json.dumps(data, indent=2))

    # Tab 2: Coin ID List (extracted columns)
    ws_list = wb.create_sheet("Coin ID List")
    ws_list.cell(row=1, column=1, value="ID")
    ws_list.cell(row=1, column=2, value="Symbol")
    ws_list.cell(row=1, column=3, value="Name")
    ws_list.cell(row=1, column=4, value="Platform Name")
    ws_list.cell(row=1, column=5, value="Platform Address")

    row_num = 2
    for coin in data:
        coin_id = coin.get("id", "")
        symbol = coin.get("symbol", "")
        name = coin.get("name", "")
        platforms = coin.get("platforms", {})

        if platforms:
            for platform_name, platform_address in platforms.items():
                if platform_name and platform_address:
                    ws_list.cell(row=row_num, column=1, value=coin_id)
                    ws_list.cell(row=row_num, column=2, value=symbol)
                    ws_list.cell(row=row_num, column=3, value=name)
                    ws_list.cell(row=row_num, column=4, value=platform_name)
                    ws_list.cell(row=row_num, column=5, value=platform_address)
                    row_num += 1
        else:
            # Coin with no platform info - still include it
            ws_list.cell(row=row_num, column=1, value=coin_id)
            ws_list.cell(row=row_num, column=2, value=symbol)
            ws_list.cell(row=row_num, column=3, value=name)
            ws_list.cell(row=row_num, column=4, value="")
            ws_list.cell(row=row_num, column=5, value="")
            row_num += 1

    filename = "Coingecko Coin ID List.xlsx"
    wb.save(filename)
    print(f"Saved to {filename}")
    print(f"  Tab 1: Raw Data")
    print(f"  Tab 2: Coin ID List - {row_num - 2} rows")
    return filename


if __name__ == "__main__":
    data = fetch_coingecko_coin_list()
    export_to_excel(data)
