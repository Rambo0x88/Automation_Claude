#!/usr/bin/env python3
"""
Query SimDune API for a specific wallet address and export to Excel
Usage: python3 query_simdune_address.py <wallet_address>
"""

import subprocess
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import sys
import os

def query_simdune_api(wallet_address):
    """Query SimDune API for wallet balances"""
    print(f"🔗 Querying SimDune API for: {wallet_address}")

    try:
        result = subprocess.run([
            "curl", "-s",
            "-H", "User-Agent: insomnia/11.6.2",
            "-H", "X-Sim-Api-Key: sim_z82dBaaNeG3Y1elSbDfAGbaHYjb4RTyE",
            f"https://api.sim.dune.com/v1/evm/balances/{wallet_address}"
        ], capture_output=True, text=True, timeout=30)

        if result.returncode == 0:
            data = json.loads(result.stdout)

            # Check for error message
            if 'message' in data and 'balances' not in data:
                print(f"❌ API Error: {data['message']}")
                return None

            total_value = sum(float(b.get('value_usd', 0) or 0) for b in data.get('balances', []))
            token_count = len(data.get('balances', []))
            print(f"✅ SimDune: {token_count} tokens, ${total_value:,.2f}")
            return data
        else:
            print(f"⚠️  SimDune API error")
            return None
    except Exception as e:
        print(f"❌ SimDune API failed: {e}")
        return None

def export_to_excel(wallet_address, simdune_data):
    """Export SimDune API response to Excel"""
    print(f"\n📊 Creating Excel report...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: SimDune API Full Response (JSON)
    ws_json = wb.create_sheet("API Response (JSON)")
    ws_json.append(["Complete SimDune API Response"])
    ws_json.append([])
    ws_json.append([json.dumps(simdune_data, indent=2)])
    print(f"   ✅ Added full JSON response")

    # Sheet 2: SimDune Balances Table
    ws_balances = wb.create_sheet("Balances", 0)
    headers = ["Wallet", "Chain", "Chain_ID", "Symbol", "Name", "Amount", "Decimals", "Price_USD", "Value_USD", "Contract_Address"]
    ws_balances.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_balances[1]:
        cell.fill = header_fill
        cell.font = header_font

    if simdune_data and 'balances' in simdune_data:
        for balance in simdune_data['balances']:
            # Convert amount from wei to token amount
            decimals = int(balance.get('decimals', 18))
            amount_wei = float(balance.get('amount', 0))
            amount_tokens = amount_wei / (10 ** decimals)

            row = [
                wallet_address,
                balance.get('chain', ''),
                balance.get('chain_id', ''),
                balance.get('symbol', ''),
                balance.get('name', ''),
                amount_tokens,
                balance.get('decimals', ''),
                balance.get('price_usd', ''),
                balance.get('value_usd', ''),
                balance.get('address', '')  # Contract address
            ]
            ws_balances.append(row)

        print(f"   ✅ Added {len(simdune_data['balances'])} token balances")

    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_dir = "test-results/excel-reports"
    os.makedirs(excel_dir, exist_ok=True)

    filename = f"{excel_dir}/simdune_response_{wallet_address[:10]}_{timestamp}.xlsx"
    wb.save(filename)

    print(f"\n✅ Excel report saved: {filename}")
    return filename

def main():
    print("=" * 80)
    print("SIMDUNE API QUERY AND EXPORT")
    print("=" * 80)

    if len(sys.argv) < 2:
        print("\nUsage: python3 query_simdune_address.py <wallet_address>")
        print("\nExample:")
        print("  python3 query_simdune_address.py 0xE4a6a1AfeED468543CC766E7357Cabbe73ab6969")
        sys.exit(1)

    wallet_address = sys.argv[1]

    # Validate address format
    if not wallet_address.startswith('0x'):
        print(f"❌ Invalid address format. Must start with '0x'")
        sys.exit(1)

    if len(wallet_address) != 42:
        print(f"❌ Invalid address length. Expected 42 characters, got {len(wallet_address)}")
        print(f"   Address: {wallet_address}")
        sys.exit(1)

    print(f"\n📋 Wallet Address: {wallet_address}\n")

    # Query SimDune API
    simdune_data = query_simdune_api(wallet_address)

    if not simdune_data:
        print("\n❌ Failed to retrieve data from SimDune API")
        sys.exit(1)

    # Export to Excel
    excel_file = export_to_excel(wallet_address, simdune_data)

    print("\n" + "=" * 80)
    print("EXPORT COMPLETE")
    print("=" * 80)
    print(f"📁 File: {excel_file}")
    print("=" * 80)

if __name__ == "__main__":
    main()
