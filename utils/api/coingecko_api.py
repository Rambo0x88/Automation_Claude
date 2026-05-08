"""
CoinGecko API — Price fetching, 24H change, and coin ID lookup.

Usage:
    from utils.api.coingecko_api import (
        load_coingecko_coin_list,
        fetch_coingecko_prices_batch,
        fetch_coingecko_price_change_batch,
        CHAIN_TO_PLATFORM,
    )
"""

import os
import time
import requests


# API key for CoinGecko demo tier
COINGECKO_API_KEY = "CG-F3KENg4b1mcvyeg6eo6LGDQU"

# Chain name mapping: SimDune/DAM chain name -> CoinGecko platform name
CHAIN_TO_PLATFORM = {
    "ethereum": "ethereum",
    "binance smart chain": "binance-smart-chain",
    "bsc": "binance-smart-chain",
    "bnb": "binance-smart-chain",
    "polygon": "polygon-pos",
    "avalanche": "avalanche",
    "fantom": "fantom",
    "arbitrum": "arbitrum-one",
    "optimism": "optimistic-ethereum",
    "base": "base",
    "polygon zkevm": "polygon-zkevm",
    "linea": "linea",
    "scroll": "scroll",
    "zksync": "zksync",
    "gnosis": "xdai",
}


def load_coingecko_coin_list(coingecko_file="Coingecko Coin ID List.xlsx"):
    """
    Load CoinGecko Coin ID List from Excel and create lookup maps.

    Args:
        coingecko_file: Path to the Excel file (checks cwd first, then ../automation/)

    Returns:
        tuple: (coingecko_map, coingecko_native_map, coingecko_addr_map)
            - coingecko_map: {(platform_address_lower, platform_name_lower): coin_id}
            - coingecko_native_map: {symbol_lower: coin_id} for native tokens
            - coingecko_addr_map: {platform_address_lower: coin_id} address-only fallback
    """
    from openpyxl import load_workbook

    coingecko_map = {}
    coingecko_native_map = {}
    coingecko_addr_map = {}

    if not os.path.exists(coingecko_file):
        alt_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'Coingecko Coin ID List.xlsx')
        if os.path.exists(alt_path):
            coingecko_file = alt_path
        else:
            alt_path2 = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', '..', 'automation', 'Coingecko Coin ID List.xlsx')
            if os.path.exists(alt_path2):
                coingecko_file = alt_path2

    if not os.path.exists(coingecko_file):
        print(f"   ⚠️  {coingecko_file} not found, CoinGecko ID lookup will be skipped")
        return coingecko_map, coingecko_native_map, coingecko_addr_map

    try:
        print(f"   📂 Loading CoinGecko Coin ID List (~26k rows)...")
        wb = load_workbook(coingecko_file, read_only=True, data_only=True)
        ws = wb["Coin ID List"]

        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                coin_id = row[0]
                symbol = row[1]
                platform_name = row[3]
                platform_address = row[4]

                if coin_id:
                    if platform_name and platform_address:
                        key = (str(platform_address).lower().strip(), str(platform_name).lower().strip())
                        coingecko_map[key] = str(coin_id).strip()
                        row_count += 1

                    if symbol:
                        pn = str(platform_name).strip().lower() if platform_name else ""
                        if not pn or pn == "ethereum":
                            sk = str(symbol).lower().strip()
                            if sk not in coingecko_native_map or not pn:
                                coingecko_native_map[sk] = str(coin_id).strip()

                    if platform_address:
                        ak = str(platform_address).lower().strip()
                        if ak not in coingecko_addr_map:
                            coingecko_addr_map[ak] = str(coin_id).strip()

            if row_count > 0 and row_count % 5000 == 0:
                print(f"      ... loaded {row_count} mappings")

        wb.close()
        print(f"   ✅ Loaded CoinGecko: {len(coingecko_map)} platform, "
              f"{len(coingecko_native_map)} native, {len(coingecko_addr_map)} addr-only")
        return coingecko_map, coingecko_native_map, coingecko_addr_map
    except Exception as e:
        print(f"   ⚠️  Failed to load CoinGecko data: {e}")
        return {}, {}, {}


def fetch_coingecko_prices_batch(coin_ids, raw_collector=None):
    """
    Fetch USD prices for multiple coin IDs from CoinGecko simple/price API.

    Args:
        coin_ids: List of coin IDs to fetch
        raw_collector: Optional list to append raw API responses to

    Returns:
        dict: {coin_id: usd_price or None}
    """
    if not coin_ids:
        return {}

    price_map = {}
    headers = {"x-cg-demo-api-key": COINGECKO_API_KEY}

    print(f"   💰 Fetching USD prices for {len(coin_ids)} coin(s)...")

    batch_size = 250
    for batch_start in range(0, len(coin_ids), batch_size):
        batch_ids = coin_ids[batch_start:batch_start + batch_size]
        ids_param = ",".join(batch_ids)

        try:
            url = f"https://api.coingecko.com/api/v3/simple/price?vs_currencies=usd&ids={ids_param}"
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            data = response.json()

            if raw_collector is not None:
                raw_collector.append({"endpoint": "simple/price", "url": url, "response": data})

            for coin_id in batch_ids:
                if coin_id in data and "usd" in data[coin_id]:
                    price_map[coin_id] = data[coin_id]["usd"]
                else:
                    price_map[coin_id] = None

            print(f"      ... batch {batch_start + 1}-{min(batch_start + batch_size, len(coin_ids))}/{len(coin_ids)}")

            if batch_start + batch_size < len(coin_ids):
                time.sleep(1)

        except Exception as e:
            print(f"      ⚠️  Batch failed: {str(e)[:50]}")
            for coin_id in batch_ids:
                price_map[coin_id] = None

    ok = len([v for v in price_map.values() if v is not None])
    print(f"   ✅ Prices: {ok}/{len(coin_ids)}")
    return price_map


def fetch_coingecko_price_change_batch(coin_ids, raw_collector=None):
    """
    Fetch 24H price change for multiple coin IDs from CoinGecko coins/{id} API.

    Args:
        coin_ids: List of coin IDs to fetch
        raw_collector: Optional list to append raw API responses to

    Returns:
        dict: {coin_id: price_change_24h or "null" or None}
    """
    if not coin_ids:
        return {}

    price_change_map = {}
    headers = {"x-cg-demo-api-key": COINGECKO_API_KEY}

    print(f"   📊 Fetching 24H price change for {len(coin_ids)} coin(s)...")

    for idx, coin_id in enumerate(coin_ids, 1):
        try:
            url = f"https://api.coingecko.com/api/v3/coins/{coin_id}"
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            data = response.json()

            if raw_collector is not None:
                raw_collector.append({"endpoint": f"coins/{coin_id}", "url": url, "response": data})

            price_change = data.get("market_data", {}).get("price_change_percentage_24h")
            if "market_data" in data and "price_change_percentage_24h" in (data.get("market_data") or {}):
                price_change_map[coin_id] = price_change if price_change is not None else "null"
            else:
                price_change_map[coin_id] = None

            if idx % 10 == 0 or idx == len(coin_ids):
                print(f"      ... {idx}/{len(coin_ids)}")

            if idx < len(coin_ids):
                time.sleep(0.5)

        except Exception as e:
            print(f"      ⚠️  {coin_id}: {str(e)[:50]}")
            price_change_map[coin_id] = None

    ok = len([v for v in price_change_map.values() if v is not None])
    print(f"   ✅ 24H change: {ok}/{len(coin_ids)}")
    return price_change_map
