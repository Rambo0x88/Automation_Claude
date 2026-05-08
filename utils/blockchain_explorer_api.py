"""Blockchain Explorer API handlers for Etherscan, BSCScan, BaseScan."""

import requests
import time
from typing import Dict, List, Tuple, Optional
import json

# API Configuration
EXPLORER_APIS = {
    "ethereum": {
        "name": "Etherscan",
        "base_url": "https://api.etherscan.io/api",
        "chain_id": 1,
    },
    "bsc": {
        "name": "BSCScan",
        "base_url": "https://api.bscscan.com/api",
        "chain_id": 56,
    },
    "base": {
        "name": "BaseScan",
        "base_url": "https://api.basescan.org/api",
        "chain_id": 8453,
    },
}

# Rate limiting
RATE_LIMIT_DELAY = 0.2  # 5 calls/second = 0.2 seconds between calls
MAX_RETRIES = 3
RETRY_DELAY = 2


def _rate_limited_request(url: str, params: Dict, max_retries: int = MAX_RETRIES) -> Optional[Dict]:
    """Make rate-limited API request with retry logic."""
    for attempt in range(max_retries):
        try:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            # Check for API errors
            if data.get("status") == "0":
                error_msg = data.get("message", "Unknown error")
                if "rate limit" in error_msg.lower():
                    if attempt < max_retries - 1:
                        print(f"      ⚠️  Rate limited, retrying in {RETRY_DELAY}s...")
                        time.sleep(RETRY_DELAY)
                        continue
                    else:
                        print(f"      ❌ Rate limit exceeded after {max_retries} attempts")
                        return None
                else:
                    print(f"      ⚠️  API error: {error_msg}")
                    return None
            
            # Success
            time.sleep(RATE_LIMIT_DELAY)
            return data
            
        except requests.exceptions.RequestException as e:
            if attempt < max_retries - 1:
                print(f"      ⚠️  Request failed: {e}, retrying...")
                time.sleep(RETRY_DELAY)
            else:
                print(f"      ❌ Request failed after {max_retries} attempts: {e}")
                return None
    
    return None


def fetch_transactions(address: str, chain: str = "ethereum", api_key: str = None) -> Tuple[str, List[Dict], bool]:
    """
    Fetch all transactions for an address from blockchain explorer API.
    
    Args:
        address: EVM address (0x...)
        chain: "ethereum", "bsc", or "base"
        api_key: Explorer API key
    
    Returns:
        (address, transactions_list, success)
    """
    if not api_key:
        print(f"      ❌ No API key provided for {chain}")
        return (address, [], False)
    
    explorer = EXPLORER_APIS.get(chain)
    if not explorer:
        print(f"      ❌ Unknown chain: {chain}")
        return (address, [], False)
    
    print(f"      🔄 Fetching transactions from {explorer['name']}...")
    
    params = {
        "module": "account",
        "action": "txlist",
        "address": address,
        "startblock": 0,
        "endblock": 99999999,
        "sort": "asc",
        "apikey": api_key,
    }
    
    data = _rate_limited_request(explorer["base_url"], params)
    if not data or data.get("status") != "1":
        return (address, [], False)
    
    transactions = data.get("result", [])
    print(f"      ✅ Fetched {len(transactions)} transactions")
    return (address, transactions, True)


def fetch_internal_transactions(address: str, chain: str = "ethereum", api_key: str = None) -> Tuple[str, List[Dict], bool]:
    """
    Fetch all internal transactions for an address.
    
    Args:
        address: EVM address (0x...)
        chain: "ethereum", "bsc", or "base"
        api_key: Explorer API key
    
    Returns:
        (address, internal_transactions_list, success)
    """
    if not api_key:
        return (address, [], False)
    
    explorer = EXPLORER_APIS.get(chain)
    if not explorer:
        return (address, [], False)
    
    print(f"      🔄 Fetching internal transactions from {explorer['name']}...")
    
    params = {
        "module": "account",
        "action": "txlistinternal",
        "address": address,
        "startblock": 0,
        "endblock": 99999999,
        "sort": "asc",
        "apikey": api_key,
    }
    
    data = _rate_limited_request(explorer["base_url"], params)
    if not data or data.get("status") != "1":
        return (address, [], False)
    
    transactions = data.get("result", [])
    print(f"      ✅ Fetched {len(transactions)} internal transactions")
    return (address, transactions, True)


def fetch_token_transfers(address: str, chain: str = "ethereum", api_key: str = None) -> Tuple[str, List[Dict], bool]:
    """
    Fetch all ERC-20 token transfers for an address.
    
    Args:
        address: EVM address (0x...)
        chain: "ethereum", "bsc", or "base"
        api_key: Explorer API key
    
    Returns:
        (address, token_transfers_list, success)
    """
    if not api_key:
        return (address, [], False)
    
    explorer = EXPLORER_APIS.get(chain)
    if not explorer:
        return (address, [], False)
    
    print(f"      🔄 Fetching ERC-20 token transfers from {explorer['name']}...")
    
    params = {
        "module": "account",
        "action": "tokentx",
        "address": address,
        "startblock": 0,
        "endblock": 99999999,
        "sort": "asc",
        "apikey": api_key,
    }
    
    data = _rate_limited_request(explorer["base_url"], params)
    if not data or data.get("status") != "1":
        return (address, [], False)
    
    transfers = data.get("result", [])
    print(f"      ✅ Fetched {len(transfers)} ERC-20 token transfers")
    return (address, transfers, True)


def fetch_nft_transfers(address: str, chain: str = "ethereum", api_key: str = None) -> Tuple[str, List[Dict], bool]:
    """
    Fetch all NFT transfers (ERC-721 + ERC-1155) for an address.
    
    Args:
        address: EVM address (0x...)
        chain: "ethereum", "bsc", or "base"
        api_key: Explorer API key
    
    Returns:
        (address, nft_transfers_list, success)
    """
    if not api_key:
        return (address, [], False)
    
    explorer = EXPLORER_APIS.get(chain)
    if not explorer:
        return (address, [], False)
    
    print(f"      🔄 Fetching NFT transfers from {explorer['name']}...")
    
    params = {
        "module": "account",
        "action": "tokennfttx",
        "address": address,
        "startblock": 0,
        "endblock": 99999999,
        "sort": "asc",
        "apikey": api_key,
    }
    
    data = _rate_limited_request(explorer["base_url"], params)
    if not data or data.get("status") != "1":
        return (address, [], False)
    
    transfers = data.get("result", [])
    print(f"      ✅ Fetched {len(transfers)} NFT transfers")
    return (address, transfers, True)


def fetch_all_explorer_data(address: str, chain: str = "ethereum", api_key: str = None) -> Dict:
    """
    Fetch all available data for an address from blockchain explorer.
    
    Args:
        address: EVM address (0x...)
        chain: "ethereum", "bsc", or "base"
        api_key: Explorer API key
    
    Returns:
        {
            "address": address,
            "chain": chain,
            "transactions": [...],
            "internal_transactions": [...],
            "token_transfers": [...],
            "nft_transfers": [...],
            "success": bool
        }
    """
    explorer = EXPLORER_APIS.get(chain)
    if not explorer:
        return {
            "address": address,
            "chain": chain,
            "transactions": [],
            "internal_transactions": [],
            "token_transfers": [],
            "nft_transfers": [],
            "success": False,
        }
    
    print(f"\n📊 Fetching all data for {address} from {explorer['name']}...")
    
    # Fetch all data types in parallel (or sequentially with rate limiting)
    _, transactions, tx_success = fetch_transactions(address, chain, api_key)
    _, internal_txs, itx_success = fetch_internal_transactions(address, chain, api_key)
    _, token_transfers, tt_success = fetch_token_transfers(address, chain, api_key)
    _, nft_transfers, nft_success = fetch_nft_transfers(address, chain, api_key)
    
    success = tx_success or itx_success or tt_success or nft_success
    
    return {
        "address": address,
        "chain": chain,
        "transactions": transactions,
        "internal_transactions": internal_txs,
        "token_transfers": token_transfers,
        "nft_transfers": nft_transfers,
        "success": success,
    }


def export_explorer_data_to_excel(explorer_data: Dict, output_path: str) -> bool:
    """
    Export blockchain explorer data to Excel file.
    
    Args:
        explorer_data: Data from fetch_all_explorer_data()
        output_path: Path to save Excel file
    
    Returns:
        True if successful, False otherwise
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        address = explorer_data["address"]
        chain = explorer_data["chain"]
        explorer = EXPLORER_APIS.get(chain, {})
        
        # Sheet 1: Transactions
        if explorer_data["transactions"]:
            ws = wb.create_sheet("Transactions")
            _write_transactions_sheet(ws, explorer_data["transactions"])
        
        # Sheet 2: Internal Transactions
        if explorer_data["internal_transactions"]:
            ws = wb.create_sheet("Internal Transactions")
            _write_internal_transactions_sheet(ws, explorer_data["internal_transactions"])
        
        # Sheet 3: Token Transfers
        if explorer_data["token_transfers"]:
            ws = wb.create_sheet("Token Transfers")
            _write_token_transfers_sheet(ws, explorer_data["token_transfers"])
        
        # Sheet 4: NFT Transfers
        if explorer_data["nft_transfers"]:
            ws = wb.create_sheet("NFT Transfers")
            _write_nft_transfers_sheet(ws, explorer_data["nft_transfers"])
        
        # Sheet 5: Summary
        ws = wb.create_sheet("Summary", 0)
        _write_summary_sheet(ws, explorer_data)
        
        wb.save(output_path)
        print(f"   ✅ Exported to {output_path}")
        return True
        
    except Exception as e:
        print(f"   ❌ Error exporting to Excel: {e}")
        return False


def _write_summary_sheet(ws, explorer_data: Dict):
    """Write summary sheet with statistics."""
    from openpyxl.styles import Font, Alignment
    
    ws.append(["Blockchain Explorer Data Export"])
    ws.append([])
    ws.append(["Address", explorer_data["address"]])
    ws.append(["Chain", explorer_data["chain"]])
    ws.append([])
    ws.append(["Data Summary"])
    ws.append(["Transactions", len(explorer_data["transactions"])])
    ws.append(["Internal Transactions", len(explorer_data["internal_transactions"])])
    ws.append(["Token Transfers", len(explorer_data["token_transfers"])])
    ws.append(["NFT Transfers", len(explorer_data["nft_transfers"])])
    ws.append([])
    ws.append(["Total Records", 
               len(explorer_data["transactions"]) + 
               len(explorer_data["internal_transactions"]) + 
               len(explorer_data["token_transfers"]) + 
               len(explorer_data["nft_transfers"])])
    
    # Format
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50


def _write_transactions_sheet(ws, transactions: List[Dict]):
    """Write transactions to sheet."""
    if not transactions:
        return
    
    # Header
    headers = list(transactions[0].keys())
    ws.append(headers)
    
    # Data
    for tx in transactions:
        row = [tx.get(h, "") for h in headers]
        ws.append(row)
    
    # Auto-fit columns
    for col in ws.columns:
        max_length = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)


def _write_internal_transactions_sheet(ws, transactions: List[Dict]):
    """Write internal transactions to sheet."""
    if not transactions:
        return
    
    headers = list(transactions[0].keys())
    ws.append(headers)
    
    for tx in transactions:
        row = [tx.get(h, "") for h in headers]
        ws.append(row)
    
    for col in ws.columns:
        max_length = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)


def _write_token_transfers_sheet(ws, transfers: List[Dict]):
    """Write token transfers to sheet."""
    if not transfers:
        return
    
    headers = list(transfers[0].keys())
    ws.append(headers)
    
    for transfer in transfers:
        row = [transfer.get(h, "") for h in headers]
        ws.append(row)
    
    for col in ws.columns:
        max_length = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)


def _write_nft_transfers_sheet(ws, transfers: List[Dict]):
    """Write NFT transfers to sheet."""
    if not transfers:
        return
    
    headers = list(transfers[0].keys())
    ws.append(headers)
    
    for transfer in transfers:
        row = [transfer.get(h, "") for h in headers]
        ws.append(row)
    
    for col in ws.columns:
        max_length = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)
