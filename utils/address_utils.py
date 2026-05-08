"""
Address Utilities — Shared address detection, validation, and lookup functions.

Reusable across all scripts that work with wallet addresses.

Usage:
    from utils.address_utils import (
        is_valid_evm_address,
        is_valid_trx_address,
        classify_address,
        detect_evm_addresses,
        detect_tron_addresses,
        clean_currency_symbols,
        lookup_portfolio_in_excel,
        lookup_address_in_excel,
    )
"""

import os
import re


def clean_currency_symbols(text):
    """Remove $ and % symbols from text data."""
    if isinstance(text, str):
        return text.replace('$', '').replace('%', '').strip()
    return text


def is_valid_evm_address(address_str):
    """
    Check if address is a valid EVM address (42 chars, starts with 0x).

    Args:
        address_str: Address string to validate

    Returns:
        bool: True if valid EVM address
    """
    if not address_str:
        return False
    address_str = str(address_str).strip()
    return len(address_str) == 42 and address_str.lower().startswith("0x")


def is_valid_trx_address(address_str):
    """
    Check if address is a valid TRON address (34 chars, starts with T, alphanumeric).

    Args:
        address_str: Address string to validate

    Returns:
        bool: True if valid TRX address
    """
    if not address_str:
        return False
    address_str = str(address_str).strip()
    return len(address_str) == 34 and address_str.startswith("T") and address_str.isalnum()


def classify_address(addr):
    """
    Classify an address as EVM, TRX, or unknown.

    Args:
        addr: Address string

    Returns:
        str: "evm", "trx", or "unknown"
    """
    addr = str(addr).strip()
    if is_valid_evm_address(addr):
        return "evm"
    elif is_valid_trx_address(addr):
        return "trx"
    return "unknown"


def detect_evm_addresses(data):
    """
    Detect EVM addresses (0x...) from a list of rows.
    First column of each row is checked.

    Args:
        data: List of rows (lists), where row[0] may contain an address

    Returns:
        list: Unique EVM addresses found (case-insensitive deduplication)
    """
    seen = set()
    addresses = []
    for row in data:
        if len(row) >= 1:
            address = str(row[0]).strip()
            if is_valid_evm_address(address):
                if address.lower() not in seen:
                    seen.add(address.lower())
                    addresses.append(address)
    return addresses


def detect_tron_addresses(data):
    """
    Detect Tron addresses (T..., 34 chars) from a list of rows.
    First column of each row is checked.

    Args:
        data: List of rows (lists), where row[0] may contain an address

    Returns:
        list: Unique Tron addresses found (case-insensitive deduplication)
    """
    seen = set()
    addresses = []
    for row in data:
        if len(row) >= 1:
            address = str(row[0]).strip()
            if is_valid_trx_address(address):
                if address.lower() not in seen:
                    seen.add(address.lower())
                    addresses.append(address)
    return addresses


def lookup_portfolio_in_excel(portfolio_name, excel_path="test_data/DAM addresses.xlsx"):
    """
    Look up a portfolio name in the DAM addresses Excel file.

    Args:
        portfolio_name: Portfolio name to search for
        excel_path: Path to the Excel file

    Returns:
        dict or None: Portfolio info dict with 'name', 'addresses', etc.
    """
    if not os.path.exists(excel_path):
        print(f"   ⚠️  Excel file not found: {excel_path}")
        return None

    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2:
                name = str(row[0]).strip() if row[0] else ""
                if name.lower() == portfolio_name.lower():
                    addresses = []
                    for cell in row[1:]:
                        if cell:
                            addr = str(cell).strip()
                            if addr and (is_valid_evm_address(addr) or is_valid_trx_address(addr)):
                                addresses.append(addr)
                    wb.close()
                    return {"name": name, "addresses": addresses}

        wb.close()
        return None
    except Exception as e:
        print(f"   ⚠️  Error reading Excel: {e}")
        return None


def lookup_address_in_excel(address, excel_path="test_data/DAM addresses.xlsx"):
    """
    Look up an address in the DAM addresses Excel file to find its portfolio.

    Args:
        address: Wallet address to search for
        excel_path: Path to the Excel file

    Returns:
        dict or None: Portfolio info dict with 'name', 'addresses', etc.
    """
    if not os.path.exists(excel_path):
        return None

    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        address_lower = address.lower()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row:
                for cell in row[1:]:
                    if cell and str(cell).strip().lower() == address_lower:
                        name = str(row[0]).strip() if row[0] else ""
                        addresses = []
                        for c in row[1:]:
                            if c:
                                addr = str(c).strip()
                                if addr and (is_valid_evm_address(addr) or is_valid_trx_address(addr)):
                                    addresses.append(addr)
                        wb.close()
                        return {"name": name, "addresses": addresses}

        wb.close()
        return None
    except Exception as e:
        return None
