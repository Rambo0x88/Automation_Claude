import pytest
import requests
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from datetime import datetime
import os
import glob


def test_trx_balance():
    """
    Test TRX Balance API endpoints and export to Excel
    Based on requirements in TRX Balance.md
    """

    # Input address - can be modified
    address = "TGn1uvntAVntT1pG8o7qoKkbViiYfeg6Gj"

    print(f"\n{'='*80}", flush=True)
    print(f"TRX Balance API Test", flush=True)
    print(f"{'='*80}", flush=True)
    print(f"Address: {address}\n", flush=True)

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Initialize token_info_responses for storing API responses from Contract Address's Token API
    token_info_responses = []

    # ========================================================================
    # STEP 1: LOAD EXISTING TOKEN LIST FROM "Token List" FOLDER
    # ========================================================================
    print(f"1️⃣  Loading existing Token List from 'Token List' folder...", flush=True)

    # Build contract address -> token info mapping (abbr, decimal, symbolShow, and canShow)
    contract_to_token = {}  # {contract_address: abbr}
    contract_to_decimal = {}  # {contract_address: decimal}
    contract_to_symbolshow = {}  # {contract_address: symbolShow}
    contract_to_canshow = {}  # {contract_address: canShow from Token List API}

    # Check if Token List folder exists and load "Token List.xlsx"
    token_list_folder = os.path.join(os.getcwd(), "Token List")
    token_list_file = os.path.join(token_list_folder, "Token List.xlsx")

    if os.path.exists(token_list_file):
        print(f"   📂 Loading existing Token List.xlsx...", flush=True)
        try:
            wb_existing = load_workbook(token_list_file, read_only=True, data_only=True)
            ws_existing = wb_existing.active

            # Load existing tokens (skip header row)
            row_count = 0
            for row in ws_existing.iter_rows(min_row=2, values_only=True):
                abbr, decimal, contract_addr = row[0], row[1], row[2]
                symbol_show = row[3] if len(row) > 3 else ''
                if contract_addr:
                    if abbr:
                        contract_to_token[contract_addr] = abbr
                    if decimal != '' and decimal is not None:
                        contract_to_decimal[contract_addr] = str(decimal)
                    if symbol_show != '' and symbol_show is not None:
                        contract_to_symbolshow[contract_addr] = symbol_show
                row_count += 1
                if row_count % 1000 == 0:
                    print(f"   ⏳ Processed {row_count} rows...", flush=True)

            wb_existing.close()
            print(f"   ✅ Loaded {len(contract_to_token)} tokens from Token List.xlsx")
        except Exception as e:
            print(f"   ⚠️  Error loading Token List.xlsx: {str(e)}")
    else:
        print(f"   ⚠️  Token List.xlsx does not exist in 'Token List' folder")

    # ========================================================================
    # STEP 2: FETCH TOKEN LIST API (to supplement mapping)
    # ========================================================================
    print(f"\n2️⃣  Fetching Token List API for supplemental mapping...")

    token_list_url = "https://apilist.tronscanapi.com/api/tokens/overview?start=0&limit=500&verifier=all&order=desc&filter=top&sort=&showAll=1&field="
    token_data = None

    try:
        token_response = requests.get(token_list_url, timeout=30)
        token_response.raise_for_status()
        token_data = token_response.json()

        tokens = token_data.get('tokens', [])
        if tokens:
            new_tokens_added = 0
            for token in tokens:
                contract_addr = token.get('contractAddress', '')
                abbr = token.get('abbr', '')
                decimal = token.get('decimal', '')
                can_show = token.get('canShow', '')

                if contract_addr:
                    # Add to mapping if not already present
                    if abbr and contract_addr not in contract_to_token:
                        contract_to_token[contract_addr] = abbr
                        new_tokens_added += 1
                    if decimal != '' and contract_addr not in contract_to_decimal:
                        contract_to_decimal[contract_addr] = decimal
                    # Store canShow from Token List API
                    if can_show != '' and can_show is not None:
                        contract_to_canshow[contract_addr] = can_show

            print(f"   ✅ Token List fetched: {len(contract_to_token)} token mappings")
        else:
            print(f"   ⚠️  No tokens found in Token List")

    except Exception as e:
        print(f"   ⚠️  Error fetching Token List: {str(e)}")

    # ========================================================================
    # STEP 3: ACCOUNT BALANCE
    # ========================================================================
    print(f"\n3️⃣  Fetching Account Balance...")

    balance_url = f"https://api.trongrid.io/v1/accounts/{address}"

    try:
        balance_response = requests.get(balance_url, timeout=30)
        balance_response.raise_for_status()
        balance_data = balance_response.json()

        print(f"   ✅ Account Balance API response received")

        # Tab 1: API - TRX Balance (Full API Response)
        ws_balance_api = wb.create_sheet("API - TRX Balance")
        ws_balance_api.append(["Raw API Response"])
        ws_balance_api.append([json.dumps(balance_data, indent=2)])

        # Tab 2: TRX Balance, Price (Parsed Data)
        ws_balance = wb.create_sheet("TRX Balance, Price")

        # Extract TRX Balance and TRC20 balances
        account_data = balance_data.get('data', [])

        if account_data and len(account_data) > 0:
            account = account_data[0]

            # Use the input address (not the hex address from API response)
            account_address = address

            # TRX Balance (in SUN, 1 TRX = 1,000,000 SUN)
            # Calculate Balance (Raw) = balance + frozen_balance + unfrozenV2.unfreeze_amount +
            #                           delegated_frozenV2_balance_for_energy + delegated_frozenV2_balance_for_bandwidth + frozenV2.amount
            balance_sun = account.get('balance', 0)

            # Get frozen_balance from frozen array
            frozen_balance = 0
            frozen = account.get('frozen', [])
            if isinstance(frozen, list):
                for item in frozen:
                    if isinstance(item, dict):
                        frozen_balance += item.get('frozen_balance', 0)

            # Get unfreeze_amount from unfrozenV2 array
            unfreeze_amount = 0
            unfrozen_v2 = account.get('unfrozenV2', [])
            if isinstance(unfrozen_v2, list):
                for item in unfrozen_v2:
                    if isinstance(item, dict):
                        unfreeze_amount += item.get('unfreeze_amount', 0)

            # Get delegated frozen balances
            # Note: delegated_frozenV2_balance_for_energy is nested in account_resource
            account_resource = account.get('account_resource', {})
            delegated_frozen_v2_energy = account_resource.get('delegated_frozenV2_balance_for_energy', 0)
            # delegated_frozenV2_balance_for_bandwidth is at root level
            delegated_frozen_v2_bandwidth = account.get('delegated_frozenV2_balance_for_bandwidth', 0)

            # Get amount from frozenV2 array
            frozen_v2_amount = 0
            frozen_v2 = account.get('frozenV2', [])
            if isinstance(frozen_v2, list):
                for item in frozen_v2:
                    if isinstance(item, dict):
                        frozen_v2_amount += item.get('amount', 0)

            trx_balance_sun = (balance_sun + frozen_balance + unfreeze_amount +
                             delegated_frozen_v2_energy + delegated_frozen_v2_bandwidth + frozen_v2_amount)

            # Store Balance (Raw) as text
            trx_balance_sun_text = str(int(trx_balance_sun))

            # Balance = (Balance (Raw)*1) × (10^-6)
            trx_balance = int(trx_balance_sun) * (10 ** -6)

            # Headers: Address, Balance, Balance (Raw), Decimal Places, Contract Token, Contract Address, Contract Balance, Contract Balance (Raw), Price, Price (24h), Symbol Show
            headers = ["Address", "Balance", "Balance (Raw)", "Decimal Places", "Contract Token", "Contract Address", "Contract Balance", "Contract Balance (Raw)", "Price", "Price (24h)", "Symbol Show"]
            ws_balance.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws_balance[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Add TRX balance row
            # For TRX: Address, Balance (calculated), Balance (Raw) as text, empty Decimal Places, empty Contract Token, empty Contract Address, empty Contract Balance, empty Contract Balance (Raw), empty Price, empty Price (24h), empty Symbol Show
            ws_balance.append([account_address, trx_balance, trx_balance_sun_text, "", "", "", "", "", "", "", ""])

            # Set Balance (Raw) column (column C, row 2) as text format
            ws_balance.cell(row=2, column=3).number_format = '@'

            # Add TRC20 tokens
            trc20_tokens = account.get('trc20', [])

            # trc20 is a list of dictionaries, each dict has contract_address: balance
            trc20_count = 0
            missing_tokens_fetched = 0

            if isinstance(trc20_tokens, list):
                for token_dict in trc20_tokens:
                    if isinstance(token_dict, dict):
                        # Each dict contains one or more contract addresses as keys
                        for contract_address, contract_balance_raw_str in token_dict.items():
                            # Get decimal places from mapping
                            decimal_places = contract_to_decimal.get(contract_address, "")
                            contract_token = contract_to_token.get(contract_address, "")

                            # Initialize price and symbol_show variables
                            price = ""
                            price_24h = ""
                            symbol_show = ""

                            # Symbol Show Priority: Token List.xlsx → All Token Info's canShow → TRC20 Token Detail's symbolShow
                            # First, check if symbolShow exists in Token List.xlsx
                            if contract_address in contract_to_symbolshow:
                                symbol_show = contract_to_symbolshow[contract_address]
                            # Second, check if canShow exists in All Token Info API
                            elif contract_address in contract_to_canshow:
                                symbol_show = contract_to_canshow[contract_address]

                            # Fetch token info from Contract Address's Token API for price data and symbol_show
                            # This is done for all tokens to get price information
                            try:
                                token_api_url = f"https://apilist.tronscanapi.com/api/token_trc20?contract={contract_address}&showAll=1&start=&limit="
                                token_api_response = requests.get(token_api_url, timeout=10)
                                token_api_response.raise_for_status()
                                token_api_data = token_api_response.json()

                                # Store API response for API - TRC20 Token Detail tab
                                token_info_responses.append({
                                    'contract_address': contract_address,
                                    'response': token_api_data
                                })

                                # Extract symbol, decimals, priceInUsd, gain, and symbol_show from trc20_tokens array
                                trc20_tokens_data = token_api_data.get('trc20_tokens', [])
                                if trc20_tokens_data and len(trc20_tokens_data) > 0:
                                    token_info = trc20_tokens_data[0]
                                    if decimal_places == "":
                                        decimal_places = token_info.get('decimals', '')
                                        if decimal_places != '':
                                            contract_to_decimal[contract_address] = decimal_places
                                    if contract_token == "":
                                        contract_token = token_info.get('symbol', '')
                                        if contract_token != '':
                                            contract_to_token[contract_address] = contract_token

                                    # Get price data from market_info object
                                    market_info = token_info.get('market_info', {})
                                    if isinstance(market_info, dict):
                                        price = market_info.get('priceInUsd', '')
                                        gain = market_info.get('gain', '')
                                        if gain != '' and gain is not None:
                                            try:
                                                price_24h = float(gain) * 100
                                            except:
                                                price_24h = ""

                                    # Get symbolShow from TRC20 Token Detail API (lowest priority)
                                    # Only use if not already set from Token List.xlsx or canShow
                                    if symbol_show == '':
                                        api_symbol_show = token_info.get('symbolShow', '')
                                        if api_symbol_show != '':
                                            symbol_show = api_symbol_show
                                            contract_to_symbolshow[contract_address] = api_symbol_show

                                    if decimal_places == "" or contract_token == "":
                                        missing_tokens_fetched += 1
                                        print(f"   ✅ Fetched token info: {contract_token} (Decimals: {decimal_places})")

                            except Exception as e:
                                print(f"   ⚠️  Error fetching token info for {contract_address}: {str(e)}")

                            # Convert contract balance string to number and calculate balance
                            try:
                                # Store raw balance as text directly (avoid float conversion for precision)
                                contract_balance_raw = contract_balance_raw_str.strip()

                                # Calculate Contract Balance using decimal places
                                # Contract Balance = Contract Balance (Raw) * 10^(-decimal_places)
                                if decimal_places != '':
                                    try:
                                        decimal_int = int(decimal_places)
                                        # Use int for raw value to maintain precision, then convert to float for calculation
                                        contract_balance_raw_int = int(contract_balance_raw)
                                        contract_balance = contract_balance_raw_int * (10 ** (-decimal_int))
                                    except:
                                        contract_balance = ""
                                else:
                                    contract_balance = ""
                            except:
                                contract_balance_raw = contract_balance_raw_str
                                contract_balance = ""

                            # For TRC20: Address, empty Balance, empty Balance (Raw), Decimal Places, Contract Token, Contract Address, Contract Balance (calculated), Contract Balance (Raw) as text, Price, Price (24h), Symbol Show
                            ws_balance.append([account_address, "", "", decimal_places, contract_token, contract_address, contract_balance, contract_balance_raw, price, price_24h, symbol_show])

                            # Format Contract Balance (Raw) column (column H) as text format
                            current_row = ws_balance.max_row
                            ws_balance.cell(row=current_row, column=8).number_format = '@'

                            # Format Contract Balance column (column G) to number format with 8 decimal places
                            if contract_balance != "":
                                ws_balance.cell(row=current_row, column=7).number_format = '0.00000000'

                            trc20_count += 1
                print(f"   ✅ TRX Balance: {trx_balance:,.6f} TRX")
                print(f"   ✅ TRC20 Tokens: {trc20_count} token(s)")
                if missing_tokens_fetched > 0:
                    print(f"   ✅ Missing tokens fetched from Contract Address's Token API: {missing_tokens_fetched}")
            else:
                print(f"   ✅ TRX Balance: {trx_balance:,.6f} TRX")
                print(f"   ⚠️  No TRC20 tokens found")

            # TRC10/AssetV2 support removed as per updated requirements
            # Only TRC20 tokens are processed
        else:
            print(f"   ⚠️  No account data found")
            ws_balance.append(["No account data available"])

    except Exception as e:
        print(f"   ❌ Error fetching Account Balance: {str(e)}")
        ws_balance_api = wb.create_sheet("API - TRX Balance")
        ws_balance_api.append(["Error", str(e)])
        ws_balance = wb.create_sheet("TRX Balance, Price")
        ws_balance.append(["Error", str(e)])

    # ========================================================================
    # STEP 4: TRANSACTION
    # ========================================================================
    print(f"\n4️⃣  Fetching Transactions...")

    transaction_url = f"https://api.trongrid.io/v1/accounts/{address}/transactions"

    try:
        transaction_response = requests.get(transaction_url, timeout=30)
        transaction_response.raise_for_status()
        transaction_data = transaction_response.json()

        print(f"   ✅ Transaction API response received")

        # Tab 1: API - TRX Transaction (Full API Response)
        ws_transaction_api = wb.create_sheet("API - TRX Transaction")
        ws_transaction_api.append(["Raw API Response"])
        ws_transaction_api.append([json.dumps(transaction_data, indent=2)])

        # Tab 2: Transaction Data (Parsed)
        ws_transaction = wb.create_sheet("Transaction")

        transactions = transaction_data.get('data', [])

        if transactions and len(transactions) > 0:
            # Get all unique keys from all transactions
            all_keys = set()
            for tx in transactions:
                all_keys.update(flatten_dict(tx).keys())

            headers = sorted(list(all_keys))
            ws_transaction.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws_transaction[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Add transaction rows
            for tx in transactions:
                flattened = flatten_dict(tx)
                row = [flattened.get(key, "") for key in headers]
                ws_transaction.append(row)

            print(f"   ✅ Transactions: {len(transactions)} transaction(s)")
        else:
            print(f"   ⚠️  No transactions found")
            ws_transaction.append(["No transactions available"])

    except Exception as e:
        print(f"   ❌ Error fetching Transactions: {str(e)}")
        ws_transaction_api = wb.create_sheet("API - TRX Transaction")
        ws_transaction_api.append(["Error", str(e)])
        ws_transaction = wb.create_sheet("Transaction")
        ws_transaction.append(["Error", str(e)])

    # ========================================================================
    # STEP 5: TOKEN LIST (Export to Excel)
    # ========================================================================
    print(f"\n5️⃣  Exporting Token List to Excel...")

    # Note: token_data was already fetched at the beginning
    if token_data:
        print(f"   ✅ Token List data available")

        # Tab 1: API - All Token Info (Full API Response)
        ws_token_api = wb.create_sheet("API - All Token Info")
        ws_token_api.append(["Raw API Response"])
        ws_token_api.append([json.dumps(token_data, indent=2)])

        # Tab 2: Token List (Parsed - only from Token List API)
        ws_token = wb.create_sheet("Token List")

        tokens = token_data.get('tokens', [])

        if tokens and len(tokens) > 0:
            # Export abbr, decimal, contractAddress, canShow columns
            headers = ["abbr", "decimal", "contractAddress", "canShow"]
            ws_token.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws_token[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Add token rows with abbr, decimal, contractAddress, canShow (from Token List API)
            for token in tokens:
                abbr = token.get('abbr', '')
                decimal = token.get('decimal', '')
                contract_address = token.get('contractAddress', '')
                can_show = token.get('canShow', '')
                ws_token.append([abbr, decimal, contract_address, can_show])

            print(f"   ✅ Tokens: {len(tokens)} token(s) exported to Excel")
        else:
            print(f"   ⚠️  No tokens found")
            ws_token.append(["No tokens available"])

    else:
        print(f"   ⚠️  No token data available")
        ws_token_api = wb.create_sheet("API - All Token Info")
        ws_token_api.append(["No token data available"])
        ws_token = wb.create_sheet("Token List")
        ws_token.append(["No tokens available"])

    # ========================================================================
    # STEP 6: API - TOKEN INFO (TRC20 Token Detail API responses)
    # ========================================================================
    if token_info_responses:
        print(f"\n6️⃣  Exporting TRC20 Token Detail to Excel...")

        ws_token_info = wb.create_sheet("API - TRC20 Token Detail")
        ws_token_info.append(["Contract Address", "Full API Response"])

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_token_info[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Add API responses for each contract address
        for token_info in token_info_responses:
            contract_addr = token_info['contract_address']
            response = token_info['response']
            ws_token_info.append([contract_addr, json.dumps(response, indent=2)])

        print(f"   ✅ API - TRC20 Token Detail tab created with {len(token_info_responses)} response(s)")

    # ========================================================================
    # STEP 7: API - TRC10 TOKEN DETAIL (TRC10 Token Detail API responses)
    # TRC10 Token Detail export removed as per updated requirements

    # ========================================================================
    # SAVE EXCEL FILE
    # ========================================================================
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"trx_balance_{timestamp}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)

    wb.save(filepath)

    print(f"\n{'='*80}")
    print(f"✅ Excel file saved: {filename}")
    print(f"{'='*80}\n")

    # ========================================================================
    # STEP 7: OVERWRITE TOKEN LIST FILE IN "Token List" FOLDER
    # ========================================================================
    if token_info_responses:
        print(f"7️⃣  Overwriting Token List file in 'Token List' folder...")

        # Create "Token List" folder if it doesn't exist
        token_list_folder = os.path.join(os.getcwd(), "Token List")
        os.makedirs(token_list_folder, exist_ok=True)

        # Create a new workbook with just the Token List (this will overwrite existing)
        wb_token_list = Workbook()
        ws_token_list_export = wb_token_list.active
        ws_token_list_export.title = "Token List"

        # Add headers
        headers = ["abbr", "decimal", "contractAddress", "symbolShow"]
        ws_token_list_export.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_token_list_export[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Add all tokens from contract_to_token, contract_to_decimal, and contract_to_symbolshow mappings
        all_contract_addresses = set(list(contract_to_token.keys()) + list(contract_to_decimal.keys()) + list(contract_to_symbolshow.keys()))

        for contract_addr in all_contract_addresses:
            symbol = contract_to_token.get(contract_addr, '')
            decimals = contract_to_decimal.get(contract_addr, '')
            symbol_show = contract_to_symbolshow.get(contract_addr, '')
            ws_token_list_export.append([symbol, decimals, contract_addr, symbol_show])

        # Overwrite the existing "Token List.xlsx" file (not timestamped)
        token_list_filename = "Token List.xlsx"
        token_list_filepath = os.path.join(token_list_folder, token_list_filename)
        wb_token_list.save(token_list_filepath)

        print(f"   ✅ Token List overwritten: Token List/{token_list_filename}")
        print(f"   ✅ Total tokens in Token List: {len(all_contract_addresses)}")
        print(f"   ✅ New tokens added from Contract Address's Token API: {len(token_info_responses)}")
        print(f"{'='*80}\n")

    assert os.path.exists(filepath), f"Excel file was not created: {filepath}"


def flatten_dict(d, parent_key='', sep='_'):
    """
    Flatten nested dictionary structure
    Example: {"a": {"b": 1}} -> {"a_b": 1}
    """
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k

        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Convert list to JSON string
            items.append((new_key, json.dumps(v)))
        else:
            items.append((new_key, v))

    return dict(items)


if __name__ == "__main__":
    test_trx_balance()
