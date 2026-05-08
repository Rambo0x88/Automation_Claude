"""
DAM Full Overview Extraction Test

Extracts ALL data from the DAM portfolio Overview page into a SINGLE Excel tab:
  - Token Holdings (main wallet table)
  - Address Breakdown
  - All DeFi protocol sections (Aave V3, Morpho, Compound, Hyperliquid, etc.)

Skipped: Allocation pie charts, Chain Breakdown, Chain Activity (UI removed).

Run:
  pytest tests/test_dam_data_extraction.py -v -s --headed
  pytest tests/test_dam_data_extraction.py -v -s --headed --portfolio "my_portfolio"

Test ID: TC_DAM_DATA_EXTRACTION_001
"""
import json
import os
import re
from datetime import datetime
from typing import Optional

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page

from config.config import Config
from pages.sign_in_page import SignInPage

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_TC1_PATH = os.path.join(_ROOT, "test_data", "tc1_account.json")
_DUNE_PATH = os.path.join(_ROOT, "test_data", "tc_dune_wallet.json")
_OUTPUT_DIR = os.path.join(_ROOT, "test-results", "excel-exports")
_SCREENSHOT_DIR = os.path.join(_ROOT, "test-results", "screenshots")

# ---------------------------------------------------------------------------
# Skip keywords (allocation pie charts only)
# ---------------------------------------------------------------------------
_SKIP_KEYWORDS = {
    'allocation', 'token allocation', 'chain allocation', 'platform allocation',
}

# DeFi protocol keywords
_DEFI_KEYWORDS = [
    'aave', 'morpho', 'compound', 'uniswap', 'curve', 'lido',
    'maker', 'spark', 'euler', 'pendle', 'yearn', 'balancer', 'convex',
    'virtuals', 'hyperliquid', 'merkl',
]

_CHAIN_NAMES = [
    'Ethereum', 'Base', 'Arbitrum', 'Polygon', 'Optimism', 'BSC',
    'Avalanche', 'Fantom', 'zkSync', 'Scroll',
]

# ---------------------------------------------------------------------------
# Excel column schema (single flat tab)
# ---------------------------------------------------------------------------
COLUMNS = [
    "Section",
    "Chain",
    "Name / Pool",
    "Price",
    "Price (24h)",
    "Share",
    "Amount",
    "Amount Tooltip",
    "Value",
    "Position Type",
]

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
_BOLD = Font(bold=True)


def _style_header_row(ws):
    for cell in ws[1]:
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")


def _auto_col_widths(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 50)


# ---------------------------------------------------------------------------
# Credentials / portfolio name
# ---------------------------------------------------------------------------

def _load_credentials():
    if os.path.exists(_TC1_PATH):
        with open(_TC1_PATH) as f:
            acc = json.load(f)
        return acc["email"], acc["password"]
    return Config.TEST_EMAIL, Config.TEST_PASSWORD


def _load_portfolio_name(pytestconfig) -> "Optional[str]":
    try:
        arg = pytestconfig.getoption("portfolio")
        if arg:
            return arg.strip()
    except Exception:
        pass
    if os.path.exists(_DUNE_PATH):
        with open(_DUNE_PATH) as f:
            data = json.load(f)
        name = data.get("portfolio_name", "").strip()
        if name:
            return name
    return None


def _load_portfolio_id(pytestconfig) -> "Optional[str]":
    try:
        arg = pytestconfig.getoption("portfolioId")
        if arg:
            return arg.strip()
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Extraction helpers
# ---------------------------------------------------------------------------

def _get_tooltip(page: Page, cell) -> str:
    try:
        trigger = cell.locator("[data-tooltip-id]").first
        if trigger.count() == 0 or not trigger.is_visible(timeout=300):
            return ""
        tooltip_id = trigger.get_attribute("data-tooltip-id")
        if not tooltip_id:
            return ""
        trigger.hover(timeout=2000)
        page.wait_for_timeout(300)
        tip = page.locator(f'[id="{tooltip_id}"]').first
        if tip.count() > 0 and tip.is_visible(timeout=300):
            return tip.inner_text().strip()
    except Exception:
        pass
    return ""


def _detect_section_name(section) -> "Optional[str]":
    selectors = [
        ".typography-title.font-semibold.text-mono-900",
        ".typography-tab.font-semibold.text-mono-900",
        '[class*="font-semibold"][class*="text-mono-900"]',
        '[class*="font-semibold"]',
    ]
    for sel in selectors:
        elems = section.locator(sel).all()
        for elem in elems[:3]:
            try:
                if not elem.is_visible(timeout=200):
                    continue
                text = elem.text_content().strip()
                if 2 < len(text) < 60 and not text.startswith("$"):
                    return text
            except Exception:
                continue
    return None


def _detect_chain(section) -> "Optional[str]":
    for chain in _CHAIN_NAMES:
        img = section.locator(f'img[alt="{chain}"]').first
        if img.count() > 0:
            try:
                if img.is_visible(timeout=200):
                    return chain
            except Exception:
                pass
    return None


def _extract_token_holdings(page: Page) -> list[dict]:
    """
    Extract the main Overview wallet/token table using DAMDataExtractor,
    which handles virtual-scroll / lazy-loading automatically.
    Returns rows mapped to the flat schema.
    dict keys from DAMDataExtractor: Row_Number, Chain, Name, Price, Price_24h,
    Share, Amount, Amount_Tooltip, Value
    """
    from utils.dam_data_extractor import DAMDataExtractor
    rows = []
    try:
        extractor = DAMDataExtractor(page)
        raw = extractor.extract_overview_data()
        for r in raw:
            rows.append({
                "Section": "Token Holdings",
                "Chain": r.get("Chain", ""),
                "Name / Pool": r.get("Name", ""),
                "Price": r.get("Price", ""),
                "Price (24h)": r.get("Price_24h", ""),
                "Share": r.get("Share", ""),
                "Amount": r.get("Amount", ""),
                "Amount Tooltip": r.get("Amount_Tooltip", ""),
                "Value": r.get("Value", ""),
                "Position Type": "",
            })
    except Exception as e:
        print(f"   ⚠️  Token holdings error: {e}")
    return rows


def _extract_address_breakdown(page: Page) -> list[dict]:
    """Extract Combined Net Worth / Address Breakdown section."""
    rows = []
    _skip_words = {"addresses", "value", "combined net worth"}
    try:
        sections = page.locator("div.bg-grey-30.rounded-sm").all()
        for section in sections:
            try:
                text = section.text_content()
                if "combined net worth" not in text.lower() and "address" not in text.lower():
                    continue

                table = section.locator("table").first
                if table.count() > 0 and table.is_visible(timeout=1000):
                    for tr in table.locator("tbody tr").all():
                        cells = tr.locator("td").all()
                        texts = [c.text_content().strip() for c in cells]
                        if len(texts) >= 2 and any(texts):
                            rows.append({
                                "Section": "Address Breakdown",
                                "Chain": "",
                                "Name / Pool": texts[0],
                                "Price": "",
                                "Price (24h)": "",
                                "Share": "",
                                "Amount": "",
                                "Amount Tooltip": "",
                                "Value": texts[1] if len(texts) > 1 else "",
                                "Position Type": "",
                            })
                else:
                    all_text = section.inner_text().strip()
                    lines = [l.strip() for l in all_text.splitlines()
                             if l.strip() and l.strip().lower() not in _skip_words]
                    i = 0
                    while i < len(lines) - 1:
                        rows.append({
                            "Section": "Address Breakdown",
                            "Chain": "",
                            "Name / Pool": lines[i],
                            "Price": "",
                            "Price (24h)": "",
                            "Share": "",
                            "Amount": "",
                            "Amount Tooltip": "",
                            "Value": lines[i + 1],
                            "Position Type": "",
                        })
                        i += 2
                if rows:
                    break
            except Exception:
                continue
    except Exception as e:
        print(f"   ⚠️  Address breakdown error: {e}")
    return rows


def _extract_defi_section(page: Page, section, section_name: str) -> list[dict]:
    """Extract a standard DeFi protocol section (Position Type | Pool | Amount | Value)."""
    rows = []
    table = section.locator("table").first
    if table.count() == 0:
        return rows
    try:
        table.wait_for(state="visible", timeout=1000)
    except Exception:
        return rows

    # Expand collapsible rows
    try:
        for btn in table.locator("tbody tr td:first-child button").all():
            if btn.is_visible(timeout=300):
                btn.click(timeout=1000)
                page.wait_for_timeout(500)
    except Exception:
        pass

    current_pos_type = ""
    is_hyperliquid = "hyperliquid" in section_name.lower()

    for tr in table.locator("tr").all():
        if tr.locator("th").count() > 0:
            continue
        cells = tr.locator("td").all()
        if not cells:
            continue
        texts = [c.text_content().strip() for c in cells]
        if not any(texts):
            continue

        if len(texts) >= 4:
            pos, pool, amount_raw, value = texts[0], texts[1], texts[2], texts[3]
            amount_cell = cells[2]
        elif len(texts) == 3:
            pos, pool, amount_raw, value = "", texts[0], texts[1], texts[2]
            amount_cell = cells[1]
        elif len(texts) == 2:
            pos, pool, amount_raw, value = "", texts[0], texts[1], ""
            amount_cell = cells[1]
        else:
            continue

        if pos and pos not in ("", "Position Type"):
            current_pos_type = pos

        if pool in ("", "Pool", "Amount", "Value", "Position Type"):
            continue

        # For Hyperliquid Perpetuals, use generic col extraction
        if is_hyperliquid and "perpetual" in current_pos_type.lower():
            rows.append({
                "Section": section_name,
                "Chain": "",
                "Name / Pool": " | ".join(t for t in texts if t),
                "Price": "",
                "Price (24h)": "",
                "Share": "",
                "Amount": amount_raw,
                "Amount Tooltip": "",
                "Value": value.replace("$", "").replace(",", "").strip(),
                "Position Type": current_pos_type,
            })
        else:
            tooltip = _get_tooltip(page, amount_cell)
            rows.append({
                "Section": section_name,
                "Chain": "",
                "Name / Pool": pool,
                "Price": "",
                "Price (24h)": "",
                "Share": "",
                "Amount": amount_raw,
                "Amount Tooltip": tooltip,
                "Value": value.replace("$", "").replace(",", "").strip(),
                "Position Type": current_pos_type,
            })

    return rows


def _scroll_to_load_all(page: Page):
    """Scroll the page to trigger lazy-loading of all DeFi sections."""
    last_height = 0
    for _ in range(15):
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(800)
        height = page.evaluate("document.body.scrollHeight")
        if height == last_height:
            break
        last_height = height
    page.evaluate("window.scrollTo(0, 0)")
    page.wait_for_timeout(1000)


def _get_individual_wallet_addresses(page: Page) -> list[str]:
    """
    Get all full EVM/Tron wallet addresses visible in the Combined Net Worth section.
    Two methods with fallback:
      1. JS evaluate – searches the ENTIRE page for data-tooltip-id^="address-display-tooltip-"
      2. Raw HTML regex – scans page.content() for address patterns in tooltip-id attributes
    Only returns actual wallet addresses (0x... 42-char or T... 34-char).
    """
    addresses: list[str] = []
    seen: set[str] = set()

    _evm_re  = re.compile(r'^0x[A-Fa-f0-9]{40}$')
    _trx_re  = re.compile(r'^[Tt][A-Za-z0-9]{33}$')

    def _add(addr: str) -> None:
        a = addr.strip()
        if not a or a in seen:
            return
        # Only accept valid EVM (0x + 40 hex) or TRX (T + 33 base58) addresses
        if not (_evm_re.match(a) or _trx_re.match(a)):
            return
        seen.add(a)
        addresses.append(a)

    # ── Method 1: JS evaluate across entire page (no section-scoping) ──────
    try:
        found = page.evaluate('''() => {
            const results = [];
            const els = document.querySelectorAll(
                '[data-tooltip-id^="address-display-tooltip-"]'
            );
            for (const el of els) {
                let addr = "";
                // data-highlight-target on a descendant holds the full address
                const hlEl = el.querySelector("[data-highlight-target]");
                if (hlEl) {
                    addr = hlEl.getAttribute("data-highlight-target") || "";
                }
                // Fallback: strip the prefix from the tooltip-id itself
                if (!addr) {
                    const tid = el.getAttribute("data-tooltip-id") || "";
                    addr = tid.replace("address-display-tooltip-", "");
                }
                if (addr.trim()) results.push(addr.trim());
            }
            return results;
        }''')
        for addr in found:
            _add(addr)
        print(f"      Method 1 (JS query): {len(addresses)} address(es) found")
    except Exception as e:
        print(f"      Method 1 (JS query) failed: {e}")

    # ── Method 2: raw HTML regex fallback ─────────────────────────────────
    if not addresses:
        try:
            content = page.content()
            evm  = re.findall(r'address-display-tooltip-(0x[A-Fa-f0-9]{40})', content)
            tron = re.findall(r'address-display-tooltip-([Tt][A-Za-z0-9]{33})', content)
            for addr in evm + tron:
                _add(addr)
            print(f"      Method 2 (HTML regex): {len(addresses)} address(es) found")
        except Exception as e:
            print(f"      Method 2 (HTML regex) failed: {e}")

    if not addresses:
        print("      ⚠️  No wallet addresses found in Combined Net Worth section")

    return addresses


def _make_individual_tab_name(address: str) -> str:
    """
    Build an Excel-safe tab name: 'Individual - <short_address>'.
    Excel tab names are limited to 31 characters.
    'Individual - ' = 13 chars → 18 chars left for the address portion.
    """
    prefix = "Individual - "
    max_addr_len = 31 - len(prefix)  # 18
    if len(address) <= max_addr_len:
        short = address
    else:
        # e.g. "0x4e14fc...c0eab" or "TUqEg3dz...hmao8D"
        head = 8
        tail = max_addr_len - head - 3  # 3 for "..."
        short = f"{address[:head]}...{address[-tail:]}"
    return prefix + short


def _extract_individual_wallet_header(page: Page) -> list[dict]:
    """
    Extract the protocol summary tabs shown at the top of an individual wallet page
    (e.g. WALLET $20.32 | Morpho $1,887,525.15 | Merkl $589.92 …).
    These are rendered as scrollable filter buttons above the wallet table.
    """
    rows: list[dict] = []
    try:
        entries = page.evaluate(r'''() => {
            // The protocol tabs live in a scrollable flex row above the wallet table.
            // They have the pattern:  <button ...> ProtocolName … $Value </button>
            // Try both: wallet-section ancestor siblings and generic scrollable rows.
            const dollarRe = /\$([\d,]+(?:\.\d+)?)/;
            const results = [];
            const seen = new Set();

            // Selector: any button/div that (a) is visible, (b) contains a $ amount,
            // and (c) looks like a compact filter chip (short text).
            const candidates = document.querySelectorAll(
                'button, div[role="tab"], div[class*="cursor-pointer"]'
            );
            for (const el of candidates) {
                const raw = el.textContent.trim();
                if (!raw || raw.length > 60) continue;
                const m = raw.match(dollarRe);
                if (!m) continue;
                const value = m[1];
                const name = raw.replace(dollarRe, '').replace(/\s+/g, ' ').trim();
                if (!name || seen.has(name + value)) continue;
                seen.add(name + value);
                results.push({ name, value });
            }
            return results;
        }''')
        for e in entries:
            name = e.get("name", "").strip()
            value = e.get("value", "").strip()
            if name:
                rows.append({
                    "Section": "Token Holdings Header",
                    "Chain": "",
                    "Name / Pool": name,
                    "Price": "",
                    "Price (24h)": "",
                    "Share": "",
                    "Amount": "",
                    "Amount Tooltip": "",
                    "Value": value,
                    "Position Type": "",
                })
    except Exception as e:
        print(f"         ⚠️  Header extraction error: {e}")
    return rows


def _extract_individual_wallet_token_table(page: Page) -> list[dict]:
    """
    Extract token rows from the individual wallet page, scoped to #wallet-section.

    WHY NOT _extract_token_holdings / DAMDataExtractor:
      DAMDataExtractor.extract_overview_data() uses page-wide
      `page.locator("tbody tr").all()`, which on the individual page collects
      rows from ALL tables (wallet + Morpho + Merkl + Compound + Aave), producing
      mixed garbage.  Scoping to #wallet-section fixes this.

    Column order on the individual wallet page (confirmed from DevTools screenshot):
      0:Chain  1:Name  2:Price(24h)  3:Amount  4:Share  5:Value
    This differs from the overview page (where Share is col-3, Amount col-4).
    """
    rows: list[dict] = []
    try:
        ws = page.locator("#wallet-section").first
        if ws.count() == 0:
            print("         ⚠️  #wallet-section not found")
            return rows

        table = ws.locator("table").first
        if table.count() == 0:
            print("         ⚠️  No table inside #wallet-section")
            return rows

        for tr in table.locator("tbody tr").all():
            try:
                cells = tr.locator("td").all()
                if not cells:
                    continue
                texts = []
                for cell in cells:
                    try:
                        raw = cell.inner_text().strip()
                        lines = [l.strip() for l in raw.split("\n") if l.strip()]
                        texts.append(lines[0] if lines else "")
                    except Exception:
                        texts.append("")

                if len(texts) < 2:
                    continue
                # Skip header-like rows
                if texts[0].lower() in ("chain", "") and texts[1].lower() in ("name", ""):
                    continue
                if not texts[1]:   # no token name → not a data row
                    continue

                rows.append({
                    "Section": "Token Holdings",
                    "Chain":       texts[0] if len(texts) > 0 else "",
                    "Name / Pool": texts[1] if len(texts) > 1 else "",
                    "Price":       texts[2] if len(texts) > 2 else "",
                    "Price (24h)": "",   # price-change % is embedded in col-2 text
                    "Share":       texts[4] if len(texts) > 4 else "",
                    "Amount":      texts[3] if len(texts) > 3 else "",
                    "Amount Tooltip": "",
                    "Value":       texts[5] if len(texts) > 5 else "",
                    "Position Type": "",
                })
            except Exception:
                continue
    except Exception as e:
        print(f"         ⚠️  Wallet table error: {e}")
    return rows


def _extract_individual_wallet_page(page: Page) -> list[dict]:
    """
    Extract all data from the current individual wallet detail page.
    Sections exported (same flat COLUMNS schema as overview):
      • Token Holdings Header  – protocol summary filter chips at top
      • Token Holdings         – wallet token table (scoped to #wallet-section)
      • <Protocol Name>        – each DeFi / protocol section below
    """
    rows: list[dict] = []

    _scroll_to_load_all(page)

    # 1. Protocol summary header chips (WALLET $x | Morpho $x | Merkl $x …)
    header_rows = _extract_individual_wallet_header(page)
    rows.extend(header_rows)
    print(f"         Token Holdings Header: {len(header_rows)} rows")

    # 2. Wallet token table — MUST be scoped to #wallet-section only.
    #    Using page-wide DAMDataExtractor here would also pick up all DeFi rows.
    token_rows = _extract_individual_wallet_token_table(page)
    rows.extend(token_rows)
    print(f"         Token Holdings: {len(token_rows)} rows")

    # 3. Protocol / DeFi sections (Morpho, Merkl, Compound V3, Aave V3 …)
    #    Class "bg-grey-30 rounded-sm" matches the protocol cards regardless
    #    of Tailwind class declaration order.
    seen_names: set[str] = set()
    candidates = page.locator("div.bg-grey-30.rounded-sm").all()
    print(f"         Found {len(candidates)} protocol section candidate(s)")

    for section in candidates:
        try:
            name = _detect_section_name(section)
            if not name:
                continue
            name_lower = name.lower()
            if any(skip in name_lower for skip in _SKIP_KEYWORDS):
                continue
            if name_lower in ("wallet", "token holdings", "holdings"):
                continue
            chain = _detect_chain(section)
            display_name = f"{name} ({chain})" if chain else name
            if display_name in seen_names:
                continue
            seen_names.add(display_name)
            defi_rows = _extract_defi_section(page, section, display_name)
            if defi_rows:
                rows.extend(defi_rows)
                print(f"         {display_name}: {len(defi_rows)} rows")
            else:
                print(f"         {display_name}: 0 rows (no table or empty)")
        except Exception as ex:
            print(f"         Section error: {ex}")
            continue

    return rows


# ---------------------------------------------------------------------------
# Test
# ---------------------------------------------------------------------------

@pytest.mark.dam_extraction
class TestDAMDataExtraction:
    """Extract all DAM Overview data into a single Excel tab."""

    def test_extract_dam_overview_holdings_data(self, page: Page, config, pytestconfig):
        """
        TC_DAM_DATA_EXTRACTION_001
        Extract token holdings + address breakdown + all DeFi sections
        into a single Excel tab. Allocation sections are skipped.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        print(f"\n{'='*80}")
        print("TC_DAM_DATA_EXTRACTION_001: DAM Full Overview Extraction")
        print(f"{'='*80}")

        # ----------------------------------------------------------------
        # STEP 1: Credentials + portfolio name
        # ----------------------------------------------------------------
        email, password = _load_credentials()
        portfolio_name = _load_portfolio_name(pytestconfig)
        portfolio_id   = _load_portfolio_id(pytestconfig)
        print(f"   Account  : {email}")
        print(f"   Portfolio: {portfolio_name or '(all portfolios)'}")
        if portfolio_id:
            print(f"   PortfolioId: {portfolio_id}")

        # ----------------------------------------------------------------
        # STEP 2: Sign in
        # ----------------------------------------------------------------
        sign_in = SignInPage(page)
        sign_in.navigate()
        sign_in.sign_in(email, password)
        assert sign_in.wait_for_successful_sign_in(), "Sign-in failed"
        print("   ✅ Signed in")

        # ----------------------------------------------------------------
        # STEP 3: Navigate to portfolio
        # ----------------------------------------------------------------
        from config.config import Config as _Config

        if portfolio_id:
            # Direct navigation by portfolioId — most reliable
            target_url = f"{_Config.BASE_URL}/portfolio?portfolioId={portfolio_id}"
            page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_timeout(5000)
            print(f"   ✅ Navigated directly to portfolioId: {portfolio_id}")
        else:
            page.goto(_Config.PORTFOLIO_URL)
            page.wait_for_load_state("networkidle")
            page.wait_for_timeout(3000)

        if not portfolio_id and portfolio_name:
            navigated_portfolio = False

            # Open portfolio dropdown (button at top of page showing current portfolio name)
            dropdown_opened = False
            try:
                btn = page.locator('button:has-text("Portfolio")').first
                if btn.count() > 0:
                    btn.click(timeout=5000)
                    page.wait_for_timeout(2000)
                    dropdown_opened = True
                    print(f"   🔽 Portfolio dropdown opened")
            except Exception:
                pass

            if not dropdown_opened:
                # Fallback: coordinate click on the portfolio switcher chevron area
                try:
                    page.mouse.click(395, 141)
                    page.wait_for_timeout(2000)
                    dropdown_opened = True
                    print(f"   🔽 Portfolio dropdown opened (coordinate click)")
                except Exception as e:
                    print(f"   ⚠️  Could not open dropdown: {e}")

            # Scroll through dropdown in chunks, checking for the portfolio at each step
            matched_element = None
            if dropdown_opened:
                page.mouse.move(490, 450)
                pname_lower = portfolio_name.lower()
                for chunk in range(50):  # up to 50 scroll chunks
                    # Check currently visible menuitems
                    try:
                        for item in page.get_by_role("menuitem").all():
                            if not item.is_visible():
                                continue
                            txt = (item.text_content() or "").lower()
                            if pname_lower in txt and "create portfolio" not in txt:
                                matched_element = item
                                break
                    except Exception:
                        pass
                    if matched_element:
                        break
                    page.mouse.wheel(0, 300)
                    page.wait_for_timeout(80)

            if matched_element:
                matched_element.click()
                page.wait_for_load_state("networkidle")
                page.wait_for_timeout(5000)
                navigated_portfolio = True
                print(f"   ✅ Navigated to: {portfolio_name}")
            else:
                print(f"   ⚠️  Portfolio '{portfolio_name}' not found in dropdown — using current portfolio")
        elif not portfolio_id:
            print("   ℹ️  No portfolio — staying on current view")

        # ----------------------------------------------------------------
        # STEP 4: Ensure Overview tab is active
        # ----------------------------------------------------------------
        try:
            overview_tab = page.locator('text="Overview"').first
            if overview_tab.is_visible(timeout=3000):
                overview_tab.click()
                page.wait_for_timeout(2000)
        except Exception:
            pass

        # Store current URL so we can return after individual wallet navigation
        portfolio_overview_url = page.url

        # ----------------------------------------------------------------
        # STEP 5: Scroll to load all lazy sections
        # ----------------------------------------------------------------
        print("\n   🔄 Scrolling to load all DeFi sections...")
        _scroll_to_load_all(page)
        print("   ✅ Fully loaded")

        # ----------------------------------------------------------------
        # STEP 6: Full-page screenshot
        # ----------------------------------------------------------------
        os.makedirs(_SCREENSHOT_DIR, exist_ok=True)
        os.makedirs(_OUTPUT_DIR, exist_ok=True)
        safe_name = re.sub(r"[^A-Za-z0-9_\-]", "_", portfolio_name or "portfolio")
        screenshot_path = os.path.join(
            _SCREENSHOT_DIR, f"DAM_Overview_{safe_name}_{timestamp}.png"
        )
        try:
            page.screenshot(path=screenshot_path, full_page=True, timeout=120000)
            print(f"   📸 Screenshot: {screenshot_path}")
        except Exception as e:
            print(f"   ⚠️  Could not take screenshot: {e}")

        # ----------------------------------------------------------------
        # STEP 7: Extract all data into flat rows
        # ----------------------------------------------------------------
        all_rows: list[dict] = []

        # 7a. Token Holdings (main table — scroll-aware via DAMDataExtractor)
        print("\n   📋 Extracting Token Holdings...")
        token_rows = _extract_token_holdings(page)
        all_rows.extend(token_rows)
        print(f"      ✅ {len(token_rows)} rows")

        # 7b. Address Breakdown
        print("\n   📋 Extracting Address Breakdown...")
        addr_rows = _extract_address_breakdown(page)
        all_rows.extend(addr_rows)
        print(f"      ✅ {len(addr_rows)} rows")

        # 7c. DeFi sections
        print("\n   🔍 Scanning DeFi sections...")
        sections = page.locator("div.bg-grey-30.rounded-sm").all()
        print(f"   Found {len(sections)} candidate sections")

        for idx, section in enumerate(sections):
            try:
                name = _detect_section_name(section)
                if not name:
                    continue

                name_lower = name.lower()
                if any(skip in name_lower for skip in _SKIP_KEYWORDS):
                    print(f"   ⏭️  Skip: {name}")
                    continue

                # Skip wallet/token-holdings section (already extracted above)
                if name_lower in ("wallet", "token holdings", "holdings"):
                    continue

                chain = _detect_chain(section)
                display_name = f"{name} ({chain})" if chain else name

                defi_rows = _extract_defi_section(page, section, display_name)
                if defi_rows:
                    all_rows.extend(defi_rows)
                    print(f"   📋 {display_name}: {len(defi_rows)} rows")
                else:
                    print(f"   ⚠️  {display_name}: no rows (no table or empty)")

            except Exception as e:
                print(f"   ⚠️  Error in section {idx}: {e}")

        print(f"\n   📊 Total rows extracted: {len(all_rows)}")

        # ----------------------------------------------------------------
        # STEP 7d: Individual wallet detail pages
        # Reload the overview to get a fresh DOM before finding addresses,
        # then navigate to each wallet URL and extract its data.
        # ----------------------------------------------------------------
        print("\n   👛 Extracting individual wallet pages...")
        individual_wallet_tabs: dict[str, list[dict]] = {}

        # Reload overview so DOM is clean (STEP 7c may have mutated it)
        print(f"      Reloading overview: {portfolio_overview_url}")
        page.goto(portfolio_overview_url)
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(3000)
        _scroll_to_load_all(page)

        wallet_addresses = _get_individual_wallet_addresses(page)
        print(f"      Found {len(wallet_addresses)} wallet address(es): {wallet_addresses}")

        for address in wallet_addresses:
            last8 = address[-8:]
            wallet_tab = f"{last8} Wallet"
            defi_tab   = f"{last8} De-Fi"
            print(f"\n      ── {last8} ({address}) ──")

            navigated = False
            try:
                # Scroll the page to bring the Combined Net Worth section into view
                # then click the address row to navigate to the individual wallet page
                # Find element case-insensitively (EVM addresses in DOM may be checksum-cased)
                addr_lower = address.lower()
                click_result = page.evaluate(f'''() => {{
                    const addrLower = "{addr_lower}";
                    const all = document.querySelectorAll('[data-tooltip-id^="address-display-tooltip-"]');
                    let el = null;
                    for (const candidate of all) {{
                        const tid = (candidate.getAttribute("data-tooltip-id") || "").toLowerCase();
                        if (tid === "address-display-tooltip-" + addrLower) {{
                            el = candidate;
                            break;
                        }}
                    }}
                    if (!el) return {{ found: false }};
                    el.scrollIntoView({{ behavior: "instant", block: "center" }});
                    // Walk up to find cursor-pointer ancestor
                    let target = el;
                    for (let i = 0; i < 15; i++) {{
                        target = target.parentElement;
                        if (!target) break;
                        const cls = target.className || "";
                        if (cls.includes("cursor-pointer")) {{
                            target.click();
                            return {{ found: true, clicked: "cursor-pointer-ancestor" }};
                        }}
                    }}
                    el.click();
                    return {{ found: true, clicked: "element-itself" }};
                }}''')
                print(f"      click_result: {click_result}")

                if click_result.get("found"):
                    page.wait_for_load_state("networkidle")
                    page.wait_for_timeout(3000)
                    current_url = page.url
                    print(f"      Landed → {current_url}")
                    navigated = addr_lower in current_url.lower()
                    if not navigated:
                        print(f"      ⚠️  URL does not contain address param")
                else:
                    print(f"      ⚠️  Element not found for {address}")
            except Exception as e:
                print(f"      ⚠️  Click failed: {e}")

            try:
                all_ind_rows = _extract_individual_wallet_page(page)
                print(f"      Extracted {len(all_ind_rows)} rows total")
                if all_ind_rows:
                    # Split into Wallet rows and De-Fi rows
                    ind_token_rows = [r for r in all_ind_rows if r["Section"] == "Token Holdings"]
                    ind_defi_rows  = [r for r in all_ind_rows if r["Section"] not in ("Token Holdings", "Token Holdings Header")]
                    if ind_token_rows:
                        individual_wallet_tabs[wallet_tab] = ind_token_rows
                        print(f"      ✅ '{wallet_tab}': {len(ind_token_rows)} rows")
                    if ind_defi_rows:
                        individual_wallet_tabs[defi_tab] = ind_defi_rows
                        print(f"      ✅ '{defi_tab}': {len(ind_defi_rows)} rows")
                else:
                    print(f"      ⚠️  No rows returned — check extraction functions")

            except Exception as e:
                print(f"      ❌ Error extracting: {e}")
                import traceback; traceback.print_exc()

            # Return to portfolio overview for next iteration
            try:
                page.goto(portfolio_overview_url)
                page.wait_for_load_state("networkidle")
                page.wait_for_timeout(2000)
                _scroll_to_load_all(page)
            except Exception:
                pass

        print(f"\n      ✅ {len(individual_wallet_tabs)} individual wallet tab(s) ready")

        # ----------------------------------------------------------------
        # STEP 8: Export to single-tab Excel
        # ----------------------------------------------------------------
        os.makedirs(_OUTPUT_DIR, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "DAM Overview"

        ws.append(COLUMNS)
        _style_header_row(ws)

        prev_section = None
        for row in all_rows:
            # Add a blank separator row between sections
            if prev_section and row["Section"] != prev_section:
                ws.append([""] * len(COLUMNS))
            ws.append([row.get(col, "") for col in COLUMNS])
            prev_section = row["Section"]

        _auto_col_widths(ws)

        # Add one tab per individual wallet
        for tab_name, wallet_rows in individual_wallet_tabs.items():
            ws_ind = wb.create_sheet(tab_name)
            ws_ind.append(COLUMNS)
            _style_header_row(ws_ind)
            prev_sec = None
            for row in wallet_rows:
                if prev_sec and row["Section"] != prev_sec:
                    ws_ind.append([""] * len(COLUMNS))
                ws_ind.append([row.get(col, "") for col in COLUMNS])
                prev_sec = row["Section"]
            _auto_col_widths(ws_ind)

        filename = f"DAM_Overview_{safe_name}_{timestamp}.xlsx"
        filepath = os.path.join(_OUTPUT_DIR, filename)
        wb.save(filepath)

        assert os.path.exists(filepath), f"Excel not created: {filepath}"

        print(f"\n{'='*80}")
        print("✅ TC_DAM_DATA_EXTRACTION_001 COMPLETE")
        print(f"   Screenshot : {screenshot_path}")
        print(f"   Excel      : {filepath}")
        print(f"   Total rows : {len(all_rows)}")
        # Section summary
        from collections import Counter
        for sec, count in Counter(r["Section"] for r in all_rows).items():
            print(f"     • {sec}: {count} rows")
        if individual_wallet_tabs:
            print(f"   Individual wallet tabs ({len(individual_wallet_tabs)}):")
            for tab, rows in individual_wallet_tabs.items():
                print(f"     • {tab}: {len(rows)} rows")
        print(f"{'='*80}")
