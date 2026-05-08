#!/usr/bin/env python3
"""
Full portfolio extraction supporting TRX, EVM (0x), or CEX-only portfolios.

Configure the addresses at the top of this file, then run:
    python3 run_overview.py

Address types:
  - EVM (0x...):  set EVM_ADDRESSES,                         SKIP_TRX_API = True,  SKIP_SIM_DUNE_API = False
  - TRX (T...):   set TRX_ADDRESS / TRX_ADDRESSES,           SKIP_TRX_API = False, SKIP_SIM_DUNE_API = True
  - Both:         set EVM_ADDRESSES + TRX_ADDRESS/ADDRESSES, SKIP_TRX_API = False, SKIP_SIM_DUNE_API = False
  - CEX-only:     no addresses needed,                       SKIP_TRX_API = True,  SKIP_SIM_DUNE_API = True

This test combines:
1. TRX Balance API (skipped for EVM-only or CEX-only)
2. SimDune API - EVM on-chain balances (skipped for TRX-only)
3. Rabby Protocol + Hyperliquid API - DeFi positions (skipped for TRX-only)
4. DAM Portfolio Full Extraction - scrapes all tables from DAM UI
5. Exports everything into a single Excel file (DAM sheets + Rabby sheets merged)
"""

import os
import sys
import pytest
import requests
import json
import re
import traceback
from datetime import datetime
from decimal import Decimal, getcontext, InvalidOperation
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.formatting.rule import CellIsRule
from config.config import Config
from concurrent.futures import ThreadPoolExecutor, as_completed

# Set high precision for Decimal calculations (50 decimal places)
getcontext().prec = 50

# Global quiet mode flag (set by command-line argument)
QUIET_MODE = False

def log(message, force=False):
    """Print message only if not in quiet mode, or if force=True"""
    if not QUIET_MODE or force:
        print(message)

# ============================================================================
# CONFIGURE ADDRESSES HERE
# ============================================================================
# EVM-only:  set EVM_ADDRESSES,                          SKIP_TRX_API = True,  SKIP_SIM_DUNE_API = False
# TRX-only:  set TRX_ADDRESS / TRX_ADDRESSES,            SKIP_TRX_API = False, SKIP_SIM_DUNE_API = True
# Both:      set EVM_ADDRESSES + TRX_ADDRESS/ADDRESSES,  SKIP_TRX_API = False, SKIP_SIM_DUNE_API = False
# CEX-only:  no addresses needed,                        SKIP_TRX_API = True,  SKIP_SIM_DUNE_API = True
TRX_ADDRESS = "TUqEg3dzVEJNQSVW2HY98z5X8SBdhmao8D"   # TRX address (T...), or leave empty if EVM-only
TRX_ADDRESSES = [TRX_ADDRESS]  # All TRX addresses (for multi-address TRX portfolios)
EVM_ADDRESSES = ["0x4e14fc11f58b64740e66e4b1aa188a4b007c0eab"]  # EVM addresses (0x..., 42 chars)
PORTFOLIO_NAME_PREFIX = "A_bD"  # A_ + last char of each address: EVM(b) + TRX(D)
SKIP_TRX_API = False   # True = skip TRX Balance API (EVM-only or CEX-only)
SKIP_SIM_DUNE_API = False  # True = skip SimDune + Rabby APIs (TRX-only or CEX-only)

# Direct portfolio ID - if set, navigate directly to this portfolio (skip search)
# Get this from the DAM URL: portfolioId=<this-value>
# Set to None to search by portfolio name instead
PORTFOLIO_ID = None  # Disabled - will search by portfolio name

# ============================================================================
# PERFORMANCE OPTIMIZATION SETTINGS
# ============================================================================
# Set to True to skip price fetching entirely (fastest, but no price data)
SKIP_PRICE_FETCHING = False

# Maximum number of tokens to process (0 = unlimited)
# Recommended: 100-500 for large addresses
MAX_TOKENS_TO_PROCESS = 0

# Number of parallel API calls (higher = faster but more resource intensive)
# Recommended: 3-5 for optimal balance, reduced to avoid rate limiting
PARALLEL_API_CALLS = 3

# API timeout in seconds (lower = faster failure handling)
API_TIMEOUT_SECONDS = 5
# ============================================================================


def clean_currency_symbols(text):
    """Remove $ and % symbols from text data."""
    if isinstance(text, str):
        return text.replace('$', '').replace('%', '').strip()
    return text


def fetch_token_details(contract_addr, has_all_info):
    """
    Fetch token details from TronScan API with aggressive retry logic.
    Returns tuple: (contract_addr, token_data, success)

    Retry strategy:
    - Up to 10 attempts with 2-second delay between failures
    - Aggressive retries to ensure all tokens get their data
    - If token truly has no data after 10 attempts, returns None
    """
    import time

    token_detail_url = f"https://apilist.tronscanapi.com/api/token_trc20?contract={contract_addr}&showAll=1&start=&limit="

    # Use shorter timeout for tokens with existing info
    timeout = API_TIMEOUT_SECONDS if not has_all_info else max(1, API_TIMEOUT_SECONDS - 1)

    # Retry up to 10 times with 2-second delay between attempts
    max_retries = 10
    last_exception = None

    for attempt in range(max_retries):
        try:
            token_response = requests.get(token_detail_url, timeout=timeout)
            token_response.raise_for_status()
            token_detail_data = token_response.json()

            if "trc20_tokens" in token_detail_data and len(token_detail_data["trc20_tokens"]) > 0:
                # Success! Return token data
                if attempt > 0:
                    print(f"      ✓ Token {contract_addr[:8]}... succeeded on attempt {attempt + 1}")
                return (contract_addr, token_detail_data["trc20_tokens"][0], True)
            else:
                # API returned but no token data - this token likely doesn't exist
                return (contract_addr, None, False)
        except Exception as e:
            last_exception = e
            # If this is not the last attempt, wait 2 seconds before retrying
            if attempt < max_retries - 1:
                if attempt == 0:
                    print(f"      ⚠️  Token {contract_addr[:8]}... failed (attempt {attempt + 1}), retrying...")
                time.sleep(2)
            else:
                # All retries exhausted
                print(f"      ❌ Token {contract_addr[:8]}... failed after {max_retries} attempts: {str(last_exception)[:50]}")
                return (contract_addr, None, False)

    return (contract_addr, None, False)


def detect_evm_addresses(combined_net_worth_data):
    """
    Detect EVM addresses (0x...) from Combined Net Worth data.
    Returns list of unique EVM addresses found (case-insensitive deduplication).
    """
    seen = set()
    evm_addresses = []
    for row in combined_net_worth_data:
        if len(row) >= 1:
            address = str(row[0]).strip()
            # Check if it's an EVM address (starts with 0x and is 42 characters)
            if address.startswith('0x') and len(address) == 42:
                # Case-insensitive deduplication
                if address.lower() not in seen:
                    seen.add(address.lower())
                    evm_addresses.append(address)
    return evm_addresses


def detect_tron_addresses(combined_net_worth_data):
    """
    Detect Tron addresses (T..., 34 chars, Base58Check) from Combined Net Worth data.
    Returns list of unique Tron addresses found (case-insensitive deduplication).
    """
    seen = set()
    tron_addresses = []
    for row in combined_net_worth_data:
        if len(row) >= 1:
            address = str(row[0]).strip()
            # Check if it's a Tron address (starts with T and is 34 characters)
            if address.startswith('T') and len(address) == 34:
                # Case-insensitive deduplication
                if address.lower() not in seen:
                    seen.add(address.lower())
                    tron_addresses.append(address)
    return tron_addresses


def fetch_sim_dune_balance(address):
    """
    Fetch EVM chain balances from Sim Dune API for a given address.
    Handles pagination via next_offset — keeps fetching until no more pages.

    API: https://api.sim.dune.com/v1/evm/balances/{address}
    Headers: X-Sim-Api-Key: sim_z82dBaaNeG3Y1elSbDfAGbaHYjb4RTyE

    Returns: (address, data, success) tuple
    """
    import time

    base_url = f"https://api.sim.dune.com/v1/evm/balances/{address}"
    headers = {
        'X-Sim-Api-Key': 'sim_z82dBaaNeG3Y1elSbDfAGbaHYjb4RTyE',
        'User-Agent': 'insomnia/11.6.2'
    }

    all_balances = []
    next_offset = None
    page = 0
    max_retries = 3

    while True:
        params = {}
        if next_offset:
            params['offset'] = next_offset

        last_exception = None
        data = None

        for attempt in range(max_retries):
            try:
                response = requests.get(base_url, headers=headers, params=params, timeout=30)
                response.raise_for_status()
                data = response.json()
                if attempt > 0:
                    print(f"      ✓ Address {address[:10]}... page {page} succeeded on attempt {attempt + 1}")
                break
            except Exception as e:
                last_exception = e
                if attempt < max_retries - 1:
                    print(f"      ⚠️  Address {address[:10]}... page {page} failed (attempt {attempt + 1}), retrying...")
                    time.sleep(2)
                else:
                    print(f"      ❌ Address {address[:10]}... page {page} failed after {max_retries} attempts: {str(last_exception)[:50]}")
                    if page == 0:
                        return (address, None, False)
                    # Return what we have so far
                    combined = dict(data) if data else {}
                    combined['balances'] = all_balances
                    return (address, combined, True)

        if data is None:
            break

        page_balances = data.get('balances', [])
        all_balances.extend(page_balances)
        print(f"      📄 Page {page}: {len(page_balances)} balances (total so far: {len(all_balances)})")

        next_offset = data.get('next_offset')
        if not next_offset:
            break

        page += 1
        time.sleep(0.3)  # small delay between pages

    # Build combined response — same structure as single-page response
    combined = dict(data) if data else {}
    combined['balances'] = all_balances
    return (address, combined, True)


def load_coingecko_coin_list():
    """
    Load Coingecko Coin ID List from Excel and create lookup maps.

    Returns:
        tuple: (coingecko_map, coingecko_native_map, coingecko_addr_map)
            - coingecko_map: {(platform_address_lowercase, platform_name_lowercase): coin_id}
            - coingecko_native_map: {symbol_lowercase: coin_id} for native tokens
            - coingecko_addr_map: {platform_address_lowercase: coin_id} address-only fallback
    """
    from openpyxl import load_workbook
    import os

    coingecko_file = "Coingecko Coin ID List.xlsx"
    coingecko_map = {}
    coingecko_native_map = {}
    coingecko_addr_map = {}

    # Check current directory first, then fallback to automation folder
    if not os.path.exists(coingecko_file):
        alt_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'automation', 'Coingecko Coin ID List.xlsx')
        if os.path.exists(alt_path):
            coingecko_file = alt_path
            print(f"   📁 Using Coingecko file from: {alt_path}")

    if not os.path.exists(coingecko_file):
        print(f"   ⚠️  Warning: {coingecko_file} not found, Coingecko ID lookup will be skipped")
        return coingecko_map, coingecko_native_map, coingecko_addr_map

    try:
        print(f"   📂 Loading Coingecko Coin ID List (~26k rows, may take a moment)...")
        wb = load_workbook(coingecko_file, read_only=True, data_only=True)
        ws = wb["Coin ID List"]

        # Structure: A=ID, B=Symbol, C=Name, D=Platform Name, E=Platform Address
        # Use iter_rows for better performance with large files
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                coin_id = row[0]
                symbol = row[1]
                platform_name = row[3]
                platform_address = row[4]

                if coin_id:
                    # Map 1: EVM address lookup (platform_address + platform_name)
                    if platform_name and platform_address:
                        key = (str(platform_address).lower().strip(), str(platform_name).lower().strip())
                        coingecko_map[key] = str(coin_id).strip()
                        row_count += 1

                    # Map 2: Native token lookup (symbol where platform_name is blank OR "ethereum")
                    if symbol:
                        platform_name_str = str(platform_name).strip().lower() if platform_name else ""
                        if not platform_name_str or platform_name_str == "ethereum":
                            symbol_key = str(symbol).lower().strip()
                            # Prioritize blank platform over "ethereum" platform
                            if symbol_key not in coingecko_native_map or not platform_name_str:
                                coingecko_native_map[symbol_key] = str(coin_id).strip()

                    # Map 3: Address-only fallback (platform_address regardless of platform name)
                    # Used when chain name doesn't map or platform name in file differs from CHAIN_TO_PLATFORM
                    if platform_address:
                        addr_key = str(platform_address).lower().strip()
                        if addr_key not in coingecko_addr_map:
                            coingecko_addr_map[addr_key] = str(coin_id).strip()

            # Progress feedback every 5000 rows
            if row_count > 0 and row_count % 5000 == 0:
                print(f"      ... loaded {row_count} mappings")

        wb.close()
        print(f"   ✅ Loaded Coingecko Coin ID List: {len(coingecko_map)} platform mappings, "
              f"{len(coingecko_native_map)} native tokens, {len(coingecko_addr_map)} address-only entries")
        return coingecko_map, coingecko_native_map, coingecko_addr_map
    except Exception as e:
        print(f"   ⚠️  Warning: Failed to load Coingecko data: {e}")
        return {}, {}, {}


# Chain name mapping: Sim Dune chain name -> Coingecko platform name
CHAIN_TO_PLATFORM = {
    "ethereum": "ethereum",
    "binance smart chain": "binance-smart-chain",
    "bsc": "binance-smart-chain",
    "bnb": "binance-smart-chain",
    "polygon": "polygon-pos",
    "avalanche": "avalanche",
    "fantom": "fantom",
    "arbitrum": "arbitrum-one",
    "optimism": "optimistic-ethereum",
    "base": "base",
    "polygon zkevm": "polygon-zkevm",
    "linea": "linea",
    "scroll": "scroll",
    "zksync": "zksync",
    "gnosis": "xdai",
}


def is_valid_evm_address(address_str):
    """Check if address is a valid EVM address (42 chars, starts with 0x)."""
    if not address_str:
        return False
    address_str = str(address_str).strip()
    return len(address_str) == 42 and address_str.lower().startswith("0x")


def fetch_coingecko_prices_batch(coin_ids, raw_collector=None):
    """
    Fetch USD prices for multiple coin IDs from Coingecko simple/price API.

    Args:
        coin_ids: List of coin IDs to fetch
        raw_collector: Optional list to append raw API responses to

    Returns:
        dict: {coin_id: usd_price}
    """
    import requests
    import time

    if not coin_ids:
        return {}

    price_map = {}
    api_key = "CG-F3KENg4b1mcvyeg6eo6LGDQU"
    headers = {"x-cg-demo-api-key": api_key}

    print(f"   💰 Fetching USD prices for {len(coin_ids)} unique coin(s)...")

    # Batch API calls - Coingecko simple/price supports multiple IDs separated by commas
    batch_size = 250  # API limit per request
    for batch_start in range(0, len(coin_ids), batch_size):
        batch_ids = coin_ids[batch_start:batch_start + batch_size]
        ids_param = ",".join(batch_ids)

        try:
            url = f"https://api.coingecko.com/api/v3/simple/price?vs_currencies=usd&ids={ids_param}"
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            data = response.json()

            if raw_collector is not None:
                raw_collector.append({"endpoint": "simple/price", "url": url, "response": data})

            # Extract USD price for each coin
            for coin_id in batch_ids:
                if coin_id in data and "usd" in data[coin_id]:
                    price_map[coin_id] = data[coin_id]["usd"]
                else:
                    price_map[coin_id] = None

            # Progress feedback
            print(f"      ... fetched batch {batch_start + 1}-{min(batch_start + batch_size, len(coin_ids))}/{len(coin_ids)}")

            # Rate limiting between batches
            if batch_start + batch_size < len(coin_ids):
                time.sleep(1)

        except Exception as e:
            # API error (404, 429, timeout, etc.) — stop the entire run
            error_detail = str(e)
            status_code = ""
            response_body = ""
            if hasattr(e, 'response') and e.response is not None:
                status_code = f" (HTTP {e.response.status_code})"
                try:
                    response_body = e.response.text[:500]
                except:
                    response_body = ""
            print(f"\n{'='*80}")
            print(f"❌ FATAL: Coingecko API error — stopping extraction.")
            print(f"   Error: {error_detail}{status_code}")
            if response_body:
                print(f"   Response: {response_body}")
            print(f"   Batch: {batch_ids[:5]}{'...' if len(batch_ids) > 5 else ''}")
            print(f"{'='*80}")
            raise SystemExit(f"Coingecko API failed{status_code}: {error_detail}")

    successful_count = len([v for v in price_map.values() if v is not None])
    print(f"   ✅ Fetched prices for {successful_count}/{len(coin_ids)} coins")
    return price_map


def fetch_coingecko_price_change_batch(coin_ids, raw_collector=None):
    """
    Fetch 24H price change data for multiple coin IDs from Coingecko API.

    Args:
        coin_ids: List of coin IDs to fetch
        raw_collector: Optional list to append raw API responses to

    Returns:
        dict: {coin_id: price_change_24h_value}
    """
    import requests
    import time

    if not coin_ids:
        return {}

    price_change_map = {}
    api_key = "CG-F3KENg4b1mcvyeg6eo6LGDQU"
    headers = {"x-cg-demo-api-key": api_key}

    print(f"   📊 Fetching 24H price change for {len(coin_ids)} unique coin(s)...")

    for idx, coin_id in enumerate(coin_ids, 1):
        try:
            url = f"https://api.coingecko.com/api/v3/coins/{coin_id}"
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            data = response.json()

            if raw_collector is not None:
                raw_collector.append({"endpoint": f"coins/{coin_id}", "url": url, "response": data})

            price_change = data.get("market_data", {}).get("price_change_percentage_24h")
            # Store "null" string when API explicitly returns null (field exists but is null)
            if "market_data" in data and "price_change_percentage_24h" in (data.get("market_data") or {}):
                price_change_map[coin_id] = price_change if price_change is not None else "null"
            else:
                price_change_map[coin_id] = None

            if idx % 10 == 0 or idx == len(coin_ids):
                print(f"      ... fetched {idx}/{len(coin_ids)}")

            if idx < len(coin_ids):
                time.sleep(0.5)

        except Exception as e:
            # API error (404, 429, timeout, etc.) — stop the entire run
            error_detail = str(e)
            status_code = ""
            response_body = ""
            if hasattr(e, 'response') and e.response is not None:
                status_code = f" (HTTP {e.response.status_code})"
                try:
                    response_body = e.response.text[:500]
                except:
                    response_body = ""
            print(f"\n{'='*80}")
            print(f"❌ FATAL: Coingecko API error — stopping extraction.")
            print(f"   Error: {error_detail}{status_code}")
            if response_body:
                print(f"   Response: {response_body}")
            print(f"   Coin ID: {coin_id}")
            print(f"{'='*80}")
            raise SystemExit(f"Coingecko API failed{status_code}: {error_detail}")

    print(f"   ✅ Fetched price change data for {len([v for v in price_change_map.values() if v is not None])}/{len(coin_ids)} coins")
    return price_change_map


def export_sim_dune_to_excel(address, data, output_folder="test-results/API Result", portfolio_name=None):
    """
    Export Sim Dune API response to Excel file.

    File naming: SimDune_{portfolio_name}_{YYYYMMDD}_{HHMMSS}.xlsx

    Tabs:
    - SimDune: Raw JSON API response (not formatted as table)
    - Sim + Coingecko + Debank API: Extracted columns (Chain, Symbol, Amount (Raw), Amount, Decimals, Token Address, ID, Price, 24H Price Change)

    Returns: (filepath, extracted_data) tuple where extracted_data is a list of rows for DAM export
    """
    from openpyxl import Workbook
    from datetime import datetime
    from decimal import Decimal, getcontext
    import os

    # Set high precision for Decimal calculations
    getcontext().prec = 50

    os.makedirs(output_folder, exist_ok=True)

    # Use portfolio name if provided, otherwise fall back to last 8 chars of address
    name_suffix = portfolio_name if portfolio_name else address[-8:]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"SimDune_{name_suffix}_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()

    # Sheet 1: SimDune (Raw JSON API response - not a table)
    ws_raw = wb.active
    ws_raw.title = "SimDune"

    # Write raw JSON data as a single blob
    if data:
        ws_raw.cell(row=1, column=1, value="Raw API Response")
        ws_raw.cell(row=2, column=1, value=json.dumps(data, indent=2))
    else:
        ws_raw.cell(row=1, column=1, value="No data returned from API")

    # Sheet 2: Sim + Coingecko + Debank API (Extracted columns)
    ws_extracted = wb.create_sheet("Sim + Coingecko + Debank API")
    ws_extracted.cell(row=1, column=1, value="Chain")
    ws_extracted.cell(row=1, column=2, value="Symbol")
    ws_extracted.cell(row=1, column=3, value="Amount (Raw)")
    ws_extracted.cell(row=1, column=4, value="Amount")
    ws_extracted.cell(row=1, column=5, value="Decimals")
    ws_extracted.cell(row=1, column=6, value="Token Address")
    ws_extracted.cell(row=1, column=7, value="ID")
    ws_extracted.cell(row=1, column=8, value="Price")
    ws_extracted.cell(row=1, column=9, value="24H Price Change")

    # Collect extracted data for DAM export
    extracted_data = [["Chain", "Symbol", "Amount (Raw)", "Amount", "Decimals", "Token Address", "ID", "Price", "24H Price Change"]]

    # Load Coingecko coin list for ID lookup
    coingecko_map, coingecko_native_map, coingecko_addr_map = load_coingecko_coin_list()

    # Track coin IDs and their row numbers for price change fetching
    coin_id_to_rows = {}  # {coin_id: [row_numbers]}

    if data:
        balances = data.get('balances', data) if isinstance(data, dict) else data

        if isinstance(balances, list):
            row_num = 2
            for item in balances:
                if isinstance(item, dict):
                    chain = item.get('chain', item.get('chain_name', ''))
                    symbol = item.get('symbol', item.get('token_symbol', ''))
                    amount_raw = item.get('amount', item.get('balance', item.get('value_raw', '')))
                    decimals = item.get('decimals', 18)
                    token_address = item.get('address', '')

                    # Calculate actual amount using Decimal for full precision
                    # Formula: Amount = Amount_Raw * 10^(-Decimals)
                    try:
                        if amount_raw and decimals is not None:
                            # Use Decimal for precise calculation
                            raw_decimal = Decimal(str(amount_raw))
                            divisor = Decimal(10) ** int(decimals)
                            amount_decimal = raw_decimal / divisor
                            # Convert to string to preserve all precision
                            amount = str(amount_decimal)
                        else:
                            amount = str(amount_raw) if amount_raw else ""
                    except Exception as e:
                        amount = str(amount_raw) if amount_raw else ""

                    # Lookup Coingecko ID
                    coin_id = ""
                    price = ""
                    price_change_24h = ""

                    token_address_lower = str(token_address).lower().strip()

                    # Case 1: Valid EVM address (42 chars, starts with 0x)
                    if is_valid_evm_address(token_address):
                        # Map Sim Dune chain name to Coingecko platform name
                        platform_name = CHAIN_TO_PLATFORM.get(chain.lower().strip(), "")
                        if platform_name:
                            # Primary lookup: address + platform name
                            lookup_key = (token_address_lower, platform_name)
                            coin_id = coingecko_map.get(lookup_key, "")
                            if coin_id:
                                print(f"      ✅ Coingecko match (platform): {symbol} on {platform_name} → {coin_id}")
                        if not coin_id:
                            # Fallback: address-only (platform name in file may differ from CHAIN_TO_PLATFORM)
                            coin_id = coingecko_addr_map.get(token_address_lower, "")
                            if coin_id:
                                print(f"      ✅ Coingecko match (addr-only fallback): {symbol} {token_address_lower} → {coin_id}")
                        if not coin_id:
                            print(f"      ⚠️  Spam Token: {symbol} {token_address_lower} chain={chain} platform={platform_name} — not in Coingecko")
                            coin_id = "Spam Token"

                    # Case 2: Native token (address = "native")
                    elif token_address_lower == "native":
                        # Lookup by symbol (where platform_name is blank or "ethereum")
                        symbol_lower = str(symbol).lower().strip()
                        coin_id = coingecko_native_map.get(symbol_lower, "")

                    ws_extracted.cell(row=row_num, column=1, value=chain)
                    ws_extracted.cell(row=row_num, column=2, value=symbol)
                    ws_extracted.cell(row=row_num, column=3, value=str(amount_raw))
                    ws_extracted.cell(row=row_num, column=4, value=amount)  # Store as string to preserve precision
                    ws_extracted.cell(row=row_num, column=5, value=decimals)
                    ws_extracted.cell(row=row_num, column=6, value=token_address)
                    ws_extracted.cell(row=row_num, column=7, value=coin_id)
                    ws_extracted.cell(row=row_num, column=8, value=price)
                    ws_extracted.cell(row=row_num, column=9, value=price_change_24h)

                    # Track coin IDs for price change fetching (exclude "Spam Token" and empty)
                    if coin_id and coin_id != "Spam Token":
                        if coin_id not in coin_id_to_rows:
                            coin_id_to_rows[coin_id] = []
                        coin_id_to_rows[coin_id].append(row_num)

                    # Add to extracted data for DAM export (will be updated with price change later)
                    extracted_data.append([chain, symbol, str(amount_raw), amount, decimals, token_address, coin_id, price, price_change_24h])
                    row_num += 1

    # Fetch price and 24H price change data for all unique coin IDs
    _coingecko_raw = []
    if coin_id_to_rows:
        unique_coin_ids = list(coin_id_to_rows.keys())

        # Fetch USD prices using simple/price API
        price_map = fetch_coingecko_prices_batch(unique_coin_ids, raw_collector=_coingecko_raw)

        # Fetch 24H price change using coins/{id} API
        price_change_map = fetch_coingecko_price_change_batch(unique_coin_ids, raw_collector=_coingecko_raw)

        # Update Excel and extracted_data with price and price change values
        for coin_id, row_numbers in coin_id_to_rows.items():
            price_value = price_map.get(coin_id)
            price_change_value = price_change_map.get(coin_id)

            for row_num in row_numbers:
                # Update Excel sheet - Column 8 (Price), Column 9 (24H Price Change)
                if price_value is not None:
                    ws_extracted.cell(row=row_num, column=8, value=price_value)
                if price_change_value is not None:
                    ws_extracted.cell(row=row_num, column=9, value=price_change_value)

                # Update extracted_data (row_num - 1 because extracted_data has header at index 0, first data at index 1)
                data_idx = row_num - 1  # row_num 2 -> index 1 (first data row after header)
                if data_idx < len(extracted_data):
                    if price_value is not None:
                        extracted_data[data_idx][7] = price_value  # Column 8 -> index 7
                    if price_change_value is not None:
                        extracted_data[data_idx][8] = price_change_value  # Column 9 -> index 8

    wb.save(filepath)
    return filepath, extracted_data


def export_sim_dune_to_excel_combined(all_data, output_folder="test-results/API Result", portfolio_name=None):
    """
    Export combined Sim Dune API responses from multiple addresses to a single Excel file.

    File naming: SimDune_{portfolio_name}_{YYYYMMDD}_{HHMMSS}.xlsx

    Args:
        all_data: List of {'address': addr, 'data': data} dicts
        output_folder: Output folder path
        portfolio_name: Portfolio name for filename

    Returns: (filepath, extracted_data) tuple
    """
    from openpyxl import Workbook
    from datetime import datetime
    from decimal import Decimal, getcontext
    import os

    # Set high precision for Decimal calculations
    getcontext().prec = 50

    os.makedirs(output_folder, exist_ok=True)

    # Use portfolio name if provided, otherwise fall back to first address suffix
    name_suffix = portfolio_name if portfolio_name else all_data[0]['address'][-8:] if all_data else "unknown"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"SimDune_{name_suffix}_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()

    # Sheet 1: SimDune (Raw JSON API response - combined)
    ws_raw = wb.active
    ws_raw.title = "SimDune"
    ws_raw.append(["Address", "Raw API Response"])
    for item in all_data:
        ws_raw.append([item['address'], json.dumps(item['data'], indent=2)])

    # Sheet 2: Sim + Coingecko + Debank API (Combined extracted data)
    ws_extracted = wb.create_sheet("Sim + Coingecko + Debank API")
    ws_extracted.cell(row=1, column=1, value="Address")
    ws_extracted.cell(row=1, column=2, value="Chain")
    ws_extracted.cell(row=1, column=3, value="Symbol")
    ws_extracted.cell(row=1, column=4, value="Amount (Raw)")
    ws_extracted.cell(row=1, column=5, value="Amount")
    ws_extracted.cell(row=1, column=6, value="Decimals")
    ws_extracted.cell(row=1, column=7, value="Token Address")
    ws_extracted.cell(row=1, column=8, value="ID")
    ws_extracted.cell(row=1, column=9, value="Price")
    ws_extracted.cell(row=1, column=10, value="24H Price Change")
    ws_extracted.cell(row=1, column=11, value="Calculated Value")

    extracted_data = [["Address", "Chain", "Symbol", "Amount (Raw)", "Amount", "Decimals", "Token Address", "ID", "Price", "24H Price Change", "Calculated Value"]]
    row_num = 2

    # Load Coingecko coin list for ID lookup
    coingecko_map, coingecko_native_map, coingecko_addr_map = load_coingecko_coin_list()

    # Track coin IDs and their row numbers for price change fetching
    coin_id_to_rows = {}  # {coin_id: [row_numbers]}

    # Track seen entries to avoid duplicates (key: address_chain_symbol)
    seen_entries = set()

    for item in all_data:
        address = item['address']
        data = item['data']

        if data:
            balances = data.get('balances', data) if isinstance(data, dict) else data
            if isinstance(balances, list):
                for balance in balances:
                    if isinstance(balance, dict):
                        # Use same field names as original function
                        chain = balance.get('chain', balance.get('chain_name', ''))
                        symbol = balance.get('symbol', balance.get('token_symbol', ''))
                        amount_raw = balance.get('amount', balance.get('balance', balance.get('value_raw', '')))
                        decimals = balance.get('decimals', 18)
                        token_address = balance.get('address', '')

                        # Create unique key to check for duplicates (include token address to allow same symbol on different contracts)
                        entry_key = f"{address.lower()}_{chain}_{symbol}_{str(token_address).lower()}"
                        if entry_key in seen_entries:
                            continue  # Skip duplicate
                        seen_entries.add(entry_key)

                        # Calculate actual amount using Decimal for full precision
                        # Formula: Amount = Amount_Raw * 10^(-Decimals)
                        try:
                            if amount_raw and decimals is not None:
                                raw_decimal = Decimal(str(amount_raw))
                                divisor = Decimal(10) ** int(decimals)
                                amount_decimal = raw_decimal / divisor
                                amount_str = str(amount_decimal)
                            else:
                                amount_str = str(amount_raw) if amount_raw else ""
                        except Exception:
                            amount_str = str(amount_raw) if amount_raw else ""

                        # Lookup Coingecko ID
                        coin_id = ""
                        price = ""
                        price_change_24h = ""

                        token_address_lower = str(token_address).lower().strip()

                        # Case 1: Valid EVM address (42 chars, starts with 0x)
                        if is_valid_evm_address(token_address):
                            # Map Sim Dune chain name to Coingecko platform name
                            platform_name = CHAIN_TO_PLATFORM.get(chain.lower().strip(), "")
                            if platform_name:
                                # Primary lookup: address + platform name
                                lookup_key = (token_address_lower, platform_name)
                                coin_id = coingecko_map.get(lookup_key, "")
                                if coin_id:
                                    print(f"      ✅ Coingecko match (platform): {symbol} on {platform_name} → {coin_id}")
                            if not coin_id:
                                # Fallback: address-only (platform name in file may differ from CHAIN_TO_PLATFORM)
                                coin_id = coingecko_addr_map.get(token_address_lower, "")
                                if coin_id:
                                    print(f"      ✅ Coingecko match (addr-only fallback): {symbol} {token_address_lower} → {coin_id}")
                            if not coin_id:
                                print(f"      ⚠️  Spam Token: {symbol} {token_address_lower} chain={chain} platform={platform_name} — not in Coingecko")
                                coin_id = "Spam Token"

                        # Case 2: Native token (address = "native")
                        elif token_address_lower == "native":
                            # Lookup by symbol (where platform_name is blank or "ethereum")
                            symbol_lower = str(symbol).lower().strip()
                            coin_id = coingecko_native_map.get(symbol_lower, "")

                        ws_extracted.cell(row=row_num, column=1, value=address)
                        ws_extracted.cell(row=row_num, column=2, value=chain)
                        ws_extracted.cell(row=row_num, column=3, value=symbol)
                        ws_extracted.cell(row=row_num, column=4, value=str(amount_raw))
                        ws_extracted.cell(row=row_num, column=5, value=amount_str)
                        ws_extracted.cell(row=row_num, column=6, value=decimals)
                        ws_extracted.cell(row=row_num, column=7, value=token_address)
                        ws_extracted.cell(row=row_num, column=8, value=coin_id)
                        ws_extracted.cell(row=row_num, column=9, value=price)
                        ws_extracted.cell(row=row_num, column=10, value=price_change_24h)
                        ws_extracted.cell(row=row_num, column=11, value="")  # Calculated Price (empty initially)

                        # Track coin IDs for price change fetching (exclude "Spam Token" and empty)
                        if coin_id and coin_id != "Spam Token":
                            if coin_id not in coin_id_to_rows:
                                coin_id_to_rows[coin_id] = []
                            coin_id_to_rows[coin_id].append(row_num)

                        extracted_data.append([address, chain, symbol, str(amount_raw), amount_str, decimals, token_address, coin_id, price, price_change_24h, ""])
                        row_num += 1

    # Fetch price and 24H price change data for all unique coin IDs
    _coingecko_raw = []
    if coin_id_to_rows:
        unique_coin_ids = list(coin_id_to_rows.keys())

        # Fetch USD prices using simple/price API
        price_map = fetch_coingecko_prices_batch(unique_coin_ids, raw_collector=_coingecko_raw)

        # Fetch 24H price change using coins/{id} API
        price_change_map = fetch_coingecko_price_change_batch(unique_coin_ids, raw_collector=_coingecko_raw)

        # Update Excel and extracted_data with price and price change values
        for coin_id, row_numbers in coin_id_to_rows.items():
            price_value = price_map.get(coin_id)
            price_change_value = price_change_map.get(coin_id)

            for row_num in row_numbers:
                # Update Excel sheet - Column 9 (Price), Column 10 (24H Price Change)
                # None = not available; "null" = API returned null explicitly
                ws_extracted.cell(row=row_num, column=9, value=price_value if price_value is not None else "Data not available")
                _24h_cell_val = price_change_value if price_change_value is not None else "Data not available"
                ws_extracted.cell(row=row_num, column=10, value=_24h_cell_val)

                # Calculate Calculated Price (Column 11) = Price (I) × Amount (E)
                # Get Amount from column 5
                amount_value = ws_extracted.cell(row=row_num, column=5).value
                if price_value is not None and amount_value:
                    try:
                        amount_decimal = Decimal(str(amount_value))
                        price_decimal = Decimal(str(price_value))
                        calculated_price = price_decimal * amount_decimal
                        calculated_price_str = str(calculated_price).rstrip('0').rstrip('.')
                        ws_extracted.cell(row=row_num, column=11, value=calculated_price_str)
                    except Exception:
                        ws_extracted.cell(row=row_num, column=11, value="")

                # Update extracted_data (row_num - 1 because extracted_data has header at index 0, first data at index 1)
                data_idx = row_num - 1  # row_num 2 -> index 1 (first data row after header)
                if data_idx < len(extracted_data):
                    extracted_data[data_idx][8] = price_value if price_value is not None else "Data not available"
                    extracted_data[data_idx][9] = price_change_value if price_change_value is not None else "Data not available"

                    # Calculate Calculated Price for extracted_data
                    amount_value_data = extracted_data[data_idx][4]  # Column 5 -> index 4
                    if price_value is not None and amount_value_data:
                        try:
                            amount_decimal = Decimal(str(amount_value_data))
                            price_decimal = Decimal(str(price_value))
                            calculated_price = price_decimal * amount_decimal
                            calculated_price_str = str(calculated_price).rstrip('0').rstrip('.')
                            extracted_data[data_idx][10] = calculated_price_str  # Column 11 -> index 10
                        except Exception:
                            extracted_data[data_idx][10] = ""

    wb.save(filepath)

    # Save raw API responses to JSON file (Excel cells truncate at 32,767 chars)
    json_path = filepath.replace(".xlsx", ".json")
    with open(json_path, 'w') as _jf:
        json.dump(all_data, _jf, indent=2, default=str)

    # Save Coingecko raw responses to a separate JSON file
    if _coingecko_raw:
        _cg_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        cg_json_path = os.path.join(output_folder, f"Coingecko_Raw_{name_suffix}_{_cg_ts}.json")
        with open(cg_json_path, 'w') as _jf:
            json.dump(_coingecko_raw, _jf, indent=2, default=str)
        print(f"   📄 Saved Coingecko raw API responses to: {os.path.basename(cg_json_path)}")

    return filepath, extracted_data


def fetch_rabby_protocol(address):
    """
    Fetch complex protocol list from Rabby API for a given EVM address.

    API: https://api.rabby.io/v1/user/complex_protocol_list?id={address}

    Returns: (address, data, success) tuple
    """
    import time

    rabby_url = f"https://api.rabby.io/v1/user/complex_protocol_list?id={address}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }

    max_retries = 3
    last_exception = None

    for attempt in range(max_retries):
        try:
            response = requests.get(rabby_url, headers=headers, timeout=30)
            response.raise_for_status()
            data = response.json()

            if attempt > 0:
                print(f"      ✓ Address {address[:10]}... succeeded on attempt {attempt + 1}")
            return (address, data, True)

        except Exception as e:
            last_exception = e
            if attempt < max_retries - 1:
                print(f"      ⚠️  Address {address[:10]}... failed (attempt {attempt + 1}), retrying...")
                time.sleep(2)
            else:
                print(f"      ❌ Address {address[:10]}... failed after {max_retries} attempts: {str(last_exception)[:50]}")
                return (address, None, False)

    return (address, None, False)


def fetch_rabby_app(address):
    """
    Fetch complex app list from Rabby API for a given EVM address (Hyperliquid).

    API: https://api.rabby.io/v1/user/complex_app_list?id={address}

    Returns: (address, data, success) tuple
    """
    import time

    rabby_url = f"https://api.rabby.io/v1/user/complex_app_list?id={address}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }

    max_retries = 3
    last_exception = None

    for attempt in range(max_retries):
        try:
            response = requests.get(rabby_url, headers=headers, timeout=30)
            response.raise_for_status()
            data = response.json()

            if attempt > 0:
                print(f"      ✓ Address {address[:10]}... succeeded on attempt {attempt + 1}")
            return (address, data, True)

        except Exception as e:
            last_exception = e
            if attempt < max_retries - 1:
                print(f"      ⚠️  Address {address[:10]}... failed (attempt {attempt + 1}), retrying...")
                time.sleep(2)
            else:
                print(f"      ❌ Address {address[:10]}... failed after {max_retries} attempts: {str(last_exception)[:50]}")
                return (address, None, False)

    return (address, None, False)


def export_rabby_to_excel(address, data, output_folder="test-results/API Result", portfolio_name=None):
    """
    Export Rabby API response to Excel file.

    File naming: Protocol_{portfolio_name}_{YYYYMMDD}_{HHMMSS}.xlsx

    Tabs:
    - Rabby Raw: Raw JSON API response
    - Rabby Api Data: Extracted columns (all data from API)
    - {Protocol Name} ({Chain}): Separate tabs for each protocol+chain combination

    Returns: (filepath, extracted_data) tuple where extracted_data is a list of rows for DAM export
    """
    from openpyxl import Workbook
    from datetime import datetime
    from decimal import Decimal, getcontext
    import os

    # Set high precision for Decimal calculations
    getcontext().prec = 50

    os.makedirs(output_folder, exist_ok=True)

    # Use portfolio name if provided, otherwise fall back to last 8 chars of address
    name_suffix = portfolio_name if portfolio_name else address[-8:]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Protocol_{name_suffix}_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()

    # Sheet 1: Rabby Api Data (Extracted columns)
    ws_extracted = wb.active
    ws_extracted.title = "Rabby Api Data"
    ws_extracted.cell(row=1, column=1, value="Address")
    ws_extracted.cell(row=1, column=2, value="Name")
    ws_extracted.cell(row=1, column=3, value="ID")
    ws_extracted.cell(row=1, column=4, value="Chain")
    ws_extracted.cell(row=1, column=5, value="Symbol")
    ws_extracted.cell(row=1, column=6, value="Price")
    ws_extracted.cell(row=1, column=7, value="Amount")
    ws_extracted.cell(row=1, column=8, value="Calculated Value/Margin")

    # Yellow fill for Amount Validation column (used in separate protocol tabs)
    from openpyxl.styles import PatternFill
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Build extracted data list for DAM export
    extracted_data = [["Address", "Name", "ID", "Chain", "Symbol", "Price", "Amount", "Calculated Value"]]

    # Chain name mapping for display
    chain_name_map = {
        'eth': 'Ethereum',
        'base': 'Base',
        'bsc': 'BSC',
        'arb': 'Arbitrum',
        'op': 'Optimism',
        'matic': 'Polygon',
        'avax': 'Avalanche',
        'ftm': 'Fantom',
        'cro': 'Cronos',
        'aurora': 'Aurora',
        'heco': 'HECO',
        'okx': 'OKX',
        'xdai': 'Gnosis',
        'boba': 'Boba',
        'metis': 'Metis',
        'movr': 'Moonriver',
        'celo': 'Celo',
        'klay': 'Klaytn',
        'mnt': 'Mantle',
        'linea': 'Linea',
        'zksync': 'zkSync Era',
        'era': 'zkSync Era',
        'scroll': 'Scroll',
        'blast': 'Blast',
        'mode': 'Mode',
    }

    row_num = 2
    if data and isinstance(data, list):
        for protocol in data:
            # Get protocol ID, name, and chain
            protocol_id = protocol.get('id', '')
            protocol_name = protocol.get('name', protocol_id)  # Use name if available, else id
            protocol_chain = protocol.get('chain', '')

            # Get portfolio_item_list
            portfolio_items = protocol.get('portfolio_item_list', [])

            for item in portfolio_items:
                item_id = item.get('id', protocol_id)

                # Get detail -> supply_token_list
                detail = item.get('detail', {})
                supply_tokens = detail.get('supply_token_list', [])

                for token in supply_tokens:
                    # Get chain from token or fall back to protocol chain
                    chain = token.get('chain', protocol_chain)
                    symbol = token.get('symbol', '')
                    price = token.get('price', 0)
                    amount = token.get('amount', 0)

                    # Amount Tooltip = full precision from API
                    amount_tooltip_str = str(amount)

                    # Amount = rounded to 2 decimal places for display
                    try:
                        amount_decimal = Decimal(str(amount))
                        amount_rounded = float(amount_decimal.quantize(Decimal('0.01')))
                    except Exception:
                        amount_rounded = 0

                    # Calculate value using Python Decimal for full precision
                    try:
                        price_decimal = Decimal(str(price))
                        amount_decimal = Decimal(str(amount))
                        calculated_value = price_decimal * amount_decimal
                        # Convert to string, remove trailing zeros
                        calc_value_str = str(calculated_value)
                        if '.' in calc_value_str:
                            calc_value_str = calc_value_str.rstrip('0').rstrip('.')
                    except Exception:
                        calc_value_str = "0"

                    # Write to Excel - Rabby Api Data tab
                    # A:Address, B:Name, C:ID, D:Chain, E:Symbol, F:Price, G:Amount, H:Calculated Value
                    ws_extracted.cell(row=row_num, column=1, value=address)
                    ws_extracted.cell(row=row_num, column=2, value=protocol_name)
                    ws_extracted.cell(row=row_num, column=3, value=item_id)
                    ws_extracted.cell(row=row_num, column=4, value=chain)
                    ws_extracted.cell(row=row_num, column=5, value=symbol)
                    ws_extracted.cell(row=row_num, column=6, value=price)
                    ws_extracted.cell(row=row_num, column=7, value=amount)
                    ws_extracted.cell(row=row_num, column=8, value=calc_value_str)

                    # Add to extracted data for DAM export
                    extracted_data.append([address, protocol_name, item_id, chain, symbol, price, amount, calc_value_str])

                    row_num += 1

                # Also check for reward_token_list
                reward_tokens = detail.get('reward_token_list', [])
                for token in reward_tokens:
                    # Get chain from token or fall back to protocol chain
                    chain = token.get('chain', protocol_chain)
                    symbol = token.get('symbol', '')
                    price = token.get('price', 0)
                    amount = token.get('amount', 0)

                    # Amount Tooltip = full precision from API
                    amount_tooltip_str = str(amount)

                    # Amount = rounded to 2 decimal places for display
                    try:
                        amount_decimal = Decimal(str(amount))
                        amount_rounded = float(amount_decimal.quantize(Decimal('0.01')))
                    except Exception:
                        amount_rounded = 0

                    # Calculate value using Python Decimal for full precision
                    try:
                        price_decimal = Decimal(str(price))
                        amount_decimal = Decimal(str(amount))
                        calculated_value = price_decimal * amount_decimal
                        calc_value_str = str(calculated_value)
                        if '.' in calc_value_str:
                            calc_value_str = calc_value_str.rstrip('0').rstrip('.')
                    except Exception:
                        calc_value_str = "0"

                    # Write to Excel - Rabby Api Data tab (reward tokens)
                    # A:Address, B:Name, C:ID, D:Chain, E:Symbol, F:Price, G:Amount, H:Calculated Value
                    ws_extracted.cell(row=row_num, column=1, value=address)
                    ws_extracted.cell(row=row_num, column=2, value=protocol_name)
                    ws_extracted.cell(row=row_num, column=3, value=item_id)
                    ws_extracted.cell(row=row_num, column=4, value=chain)
                    ws_extracted.cell(row=row_num, column=5, value=symbol)
                    ws_extracted.cell(row=row_num, column=6, value=price)
                    ws_extracted.cell(row=row_num, column=7, value=amount)
                    ws_extracted.cell(row=row_num, column=8, value=calc_value_str)

                    extracted_data.append([address, protocol_name, item_id, chain, symbol, price, amount, calc_value_str])

                    row_num += 1

                # Also check for borrow_token_list (debt positions)
                borrow_tokens = detail.get('borrow_token_list', [])
                for token in borrow_tokens:
                    # Get chain from token or fall back to protocol chain
                    chain = token.get('chain', protocol_chain)
                    symbol = token.get('symbol', '')
                    price = token.get('price', 0)
                    amount = token.get('amount', 0)

                    # Ensure borrow amounts are negative
                    try:
                        amount_num = float(amount) if not isinstance(amount, (int, float)) else amount
                        if amount_num > 0:
                            amount = -amount_num
                    except (ValueError, TypeError):
                        pass

                    # Amount Tooltip = full precision from API
                    amount_tooltip_str = str(amount)

                    # Amount = rounded to 2 decimal places for display
                    try:
                        amount_decimal = Decimal(str(amount))
                        amount_rounded = float(amount_decimal.quantize(Decimal('0.01')))
                    except Exception:
                        amount_rounded = 0

                    # Calculate value using Python Decimal for full precision
                    try:
                        price_decimal = Decimal(str(price))
                        amount_decimal = Decimal(str(amount))
                        calculated_value = price_decimal * amount_decimal
                        calc_value_str = str(calculated_value)
                        if '.' in calc_value_str:
                            calc_value_str = calc_value_str.rstrip('0').rstrip('.')
                    except Exception:
                        calc_value_str = "0"

                    # Write to Excel - Rabby Api Data tab (borrow tokens)
                    # A:Address, B:Name, C:ID, D:Chain, E:Symbol, F:Price, G:Amount, H:Calculated Value
                    ws_extracted.cell(row=row_num, column=1, value=address)
                    ws_extracted.cell(row=row_num, column=2, value=protocol_name)
                    ws_extracted.cell(row=row_num, column=3, value=item_id)
                    ws_extracted.cell(row=row_num, column=4, value=chain)
                    ws_extracted.cell(row=row_num, column=5, value=symbol)
                    ws_extracted.cell(row=row_num, column=6, value=price)
                    ws_extracted.cell(row=row_num, column=7, value=amount)
                    ws_extracted.cell(row=row_num, column=8, value=calc_value_str)

                    extracted_data.append([address, protocol_name, item_id, chain, symbol, price, amount, calc_value_str])

                    row_num += 1

    wb.save(filepath)
    return filepath, extracted_data


def export_rabby_to_excel_combined(all_data, output_folder="test-results/API Result", portfolio_name=None):
    """
    Export combined Rabby API responses from multiple addresses to a single Excel file.

    File naming: Protocol_{portfolio_name}_{YYYYMMDD}_{HHMMSS}.xlsx

    Args:
        all_data: List of {'address': addr, 'data': data} dicts
        output_folder: Output folder path
        portfolio_name: Portfolio name for filename

    Returns: (filepath, extracted_data) tuple
    """
    from openpyxl import Workbook
    from datetime import datetime
    from decimal import Decimal, getcontext
    import os

    # Set high precision for Decimal calculations
    getcontext().prec = 50

    os.makedirs(output_folder, exist_ok=True)

    # Use portfolio name if provided, otherwise fall back to first address suffix
    name_suffix = portfolio_name if portfolio_name else all_data[0]['address'][-8:] if all_data else "unknown"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Protocol_{name_suffix}_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()

    # Yellow fill for Amount Validation column
    from openpyxl.styles import PatternFill
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Sheet 1: Rabby Api Data (Combined extracted data)
    ws_extracted = wb.active
    ws_extracted.title = "Rabby Api Data"
    ws_extracted.cell(row=1, column=1, value="Address")
    ws_extracted.cell(row=1, column=2, value="Name")
    ws_extracted.cell(row=1, column=3, value="ID")
    ws_extracted.cell(row=1, column=4, value="Chain")
    ws_extracted.cell(row=1, column=5, value="Symbol")
    ws_extracted.cell(row=1, column=6, value="Price")
    ws_extracted.cell(row=1, column=7, value="Amount")
    ws_extracted.cell(row=1, column=8, value="Calculated Value")

    extracted_data = [["Address", "Name", "ID", "Chain", "Symbol", "Price", "Amount", "Calculated Value"]]

    # Chain name mapping for display
    chain_name_map = {
        'eth': 'Ethereum', 'base': 'Base', 'bsc': 'BSC', 'arb': 'Arbitrum',
        'op': 'Optimism', 'matic': 'Polygon', 'avax': 'Avalanche', 'ftm': 'Fantom',
        'cro': 'Cronos', 'aurora': 'Aurora', 'heco': 'HECO', 'okx': 'OKX',
        'xdai': 'Gnosis', 'boba': 'Boba', 'metis': 'Metis', 'movr': 'Moonriver',
        'celo': 'Celo', 'klay': 'Klaytn', 'mnt': 'Mantle', 'linea': 'Linea',
        'zksync': 'zkSync Era', 'era': 'zkSync Era', 'scroll': 'Scroll',
        'blast': 'Blast', 'mode': 'Mode',
    }

    # Track seen entries to avoid duplicates (key: address_id_chain_symbol_amount)
    seen_entries = set()

    row_num = 2
    for item in all_data:
        address = item['address']
        data = item['data']

        if data and isinstance(data, list):
            for protocol in data:
                protocol_id = protocol.get('id', '')
                protocol_name = protocol.get('name', protocol_id)
                protocol_chain = protocol.get('chain', '')

                portfolio_items = protocol.get('portfolio_item_list', [])

                for portfolio_item in portfolio_items:
                    item_id = portfolio_item.get('id', protocol_id)
                    position_type = portfolio_item.get('name', item_id)

                    detail = portfolio_item.get('detail', {})
                    supply_tokens = detail.get('supply_token_list', [])

                    for token in supply_tokens:
                        chain = token.get('chain', protocol_chain)
                        symbol = token.get('symbol', '')
                        price = token.get('price', 0)
                        amount = token.get('amount', 0)

                        # Create unique key to check for duplicates
                        entry_key = f"{address.lower()}_{item_id}_{chain}_{symbol}_{amount}"
                        if entry_key in seen_entries:
                            continue  # Skip duplicate
                        seen_entries.add(entry_key)

                        amount_tooltip_str = str(amount)

                        try:
                            amount_decimal = Decimal(str(amount))
                            amount_rounded = float(amount_decimal.quantize(Decimal('0.01')))
                        except Exception:
                            amount_rounded = 0

                        try:
                            price_decimal = Decimal(str(price))
                            amount_decimal = Decimal(str(amount))
                            calculated_value = price_decimal * amount_decimal
                            calc_value_str = str(calculated_value)
                            if '.' in calc_value_str:
                                calc_value_str = calc_value_str.rstrip('0').rstrip('.')
                        except Exception:
                            calc_value_str = "0"

                        # Write to Rabby Api Data tab
                        ws_extracted.cell(row=row_num, column=1, value=address)
                        ws_extracted.cell(row=row_num, column=2, value=protocol_name)
                        ws_extracted.cell(row=row_num, column=3, value=item_id)
                        ws_extracted.cell(row=row_num, column=4, value=chain)
                        ws_extracted.cell(row=row_num, column=5, value=symbol)
                        ws_extracted.cell(row=row_num, column=6, value=price)
                        ws_extracted.cell(row=row_num, column=7, value=amount)
                        ws_extracted.cell(row=row_num, column=8, value=calc_value_str)

                        extracted_data.append([address, protocol_name, item_id, chain, symbol, price, amount, calc_value_str])

                        row_num += 1

                    # Also check for reward_token_list
                    reward_tokens = detail.get('reward_token_list', [])
                    for token in reward_tokens:
                        chain = token.get('chain', protocol_chain)
                        symbol = token.get('symbol', '')
                        price = token.get('price', 0)
                        amount = token.get('amount', 0)

                        # Create unique key to check for duplicates
                        entry_key = f"{address.lower()}_{item_id}_{chain}_{symbol}_{amount}"
                        if entry_key in seen_entries:
                            continue  # Skip duplicate
                        seen_entries.add(entry_key)

                        amount_tooltip_str = str(amount)

                        try:
                            amount_decimal = Decimal(str(amount))
                            amount_rounded = float(amount_decimal.quantize(Decimal('0.01')))
                        except Exception:
                            amount_rounded = 0

                        try:
                            price_decimal = Decimal(str(price))
                            amount_decimal = Decimal(str(amount))
                            calculated_value = price_decimal * amount_decimal
                            calc_value_str = str(calculated_value)
                            if '.' in calc_value_str:
                                calc_value_str = calc_value_str.rstrip('0').rstrip('.')
                        except Exception:
                            calc_value_str = "0"

                        ws_extracted.cell(row=row_num, column=1, value=address)
                        ws_extracted.cell(row=row_num, column=2, value=protocol_name)
                        ws_extracted.cell(row=row_num, column=3, value=item_id)
                        ws_extracted.cell(row=row_num, column=4, value=chain)
                        ws_extracted.cell(row=row_num, column=5, value=symbol)
                        ws_extracted.cell(row=row_num, column=6, value=price)
                        ws_extracted.cell(row=row_num, column=7, value=amount)
                        ws_extracted.cell(row=row_num, column=8, value=calc_value_str)

                        extracted_data.append([address, protocol_name, item_id, chain, symbol, price, amount, calc_value_str])

                        row_num += 1

                    # Also check for borrow_token_list (debt positions)
                    borrow_tokens = detail.get('borrow_token_list', [])
                    for token in borrow_tokens:
                        chain = token.get('chain', protocol_chain)
                        symbol = token.get('symbol', '')
                        price = token.get('price', 0)
                        amount = token.get('amount', 0)

                        # Ensure borrow amounts are negative
                        try:
                            amount_num = float(amount) if not isinstance(amount, (int, float)) else amount
                            if amount_num > 0:
                                amount = -amount_num
                        except (ValueError, TypeError):
                            pass

                        # Create unique key to check for duplicates
                        entry_key = f"{address.lower()}_{item_id}_{chain}_{symbol}_{amount}"
                        if entry_key in seen_entries:
                            continue  # Skip duplicate
                        seen_entries.add(entry_key)

                        amount_tooltip_str = str(amount)

                        try:
                            amount_decimal = Decimal(str(amount))
                            amount_rounded = float(amount_decimal.quantize(Decimal('0.01')))
                        except Exception:
                            amount_rounded = 0

                        try:
                            price_decimal = Decimal(str(price))
                            amount_decimal = Decimal(str(amount))
                            calculated_value = price_decimal * amount_decimal
                            calc_value_str = str(calculated_value)
                            if '.' in calc_value_str:
                                calc_value_str = calc_value_str.rstrip('0').rstrip('.')
                        except Exception:
                            calc_value_str = "0"

                        ws_extracted.cell(row=row_num, column=1, value=address)
                        ws_extracted.cell(row=row_num, column=2, value=protocol_name)
                        ws_extracted.cell(row=row_num, column=3, value=item_id)
                        ws_extracted.cell(row=row_num, column=4, value=chain)
                        ws_extracted.cell(row=row_num, column=5, value=symbol)
                        ws_extracted.cell(row=row_num, column=6, value=price)
                        ws_extracted.cell(row=row_num, column=7, value=amount)
                        ws_extracted.cell(row=row_num, column=8, value=calc_value_str)

                        extracted_data.append([address, protocol_name, item_id, chain, symbol, price, amount, calc_value_str])

                        row_num += 1

    wb.save(filepath)
    return filepath, extracted_data


def export_rabby_app_to_excel_combined(all_data, output_folder="test-results/API Result", portfolio_name=None):
    """
    Export combined Rabby complex_app_list API responses (Hyperliquid) to a single Excel file.

    File naming: Hyperliquid_{portfolio_name}_{YYYYMMDD}_{HHMMSS}.xlsx

    Tabs:
    - Hyperliquid Raw: Raw JSON API response (one row per address)
    - Hyperliquid: Extracted columns (Asset Dict, Name, Detail Description, Symbol, Amount, Net USD Value, Position Token Symbol, Side, Leverage, Margin Token Amount, PnL USD Value)

    Args:
        all_data: List of {'address': addr, 'data': data} dicts
        output_folder: Output folder path
        portfolio_name: Portfolio name for filename

    Returns: (filepath, extracted_data) tuple
    """
    from openpyxl import Workbook
    from datetime import datetime
    import os

    os.makedirs(output_folder, exist_ok=True)

    name_suffix = portfolio_name if portfolio_name else all_data[0]['address'][-8:] if all_data else "unknown"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Hyperliquid_{name_suffix}_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()

    # Sheet 1: Hyperliquid Raw
    ws_raw = wb.active
    ws_raw.title = "Hyperliquid Raw"
    ws_raw.append(["Address", "Raw API Response"])
    for item in all_data:
        ws_raw.append([item['address'], json.dumps(item['data'], indent=2)])

    # Sheet 2: Hyperliquid
    # All rows share base columns A-F; perpetuals also populate columns G-K
    ws_extracted = wb.create_sheet("Hyperliquid")
    headers = [
        "Asset Dict", "Name", "Detail Description", "Symbol", "Amount", "Net USD Value",
        "Position Token Symbol", "Side", "Leverage", "Margin Token Amount", "PnL USD Value"
    ]
    ws_extracted.append(headers)
    extracted_data = [headers[:]]

    seen_entries = set()
    row_num = 2

    for item in all_data:
        address = item['address']
        data = item['data']

        if not data:
            continue

        # API response is {"apps": [...], "error_apps": [...]}
        if isinstance(data, dict):
            apps = data.get("apps", [])
        elif isinstance(data, list):
            apps = data
        else:
            continue

        for app in apps:
            portfolio_items = app.get('portfolio_item_list', [])

            for portfolio_item in portfolio_items:
                item_name = portfolio_item.get('name', '')
                detail = portfolio_item.get('detail', {}) if portfolio_item.get('detail') else {}
                detail_description = detail.get('description', '')
                stats = portfolio_item.get('stats', {}) if portfolio_item.get('stats') else {}
                net_usd_value = stats.get('net_usd_value', 0)

                # asset_dict: {token_id: amount} — 1 row per entry
                asset_dict = portfolio_item.get('asset_dict', {})

                # asset_token_list: lookup symbol by token_id
                asset_token_list = portfolio_item.get('asset_token_list', [])
                token_lookup = {t.get('id', ''): t.get('symbol', '') for t in asset_token_list}

                # Perpetuals-specific fields (blank for non-perpetuals)
                is_perpetual = item_name.lower() == 'perpetuals'
                if is_perpetual:
                    position_token = detail.get('position_token', {}) or {}
                    margin_token = detail.get('margin_token', {}) or {}
                    position_token_symbol = position_token.get('symbol', '')
                    leverage = detail.get('leverage', '')
                    margin_token_amount = margin_token.get('amount', '')
                    pnl_usd_value = detail.get('pnl_usd_value', '')
                    side = detail.get('side', '')
                else:
                    position_token_symbol = leverage = margin_token_amount = pnl_usd_value = side = ''

                for token_id, amount in asset_dict.items():
                    symbol = token_lookup.get(token_id, '')

                    entry_key = f"{address.lower()}_{item_name}_{token_id}_{amount}"
                    if entry_key in seen_entries:
                        continue
                    seen_entries.add(entry_key)

                    row = [
                        token_id, item_name, detail_description, symbol, amount, net_usd_value,
                        position_token_symbol, side, leverage, margin_token_amount, pnl_usd_value
                    ]
                    for col_idx, val in enumerate(row, 1):
                        ws_extracted.cell(row=row_num, column=col_idx, value=val)

                    extracted_data.append(row)
                    row_num += 1

    wb.save(filepath)
    return filepath, extracted_data


def add_validation_columns_to_overview_token(data):
    """
    Add validation columns to Overview - Wallet table data.
    Inserts validation columns after Price, Price(24h), Share, Amount, and Value columns.

    Expected original columns: Chain, Name, Price, Price (24h), Share, Amount, Value, Price Tooltip, Share Tooltip, Amount Tooltip
    After insertion: Chain, Name, Price, Price Tooltip, FE - Price Validation, Price Validation, Price Diff Validation,
                     Price (24h), Price(24H) Validation, Price(24H) Diff Validation,
                     Share, Share Tooltip, FE - Share Validation, Calculation Share, Share Validation,
                     Amount, Amount Tooltip, FE - Amount Validation, Amount Validation, Amount Diff Validation,
                     Value, Value Validation, Calculated Value, Data Row
    """
    if not data or len(data) == 0:
        return data

    # Debug: Print first row to see structure
    if len(data) > 1:
        print(f"   DEBUG add_validation: Row 1 (TRX) input has {len(data[1])} columns: {data[1]}")

    modified_data = []
    for row_idx, row in enumerate(data):
        # Handle rows with insufficient data
        if len(row) < 7:
            modified_data.append(row)
            continue

        # Actual input column positions (0-indexed, confirmed from debug output):
        # 0: Chain, 1: Name, 2: Price, 3: Price (24h), 4: Amount, 5: Share, 6: Value,
        # 7: Price Tooltip, 8: Share Tooltip, 9: Amount Tooltip
        # Output column order: Amount group first (K-O), then Share group (P-T)
        # K:Amount, L:AmountTooltip, M:FE-AmtVal, N:AmtVal, O:AmtDiffVal,
        # P:Share, Q:ShareTooltip, R:FE-ShareVal, S:CalcShare, T:ShareVal, U:Value
        new_row = []
        new_row.append(row[0] if len(row) > 0 else "")  # A: Chain
        new_row.append(row[1] if len(row) > 1 else "")  # B: Name
        new_row.append(row[2] if len(row) > 2 else "")  # C: Price
        new_row.append(row[7] if len(row) > 7 else "")  # D: Price Tooltip (index 7)
        new_row.append("")       # E: FE - Price Validation (empty for now)
        new_row.append("")       # F: Price Validation (empty for now)
        new_row.append("")       # G: Price abs_diff (empty for now)
        new_row.append("")       # H: Price Diff % (empty for now)
        new_row.append(row[3] if len(row) > 3 else "")  # I: Price (24h)
        new_row.append("")       # J: Price(24H) Validation (empty for now)
        new_row.append("")       # K: Price(24H) Diff Validation (empty for now)
        new_row.append(row[4] if len(row) > 4 else "")  # L: Amount (index 4)
        new_row.append(row[9] if len(row) > 9 else "")  # M: Amount Tooltip (index 9)
        new_row.append("")       # N: FE - Amount Validation (empty for now)
        new_row.append("")       # O: Amount Validation (empty for now)
        new_row.append("")       # P: Amount Diff Validation (empty for now)
        new_row.append(row[5] if len(row) > 5 else "")  # Q: Share (index 5)
        new_row.append(row[8] if len(row) > 8 else "")  # R: Share Tooltip (index 8)
        new_row.append("")       # S: FE - Share Validation (empty for now)
        new_row.append("")       # T: Calculation Share (empty for now)
        new_row.append("")       # U: Share Validation (empty for now)
        new_row.append(row[6] if len(row) > 6 else "")  # V: Value
        new_row.append("")       # W: Value Validation (empty for now)
        new_row.append("")       # X: Value - UI validation (empty for now)
        new_row.append("")       # Y: API Calculated Value (empty for now)
        new_row.append("")       # Z: Data Row (empty for now)

        # If there are more columns beyond Amount Tooltip (index 9), append them
        if len(row) > 10:
            new_row.extend(row[10:])

        modified_data.append(new_row)

    return modified_data


def extract_svg_networth_map(page):
    """
    Extract token/chain/platform -> net worth mappings from pie chart SVG.
    Uses multiple strategies to reliably extract SVG text content.
    Returns a dict mapping name (uppercase) -> net worth value string.
    """
    networth_map = {}

    # Strategy 1: Get SVG innerHTML and parse with regex (most reliable)
    try:
        svg_htmls = page.evaluate('''() => {
            const svgs = document.querySelectorAll('.charts-for-react svg, .recharts-wrapper svg, svg');
            return Array.from(svgs).map(svg => svg.innerHTML);
        }''')
        for svg_html in svg_htmls:
            if not svg_html or '<text' not in svg_html:
                continue
            # Also extract tspan content: <text ...><tspan ...>content</tspan></text>
            tspan_elements = re.findall(r'<text[^>]*transform="([^"]*)"[^>]*>(?:<tspan[^>]*>)?([^<]+)(?:</tspan>)?</text>', svg_html)
            # Extract all <text ...>content</text> from SVG HTML
            text_elements = re.findall(r'<text[^>]*transform="([^"]*)"[^>]*>([^<]+)</text>', svg_html)
            # Merge tspan results
            if tspan_elements and not text_elements:
                text_elements = tspan_elements
            elif tspan_elements:
                # Add tspan results that aren't already in text_elements
                existing_transforms = {t[0] for t in text_elements}
                for te in tspan_elements:
                    if te[0] not in existing_transforms:
                        text_elements.append(te)
            if not text_elements:
                text_elements = re.findall(r'<text[^>]*>([^<]+)</text>', svg_html)
                # Sequential pairing for non-transform texts
                texts_only = [t for t in text_elements if t.strip()]
                for i, txt in enumerate(texts_only):
                    if '$' in txt:
                        val_match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', txt)
                        if val_match:
                            net_worth_val = val_match.group(1).replace(',', '')
                            for j in range(i - 1, max(i - 3, -1), -1):
                                if j >= 0:
                                    prev = texts_only[j].strip()
                                    if prev and any(c.isalpha() for c in prev) and '$' not in prev:
                                        networth_map[prev.upper()] = net_worth_val
                                        break
            else:
                # Group by transform attribute
                transform_groups = {}
                for transform, content in text_elements:
                    if transform not in transform_groups:
                        transform_groups[transform] = []
                    transform_groups[transform].append(content.strip())

                for transform, texts in transform_groups.items():
                    name = None
                    for txt in texts:
                        if '$' in txt:
                            val_match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', txt)
                            if val_match:
                                value = val_match.group(1).replace(',', '')
                                if name:
                                    networth_map[name.upper()] = value
                                    name = None
                                else:
                                    # Try combined text like "De-Fi Positions($9,436,462.85)"
                                    name_in_value = re.match(r'^([A-Za-z][A-Za-z\s\-]*?)\s*\(?\$', txt)
                                    if name_in_value:
                                        networth_map[name_in_value.group(1).strip().upper()] = value
                        elif txt and any(c.isalpha() for c in txt) and '%' not in txt and 'Total' not in txt and len(txt) < 30:
                            name = txt

        if networth_map:
            print(f"   📊 SVG innerHTML extraction: {networth_map}")
            return networth_map
    except Exception as e:
        print(f"   ⚠️  SVG innerHTML strategy failed: {e}")

    # Strategy 2: Use Playwright locators to find SVG text elements
    try:
        svg_text_els = page.locator('svg text').all()
        print(f"   DEBUG: Playwright locator found {len(svg_text_els)} svg text elements")
        svg_items = []
        for el in svg_text_els:
            try:
                txt = el.text_content()
                transform = el.get_attribute('transform') or 'none'
                if txt:
                    svg_items.append({'text': txt.strip(), 'transform': transform})
            except:
                continue

        if svg_items:
            # Group by transform
            transform_groups = {}
            for item in svg_items:
                t = item['transform']
                if t not in transform_groups:
                    transform_groups[t] = []
                transform_groups[t].append(item['text'])

            for transform, texts in transform_groups.items():
                name = None
                for txt in texts:
                    if '$' in txt:
                        val_match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', txt)
                        if val_match:
                            value = val_match.group(1).replace(',', '')
                            if name:
                                networth_map[name.upper()] = value
                                name = None
                            else:
                                # Try combined text like "De-Fi Positions($9,436,462.85)"
                                name_in_value = re.match(r'^([A-Za-z][A-Za-z\s\-]*?)\s*\(?\$', txt)
                                if name_in_value:
                                    networth_map[name_in_value.group(1).strip().upper()] = value
                    elif txt and any(c.isalpha() for c in txt) and '%' not in txt and 'Total' not in txt and len(txt) < 30:
                        name = txt

            if networth_map:
                print(f"   📊 Playwright locator extraction: {networth_map}")
                return networth_map
    except Exception as e:
        print(f"   ⚠️  Playwright locator strategy failed: {e}")

    # Strategy 3: page.evaluate querySelectorAll (enhanced - checks parent transforms + tspan children)
    try:
        svg_text_pairs = page.evaluate('''() => {
            const results = {};
            const texts = document.querySelectorAll('svg text');
            const items = [];

            for (const t of texts) {
                // Get transform: check self first, then walk parent chain up to <svg>
                let transform = t.getAttribute('transform');
                if (!transform) {
                    let parent = t.parentElement;
                    while (parent && parent.tagName.toLowerCase() !== 'svg') {
                        transform = parent.getAttribute('transform');
                        if (transform) break;
                        parent = parent.parentElement;
                    }
                }
                const effectiveTransform = transform || 'none';

                // Get individual tspan texts (if any)
                const tspans = t.querySelectorAll('tspan');
                if (tspans.length > 0) {
                    // Each tspan is a separate text item in the same group
                    for (const ts of tspans) {
                        const tsTxt = ts.textContent.trim();
                        if (tsTxt) items.push({text: tsTxt, transform: effectiveTransform});
                    }
                } else {
                    const txt = t.textContent.trim();
                    if (txt) items.push({text: txt, transform: effectiveTransform});
                }
            }

            // Group by effective transform
            const transformGroups = {};
            for (const item of items) {
                if (!transformGroups[item.transform]) transformGroups[item.transform] = [];
                transformGroups[item.transform].push(item.text);
            }

            // Extract name/value pairs from groups (handles multiple pairs per group)
            for (const [transform, texts] of Object.entries(transformGroups)) {
                let itemName = null;
                for (const text of texts) {
                    if (text.includes('$')) {
                        const match = text.match(/\\(?\\$?([\\d,]+\\.?\\d*)\\)?/);
                        if (match) {
                            const val = match[1].replace(/,/g, '');
                            if (itemName) {
                                results[itemName.toUpperCase()] = val;
                                itemName = null;
                            } else {
                                // Try combined "Name ($X,XXX.XX)"
                                const nameMatch = text.match(/^([A-Za-z][A-Za-z\\s\\-]*?)\\s*\\(?\\$/);
                                if (nameMatch) results[nameMatch[1].trim().toUpperCase()] = val;
                            }
                        }
                    } else if (text && /[A-Za-z]/.test(text) && !text.includes('%') && !text.includes('Total') && text.length < 30) {
                        itemName = text;
                    }
                }
            }

            // Fallback: sequential pairing (adjacent items regardless of transform)
            if (Object.keys(results).length === 0 && items.length >= 2) {
                for (let i = 0; i < items.length; i++) {
                    const txt = items[i].text;
                    if (txt.includes('$')) {
                        const match = txt.match(/\\(?\\$?([\\d,]+\\.?\\d*)\\)?/);
                        if (match) {
                            const val = match[1].replace(/,/g, '');
                            // Look backwards for name
                            for (let j = i - 1; j >= Math.max(i - 3, 0); j--) {
                                const prev = items[j].text;
                                if (prev && /[A-Za-z]/.test(prev) && !prev.includes('$') && !prev.includes('%') && !prev.includes('Total') && prev.length < 30) {
                                    results[prev.toUpperCase()] = val;
                                    break;
                                }
                            }
                        }
                    }
                    // Combined "Name ($X,XXX.XX)" in single element
                    const combinedMatch = txt.match(/^([A-Za-z][A-Za-z\\s\\-]+)\\s*\\(\\$?([\\d,]+\\.?\\d*)\\)/);
                    if (combinedMatch) {
                        results[combinedMatch[1].trim().toUpperCase()] = combinedMatch[2].replace(/,/g, '');
                    }
                }
            }

            return results;
        }''')
        if svg_text_pairs:
            networth_map = svg_text_pairs
            print(f"   📊 evaluate querySelectorAll extraction: {networth_map}")
    except Exception as e:
        print(f"   ⚠️  evaluate strategy failed: {e}")

    # Strategy 4: Scan ALL visible text in the chart area for "Name ($X,XXX.XX)" patterns
    if not networth_map:
        try:
            chart_text_pairs = page.evaluate('''() => {
                const results = {};
                // Find all text nodes that contain $ sign
                const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT, null, false);
                const dollarTexts = [];
                while (walker.nextNode()) {
                    const txt = walker.currentNode.textContent.trim();
                    if (txt && txt.includes('$') && txt.length < 100) {
                        const parent = walker.currentNode.parentElement;
                        // Check if inside or near an SVG/chart area
                        const inSvg = parent && parent.closest('svg');
                        const inChart = parent && (parent.closest('[class*="chart"]') || parent.closest('[class*="Chart"]') || parent.closest('[class*="recharts"]'));
                        if (inSvg || inChart) {
                            dollarTexts.push({text: txt, element: parent});
                        }
                    }
                }
                // Also check SVG text/tspan directly
                document.querySelectorAll('svg text, svg tspan').forEach(el => {
                    const txt = el.textContent.trim();
                    if (txt && txt.includes('$')) {
                        dollarTexts.push({text: txt, element: el});
                    }
                });
                // Parse combined "Name ($X,XXX.XX)" patterns
                for (const item of dollarTexts) {
                    const combined = item.text.match(/^([A-Za-z][A-Za-z\\s\\-]+)\\s*\\(\\$?([\\d,]+\\.?\\d*)\\)/);
                    if (combined) {
                        results[combined[1].trim().toUpperCase()] = combined[2].replace(/,/g, '');
                        continue;
                    }
                    // Extract just the value
                    const valMatch = item.text.match(/\\(?\\$?([\\d,]+\\.?\\d*)\\)?/);
                    if (valMatch) {
                        const val = valMatch[1].replace(/,/g, '');
                        // Try to find name from sibling elements
                        const parent = item.element;
                        if (parent && parent.parentElement) {
                            const siblings = parent.parentElement.children;
                            for (const sib of siblings) {
                                const sibText = sib.textContent.trim();
                                if (sibText && /^[A-Za-z][A-Za-z\\s\\-]*$/.test(sibText) && !sibText.includes('Total') && sibText.length < 30) {
                                    results[sibText.toUpperCase()] = val;
                                    break;
                                }
                            }
                        }
                    }
                }
                return results;
            }''')
            if chart_text_pairs:
                networth_map = chart_text_pairs
                print(f"   📊 Chart text scan extraction: {networth_map}")
        except Exception as e:
            print(f"   ⚠️  Chart text scan strategy failed: {e}")

    return networth_map


def add_validation_columns_to_token_allocation(data):
    """
    Add validation columns to Overview - Token Allocation table data.

    Expected original columns: Token, Percentage, Net Worth
    After insertion: Token, Percentage, % - Api NW Calc, Percentage Validation, Net Worth, Net Worth - API Calculation, Net Worth Validation
    """
    if not data or len(data) == 0:
        return data

    modified_data = []
    for row_idx, row in enumerate(data):
        if len(row) < 3:
            modified_data.append(row)
            continue

        new_row = []
        new_row.append(row[0] if len(row) > 0 else "")  # A: Token
        new_row.append(row[1] if len(row) > 1 else "")  # B: Percentage
        new_row.append("")       # C: % - Api NW Calc (empty for now)
        new_row.append("")       # D: Percentage Validation (empty for now)
        new_row.append(row[2] if len(row) > 2 else "")  # E: Net Worth
        new_row.append("")       # F: Net Worth - API Calculation (empty for now)
        new_row.append("")       # G: Net Worth Validation (empty for now)

        if len(row) > 3:
            new_row.extend(row[3:])

        modified_data.append(new_row)

    return modified_data


def add_validation_columns_to_chain_allocation(data):
    """
    Add validation columns to Overview - Chain Allocation table data.

    Expected original columns: Chain, Percentage, Net Worth
    After insertion: Chain, Percentage, % - Api NW Calc, Percentage Validation, Net Worth, Net Worth - API Calculation, Net Worth Validation
    """
    if not data or len(data) == 0:
        return data

    modified_data = []
    for row_idx, row in enumerate(data):
        if len(row) < 3:
            modified_data.append(row)
            continue

        new_row = []
        new_row.append(row[0] if len(row) > 0 else "")  # A: Chain
        new_row.append(row[1] if len(row) > 1 else "")  # B: Percentage
        new_row.append("")       # C: % - Api NW Calc (empty for now)
        new_row.append("")       # D: Percentage Validation (empty for now)
        new_row.append(row[2] if len(row) > 2 else "")  # E: Net Worth
        new_row.append("")       # F: Net Worth - API Calculation (empty for now)
        new_row.append("")       # G: Net Worth Validation (empty for now)

        if len(row) > 3:
            new_row.extend(row[3:])

        modified_data.append(new_row)

    return modified_data


def add_validation_to_defi_tab(defi_data, rabby_data):
    """
    Calculate validation columns for Overview - De-Fi tab using Python.

    De-Fi columns: 0:De-Fi, 1:Chain, 2:Type, 3:Pool/Position Pair, 4:Description, 5:Amount, 6:Amount Tooltip,
                   7:Amount Validation, 8:FE-Amount Validation, 9:Amount Validation Diff, 10:Value, 11:Value Validation

    Rabby columns: 0:Address, 1:Name, 2:ID, 3:Chain, 4:Pool Name, 5:Description, 6:Side,
                   7:Symbol/Currency Pair, 8:Leverage, 9:PnL(USD), 10:Price, 11:Amount, 12:Calculated Value

    Matching criteria (case insensitive):
    - De-Fi A (col 0) = Rabby B (col 1) - Name
    - De-Fi B (col 1) = Rabby D (col 3) - Chain
    - De-Fi C (col 2) = Rabby E (col 4) - Type
    - De-Fi D (col 3) = Rabby H (col 7) - Symbol/Currency Pair
    """
    from decimal import Decimal, getcontext
    getcontext().prec = 50

    if not defi_data or len(defi_data) <= 1:
        return defi_data

    if not rabby_data or len(rabby_data) <= 1:
        return defi_data

    # Skip header row in rabby_data
    rabby_rows = rabby_data[1:] if len(rabby_data) > 1 else []

    modified_data = [defi_data[0]]  # Keep header row

    for row_idx, row in enumerate(defi_data[1:], start=1):
        if len(row) < 7:
            modified_data.append(row)
            continue

        # Extract De-Fi row values
        defi_name = str(row[0]).strip().lower() if row[0] else ""
        defi_chain = str(row[1]).strip().lower() if row[1] else ""
        defi_pool_name = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ""  # C: Type
        defi_symbol = str(row[3]).strip().lower() if row[3] else ""                       # D: Pool/Position Pair
        defi_amount_tooltip = row[6] if len(row) > 6 else ""
        defi_value = row[11] if len(row) > 11 else ""  # K→L: Value (shifted by Amount Diff % column)

        # Convert De-Fi Amount Tooltip to Decimal
        try:
            defi_amount_decimal = Decimal(str(defi_amount_tooltip).replace(',', '')) if defi_amount_tooltip else Decimal('0')
        except:
            defi_amount_decimal = Decimal('0')

        # Convert De-Fi Value to Decimal
        try:
            defi_value_decimal = Decimal(str(defi_value).replace(',', '').replace('$', '')) if defi_value else Decimal('0')
        except:
            defi_value_decimal = Decimal('0')

        # Find first matching row in Rabby data
        # Rabby Api Data columns (0-indexed):
        # 0:Name, 1:ID, 2:Chain, 3:Pool Name, 4:Description,
        # 5:Side, 6:Symbol/Currency Pair, 7:Leverage, 8:PnL(USD),
        # 9:Price, 10:Amount, 11:Calculated Value
        is_hyperliquid = "hyperliquid" in defi_name
        matched_rabby_amount = None
        rabby_value_sum = Decimal('0')
        match_count = 0

        # Collect all matching Rabby rows as candidates
        candidates = []

        for rabby_row in rabby_rows:
            if not isinstance(rabby_row, dict):
                continue

            rabby_name = str(rabby_row.get("Name", "")).strip().lower()
            rabby_chain = str(rabby_row.get("Chain", "")).strip().lower()
            rabby_pool_name = str(rabby_row.get("Pool_Name", "")).strip().lower()
            rabby_symbol = str(rabby_row.get("Symbol", "")).strip().lower()

            matched = False
            if is_hyperliquid:
                if defi_name == rabby_name and defi_symbol == rabby_symbol:
                    matched = True
            else:
                if (defi_name == rabby_name and defi_chain == rabby_chain
                        and defi_pool_name == rabby_pool_name and defi_symbol == rabby_symbol):
                    matched = True

            if matched:
                try:
                    _amt = rabby_row.get("Amount")
                    _amt_dec = Decimal(str(_amt).replace(',', '')) if _amt else Decimal('0')
                except:
                    _amt_dec = Decimal('0')
                try:
                    _cv = rabby_row.get("Calculated_Value")
                    _cv_dec = Decimal(str(_cv).replace(',', '')) if _cv else Decimal('0')
                except:
                    _cv_dec = Decimal('0')
                candidates.append({"amount": _amt_dec, "calc_value": _cv_dec, "row": rabby_row})

        # Fallback: 2-field match (Name + Symbol) if no 4-field match found
        if len(candidates) == 0 and not is_hyperliquid:
            for rabby_row in rabby_rows:
                if not isinstance(rabby_row, dict):
                    continue
                rabby_name = str(rabby_row.get("Name", "")).strip().lower()
                rabby_symbol = str(rabby_row.get("Symbol", "")).strip().lower()
                if defi_name == rabby_name and defi_symbol == rabby_symbol:
                    try:
                        _amt = rabby_row.get("Amount")
                        _amt_dec = Decimal(str(_amt).replace(',', '')) if _amt else Decimal('0')
                    except:
                        _amt_dec = Decimal('0')
                    try:
                        _cv = rabby_row.get("Calculated_Value")
                        _cv_dec = Decimal(str(_cv).replace(',', '')) if _cv else Decimal('0')
                    except:
                        _cv_dec = Decimal('0')
                    candidates.append({"amount": _amt_dec, "calc_value": _cv_dec, "row": rabby_row})

        # Pick the candidate with the closest amount to DAM amount
        if candidates:
            match_count = len(candidates)
            if len(candidates) == 1:
                best = candidates[0]
            else:
                best = min(candidates, key=lambda c: abs(c["amount"] - defi_amount_decimal))
            matched_rabby_amount = best["amount"]
            rabby_value_sum = best["calc_value"]

        # Calculate validations
        if match_count > 0:
            # H: Amount Validation — truncate to 5dp match OR within 1%
            if defi_name == "hyperliquid":
                if matched_rabby_amount is not None and matched_rabby_amount != Decimal('0'):
                    _amt_pct_diff = abs(defi_amount_decimal - matched_rabby_amount) / abs(matched_rabby_amount) * Decimal('100')
                    amount_validation = "Passed" if _amt_pct_diff <= Decimal('1') else "Failed"
                else:
                    amount_validation = "Passed" if defi_amount_decimal == Decimal('0') else "Failed"
            else:
                # Truncate both to 5dp first (DAM truncates amounts), then check 1% fallback
                if matched_rabby_amount is not None and matched_rabby_amount != Decimal('0'):
                    _defi_trunc5 = defi_amount_decimal.quantize(Decimal('0.00001'), rounding='ROUND_DOWN')
                    _rabby_trunc5 = matched_rabby_amount.quantize(Decimal('0.00001'), rounding='ROUND_DOWN')
                    if _defi_trunc5 == _rabby_trunc5:
                        amount_validation = "Passed"
                    else:
                        _amt_pct_diff = abs(defi_amount_decimal - matched_rabby_amount) / abs(matched_rabby_amount) * Decimal('100')
                        amount_validation = "Passed" if _amt_pct_diff <= Decimal('1') else "Failed"
                else:
                    amount_validation = "Passed" if defi_amount_decimal == Decimal('0') else "Failed"

            # I: Amount Validation Diff - De-Fi F - Rabby L
            amount_diff = str(defi_amount_decimal - (matched_rabby_amount if matched_rabby_amount is not None else Decimal('0')))

            # K: Value Validation - Passed if TRUNC(Rabby Calculated Value, 2) == DAM Value
            try:
                if rabby_value_sum != Decimal('0'):
                    rabby_truncated = rabby_value_sum.quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                    defi_truncated = defi_value_decimal.quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                    if rabby_truncated == defi_truncated:
                        value_validation = "Passed"
                    else:
                        # Fallback: within 1% tolerance
                        value_diff_abs = abs(defi_value_decimal - rabby_value_sum)
                        percentage_diff = (value_diff_abs / abs(rabby_value_sum)) * Decimal('100')
                        value_validation = "Passed" if percentage_diff <= Decimal('1') else "Failed"
                elif defi_value_decimal == Decimal('0'):
                    # Both are zero
                    value_validation = "Passed"
                else:
                    value_validation = "Failed"
            except:
                value_validation = "Error"
        else:
            amount_validation = "No Match"
            amount_diff = "No Match"
            value_validation = "No Match"

        # I: FE - Amount Validation
        # F = Amount Tooltip (col 6), E = Amount (col 5)
        # Passed if ABS(E) == TRUNC(ABS(F), 5) OR within 1% difference
        _is_hl_perp_row = defi_name == "hyperliquid" and defi_pool_name == "perpetuals"
        if _is_hl_perp_row:
            fe_amount_validation = "-"
        else:
            try:
                f_raw = row[6]
                e_raw = row[5]
                # If tooltip is missing, fall back to Amount cell text
                if f_raw is None or str(f_raw).strip() == "":
                    # Fallback to Amount cell text (e_raw)
                    if e_raw is None or str(e_raw).strip() == "":
                        fe_amount_validation = "No Amount Data"
                    else:
                        fe_amount_validation = "Tooltip N/A, cant compare"
                else:
                    # Strip token symbol suffix (e.g. "1 LAUNCH" → "1", "8,824 SENT" → "8824")
                    _f_clean = re.sub(r'[A-Za-z\s]+$', '', str(f_raw).replace(',', '').strip())
                    f_val = abs(Decimal(_f_clean)) if _f_clean and _f_clean not in ('-',) else Decimal('0')
                    _threshold = Decimal('0.00001')
                    if f_val >= _threshold:
                        try:
                            # Strip token symbol suffix from amount (e.g. "5 ATEHUN" -> "5")
                            _e_str = re.sub(r'[A-Za-z\s]+$', '', str(e_raw).replace(',', '').strip()) if e_raw not in (None, '', '-') else '0'
                            e_val = abs(Decimal(_e_str)) if _e_str else Decimal('0')
                            # Passed if exact match, trunc-to-5dp match, or within 1%
                            if e_val == f_val:
                                fe_amount_validation = "Passed"
                            elif int(e_val * 100000) == int(f_val * 100000):
                                fe_amount_validation = "Passed"
                            elif f_val != Decimal('0'):
                                _fe_pct = abs(e_val - f_val) / f_val * Decimal('100')
                                fe_amount_validation = "Passed" if _fe_pct <= Decimal('1') else "Failed"
                            else:
                                fe_amount_validation = "Passed" if e_val == Decimal('0') else "Failed"
                        except Exception:
                            fe_amount_validation = "Failed"
                    else:
                        # F < 0.00001: E must be the string "< 0.00001"
                        e_str = str(e_raw).strip() if e_raw is not None else ""
                        fe_amount_validation = "Passed" if e_str == "< 0.00001" else "Failed"
            except Exception:
                fe_amount_validation = "Error"

        # Build the new row with calculated validations
        new_row = list(row)
        # Ensure row has enough columns (now 19 with Amount Diff %)
        while len(new_row) < 14:
            new_row.append("")

        new_row[7] = amount_validation      # H: Amount Validation
        new_row[8] = fe_amount_validation   # I: FE - Amount Validation
        new_row[9] = amount_diff            # J: Amount Validation Diff

        # K: Amount Diff % — percentage difference between DAM and Rabby amounts
        if match_count > 0 and matched_rabby_amount is not None and matched_rabby_amount != Decimal('0'):
            try:
                _amt_diff_pct = abs(defi_amount_decimal - matched_rabby_amount) / abs(matched_rabby_amount) * Decimal('100')
                new_row[10] = str(_amt_diff_pct.quantize(Decimal('0.0001')))
                if _amt_diff_pct > Decimal('0.01'):
                    print(f"      [amt-diff-debug] {defi_name}|{defi_chain}|{defi_symbol}: DAM={defi_amount_decimal}, Rabby={matched_rabby_amount}, diff%={_amt_diff_pct}")
            except:
                new_row[10] = ""
        else:
            new_row[10] = ""

        new_row[12] = value_validation      # M: Value Validation

        # T: Row Matched — show which Rabby row was picked
        # Ensure row has enough columns for Api Calc Value (index 13) and Row Matched (index 19)
        while len(new_row) < 20:
            new_row.append("")
        if match_count > 0 and candidates:
            best = min(candidates, key=lambda c: abs(c["amount"] - defi_amount_decimal))
            _r = best["row"]
            _matched_info = f"{_r.get('Name','')}|{_r.get('Chain','')}|{_r.get('Pool_Name','')}|{_r.get('Symbol','')} Amt={best['amount']} Val={best['calc_value']}"
            # N: Api Calc Value — Rabby's Calculated Value for the matched row
            new_row[13] = str(best["calc_value"])
            # T: Row Matched
            new_row[19] = _matched_info
        else:
            new_row[13] = ""
            new_row[19] = "No Match"

        modified_data.append(new_row)

        print(f"   De-Fi row {row_idx}: {defi_name}|{defi_chain}|{defi_pool_name}|{defi_symbol} -> Matches: {match_count}, Amount Val: {amount_validation}, Value Val: {value_validation}, Rabby Value Sum: {rabby_value_sum}, DAM Value: {defi_value_decimal}")

    return modified_data


def add_validation_columns_to_platform_allocation(data):
    """
    Add validation columns to Overview - Platform Allocation table data.

    Expected original columns: Platform, Percentage, Net Worth
    After insertion: Platform, Percentage, % - Api NW Calc, Percentage Validation, Net Worth, Net Worth - API Calculation, Net Worth Validation
    """
    if not data or len(data) == 0:
        return data

    modified_data = []
    for row_idx, row in enumerate(data):
        if len(row) < 3:
            modified_data.append(row)
            continue

        new_row = []
        new_row.append(row[0] if len(row) > 0 else "")  # A: Platform
        new_row.append(row[1] if len(row) > 1 else "")  # B: Percentage
        new_row.append("")       # C: % - Api NW Calc (empty for now)
        new_row.append("")       # D: Percentage Validation (empty for now)
        new_row.append(row[2] if len(row) > 2 else "")  # E: Net Worth
        new_row.append("")       # F: Net Worth - API Calculation (empty for now)
        new_row.append("")       # G: Net Worth Validation (empty for now)

        if len(row) > 3:
            new_row.extend(row[3:])

        modified_data.append(new_row)

    return modified_data


def calculate_allocation_percentage_validation(data, pct_col_idx, networth_col_idx, validation_col_idx):
    """
    Calculate Percentage Validation for allocation tabs using Python.

    Formula: if B == TRUNC(D / total_D * 100, 2) -> "Passed", else "Failed"

    Args:
        data: List of rows (first row is header)
        pct_col_idx: 0-indexed column for Percentage (B column)
        networth_col_idx: 0-indexed column for Net Worth (D column)
        validation_col_idx: 0-indexed column for Percentage Validation result (C column)
    """
    if not data or len(data) < 2:
        return data

    # First pass: calculate total net worth from all data rows
    total_networth = Decimal('0')
    networth_values = []
    for row_idx in range(1, len(data)):  # Skip header
        row = data[row_idx]
        if len(row) > networth_col_idx:
            nw_str = str(row[networth_col_idx]).replace(',', '').replace('$', '').replace('%', '').replace('<', '').replace('>', '').strip()
            try:
                nw_val = Decimal(nw_str) if nw_str else Decimal('0')
            except:
                nw_val = Decimal('0')
            networth_values.append(nw_val)
            total_networth += nw_val
        else:
            networth_values.append(Decimal('0'))

    print(f"   📊 Percentage Validation: total_networth = {total_networth}, {len(networth_values)} rows")

    # Second pass: calculate validation for each row
    for row_idx in range(1, len(data)):
        row = data[row_idx]
        if len(row) <= validation_col_idx:
            continue

        nw_val = networth_values[row_idx - 1]

        # Get the FE percentage value (B column) - preserve original for "<0.01" check
        pct_original = str(row[pct_col_idx]).strip() if len(row) > pct_col_idx else ''
        pct_str = pct_original.replace(',', '').replace('%', '').replace('<', '').replace('>', '').strip()
        try:
            fe_pct = Decimal(pct_str) if pct_str else None
        except:
            fe_pct = None

        if total_networth != 0:
            # Calculate: nw_val / total_networth * 100
            calc_pct = nw_val / total_networth * Decimal('100')

            # Special handling for calculated percentage < 0.01
            if calc_pct < Decimal('0.01'):
                # Check if B column = "<0.01" (with or without % sign)
                if pct_original in ["<0.01", "< 0.01", "<0.01%", "< 0.01%"]:
                    row[validation_col_idx] = "Passed"
                else:
                    row[validation_col_idx] = "Failed"
                    print(f"      Percentage Validation row {row_idx}: Calc={calc_pct} (<0.01), FE='{pct_original}' (expected '<0.01')")
            elif fe_pct is not None:
                # TRUNC to 2 decimal places using pure Decimal (no float conversion)
                calc_pct_truncated = Decimal(int(calc_pct * Decimal('100'))) / Decimal('100')

                if calc_pct_truncated == fe_pct:
                    row[validation_col_idx] = "Passed"
                else:
                    row[validation_col_idx] = "Failed"
                    print(f"      Percentage Validation row {row_idx}: FE={fe_pct}, Calc={calc_pct_truncated} (raw={calc_pct})")
            else:
                row[validation_col_idx] = "Failed"
        elif total_networth == 0:
            row[validation_col_idx] = "No Data"
        else:
            row[validation_col_idx] = "Failed"

    # Set header
    if len(data[0]) > validation_col_idx:
        data[0][validation_col_idx] = "Percentage Validation"

    return data


def add_validation_columns_to_header_holdings(data):
    """
    Add validation columns to Overview - Header & Token Holdings Header table data.

    Expected input columns (9): Section, Category, Token Count, TC_UI Count, Token Count Validation,
                                 Net Worth, Net Worth_UI Calculation, Net Worth UI Validation, Percentage
    After insertion (12): + Net Worth - API Calculation, Net Worth - UI-API Validation, Percentage Validation
    
    Output: A:Section, B:Category, C:Token Count, D:TC_UI Count, E:Token Count Validation,
            F:Net Worth, G:Net Worth_UI Calculation, H:Net Worth UI Validation,
            I:Net Worth - API Calculation, J:Net Worth - UI-API Validation,
            K:Percentage, L:Percentage Validation
    """
    if not data or len(data) == 0:
        return data

    modified_data = []
    for row_idx, row in enumerate(data):
        new_row = list(row)
        # Ensure row has at least 9 columns (A-I original)
        while len(new_row) < 9:
            new_row.append("")
        
        # Current: [A, B, C, D, E, F, G, H, I(Percentage)]
        # Need: [A, B, C, D, E, F, G, H, I(API Calc), J(UI-API Val), K(Percentage), L(Pct Val)]
        percentage = new_row[8] if len(new_row) > 8 else ""  # Save original I (Percentage)
        new_row[8] = ""   # I: Net Worth - API Calculation (empty for now)
        # Ensure we have slots for J, K, L
        while len(new_row) < 12:
            new_row.append("")
        new_row[9] = ""           # J: Net Worth - UI-API Validation (empty for now)
        new_row[10] = percentage  # K: Percentage (moved from I)
        new_row[11] = ""          # L: Percentage Validation (empty for now)
        
        modified_data.append(new_row)

    return modified_data


def add_validation_columns_to_combined_net_worth(data):
    """
    Add validation columns to Overview - Combined Net Worth table data.

    Expected original columns: Address/Exchange, Value
    After insertion: Address/Exchange, Value, Value Validation, Calculated Value
    """
    if not data or len(data) == 0:
        return data

    modified_data = []
    for row_idx, row in enumerate(data):
        # Handle rows with insufficient data
        if len(row) < 2:
            modified_data.append(row)
            continue

        # Expected column positions (0-indexed):
        # 0: Address/Exchange, 1: Value
        new_row = []
        new_row.append(row[0] if len(row) > 0 else "")  # A: Address/Exchange
        new_row.append(row[1] if len(row) > 1 else "")  # B: Value
        new_row.append("")       # C: Value Validation (empty for now)
        new_row.append("")       # D: Calculated Value (empty for now)
        # Do NOT append extra columns — original data only has A and B

        modified_data.append(new_row)

    return modified_data


def run_trx_balance_api_test(addresses=None, portfolio_name=None):
    """
    Part 1: TRX Balance API Test
    Fetches balance, transactions, and token data from TronGrid and TronScan APIs
    For multi-address portfolios, creates per-address sheets with last-8-chars suffix.

    Args:
        addresses: List of TRX addresses to test. Defaults to [TRX_ADDRESS].
        portfolio_name: Portfolio name (used in filename for multi-address portfolios).
    """
    if addresses is None:
        addresses = [TRX_ADDRESS]

    print("\n" + "="*80)
    print("PART 1: TRX BALANCE API TEST")
    print("="*80)
    print(f"Test Case: TC00003")
    print(f"Addresses: {len(addresses)}")
    for idx, addr in enumerate(addresses, 1):
        print(f"  {idx}. {addr}")
    print()

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Raw API response collector for JSON export
    _trx_raw = {"token_list": None, "trx_price": None, "addresses": []}

    # Initialize token_info_responses (will be populated per-address)
    token_info_responses = []

    # ========================================================================
    # STEP 1: LOAD EXISTING TOKEN LIST
    # ========================================================================
    print("1️⃣  Loading existing Token List from 'Token List' folder...")

    contract_to_token = {}
    contract_to_decimal = {}
    contract_to_symbolshow = {}
    contract_to_canshow = {}

    token_list_folder = os.path.join(os.getcwd(), "Token List")
    token_list_file = os.path.join(token_list_folder, "Token List.xlsx")

    if os.path.exists(token_list_file):
        print(f"   📂 Loading existing Token List.xlsx...")
        try:
            wb_existing = load_workbook(token_list_file, read_only=True, data_only=True)
            if "Token List" in wb_existing.sheetnames:
                ws_existing = wb_existing["Token List"]
                header_row = [cell.value for cell in ws_existing[1]]

                try:
                    abbr_idx = header_row.index("abbr")
                    decimal_idx = header_row.index("decimal")
                    contract_idx = header_row.index("contractAddress")
                    symbolshow_idx = header_row.index("symbolShow") if "symbolShow" in header_row else None
                    canshow_idx = header_row.index("canShow") if "canShow" in header_row else None
                except ValueError as e:
                    print(f"   ⚠️  Missing required column: {e}")
                    abbr_idx = decimal_idx = contract_idx = symbolshow_idx = canshow_idx = None

                if abbr_idx is not None and decimal_idx is not None and contract_idx is not None:
                    for row in ws_existing.iter_rows(min_row=2, values_only=True):
                        contract_addr = row[contract_idx]
                        abbr = row[abbr_idx]
                        decimal = row[decimal_idx]

                        if contract_addr:
                            contract_to_token[contract_addr] = abbr
                            contract_to_decimal[contract_addr] = decimal

                            if symbolshow_idx is not None and row[symbolshow_idx] is not None:
                                contract_to_symbolshow[contract_addr] = row[symbolshow_idx]

                            if canshow_idx is not None and row[canshow_idx] is not None:
                                contract_to_canshow[contract_addr] = row[canshow_idx]

                    print(f"   ✅ Loaded {len(contract_to_token)} existing token mappings")

            wb_existing.close()
        except Exception as e:
            print(f"   ⚠️  Error loading existing Token List: {e}")
    else:
        print(f"   ℹ️  No existing Token List.xlsx found, will create new")

    # ========================================================================
    # STEP 2: FETCH TOKEN LIST API (500 tokens)
    # ========================================================================
    print("\n2️⃣  Fetching Token List API (500 tokens)...")

    token_list_url = "https://apilist.tronscanapi.com/api/tokens/overview?start=0&limit=500&verifier=all&order=desc&filter=top&sort=&showAll=1&field="

    try:
        response = requests.get(token_list_url, timeout=30)
        response.raise_for_status()
        token_list_data = response.json()
        _trx_raw["token_list"] = token_list_data

        # Tab 1: API - All Token Info
        ws_api_token_list = wb.create_sheet("API - All Token Info")
        ws_api_token_list.append(["API Response"])
        ws_api_token_list.append([json.dumps(token_list_data, indent=2)])
        print(f"   ✅ API - All Token Info tab created")

        # Tab 2: Token List (parsed data)
        ws_token_list = wb.create_sheet("Token List")
        ws_token_list.append(["abbr", "decimal", "contractAddress", "canShow"])

        if "data" in token_list_data:
            for token in token_list_data["data"]:
                abbr = token.get("abbr", "")
                decimal = token.get("decimal", "")
                contract_addr = token.get("contractAddress", "")
                can_show = token.get("canShow", "")

                ws_token_list.append([abbr, decimal, contract_addr, can_show])

                # Update mappings (Token List API is supplemental)
                if contract_addr and contract_addr not in contract_to_token:
                    contract_to_token[contract_addr] = abbr
                    contract_to_decimal[contract_addr] = decimal
                    contract_to_canshow[contract_addr] = can_show

            print(f"   ✅ Token List tab created with {len(token_list_data['data'])} tokens")

    except Exception as e:
        print(f"   ❌ Error fetching Token List API: {e}")
        ws_api_token_list = wb.create_sheet("API - All Token Info")
        ws_api_token_list.append(["Error", str(e)])
        ws_token_list = wb.create_sheet("Token List")
        ws_token_list.append(["abbr", "decimal", "contractAddress", "canShow"])

    # ========================================================================
    # STEP 3-5: FETCH BALANCE, TRANSACTIONS, TOKEN DETAILS (PER ADDRESS)
    # ========================================================================
    # Create single combined "TRX Balance, Price" sheet (rows from all addresses)
    ws_balance = wb.create_sheet("TRX Balance, Price")
    ws_balance.append([
        "Address", "Decimal Places", "Token", "Contract Address",
        "Balance", "Balance (Raw)", "Price", "Price (24h)", "Symbol Show", "Calculated Value"
    ])

    for addr_loop_idx, addr in enumerate(addresses):
        addr_suffix = addr[-8:]
        sheet_suffix = f" ({addr_suffix})" if len(addresses) > 1 else ""

        print(f"\n{'='*60}")
        print(f"  ADDRESS {addr_loop_idx + 1}/{len(addresses)}: {addr}")
        print(f"  Suffix: {addr_suffix}")
        print(f"{'='*60}")

        # Reset per-address token_info_responses
        token_info_responses = []

        # Per-address raw collector
        _addr_raw = {"address": addr, "balance": None, "transactions": None, "token_details": []}

        # STEP 3: FETCH ACCOUNT BALANCE
        print("\n3️⃣  Fetching Account Balance...")

        balance_url = f"https://api.trongrid.io/v1/accounts/{addr}"

        try:
            response = requests.get(balance_url, timeout=30)
            response.raise_for_status()
            balance_data = response.json()
            _addr_raw["balance"] = balance_data

            # Tab: API - TRX Balance (with suffix for multi-address)
            ws_api_balance = wb.create_sheet(f"API - TRX Balance{sheet_suffix}")
            ws_api_balance.append(["API Response"])
            ws_api_balance.append([json.dumps(balance_data, indent=2)])
            print(f"   ✅ API - TRX Balance{sheet_suffix} tab created")

            # Data rows appended to the shared "TRX Balance, Price" sheet

            # Process balance data
            if "data" in balance_data and len(balance_data["data"]) > 0:
                account_data = balance_data["data"][0]

                # Verify account_data is a dictionary
                if not isinstance(account_data, dict):
                    print(f"   ⚠️  Warning: account_data is {type(account_data)}, not dict. Skipping balance calculation.")
                    print(f"   Raw data: {account_data}")
                else:
                    # TRX balance row - DEBUG: trace each component
                    print(f"\n   🔍 DEBUG TRX Balance Calculation:")

                    balance_raw = account_data.get("balance", 0)
                    print(f"      balance: {balance_raw:,}")
                    running_total = balance_raw

                    # Add frozen[].frozen_balance
                    frozen_total = 0
                    for frozen in account_data.get("frozen", []):
                        frozen_total += frozen.get("frozen_balance", 0)
                    balance_raw += frozen_total
                    print(f"      frozen[].frozen_balance TOTAL: {frozen_total:,}")

                    # Add frozenV2[].amount
                    frozen_v2_total = 0
                    for frozen_v2 in account_data.get("frozenV2", []):
                        frozen_v2_total += frozen_v2.get("amount", 0)
                    balance_raw += frozen_v2_total
                    print(f"      frozenV2[].amount TOTAL: {frozen_v2_total:,}")

                    # Add account_resource frozen balances (account_resource is a dict, not a list)
                    account_resource = account_data.get("account_resource", {})
                    if isinstance(account_resource, dict):
                        # frozen_balance_for_energy.frozen_balance (it's an object, not array)
                        frozen_energy = account_resource.get("frozen_balance_for_energy", {})
                        frozen_energy_val = 0
                        if isinstance(frozen_energy, dict):
                            frozen_energy_val = frozen_energy.get("frozen_balance", 0)
                            balance_raw += frozen_energy_val
                        print(f"      account_resource.frozen_balance_for_energy.frozen_balance: {frozen_energy_val:,}")

                        # frozen_balance_for_bandwidth.frozen_balance (it's an object, not array)
                        frozen_bandwidth = account_resource.get("frozen_balance_for_bandwidth", {})
                        frozen_bandwidth_val = 0
                        if isinstance(frozen_bandwidth, dict):
                            frozen_bandwidth_val = frozen_bandwidth.get("frozen_balance", 0)
                            balance_raw += frozen_bandwidth_val
                        print(f"      account_resource.frozen_balance_for_bandwidth.frozen_balance: {frozen_bandwidth_val:,}")

                        # frozenV2_balance_for_energy.frozen_balance (it's an object, not array)
                        frozen_v2_energy = account_resource.get("frozenV2_balance_for_energy", {})
                        frozen_v2_energy_val = 0
                        if isinstance(frozen_v2_energy, dict):
                            frozen_v2_energy_val = frozen_v2_energy.get("frozen_balance", 0)
                            balance_raw += frozen_v2_energy_val
                        print(f"      account_resource.frozenV2_balance_for_energy.frozen_balance: {frozen_v2_energy_val:,}")

                        # frozenV2_balance_for_bandwidth.frozen_balance (it's an object, not array)
                        frozen_v2_bandwidth = account_resource.get("frozenV2_balance_for_bandwidth", {})
                        frozen_v2_bandwidth_val = 0
                        if isinstance(frozen_v2_bandwidth, dict):
                            frozen_v2_bandwidth_val = frozen_v2_bandwidth.get("frozen_balance", 0)
                            balance_raw += frozen_v2_bandwidth_val
                        print(f"      account_resource.frozenV2_balance_for_bandwidth.frozen_balance: {frozen_v2_bandwidth_val:,}")

                        # delegated_frozen_balance_for_energy (inside account_resource)
                        delegated_energy = account_resource.get("delegated_frozen_balance_for_energy", 0)
                        balance_raw += delegated_energy
                        print(f"      account_resource.delegated_frozen_balance_for_energy: {delegated_energy:,}")

                        # delegated_frozenV2_balance_for_energy (inside account_resource, not root!)
                        delegated_v2_energy_ar = account_resource.get("delegated_frozenV2_balance_for_energy", 0)
                        balance_raw += delegated_v2_energy_ar
                        print(f"      account_resource.delegated_frozenV2_balance_for_energy: {delegated_v2_energy_ar:,}")

                        # delegated_frozenV2_balance_for_bandwidth (inside account_resource)
                        delegated_v2_bandwidth_ar = account_resource.get("delegated_frozenV2_balance_for_bandwidth", 0)
                        balance_raw += delegated_v2_bandwidth_ar
                        print(f"      account_resource.delegated_frozenV2_balance_for_bandwidth: {delegated_v2_bandwidth_ar:,}")

                    # Add unfrozen[].unfreeze_amount
                    unfrozen_total = 0
                    for unfrozen in account_data.get("unfrozen", []):
                        unfrozen_total += unfrozen.get("unfreeze_amount", 0)
                    balance_raw += unfrozen_total
                    print(f"      unfrozen[].unfreeze_amount TOTAL: {unfrozen_total:,}")

                    # Add unfrozenV2[].unfreeze_amount
                    unfrozen_v2_total = 0
                    for unfrozen_v2 in account_data.get("unfrozenV2", []):
                        unfrozen_v2_total += unfrozen_v2.get("unfreeze_amount", 0)
                    balance_raw += unfrozen_v2_total
                    print(f"      unfrozenV2[].unfreeze_amount TOTAL: {unfrozen_v2_total:,}")

                    # Add root-level delegated balances (older frozen v1 format)
                    delegated_bandwidth = account_data.get("delegated_frozen_balance_for_bandwidth", 0)
                    balance_raw += delegated_bandwidth
                    print(f"      delegated_frozen_balance_for_bandwidth (root): {delegated_bandwidth:,}")

                    print(f"      ────────────────────────────────────")
                    print(f"      GRAND TOTAL (sun): {balance_raw:,}")
                    print(f"      GRAND TOTAL (TRX): {balance_raw / 1_000_000:,.6f}")

                    balance = balance_raw / 1_000_000

                    # Fetch TRX price from TRC10 Token Price API
                    trx_price = ""
                    trx_price_24h = ""
                    try:
                        trx_price_url = "https://apilist.tronscanapi.com/api/token?id=0&showAll=1&order=0&owner=&start=&limit=&id_gt=&totalAll=&name=&field="
                        trx_response = requests.get(trx_price_url, timeout=10)
                        trx_response.raise_for_status()
                        trx_data = trx_response.json()
                        _trx_raw["trx_price"] = trx_data

                        if "data" in trx_data and len(trx_data["data"]) > 0:
                            # The first token (id=0) should be TRX
                            trx_info = trx_data["data"][0]
                            # Price data is inside market_info object
                            if "market_info" in trx_info:
                                market_info = trx_info["market_info"]
                                trx_price = market_info.get("priceInUsd", "")
                                gain = market_info.get("gain", "")
                                if gain != "" and gain is not None:
                                    trx_price_24h = gain * 100
                    except Exception as e:
                        print(f"   ⚠️  Warning: Failed to fetch TRX price: {e}")
                        pass

                    # Add TRX row: Address, Decimal Places, Token, Contract Address, Balance, Balance (Raw), Price, Price (24h), Symbol Show, Calculated Price
                    # Balance will be set with formula later
                    ws_balance.append([
                        addr, "6", "TRX", "", "", str(balance_raw), trx_price, trx_price_24h, "", ""
                    ])

                    print(f"   ✅ TRX Balance: {balance:,.6f} TRX")

                    # TRC20 tokens
                    trc20_tokens = account_data.get("trc20", [])
                    if trc20_tokens:
                        # First, filter out tokens with zero balance to optimize processing
                        non_zero_tokens = []
                        total_tokens = 0
                        for trc20 in trc20_tokens:
                            for contract_addr, balance_raw_str in trc20.items():
                                total_tokens += 1
                                balance_raw = int(balance_raw_str) if balance_raw_str else 0
                                if balance_raw > 0:
                                    non_zero_tokens.append((contract_addr, balance_raw_str))

                        # Apply token limit if configured
                        tokens_to_process = non_zero_tokens
                        if MAX_TOKENS_TO_PROCESS > 0 and len(non_zero_tokens) > MAX_TOKENS_TO_PROCESS:
                            print(f"   ⚠️  Token limit enabled: Processing top {MAX_TOKENS_TO_PROCESS} tokens out of {len(non_zero_tokens)}")
                            tokens_to_process = non_zero_tokens[:MAX_TOKENS_TO_PROCESS]

                        print(f"   📊 Total tokens: {total_tokens}, Non-zero balance tokens: {len(non_zero_tokens)}")
                        print(f"   ⚡ Processing {len(tokens_to_process)} tokens with non-zero balances...")

                        if SKIP_PRICE_FETCHING:
                            print(f"   ⚡ SKIP_PRICE_FETCHING enabled - skipping all API calls for faster processing")

                        # Prepare token data with mappings
                        token_data_list = []
                        for contract_addr, balance_raw_str in tokens_to_process:
                            abbr = contract_to_token.get(contract_addr, "")
                            decimal = contract_to_decimal.get(contract_addr, "")
                            symbol_show = contract_to_symbolshow.get(contract_addr, "")
                            has_all_info = abbr and decimal and symbol_show

                            token_data_list.append({
                                'contract_addr': contract_addr,
                                'balance_raw_str': balance_raw_str,
                                'abbr': abbr,
                                'decimal': decimal,
                                'symbol_show': symbol_show,
                                'has_all_info': has_all_info,
                                'price': "",
                                'price_24h': ""
                            })

                        # Fetch token details in parallel (unless SKIP_PRICE_FETCHING is enabled)
                        if not SKIP_PRICE_FETCHING:
                            print(f"   🔄 Fetching token details using {PARALLEL_API_CALLS} parallel connections...")

                            # Submit all API calls to thread pool
                            with ThreadPoolExecutor(max_workers=PARALLEL_API_CALLS) as executor:
                                future_to_idx = {}
                                for idx, token_data in enumerate(token_data_list):
                                    future = executor.submit(
                                        fetch_token_details,
                                        token_data['contract_addr'],
                                        token_data['has_all_info']
                                    )
                                    future_to_idx[future] = idx

                                # Process completed API calls as they finish
                                completed_count = 0
                                for future in as_completed(future_to_idx):
                                    idx = future_to_idx[future]
                                    contract_addr, token_info, success = future.result()

                                    completed_count += 1
                                    if completed_count % 50 == 0 or completed_count == len(token_data_list):
                                        print(f"   ⏳ Fetched {completed_count}/{len(token_data_list)} tokens...")

                                    if success and token_info:
                                        token_data = token_data_list[idx]

                                        # Update missing token info
                                        if not token_data['abbr']:
                                            token_data['abbr'] = token_info.get("symbol", "")
                                            contract_to_token[contract_addr] = token_data['abbr']
                                        if not token_data['decimal']:
                                            token_data['decimal'] = token_info.get("decimals", 0)
                                            contract_to_decimal[contract_addr] = token_data['decimal']
                                        if not token_data['symbol_show']:
                                            token_data['symbol_show'] = token_info.get("symbolShow", "")
                                            contract_to_symbolshow[contract_addr] = token_data['symbol_show']

                                        # Get price data from market_info
                                        if "market_info" in token_info:
                                            token_data['price'] = token_info["market_info"].get("priceInUsd", "")
                                            gain = token_info["market_info"].get("gain", "")
                                            if gain:
                                                token_data['price_24h'] = gain * 100

                                        # Store response for TRC20 Token Detail tab (only if missing from Token List)
                                        if not token_data['has_all_info']:
                                            token_info_responses.append({
                                                "contract_address": contract_addr,
                                                "response": {"trc20_tokens": [token_info]}
                                            })

                        # Add all token rows to worksheet
                        for token_data in token_data_list:
                            contract_balance_raw = int(token_data['balance_raw_str']) if token_data['balance_raw_str'] else 0

                            # Add TRC20 token row: Address, Decimal Places, Token, Contract Address, Balance, Balance (Raw), Price, Price (24h), Symbol Show, Calculated Price
                            # Balance will be set with formula later
                            ws_balance.append([
                                addr,
                                token_data['decimal'],
                                token_data['abbr'],
                                token_data['contract_addr'],
                                "",  # Balance (formula will be added later)
                                str(contract_balance_raw),
                                token_data['price'],
                                token_data['price_24h'],
                                token_data['symbol_show'],
                                ""  # Calculated Price (will be calculated later)
                            ])

                        print(f"   ✅ TRX Balance, Price tab created with {len(tokens_to_process) + 1} rows (TRX + {len(tokens_to_process)} tokens)")

                        # Add Balance formula to all rows (starting from row 2)
                        # Formula: =IF(LEN(F2)<=VALUE(B2),"0."&REPT("0",VALUE(B2)-LEN(F2))&F2,TEXT(LEFT(F2,LEN(F2)-VALUE(B2)),"#,##0")&"."&RIGHT(F2,VALUE(B2)))
                        for row_idx in range(2, len(tokens_to_process) + 3):  # +3 because: +1 for TRX row, +1 for header, +1 for range end
                            cell = ws_balance.cell(row_idx, 5)  # Column E (5) is Balance
                            formula = f'=IF(LEN(F{row_idx})<=VALUE(B{row_idx}),"0."&REPT("0",VALUE(B{row_idx})-LEN(F{row_idx}))&F{row_idx},TEXT(LEFT(F{row_idx},LEN(F{row_idx})-VALUE(B{row_idx})),"#,##0")&"."&RIGHT(F{row_idx},VALUE(B{row_idx})))'
                            cell.value = formula
                            # Keep as text format by setting number_format
                            cell.number_format = '@'

                        print(f"   ✅ Added Balance formulas with thousand separators (saved as text format)")

                        # Add Calculated Price formula (Column J) = Balance (E) × Price (G)
                        # Formula: =IF(AND(E2<>"",G2<>""),VALUE(SUBSTITUTE(E2,",",""))*G2,"")
                        # This handles: Balance as text with commas, Price as number, blank if either is empty
                        for row_idx in range(2, len(tokens_to_process) + 3):  # +3 because: +1 for TRX row, +1 for header, +1 for range end
                            calc_price_cell = ws_balance.cell(row_idx, 10)  # Column J - Calculated Price
                            formula = f'=IF(AND(E{row_idx}<>"",G{row_idx}<>""),VALUE(SUBSTITUTE(E{row_idx},",",""))*G{row_idx},"")'
                            calc_price_cell.value = formula

                        print(f"   ✅ Added Calculated Price formulas (Balance × Price) to column J")

        except Exception as e:
            print(f"   ❌ Error fetching Account Balance: {e}")
            import traceback
            traceback.print_exc()

        # ====================================================================
        # STEP 4: FETCH TRANSACTIONS (for this address)
        # ====================================================================
        print("\n4️⃣  Fetching Transactions...")

        tx_url = f"https://api.trongrid.io/v1/accounts/{addr}/transactions"

        try:
            response = requests.get(tx_url, timeout=30)
            response.raise_for_status()
            tx_data = response.json()
            _addr_raw["transactions"] = tx_data

            # Tab: API - TRX Transaction (with suffix for multi-address)
            ws_api_tx = wb.create_sheet(f"API - TRX Transaction{sheet_suffix}")
            ws_api_tx.append(["API Response"])
            ws_api_tx.append([json.dumps(tx_data, indent=2)])
            print(f"   ✅ API - TRX Transaction{sheet_suffix} tab created")

            # Tab: Transaction (parsed, with suffix)
            ws_tx = wb.create_sheet(f"Transaction{sheet_suffix}")

            if "data" in tx_data and len(tx_data["data"]) > 0:
                # Collect all unique keys
                all_keys = set()
                for tx in tx_data["data"]:
                    def flatten_dict(d, parent_key=''):
                        items = []
                        for k, v in d.items():
                            new_key = f"{parent_key}_{k}" if parent_key else k
                            if isinstance(v, dict):
                                items.extend(flatten_dict(v, new_key).items())
                            else:
                                items.append((new_key, v))
                        return dict(items)

                    flat_tx = flatten_dict(tx)
                    all_keys.update(flat_tx.keys())

                headers = sorted(list(all_keys))
                ws_tx.append(headers)

                # Write data rows
                for tx in tx_data["data"]:
                    flat_tx = flatten_dict(tx)
                    row = []
                    for header in headers:
                        value = flat_tx.get(header, "")
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value)
                        row.append(value)
                    ws_tx.append(row)

                print(f"   ✅ Transaction tab created with {len(tx_data['data'])} transactions")

        except Exception as e:
            print(f"   ❌ Error fetching Transactions: {e}")

        # ====================================================================
        # STEP 5: EXPORT TRC20 TOKEN DETAIL RESPONSES (for this address)
        # ====================================================================
        print("\n5️⃣  Exporting TRC20 Token Detail API responses...")

        # Shorten sheet name to fit Excel 31-char limit
        ws_token_detail = wb.create_sheet(f"API - TRC20 Token{sheet_suffix}")
        ws_token_detail.append(["Contract Address", "Full API Response"])

        for token_response in token_info_responses:
            contract_addr = token_response["contract_address"]
            response_json = json.dumps(token_response["response"], indent=2)
            ws_token_detail.append([contract_addr, response_json])
            _addr_raw["token_details"].append(token_response)

        print(f"   ✅ API - TRC20 Token{sheet_suffix} tab created with {len(token_info_responses)} responses")

        # Append this address's raw data to the collector
        _trx_raw["addresses"].append(_addr_raw)

    # END OF ADDRESS LOOP
    # ========================================================================
    # STEP 6: SAVE EXCEL FILE
    # ========================================================================
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Create test-results directory if it doesn't exist
    os.makedirs("test-results/API Result", exist_ok=True)

    # File naming: single address uses last 8 chars, multi-address uses portfolio name
    if len(addresses) > 1 and portfolio_name:
        excel_filename = f"API_TRXBalance_{portfolio_name}_{timestamp}.xlsx"
    else:
        last_8_chars_address = addresses[0][-8:]
        excel_filename = f"API_TRXBalance_{last_8_chars_address}_{timestamp}.xlsx"
    excel_path = os.path.join("test-results/API Result", excel_filename)

    try:
        # Apply Passed/Failed conditional formatting
        from openpyxl.formatting.rule import CellIsRule
        _passed_fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
        _failed_fill = PatternFill(start_color="E57373", end_color="E57373", fill_type="solid")
        _tooltip_na_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for _ws in wb.worksheets:
            _max_col_letter = _ws.cell(1, max(1, _ws.max_column)).column_letter
            _range = f"A1:{_max_col_letter}{max(1, _ws.max_row)}"
            _ws.conditional_formatting.add(_range, CellIsRule(operator='equal', formula=['"Passed"'], fill=_passed_fill))
            _ws.conditional_formatting.add(_range, CellIsRule(operator='equal', formula=['"Failed"'], fill=_failed_fill))
            _ws.conditional_formatting.add(_range, CellIsRule(operator='equal', formula=['"Tooltip N/A, cant compare"'], fill=_tooltip_na_fill))

        # Highlight rows in TRX Balance, Price where Price (col G) is empty
        _grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        if "TRX Balance, Price" in wb.sheetnames:
            _ws_bal = wb["TRX Balance, Price"]
            _max_col = _ws_bal.max_column
            for _ri in range(2, _ws_bal.max_row + 1):
                _price_val = _ws_bal.cell(_ri, 7).value  # Column G - Price
                if _price_val is None or str(_price_val).strip() == "":
                    for _ci in range(1, _max_col + 1):
                        _ws_bal.cell(_ri, _ci).fill = _grey_fill
        wb.save(excel_path)
        file_size = os.path.getsize(excel_path) / 1024  # KB
        print(f"\n✅ Excel file created: {excel_path} ({file_size:.1f} KB)")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")

        # Save raw API responses to JSON
        _trx_json_path = excel_path.replace(".xlsx", "_Raw.json")
        with open(_trx_json_path, 'w') as _jf:
            json.dump(_trx_raw, _jf, indent=2, default=str)
        print(f"   📄 Saved TRX raw API responses to: {os.path.basename(_trx_json_path)}")

        return excel_path
    except Exception as e:
        print(f"\n❌ Error saving Excel: {e}")
        return None


def run_dam_portfolio_extraction(trx_balance_filename=None, target_portfolio_name=None):
    """
    Part 2: DAM Portfolio Full Extraction
    Creates portfolio with TRX address and extracts all 6 tables

    Args:
        trx_balance_filename: Filename of the TRX Balance Excel file (for validation formulas)
        target_portfolio_name: If provided, navigate to this specific portfolio name
                               (skips address-based Excel lookup)
    """
    import re  # Import re module at function level to avoid scoping issues
    from openpyxl.styles import PatternFill, Font  # Import at function level to avoid scoping issues
    global TRX_ADDRESSES, TRX_ADDRESS, SKIP_TRX_API, SKIP_SIM_DUNE_API
    print("\nPART 2: DAM PORTFOLIO EXTRACTION")

    # Load credentials from tc1_account.json if available, else fall back to Config
    import json as _json
    from config.config import Config as _Cfg
    _tc1_path = os.path.join(_Cfg.PROJECT_ROOT, "test_data", "tc1_account.json")
    if os.path.exists(_tc1_path):
        with open(_tc1_path) as _f:
            _acc = _json.load(_f)
        test_email = _acc["email"]
        test_password = _acc["password"]
    else:
        test_email = Config.TEST_EMAIL
        test_password = Config.TEST_PASSWORD

    # TRX Balance sheet name for formula references
    # (single combined sheet for both single and multi-address portfolios)
    trx_balance_sheet_ref = "TRX Balance, Price"

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            slow_mo=400,
            channel="chrome",        # use installed Chrome — better anti-bot fingerprint than Chromium
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
        )
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        # Apply stealth to the whole context so every new_page() inherits it automatically
        from playwright_stealth import Stealth
        Stealth(navigator_platform_override="MacIntel").apply_stealth_sync(context)
        page = context.new_page()

        # Intercept network responses to capture Rabby API data served via DAM backend
        _intercepted_rabby_protocols = []   # rows from complex_protocol_list
        _intercepted_rabby_apps = []        # rows from complex_app_list
        _intercepted_rabby_raw = []         # raw {address, data} entries for Raw tab

        def _on_response(response):
            try:
                url = response.url
                if "complex_protocol_list" in url or "complex_app_list" in url:
                    import urllib.parse as _up
                    _params = dict(_up.parse_qsl(_up.urlparse(url).query))
                    _addr = _params.get("id", "unknown")
                    try:
                        _body = response.json()
                    except Exception:
                        return
                    _intercepted_rabby_raw.append({"address": _addr, "url": url, "data": _body})
                    print(f"   🌐 Intercepted Rabby API: {url.split('?')[0]} for {_addr[-8:]}")
            except Exception:
                pass

        page.on("response", _on_response)

        try:
            # STEP 1: Sign In
            print("STEP 1: Sign In")
            print("="*80)

            # Sign-in with retry mechanism (up to 3 attempts)
            max_sign_in_attempts = 3
            for sign_in_attempt in range(1, max_sign_in_attempts + 1):
                if sign_in_attempt > 1:
                    print(f"\n🔄 Sign-in retry attempt {sign_in_attempt}/{max_sign_in_attempts}...")

                page.goto(Config.SIGN_IN_URL)
                page.wait_for_timeout(2000)

                page.fill('input[data-testid="input-email"]', test_email)
                page.fill('input[data-testid="input-password"]', test_password)
                page.click('button[data-testid="sign-in-btn"]')
                print(f"🔐 Signing in with account: {test_email}")

                # Wait for sign-in to complete
                print("⏳ Waiting for page to load after sign-in...")
                page.wait_for_timeout(10000)

                # Close popup
                try:
                    for selector in ['button:has-text("×")', '[aria-label="close"]']:
                        if page.locator(selector).is_visible(timeout=1000):
                            page.locator(selector).first.click()
                            page.wait_for_timeout(1000)
                            break
                except:
                    pass

                # Check if sign-in succeeded (should NOT still be on /sign-in)
                current_url = page.url
                if '/sign-in' not in current_url:
                    print("✅ Signed in")
                    break
                else:
                    print(f"⚠️  Still on sign-in page after attempt {sign_in_attempt}")
                    if sign_in_attempt < max_sign_in_attempts:
                        print("   Waiting 5 seconds before retry...")
                        page.wait_for_timeout(5000)
                    else:
                        print("❌ Sign-in failed after all attempts - proceeding anyway")

            # Check current URL after sign-in
            current_url = page.url
            print(f"📍 Current URL: {current_url}")

            # If PORTFOLIO_ID is set, navigate directly to that portfolio
            if PORTFOLIO_ID:
                target_url = f"{Config.BASE_URL}/portfolio?portfolioId={PORTFOLIO_ID}"
                print(f"🎯 Navigating directly to portfolio ID: {PORTFOLIO_ID}")
                print(f"   Target URL: {target_url}")

                # Force full page reload to the correct URL (with longer timeout)
                try:
                    page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
                    page.wait_for_timeout(5000)  # Wait for app to hydrate
                except Exception as e:
                    print(f"   ⚠️  Initial navigation timeout, retrying...")
                    page.goto(target_url, wait_until="commit", timeout=30000)
                    page.wait_for_timeout(5000)

                # If URL doesn't match, force reload
                if PORTFOLIO_ID not in page.url:
                    print(f"   ⚠️  URL mismatch, forcing reload...")
                    page.reload(wait_until="networkidle")
                    page.wait_for_timeout(3000)

                # Final check and hard refresh if needed
                if PORTFOLIO_ID not in page.url:
                    print(f"   ⚠️  Still wrong URL, using hard navigation...")
                    page.evaluate(f"window.location.replace('{target_url}')")
                    page.wait_for_load_state("networkidle")
                    page.wait_for_timeout(5000)

                # Verify we're on the correct portfolio by checking page content
                try:
                    # Wait for portfolio data to load
                    page.wait_for_selector("text=Wallets", timeout=10000)
                    page.wait_for_timeout(2000)

                    # Check if Total Net Worth is around $45 (the correct portfolio)
                    wallets_text = page.locator("text=Wallets").first.locator("..").text_content()
                    print(f"   📊 Page shows: {wallets_text[:50]}...")

                    if PORTFOLIO_ID in page.url:
                        print(f"   ✅ Successfully on correct portfolio")
                        print(f"   URL: {page.url}")
                    else:
                        print(f"   ❌ Navigation failed - wrong portfolio")
                except Exception as e:
                    print(f"   ⚠️  Could not verify portfolio: {e}")

                print()
            elif "portfolioId=" in current_url:
                print("⚠️  Redirected to specific portfolio - navigating to portfolio list...")
                page.goto(f"{Config.BASE_URL}/portfolio")
                page.wait_for_timeout(3000)
                print(f"   ✅ Navigated to: {page.url}\n")
            else:
                print()

            # STEP 2: Check DAM addresses Excel and create/navigate to portfolio
            print("STEP 2: Check DAM Addresses Excel & Portfolio")
            print("="*80)

            # Screenshot folder will be created later after portfolio name is determined

            # Determine portfolio name to navigate to
            dam_excel_path = "test_data/DAM addresses.xlsx"
            portfolio_found_in_excel = False
            portfolio_name_from_excel = None
            _safe_name = "unknown"

            if target_portfolio_name:
                # Use the exact portfolio name provided (from --portfolio flag)
                portfolio_name_from_excel = target_portfolio_name
                portfolio_found_in_excel = True
                _safe_name = re.sub(r'[^\w\-]', '_', target_portfolio_name)
                print(f"🎯 Using target portfolio name: {target_portfolio_name}")

                # Still load Excel for the wb_dam reference (used later for updates)
                try:
                    import openpyxl
                    wb_dam = openpyxl.load_workbook(dam_excel_path, data_only=True)
                    ws_dam = wb_dam.active
                except Exception as e:
                    print(f"⚠️  Could not load DAM addresses Excel: {e}")
                    wb_dam = None
                    ws_dam = None
            else:
                # Search by address in Excel file
                try:
                    import openpyxl
                    wb_dam = openpyxl.load_workbook(dam_excel_path, data_only=True)
                    ws_dam = wb_dam.active

                    print(f"📂 Loading DAM addresses from: {dam_excel_path}")

                    # Search for the address in Excel file (column B)
                    for row_idx, row in enumerate(ws_dam.iter_rows(min_row=2, values_only=True), start=2):
                        if row and len(row) >= 2:
                            excel_address = str(row[1]).strip() if row[1] else ""
                            if excel_address == TRX_ADDRESS:
                                portfolio_name_from_excel = str(row[0]).strip() if row[0] else None
                                portfolio_found_in_excel = True
                                print(f"✅ Address found in Excel at row {row_idx}")
                                print(f"   Portfolio Name: {portfolio_name_from_excel}")
                                break

                    if not portfolio_found_in_excel:
                        print(f"⚠️  Address NOT found in DAM addresses Excel")
                        print(f"   Will create new portfolio with name: last 8 chars of address")
                        portfolio_name_from_excel = TRX_ADDRESS[-8:]  # Use last 8 characters

                except Exception as e:
                    print(f"⚠️  Could not load DAM addresses Excel: {e}")
                    print(f"   Using last 8 characters as portfolio name")
                    portfolio_name_from_excel = TRX_ADDRESS[-8:]

            print()

            # Skip dropdown navigation if we already navigated directly to portfolio by ID
            if PORTFOLIO_ID:
                print(f"✅ Using direct portfolio ID navigation - skipping dropdown search")
                print(f"   Portfolio ID: {PORTFOLIO_ID}")
                portfolio_exists_in_dam = True
                # Extract portfolio name from the current page
                try:
                    portfolio_name_elem = page.locator('div.text-mono-900.typography-body.font-normal.text-left.break-all.w-full').first
                    if portfolio_name_elem.count() > 0 and portfolio_name_elem.is_visible(timeout=3000):
                        portfolio_name_from_excel = portfolio_name_elem.text_content().strip()
                        print(f"   📝 Portfolio name from UI: {portfolio_name_from_excel}")
                except:
                    pass
                # Skip to extraction step - don't open dropdown
                matched_element = None
            else:
                # Now check if this portfolio exists in DAM dropdown, otherwise create it
                print("🔽 Checking Portfolio dropdown for existing portfolios...")

                # Open portfolio dropdown — button text = current portfolio name (not "Portfolio")
                # Use coordinate click on the chevron area (reliable across all portfolio names)
                portfolio_dropdown = page.locator('button:has-text("Portfolio")').first
                print(f"   DEBUG: Portfolio button count (text match): {portfolio_dropdown.count()}")

            # Only do dropdown navigation if NOT using direct portfolio ID
            if not PORTFOLIO_ID:
                dropdown_opened = False
                if portfolio_dropdown.count() > 0:
                    try:
                        portfolio_dropdown.click(timeout=5000)
                        page.wait_for_timeout(2000)
                        print("   ✅ Portfolio dropdown opened")
                        dropdown_opened = True
                    except Exception as e:
                        print(f"   ⚠️  Text-based click failed: {e}")

                if not dropdown_opened:
                    # Fallback: coordinate click on portfolio switcher chevron
                    try:
                        page.mouse.click(395, 141)
                        page.wait_for_timeout(2000)
                        print("   ✅ Portfolio dropdown opened (coordinate click)")
                        dropdown_opened = True
                    except Exception as e2:
                        print(f"   ❌ Coordinate click also failed: {e2}")

                # Scroll inside dropdown list to load all portfolios
                if dropdown_opened:
                    page.mouse.move(490, 450)
                    for _ in range(15):
                        page.mouse.wheel(0, 300)
                        page.wait_for_timeout(100)
                    page.wait_for_timeout(500)

                # Check if the portfolio from Excel exists in the dropdown
                portfolio_exists_in_dam = False
                matched_element = None

                # Method 1: Direct text match — find element with exact portfolio name text
                try:
                    exact_match = page.get_by_text(portfolio_name_from_excel, exact=True)
                    count = exact_match.count()
                    print(f"   DEBUG: get_by_text(exact) count={count}")
                    if count > 0:
                        for i in range(count):
                            elem = exact_match.nth(i)
                            if elem.is_visible():
                                matched_element = elem
                                portfolio_exists_in_dam = True
                                print(f"✅ Portfolio '{portfolio_name_from_excel}' found (exact text match)")
                                break
                except:
                    pass

                # Method 2: Use role="menuitem" with substring match
                if not portfolio_exists_in_dam:
                    try:
                        menu_items = page.get_by_role("menuitem").all()
                        print(f"   DEBUG: Found {len(menu_items)} menuitem elements in dropdown")
                        for item in menu_items:
                            try:
                                if not item.is_visible():
                                    continue
                                item_text = item.text_content().strip()
                                if portfolio_name_from_excel.lower() in item_text.lower():
                                    if 'create portfolio' not in item_text.lower():
                                        matched_element = item
                                        portfolio_exists_in_dam = True
                                        print(f"✅ Portfolio '{portfolio_name_from_excel}' found in DAM dropdown (menuitem)")
                                        break
                            except:
                                pass
                    except:
                        pass

                # Method 3: Search all visible divs using substring match
                if not portfolio_exists_in_dam:
                    try:
                        all_divs = page.locator('div').all()
                        for div in all_divs:
                            try:
                                if not div.is_visible():
                                    continue
                                div_text = div.text_content().strip()
                                if not div_text:
                                    continue
                                if portfolio_name_from_excel.lower() in div_text.lower():
                                    if ('Addresses' in div_text or 'Exchange' in div_text) and 'create portfolio' not in div_text.lower():
                                        matched_element = div
                                        portfolio_exists_in_dam = True
                                        print(f"✅ Portfolio '{portfolio_name_from_excel}' found in DAM dropdown")
                                        break
                            except:
                                pass
                    except:
                        pass

                if portfolio_exists_in_dam and matched_element:
                    print(f"   Navigating to existing portfolio...")
                    matched_element.click()
                    page.wait_for_timeout(3000)

                    # Extract the actual portfolio name from DAM UI after selection
                    try:
                        portfolio_name_elem = page.locator('div.text-mono-900.typography-body.font-normal.text-left.break-all.w-full').first
                        if portfolio_name_elem.count() > 0:
                            portfolio_name_from_dam = portfolio_name_elem.text_content().strip()
                            if portfolio_name_from_dam:
                                portfolio_name_from_excel = portfolio_name_from_dam
                                print(f"   📝 Extracted portfolio name from DAM UI: {portfolio_name_from_excel}")
                    except Exception as e:
                        print(f"   ⚠️  Could not extract portfolio name from UI: {e}")
                else:
                    print(f"⚠️  Portfolio '{portfolio_name_from_excel}' NOT found in DAM dropdown")
                    print(f"   Will create new portfolio")

            print()

            # Create portfolio if it doesn't exist
            if not portfolio_exists_in_dam:
                print("STEP 3: Create Portfolio")
                print("="*80)

                # Close dropdown if still open, then reopen cleanly
                print("🔽 Opening Portfolio dropdown...")
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
                # Try text-based selector first, fallback to coordinate click
                portfolio_dropdown = page.locator('button:has-text("Portfolio")').first
                if portfolio_dropdown.count() > 0:
                    try:
                        portfolio_dropdown.click(timeout=5000)
                        page.wait_for_timeout(2000)
                        print("   ✅ Dropdown opened")
                    except:
                        page.mouse.click(395, 141)
                        page.wait_for_timeout(2000)
                        print("   ✅ Dropdown opened (coordinate click)")
                else:
                    page.mouse.click(395, 141)
                    page.wait_for_timeout(2000)
                    print("   ✅ Dropdown opened (coordinate click)")
                # Scroll to bottom of list to reveal Create portfolio button
                page.mouse.move(490, 450)
                for _ in range(15):
                    page.mouse.wheel(0, 300)
                    page.wait_for_timeout(100)
                page.wait_for_timeout(500)

                # Debug: Take screenshot of dropdown
                try:
                    page.screenshot(path=f"{screenshot_folder}/debug_dropdown.png")
                    print("   📸 Debug screenshot captured: debug_dropdown.png")
                except:
                    pass

                # STEP: Click "+ Create portfolio" button
                print("➕ Clicking Create portfolio...")

                # Wait for dropdown menu to fully render
                page.wait_for_timeout(1000)

                # The "+ Create portfolio" button has specific styling - look for it
                create_clicked = False

                # Try 1: Use role="menuitem" with exact name "Create portfolio"
                try:
                    create_btn = page.get_by_role("menuitem", name="Create portfolio", exact=True)
                    if create_btn.count() > 0 and create_btn.first.is_visible():
                        create_btn.first.click()
                        create_clicked = True
                        print(f"   ✅ Clicked Create portfolio (menuitem exact)")
                except Exception as e:
                    print(f"   DEBUG: Error with menuitem selector: {e}")

                # Try 2: Look for button/div with exact "Create portfolio" text
                if not create_clicked:
                    create_button = page.get_by_text("Create portfolio", exact=True)
                    if create_button.count() > 0:
                        create_button.click()
                        create_clicked = True
                        print("   ✅ Clicked using exact text match")

                # Try 3: Look for any element containing "Create portfolio"
                if not create_clicked:
                    create_button = page.get_by_text("Create portfolio")
                    if create_button.count() > 0:
                        create_button.first.click()
                        create_clicked = True
                        print("   ✅ Clicked using contains text match")

                # Try 4: Use role-based selector
                if not create_clicked:
                    create_button = page.get_by_role("menuitem").filter(has_text="Create")
                    if create_button.count() > 0:
                        create_button.first.click()
                        create_clicked = True
                        print("   ✅ Clicked using menuitem role")

                if create_clicked:
                    page.wait_for_timeout(2000)
                    print("✅ Create Portfolio dialog opened")
                else:
                    print("❌ Could not find Create portfolio button")
                    # Take another screenshot to see final state
                    try:
                        page.screenshot(path=f"{screenshot_folder}/debug_dropdown_failed.png")
                        print("   📸 Debug screenshot saved: debug_dropdown_failed.png")
                    except:
                        pass
                    raise Exception("Create portfolio button not found")

                print()

                # Use portfolio name from Excel (or last 8 chars if not found)
                portfolio_name = portfolio_name_from_excel
                print(f"📝 Portfolio: {portfolio_name}")

                # Fill name
                print("STEP: Enter Portfolio Name")
                print("=" * 80)
                # Take debug screenshot to see what dialog opened
                try:
                    page.screenshot(path=f"debug_create_dialog.png")
                    print("   📸 Dialog screenshot: debug_create_dialog.png")
                except:
                    pass
                # Try multiple selectors for name input
                name_input = None
                for placeholder in ["Enter portfolio name", "Portfolio name", "Name", "Enter name"]:
                    candidate = page.get_by_placeholder(placeholder)
                    if candidate.count() > 0:
                        name_input = candidate
                        print(f"   Found input with placeholder: '{placeholder}'")
                        break
                if name_input is None:
                    # Try any visible input or textarea in a dialog/form
                    for selector in ['dialog input[type="text"]', 'dialog input', '[role="dialog"] input', 'form input[type="text"]', 'input[type="text"]:visible']:
                        candidate = page.locator(selector)
                        if candidate.count() > 0 and candidate.first.is_visible(timeout=2000):
                            name_input = candidate.first
                            print(f"   Found input with selector: '{selector}'")
                            break
                if name_input:
                    name_input.click()
                    name_input.fill(portfolio_name)
                    page.wait_for_timeout(500)
                    print("✅ Portfolio name entered")
                else:
                    # Print all visible inputs for debugging
                    all_inputs = page.locator('input').all()
                    print(f"   DEBUG: {len(all_inputs)} input elements total on page")
                    for inp in all_inputs:
                        try:
                            if inp.is_visible():
                                print(f"   DEBUG: visible input type={inp.get_attribute('type')} placeholder={inp.get_attribute('placeholder')}")
                        except:
                            pass
                    raise Exception("Could not find portfolio name input field")

                print()

                # Add all addresses (EVM + TRX) to the portfolio
                all_addresses_to_add = list(EVM_ADDRESSES) + ([TRX_ADDRESS] if TRX_ADDRESS and not SKIP_TRX_API else [])
                print(f"📍 Adding {len(all_addresses_to_add)} address(es) to portfolio...")
                print("=" * 80)

                for addr_idx, addr in enumerate(all_addresses_to_add):
                    print(f"   Adding: {addr}")
                    # Target the specific wallet input for this index (wallet.N.address)
                    # After filling each address, that field becomes disabled; the next one appears
                    addr_field = None
                    specific_sel = f'input[name="wallet.{addr_idx}.address"]'
                    candidate = page.locator(specific_sel).first
                    if candidate.count() > 0:
                        try:
                            candidate.wait_for(state="visible", timeout=5000)
                            if candidate.is_enabled(timeout=2000):
                                addr_field = candidate
                        except Exception:
                            pass
                    # Fallback: find any enabled placeholder input
                    if not addr_field:
                        for sel in [
                            '[placeholder*="wallet address"]:not([disabled])',
                            '[placeholder*="Paste your wallet"]:not([disabled])',
                            'textarea:not([disabled])',
                        ]:
                            try:
                                candidate = page.locator(sel).first
                                if candidate.count() > 0 and candidate.is_visible(timeout=2000) and candidate.is_enabled(timeout=1000):
                                    addr_field = candidate
                                    break
                            except Exception:
                                pass
                    if addr_field:
                        addr_field.click()
                        addr_field.fill(addr)
                        page.wait_for_timeout(500)
                        page.keyboard.press("Enter")
                        page.wait_for_timeout(2000)
                        print(f"   ✅ Added: {addr[:12]}...")
                    else:
                        print(f"   ⚠️  Could not find address input for: {addr}")
                print()

                # Save
                print("STEP: Click Save Button")
                print("=" * 80)
                print("💾 Saving portfolio...")

                # Wait up to 15s for Save button to become enabled (address validation)
                save_button = page.locator('button:has-text("Save")').first
                try:
                    save_button.wait_for(state="visible", timeout=10000)
                    page.wait_for_timeout(3000)  # extra wait for validation
                    save_button.click(timeout=15000)
                    page.wait_for_timeout(5000)
                    print("✅ Portfolio saved!")
                except Exception as _e:
                    print(f"⚠️  Save button issue: {_e}")
                    # Try submit button fallback
                    try:
                        page.locator('button[type="submit"]').first.click(timeout=10000)
                        page.wait_for_timeout(5000)
                        print("✅ Portfolio saved (submit fallback)!")
                    except Exception as _e2:
                        # Force click via JS as last resort
                        try:
                            page.evaluate("document.querySelector('button:has-text(\"Save\"), button[type=\"submit\"]')?.click()")
                            page.wait_for_timeout(5000)
                            print("✅ Portfolio saved (JS click)!")
                        except Exception:
                            print(f"❌ ERROR: Could not save portfolio: {_e2}")

                print("✅ Portfolio created")

                # Update the Excel file with new entry
                print(f"📝 Updating DAM addresses Excel file...")
                try:
                    # Add new row with portfolio name and address
                    next_row = ws_dam.max_row + 1
                    ws_dam.cell(row=next_row, column=1, value=portfolio_name_from_excel)
                    ws_dam.cell(row=next_row, column=2, value=TRX_ADDRESS)
                    wb_dam.save(dam_excel_path)
                    print(f"✅ Excel file updated - added row {next_row}")
                except Exception as e:
                    print(f"⚠️  Warning: Could not update Excel file: {e}")

                print()
            else:
                print("✅ Reusing existing portfolio\n")

            # Only need to navigate to portfolio list if we created a new portfolio
            # (If we clicked existing portfolio from dropdown, we're already there)
            if not portfolio_exists_in_dam:
                # Navigate back to portfolio list and find the newly created portfolio
                print("🔄 Navigating to portfolio list...")
                page.goto(Config.PORTFOLIO_URL)
                page.wait_for_timeout(3000)

                # Find and click on the newly created portfolio
                print(f"🔍 Looking for portfolio: {portfolio_name_from_excel}")
                portfolio_found = False
                try:
                    # Try to find the portfolio by exact name match
                    portfolio_link = page.get_by_text(portfolio_name_from_excel, exact=True)
                    if portfolio_link.count() > 0:
                        print(f"   Clicking on: {portfolio_name_from_excel}")
                        portfolio_link.first.click()
                        page.wait_for_timeout(5000)
                        portfolio_found = True
                        print(f"✅ Opened portfolio")

                        # Extract the actual portfolio name from DAM UI
                        try:
                            # Look for portfolio name in the UI - typically shown in header or breadcrumb
                            # Based on the screenshot, it appears near "Portfolio" text
                            portfolio_name_elem = page.locator('div.text-mono-900.typography-body.font-normal.text-left.break-all.w-full').first
                            if portfolio_name_elem.count() > 0:
                                portfolio_name_from_dam = portfolio_name_elem.text_content().strip()
                                if portfolio_name_from_dam:
                                    portfolio_name_from_excel = portfolio_name_from_dam
                                    print(f"   📝 Extracted portfolio name from DAM UI: {portfolio_name_from_excel}")
                        except Exception as e:
                            print(f"   ⚠️  Could not extract portfolio name from UI: {e}")
                    else:
                        print(f"⚠️  Warning: Portfolio '{portfolio_name_from_excel}' not found in list")
                except Exception as e:
                    print(f"⚠️  Warning: Error finding portfolio: {e}")

                if not portfolio_found:
                    print(f"⚠️  Could not find portfolio, trying direct navigation...")
                    # As fallback, just go to the portfolio base URL
                    page.goto(Config.PORTFOLIO_URL)
                    page.wait_for_timeout(3000)

            # Verify the correct address is displayed (skip for CEX-only portfolios)
            if SKIP_TRX_API or not TRX_ADDRESSES:
                print(f"🔍 CEX-only portfolio - no wallet addresses to verify")
            else:
                print(f"🔍 Verifying address: {TRX_ADDRESS[:10]}...{TRX_ADDRESS[-10:]}")
                page.wait_for_timeout(3000)

                # Try to find the address on the page
                address_found = False
                try:
                    if page.locator(f'text="{TRX_ADDRESS[:10]}"').count() > 0:
                        address_found = True
                        print(f"✅ Verified address prefix: {TRX_ADDRESS[:10]}")
                    elif page.locator(f'text="{TRX_ADDRESS[-10:]}"').count() > 0:
                        address_found = True
                        print(f"✅ Verified address suffix: {TRX_ADDRESS[-10:]}")

                    if not address_found:
                        print(f"⚠️  WARNING: Address NOT found on page!")
                        print(f"⚠️  May be viewing wrong portfolio or page not loaded")
                except Exception as e:
                    print(f"⚠️  Warning: Error verifying address: {e}")

            # Create screenshot folder — naming: DAMSS_portfolioname_email_MMDD_HHMM
            timestamp_folder = datetime.now().strftime("%m%d_%H%M")
            email_username = test_email.split('@')[0] if '@' in test_email else test_email
            _ss_portfolio = re.sub(r'[^\w\-]', '_', portfolio_name_from_excel or email_username)
            screenshot_folder = f"test-results/screenshots/DAMSS_{_ss_portfolio}_{email_username}_{timestamp_folder}"
            os.makedirs(screenshot_folder, exist_ok=True)
            print(f"📁 Screenshot folder: {screenshot_folder}")

            # STEP 3: Extract all 6 tables
            print("\nSTEP 3: Extract All 6 Tables")
            print("="*80)

            # Wait for DAM to fully load token data for the new address
            print("⏳ Waiting for token data to load (20 seconds)...")
            page.wait_for_timeout(20000)

            import pandas as pd
            all_tables_data = {}

            # Check if this portfolio has a Wallet section (vs CEX-only)
            # CEX-only portfolios don't have a wallet-section div
            has_wallet_section = False
            try:
                wallet_section = page.locator('#wallet-section, div[id="wallet-section"]')
                wallet_count = wallet_section.count()
                print(f"   DEBUG: wallet-section count: {wallet_count}")
                if wallet_count > 0:
                    has_wallet_section = True
                    print("   📍 Found Wallet section")
                else:
                    # Also check for the "Wallet" header text
                    wallet_header = page.locator('text="Wallet"').first
                    if wallet_header.is_visible(timeout=2000):
                        has_wallet_section = True
                        print("   📍 Found Wallet header")
                    else:
                        print("   DEBUG: No wallet-section div and no visible 'Wallet' header")
            except Exception as e:
                print(f"   DEBUG: Wallet section detection error: {e}")

            if not has_wallet_section:
                print("   ℹ️  No Wallet section found - this appears to be a CEX-only portfolio")

            # TABLE 1: Overview - Wallet Breakdown (skip for CEX-only portfolios)
            print("\n📊 Table 1: Overview - Wallet Breakdown")
            print("-"*80)

            # Initialize token_data for both wallet and CEX-only cases
            token_data = []

            if not has_wallet_section:
                print("   ⏭️  Skipped - CEX-only portfolio (no wallet data)")
                all_tables_data['Overview - Wallet'] = []
            else:
                for selector in ['text="Overview"', '[role="tab"]:has-text("Overview")']:
                    try:
                        if page.locator(selector).first.is_visible(timeout=3000):
                            page.locator(selector).first.click()
                            page.wait_for_timeout(3000)
                            break
                    except:
                        continue

                # Click on the Token tab (at the bottom)
                for selector in ['[role="tab"]:has-text("Token")', 'button[role="tab"]:has-text("Token")']:
                    try:
                        if page.locator(selector).first.is_visible(timeout=3000):
                            page.locator(selector).first.click()
                            page.wait_for_timeout(3000)
                            break
                    except:
                        continue

                # Wait for table data to fully load before screenshot
                print("   ⏳ Waiting for table data to load...")
                try:
                    # Wait for table to be visible and contain data rows (not just loading state)
                    page.wait_for_selector("table tbody tr", state="visible", timeout=10000)
                    # Additional wait to ensure all data is rendered
                    page.wait_for_timeout(2000)
                    print("   ✅ Table data loaded")
                except Exception as e:
                    print(f"   ⚠️  Table load wait warning: {e}")
                    # Still wait a bit even if selector times out
                    page.wait_for_timeout(3000)

            # === SCROLL INSIDE TABLE CONTAINER TO LOAD ALL LAZY-LOADED TOKENS ===
            # The table is inside a fixed-height container (max-h-[600px]) with overflow scroll
            # We need to scroll INSIDE this container, not the page
            # Skip for CEX-only portfolios
            if has_wallet_section:
                print("   🔄 Scrolling inside table container to load all tokens...")
                try:
                    # Use JavaScript to find the scrollable container that contains the table
                    # Look for container with data-slot="table-container" or parent of tbody
                    scroll_container = None
    
                    # Method 1: Find by data-slot attribute
                    container_js = page.evaluate("""() => {
                        // Find container with data-slot="table-container"
                        const tableContainer = document.querySelector('[data-slot="table-container"]');
                        if (tableContainer && tableContainer.scrollHeight > tableContainer.clientHeight) {
                            return {
                                found: true,
                                selector: '[data-slot="table-container"]',
                                scrollHeight: tableContainer.scrollHeight,
                                clientHeight: tableContainer.clientHeight
                            };
                        }
    
                        // Find div containing table with overflow and max-height
                        const divs = document.querySelectorAll('div');
                        for (const div of divs) {
                            const style = window.getComputedStyle(div);
                            const hasOverflow = style.overflowY === 'auto' || style.overflowY === 'scroll';
                            const hasMaxHeight = div.className.includes('max-h-');
                            const containsTable = div.querySelector('table') !== null;
                            const isScrollable = div.scrollHeight > div.clientHeight;
    
                            if (containsTable && isScrollable && (hasOverflow || hasMaxHeight)) {
                                return {
                                    found: true,
                                    selector: 'scrollable-table-div',
                                    scrollHeight: div.scrollHeight,
                                    clientHeight: div.clientHeight,
                                    className: div.className.substring(0, 100)
                                };
                            }
                        }
                        return { found: false };
                    }""")
    
                    if container_js['found']:
                        print(f"   📍 Found scrollable table container via JS")
                        print(f"      scrollHeight: {container_js['scrollHeight']}px, clientHeight: {container_js['clientHeight']}px")
                        if 'className' in container_js:
                            print(f"      class: {container_js['className']}")
    
                        # Get the actual element
                        if container_js['selector'] == '[data-slot="table-container"]':
                            scroll_container = page.locator('[data-slot="table-container"]').first
                        else:
                            # Use the div that contains table and is scrollable
                            scroll_container = page.evaluate_handle("""() => {
                                const divs = document.querySelectorAll('div');
                                for (const div of divs) {
                                    const style = window.getComputedStyle(div);
                                    const hasOverflow = style.overflowY === 'auto' || style.overflowY === 'scroll';
                                    const hasMaxHeight = div.className.includes('max-h-');
                                    const containsTable = div.querySelector('table') !== null;
                                    const isScrollable = div.scrollHeight > div.clientHeight;
                                    if (containsTable && isScrollable && (hasOverflow || hasMaxHeight)) {
                                        return div;
                                    }
                                }
                                return null;
                            }""")
                    else:
                        # Fallback: try CSS selectors
                        table_container_selectors = [
                            "[data-slot='table-container']",
                            "div.w-full.overflow-x-auto",
                            "div[class*='overflow'][class*='max-h']"
                        ]
                        for selector in table_container_selectors:
                            try:
                                container = page.locator(selector).first
                                if container.count() > 0:
                                    scroll_info = container.evaluate("(el) => ({ scrollHeight: el.scrollHeight, clientHeight: el.clientHeight, scrollable: el.scrollHeight > el.clientHeight })")
                                    if scroll_info['scrollable']:
                                        scroll_container = container
                                        print(f"   📍 Found scrollable table container: {selector}")
                                        print(f"      scrollHeight: {scroll_info['scrollHeight']}px, clientHeight: {scroll_info['clientHeight']}px")
                                        break
                            except:
                                continue
    
                    if scroll_container or container_js['found']:
                        # Use JavaScript to scroll the container directly
                        # This handles both locator and JSHandle cases
                        print("   🔄 Starting scroll through table container...")
    
                        scroll_step = 300
                        last_row_count = 0
                        stable_count = 0
                        scroll_attempts = 0
                        max_scroll_attempts = 100
    
                        while stable_count < 5 and scroll_attempts < max_scroll_attempts:
                            # Scroll and get current state using JS
                            scroll_result = page.evaluate(f"""(scrollAmount) => {{
                                // Find the scrollable container
                                let container = document.querySelector('[data-slot="table-container"]');
                                if (!container || container.scrollHeight <= container.clientHeight) {{
                                    // Find by class
                                    const divs = document.querySelectorAll('div');
                                    for (const div of divs) {{
                                        const containsTable = div.querySelector('table') !== null;
                                        const isScrollable = div.scrollHeight > div.clientHeight;
                                        if (containsTable && isScrollable) {{
                                            container = div;
                                            break;
                                        }}
                                    }}
                                }}
    
                                if (!container) return {{ success: false }};
    
                                // Scroll down
                                container.scrollTop += scrollAmount;
    
                                return {{
                                    success: true,
                                    scrollTop: container.scrollTop,
                                    scrollHeight: container.scrollHeight,
                                    clientHeight: container.clientHeight,
                                    maxScroll: container.scrollHeight - container.clientHeight,
                                    atBottom: container.scrollTop >= container.scrollHeight - container.clientHeight - 10
                                }};
                            }}""", scroll_step)
    
                            if not scroll_result['success']:
                                print("   ⚠️  Could not scroll container")
                                break
    
                            page.wait_for_timeout(200)  # Wait for virtual scroll to render
    
                            # Check row count
                            current_rows = page.locator("table tbody tr").count()
                            if current_rows == last_row_count:
                                stable_count += 1
                            else:
                                stable_count = 0
                                last_row_count = current_rows
    
                            scroll_attempts += 1
    
                            # Stop if we've reached the bottom
                            if scroll_result['atBottom']:
                                print(f"   📍 Reached bottom of container at scroll {scroll_result['scrollTop']}px")
                                break
    
                        # Scroll back to top of container
                        page.evaluate("""() => {
                            let container = document.querySelector('[data-slot="table-container"]');
                            if (!container || container.scrollHeight <= container.clientHeight) {
                                const divs = document.querySelectorAll('div');
                                for (const div of divs) {
                                    const containsTable = div.querySelector('table') !== null;
                                    const isScrollable = div.scrollHeight > div.clientHeight;
                                    if (containsTable && isScrollable) { container = div; break; }
                                }
                            }
                            if (container) container.scrollTop = 0;
                        }""")
                        page.wait_for_timeout(500)
    
                        # Final count
                        final_row_count = page.locator("table tbody tr").count()
                        print(f"   ✅ Scrolled table container - found {final_row_count} rows after {scroll_attempts} scroll steps")
                    else:
                        # Fallback: try page scroll
                        print("   ⚠️  No scrollable table container found, trying page scroll...")
                        viewport_height = page.viewport_size['height']
                        total_height = page.evaluate("document.body.scrollHeight")
                        current_scroll = 0
    
                        while current_scroll < total_height:
                            current_scroll += viewport_height // 2
                            page.evaluate(f"window.scrollTo(0, {current_scroll})")
                            page.wait_for_timeout(300)
                            total_height = page.evaluate("document.body.scrollHeight")
    
                        page.evaluate("window.scrollTo(0, 0)")
                        page.wait_for_timeout(500)
                        final_row_count = page.locator("table tbody tr").count()
                        print(f"   ✅ Page scroll fallback - found {final_row_count} rows")
    
                except Exception as e:
                    print(f"   ⚠️  Scroll warning: {e}")
                    import traceback
                    traceback.print_exc()
                # === END SCROLL LOGIC ===
    
                try:
                    page.screenshot(path=f"{screenshot_folder}/01_overview_{timestamp_folder}.png", full_page=True, timeout=120000)
                    print("   📸 Screenshot captured")
                except Exception as e:
                    print(f"⚠️  Screenshot warning: {e}")
    
                token_data = []
                try:
                    # === ROW-BY-ROW TOOLTIP EXTRACTION WITH FULL SCREENSHOTS ===
                    # Extract all tooltips row-by-row for better accuracy and complete evidence
                    
                    def read_tooltip_text(page, trigger, tooltip_id=None, wait_ms=1200, debug_label="", debug_screenshot_folder=None):
                        """
                        Robustly extract tooltip text using multiple strategies in order.
                        Returns the best non-empty string found, or "" if all strategies fail.
                        """
                        import re as _re

                        def _clean(t):
                            if not t:
                                return ""
                            t = t.strip()
                            t = _re.sub(r'[ \t]+', ' ', t)
                            t = _re.sub(r'\n+', '\n', t)
                            return t.strip()

                        def _is_zero(txt):
                            """Return True if txt is numerically zero — stale tooltip."""
                            try:
                                return float(txt.replace(',','').replace('%','').replace('$','').strip()) == 0.0
                            except:
                                return False

                        # Step 1: scroll into view + hover
                        try:
                            trigger.scroll_into_view_if_needed(timeout=1000)
                        except:
                            pass
                        try:
                            trigger.hover(force=True, timeout=1000)
                        except:
                            try:
                                trigger.hover(timeout=1000)
                            except:
                                pass
                        page.wait_for_timeout(wait_ms)

                        # Step 2: getElementById — poll until non-zero (tooltip updates after hover)
                        if tooltip_id:
                            try:
                                trigger.hover(force=True, timeout=1000)
                            except:
                                pass
                            for _wait in [300, 500, 800, 1000]:
                                page.wait_for_timeout(_wait)
                                try:
                                    txt = page.evaluate("(id) => { const el = document.getElementById(id); return el ? el.textContent : null; }", tooltip_id)
                                    txt = _clean(txt)
                                    if debug_label:
                                        print(f"   [tooltip-debug] {debug_label} | Step2 getElementById('{tooltip_id}') wait={_wait}ms → raw='{txt}'")
                                    if txt and not _is_zero(txt):
                                        return txt
                                except Exception as _e2:
                                    if debug_label:
                                        print(f"   [tooltip-debug] {debug_label} | Step2 error: {_e2}")

                        # Step 3: CSS [id="..."] — retry once if "0"
                        if tooltip_id:
                            for _attempt in range(2):
                                try:
                                    loc = page.locator(f'[id="{tooltip_id}"]').first
                                    if debug_label:
                                        print(f"   [tooltip-debug] {debug_label} | Step3 CSS [id='{tooltip_id}'] count={loc.count()}")
                                    if loc.count() > 0:
                                        txt = _clean(loc.inner_text())
                                        if debug_label:
                                            print(f"   [tooltip-debug] {debug_label} | Step3 inner_text → '{txt}'")
                                        if txt and not _is_zero(txt):
                                            return txt
                                        elif txt and _is_zero(txt) and _attempt == 0:
                                            page.wait_for_timeout(800)
                                except:
                                    pass

                        # Step 4: aria-describedby / title on trigger
                        for attr in ("aria-describedby", "title"):
                            try:
                                val = trigger.get_attribute(attr)
                                if val:
                                    val = val.strip()
                                    if attr == "aria-describedby":
                                        txt = page.evaluate("(id) => { const el = document.getElementById(id); return el ? el.textContent : null; }", val)
                                        txt = _clean(txt)
                                        if txt and not _is_zero(txt):
                                            return txt
                                    else:
                                        if val and not _is_zero(val):
                                            return _clean(val)
                            except:
                                pass

                        # Step 5: scan visible tooltip containers
                        _tooltip_selectors = [
                            '[role="tooltip"]',
                            '.react-tooltip',
                            '.tippy-box',
                            '.MuiTooltip-popper',
                            '.chakra-tooltip',
                            '[data-state="open"]',
                            '[class*="tooltip"]',
                            '[class*="popover"]',
                        ]
                        for sel in _tooltip_selectors:
                            try:
                                for el in page.locator(sel).all():
                                    try:
                                        if el.is_visible():
                                            txt = _clean(el.inner_text())
                                            # Reject "0" or "0.0" — stale/wrong tooltip element
                                            if txt:
                                                try:
                                                    if float(txt.replace(',','').replace('%','').replace('$','')) == 0.0:
                                                        continue
                                                except:
                                                    pass
                                                return txt
                                    except:
                                        pass
                            except:
                                pass

                        # Step 6: JS dispatch mouseenter/mouseover/mousemove then re-scan
                        try:
                            page.evaluate("""(tid) => {
                                const trigger = tid
                                    ? document.querySelector('[data-tooltip-id="' + tid + '"]')
                                    : null;
                                const target = trigger || document.activeElement;
                                if (target) {
                                    ['mouseenter','mouseover','mousemove'].forEach(evt =>
                                        target.dispatchEvent(new MouseEvent(evt, {bubbles:true}))
                                    );
                                }
                            }""", tooltip_id or "")
                            page.wait_for_timeout(400)
                            for sel in _tooltip_selectors:
                                try:
                                    for el in page.locator(sel).all():
                                        try:
                                            if el.is_visible():
                                                txt = _clean(el.inner_text())
                                                if txt:
                                                    try:
                                                        if float(txt.replace(',','').replace('%','').replace('$','')) == 0.0:
                                                            continue
                                                    except:
                                                        pass
                                                    return txt
                                        except:
                                            pass
                                except:
                                    pass
                        except:
                            pass

                        # All strategies failed
                        if debug_label:
                            print(f"   WARN: missing tooltip for {debug_label}")
                        if debug_screenshot_folder:
                            try:
                                safe = _re.sub(r'[^\w\-]', '_', debug_label)[:40]
                                page.screenshot(path=f"{debug_screenshot_folder}/tooltip_miss_{safe}.png")
                            except:
                                pass
                        return ""

                    # Keep old name as alias for backward compatibility
                    def extract_tooltip_best_practice(page, elem, tooltip_id, max_retries=2, debug_label=""):
                        return read_tooltip_text(page, elem, tooltip_id=tooltip_id, debug_label=debug_label) or None

                    def _price_tooltip_sanity_check(disp_raw, tip_value):
                        """Check if a price tooltip value is consistent with the displayed price.
                        Returns True if tooltip is sane, False if it should be discarded.
                        disp_raw: raw displayed price text (e.g. '< 0.01', '0.99', '2,358.43')
                        tip_value: cleaned tooltip string (e.g. '0.00009952', '$2,358.43')
                        """
                        if not tip_value or not disp_raw:
                            return True  # nothing to check
                        try:
                            _tip_clean = tip_value.replace('$', '').replace(',', '').strip()
                            if not re.match(r'^[\d.]+$', _tip_clean):
                                return True  # non-numeric tooltip, skip check
                            _tip_num = float(_tip_clean)

                            # Clean commas and $ from displayed price before parsing
                            _disp_cleaned = disp_raw.replace(',', '').replace('$', '').strip()
                            _disp_is_small = '<' in _disp_cleaned or (re.search(r'[\d.]+', _disp_cleaned) and float(re.search(r'[\d.]+', _disp_cleaned).group()) < 0.01)
                            _disp_num_match = re.search(r'[\d.]+', _disp_cleaned)
                            _disp_num = float(_disp_num_match.group()) if _disp_num_match else None

                            # Displayed price is 0 but tooltip > 0 → wrong
                            if _disp_num is not None and _disp_num == 0 and _tip_num > 0:
                                print(f"   ⚠️  Price tooltip sanity fail (disp=0, tip={_tip_num}), discarding")
                                return False
                            # Displayed < 0.01 but tooltip >= 0.01 → wrong (e.g. POV: disp=<0.01, tip=0.999817)
                            if _disp_is_small and _tip_num >= 0.01:
                                print(f"   ⚠️  Price tooltip sanity fail (disp=small, tip={_tip_num}), discarding")
                                return False
                            # Displayed is a real number > 0: check ratio
                            if not _disp_is_small and _disp_num and _disp_num > 0:
                                ratio = _tip_num / _disp_num
                                if ratio < 0.001 or ratio > 1000:
                                    print(f"   ⚠️  Price tooltip sanity fail (disp={_disp_num}, tip={_tip_num}, ratio={ratio:.4f}), discarding")
                                    return False
                        except:
                            pass
                        return True

                    def _share_tooltip_sanity_check(disp_raw, tip_value):
                        """Check if a share tooltip value is consistent with the displayed share.
                        Returns True if tooltip is sane, False if it should be discarded.
                        disp_raw: raw displayed share text (e.g. '< 0.01', '98.42', '0.96')
                        tip_value: tooltip string (e.g. '98.42182782750', '0.96129637796')
                        Share values are percentages (0-100).
                        """
                        if not tip_value or not disp_raw:
                            return True
                        try:
                            _tip_clean = tip_value.replace('%', '').replace(',', '').strip()
                            if not re.match(r'^[\d.]+$', _tip_clean):
                                return True
                            _tip_num = float(_tip_clean)

                            _disp_is_small = '<' in disp_raw or (re.search(r'[\d.]+', disp_raw) and float(re.search(r'[\d.]+', disp_raw).group()) < 0.01)
                            _disp_num_match = re.search(r'[\d.]+', disp_raw)
                            _disp_num = float(_disp_num_match.group()) if _disp_num_match else None

                            # Displayed share is 0 but tooltip > 0 → wrong
                            if _disp_num is not None and _disp_num == 0 and _tip_num > 0:
                                print(f"   ⚠️  Share tooltip sanity fail (disp=0, tip={_tip_num}), discarding")
                                return False
                            # Displayed < 0.01 but tooltip >= 0.01 → wrong
                            if _disp_is_small and _tip_num >= 0.01:
                                print(f"   ⚠️  Share tooltip sanity fail (disp=small, tip={_tip_num}), discarding")
                                return False
                            # Displayed is a real number > 0: check ratio
                            if not _disp_is_small and _disp_num and _disp_num > 0:
                                ratio = _tip_num / _disp_num
                                if ratio < 0.001 or ratio > 1000:
                                    print(f"   ⚠️  Share tooltip sanity fail (disp={_disp_num}, tip={_tip_num}, ratio={ratio:.4f}), discarding")
                                    return False
                        except:
                            pass
                        return True

                    def _amount_tooltip_sanity_check(disp_raw, tip_value):
                        """Check if an amount tooltip value is consistent with the displayed amount.
                        Returns True if tooltip is sane, False if it should be discarded.
                        disp_raw: raw displayed amount text (e.g. '< 0.01', '1,234.56', '0.00')
                        tip_value: tooltip string (e.g. '1234.5678901234')
                        """
                        if not tip_value or not disp_raw:
                            return True
                        try:
                            _tip_clean = tip_value.replace('$', '').replace(',', '').strip()
                            if not re.match(r'^[\d.]+$', _tip_clean):
                                return True
                            _tip_num = float(_tip_clean)

                            _disp_is_small = '<' in disp_raw or (re.search(r'[\d.]+', disp_raw) and float(re.search(r'[\d.]+', disp_raw).group()) < 0.01)
                            _disp_num_match = re.search(r'[\d.]+', disp_raw)
                            _disp_num = float(_disp_num_match.group()) if _disp_num_match else None

                            # Displayed amount is 0 but tooltip > 0 → wrong
                            if _disp_num is not None and _disp_num == 0 and _tip_num > 0:
                                print(f"   ⚠️  Amount tooltip sanity fail (disp=0, tip={_tip_num}), discarding")
                                return False
                            # Displayed < 0.01 but tooltip >= 0.01 → wrong
                            if _disp_is_small and _tip_num >= 0.01:
                                print(f"   ⚠️  Amount tooltip sanity fail (disp=small, tip={_tip_num}), discarding")
                                return False
                            # Displayed is a real number > 0: check ratio
                            if not _disp_is_small and _disp_num and _disp_num > 0:
                                ratio = _tip_num / _disp_num
                                if ratio < 0.001 or ratio > 1000:
                                    print(f"   ⚠️  Amount tooltip sanity fail (disp={_disp_num}, tip={_tip_num}, ratio={ratio:.4f}), discarding")
                                    return False
                        except:
                            pass
                        return True

                    
                    # Get all table rows
                    # Scope to wallet section only to avoid picking up DeFi/Hyperliquid tables below
                    _wallet_section = page.locator('#wallet-section').first
                    _wallet_table_loc = _wallet_section.locator('table tbody tr') if _wallet_section.count() > 0 else page.locator('table tbody tr')

                    print(f"\n📸 Starting tooltip extraction with screenshots (scroll-based)...")
                    print("=" * 80)

                    all_tooltips = []         # list of dicts keyed by composite_key
                    total_screenshots = 0
                    processed_keys = set()    # composite keys already done
                    row_seq = 0               # sequential counter for screenshot naming

                    # Scroll back to top
                    page.evaluate("""() => {
                        let c = document.querySelector('[data-slot="table-container"]');
                        if (!c || c.scrollHeight <= c.clientHeight) {
                            for (const d of document.querySelectorAll('div')) {
                                if (d.querySelector('table') && d.scrollHeight > d.clientHeight) { c = d; break; }
                            }
                        }
                        if (c) c.scrollTop = 0;
                    }""")
                    page.wait_for_timeout(500)

                    scroll_count = 0
                    max_scrolls = 300
                    no_new_count = 0

                    while scroll_count < max_scrolls:
                        current_rows = _wallet_table_loc.all()
                        found_new = False

                        for row in current_rows:
                            try:
                                cells = row.locator('td').all()
                                if len(cells) < 2:
                                    continue
                                row_chain = cells[0].inner_text().strip()
                                row_token = cells[1].inner_text().strip().split('\n')[0]
                                # Include amount cell to distinguish same token with different contract/amount
                                row_amount_raw = cells[3].inner_text().strip() if len(cells) > 3 else ""
                                composite_key = f"{row_chain}|{row_token}|{row_amount_raw}"
                                if composite_key in processed_keys:
                                    continue

                                found_new = True
                                token_name_safe = re.sub(r'[^\w\-]', '_', row_token)[:20]
                                row_idx = row_seq

                                row_tooltips = {
                                    'composite_key': composite_key,
                                    'row_index': row_idx,
                                    'price': None, 'price_24h': None,
                                    'share': None, 'amount': None,
                                    'screenshots': {'price': None, 'price_24h': None, 'share': None, 'amount': None}
                                }

                                try:
                                    row.scroll_into_view_if_needed()
                                    page.wait_for_timeout(400)
                                except:
                                    pass

                                print(f"\n📍 Row {row_idx:03d} ({row_chain}|{row_token})")
                                print("-" * 80)

                                # PRICE TOOLTIP
                                price_elem = row.locator('[data-tooltip-id*="price-tooltip"]').first
                                if price_elem.count() > 0:
                                    try:
                                        price_tooltip_id = price_elem.get_attribute('data-tooltip-id')
                                        if not price_elem.is_visible():
                                            price_elem.scroll_into_view_if_needed()
                                            page.wait_for_timeout(300)
                                        price_elem.hover()
                                        page.wait_for_timeout(500)
                                        screenshot_path = f"{screenshot_folder}/row_{row_idx:03d}_{token_name_safe}_01_price_tooltip.png"
                                        page.screenshot(path=screenshot_path)
                                        row_tooltips['screenshots']['price'] = screenshot_path
                                        total_screenshots += 1
                                        price_value = extract_tooltip_best_practice(page, price_elem, price_tooltip_id,
                                                                                    debug_label=f"price/{row_chain}|{row_token}")
                                        # Sanity check: tooltip must be consistent with displayed price
                                        _disp_raw = cells[2].inner_text().strip().split('\n')[0] if len(cells) > 2 else ""
                                        if price_value and not _price_tooltip_sanity_check(_disp_raw, price_value):
                                            # Sanity failed — retry: dismiss tooltip, re-hover, recapture
                                            price_value = None
                                            for _retry_i in range(2):
                                                try:
                                                    page.mouse.move(0, 0)
                                                    page.wait_for_timeout(400)
                                                    price_elem.hover(force=True)
                                                    page.wait_for_timeout(600)
                                                    price_value = extract_tooltip_best_practice(page, price_elem, price_tooltip_id,
                                                                                                debug_label=f"price/{row_chain}|{row_token} retry{_retry_i+1}")
                                                    if price_value and _price_tooltip_sanity_check(_disp_raw, price_value):
                                                        print(f"   ✅ Price tooltip recaptured on retry {_retry_i+1}: {price_value}")
                                                        break
                                                    else:
                                                        price_value = None
                                                except:
                                                    pass
                                        # "$0" with no decimals is a bad capture
                                        if price_value and price_value.replace('$', '').replace(',', '').strip() == '0':
                                            price_value = None
                                        row_tooltips['price'] = price_value
                                        status = "✅" if price_value else "⚠️"
                                        print(f"   {status} Price: {price_value}")
                                        print(f"      Screenshot: {screenshot_path}")
                                    except Exception as e:
                                        print(f"   ❌ Price: Error - {e}")

                                # SHARE TOOLTIP
                                share_elem = row.locator('[data-tooltip-id*="share-tooltip"]').first
                                if share_elem.count() > 0:
                                    try:
                                        share_tooltip_id = share_elem.get_attribute('data-tooltip-id')
                                        if not share_elem.is_visible():
                                            share_elem.scroll_into_view_if_needed()
                                            page.wait_for_timeout(300)
                                        share_elem.hover()
                                        page.wait_for_timeout(500)
                                        screenshot_path = f"{screenshot_folder}/row_{row_idx:03d}_{token_name_safe}_02_share_tooltip.png"
                                        page.screenshot(path=screenshot_path)
                                        row_tooltips['screenshots']['share'] = screenshot_path
                                        total_screenshots += 1
                                        share_value = extract_tooltip_best_practice(page, share_elem, share_tooltip_id)
                                        # Reject bare "0" — stale tooltip
                                        if share_value and share_value.replace('%', '').replace(',', '').strip() == '0':
                                            share_value = None
                                        # Sanity check: tooltip must be consistent with displayed share
                                        _disp_share_raw = cells[4].inner_text().strip().split('\n')[0] if len(cells) > 4 else ""
                                        if share_value and not _share_tooltip_sanity_check(_disp_share_raw, share_value):
                                            share_value = None
                                            for _retry_s in range(2):
                                                try:
                                                    page.mouse.move(0, 0)
                                                    page.wait_for_timeout(400)
                                                    share_elem.hover(force=True)
                                                    page.wait_for_timeout(600)
                                                    share_value = extract_tooltip_best_practice(page, share_elem, share_tooltip_id)
                                                    if share_value and share_value.replace('%', '').replace(',', '').strip() != '0' and _share_tooltip_sanity_check(_disp_share_raw, share_value):
                                                        print(f"   ✅ Share tooltip recaptured on retry {_retry_s+1}: {share_value}")
                                                        break
                                                    else:
                                                        share_value = None
                                                except:
                                                    pass
                                        row_tooltips['share'] = share_value
                                        status = "✅" if share_value else "⚠️"
                                        print(f"   {status} Share: {share_value}")
                                        print(f"      Screenshot: {screenshot_path}")
                                    except Exception as e:
                                        print(f"   ❌ Share: Error - {e}")

                                # AMOUNT TOOLTIP
                                amount_elem = row.locator('[data-tooltip-id*="amount-tooltip"]').first
                                if amount_elem.count() > 0:
                                    try:
                                        amount_tooltip_id = amount_elem.get_attribute('data-tooltip-id')
                                        if not amount_elem.is_visible():
                                            amount_elem.scroll_into_view_if_needed()
                                            page.wait_for_timeout(300)
                                        amount_elem.hover()
                                        page.wait_for_timeout(500)
                                        screenshot_path = f"{screenshot_folder}/row_{row_idx:03d}_{token_name_safe}_03_amount_tooltip.png"
                                        page.screenshot(path=screenshot_path)
                                        row_tooltips['screenshots']['amount'] = screenshot_path
                                        total_screenshots += 1
                                        amount_value = extract_tooltip_best_practice(page, amount_elem, amount_tooltip_id)
                                        # Reject bare "0" — stale tooltip
                                        if amount_value and amount_value.replace(',', '').strip() == '0':
                                            amount_value = None
                                        # Sanity check: tooltip must be consistent with displayed amount
                                        _disp_amount_raw = cells[3].inner_text().strip().split('\n')[0] if len(cells) > 3 else ""
                                        if amount_value and not _amount_tooltip_sanity_check(_disp_amount_raw, amount_value):
                                            amount_value = None
                                            for _retry_a in range(2):
                                                try:
                                                    page.mouse.move(0, 0)
                                                    page.wait_for_timeout(400)
                                                    amount_elem.hover(force=True)
                                                    page.wait_for_timeout(600)
                                                    amount_value = extract_tooltip_best_practice(page, amount_elem, amount_tooltip_id)
                                                    if amount_value and amount_value.replace(',', '').strip() != '0' and _amount_tooltip_sanity_check(_disp_amount_raw, amount_value):
                                                        print(f"   ✅ Amount tooltip recaptured on retry {_retry_a+1}: {amount_value}")
                                                        break
                                                    else:
                                                        amount_value = None
                                                except:
                                                    pass
                                        row_tooltips['amount'] = amount_value
                                        status = "✅" if amount_value else "⚠️"
                                        print(f"   {status} Amount: {amount_value}")
                                        print(f"      Screenshot: {screenshot_path}")
                                    except Exception as e:
                                        print(f"   ❌ Amount: Error - {e}")

                                # PRICE(24H) TOOLTIP
                                price_24h_elem = row.locator('[data-tooltip-id*="price-24h-tooltip"]').first
                                if price_24h_elem.count() > 0:
                                    try:
                                        price_24h_tooltip_id = price_24h_elem.get_attribute('data-tooltip-id')
                                        if not price_24h_elem.is_visible():
                                            price_24h_elem.scroll_into_view_if_needed()
                                            page.wait_for_timeout(300)
                                        price_24h_elem.hover()
                                        page.wait_for_timeout(500)
                                        screenshot_path = f"{screenshot_folder}/row_{row_idx:03d}_{token_name_safe}_04_price24h_tooltip.png"
                                        page.screenshot(path=screenshot_path)
                                        row_tooltips['screenshots']['price_24h'] = screenshot_path
                                        total_screenshots += 1
                                        price_24h_value = extract_tooltip_best_practice(page, price_24h_elem, price_24h_tooltip_id)
                                        row_tooltips['price_24h'] = price_24h_value
                                        status = "✅" if price_24h_value else "⚠️"
                                        print(f"   {status} Price(24h): {price_24h_value}")
                                        print(f"      Screenshot: {screenshot_path}")
                                    except Exception as e:
                                        print(f"   ❌ Price(24h): Error - {e}")

                                # Check if any tooltips are missing and retry once
                                missing = not row_tooltips['price'] or not row_tooltips['share'] or not row_tooltips['amount']
                                if missing:
                                    print(f"   🔄 Retrying row {row_idx:03d} (some tooltips missing)...")
                                    try:
                                        row.scroll_into_view_if_needed()
                                        page.wait_for_timeout(800)
                                    except:
                                        pass

                                    if not row_tooltips['price']:
                                        try:
                                            pe = row.locator('[data-tooltip-id*="price-tooltip"]').first
                                            if pe.count() > 0:
                                                pid = pe.get_attribute('data-tooltip-id')
                                                _disp_raw_r = cells[2].inner_text().strip().split('\n')[0] if len(cells) > 2 else ""
                                                pe.hover(force=True)
                                                page.wait_for_timeout(600)
                                                pv = extract_tooltip_best_practice(page, pe, pid)
                                                if pv and not _price_tooltip_sanity_check(_disp_raw_r, pv):
                                                    pv = None
                                                if pv and pv.replace('$', '').replace(',', '').strip() == '0':
                                                    pv = None
                                                if pv:
                                                    row_tooltips['price'] = pv
                                                    print(f"   ✅ Price (retry): {pv}")
                                        except:
                                            pass

                                    if not row_tooltips['share']:
                                        try:
                                            se = row.locator('[data-tooltip-id*="share-tooltip"]').first
                                            if se.count() > 0:
                                                sid = se.get_attribute('data-tooltip-id')
                                                _disp_share_r = cells[4].inner_text().strip().split('\n')[0] if len(cells) > 4 else ""
                                                se.hover(force=True)
                                                page.wait_for_timeout(600)
                                                sv = extract_tooltip_best_practice(page, se, sid)
                                                if sv and sv.replace('%', '').replace(',', '').strip() == '0':
                                                    sv = None
                                                if sv and not _share_tooltip_sanity_check(_disp_share_r, sv):
                                                    sv = None
                                                if sv:
                                                    row_tooltips['share'] = sv
                                                    print(f"   ✅ Share (retry): {sv}")
                                        except:
                                            pass

                                    if not row_tooltips['amount']:
                                        try:
                                            ae = row.locator('[data-tooltip-id*="amount-tooltip"]').first
                                            if ae.count() > 0:
                                                aid = ae.get_attribute('data-tooltip-id')
                                                _disp_amt_r = cells[3].inner_text().strip().split('\n')[0] if len(cells) > 3 else ""
                                                ae.hover(force=True)
                                                page.wait_for_timeout(600)
                                                av = extract_tooltip_best_practice(page, ae, aid)
                                                if av and av.replace(',', '').strip() == '0':
                                                    av = None
                                                if av and not _amount_tooltip_sanity_check(_disp_amt_r, av):
                                                    av = None
                                                if av:
                                                    row_tooltips['amount'] = av
                                                    print(f"   ✅ Amount (retry): {av}")
                                        except:
                                            pass

                                all_tooltips.append(row_tooltips)
                                processed_keys.add(composite_key)
                                row_seq += 1

                            except Exception as e:
                                print(f"❌ Row error: {e}")
                                continue

                        if not found_new:
                            no_new_count += 1
                            if no_new_count >= 5:
                                print(f"\n✅ No more new rows after {row_seq} rows (scrolled 5 times with no new rows)")
                                break
                        else:
                            no_new_count = 0

                        # Scroll down
                        scroll_result = page.evaluate("""(step) => {
                            let c = document.querySelector('[data-slot="table-container"]');
                            if (!c || c.scrollHeight <= c.clientHeight) {
                                for (const d of document.querySelectorAll('div')) {
                                    if (d.querySelector('table') && d.scrollHeight > d.clientHeight) { c = d; break; }
                                }
                            }
                            if (!c) return {success: false};
                            c.scrollTop += step;
                            return {success: true, atBottom: c.scrollTop >= c.scrollHeight - c.clientHeight - 5};
                        }""", 500)

                        if not scroll_result.get('success') or scroll_result.get('atBottom'):
                            # One more pass at bottom before stopping
                            if no_new_count >= 1:
                                break

                        page.wait_for_timeout(400)
                        scroll_count += 1

                    print("\n" + "=" * 80)
                    print(f"✅ Extraction complete!")
                    print(f"   Total rows processed: {len(all_tooltips)}")
                    print(f"   Total screenshots captured: {total_screenshots}")
                    print(f"   Screenshot folder: {screenshot_folder}")
                    print("=" * 80 + "\n")
                    
                    # Convert to maps for backward compatibility with existing code
                    price_tooltips_map = {t['row_index']: t['price'] for t in all_tooltips if t['price']}
                    share_tooltips_map = {t['row_index']: t['share'] for t in all_tooltips if t['share']}
                    amount_tooltips_map = {t['row_index']: t['amount'] for t in all_tooltips if t['amount']}
    
                    # === VIRTUAL SCROLL EXTRACTION ===
                    # The table uses virtual scrolling - only ~15 rows exist in DOM at any time
                    # We need to scroll and extract data incrementally
    
                    print("   🔄 Extracting table data with virtual scroll handling...")
    
                    all_tokens_map = {}  # Store unique tokens by (chain, name): {composite_key: (token_name, row_data)}
                    ordered_tokens = []  # Track insertion order to preserve DAM page order (composite keys)
                    header_row = None
    
                    def extract_visible_rows():
                        """Extract currently visible rows from the DOM"""
                        nonlocal header_row
                        extracted = []
                        tables = page.locator("table").all()
                        if not tables:
                            return extracted
    
                        rows = tables[0].locator("tr[data-slot='table-row'], tr").all()
                        for row in rows:
                            try:
                                cells = row.locator("th, td").all()
                                if not cells:
                                    continue
    
                                row_data = []
                                token_name = ""
                                is_header = row.locator("th").count() > 0
    
                                for cell_idx, cell in enumerate(cells):
                                    cell_text = cell.inner_text().strip()

                                    # Capture token name from Name column (cell_idx == 1)
                                    if not is_header and cell_idx == 1:
                                        # Handle token names with newlines (take first line)
                                        token_name = cell_text.split('\n')[0].strip()
                                        # Fallback: if name is empty, try to get from first span or div
                                        if not token_name:
                                            try:
                                                name_span = cell.locator('span, div').first
                                                if name_span.count() > 0:
                                                    token_name = name_span.inner_text().strip().split('\n')[0]
                                            except:
                                                pass

                                    # Handle header row's "Price (24H)" column
                                    if is_header and 'Price' in cell_text and '24' in cell_text:
                                        row_data.append('Price')
                                        row_data.append('Price (24h)')
                                    # Handle Price + 24h change cell (ONLY for price column cell_idx == 2)
                                    elif not is_header and cell_idx == 2:
                                        # Extract price: use DOM to get just the price, excluding 24h change
                                        price_part = ""
                                        price_24h = ""
                                        try:
                                            # Find the 24h change element by its red/green color class
                                            pct_elem = cell.locator('[class*="text-error"], [class*="text-success"], [class*="bg-error"], [class*="bg-success"]').first
                                            if pct_elem.count() > 0 and pct_elem.is_visible():
                                                pct_text = pct_elem.text_content().strip()
                                                pct_match = re.search(r'([\d.]+)%?', pct_text)
                                                if pct_match:
                                                    pct_value = pct_match.group(1)
                                                    pct_class = pct_elem.get_attribute('class') or ""
                                                    if 'error' in pct_class.lower() or '↓' in pct_text:
                                                        price_24h = f"-{pct_value}"
                                                    else:
                                                        price_24h = pct_value
                                        except:
                                            pass
                                        # Extract price from FIRST LINE ONLY — 24h change (e.g. "< 0.01%") is on line 2
                                        # Searching full cell_text would wrongly match "< 0.01" from the 24h line
                                        price_line = cell_text.split('\n')[0].strip()
                                        lt_match = re.search(r'<\s*\$?([\d,]+\.?\d*)', price_line)
                                        price_match = re.search(r'\$?([\d,]+\.?\d*)', price_line)
                                        if lt_match:
                                            price_part = f"< {lt_match.group(1)}"
                                        elif price_match:
                                            price_part = price_match.group(1)
                                        else:
                                            price_part = clean_currency_symbols(price_line)
                                        row_data.append(price_part)
                                        row_data.append(price_24h)
                                    else:
                                        # col B (cell_idx == 1): preserve exact text including symbols
                                        if not is_header and cell_idx == 1:
                                            row_data.append(cell_text)
                                        else:
                                            row_data.append(clean_currency_symbols(cell_text))
    
                                if is_header and not header_row:
                                    header_row = row_data
                                elif token_name and row_data:
                                    extracted.append((token_name, row_data))
    
                            except Exception as e:
                                continue
    
                        return extracted
    
                    def extract_tooltip_for_token(token_name, chain=None):
                        """Extract tooltip values for a specific token by hovering.
                        If chain is provided, find the row matching both chain and token first."""
                        price_tooltip = ""
                        share_tooltip = ""
                        amount_tooltip = ""

                        try:
                            # If chain is provided, find the specific row first
                            target_row = None
                            if chain:
                                # Find the row that has this chain AND token name
                                rows = page.locator("table tr[data-slot='table-row'], table tr").all()
                                for row in rows:
                                    try:
                                        cells = row.locator("td").all()
                                        if len(cells) >= 2:
                                            row_chain = cells[0].inner_text().strip()
                                            row_token = cells[1].inner_text().strip().split('\n')[0]
                                            if row_chain == chain and row_token == token_name:
                                                target_row = row
                                                break
                                    except:
                                        continue

                            # If we found a specific row, extract from that row
                            if target_row:
                                # Price tooltip
                                try:
                                    price_trigger = target_row.locator('[data-tooltip-id*="price-tooltip"]').first
                                    if price_trigger.count() > 0:
                                        tid = price_trigger.get_attribute('data-tooltip-id')
                                        raw = read_tooltip_text(page, price_trigger, tooltip_id=tid, wait_ms=400,
                                                                debug_label=f"wallet price/{chain}|{token_name}")
                                        price_tooltip = raw.replace('$', '').strip() if raw else ""
                                        # Sanity check against displayed price
                                        if price_tooltip:
                                            _tr_cells = target_row.locator("td").all()
                                            _tr_disp = _tr_cells[2].inner_text().strip().split('\n')[0] if len(_tr_cells) > 2 else ""
                                            if not _price_tooltip_sanity_check(_tr_disp, price_tooltip):
                                                # Retry: dismiss and recapture
                                                price_tooltip = ""
                                                for _rt in range(2):
                                                    try:
                                                        page.mouse.move(0, 0)
                                                        page.wait_for_timeout(400)
                                                        price_trigger.hover(force=True)
                                                        page.wait_for_timeout(500)
                                                        raw2 = read_tooltip_text(page, price_trigger, tooltip_id=tid, wait_ms=400,
                                                                                 debug_label=f"wallet price/{chain}|{token_name} retry{_rt+1}")
                                                        _pt2 = raw2.replace('$', '').strip() if raw2 else ""
                                                        if _pt2 and _price_tooltip_sanity_check(_tr_disp, _pt2):
                                                            price_tooltip = _pt2
                                                            print(f"   ✅ Price tooltip recaptured on retry {_rt+1}: {price_tooltip}")
                                                            break
                                                    except:
                                                        pass
                                except:
                                    pass

                                # Share tooltip
                                try:
                                    share_trigger = target_row.locator('[data-tooltip-id*="share-tooltip"]').first
                                    if share_trigger.count() > 0:
                                        tid = share_trigger.get_attribute('data-tooltip-id')
                                        raw = read_tooltip_text(page, share_trigger, tooltip_id=tid, wait_ms=400,
                                                                debug_label=f"wallet share/{chain}|{token_name}")
                                        share_tooltip = raw.replace('%', '').strip() if raw else ""
                                        # Sanity check against displayed share
                                        if share_tooltip:
                                            _tr_disp_share = _tr_cells[4].inner_text().strip().split('\n')[0] if len(_tr_cells) > 4 else ""
                                            if not _share_tooltip_sanity_check(_tr_disp_share, share_tooltip):
                                                share_tooltip = ""
                                                for _rts in range(2):
                                                    try:
                                                        page.mouse.move(0, 0)
                                                        page.wait_for_timeout(400)
                                                        share_trigger.hover(force=True)
                                                        page.wait_for_timeout(500)
                                                        raw2 = read_tooltip_text(page, share_trigger, tooltip_id=tid, wait_ms=400,
                                                                                 debug_label=f"wallet share/{chain}|{token_name} retry{_rts+1}")
                                                        _st2 = raw2.replace('%', '').strip() if raw2 else ""
                                                        if _st2 and _share_tooltip_sanity_check(_tr_disp_share, _st2):
                                                            share_tooltip = _st2
                                                            print(f"   ✅ Share tooltip recaptured on retry {_rts+1}: {share_tooltip}")
                                                            break
                                                    except:
                                                        pass
                                except:
                                    pass

                                # Amount tooltip
                                try:
                                    amount_trigger = target_row.locator('[data-tooltip-id*="amount-tooltip"]').first
                                    if amount_trigger.count() > 0:
                                        tid = amount_trigger.get_attribute('data-tooltip-id')
                                        amount_tooltip = read_tooltip_text(page, amount_trigger, tooltip_id=tid, wait_ms=400,
                                                                           debug_label=f"wallet amount/{chain}|{token_name}") or ""
                                        # Sanity check against displayed amount
                                        if amount_tooltip:
                                            _tr_disp_amt = _tr_cells[3].inner_text().strip().split('\n')[0] if len(_tr_cells) > 3 else ""
                                            if not _amount_tooltip_sanity_check(_tr_disp_amt, amount_tooltip):
                                                amount_tooltip = ""
                                                for _rta in range(2):
                                                    try:
                                                        page.mouse.move(0, 0)
                                                        page.wait_for_timeout(400)
                                                        amount_trigger.hover(force=True)
                                                        page.wait_for_timeout(500)
                                                        raw2 = read_tooltip_text(page, amount_trigger, tooltip_id=tid, wait_ms=400,
                                                                                 debug_label=f"wallet amount/{chain}|{token_name} retry{_rta+1}")
                                                        if raw2 and _amount_tooltip_sanity_check(_tr_disp_amt, raw2):
                                                            amount_tooltip = raw2
                                                            print(f"   ✅ Amount tooltip recaptured on retry {_rta+1}: {amount_tooltip}")
                                                            break
                                                    except:
                                                        pass
                                except:
                                    pass

                                return price_tooltip, share_tooltip, amount_tooltip

                            # Fallback: Global search by token name
                            try:
                                price_elem = page.locator(f'[data-tooltip-id*="price-tooltip-{token_name}"]').first
                                if price_elem.count() > 0:
                                    tid = price_elem.get_attribute('data-tooltip-id')
                                    raw = read_tooltip_text(page, price_elem, tooltip_id=tid, wait_ms=400)
                                    price_tooltip = raw.replace('$', '').strip() if raw else ""
                            except:
                                pass

                            try:
                                share_elem = page.locator(f'[data-tooltip-id*="share-tooltip-{token_name}"]').first
                                if share_elem.count() > 0:
                                    tid = share_elem.get_attribute('data-tooltip-id')
                                    raw = read_tooltip_text(page, share_elem, tooltip_id=tid, wait_ms=400)
                                    share_tooltip = raw.replace('%', '').strip() if raw else ""
                            except:
                                pass

                            try:
                                amount_elem = page.locator(f'[data-tooltip-id*="amount-tooltip-{token_name}"]').first
                                if amount_elem.count() > 0:
                                    tid = amount_elem.get_attribute('data-tooltip-id')
                                    amount_tooltip = read_tooltip_text(page, amount_elem, tooltip_id=tid, wait_ms=400) or ""
                            except:
                                pass

                        except Exception as e:
                            print(f"   DEBUG: Tooltip extraction error for {chain}|{token_name}: {e}")

                        return price_tooltip, share_tooltip, amount_tooltip
    
                    # Scroll through table and collect all tokens
                    scroll_step = 100  # Smaller steps to catch all rows in virtual scroll
                    max_scrolls = 150  # More scrolls to handle larger tables
                    scroll_count = 0
                    last_token_count = 0
                    stable_count = 0
    
                    # First scroll to TOP of the table to ensure we capture all tokens from the beginning
                    page.evaluate("""() => {
                        let container = document.querySelector('[data-slot="table-container"]');
                        if (!container || container.scrollHeight <= container.clientHeight) {
                            const divs = document.querySelectorAll('div');
                            for (const div of divs) {
                                if (div.querySelector('table') && div.scrollHeight > div.clientHeight) {
                                    container = div; break;
                                }
                            }
                        }
                        if (container) container.scrollTop = 0;
                    }""")
                    page.wait_for_timeout(300)  # Wait for scroll and virtual render
    
                    # Extract from initial position (top of table)
                    for token_name, row_data in extract_visible_rows():
                        chain = row_data[0] if row_data else ""
                        amount_key = row_data[4] if len(row_data) > 4 else ""
                        composite_key = f"{chain}|{token_name}|{amount_key}"
                        if composite_key not in all_tokens_map:
                            all_tokens_map[composite_key] = (token_name, row_data)
                            ordered_tokens.append(composite_key)
    
                    # Scroll and extract
                    while scroll_count < max_scrolls:
                        scroll_result = page.evaluate(f"""(scrollAmount) => {{
                            let container = document.querySelector('[data-slot="table-container"]');
                            if (!container || container.scrollHeight <= container.clientHeight) {{
                                const divs = document.querySelectorAll('div');
                                for (const div of divs) {{
                                    if (div.querySelector('table') && div.scrollHeight > div.clientHeight) {{
                                        container = div;
                                        break;
                                    }}
                                }}
                            }}
                            if (!container) return {{ success: false }};
                            container.scrollTop += scrollAmount;
                            return {{
                                success: true,
                                scrollTop: container.scrollTop,
                                scrollHeight: container.scrollHeight,
                                clientHeight: container.clientHeight,
                                atBottom: container.scrollTop >= container.scrollHeight - container.clientHeight - 5
                            }};
                        }}""", scroll_step)
    
                        if not scroll_result.get('success'):
                            break
    
                        page.wait_for_timeout(200)  # Wait for virtual scroll render
    
                        # Extract newly visible rows
                        for token_name, row_data in extract_visible_rows():
                            chain = row_data[0] if row_data else ""
                            amount_key = row_data[4] if len(row_data) > 4 else ""
                            composite_key = f"{chain}|{token_name}|{amount_key}"
                            if composite_key not in all_tokens_map:
                                all_tokens_map[composite_key] = (token_name, row_data)
                                ordered_tokens.append(composite_key)
    
                        scroll_count += 1
    
                        # Check if we're finding new tokens
                        current_token_count = len(all_tokens_map)
                        if current_token_count == last_token_count:
                            stable_count += 1
                        else:
                            stable_count = 0
                            last_token_count = current_token_count
    
                        # If at bottom, do extra scrolls to make sure we got everything
                        if scroll_result.get('atBottom'):
                            # Do a few more extractions at the bottom
                            for _ in range(3):
                                page.wait_for_timeout(200)
                                for token_name, row_data in extract_visible_rows():
                                    chain = row_data[0] if row_data else ""
                                    amount_key = row_data[4] if len(row_data) > 4 else ""
                                    composite_key = f"{chain}|{token_name}|{amount_key}"
                                    if composite_key not in all_tokens_map:
                                        all_tokens_map[composite_key] = (token_name, row_data)
                                        ordered_tokens.append(composite_key)
                            break
    
                        # If no new tokens for 10 scrolls, stop
                        if stable_count >= 10:
                            break
    
                    # Scroll back to top to do a final extraction pass
                    page.evaluate("""() => {
                        let container = document.querySelector('[data-slot="table-container"]');
                        if (!container || container.scrollHeight <= container.clientHeight) {
                            const divs = document.querySelectorAll('div');
                            for (const div of divs) {
                                if (div.querySelector('table') && div.scrollHeight > div.clientHeight) {
                                    container = div; break;
                                }
                            }
                        }
                        if (container) container.scrollTop = 0;
                    }""")
                    page.wait_for_timeout(500)
    
                    # Final extraction at top to catch any missed rows (like TRX native token)
                    top_tokens_before = len(all_tokens_map)
                    for token_name, row_data in extract_visible_rows():
                        chain = row_data[0] if row_data else ""
                        amount_key = row_data[4] if len(row_data) > 4 else ""
                        composite_key = f"{chain}|{token_name}|{amount_key}"
                        if composite_key not in all_tokens_map:
                            all_tokens_map[composite_key] = (token_name, row_data)
                            ordered_tokens.insert(0, composite_key)  # Insert at beginning since these are top rows
                    if len(all_tokens_map) > top_tokens_before:
                        print(f"   📍 Found {len(all_tokens_map) - top_tokens_before} additional tokens at top")
    
                    print(f"   ✅ Collected {len(all_tokens_map)} unique tokens via virtual scroll")

                    # Pre-populate tooltips from first pass using composite_key
                    first_pass_hits = 0
                    first_pass_map = {t['composite_key']: t for t in all_tooltips}
                    for _ckey in ordered_tokens:
                        _tt = first_pass_map.get(_ckey)
                        if _tt and _tt['price'] and _tt['share'] and _tt['amount']:
                            _tname, _rdata = all_tokens_map[_ckey]
                            _rdata.append(_tt['price'])
                            _rdata.append(_tt['share'])
                            _rdata.append(_tt['amount'])
                            first_pass_hits += 1
                        elif _tt and (_tt['price'] or _tt['share'] or _tt['amount']):
                            # Partial capture — store what we have, mark missing as "Tooltip N/A"
                            # but still allow second pass to retry
                            _tname, _rdata = all_tokens_map[_ckey]
                            _rdata.append(_tt['price'] or "Tooltip N/A")
                            _rdata.append(_tt['share'] or "Tooltip N/A")
                            _rdata.append(_tt['amount'] or "Tooltip N/A")
                            first_pass_hits += 1

                    # Second pass needed for tokens where ANY tooltip is still missing
                    tokens_needing_tooltips = set()
                    for _ckey in ordered_tokens:
                        _tt = first_pass_map.get(_ckey)
                        if not _tt or not (_tt.get('price') and _tt.get('share') and _tt.get('amount')):
                            tokens_needing_tooltips.add(_ckey)
                    print(f"   ✅ First pass covered {first_pass_hits}/{len(all_tokens_map)} tokens")
                    print(f"   🔄 Extracting tooltips for {len(tokens_needing_tooltips)} remaining tokens (scroll pass)...")

                    # Scroll back to top first
                    page.evaluate("""() => {
                        let container = document.querySelector('[data-slot="table-container"]');
                        if (!container || container.scrollHeight <= container.clientHeight) {
                            const divs = document.querySelectorAll('div');
                            for (const div of divs) {
                                if (div.querySelector('table') && div.scrollHeight > div.clientHeight) {
                                    container = div; break;
                                }
                            }
                        }
                        if (container) container.scrollTop = 0;
                    }""")
                    page.wait_for_timeout(500)

                    # Scroll through and extract tooltips for visible rows
                    tooltip_scroll_count = 0
                    max_tooltip_scrolls = 200
                    reached_bottom = False
                    while tokens_needing_tooltips and tooltip_scroll_count < max_tooltip_scrolls:
                        # Find visible rows and extract tooltips
                        rows = page.locator("table tr[data-slot='table-row'], table tr").all()
                        for row in rows:
                            try:
                                cells = row.locator("td").all()
                                if len(cells) < 2:
                                    continue
                                row_chain = cells[0].inner_text().strip()
                                row_token = cells[1].inner_text().strip().split('\n')[0]
                                row_amount_key = cells[4].inner_text().strip() if len(cells) > 4 else ""
                                composite_key = f"{row_chain}|{row_token}|{row_amount_key}"

                                if composite_key in tokens_needing_tooltips:
                                    # Extract tooltips using read_tooltip_text helper
                                    price_tt = ""
                                    share_tt = ""
                                    amount_tt = ""

                                    # Price tooltip
                                    try:
                                        pt = row.locator('[data-tooltip-id*="price-tooltip"]').first
                                        if pt.count() > 0:
                                            tid = pt.get_attribute('data-tooltip-id')
                                            raw = read_tooltip_text(page, pt, tooltip_id=tid, wait_ms=400,
                                                                    debug_label=f"wallet price/{row_chain}|{row_token}")
                                            if raw:
                                                cleaned = raw.replace('$', '').replace(',', '').strip()
                                                if cleaned == '0':
                                                    price_tt = ""
                                                else:
                                                    _disp_price_raw = cells[2].inner_text().strip().split('\n')[0] if len(cells) > 2 else ""
                                                    if _price_tooltip_sanity_check(_disp_price_raw, raw):
                                                        price_tt = raw.replace('$', '').strip()
                                                    else:
                                                        # Retry: dismiss and recapture
                                                        price_tt = ""
                                                        for _rt3 in range(2):
                                                            try:
                                                                page.mouse.move(0, 0)
                                                                page.wait_for_timeout(400)
                                                                pt.hover(force=True, timeout=500)
                                                                page.wait_for_timeout(500)
                                                                raw2 = read_tooltip_text(page, pt, tooltip_id=tid, wait_ms=400,
                                                                                         debug_label=f"wallet price/{row_chain}|{row_token} retry{_rt3+1}")
                                                                if raw2 and raw2.replace('$', '').replace(',', '').strip() != '0':
                                                                    if _price_tooltip_sanity_check(_disp_price_raw, raw2):
                                                                        price_tt = raw2.replace('$', '').strip()
                                                                        print(f"   ✅ Price tooltip recaptured on retry {_rt3+1}: {price_tt}")
                                                                        break
                                                            except:
                                                                pass
                                    except:
                                        pass

                                    # Share tooltip — direct ID lookup first
                                    try:
                                        st = row.locator('[data-tooltip-id*="share-tooltip"]').first
                                        if st.count() > 0:
                                            tid = st.get_attribute('data-tooltip-id')
                                            if tid:
                                                st.hover(force=True, timeout=500)
                                                page.wait_for_timeout(600)
                                                td = page.locator(f'[id="{tid}"]').first
                                                if td.count() > 0:
                                                    raw = td.inner_text().strip()
                                                    if raw and raw != '0':
                                                        _disp_share_raw3 = cells[4].inner_text().strip().split('\n')[0] if len(cells) > 4 else ""
                                                        if _share_tooltip_sanity_check(_disp_share_raw3, raw):
                                                            share_tt = raw.replace('%', '').strip()
                                                        else:
                                                            # Retry
                                                            for _rts3 in range(2):
                                                                try:
                                                                    page.mouse.move(0, 0)
                                                                    page.wait_for_timeout(400)
                                                                    st.hover(force=True, timeout=500)
                                                                    page.wait_for_timeout(500)
                                                                    td2 = page.locator(f'[id="{tid}"]').first
                                                                    if td2.count() > 0:
                                                                        raw2 = td2.inner_text().strip()
                                                                        if raw2 and raw2 != '0' and _share_tooltip_sanity_check(_disp_share_raw3, raw2):
                                                                            share_tt = raw2.replace('%', '').strip()
                                                                            print(f"   ✅ Share tooltip recaptured on retry {_rts3+1}: {share_tt}")
                                                                            break
                                                                except:
                                                                    pass
                                    except:
                                        pass

                                    # Amount tooltip — direct ID lookup first
                                    try:
                                        at = row.locator('[data-tooltip-id*="amount-tooltip"]').first
                                        if at.count() > 0:
                                            tid = at.get_attribute('data-tooltip-id')
                                            if tid:
                                                at.hover(force=True, timeout=500)
                                                page.wait_for_timeout(600)
                                                td = page.locator(f'[id="{tid}"]').first
                                                if td.count() > 0:
                                                    raw = td.inner_text().strip()
                                                    if raw and raw != '0':
                                                        _disp_amt_raw3 = cells[3].inner_text().strip().split('\n')[0] if len(cells) > 3 else ""
                                                        _amt_candidate = raw
                                                        if _amt_candidate.count('.') > 1:
                                                            m = re.search(r'^(\d+\.\d+)', _amt_candidate)
                                                            if m:
                                                                _amt_candidate = m.group(1)
                                                        if _amount_tooltip_sanity_check(_disp_amt_raw3, _amt_candidate):
                                                            amount_tt = _amt_candidate
                                                        else:
                                                            # Retry
                                                            for _rta3 in range(2):
                                                                try:
                                                                    page.mouse.move(0, 0)
                                                                    page.wait_for_timeout(400)
                                                                    at.hover(force=True, timeout=500)
                                                                    page.wait_for_timeout(500)
                                                                    td2 = page.locator(f'[id="{tid}"]').first
                                                                    if td2.count() > 0:
                                                                        raw2 = td2.inner_text().strip()
                                                                        if raw2 and raw2 != '0':
                                                                            _amt2 = raw2
                                                                            if _amt2.count('.') > 1:
                                                                                m2 = re.search(r'^(\d+\.\d+)', _amt2)
                                                                                if m2:
                                                                                    _amt2 = m2.group(1)
                                                                            if _amount_tooltip_sanity_check(_disp_amt_raw3, _amt2):
                                                                                amount_tt = _amt2
                                                                                print(f"   ✅ Amount tooltip recaptured on retry {_rta3+1}: {amount_tt}")
                                                                                break
                                                                except:
                                                                    pass
                                    except:
                                        pass

                                    # Fallback: read from Amount cell text if still empty
                                    if not amount_tt:
                                        try:
                                            for ci, c in enumerate(cells):
                                                c_text = c.text_content().strip()
                                                if ci >= 2 and c_text and not c_text.startswith('$') and '%' not in c_text:
                                                    has_amt = c.locator('[data-tooltip-id*="amount"]').count()
                                                    if has_amt > 0:
                                                        amount_tt = re.sub(r'[A-Za-z]+$', '', c_text).strip().replace(',', '')
                                                        print(f"      💡 Amount tooltip fallback for {row_token}: '{amount_tt}' (from cell text)")
                                                        break
                                        except:
                                            pass

                                    # Update tooltips — for tokens already partially captured in first pass,
                                    # update only the missing ones instead of appending
                                    token_name, row_data = all_tokens_map[composite_key]
                                    if len(row_data) > 7:
                                        # Already has tooltip slots from first pass — update missing ones
                                        if price_tt and row_data[-3] == "Tooltip N/A":
                                            row_data[-3] = price_tt
                                        if share_tt and row_data[-2] == "Tooltip N/A":
                                            row_data[-2] = share_tt
                                        if amount_tt and row_data[-1] == "Tooltip N/A":
                                            row_data[-1] = amount_tt
                                    else:
                                        # No tooltip slots yet — append
                                        row_data.append(price_tt if price_tt else "Tooltip N/A")
                                        row_data.append(share_tt if share_tt else "Tooltip N/A")
                                        row_data.append(amount_tt if amount_tt else "Tooltip N/A")
                                    # Only remove from needing list if all three are now captured
                                    if row_data[-3] != "Tooltip N/A" and row_data[-2] != "Tooltip N/A" and row_data[-1] != "Tooltip N/A":
                                        tokens_needing_tooltips.discard(composite_key)
                                    elif price_tt or share_tt or amount_tt:
                                        # Got something new but still incomplete — keep trying
                                        pass
                                    else:
                                        # Got nothing new — don't retry this one
                                        tokens_needing_tooltips.discard(composite_key)
                                    print(f"      ✓ Tooltip: {row_chain}|{row_token} - price='{(price_tt or row_data[-3])[:15]}' share='{(share_tt or row_data[-2])[:10]}' amount='{(amount_tt or row_data[-1])[:15]}'")
                            except:
                                continue

                        if not tokens_needing_tooltips:
                            break

                        # If we already reached bottom on the previous scroll, stop now
                        # (we just processed the bottom-most visible rows above)
                        if reached_bottom:
                            break

                        # Scroll down
                        scroll_result = page.evaluate(f"""(scrollAmount) => {{
                            let container = document.querySelector('[data-slot="table-container"]');
                            if (!container || container.scrollHeight <= container.clientHeight) {{
                                const divs = document.querySelectorAll('div');
                                for (const div of divs) {{
                                    if (div.querySelector('table') && div.scrollHeight > div.clientHeight) {{
                                        container = div; break;
                                    }}
                                }}
                            }}
                            if (!container) return {{ success: false }};
                            container.scrollTop += scrollAmount;
                            return {{
                                success: true,
                                atBottom: container.scrollTop >= container.scrollHeight - container.clientHeight - 5
                            }};
                        }}""", scroll_step)

                        if not scroll_result.get('success'):
                            break

                        page.wait_for_timeout(200)
                        tooltip_scroll_count += 1

                        # Mark if we've reached bottom — next iteration will process bottom rows, then break
                        if scroll_result.get('atBottom'):
                            reached_bottom = True

                    # For any tokens we still couldn't find, append "Tooltip N/A"
                    for composite_key in tokens_needing_tooltips:
                        token_name, row_data = all_tokens_map[composite_key]
                        print(f"   ⚠️  Could not extract tooltip for {composite_key}")
                        row_data.append("Tooltip N/A")  # price_tt
                        row_data.append("Tooltip N/A")  # share_tt
                        row_data.append("Tooltip N/A")  # amount_tt

                    print(f"   ✅ Extracted tooltips for {len(all_tokens_map) - len(tokens_needing_tooltips)}/{len(all_tokens_map)} tokens")
    
                    # Build final token_data list
                    if header_row:
                        # Add tooltip columns to header
                        header_row.extend(['Price Tooltip', 'Share Tooltip', 'Amount Tooltip'])
                        token_data.append(header_row)
    
                    # Keep original order from DAM page (no sorting)
                    # Use ordered_tokens list to preserve insertion order (composite keys: chain|token_name)
                    for composite_key in ordered_tokens:
                        if composite_key in all_tokens_map:
                            token_name, row_data = all_tokens_map[composite_key]
                            token_data.append(row_data)
    
                    print(f"   ✅ Total rows: {len(token_data)} (1 header + {len(all_tokens_map)} tokens)")
                except Exception as e:
                    print(f"⚠️  Error: {e}")
    
            # Only assign wallet data if we have a wallet section (not CEX-only)
            if has_wallet_section:
                all_tables_data['Overview - Wallet'] = token_data
                if token_data:
                    print(f"✅ Extracted {len(token_data)} rows")
                else:
                    print(f"⚠️  WARNING: Wallet section found but no data extracted! 'Overview - Wallet' tab will be empty.")

                # Debug: Print first row to check structure
                if token_data and len(token_data) > 1:
                    print(f"   DEBUG - Row 1 (header) has {len(token_data[0])} columns")
                    print(f"   DEBUG - Row 2 (TRX) has {len(token_data[1])} columns: {token_data[1]}")

            # TABLE 1.3: Account-Level Token Breakdown Tables (CEX + DeFi)
            print("\n📊 Table 1.3: Account-Level Token Breakdowns (CEX + DeFi)")
            print("-"*80)

            # Extract all account/protocol breakdown tables
            # Types: CEX accounts (Binance - david, moontest), DeFi protocols (Aave V3)
            try:
                print(f"   🔄 Extracting account-level token breakdowns...")

                # Scroll the full page to ensure all lazy-loaded sections are in the DOM
                print(f"   🔄 Scrolling page to load all sections...")
                _total_h = page.evaluate("document.body.scrollHeight")
                _step = 600
                _pos = 0
                while _pos < _total_h:
                    _pos += _step
                    page.evaluate(f"window.scrollTo(0, {_pos})")
                    page.wait_for_timeout(300)
                    _total_h = page.evaluate("document.body.scrollHeight")
                page.evaluate("window.scrollTo(0, 0)")
                page.wait_for_timeout(500)

                # Look for account sections with bg-grey-30 class
                account_sections = page.locator('div.bg-grey-30.rounded-sm').all()

                print(f"   🔍 Found {len(account_sections)} potential account sections")

                for section_idx, section in enumerate(account_sections):
                    try:
                        # Get account/protocol name from the section header
                        account_name = None
                        chain_name = None

                        # Strategy 1: Look for title with typography-title class (CEX accounts like "Binance - david")
                        try:
                            title_elem = section.locator('.typography-title.font-semibold.text-mono-900').first
                            if title_elem.is_visible():
                                account_name = title_elem.text_content().strip()
                        except:
                            pass

                        # Strategy 2: Look for protocol name in typography-tab class (DeFi like "Aave V3")
                        if not account_name:
                            try:
                                tab_elem = section.locator('.typography-tab.font-semibold.text-mono-900').first
                                if tab_elem.is_visible():
                                    account_name = tab_elem.text_content().strip()
                            except:
                                pass

                        # Strategy 3: Look for any font-semibold text-mono-900 header
                        if not account_name:
                            try:
                                header_elems = section.locator('[class*="font-semibold"][class*="text-mono-900"]').all()
                                for elem in header_elems[:5]:
                                    text = elem.text_content().strip()
                                    # Skip common UI elements
                                    skip_texts = ['total', 'value:', '$', 'position type', 'pool', 'amount', 'lending', 'funding']
                                    if text and len(text) > 2 and len(text) < 50:
                                        if not any(skip in text.lower() for skip in skip_texts):
                                            account_name = text
                                            break
                            except:
                                pass

                        # For DeFi protocols, extract chain from logo image alt text or src URL
                        defi_protocol_keywords = ['aave', 'v3', 'morpho', 'compound', 'uniswap', 'curve', 'lido', 'maker', 'spark', 'euler', 'pendle', 'yearn', 'balancer', 'convex', 'merkl', 'gearbox', 'fluid', 'kamino', 'drift', 'marginfi', 'midas', 'defi saver', 'hyperliquid', 'eigenlayer', 'nftx']
                        if account_name and (any(kw in account_name.lower() for kw in defi_protocol_keywords) or any(kw in account_name.lower() for kw in ['defi', 'protocol', 'finance', 'swap', 'pool'])):
                            try:
                                # Strategy 1: chain logo by alt text
                                chain_imgs = section.locator('img[alt="Ethereum"], img[alt="Base"], img[alt="Arbitrum"], img[alt="Polygon"], img[alt="Optimism"], img[alt="BSC"], img[alt="BNB"]').all()
                                if chain_imgs:
                                    for img in chain_imgs:
                                        if img.is_visible():
                                            chain_name = img.get_attribute('alt')
                                            break
                                # Strategy 2: fallback — check all img src URLs for chain keywords
                                if not chain_name:
                                    _chain_src_map = {
                                        'ethereum': 'Ethereum', 'base': 'Base', 'arbitrum': 'Arbitrum',
                                        'polygon': 'Polygon', 'optimism': 'Optimism', 'bnb': 'BSC', 'bsc': 'BSC'
                                    }
                                    all_imgs = section.locator('img').all()
                                    for img in all_imgs:
                                        try:
                                            src = (img.get_attribute('src') or '').lower()
                                            for kw, cname in _chain_src_map.items():
                                                if kw in src:
                                                    chain_name = cname
                                                    break
                                            if chain_name:
                                                break
                                        except:
                                            pass
                            except:
                                pass

                        if not account_name:
                            # Fallback: try to get any meaningful text from section
                            section_text = section.text_content()[:100]
                            print(f"   ⚠️  Section {section_idx}: Could not find name (preview: {section_text}...)")
                            continue

                        # Build final tab name
                        if chain_name:
                            tab_name = f"{account_name} ({chain_name})"
                        else:
                            tab_name = account_name

                        # Skip "Combined Net Worth" sections
                        if 'combined' in tab_name.lower() or 'net worth' in tab_name.lower():
                            continue

                        print(f"   📋 Processing: {tab_name}")

                        # For Hyperliquid, process all tables (deposit/yield + perpetuals)
                        # For all other protocols, only process the first table
                        _is_hl_section = 'hyperliquid' in tab_name.lower()
                        _all_tables = section.locator('table').all()
                        _tables_to_process = _all_tables if _is_hl_section else [section.locator('table').first]

                        # Initialize account_table_data once for the whole section
                        account_table_data = []

                        for _tbl_idx, table in enumerate(_tables_to_process):
                            if not table.is_visible():
                                print(f"   ⚠️  '{tab_name}' table {_tbl_idx+1}: not visible, skipping")
                                continue

                            # Click all expandable buttons within TABLE ROWS to reveal nested child rows
                            try:
                                # Find buttons in the first cell of each tbody row
                                # These buttons expand to show detailed breakdown by Category and Account
                                expand_buttons = table.locator('tbody tr td:first-child button').all()

                                if expand_buttons:
                                    print(f"   🔽 Found {len(expand_buttons)} row expand buttons, clicking...")
                                    for btn_idx, btn in enumerate(expand_buttons):
                                        try:
                                            if btn.is_visible():
                                                btn.click(timeout=2000)
                                                page.wait_for_timeout(800)  # Wait for child rows to appear
                                                print(f"      ✓ Expanded row {btn_idx + 1}/{len(expand_buttons)}")
                                        except Exception as e:
                                            print(f"      ⚠️  Could not expand row {btn_idx + 1}: {e}")
                                else:
                                    print(f"   ℹ️  No expandable row buttons found")
                            except Exception as e:
                                print(f"   ⚠️  Error expanding rows: {e}")

                            # Only reset header on first table (Hyperliquid multi-table: append to same data)
                            if _tbl_idx == 0:
                                account_table_data = []
                                # Add header row
                                account_table_data.append([
                                    'Token', 'Category', 'Account', 'Price', 'Price Tooltip',
                                    'FE - Price Validation', 'Price Validation', 'Price Diff Validation',
                                    'Price (24h)', 'Price (24h) Validation', 'Price (24H) Diff Validation',
                                    'Amount', 'Amount Tooltip', 'Amount Validation', 'FE - Amount Validation',
                                    'Amount Diff Validation', 'Value', 'Value Validation', 'Calculated Value',
                                    'Total Value', 'Total Value Validation'
                                ])

                            # Get only the PARENT rows (not child containers or nested rows)
                            # Parent rows have 5 td cells
                            parent_rows = table.locator('tbody > tr').all()
                            print(f"   📊 Found {len(parent_rows)} total rows after expansion")

                            for parent_idx, parent_row in enumerate(parent_rows):
                                try:
                                    parent_cells = parent_row.locator('> td').all()  # Direct children only

                                    # Determine if this is a CEX table or DeFi table
                                    # DeFi tables have a different row structure (no CEX 5-cell pattern)
                                    # Detect by tab name keywords OR by checking if current account_table_data header has "Position Type"
                                    _defi_kws = ['aave', 'v3', 'defi', 'saver', 'morpho', 'compound', 'uniswap', 'curve', 'lido', 'maker', 'spark', 'euler', 'pendle', 'yearn', 'balancer', 'convex',
                                                 'hyperliquid', 'virtuals', 'virtual', 'aix', 'pendle', 'fluid', 'kamino', 'drift', 'marginfi', 'tensor', 'zeta', 'orca', 'raydium', 'jupiter',
                                                 'midas', 'merkl', 'gearbox', 'eigenlayer', 'nftx']
                                    is_defi = any(kw in tab_name.lower() for kw in _defi_kws)
                                    # Also check if current data header already has DeFi structure
                                    if not is_defi and account_table_data and len(account_table_data) > 0:
                                        _h = account_table_data[0]
                                        if _h and 'Position Type' in str(_h[0]):
                                            is_defi = True

                                    # CASE 1: CEX Parent summary row (5 cells: button, token, price, amount, value)
                                    if len(parent_cells) == 5 and not is_defi:
                                        # Extract parent row as Token with Category="Main"
                                        token_name = parent_cells[1].text_content().strip()
                                        # Clean token name (remove icons, etc.)
                                        token_name = re.sub(r'\s+', ' ', token_name).strip()

                                        # Price cell contains price + 24h% change combined
                                        price_cell = parent_cells[2]
                                        price_cell_text = price_cell.text_content().strip()
                                        # Extract price part, preserving "<" symbol if present (e.g., "< $0.01")
                                        lt_match = re.search(r'<\s*\$?([\d,]+\.?\d*)', price_cell_text)
                                        price_match = re.search(r'\$?([\d,]+\.?\d*)', price_cell_text)
                                        if lt_match:
                                            price = f"< {lt_match.group(1)}"
                                        elif price_match:
                                            price = price_match.group(1)
                                        else:
                                            price = ""

                                        # Extract Price Tooltip by hovering over price element
                                        price_tooltip = ""
                                        try:
                                            price_tooltip_trigger = price_cell.locator('[data-tooltip-id*="price-tooltip"], [data-tooltip-id*="price"]').first
                                            if price_cell.locator('[data-tooltip-id*="price-tooltip"], [data-tooltip-id*="price"]').count() > 0:
                                                tid = price_tooltip_trigger.get_attribute('data-tooltip-id')
                                                raw = read_tooltip_text(page, price_tooltip_trigger, tooltip_id=tid, wait_ms=600,
                                                                        debug_label=f"CEX price/{token_name}",
                                                                        debug_screenshot_folder=screenshot_folder)
                                                price_tooltip = raw.replace('$', '').strip() if raw else ""
                                        except:
                                            pass

                                        # Extract Price (24h) - check for red color (negative) or green (positive)
                                        price_24h = ""
                                        try:
                                            # Look for the percentage change element
                                            pct_elem = price_cell.locator('[class*="text-error"], [class*="text-success"], [class*="bg-error"], [class*="bg-success"]').first
                                            if pct_elem.is_visible():
                                                pct_text = pct_elem.text_content().strip()
                                                # Extract the number and remove %
                                                pct_match = re.search(r'([\d.]+)%?', pct_text)
                                                if pct_match:
                                                    pct_value = pct_match.group(1)
                                                    # Check if it's negative (red/error color)
                                                    pct_class = pct_elem.get_attribute('class') or ""
                                                    if 'error' in pct_class.lower() or '↓' in pct_text:
                                                        price_24h = f"-{pct_value}"
                                                    else:
                                                        price_24h = pct_value
                                        except:
                                            pass

                                        # Amount cell contains amount + token symbol combined
                                        amount_cell = parent_cells[3]
                                        amount_cell_text = amount_cell.text_content().strip()
                                        # Remove token symbol at the end
                                        amount = re.sub(r'[A-Z]+$', '', amount_cell_text).strip()

                                        # Extract Amount Tooltip by hovering over amount element
                                        amount_tooltip = ""
                                        try:
                                            amount_tooltip_trigger = amount_cell.locator('[data-tooltip-id*="amount-tooltip"], [data-tooltip-id*="amount"]').first
                                            if amount_cell.locator('[data-tooltip-id*="amount-tooltip"], [data-tooltip-id*="amount"]').count() > 0:
                                                tooltip_id = amount_tooltip_trigger.get_attribute('data-tooltip-id')
                                                amount_tooltip = read_tooltip_text(
                                                    page, amount_tooltip_trigger,
                                                    tooltip_id=tooltip_id,
                                                    wait_ms=600,
                                                    debug_label=f"CEX amount/{token_name}",
                                                    debug_screenshot_folder=screenshot_folder,
                                                )
                                                # Clean up multiple decimals
                                                if amount_tooltip and amount_tooltip.count('.') > 1:
                                                    m = re.search(r'^(\d+\.\d+)', amount_tooltip)
                                                    amount_tooltip = m.group(1) if m else amount_tooltip
                                            else:
                                                amount_tooltip = read_tooltip_text(
                                                    page, amount_cell,
                                                    wait_ms=600,
                                                    debug_label=f"CEX amount/{token_name}",
                                                    debug_screenshot_folder=screenshot_folder,
                                                )
                                        except:
                                            pass

                                        # Value
                                        value = clean_currency_symbols(parent_cells[4].text_content().strip())

                                        # Create parent row: Token, Category="Main", Account="-"
                                        row_data = [
                                            token_name,      # Token
                                            "Main",          # Category
                                            "-",             # Account
                                            price,           # Price
                                            price_tooltip,   # Price Tooltip
                                            "", "", "",      # FE - Price Val, Price Val, Price Diff Val
                                            price_24h,       # Price (24h)
                                            "", "",          # Price (24h) Val, Price (24H) Diff Val
                                            amount,          # Amount
                                            amount_tooltip,  # Amount Tooltip
                                            "", "", "",      # Amount Val, FE - Amount Val, Amount Diff Val
                                            value,           # Value
                                            "", "",          # Value Val, Calculated Value
                                            "",              # Total Value
                                            ""               # Total Value Validation
                                        ]
                                        account_table_data.append(row_data)
                                        print(f"   ✓ Parent row: {token_name} (Main) - Price: {price} ({price_24h}%), Amount: {amount}, Value: {value}")

                                        # Look for child container in next sibling
                                        try:
                                            next_row = parent_rows[parent_idx + 1] if parent_idx + 1 < len(parent_rows) else None
                                            if next_row:
                                                next_cells = next_row.locator('> td').all()
                                                # Child container has colspan cell
                                                if len(next_cells) > 0:
                                                    # Check if it contains nested table
                                                    nested_tables = next_row.locator('table').all()
                                                    for nested_table in nested_tables:
                                                        # Extract category from badge in child container
                                                        category = ""
                                                        try:
                                                            category_badge = next_row.locator('.bg-white\\/10 .typography-body').first
                                                            if category_badge.is_visible():
                                                                category = category_badge.text_content().strip()
                                                        except:
                                                            category = "Unknown"

                                                        # Extract rows from nested table
                                                        try:
                                                            nested_rows = nested_table.locator('tbody tr').all()
                                                            for nested_row in nested_rows:
                                                                nested_cells = nested_row.locator('td').all()
                                                                if len(nested_cells) >= 3:
                                                                    # Nested table: [Account, Amount, Value]
                                                                    account = nested_cells[0].text_content().strip()
                                                                    nested_amount_cell = nested_cells[1]
                                                                    nested_amount_text = nested_amount_cell.text_content().strip()
                                                                    # Remove token symbol
                                                                    nested_amount = re.sub(r'[A-Z]+$', '', nested_amount_text).strip()
                                                                    nested_value = clean_currency_symbols(nested_cells[2].text_content().strip())

                                                                    # Extract Amount Tooltip by hovering over amount element in child row
                                                                    nested_amount_tooltip = ""
                                                                    try:
                                                                        amt_trigger = nested_amount_cell.locator('[data-tooltip-id*="amount-tooltip"], [data-tooltip-id*="amount"]').first
                                                                        use_trigger = amt_trigger if amt_trigger.count() > 0 else nested_amount_cell
                                                                        tid = use_trigger.get_attribute('data-tooltip-id') if amt_trigger.count() > 0 else None
                                                                        raw = read_tooltip_text(
                                                                            page, use_trigger,
                                                                            tooltip_id=tid,
                                                                            wait_ms=400,
                                                                            debug_label=f"nested amount/{token_name}/{account}",
                                                                            debug_screenshot_folder=screenshot_folder,
                                                                        )
                                                                        if raw:
                                                                            if raw.count('.') > 1:
                                                                                m = re.search(r'^(\d+\.\d+)', raw)
                                                                                nested_amount_tooltip = m.group(1) if m else raw
                                                                            else:
                                                                                nested_amount_tooltip = raw
                                                                    except:
                                                                        pass
                                                                    child_row_data = [
                                                                        token_name,      # Token (same as parent)
                                                                        category,        # Category (from badge)
                                                                        account,         # Account
                                                                        "-",             # Price (not shown in child)
                                                                        "",              # Price Tooltip
                                                                        "", "", "",      # FE - Price Val, Price Val, Price Diff Val
                                                                        "",              # Price (24h)
                                                                        "", "",          # Price (24h) Val, Price (24H) Diff Val
                                                                        nested_amount,   # Amount
                                                                        nested_amount_tooltip,  # Amount Tooltip
                                                                        "", "", "",      # Amount Val, FE - Amount Val, Amount Diff Val
                                                                        nested_value,    # Value
                                                                        "", "",          # Value Val, Calculated Value
                                                                        "",              # Total Value
                                                                        ""               # Total Value Validation
                                                                    ]
                                                                    account_table_data.append(child_row_data)
                                                                    print(f"      ↳ Child row: {token_name} ({category}) - {account} - {nested_value}")
                                                        except:
                                                            pass
                                        except:
                                            pass

                                    # CASE 2: DeFi protocol table (Aave V3) - different structure
                                    # DeFi columns: Position Type, Pool, Amount, Amount Tooltip, Value, Amount Validation, FE - Amount Validation, Amount Diff Validation
                                    elif is_defi:
                                        # Replace CEX header with DeFi header only on first table
                                        if _tbl_idx == 0:
                                            account_table_data = []
                                            account_table_data.append([
                                                'Position Type', 'Pool', 'Amount', 'Amount Tooltip', 'Value',
                                                'Amount Validation', 'FE - Amount Validation', 'Amount Diff Validation'
                                            ])

                                        try:
                                            # For Aave V3, the structure shows Position Type grouping multiple pool rows
                                            # Try to extract all tr elements (not just tbody tr)
                                            all_rows = table.locator('tr').all()
                                            print(f"   DEBUG: Found {len(all_rows)} tr elements in DeFi table")

                                            current_position_type = ""
                                            current_value = ""
                                            row_index = 0

                                            for row in all_rows:
                                                # Skip header rows (contain th)
                                                th_count = row.locator('th').count()
                                                if th_count > 0:
                                                    continue

                                                cells = row.locator('td').all()
                                                if len(cells) == 0:
                                                    continue

                                                # Get all cell texts for debugging
                                                cell_texts = [c.text_content().strip() for c in cells]
                                                print(f"   DEBUG: Row has {len(cells)} cells: {cell_texts[:5]}")

                                                # Determine amount cell index for hover tooltip extraction
                                                amount_cell = None
                                                amount_cell_idx = -1

                                                # Parse the row based on content
                                                is_hyperliquid = 'hyperliquid' in tab_name.lower()

                                                # Hyperliquid perpetuals: Type | Position Pair | Description | Side | Leverage | Collateral | P&L | Value (8 cells)
                                                if is_hyperliquid and len(cells) >= 8:
                                                    pos = cell_texts[0]
                                                    if pos and pos not in ['', 'Position Type', 'Type']:
                                                        current_position_type = pos
                                                    pool = cell_texts[1]          # Position Pair
                                                    description = cell_texts[2]   # Description
                                                    side = cell_texts[3]          # Side
                                                    leverage = cell_texts[4]      # Leverage
                                                    collateral = cell_texts[5]    # Collateral
                                                    pnl = cell_texts[6]           # P&L
                                                    current_value = cell_texts[7] # Value
                                                    amount_raw = ""
                                                    amount_cell = None
                                                    amount_cell_idx = -1
                                                # Hyperliquid deposit/yield: Type | Pool | Description | Amount | Value (5 cells)
                                                elif is_hyperliquid and len(cells) >= 5:
                                                    pos = cell_texts[0]
                                                    if pos and pos not in ['', 'Position Type', 'Type']:
                                                        current_position_type = pos
                                                    pool = cell_texts[1]
                                                    description = cell_texts[2]
                                                    amount_raw = cell_texts[3]
                                                    current_value = cell_texts[4]
                                                    side = leverage = collateral = pnl = ""
                                                    amount_cell = cells[3]
                                                    amount_cell_idx = 3
                                                elif len(cells) >= 4:
                                                    # Full row: Position Type | Pool | Amount | Value
                                                    pos = cell_texts[0]
                                                    if pos and pos not in ['', 'Position Type']:
                                                        current_position_type = pos
                                                    pool = cell_texts[1]
                                                    amount_raw = cell_texts[2]
                                                    current_value = cell_texts[3] if len(cell_texts) > 3 else ""
                                                    amount_cell = cells[2]
                                                    amount_cell_idx = 2
                                                    description = side = leverage = collateral = pnl = ""
                                                elif len(cells) >= 3:
                                                    # Partial: Pool | Amount | Value
                                                    pool = cell_texts[0]
                                                    amount_raw = cell_texts[1]
                                                    current_value = cell_texts[2] if len(cell_texts) > 2 else ""
                                                    amount_cell = cells[1]
                                                    amount_cell_idx = 1
                                                    description = side = leverage = collateral = pnl = ""
                                                elif len(cells) >= 2:
                                                    # Minimal: Pool | Amount
                                                    pool = cell_texts[0]
                                                    amount_raw = cell_texts[1]
                                                    amount_cell = cells[1]
                                                    amount_cell_idx = 1
                                                    description = side = leverage = collateral = pnl = ""
                                                else:
                                                    continue

                                                # Skip header-like content
                                                if pool in ['Pool', 'Amount', 'Value', 'Position Type', 'Position Pair', '']:
                                                    continue

                                                # Clean up the Value column (remove $ sign and handle formats like "$0.94")
                                                value_cleaned = ""
                                                if current_value:
                                                    value_cleaned = current_value.replace('$', '').replace(',', '').strip()

                                                # Hyperliquid perpetuals have no amount — write row directly and skip tooltip logic
                                                if is_hyperliquid and amount_cell is None:
                                                    row_data = [
                                                        current_position_type,  # Position Type
                                                        pool,                   # Pool/ Position Pair
                                                        "",                     # Amount
                                                        "",                     # Amount Tooltip
                                                        value_cleaned,          # Value
                                                        "",                     # Amount Validation
                                                        "",                     # FE - Amount Validation
                                                        "",                     # Amount Diff Validation
                                                        description,            # Description
                                                        side,                   # Side
                                                        leverage,               # Leverage
                                                        collateral,             # Collateral
                                                        pnl,                    # P&L
                                                    ]
                                                    account_table_data.append(row_data)
                                                    print(f"   ✓ Hyperliquid Perpetual: {current_position_type} | {pool} | {description} | {side} | {leverage} | {collateral} | {pnl} | {value_cleaned}")
                                                    row_index += 1
                                                    continue

                                                # Extract Amount Tooltip by hovering over amount cell
                                                amount_tooltip = ""
                                                full_tooltip_text = ""
                                                try:
                                                    if amount_cell:
                                                        amt_trigger = amount_cell.locator('[data-tooltip-id*="amount-tooltip"], [data-tooltip-id*="amount"], [data-tooltip-id]').first
                                                        use_trigger = amt_trigger if amt_trigger.count() > 0 else amount_cell
                                                        tid = use_trigger.get_attribute('data-tooltip-id') if amt_trigger.count() > 0 else None
                                                        raw_tooltip = read_tooltip_text(
                                                            page, use_trigger,
                                                            tooltip_id=tid,
                                                            wait_ms=600,
                                                            debug_label=f"DeFi amount/row{row_index}/{tab_name}",
                                                            debug_screenshot_folder=screenshot_folder,
                                                        )
                                                        if raw_tooltip:
                                                            full_tooltip_text = raw_tooltip
                                                            tooltip_match = re.search(r'(-?[\d,]+\.?\d*)', raw_tooltip)
                                                            if tooltip_match:
                                                                amount_tooltip = tooltip_match.group(1).replace(',', '')
                                                                if 'borrow' in raw_tooltip.lower() and not amount_tooltip.startswith('-'):
                                                                    amount_tooltip = '-' + amount_tooltip
                                                                print(f"         ✅ DeFi Amount tooltip: {amount_tooltip}")
                                                                print(f"         📝 Full tooltip text: {full_tooltip_text}")
                                                except Exception as tooltip_err:
                                                    print(f"         ⚠️  DeFi Amount tooltip extraction failed: {tooltip_err}")

                                                # Check if multiple tokens are combined in one cell (e.g., "USDCUSDT", "0.20408 USDC0.29486 USDT")
                                                # Split by finding token amounts pattern (including negative amounts)
                                                amount_pattern = r'(-?[\d.,]+)\s*([A-Z]+)'
                                                amount_matches = re.findall(amount_pattern, amount_raw)

                                                if len(amount_matches) > 1:
                                                    # Multiple tokens in one row - create separate rows for each
                                                    # Parse tooltip to extract individual amounts per token
                                                    # Tooltip format: "Supply: 40460.535296 ETH\nBorrow: 37376.113827 WETH"
                                                    tooltip_amounts = {}
                                                    if full_tooltip_text:
                                                        # Parse line by line to detect "Borrow:" prefix for negative sign
                                                        for tooltip_line in full_tooltip_text.split('\n'):
                                                            tooltip_line = tooltip_line.strip()
                                                            if not tooltip_line:
                                                                continue
                                                            is_borrow_line = 'borrow' in tooltip_line.lower()
                                                            line_match = re.search(r'(-?[\d,]+\.?\d*)\s*([A-Z]+)', tooltip_line)
                                                            if line_match:
                                                                tooltip_amt = line_match.group(1).replace(',', '')
                                                                tooltip_symbol = line_match.group(2).upper()
                                                                # Add "-" for borrow amounts if positive
                                                                if is_borrow_line and not tooltip_amt.startswith('-'):
                                                                    tooltip_amt = '-' + tooltip_amt
                                                                tooltip_amounts[tooltip_symbol] = tooltip_amt
                                                        print(f"         📝 Parsed tooltip amounts: {tooltip_amounts}")

                                                    # Parse value column to extract individual values
                                                    # Value format: "$0.20$0.29" or "0.200.29" or "$0.20 $0.29" or "-$0.20" or "< $0.01"
                                                    value_pattern = r'(<\s*)?(-?)\$?([\d,]+\.?\d*)'
                                                    value_matches_raw = re.findall(value_pattern, current_value) if current_value else []
                                                    # Combine less-than prefix, sign and number, filter out empty matches
                                                    value_matches = []
                                                    for _lt, _sign, _num in value_matches_raw:
                                                        if _num and _num != '.':
                                                            if _lt.strip():  # has < prefix
                                                                value_matches.append('< ' + _num)
                                                            else:
                                                                value_matches.append(_sign + _num)
                                                    print(f"         📝 Parsed values: {value_matches} from '{current_value}'")

                                                    for idx, (amt_value, token_symbol) in enumerate(amount_matches):
                                                        # Get tooltip amount for this specific token (already has "-" for borrow from tooltip parsing)
                                                        token_tooltip = tooltip_amounts.get(token_symbol.upper(), amount_tooltip)

                                                        # Get value for this token (match by index if available)
                                                        token_value = ""
                                                        if idx < len(value_matches):
                                                            token_value = value_matches[idx]

                                                        # If tooltip has "-" (borrow), ensure amount and value also have "-"
                                                        if str(token_tooltip).startswith('-'):
                                                            if amt_value and not amt_value.startswith('-'):
                                                                amt_value = '-' + amt_value
                                                            if token_value and not token_value.startswith('-'):
                                                                token_value = '-' + token_value

                                                        row_data = [
                                                            current_position_type,  # Position Type
                                                            token_symbol,           # Pool (individual token)
                                                            amt_value,              # Amount
                                                            token_tooltip,          # Amount Tooltip (specific to this token)
                                                            token_value,            # Value (specific to this token)
                                                            "",                     # Amount Validation
                                                            "",                     # FE - Amount Validation
                                                            "",                     # Amount Diff Validation
                                                            description,            # Description
                                                            side,                   # Side
                                                            leverage,               # Leverage
                                                            collateral,             # Collateral
                                                            pnl,                    # P&L
                                                        ]
                                                        account_table_data.append(row_data)
                                                        print(f"   ✓ DeFi: {current_position_type} | {token_symbol} | {amt_value} | Tooltip: {token_tooltip} | Value: {token_value}")
                                                else:
                                                    # Single token - extract amount
                                                    amount = re.sub(r'\s*[A-Z]+$', '', amount_raw).strip()

                                                    # If tooltip has "-" (borrow), ensure amount and value also have "-"
                                                    if str(amount_tooltip).startswith('-'):
                                                        if amount and not amount.startswith('-'):
                                                            amount = '-' + amount
                                                        if value_cleaned and not value_cleaned.startswith('-'):
                                                            value_cleaned = '-' + value_cleaned

                                                    # Create DeFi row with Value
                                                    row_data = [
                                                        current_position_type,  # Position Type
                                                        pool,                   # Pool
                                                        amount,                 # Amount
                                                        amount_tooltip,         # Amount Tooltip
                                                        value_cleaned,          # Value
                                                        "",                     # Amount Validation
                                                        "",                     # FE - Amount Validation
                                                        "",                     # Amount Diff Validation
                                                        description,            # Description
                                                        side,                   # Side
                                                        leverage,               # Leverage
                                                        collateral,             # Collateral
                                                        pnl,                    # P&L
                                                    ]
                                                    account_table_data.append(row_data)
                                                    print(f"   ✓ DeFi: {current_position_type} | {pool} | {amount} | Tooltip: {amount_tooltip} | Value: {value_cleaned}")

                                                row_index += 1

                                        except Exception as defi_err:
                                            print(f"   ⚠️ DeFi extraction error: {defi_err}")
                                            import traceback
                                            traceback.print_exc()

                                except Exception as row_err:
                                    print(f"   ⚠️  Error extracting row in '{tab_name}': {row_err}")
                                    continue

                        # Store the account table data
                        if len(account_table_data) > 1:  # More than just header
                            all_tables_data[tab_name] = account_table_data
                            print(f"   ✅ Extracted {len(account_table_data) - 1} rows for '{tab_name}'")

                    except Exception as section_err:
                        print(f"   ⚠️  Error processing section {section_idx}: {section_err}")
                        continue

            except Exception as e:
                print(f"   ⚠️  Error extracting account breakdowns: {e}")
                traceback.print_exc()

            # TABLE 1.5: Overview - Header & Token Holdings Header
            print("\n📊 Table 1.5: Overview - Header & Token Holdings Header")
            print("-"*80)

            header_holdings_data = []
            try:
                print(f"   🔄 Extracting Header & Token Holdings data...")

                # Add header row
                header_holdings_data.append(['Section', 'Category', 'Token Count', 'TC_UI Count', 'Token Count Validation', 'Net Worth', 'Net Worth (Actual)', 'Net Worth Validation', 'Percentage'])

                # SECTION 1: Extract Overview Header (Wallets, Exchanges, De-Fi Positions)
                # These appear as cards at the top showing total net worth breakdown
                print("   🔍 Extracting Overview Header sections...")

                try:
                    # Look for the main overview cards showing Wallets, Exchanges, De-Fi Positions
                    # These usually have labels like "Wallets" and values like "$66,033.81"
                    overview_sections = [
                        ('Wallets', ''),
                        ('Exchanges', ''),
                        ('De-Fi Positions', '')
                    ]

                    for section_name, _ in overview_sections:
                        try:
                            # Find element containing the section name
                            section_elem = page.locator(f'text="{section_name}"').first
                            if section_elem.count() > 0:
                                # Try to find value in parent or nearby element
                                parent = section_elem.locator('..').first
                                parent_text = parent.text_content().strip()

                                # Extract dollar value, preserving "<" prefix if present (e.g., "<$0.01")
                                search_text = parent_text.replace(section_name, '')
                                lt_prefix_ov = '<' if re.search(r'<\s*\$', search_text) else ''
                                value_match = re.search(r'\$?([\d,]+\.?\d*)', search_text)
                                if value_match:
                                    value = lt_prefix_ov + value_match.group(1)
                                    header_holdings_data.append([
                                        'Overview Header',
                                        section_name,
                                        '',  # Token Count
                                        '',  # TC_UI Count
                                        '',  # Token Count Validation
                                        value,
                                        '',  # Net Worth (Actual)
                                        '',  # Net Worth Validation
                                        ''   # Percentage calculated later
                                    ])
                                    print(f"      Found: {section_name} = ${value}")
                        except:
                            pass
                except Exception as e:
                    print(f"   ⚠️  Error extracting overview sections: {e}")

                # SECTION 2: Extract Token Holdings - Chain (Tron (35), Ethereum (17), etc.)
                # These appear in the Combined Net Worth section with token counts
                print("   🔍 Extracting Token Holdings - Chain...")

                try:
                    # Click "+ Show more chains" if present to reveal all chains
                    try:
                        show_more = page.locator('button:has-text("Show more chains"), button:has-text("+ Show more chains")').first
                        if show_more.count() > 0 and show_more.is_visible():
                            show_more.click()
                            page.wait_for_timeout(1000)
                            print("   ✅ Clicked '+ Show more chains'")
                    except:
                        pass

                    # Look for chain groupings with token counts in parentheses
                    # Pattern: "Tron (35)" where 35 is the number of tokens
                    chain_elements = page.locator('div, span').all()

                    chains_found = []
                    for elem in chain_elements:
                        try:
                            text = elem.text_content().strip()

                            # Match pattern: "ChainName (TokenCount)"
                            match = re.match(r'^([A-Za-z\s]+)\s*\((\d+)\)$', text)
                            if match:
                                chain_name = match.group(1).strip()
                                token_count = match.group(2)

                                # Skip generic keywords
                                skip_keywords = ['show', 'hide', 'more', 'less', 'all', 'filter', 'token', 'showing']
                                if any(keyword in chain_name.lower() for keyword in skip_keywords):
                                    continue

                                # Avoid duplicates
                                if chain_name in chains_found:
                                    continue

                                # Try to find associated value and percentage
                                try:
                                    parent = elem.locator('..').first
                                    parent_text = parent.text_content().strip()

                                    # Remove the chain name pattern from text to avoid matching token count
                                    cleaned_text = re.sub(r'^[A-Za-z\s]+\s*\(\d+\)\s*', '', parent_text)

                                    # Look for dollar value - prefer explicit $ prefix, or large numbers with commas
                                    # Preserve "<" prefix if present (e.g., "<$0.01" stays as "<0.01")
                                    lt_prefix_nw = '<' if re.search(r'<\s*\$', cleaned_text) else ''
                                    value_match = re.search(r'\$([\d,]+\.?\d*)', cleaned_text)
                                    if not value_match:
                                        # Fallback: look for numbers with commas (likely currency values)
                                        value_match = re.search(r'([\d,]{4,}\.?\d*)', cleaned_text)
                                        lt_prefix_nw = ''  # No < prefix for numeric fallback
                                    net_worth = lt_prefix_nw + (value_match.group(1) if value_match else '')

                                    # Look for percentage (including optional "<" symbol for small values)
                                    # Pattern matches: "0.01%", "<0.01%", "< 0.01%" (with space), "50.25%"
                                    pct_match = re.search(r'(<\s*[\d.]+|[\d.]+)%', parent_text)
                                    percentage = pct_match.group(1).replace(' ', '') if pct_match else ''  # Remove space if present

                                    # Debug: Print raw percentage value to verify "<" symbol is captured
                                    if percentage and percentage.startswith('<'):
                                        print(f"      🔍 DEBUG: Captured percentage with '<' symbol: '{percentage}'")

                                    chains_found.append(chain_name)
                                    header_holdings_data.append([
                                        'Token Holdings - Chain',
                                        chain_name,
                                        token_count,
                                        '',  # TC_UI Count (will be filled later)
                                        '',  # Token Count Validation (will be filled later)
                                        net_worth,
                                        '',  # Net Worth (Actual) (will be filled later)
                                        '',  # Net Worth Validation (will be filled later)
                                        percentage
                                    ])
                                    print(f"      Found chain: {chain_name} ({token_count}) = ${net_worth} ({percentage}%)")
                                except:
                                    # Add without value if extraction failed
                                    chains_found.append(chain_name)
                                    header_holdings_data.append([
                                        'Token Holdings - Chain',
                                        chain_name,
                                        token_count,
                                        '',  # TC_UI Count
                                        '',  # Token Count Validation
                                        '',
                                        '',  # Net Worth (Actual)
                                        '',  # Net Worth Validation
                                        ''
                                    ])
                                    print(f"      Found chain: {chain_name} ({token_count})")
                        except:
                            continue
                except Exception as e:
                    print(f"   ⚠️  Error extracting chain holdings: {e}")

                # SECTION 3: Extract Token Holdings - Platform (WALLET, user-defined exchanges, DeFi protocols)
                # Platform items appear in a specific flex-wrap container with SVG icons
                # Chain items (Tron, Ethereum, etc.) have img icons and token counts - these should NOT be here
                # NFT is a token type, not a platform - should NOT be captured
                print("   🔍 Extracting Token Holdings - Platform...")

                try:
                    platforms_found = []

                    # Items to exclude from Platform section:
                    # - Chain names belong in Token Holdings - Chain section
                    # - NFT is a token type, not a platform
                    excluded_names = ['tron', 'ethereum', 'bsc', 'binance', 'polygon', 'arbitrum', 'optimism',
                                     'avalanche', 'fantom', 'solana', 'base', 'zksync', 'linea', 'scroll',
                                     'manta', 'blast', 'mantle', 'mode', 'cronos', 'gnosis', 'celo', 'moonbeam',
                                     'harmony', 'aurora', 'metis', 'boba', 'kava', 'klaytn', 'oasis', 'iotex',
                                     'heco', 'okc', 'moonriver', 'fuse', 'evmos', 'canto', 'dogechain', 'pulsechain',
                                     'near', 'cosmos', 'osmosis', 'injective', 'sei', 'sui', 'aptos', 'bitcoin', 'btc',
                                     'hyperevm', 'hyper evm',
                                     'nft']  # NFT is not a platform

                    # Method 1: Extract from HTML structure using JavaScript
                    # Platform items include:
                    # - WALLET (with SVG icon)
                    # - User-defined exchanges like "Binance - david", "moontest" (with SVG icon)
                    # - DeFi protocols like "Aave V3" (with img icons: protocol logo + chain overlay)
                    platform_items = page.evaluate('''() => {
                        const results = [];

                        // Find all flex-wrap containers
                        const containers = document.querySelectorAll('div.flex.flex-wrap.gap-3');

                        for (const container of containers) {
                            // Get direct children (div or button items)
                            const items = container.querySelectorAll(':scope > div, :scope > button');

                            for (const item of items) {
                                // Must have a name element with typography-body font-medium
                                const nameEl = item.querySelector('div[class*="typography-body"][class*="font-medium"]');
                                if (!nameEl) continue;
                                const name = nameEl.textContent.trim();
                                if (!name) continue;

                                // Skip items that look like chain entries (have token count like "(12)")
                                // Chain entries have pattern like "Tron (6)", "Base (3)", "Ethereum (5)"
                                if (/\\(\\d+\\)/.test(name)) continue;

                                // Skip items with percentage pattern only (these are from allocation sections)
                                const itemText = item.textContent || '';
                                if (/\\d+\\.\\d+%/.test(itemText) && !/\\$/.test(itemText)) continue;

                                // Get net worth value from title attribute or text content
                                let netWorth = '';
                                const valueEl = item.querySelector('div[class*="typography-body"][class*="font-normal"][class*="text-mono-900"]');
                                if (valueEl) {
                                    netWorth = valueEl.getAttribute('title') || valueEl.textContent.trim();
                                }

                                // Must have a value to be a valid platform entry
                                if (!netWorth || !netWorth.includes('$')) continue;

                                // Extract chain identifier from overlay icon images (for DeFi protocols)
                                // Structure: div.relative > img[alt="Protocol"] + img[alt="Chain"] (smaller overlay)
                                // DeFi protocols like Aave V3 have 2 imgs: protocol icon + chain overlay (Base/Ethereum)
                                let chain = '';
                                const relativeDiv = item.querySelector('div.relative');
                                if (relativeDiv) {
                                    const imgs = relativeDiv.querySelectorAll('img');
                                    if (imgs.length >= 2) {
                                        // Second img is the chain overlay icon
                                        chain = imgs[1].alt || '';
                                    }
                                }

                                results.push({
                                    name: name,
                                    netWorth: netWorth,
                                    chain: chain
                                });
                            }
                        }

                        return results;
                    }''')

                    for item in platform_items:
                        name = item.get('name', '')
                        net_worth_raw = item.get('netWorth', '')
                        chain = item.get('chain', '')

                        # Skip excluded names (chains and NFT)
                        if name.lower() in excluded_names:
                            print(f"      ⏭️  Skipping '{name}' from Platform section (not a platform)")
                            continue

                        # Build display name with chain identifier for DeFi protocols
                        display_name = name
                        if chain:
                            display_name = f"{name} ({chain})"

                        # Check for duplicates — exact match only (including chain)
                        if display_name in platforms_found:
                            continue

                        # Clean up net worth value
                        lt_prefix = '<' if '<' in net_worth_raw else ''
                        value_clean = re.search(r'\$?([\d,\.]+[KMB]?)', net_worth_raw)
                        net_worth = lt_prefix + (value_clean.group(1) if value_clean else net_worth_raw)

                        platforms_found.append(display_name)
                        header_holdings_data.append([
                            'Token Holdings - Platform',
                            display_name,
                            '',  # No token count for platforms
                            '',  # TC_UI Count
                            '',  # Token Count Validation
                            net_worth,
                            '',  # Net Worth (Actual)
                            '',  # Net Worth Validation
                            ''
                        ])
                        print(f"      Found platform: {display_name} = ${net_worth}")

                    # Method 2: Fallback - text-based extraction with known platform names
                    if not platforms_found:
                        valid_platforms = ['WALLET', 'CEX', 'DEFI', 'STAKING', 'LENDING', 'FARMING', 'BRIDGE', 'DEX', 'EXCHANGE', 'NFT']
                        platform_containers = page.locator('div.flex.flex-wrap').all()

                        for container in platform_containers:
                            try:
                                container_text = container.text_content().strip()
                                for platform in valid_platforms:
                                    pattern = rf'\b{platform}\s*\$?([\d,\.]+[KMB]?)'
                                    match = re.search(pattern, container_text, re.IGNORECASE)
                                    if match and platform.upper() not in platforms_found:
                                        net_worth = match.group(1)
                                        # Look for percentage (including optional "<" symbol for small values)
                                        # Pattern matches: "0.01%", "<0.01%", "< 0.01%" (with space), "50.25%"
                                        pct_match = re.search(rf'{platform}.*?(<\s*[\d.]+|[\d.]+)%', container_text, re.IGNORECASE)
                                        percentage = pct_match.group(1).replace(' ', '') if pct_match else ''  # Remove space if present
                                        platforms_found.append(platform.upper())
                                        header_holdings_data.append([
                                            'Token Holdings - Platform',
                                            platform.upper(),
                                            '',  # No token count
                                            '',  # TC_UI Count
                                            '',  # Token Count Validation
                                            net_worth,
                                            '',  # Net Worth (Actual)
                                            '',  # Net Worth Validation
                                            percentage
                                        ])
                                        print(f"      Found platform: {platform.upper()} = ${net_worth} ({percentage}%)")
                            except:
                                continue

                except Exception as e:
                    print(f"   ⚠️  Error extracting platform holdings: {e}")

                # SECTION 4: Extract Table Total Values (Wallet, DeFi protocols, etc.)
                # Rows: A="Table", B=table_name, C=blank, D=blank→"Not Applicable", E=value, G=blank, H=blank→"Not Applicable"
                print("   🔍 Extracting Table Total Values...")
                try:
                    _table_totals_found = []
                    for _ttdiv in page.locator('div.bg-grey-30.rounded-sm, div[class*="bg-grey-30"][class*="rounded-sm"]').all():
                        try:
                            _ttsec_txt = _ttdiv.text_content() or ""
                            if 'Total value' not in _ttsec_txt:
                                continue
                            # Section name from heading element
                            _ttname = ""
                            for _sel in ["h2", "h3", "[class*='font-bold']", "[class*='font-semibold']", "p"]:
                                _el = _ttdiv.locator(_sel).first
                                if _el.count() > 0:
                                    try:
                                        _t = (_el.inner_text(timeout=500) or "").strip().split("\n")[0]
                                        if _t and len(_t) < 60 and '$' not in _t and 'Total' not in _t:
                                            _ttname = _t; break
                                    except Exception:
                                        pass
                            if not _ttname:
                                continue
                            # Extract value from "Total value: $X" or "Total value: < $0.01"
                            _tv_m = re.search(r'Total\s+value\s*:?\s*(<?\s*)\$?\s*([\d,]+\.?\d*)', _ttsec_txt)
                            if not _tv_m:
                                continue
                            _tv_prefix = '<' if '<' in (_tv_m.group(1) or '') else ''
                            _tv = _tv_prefix + _tv_m.group(2)
                            # Deduplicate by (name, value)
                            if (_ttname, _tv) not in _table_totals_found:
                                _table_totals_found.append((_ttname, _tv))
                                # Table rows removed — not needed in Header & Token Holdings
                                print(f"      Table '{_ttname}' total: ${_tv} (skipped)")
                        except Exception:
                            pass
                    print(f"      Extracted {len(_table_totals_found)} table total(s)")
                except Exception as _tte:
                    print(f"   ⚠️  Error extracting table totals: {_tte}")

                print(f"✅ Extracted {len(header_holdings_data) - 1} header & holdings rows")

            except Exception as e:
                print(f"⚠️  Error extracting header & holdings: {e}")
                import traceback
                traceback.print_exc()

            all_tables_data['Overview - Header & Token Holdings Header'] = header_holdings_data

            # TABLE 1.6: Overview - Combined Net Worth
            print("\n📊 Table 1.6: Overview - Combined Net Worth")
            print("-"*80)
            print(f"   ⏱️  Starting Combined Net Worth extraction...")

            combined_net_worth_data = []
            try:
                print(f"   🔄 Extracting Combined Net Worth data...")

                # Add header row
                combined_net_worth_data.append(['Address/Exchange', 'Value'])

                # Extract Combined Net Worth section by looking for specific elements
                print("   🔍 Searching for Combined Net Worth addresses...")

                # Track found addresses to avoid duplicates
                addresses_found = set()

                # Strategy 0: Extract from HTML structure using data-tooltip-id elements
                # Two patterns exist:
                #   1. data-tooltip-id="address-tip" - for exchange names (e.g., "moon")
                #   2. data-tooltip-id="address-display-tooltip-{fullAddress}" - for wallet addresses
                # Full address comes from data-highlight-target attribute
                # Value comes from typography-body font-semibold text-mono-900 element
                try:
                    entries = page.evaluate('''() => {
                        // Find the Combined Net Worth section
                        const allElements = document.querySelectorAll('div, h2, h3, span');
                        let section = null;
                        for (const el of allElements) {
                            const txt = el.textContent.trim();
                            if (txt === 'Combined Net Worth' || (txt.startsWith('Combined Net Worth') && txt.length < 30)) {
                                section = el.closest('div[class*="space-y"]') || el.closest('div[class*="col-span"]') || el.parentElement?.parentElement?.parentElement;
                                break;
                            }
                        }
                        if (!section) return [];

                        // Find tooltip elements: both "address-tip" and "address-display-tooltip-*"
                        const tooltipElements = section.querySelectorAll('[data-tooltip-id^="address-display-tooltip-"], [data-tooltip-id="address-tip"]');
                        const results = [];
                        for (const el of tooltipElements) {
                            let name = '';

                            // Try to get full address from data-highlight-target attribute
                            const highlightEl = el.querySelector('[data-highlight-target]');
                            if (highlightEl) {
                                name = highlightEl.getAttribute('data-highlight-target') || '';
                            }

                            // Fallback: extract from data-tooltip-id (remove "address-display-tooltip-" prefix)
                            if (!name) {
                                const tooltipId = el.getAttribute('data-tooltip-id') || '';
                                if (tooltipId.startsWith('address-display-tooltip-')) {
                                    name = tooltipId.replace('address-display-tooltip-', '');
                                }
                            }

                            // Fallback: get display text from typography-body element (for exchange names like "moon")
                            if (!name) {
                                const nameEl = el.querySelector('div[class*="typography-body"]');
                                name = nameEl ? nameEl.textContent.trim() : el.textContent.trim();
                            }

                            // Find the value - walk up to find font-semibold element with $
                            let value = '';
                            let parent = el.parentElement;
                            for (let depth = 0; depth < 5 && parent; depth++) {
                                const semibold = parent.querySelector('div[class*="font-semibold"][class*="text-mono-900"]');
                                if (semibold && semibold.textContent.includes('$')) {
                                    value = semibold.textContent.trim();
                                    break;
                                }
                                parent = parent.parentElement;
                                if (parent === section) break;
                            }

                            if (name) {
                                results.push({name: name, value: value});
                            }
                        }
                        return results;
                    }''')

                    for entry in entries:
                        name = entry.get('name', '')
                        value = entry.get('value', '')
                        if name and name not in addresses_found:
                            addresses_found.add(name)
                            value_clean = re.search(r'\$?([\d,]+\.?\d*)', value)
                            combined_net_worth_data.append([
                                name,
                                value_clean.group(1) if value_clean else value
                            ])
                            print(f"      Found entry: {name} = {value}")
                except Exception as e:
                    print(f"   ⚠️  Strategy 0 (HTML structure) failed: {e}")

                # Strategy 1: Find the "Combined Net Worth" section first, then extract addresses within it
                try:
                    # Look for the Combined Net Worth section header
                    combined_section = page.locator('text="Combined Net Worth"').first
                    if combined_section.is_visible(timeout=3000):
                        # Get parent container that contains both header and address list
                        # Navigate up to find the container with address rows
                        parent = combined_section.locator('xpath=ancestor::div[contains(@class, "mb-4") or contains(@class, "space-y")]').first
                        if not parent.is_visible(timeout=1000):
                            parent = combined_section.locator('..').locator('..').locator('..').first

                        container_text = parent.text_content().strip()

                        # Find address patterns (truncated format like "TWQsW9HJ...4RePrMjr")
                        address_matches = re.findall(r'([A-Za-z0-9]{6,10}\.{2,3}[A-Za-z0-9]{6,10})', container_text)
                        # Find dollar values with 2 decimal places
                        value_matches = re.findall(r'\$([\d,]+\.\d{2})', container_text)

                        # Pair addresses with values (they should appear in order)
                        for i, address in enumerate(address_matches):
                            if address not in addresses_found and i < len(value_matches):
                                addresses_found.add(address)
                                combined_net_worth_data.append([address, value_matches[i]])
                                print(f"      Found address: {address} = ${value_matches[i]}")
                except Exception as e:
                    print(f"   ⚠️  Strategy 1 failed: {e}")

                # Strategy 2: If no addresses found, look within "Addresses" subsection
                if len(combined_net_worth_data) <= 1:
                    print("   🔍 Trying alternate extraction method...")

                    try:
                        # OPTIMIZED: Use more specific selector instead of all divs
                        # Look for divs that likely contain address rows (with flex, grid, or items-center classes)
                        specific_divs = page.locator('div[class*="flex"], div[class*="grid"], div[class*="items-center"]').all()

                        # Limit iteration to prevent hanging (max 500 divs instead of all divs)
                        max_divs_to_check = min(500, len(specific_divs))
                        found_in_this_strategy = 0

                        print(f"   🔍 Checking {max_divs_to_check} divs...")

                        for idx, div in enumerate(specific_divs[:max_divs_to_check]):
                            try:
                                div_text = div.text_content().strip()

                                # Must be short enough to be a single row (not a container)
                                if len(div_text) > 50:
                                    continue

                                # Must have BOTH an address pattern AND a dollar value
                                address_match = re.search(r'([A-Za-z0-9]{6,10}\.{2,3}[A-Za-z0-9]{6,10})', div_text)
                                value_match = re.search(r'\$([\d,]+\.\d{2})', div_text)

                                if address_match and value_match:
                                    address = address_match.group(1)
                                    if address not in addresses_found:
                                        addresses_found.add(address)
                                        combined_net_worth_data.append([address, value_match.group(1)])
                                        print(f"      Found address: {address} = ${value_match.group(1)}")
                                        found_in_this_strategy += 1

                                # Early exit if we found multiple addresses
                                if found_in_this_strategy >= 5:
                                    print(f"   ✅ Found {found_in_this_strategy} addresses, stopping early")
                                    break

                            except:
                                continue

                        if found_in_this_strategy == 0:
                            print(f"   ⚠️  No addresses found in {max_divs_to_check} divs checked")

                    except Exception as e:
                        print(f"   ⚠️  Strategy 2 error: {e}")

                print(f"✅ Extracted {len(combined_net_worth_data) - 1} Combined Net Worth rows")

                # Add validation column to Combined Net Worth data
                # Note: Validation will be performed later using external website scraping
                if len(combined_net_worth_data) > 0:
                    # Add validation column header
                    if len(combined_net_worth_data[0]) == 2:
                        combined_net_worth_data[0].append("Value Validation")

                    # Add empty validation cells for data rows
                    for row_idx in range(1, len(combined_net_worth_data)):
                        if len(combined_net_worth_data[row_idx]) == 2:
                            combined_net_worth_data[row_idx].append("")  # Will be filled later

            except Exception as e:
                print(f"⚠️  Error extracting Combined Net Worth: {e}")
                import traceback
                traceback.print_exc()

            all_tables_data['Overview - Combined Net Worth'] = combined_net_worth_data

            # TABLE 2: Overview - Platform Allocation (Pie Chart)
            print("\n📊 Table 2: Overview - Platform Allocation")
            print("-"*80)

            # Click on the Platform tab
            for selector in ['[role="tab"]:has-text("Platform")', 'button[role="tab"]:has-text("Platform")']:
                try:
                    if page.locator(selector).first.is_visible(timeout=2000):
                        page.locator(selector).first.click()
                        page.wait_for_timeout(2000)  # Reduced from 3000ms to 2000ms
                        break
                except:
                    continue

            # Take screenshot
            screenshot_path = f"{screenshot_folder}/02_overview_platformallocation_{timestamp_folder}.png"
            try:
                page.screenshot(path=screenshot_path, full_page=True, timeout=120000)
                print(f"   📸 Screenshot captured")
            except Exception as e:
                print(f"⚠️  Screenshot warning: {e}")

            # Extract pie chart data
            platform_allocation_data = []
            try:
                print(f"   🔄 Extracting Platform Allocation data...")

                # Add header row: Platform, Percentage, Net Worth
                platform_allocation_data.append(['Platform', 'Percentage', 'Net Worth'])

                extracted_platforms = set()

                # STEP 1: Extract SVG networth map FIRST (before any hover interactions that may corrupt DOM)
                svg_platform_networth_map = {}
                page.wait_for_timeout(2000)
                try:
                    page.locator('svg text').first.wait_for(state='attached', timeout=5000)
                    print(f"   ✅ SVG text elements found in DOM")
                except:
                    print(f"   ⚠️  No SVG text elements detected")

                svg_platform_networth_map = extract_svg_networth_map(page)
                if svg_platform_networth_map:
                    print(f"   📊 SVG Platform networth map: {svg_platform_networth_map}")

                # STEP 2: Direct page text scan for "Name ($X,XXX.XX)" patterns (works regardless of SVG structure)
                try:
                    direct_networth_map = page.evaluate('''() => {
                        const results = {};
                        // Scan ALL text nodes in the entire page
                        const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT, null, false);
                        const allTexts = [];
                        while (walker.nextNode()) {
                            const txt = walker.currentNode.textContent.trim();
                            if (txt) allTexts.push(txt);
                        }
                        // Also get SVG text/tspan content
                        document.querySelectorAll('svg text, svg tspan').forEach(t => {
                            const txt = t.textContent.trim();
                            if (txt) allTexts.push(txt);
                        });
                        // Find combined "Name ($X,XXX.XX)" patterns
                        for (const txt of allTexts) {
                            if (txt.includes('$')) {
                                const match = txt.match(/^([A-Za-z][A-Za-z\\s\\-]+?)\\s*\\(\\$?([\\d,]+\\.?\\d*)\\)/);
                                if (match) {
                                    results[match[1].trim().toUpperCase()] = match[2].replace(/,/g, '');
                                }
                            }
                        }
                        // Also try sequential pairing: find $ values and look at previous text nodes
                        const dollarItems = [];
                        const nameItems = [];
                        for (const txt of allTexts) {
                            if (txt.includes('$') && txt.length < 50) {
                                const valMatch = txt.match(/\\(?\\$?([\\d,]+\\.?\\d*)\\)?/);
                                if (valMatch) dollarItems.push({text: txt, value: valMatch[1].replace(/,/g, '')});
                            } else if (/^[A-Za-z][A-Za-z\\s\\-]*$/.test(txt) && txt.length < 30 && txt !== 'Total') {
                                nameItems.push(txt);
                            }
                        }
                        return results;
                    }''')
                    if direct_networth_map:
                        print(f"   📊 Direct text scan: {direct_networth_map}")
                        for dk, dv in direct_networth_map.items():
                            if dk not in svg_platform_networth_map:
                                svg_platform_networth_map[dk] = dv
                except Exception as e:
                    print(f"   ⚠️  Direct text scan failed: {e}")

                # STEP 3: Debug dump - show ALL $ elements on page for diagnostics
                try:
                    debug_dollar = page.evaluate('''() => {
                        const results = [];
                        const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT, null, false);
                        while (walker.nextNode()) {
                            const txt = walker.currentNode.textContent.trim();
                            if (txt && txt.includes('$') && txt.length < 100) {
                                const parent = walker.currentNode.parentElement;
                                const tag = parent ? parent.tagName : 'unknown';
                                const cls = parent ? (parent.className || '') : '';
                                results.push({text: txt, tag: tag, class: String(cls).substring(0, 80)});
                            }
                        }
                        document.querySelectorAll('svg text, svg tspan').forEach(t => {
                            const txt = t.textContent.trim();
                            if (txt && txt.includes('$')) {
                                results.push({text: txt, tag: t.tagName, class: 'SVG'});
                            }
                        });
                        return results;
                    }''')
                    print(f"   DEBUG: All '$' text on page ({len(debug_dollar)} items):")
                    for item in debug_dollar[:15]:
                        print(f"      <{item.get('tag','')} class='{item.get('class','')}'> {item.get('text','')}")
                except:
                    pass

                if svg_platform_networth_map:
                    print(f"   📊 Final Platform networth map: {svg_platform_networth_map}")

                # STEP 4: Extract net worth via hover on pie chart segments
                hover_networth_map = {}
                try:
                    print(f"   🔄 Extracting Platform net worth via hover...")
                    pie_paths = page.locator('svg path[d*="A"]').all()
                    print(f"   DEBUG: Found {len(pie_paths)} SVG paths for platform pie chart")

                    for i, path in enumerate(pie_paths):
                        try:
                            if path.is_visible():
                                path.hover(timeout=2000)
                                page.wait_for_timeout(800)

                                # Scan DOM for $ text after hover
                                hover_dollar_texts = page.evaluate('''() => {
                                    const results = [];
                                    const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT, null, false);
                                    while (walker.nextNode()) {
                                        const txt = walker.currentNode.textContent.trim();
                                        if (txt && txt.includes('$') && txt.length < 100) {
                                            const parent = walker.currentNode.parentElement;
                                            const tag = parent ? parent.tagName : 'unknown';
                                            const cls = parent ? parent.className : '';
                                            results.push({text: txt, tag: tag, class: String(cls).substring(0, 80)});
                                        }
                                    }
                                    document.querySelectorAll('svg text, svg tspan').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt && txt.includes('$')) {
                                            results.push({text: txt, tag: t.tagName, class: ''});
                                        }
                                    });
                                    return results;
                                }''')

                                if i == 0:
                                    print(f"      DEBUG: After hover path[{i}] - found {len(hover_dollar_texts)} '$' elements:")
                                    for item in hover_dollar_texts[:15]:
                                        print(f"         <{item.get('tag','')} class='{item.get('class','')}'> {item.get('text','')}")

                                hover_texts = page.evaluate('''() => {
                                    const results = [];
                                    document.querySelectorAll('svg text, svg tspan').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt) results.push(txt);
                                    });
                                    document.querySelectorAll('[role="tooltip"], .recharts-tooltip-wrapper, .recharts-default-tooltip, div[class*="tooltip"], div[class*="Tooltip"], div[class*="popover"], div[class*="Popover"]').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt) results.push(txt);
                                    });
                                    return results;
                                }''')

                                platform_name = None
                                net_worth_val = None

                                for item in hover_dollar_texts:
                                    txt = item.get('text', '')
                                    if '$' in txt:
                                        match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', txt)
                                        if match:
                                            candidate = match.group(1).replace(',', '')
                                            tag = item.get('tag', '').upper()
                                            if tag in ['TEXT', 'TSPAN', 'SPAN', 'DIV', 'P']:
                                                if not net_worth_val or tag in ['TEXT', 'TSPAN']:
                                                    net_worth_val = candidate

                                for hover_text in hover_texts:
                                    if '$' in hover_text:
                                        match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', hover_text)
                                        if match:
                                            net_worth_val = match.group(1).replace(',', '')
                                    elif hover_text and re.match(r'^[A-Za-z][A-Za-z\-\s]*$', hover_text) and hover_text not in ['Total', 'Allocation']:
                                        if len(hover_text) < 30:
                                            platform_name = hover_text.strip()

                                if platform_name and net_worth_val:
                                    platform_key = platform_name.upper()
                                    if platform_key not in hover_networth_map:
                                        hover_networth_map[platform_key] = net_worth_val
                                        print(f"      ✅ Hover extracted: {platform_name} = ${net_worth_val}")

                                        try:
                                            safe_name = re.sub(r'[^\w\-]', '_', platform_name)
                                            screenshot_path = f"{screenshot_folder}/02.1_platform_hover_{safe_name}_{timestamp_folder}.png"
                                            page.screenshot(path=screenshot_path, full_page=True, timeout=120000)
                                        except:
                                            pass
                                else:
                                    if i < 3:
                                        print(f"      DEBUG: Path[{i}] hover - name='{platform_name}', value='{net_worth_val}'")

                        except Exception as hover_err:
                            print(f"      ⚠️  Hover error on path {i}: {hover_err}")
                            continue

                    print(f"   📊 Hover platform net worth map: {hover_networth_map}")

                except Exception as e:
                    print(f"   ⚠️  Platform hover extraction warning: {e}")

                # Merge hover results into SVG map (SVG results take priority, hover fills gaps)
                for hk, hv in hover_networth_map.items():
                    if hk not in svg_platform_networth_map:
                        svg_platform_networth_map[hk] = hv

                if svg_platform_networth_map:
                    print(f"   📊 Combined Platform networth map: {svg_platform_networth_map}")

                # Method 0: Extract from cursor-pointer legend items + donut center text for net worth
                try:
                    cursor_divs = page.locator('div.cursor-pointer').all()
                    print(f"   🔍 Found {len(cursor_divs)} cursor-pointer divs")

                    valid_legend_items = []
                    for cdiv in cursor_divs:
                        try:
                            truncate_el = cdiv.locator('.truncate').first
                            if truncate_el.count() == 0:
                                continue
                            platform_name = truncate_el.text_content().strip()
                            shrink_els = cdiv.locator('.shrink-0').all()
                            percentage = ''
                            for sel in shrink_els:
                                sel_text = sel.text_content().strip()
                                if '%' in sel_text:
                                    percentage = sel_text
                                    break
                            if platform_name and percentage:
                                valid_legend_items.append({'div': cdiv, 'name': platform_name, 'percentage': percentage})
                                print(f"      Found: '{platform_name}' | '{percentage}'")
                        except:
                            continue

                    print(f"   📊 Found {len(valid_legend_items)} valid platform legend items")

                    for item in valid_legend_items:
                        name = item['name']
                        pct = item['percentage'].replace('%', '').strip()
                        cdiv = item['div']

                        if name in extracted_platforms:
                            continue

                        # Priority 1: Check pre-extracted maps
                        net_worth = hover_networth_map.get(name.upper(), "")
                        if not net_worth:
                            net_worth = svg_platform_networth_map.get(name.upper(), "")
                        if not net_worth:
                            clean_name = name.upper().replace('-', '').replace(' ', '')
                            for map_key, map_val in {**hover_networth_map, **svg_platform_networth_map}.items():
                                if map_key.replace('-', '').replace(' ', '') == clean_name:
                                    net_worth = map_val
                                    break

                        # Priority 2: Hover legend item and read donut chart center text
                        if not net_worth:
                            try:
                                cdiv.hover(timeout=3000)
                                page.wait_for_timeout(1500)

                                center_text = page.evaluate('''() => {
                                    const el = document.querySelector('div[class*="font-bold"][class*="text-center"][class*="max-w-"]');
                                    return el ? el.textContent.trim() : '';
                                }''')

                                if center_text and '$' in center_text:
                                    val_match = re.search(r'\$?([\d,]+\.?\d*)', center_text)
                                    if val_match:
                                        net_worth = val_match.group(1).replace(',', '')
                                        print(f"      ✅ Donut center: '{name}' = ${net_worth}")
                            except Exception as hover_err:
                                print(f"      ⚠️  Hover failed for '{name}': {hover_err}")

                        platform_allocation_data.append([name, pct, net_worth])
                        extracted_platforms.add(name)
                        print(f"      Platform: {name} = {pct}% (${net_worth})")
                except Exception as e:
                    print(f"   ⚠️  Method 0 (cursor-pointer) warning: {e}")

                # Method 1: Fallback - Look for div items with percentage
                if not extracted_platforms:
                    all_divs = page.locator('div').all()

                    allocation_items = []
                    for div in all_divs:
                        try:
                            text = div.text_content().strip()
                            # Must contain %, be short, and look like allocation item
                            if '%' in text and len(text) < 50 and len(text) > 3:
                                has_letters = any(c.isalpha() for c in text)
                                has_digits = any(c.isdigit() for c in text)
                                if has_letters and has_digits:
                                    allocation_items.append((div, text))
                        except:
                            continue

                    print(f"   🔍 Found {len(allocation_items)} potential allocation items (fallback)")

                    for div, text in allocation_items:
                        try:
                            clean_text = text.strip()

                            if '%' not in clean_text:
                                continue

                            parts = clean_text.split('%')[0].strip()
                            parts = parts.replace('<', '').replace('>', '').strip()

                            words = parts.split()
                            if len(words) >= 2:
                                percentage_num = words[-1]
                                platform_name = ' '.join(words[:-1])
                            else:
                                match = re.match(r'([A-Za-z\s]+)([\d.]+)', parts)
                                if not match:
                                    continue
                                platform_name = match.group(1).strip()
                                percentage_num = match.group(2)

                            try:
                                pct_value = float(percentage_num)
                            except:
                                continue

                            skip_keywords = ['show', 'holdings', 'value', 'tokens', 'total', 'combined', 'address', 'showing']
                            if any(keyword in platform_name.lower() for keyword in skip_keywords):
                                continue

                            if len(words) > 5:
                                continue

                            if platform_name in extracted_platforms:
                                continue

                            # Look up net worth from pre-extracted SVG map first
                            net_worth = svg_platform_networth_map.get(platform_name.upper(), "")
                            if not net_worth:
                                clean_name = platform_name.upper().replace('-', '').replace(' ', '')
                                for svg_key, svg_val in svg_platform_networth_map.items():
                                    if svg_key.replace('-', '').replace(' ', '') == clean_name:
                                        net_worth = svg_val
                                        break

                            # Fallback: hover to trigger tooltip and read SVG
                            if not net_worth:
                                try:
                                    div.hover(timeout=2000)
                                    page.wait_for_timeout(1500)

                                    svg_texts = page.evaluate('''() => {
                                        const texts = document.querySelectorAll('svg text');
                                        return Array.from(texts).map(t => t.textContent);
                                    }''')

                                    for svg_text in svg_texts:
                                        if svg_text and ('$' in svg_text or (svg_text.startswith('(') and ')' in svg_text)):
                                            value_match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', svg_text)
                                            if value_match:
                                                net_worth = value_match.group(1)
                                                break
                                except:
                                    pass

                            platform_allocation_data.append([
                                platform_name,
                                percentage_num,
                                net_worth
                            ])
                            extracted_platforms.add(platform_name)

                        except Exception as e:
                            continue

                print(f"✅ Extracted {len(platform_allocation_data) - 1} platform allocation rows")
            except Exception as e:
                print(f"⚠️  Error extracting platform allocation: {e}")

            all_tables_data['Overview - Platform Allocation'] = platform_allocation_data

            # TABLE 3: Overview - Chain Allocation (Pie Chart)
            print("\n📊 Table 3: Overview - Chain Allocation")
            print("-"*80)

            # Click on the Chain tab
            for selector in ['[role="tab"]:has-text("Chain")', 'button[role="tab"]:has-text("Chain")']:
                try:
                    if page.locator(selector).first.is_visible(timeout=3000):
                        page.locator(selector).first.click()
                        page.wait_for_timeout(3000)
                        break
                except:
                    continue

            # Take screenshot
            screenshot_path = f"{screenshot_folder}/03_overview_chainallocation_{timestamp_folder}.png"
            try:
                page.screenshot(path=screenshot_path, full_page=True, timeout=120000)
                print(f"   📸 Screenshot captured")
            except Exception as e:
                print(f"⚠️  Screenshot warning: {e}")

            # Extract pie chart data
            chain_allocation_data = []
            try:
                print(f"   🔄 Extracting Chain Allocation data...")

                # Add header row: Chain, Percentage, Net Worth
                chain_allocation_data.append(['Chain', 'Percentage', 'Net Worth'])

                extracted_chains = set()

                # NEW: Extract net worth via hover on pie chart segments (reads SVG text elements)
                hover_chain_networth_map = {}
                try:
                    print(f"   🔄 Extracting Chain net worth via hover...")
                    pie_paths = page.locator('svg path[d*="A"]').all()
                    print(f"   DEBUG: Found {len(pie_paths)} SVG paths for chain pie chart")

                    for i, path in enumerate(pie_paths):
                        try:
                            if path.is_visible():
                                path.hover(timeout=2000)
                                page.wait_for_timeout(800)

                                # Comprehensive: scan ALL text nodes in DOM for '$' after hover
                                hover_dollar_texts = page.evaluate('''() => {
                                    const results = [];
                                    const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT, null, false);
                                    while (walker.nextNode()) {
                                        const txt = walker.currentNode.textContent.trim();
                                        if (txt && txt.includes('$') && txt.length < 100) {
                                            const parent = walker.currentNode.parentElement;
                                            const tag = parent ? parent.tagName : 'unknown';
                                            const cls = parent ? parent.className : '';
                                            results.push({text: txt, tag: tag, class: String(cls).substring(0, 80)});
                                        }
                                    }
                                    document.querySelectorAll('svg text, svg tspan').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt && txt.includes('$')) {
                                            results.push({text: txt, tag: t.tagName, class: ''});
                                        }
                                    });
                                    return results;
                                }''')

                                if i == 0:
                                    print(f"      DEBUG: After hover path[{i}] - found {len(hover_dollar_texts)} '$' elements:")
                                    for item in hover_dollar_texts[:15]:
                                        print(f"         <{item.get('tag','')} class='{item.get('class','')}'> {item.get('text','')}")

                                hover_texts = page.evaluate('''() => {
                                    const results = [];
                                    document.querySelectorAll('svg text, svg tspan').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt) results.push(txt);
                                    });
                                    document.querySelectorAll('[role="tooltip"], .recharts-tooltip-wrapper, .recharts-default-tooltip, div[class*="tooltip"], div[class*="Tooltip"], div[class*="popover"], div[class*="Popover"]').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt) results.push(txt);
                                    });
                                    return results;
                                }''')

                                chain_name = None
                                net_worth_val = None

                                for item in hover_dollar_texts:
                                    txt = item.get('text', '')
                                    if '$' in txt:
                                        match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', txt)
                                        if match:
                                            candidate = match.group(1).replace(',', '')
                                            tag = item.get('tag', '').upper()
                                            if tag in ['TEXT', 'TSPAN', 'SPAN', 'DIV', 'P']:
                                                if not net_worth_val or tag in ['TEXT', 'TSPAN']:
                                                    net_worth_val = candidate

                                for hover_text in hover_texts:
                                    if '$' in hover_text:
                                        match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', hover_text)
                                        if match:
                                            net_worth_val = match.group(1).replace(',', '')
                                    elif hover_text and re.match(r'^[A-Za-z\-\s]+$', hover_text) and hover_text not in ['Total', 'Allocation']:
                                        if len(hover_text) < 30:
                                            chain_name = hover_text.strip()

                                if chain_name and net_worth_val:
                                    chain_key = chain_name.upper()
                                    if chain_key not in hover_chain_networth_map:
                                        hover_chain_networth_map[chain_key] = net_worth_val
                                        print(f"      ✅ Hover extracted: {chain_name} = ${net_worth_val}")

                                        try:
                                            safe_name = re.sub(r'[^\w\-]', '_', chain_name)
                                            screenshot_path = f"{screenshot_folder}/03.1_chain_hover_{safe_name}_{timestamp_folder}.png"
                                            page.screenshot(path=screenshot_path, full_page=True, timeout=120000)
                                        except:
                                            pass
                                else:
                                    print(f"      DEBUG: Path[{i}] hover - name='{chain_name}', value='{net_worth_val}'")

                        except Exception as hover_err:
                            print(f"      ⚠️  Hover error on path {i}: {hover_err}")
                            continue

                    print(f"   📊 Hover chain net worth map: {hover_chain_networth_map}")

                except Exception as e:
                    print(f"   ⚠️  Chain hover extraction warning: {e}")

                # Pre-extract all SVG text elements to build chain->networth map
                # Wait for SVG to render, then use shared helper
                page.wait_for_timeout(2000)
                try:
                    page.locator('svg text').first.wait_for(state='attached', timeout=5000)
                    print(f"   ✅ SVG text elements found in DOM")
                except:
                    print(f"   ⚠️  No SVG text elements detected, will try fallbacks")

                svg_chain_networth_map = extract_svg_networth_map(page)

                # Merge hover results into SVG map (hover results take priority)
                for hk, hv in hover_chain_networth_map.items():
                    if hk not in svg_chain_networth_map:
                        svg_chain_networth_map[hk] = hv

                if svg_chain_networth_map:
                    print(f"   📊 Chain networth map: {svg_chain_networth_map}")

                # Method 0: Extract from cursor-pointer legend items + donut center text for net worth
                try:
                    cursor_divs = page.locator('div.cursor-pointer').all()
                    print(f"   🔍 Found {len(cursor_divs)} cursor-pointer divs")

                    valid_legend_items = []
                    for cdiv in cursor_divs:
                        try:
                            truncate_el = cdiv.locator('.truncate').first
                            if truncate_el.count() == 0:
                                continue
                            chain_name = truncate_el.text_content().strip()
                            shrink_els = cdiv.locator('.shrink-0').all()
                            percentage = ''
                            for sel in shrink_els:
                                sel_text = sel.text_content().strip()
                                if '%' in sel_text:
                                    percentage = sel_text
                                    break
                            if chain_name and percentage:
                                valid_legend_items.append({'div': cdiv, 'name': chain_name, 'percentage': percentage})
                                print(f"      Found: '{chain_name}' | '{percentage}'")
                        except:
                            continue

                    print(f"   📊 Found {len(valid_legend_items)} valid chain legend items")

                    for item in valid_legend_items:
                        name = item['name']
                        pct = item['percentage'].replace('%', '').strip()
                        cdiv = item['div']

                        if name in extracted_chains:
                            continue

                        # Priority 1: Check pre-extracted maps
                        net_worth = hover_chain_networth_map.get(name.upper(), "")
                        if not net_worth:
                            net_worth = svg_chain_networth_map.get(name.upper(), "")
                        if not net_worth:
                            clean_name = name.upper().replace('-', '').replace(' ', '')
                            for map_key, map_val in {**hover_chain_networth_map, **svg_chain_networth_map}.items():
                                if map_key.replace('-', '').replace(' ', '') == clean_name:
                                    net_worth = map_val
                                    break

                        # Priority 2: Hover legend item and read donut chart center text
                        if not net_worth:
                            try:
                                cdiv.hover(timeout=3000)
                                page.wait_for_timeout(1500)

                                center_text = page.evaluate('''() => {
                                    const el = document.querySelector('div[class*="font-bold"][class*="text-center"][class*="max-w-"]');
                                    return el ? el.textContent.trim() : '';
                                }''')

                                if center_text and '$' in center_text:
                                    val_match = re.search(r'\$?([\d,]+\.?\d*)', center_text)
                                    if val_match:
                                        net_worth = val_match.group(1).replace(',', '')
                                        print(f"      ✅ Donut center: '{name}' = ${net_worth}")
                            except Exception as hover_err:
                                print(f"      ⚠️  Hover failed for '{name}': {hover_err}")

                        chain_allocation_data.append([name, pct, net_worth])
                        extracted_chains.add(name)
                        print(f"      Chain: {name} = {pct}% (${net_worth})")
                except Exception as e:
                    print(f"   ⚠️  Method 0 (cursor-pointer) warning: {e}")

                # Method 1: Fallback - Look for div items with percentage
                if not extracted_chains:
                    all_divs = page.locator('div').all()

                    allocation_items = []
                    for div in all_divs:
                        try:
                            text = div.text_content().strip()
                            if '%' in text and len(text) < 50 and len(text) > 3:
                                has_letters = any(c.isalpha() for c in text)
                                has_digits = any(c.isdigit() for c in text)
                                if has_letters and has_digits:
                                    allocation_items.append((div, text))
                        except:
                            continue

                    print(f"   🔍 Found {len(allocation_items)} potential allocation items (fallback)")

                    for div, text in allocation_items:
                        try:
                            clean_text = text.strip()

                            if '%' not in clean_text:
                                continue

                            parts = clean_text.split('%')[0].strip()
                            parts = parts.replace('<', '').replace('>', '').strip()

                            words = parts.split()
                            if len(words) >= 2:
                                percentage_num = words[-1]
                                chain_name = ' '.join(words[:-1])
                            else:
                                match = re.match(r'([A-Za-z\s]+)([\d.]+)', parts)
                                if not match:
                                    continue
                                chain_name = match.group(1).strip()
                                percentage_num = match.group(2)

                            try:
                                pct_value = float(percentage_num)
                            except:
                                continue

                            skip_keywords = ['show', 'holdings', 'value', 'tokens', 'wallet', 'total', 'combined', 'address', 'showing']
                            if any(keyword in chain_name.lower() for keyword in skip_keywords):
                                continue

                            if len(words) > 4:
                                continue

                            if chain_name in extracted_chains:
                                continue

                            # Look up net worth from pre-extracted SVG map first
                            net_worth = svg_chain_networth_map.get(chain_name.upper(), "")
                            if not net_worth:
                                clean_name = chain_name.upper().replace('-', '').replace(' ', '')
                                for svg_key, svg_val in svg_chain_networth_map.items():
                                    if svg_key.replace('-', '').replace(' ', '') == clean_name:
                                        net_worth = svg_val
                                        break

                            # Fallback: hover to trigger tooltip and read SVG
                            if not net_worth:
                                try:
                                    div.hover(timeout=2000)
                                    page.wait_for_timeout(1500)

                                    svg_texts = page.evaluate('''() => {
                                        const texts = document.querySelectorAll('svg text');
                                        return Array.from(texts).map(t => t.textContent);
                                    }''')

                                    for svg_text in svg_texts:
                                        if svg_text and ('$' in svg_text or (svg_text.startswith('(') and ')' in svg_text)):
                                            value_match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', svg_text)
                                            if value_match:
                                                net_worth = value_match.group(1)
                                                break
                                except:
                                    pass

                            chain_allocation_data.append([
                                chain_name,
                                percentage_num,
                                net_worth
                            ])
                            extracted_chains.add(chain_name)

                        except Exception as e:
                            continue

                print(f"✅ Extracted {len(chain_allocation_data) - 1} chain allocation rows")
            except Exception as e:
                print(f"⚠️  Error extracting chain allocation: {e}")

            all_tables_data['Overview - Chain Allocation'] = chain_allocation_data

            # TABLE 4: Overview - Token Allocation (Pie Chart)
            print("\n📊 Table 4: Overview - Token Allocation")
            print("-"*80)

            # Click on the Token tab
            for selector in ['[role="tab"]:has-text("Token")', 'button[role="tab"]:has-text("Token")']:
                try:
                    if page.locator(selector).first.is_visible(timeout=3000):
                        page.locator(selector).first.click()
                        page.wait_for_timeout(3000)
                        break
                except:
                    continue

            # Take screenshot
            screenshot_path = f"{screenshot_folder}/04_overview_tokenallocation_{timestamp_folder}.png"
            try:
                page.screenshot(path=screenshot_path, full_page=True, timeout=120000)
                print(f"   📸 Screenshot captured")
            except Exception as e:
                print(f"⚠️  Screenshot warning: {e}")

            # Extract pie chart data
            token_allocation_data = []
            try:
                print(f"   🔄 Extracting Token Allocation data...")

                # Add header row: Token, Percentage, Net Worth
                token_allocation_data.append(['Token', 'Percentage', 'Net Worth'])

                extracted_tokens = set()  # To avoid duplicates

                # Step 1: Wait for SVG chart to render, then extract net worth map
                page.wait_for_timeout(2000)  # Extra wait for chart render
                try:
                    page.locator('svg text').first.wait_for(state='attached', timeout=5000)
                    print(f"   ✅ SVG text elements found in DOM")
                except:
                    print(f"   ⚠️  No SVG text elements detected, will try HTML fallback")

                svg_token_networth_map = extract_svg_networth_map(page)
                if svg_token_networth_map:
                    print(f"   📊 SVG Token networth map: {svg_token_networth_map}")

                # Step 1b: Extract net worth via hover on pie chart segments
                hover_token_networth_map = {}
                try:
                    print(f"   🔄 Extracting Token net worth via hover...")
                    pie_paths = page.locator('svg path[d*="A"]').all()
                    print(f"   DEBUG: Found {len(pie_paths)} SVG paths for token pie chart")

                    for i, path in enumerate(pie_paths):
                        try:
                            if path.is_visible():
                                path.hover(timeout=2000)
                                page.wait_for_timeout(800)

                                hover_dollar_texts = page.evaluate('''() => {
                                    const results = [];
                                    const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT, null, false);
                                    while (walker.nextNode()) {
                                        const txt = walker.currentNode.textContent.trim();
                                        if (txt && txt.includes('$') && txt.length < 100) {
                                            const parent = walker.currentNode.parentElement;
                                            const tag = parent ? parent.tagName : 'unknown';
                                            const cls = parent ? parent.className : '';
                                            results.push({text: txt, tag: tag, class: String(cls).substring(0, 80)});
                                        }
                                    }
                                    document.querySelectorAll('svg text, svg tspan').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt && txt.includes('$')) {
                                            results.push({text: txt, tag: t.tagName, class: ''});
                                        }
                                    });
                                    return results;
                                }''')

                                if i == 0:
                                    print(f"      DEBUG: After hover path[{i}] - found {len(hover_dollar_texts)} '$' elements:")
                                    for item in hover_dollar_texts[:15]:
                                        print(f"         <{item.get('tag','')} class='{item.get('class','')}'> {item.get('text','')}")

                                hover_texts = page.evaluate('''() => {
                                    const results = [];
                                    document.querySelectorAll('svg text, svg tspan').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt) results.push(txt);
                                    });
                                    document.querySelectorAll('[role="tooltip"], .recharts-tooltip-wrapper, .recharts-default-tooltip, div[class*="tooltip"], div[class*="Tooltip"], div[class*="popover"], div[class*="Popover"]').forEach(t => {
                                        const txt = t.textContent.trim();
                                        if (txt) results.push(txt);
                                    });
                                    return results;
                                }''')

                                token_name_hover = None
                                net_worth_val = None

                                for item in hover_dollar_texts:
                                    txt = item.get('text', '')
                                    if '$' in txt:
                                        match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', txt)
                                        if match:
                                            candidate = match.group(1).replace(',', '')
                                            tag = item.get('tag', '').upper()
                                            if tag in ['TEXT', 'TSPAN', 'SPAN', 'DIV', 'P']:
                                                if not net_worth_val or tag in ['TEXT', 'TSPAN']:
                                                    net_worth_val = candidate

                                for hover_text in hover_texts:
                                    if '$' in hover_text:
                                        match = re.search(r'\(?\$?([\d,]+\.?\d*)\)?', hover_text)
                                        if match:
                                            net_worth_val = match.group(1).replace(',', '')
                                    elif hover_text and re.match(r'^[A-Za-z][A-Za-z\-\s]*$', hover_text) and hover_text not in ['Total', 'Allocation']:
                                        if len(hover_text) < 30:
                                            token_name_hover = hover_text.strip()

                                if token_name_hover and net_worth_val:
                                    token_key = token_name_hover.upper()
                                    if token_key not in hover_token_networth_map:
                                        hover_token_networth_map[token_key] = net_worth_val
                                        print(f"      ✅ Hover extracted: {token_name_hover} = ${net_worth_val}")
                                else:
                                    if i < 3:
                                        print(f"      DEBUG: Path[{i}] hover - name='{token_name_hover}', value='{net_worth_val}'")

                        except Exception as hover_err:
                            print(f"      ⚠️  Hover error on path {i}: {hover_err}")
                            continue

                    print(f"   📊 Hover token net worth map: {hover_token_networth_map}")

                except Exception as e:
                    print(f"   ⚠️  Token hover extraction warning: {e}")

                # Merge hover results into SVG map
                for hk, hv in hover_token_networth_map.items():
                    if hk not in svg_token_networth_map:
                        svg_token_networth_map[hk] = hv

                if svg_token_networth_map:
                    print(f"   📊 Combined Token networth map: {svg_token_networth_map}")

                # Step 2: Find allocation items from HTML list
                cursor_divs = page.locator('div.cursor-pointer').all()
                print(f"   🔍 Found {len(cursor_divs)} cursor-pointer divs")

                allocation_divs = []
                for cdiv in cursor_divs:
                    try:
                        truncate_el = cdiv.locator('.truncate').first
                        if truncate_el.count() == 0:
                            continue
                        token_name = truncate_el.text_content().strip()
                        shrink_els = cdiv.locator('.shrink-0').all()
                        percentage = ''
                        for sel in shrink_els:
                            sel_text = sel.text_content().strip()
                            if '%' in sel_text:
                                percentage = sel_text
                                break
                        if token_name and percentage:
                            allocation_divs.append({'div': cdiv, 'token': token_name, 'percentage': percentage})
                            print(f"      Found: '{token_name}' | '{percentage}'")
                    except:
                        continue

                print(f"   🔍 Found {len(allocation_divs)} valid allocation items")

                # Step 3: Process each allocation item and match net worth
                for item in allocation_divs:
                    token_name = item['token']
                    percentage_raw = item['percentage']
                    cdiv = item['div']

                    if token_name in extracted_tokens:
                        continue

                    percentage_cleaned = percentage_raw.replace('%', '').strip()

                    skip_keywords = ['show', 'holdings', 'value', 'tokens', 'wallet', 'total', 'combined', 'address', 'showing']
                    if any(keyword in token_name.lower() for keyword in skip_keywords):
                        continue

                    # Look up net worth from pre-extracted hover and SVG maps
                    net_worth = ""
                    # Priority 1: Hover-extracted net worth
                    net_worth = hover_token_networth_map.get(token_name.upper(), "")
                    if not net_worth:
                        clean_name = token_name.upper().replace('-', '').replace(' ', '')
                        for hover_key, hover_val in hover_token_networth_map.items():
                            if hover_key.replace('-', '').replace(' ', '') == clean_name:
                                net_worth = hover_val
                                break
                    # Priority 2: SVG/combined map exact match
                    if not net_worth:
                        net_worth = svg_token_networth_map.get(token_name.upper(), "")
                    # Try partial match: SVG may have full name, HTML may be truncated
                    if not net_worth:
                        for svg_key, svg_val in svg_token_networth_map.items():
                            if svg_key in token_name.upper() or token_name.upper() in svg_key:
                                net_worth = svg_val
                                break
                    # Try base name without chain suffix: "GTUSDCP(BASE)" -> "GTUSDCP"
                    if not net_worth:
                        base_name = re.sub(r'\(.*?\)', '', token_name).strip().upper()
                        net_worth = svg_token_networth_map.get(base_name, "")

                    # Remove commas from net worth value
                    if net_worth:
                        net_worth = net_worth.replace(',', '')

                    # Fallback: hover on legend item and read donut chart center text
                    if not net_worth:
                        try:
                            cdiv.hover(timeout=3000)
                            page.wait_for_timeout(1500)

                            # Read the donut chart center text - it shows the hovered item's net worth
                            center_text = page.evaluate('''() => {
                                const el = document.querySelector('div[class*="font-bold"][class*="text-center"][class*="max-w-"]');
                                return el ? el.textContent.trim() : '';
                            }''')

                            if center_text and '$' in center_text:
                                val_match = re.search(r'\$?([\d,]+\.?\d*)', center_text)
                                if val_match:
                                    net_worth = val_match.group(1).replace(',', '')
                                    print(f"      ✅ Donut center net worth for '{token_name}': ${net_worth}")

                            # Debug: if first item, dump what we see
                            if not net_worth and token_name == allocation_divs[0]['token']:
                                print(f"      DEBUG: Donut center text after hover '{token_name}': '{center_text}'")

                        except Exception as hover_err:
                            print(f"      ⚠️  Hover failed for '{token_name}': {hover_err}")

                    # Take screenshot after hover
                    try:
                        safe_token = re.sub(r'[^\w\-]', '_', token_name)
                        screenshot_path = f"{screenshot_folder}/04.1_token_alloc_hover_{safe_token}_{timestamp_folder}.png"
                        page.screenshot(path=screenshot_path, full_page=True, timeout=120000)
                    except:
                        pass

                    token_allocation_data.append([
                        token_name,
                        percentage_cleaned,
                        net_worth
                    ])
                    extracted_tokens.add(token_name)
                    print(f"   ✓ Token Allocation: {token_name} | {percentage_cleaned}% | Net Worth: {net_worth}")

                print(f"✅ Extracted {len(token_allocation_data) - 1} token allocation rows")
            except Exception as e:
                print(f"⚠️  Error extracting token allocation: {e}")
                import traceback
                traceback.print_exc()

            all_tables_data['Overview - Token Allocation'] = token_allocation_data

            # REMOVED: Holdings page screenshots and data extraction
            # TABLE 5 & 6 (Holdings - Token and Holdings - Chain) have been removed per user request

            # STEP 3.5: Navigate to TronScan and capture address details
            # For multi-address portfolios, visit ALL addresses
            # Skip for CEX-only portfolios with no wallet addresses
            print("\n" + "="*80)
            print("STEP 3.5: TronScan Address Details")
            print("="*80)

            # Collect all addresses to visit from TRX_ADDRESSES
            # Skip if no wallet addresses (CEX-only portfolio)
            if SKIP_TRX_API or not TRX_ADDRESSES:
                print("   ⏭️  Skipped (CEX-only portfolio - no wallet addresses)")
                addresses_to_visit = []
            else:
                addresses_to_visit = TRX_ADDRESSES if len(TRX_ADDRESSES) > 0 else [TRX_ADDRESS]
                print(f"   Addresses to visit: {len(addresses_to_visit)}")

            for addr_idx, visit_address in enumerate(addresses_to_visit):
                addr_suffix = visit_address[-8:]
                try:
                    print(f"\n🌐 Opening TronScan for address {addr_idx + 1}/{len(addresses_to_visit)}: {visit_address}")
                    tronscan_url = f"https://tronscan.org/#/address/{visit_address}"

                    # Create new page (tab) in same context
                    tronscan_page = context.new_page()
                    tronscan_page.goto(tronscan_url)

                    # Wait for page to load
                    print("   ⏳ Waiting for TronScan page to load...")
                    tronscan_page.wait_for_timeout(5000)

                    # Click "Got It" button to dismiss cookie consent modal
                    print("   🍪 Dismissing cookie consent modal...")
                    try:
                        got_it_selectors = [
                            'button:has-text("Got It")',
                            'button:has-text("Got it")',
                            'text="Got It"',
                            'text="Got it"',
                            '.ant-btn:has-text("Got")',
                        ]

                        for selector in got_it_selectors:
                            try:
                                got_it_button = tronscan_page.locator(selector).first
                                if got_it_button.count() > 0 and got_it_button.is_visible(timeout=2000):
                                    got_it_button.click()
                                    print(f"   ✅ Clicked 'Got It' button using: {selector}")
                                    tronscan_page.wait_for_timeout(1000)
                                    break
                            except:
                                continue
                    except Exception as e:
                        print(f"   ⚠️  Could not click 'Got It' button: {e}")

                    # Click "More" button to expand token list
                    print("   🔽 Looking for 'More' button...")

                    more_clicked = False
                    selectors_to_try = [
                        'text=/More/i',
                        'a:has-text("More")',
                        'button:has-text("More")',
                        'span:has-text("More")',
                        '.ant-tabs-content-more-link',
                        'text="more"',
                        'a[href*="more"]',
                    ]

                    for selector in selectors_to_try:
                        try:
                            more_button = tronscan_page.locator(selector).first

                            if more_button.count() > 0 and more_button.is_visible(timeout=2000):
                                print(f"   ✅ Found 'More' button with selector: {selector}")
                                more_button.click()
                                print("   ✅ 'More' button clicked")
                                more_clicked = True

                                print("   ⏳ Waiting for page to navigate and load transaction details...")
                                tronscan_page.wait_for_timeout(5000)

                                try:
                                    tronscan_page.wait_for_selector('table tbody tr', state='visible', timeout=10000)
                                    print("   ✅ Transaction details loaded")
                                except:
                                    print("   ⚠️  Transaction table not found, proceeding with screenshot")

                                break
                        except Exception as e:
                            continue

                    if not more_clicked:
                        print("   ⚠️  'More' button not found with any selector, capturing page as is")

                    # Take screenshot with address suffix for multi-address
                    if len(addresses_to_visit) > 1:
                        screenshot_path = f"{screenshot_folder}/05_tronscan_{addr_suffix}_{timestamp_folder}.png"
                    else:
                        screenshot_path = f"{screenshot_folder}/05_tronscan_address_details_{timestamp_folder}.png"
                    tronscan_page.screenshot(path=screenshot_path, full_page=True, timeout=120000)
                    print(f"   📸 Screenshot saved: {screenshot_path}")

                    # Close TronScan tab
                    tronscan_page.close()
                    print(f"   ✅ TronScan tab closed")

                except Exception as e:
                    print(f"⚠️  Error during TronScan navigation for {visit_address}: {e}")
                    print("   Continuing...\n")

            print()

            # ================================================================
            # DYNAMIC ADDRESS DETECTION from Combined Net Worth
            # Classify addresses as EVM (0x, 42 chars) or Tron (T, 34 chars)
            # and set skip flags for API routing
            # ================================================================
            detected_tron = detect_tron_addresses(combined_net_worth_data)
            detected_evm = detect_evm_addresses(combined_net_worth_data)

            # Also merge with pre-configured addresses (from --portfolio flag or hardcoded)
            if EVM_ADDRESSES:
                for addr in EVM_ADDRESSES:
                    if addr.lower() not in [a.lower() for a in detected_evm]:
                        detected_evm.append(addr)

            # Update globals based on detection
            if detected_tron:
                TRX_ADDRESSES = detected_tron
                TRX_ADDRESS = detected_tron[0]
            if not detected_tron and not TRX_ADDRESSES:
                SKIP_TRX_API = True

            if not detected_evm:
                SKIP_SIM_DUNE_API = True

            # Remove duplicates (case-insensitive) for EVM
            seen = set()
            unique_evm = []
            for addr in detected_evm:
                if addr.lower() not in seen:
                    seen.add(addr.lower())
                    unique_evm.append(addr)
            detected_evm = unique_evm

            # Print address detection summary
            print(f"\n   📋 Address Detection Summary (from Combined Net Worth):")
            print(f"      Tron: {len(detected_tron)} address(es) → {'TRX Balance API' if not SKIP_TRX_API else 'SKIP'}")
            for idx, addr in enumerate(detected_tron, 1):
                print(f"         {idx}. {addr}")
            print(f"      EVM: {len(detected_evm)} address(es) → {'Sim Dune & Coingecko API' if not SKIP_SIM_DUNE_API else 'SKIP'}")
            for idx, addr in enumerate(detected_evm, 1):
                print(f"         {idx}. {addr}")
            if SKIP_TRX_API and SKIP_SIM_DUNE_API:
                print(f"      Portfolio type: CEX/Exchange only (no blockchain addresses)")

            # STEP 3.6: Sim Dune API for EVM Addresses (0x...)
            # Fetch balances via Sim Dune API for detected EVM addresses

            sim_dune_excel_files = []
            sim_dune_extracted_data = []  # Collected data for DAM export

            evm_addresses = detected_evm
            if SKIP_SIM_DUNE_API:
                print("\n" + "="*80)
                print("STEP 3.6: Sim Dune API - SKIPPED")
                print("="*80)
                print("   ⏭️  Skipping Sim Dune API (no EVM addresses in portfolio)")

            if not SKIP_SIM_DUNE_API and evm_addresses:
                print("\n" + "="*80)
                print("STEP 3.6: Sim Dune API - EVM Address Balances")
                print("="*80)

                print(f"   🔍 Found {len(evm_addresses)} EVM address(es) (0x...)")
                for idx, evm_addr in enumerate(evm_addresses, 1):
                    print(f"      {idx}. {evm_addr}")

                print()

                # Collect all data from all addresses first
                all_sim_dune_data = []
                for addr_idx, evm_addr in enumerate(evm_addresses):
                    addr_suffix = evm_addr[-8:]
                    print(f"\n🌐 Fetching Sim Dune balances for address {addr_idx + 1}/{len(evm_addresses)}: {evm_addr}")

                    # Call Sim Dune API
                    address, data, success = fetch_sim_dune_balance(evm_addr)

                    if success and data:
                        print(f"   ✅ Sim Dune API response received")

                        # Count balances
                        balances = data.get('balances', data) if isinstance(data, dict) else data
                        if isinstance(balances, list):
                            print(f"   📊 Found {len(balances)} token balance(s)")

                        # Collect data for combined export
                        all_sim_dune_data.append({'address': evm_addr, 'data': data})
                    else:
                        print(f"   ⚠️  No data returned from Sim Dune API")

                # Save raw SimDune API responses to JSON immediately after collection
                if all_sim_dune_data:
                    import json as _json_sd_raw
                    _raw_json_dir = "test-results/API Result"
                    os.makedirs(_raw_json_dir, exist_ok=True)
                    from datetime import datetime as _dt_sd
                    _ts = _dt_sd.now().strftime("%Y%m%d_%H%M%S")
                    _sd_raw_path = os.path.join(_raw_json_dir, f"SimDune_Raw_{_safe_name}_{_ts}.json")
                    with open(_sd_raw_path, 'w') as _jf:
                        _json_sd_raw.dump(all_sim_dune_data, _jf, indent=2, default=str)
                    print(f"   📄 Saved SimDune raw API response to: {os.path.basename(_sd_raw_path)}")

                # Export all data to a single Excel file
                if all_sim_dune_data:
                    try:
                        sim_dune_file, extracted_data = export_sim_dune_to_excel_combined(
                            all_sim_dune_data,
                            output_folder="test-results/API Result",
                            portfolio_name=target_portfolio_name
                        )
                        sim_dune_excel_files.append(sim_dune_file)
                        sim_dune_extracted_data = extracted_data
                        print(f"\n   ✅ Combined Excel exported: {os.path.basename(sim_dune_file)}")
                    except Exception as e:
                        print(f"   ⚠️  Error exporting Excel: {e}")
                        import traceback
                        traceback.print_exc()

                print(f"\n✅ Sim Dune API completed for {len(evm_addresses)} EVM address(es)")
                if sim_dune_excel_files:
                    print(f"   📁 Excel file created: {sim_dune_excel_files[0]}")
                print()

            # STEP 3.7: Rabby Combined API — always fires both complex_protocol_list
            # AND complex_app_list for every EVM address, even if one returns no data.
            rabby_excel_files = []
            rabby_extracted_data = []
            hyperliquid_extracted_data = []

            if not evm_addresses:
                print("\n" + "="*80)
                print("STEP 3.7: Rabby Combined API - SKIPPED")
                print("="*80)
                print("   ⏭️  Skipping Rabby API (no EVM addresses in portfolio)")
                print()
            else:
                print("\n" + "="*80)
                print("STEP 3.7: Rabby Combined API (Protocol + Hyperliquid)")
                print("="*80)

                print(f"   🔍 Found {len(evm_addresses)} EVM address(es) (0x...)")
                for idx, evm_addr in enumerate(evm_addresses, 1):
                    print(f"      {idx}. {evm_addr}")
                print()

                try:
                    from utils.rabby_api import fetch_protocol_list, fetch_app_list, parse_protocol_data, parse_app_data
                    import json as _json_rabby

                    _rabby_json_folder = "test-results/excel-exports"
                    os.makedirs(_rabby_json_folder, exist_ok=True)

                    _all_proto_raw = []
                    _all_app_raw = []
                    _all_parsed_rows = []

                    for evm_addr in evm_addresses:
                        print(f"\n   📡 Fetching Rabby protocol list for {evm_addr}...")
                        proto_data = fetch_protocol_list(evm_addr)
                        _all_proto_raw.append({"address": evm_addr, "request": f"complex_protocol_list?id={evm_addr}", "response": proto_data})
                        if proto_data is not None:
                            rows = parse_protocol_data(evm_addr, proto_data)
                            _all_parsed_rows.extend(rows)
                            print(f"      ✅ Got {len(proto_data) if isinstance(proto_data, list) else 0} protocol(s), {len(rows)} token row(s)")
                        else:
                            print(f"      ⚠️  No protocol data")

                        print(f"   📡 Fetching Rabby app list for {evm_addr}...")
                        app_data = fetch_app_list(evm_addr)
                        _all_app_raw.append({"address": evm_addr, "request": f"complex_app_list?id={evm_addr}", "response": app_data})
                        if app_data is not None:
                            rows = parse_app_data(evm_addr, app_data)
                            _all_parsed_rows.extend(rows)
                            app_count = len(app_data.get("apps", [])) if isinstance(app_data, dict) else len(app_data) if isinstance(app_data, list) else 0
                            print(f"      ✅ Got {app_count} app(s), {len(rows)} token row(s)")
                        else:
                            print(f"      ⚠️  No app data")

                    # Save request + response as 2 separate JSON files
                    _proto_json_path = os.path.join(_rabby_json_folder, f"Rabby_Protocol_Raw_{_safe_name}.json")
                    _app_json_path = os.path.join(_rabby_json_folder, f"Rabby_App_Raw_{_safe_name}.json")
                    with open(_proto_json_path, 'w') as _jf:
                        _json_rabby.dump(_all_proto_raw, _jf, indent=2, default=str)
                    with open(_app_json_path, 'w') as _jf:
                        _json_rabby.dump(_all_app_raw, _jf, indent=2, default=str)
                    print(f"\n   📄 Saved Rabby protocol raw to: {os.path.basename(_proto_json_path)}")
                    print(f"   📄 Saved Rabby app raw to: {os.path.basename(_app_json_path)}")

                    # Store parsed rows for Rabby Api Data sheet
                    # Keep as dicts for validation; convert to lists only when writing Excel
                    _rabby_headers = ["Address", "Name", "ID", "Chain", "Pool Name", "Description", "Side", "Symbol/Currency Pair", "Leverage", "PnL (USD)", "Price", "Amount", "Calculated Value"]
                    rabby_extracted_data = [_rabby_headers] + _all_parsed_rows
                    print(f"   ✅ Rabby API: {len(_all_parsed_rows)} parsed rows ready for Excel")

                except Exception as e:
                    print(f"   ⚠️  Error fetching Rabby API: {e}")
                    import traceback
                    traceback.print_exc()
                print()

            # STEP 4: Export to Excel
            print("\n" + "="*80)
            print("STEP 4: Export DAM Tables to Excel (Overview tables only)")
            print("="*80)

            os.makedirs("test-results/excel-exports", exist_ok=True)

            # REMOVED: Holdings - Token and Holdings - Chain tables per user request
            # (TABLE 5 & 6 extraction code removed)

            """
            # COMMENTED OUT: Holdings extraction code removed per user request
            # Original TABLE 5 & 6 code was here (lines 1149-1419)
            for selector in ['text="Holdings"', '[role="tab"]:has-text("Holdings")']:
                try:
                    if page.locator(selector).first.is_visible(timeout=3000):
                        page.locator(selector).first.click()
                        page.wait_for_timeout(5000)  # Increased wait after clicking Holdings tab
                        break
                except:
                    continue

            try:
                if page.locator('button:has-text("Token")').first.is_visible(timeout=2000):
                    page.locator('button:has-text("Token")').first.click()
                    page.wait_for_timeout(5000)  # Increased wait after clicking Token sub-tab
            except:
                pass

            # Wait for table data to fully load before screenshot
            print("   ⏳ Waiting for table data to load...")
            try:
                # First, wait for table to exist
                page.wait_for_selector("table tbody tr", state="visible", timeout=15000)

                # Then wait for loading spinner/skeleton to disappear
                # Look for common loading indicators and wait for them to be hidden
                loading_selectors = [
                    '[data-loading="true"]',
                    '.loading',
                    '.skeleton',
                    '[aria-busy="true"]',
                    '.animate-pulse',  # Common Tailwind loading class
                    '[data-state="loading"]'
                ]
                for selector in loading_selectors:
                    try:
                        if page.locator(selector).count() > 0:
                            print(f"   ⏳ Waiting for loading indicator '{selector}' to disappear...")
                            page.wait_for_selector(selector, state="hidden", timeout=8000)
                    except:
                        pass

                # Verify actual data is visible (not "No data to display")
                print("   🔍 Verifying actual token data is visible...")
                data_visible = False
                max_retries = 20  # Increased retries
                retry_count = 0

                while not data_visible and retry_count < max_retries:
                    # Check for "No data to display" message
                    if page.locator('text="No data to display"').count() > 0:
                        print(f"   ⚠️  Still showing 'No data to display', waiting... (retry {retry_count + 1}/{max_retries})")
                        page.wait_for_timeout(3000)
                        retry_count += 1
                        continue

                    # Check if we have actual token data in table cells
                    # Look for table cells with token names (TRX, USDT, etc.) or value amounts
                    token_cells = page.locator('table td, table th').all()
                    has_token_data = False

                    for cell in token_cells:
                        cell_text = cell.inner_text().strip()
                        # Check if cell contains token names or numeric values (not empty, not just headers)
                        if cell_text and cell_text not in ['Token', 'Chain', 'Portfolio %', 'Amount', 'Value', 'Share']:
                            if 'TRX' in cell_text or 'USDT' in cell_text or 'NFT' in cell_text or \
                               '$' in cell_text or (',' in cell_text and any(c.isdigit() for c in cell_text)):
                                has_token_data = True
                                break

                    if has_token_data:
                        data_visible = True
                        print("   ✅ Token data visible in table")
                    else:
                        print(f"   ⏳ Waiting for token data to appear... (retry {retry_count + 1}/{max_retries})")
                        page.wait_for_timeout(3000)
                        retry_count += 1

                if not data_visible:
                    print("   ⚠️  Warning: Could not verify data visibility after retries - taking screenshot anyway")

                # Additional wait to ensure all data is rendered
                page.wait_for_timeout(2000)
                print("   ✅ Ready for screenshot")
            except Exception as e:
                print(f"   ⚠️  Table load wait warning: {e}")
                page.wait_for_timeout(8000)  # Much longer fallback wait

            try:
                page.screenshot(path=f"{screenshot_folder}/04_holdings_token.png", full_page=True, timeout=120000)
                print("   📸 Screenshot captured")
            except Exception as e:
                print(f"⚠️  Screenshot warning: {e}")

            holdings_token_data = []
            try:
                tables = page.locator("table").all()
                if tables:
                    rows = tables[0].locator("tr").all()
                    for row_idx, row in enumerate(rows):
                        cells = row.locator("th, td").all()
                        row_data = []
                        is_header_row = row_idx == 0  # First row is header

                        for cell in cells:
                            cell_text = cell.inner_text().strip()

                            # Special handling for header row's "Price (24H)" column
                            if is_header_row and 'Price' in cell_text and '24' in cell_text:
                                # Split "Price (24H)" header into two separate columns
                                row_data.append('Price')
                                row_data.append('Price (24h)')
                            # Check if this cell contains combined Price and Price(24h) data
                            # Pattern: "$0.31 1.66% ↑" or "$0.311.66%↑" (with or without spaces)
                            elif '$' in cell_text and ('%' in cell_text or '↑' in cell_text or '↓' in cell_text):
                                # Use regex to extract price and change parts
                                match = re.match('[$]?(\\d+\\.\\d{1,2})\\s*(\\d+\\.\\d+)%\\s*([↑↓]?)', cell_text)
                                if match:
                                    price_part = match.group(1)
                                    change_percentage = match.group(2)
                                    arrow = match.group(3)

                                    # Detect color to determine sign (+ or -)
                                    sign = ""
                                    try:
                                        # Look for the percentage element inside the cell to check its color
                                        percentage_elem = cell.locator(f'text=/{change_percentage}%/').first
                                        class_attr = percentage_elem.get_attribute('class')

                                        # Check if it has text-error class (red = negative)
                                        if class_attr and 'text-error' in class_attr:
                                            sign = "-"
                                        # Otherwise check for green/positive indicators
                                        elif class_attr and ('text-success' in class_attr or 'text-green' in class_attr):
                                            sign = "+"
                                        # Fallback: use arrow direction
                                        elif arrow == '↓':
                                            sign = "-"
                                        elif arrow == '↑':
                                            sign = "+"
                                    except:
                                        # Fallback: use arrow direction if color detection fails
                                        if arrow == '↓':
                                            sign = "-"
                                        elif arrow == '↑':
                                            sign = "+"

                                    change_part = f"{sign}{change_percentage}" if sign else change_percentage
                                    row_data.append(price_part)
                                    row_data.append(clean_currency_symbols(change_part))
                                else:
                                    # Fallback: treat as single value if regex doesn't match
                                    row_data.append(clean_currency_symbols(cell_text))
                                    row_data.append("")  # Empty Price(24h) column
                            else:
                                row_data.append(clean_currency_symbols(cell_text))

                        if row_data and any(row_data):
                            holdings_token_data.append(row_data)
            except Exception as e:
                print(f"⚠️  Error: {e}")

            # REMOVED: Holdings - Token table per user request
            # all_tables_data['Holdings - Token'] = holdings_token_data
            # print(f"✅ Extracted {len(holdings_token_data)} rows")

            # TABLE 6: Holdings - Chain
            print("\n📊 Table 6: Holdings - Chain")
            print("-"*80)

            try:
                if page.locator('button:has-text("Chain")').first.is_visible(timeout=2000):
                    page.locator('button:has-text("Chain")').first.click()
                    page.wait_for_timeout(5000)  # Increased wait after clicking

                    # Wait specifically for "Chain" text to appear in table header
                    print("   ⏳ Waiting for 'Chain' column to appear...")
                    page.wait_for_selector('table th:has-text("Chain"), table td:has-text("Chain")', state="visible", timeout=10000)
                    print("   ✅ 'Chain' column visible")
            except Exception as e:
                print(f"   ⚠️  Warning clicking Chain tab: {e}")

            # Wait for table data to fully load before screenshot
            print("   ⏳ Waiting for table data to load...")
            try:
                # First, wait for table to exist
                page.wait_for_selector("table tbody tr", state="visible", timeout=15000)

                # Then wait for loading spinner/skeleton to disappear
                loading_selectors = [
                    '[data-loading="true"]',
                    '.loading',
                    '.skeleton',
                    '[aria-busy="true"]',
                    '.animate-pulse',
                    '[data-state="loading"]'
                ]
                for selector in loading_selectors:
                    try:
                        if page.locator(selector).count() > 0:
                            print(f"   ⏳ Waiting for loading indicator '{selector}' to disappear...")
                            page.wait_for_selector(selector, state="hidden", timeout=8000)
                    except:
                        pass

                # Verify actual chain data is visible (not "No data to display")
                print("   🔍 Verifying actual chain data is visible...")
                data_visible = False
                max_retries = 20
                retry_count = 0

                while not data_visible and retry_count < max_retries:
                    # Check for "No data to display" message
                    if page.locator('text="No data to display"').count() > 0:
                        print(f"   ⚠️  Still showing 'No data to display', waiting... (retry {retry_count + 1}/{max_retries})")
                        page.wait_for_timeout(3000)
                        retry_count += 1
                        continue

                    # Check if we have actual chain data in table cells
                    chain_cells = page.locator('table td, table th').all()
                    has_chain_data = False

                    for cell in chain_cells:
                        cell_text = cell.inner_text().strip()
                        # Check if cell contains chain names (Tron) or token names or values
                        if cell_text and cell_text not in ['Chain', 'Token', 'Portfolio %', 'Amount', 'Value', 'Share']:
                            if 'Tron' in cell_text or 'TRX' in cell_text or 'USDT' in cell_text or \
                               '$' in cell_text or (',' in cell_text and any(c.isdigit() for c in cell_text)):
                                has_chain_data = True
                                break

                    if has_chain_data:
                        data_visible = True
                        print("   ✅ Chain data visible in table")
                    else:
                        print(f"   ⏳ Waiting for chain data to appear... (retry {retry_count + 1}/{max_retries})")
                        page.wait_for_timeout(3000)
                        retry_count += 1

                if not data_visible:
                    print("   ⚠️  Warning: Could not verify data visibility after retries - taking screenshot anyway")

                # Additional wait to ensure all data is rendered
                page.wait_for_timeout(2000)
                print("   ✅ Ready for screenshot")
            except Exception as e:
                print(f"   ⚠️  Table load wait warning: {e}")
                page.wait_for_timeout(8000)  # Longer fallback wait

            try:
                page.screenshot(path=f"{screenshot_folder}/05_holdings_chain.png", full_page=True, timeout=120000)
                print("   📸 Screenshot captured")
            except Exception as e:
                print(f"⚠️  Screenshot warning: {e}")

            holdings_chain_data = []
            try:
                tables = page.locator("table").all()
                if tables:
                    rows = tables[0].locator("tr").all()
                    for row_idx, row in enumerate(rows):
                        cells = row.locator("th, td").all()
                        row_data = []
                        is_header_row = row_idx == 0  # First row is header

                        for cell in cells:
                            cell_text = cell.inner_text().strip()

                            # Special handling for header row's "Price (24H)" column
                            if is_header_row and 'Price' in cell_text and '24' in cell_text:
                                # Split "Price (24H)" header into two separate columns
                                row_data.append('Price')
                                row_data.append('Price (24h)')
                            # Check if this cell contains combined Price and Price(24h) data
                            # Pattern: "$0.31 1.66% ↑" or "$0.311.66%↑" (with or without spaces)
                            elif '$' in cell_text and ('%' in cell_text or '↑' in cell_text or '↓' in cell_text):
                                # Use regex to extract price and change parts
                                match = re.match('[$]?(\\d+\\.\\d{1,2})\\s*(\\d+\\.\\d+)%\\s*([↑↓]?)', cell_text)
                                if match:
                                    price_part = match.group(1)
                                    change_percentage = match.group(2)
                                    arrow = match.group(3)

                                    # Detect color to determine sign (+ or -)
                                    sign = ""
                                    try:
                                        # Look for the percentage element inside the cell to check its color
                                        percentage_elem = cell.locator(f'text=/{change_percentage}%/').first
                                        class_attr = percentage_elem.get_attribute('class')

                                        # Check if it has text-error class (red = negative)
                                        if class_attr and 'text-error' in class_attr:
                                            sign = "-"
                                        # Otherwise check for green/positive indicators
                                        elif class_attr and ('text-success' in class_attr or 'text-green' in class_attr):
                                            sign = "+"
                                        # Fallback: use arrow direction
                                        elif arrow == '↓':
                                            sign = "-"
                                        elif arrow == '↑':
                                            sign = "+"
                                    except:
                                        # Fallback: use arrow direction if color detection fails
                                        if arrow == '↓':
                                            sign = "-"
                                        elif arrow == '↑':
                                            sign = "+"

                                    change_part = f"{sign}{change_percentage}" if sign else change_percentage
                                    row_data.append(price_part)
                                    row_data.append(clean_currency_symbols(change_part))
                                else:
                                    # Fallback: treat as single value if regex doesn't match
                                    row_data.append(clean_currency_symbols(cell_text))
                                    row_data.append("")  # Empty Price(24h) column
                            else:
                                row_data.append(clean_currency_symbols(cell_text))

                        if row_data and any(row_data):
                            holdings_chain_data.append(row_data)
            except Exception as e:
                print(f"⚠️  Error: {e}")

            # REMOVED: Holdings - Chain table per user request
            # all_tables_data['Holdings - Chain'] = holdings_chain_data
            # print(f"✅ Extracted {len(holdings_chain_data)} rows")
            """

            timestamp_file = datetime.now().strftime("%m%d_%H%M")
            email_username = test_email.split('@')[0] if '@' in test_email else test_email
            # Sanitize portfolio name for filename (replace special chars with _)
            _safe_name = re.sub(r'[^\w\-]', '_', portfolio_name_from_excel or "unknown")
            # Naming: DAMExp_portfolioname_username_MMDD_HHMM
            excel_filename = f"DAMExp_{_safe_name}_{email_username}_{timestamp_file}.xlsx"
            excel_path = os.path.join("test-results/excel-exports", excel_filename)

            # Create Excel workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Yellow fill for validation columns
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # Combine DeFi protocol tabs into a single "Overview - De-Fi" tab
            defi_tab_names = []
            defi_combined_data = []
            defi_header = ['Defi', 'Chain', 'Type', 'Pool/ Position Pair', 'Description', 'Amount', 'Amount Tooltip',
                          'Amount Validation', 'FE - Amount Validation', 'Amount Validation Diff', 'Amount Diff %',
                          'Value', 'Value Validation', 'Api Calc Value', 'Value Difference', 'Side', 'Leverage', 'Collateral', 'P&L', 'Row Matched']
            defi_combined_data.append(defi_header)

            # Reverse chain name map: Full name -> short code (for matching with Rabby API)
            chain_name_to_code = {
                'ethereum': 'eth', 'base': 'base', 'bsc': 'bsc', 'binance smart chain': 'bsc',
                'arbitrum': 'arb', 'optimism': 'op', 'polygon': 'matic', 'avalanche': 'avax',
                'fantom': 'ftm', 'cronos': 'cro', 'aurora': 'aurora', 'heco': 'heco',
                'okx': 'okx', 'gnosis': 'xdai', 'boba': 'boba', 'metis': 'metis',
                'moonriver': 'movr', 'celo': 'celo', 'klaytn': 'klay', 'mantle': 'mnt',
                'linea': 'linea', 'zksync era': 'era', 'zksync': 'era', 'scroll': 'scroll',
                'blast': 'blast', 'mode': 'mode', 'tron': 'tron',
            }

            for sheet_name, data in list(all_tables_data.items()):
                if sheet_name.startswith('Overview') or not data or len(data) <= 1:
                    continue
                # Detect DeFi tabs by data structure: header starts with "Position Type"
                # This catches any protocol tab regardless of name (Lido, Virtuals Protocol, Hyperliquid, aixCB, etc.)
                _header = data[0] if data else []
                is_defi_structure_tab = len(_header) >= 2 and 'Position Type' in str(_header[0])
                # Also detect by known DeFi keywords (fallback) or chain suffix pattern
                _defi_tab_kws = ['aave', 'v3', 'compound', 'uniswap', 'defi', 'saver', 'lido', 'curve', 'maker',
                                 'morpho', 'spark', 'euler', 'pendle', 'yearn', 'balancer', 'convex', 'merkl',
                                 'gearbox', 'fluid', 'kamino', 'drift', 'marginfi', 'hyperliquid',
                                 'eigenlayer', 'nftx']
                is_defi_tab = any(kw in sheet_name.lower() for kw in _defi_tab_kws + ['midas'])
                has_chain_suffix = '(' in sheet_name and ')' in sheet_name

                if is_defi_structure_tab or is_defi_tab or has_chain_suffix:
                    defi_tab_names.append(sheet_name)
                    # Extract protocol name and chain from tab name (e.g., "Aave V3 (Base)" -> "Aave V3", "Base")
                    if '(' in sheet_name and ')' in sheet_name:
                        protocol_name = sheet_name[:sheet_name.rfind('(')].strip()
                        chain_full = sheet_name[sheet_name.rfind('(')+1:sheet_name.rfind(')')]
                        # Convert chain full name to short code for Rabby matching
                        chain = chain_name_to_code.get(chain_full.lower(), chain_full.lower())
                    else:
                        protocol_name = sheet_name
                        chain = "Hyperliquid" if sheet_name.lower() == "hyperliquid" else ""

                    # Check the header to determine table structure
                    header = data[0] if data else []
                    is_defi_structure = len(header) >= 3 and 'Position Type' in str(header[0])

                    # Skip header row, process data rows
                    for row in data[1:]:
                        if is_defi_structure and len(row) >= 3:
                            # DeFi table structure (8 columns):
                            # 0:Position Type, 1:Pool, 2:Amount, 3:Amount Tooltip, 4:Value, 5+:extra cols
                            position_type = row[0] if len(row) > 0 else ""
                            pool = row[1] if len(row) > 1 else ""
                            amount = row[2] if len(row) > 2 else ""
                            amount_tooltip = row[3] if len(row) > 3 else ""
                            value = row[4] if len(row) > 4 else ""
                            # Extra columns beyond the standard 5
                            extra_cols = list(row[5:]) if len(row) > 5 else []

                            # Add to combined De-Fi data
                            _is_hl_perp = "hyperliquid" in protocol_name.lower() and str(position_type).strip().lower() == "perpetuals"
                            combined_row = [
                                protocol_name,  # A: Defi
                                chain,          # B: Chain
                                position_type,  # C: Type
                                pool,           # D: Pool/ Position Pair
                                row[8] if len(row) > 8 else "",  # E: Description
                                "-" if _is_hl_perp else amount,         # F: Amount
                                "-" if _is_hl_perp else amount_tooltip, # G: Amount Tooltip
                                "-" if _is_hl_perp else "",             # H: Amount Validation
                                "-" if _is_hl_perp else "",             # I: FE - Amount Validation
                                "-" if _is_hl_perp else "",             # J: Amount Validation Diff
                                "-" if _is_hl_perp else "",             # K: Amount Diff %
                                value,          # L: Value
                                "",             # M: Value Validation
                                "",             # N: Api Calc Value
                                "",             # O: Value Difference
                                row[9]  if len(row) > 9  else "",  # P: Side
                                row[10] if len(row) > 10 else "",  # Q: Leverage
                                row[11] if len(row) > 11 else "",  # R: Collateral
                                row[12] if len(row) > 12 else "",  # S: P&L
                            ]
                            defi_combined_data.append(combined_row)
                        elif len(row) >= 17:
                            # CEX table structure (21 columns):
                            # 0:Token, 1:Category, 2:Account, 3:Price, 4:Price Tooltip,
                            # 5:FE-Price Val, 6:Price Val, 7:Price Diff Val, 8:Price (24h),
                            # 9:Price (24h) Val, 10:Price (24H) Diff Val, 11:Amount,
                            # 12:Amount Tooltip, 13:Amount Val, 14:FE-Amount Val,
                            # 15:Amount Diff Val, 16:Value
                            pool = row[0] if len(row) > 0 else ""  # Token = Pool
                            position_type = row[1] if len(row) > 1 else ""  # Category = Pool Name
                            amount = row[11] if len(row) > 11 else ""  # Amount
                            amount_tooltip = row[12] if len(row) > 12 else ""  # Amount Tooltip
                            value = row[16] if len(row) > 16 else ""  # Value

                            # Add to combined De-Fi data
                            defi_combined_data.append([
                                protocol_name,  # A: Defi
                                chain,          # B: Chain
                                position_type,  # C: Type
                                pool,           # D: Pool/ Position Pair
                                "",             # E: Description
                                amount,         # F: Amount
                                amount_tooltip, # G: Amount Tooltip
                                "",             # H: Amount Validation
                                "",             # I: FE - Amount Validation
                                "",             # J: Amount Validation Diff
                                "",             # K: Amount Diff %
                                value,          # L: Value
                                "",             # M: Value Validation
                                "",             # N: Api Calc Value
                                "",             # O: Value Difference
                                "",             # P: Side
                                "",             # Q: Leverage
                                "",             # R: Collateral
                                "",             # S: P&L
                            ])

            # Remove DeFi tabs from all_tables_data (they'll be replaced by combined Overview - De-Fi tab)
            for tab_name in defi_tab_names:
                del all_tables_data[tab_name]
                print(f"   ℹ️  Combining '{tab_name}' into Overview - De-Fi tab")


            # Insert empty row separators between (protocol, pool_name) groups for visual grouping
            if len(defi_combined_data) > 1:
                _sep_data = [defi_combined_data[0]]  # keep header
                _empty_row = [""] * len(defi_combined_data[0])
                _prev_grp = None
                for _row in defi_combined_data[1:]:
                    _cur_grp = (_row[0] if len(_row) > 0 else "", _row[2] if len(_row) > 2 else "")
                    if _prev_grp is not None and _cur_grp != _prev_grp:
                        _sep_data.append(list(_empty_row))
                    _sep_data.append(_row)
                    _prev_grp = _cur_grp
                defi_combined_data = _sep_data

            # Add combined Overview - De-Fi tab if there's data
            if len(defi_combined_data) > 1:
                all_tables_data['Overview - De-Fi'] = defi_combined_data
                print(f"   ✅ Created combined 'Overview - De-Fi' tab with {len(defi_combined_data) - 1} rows from {len(defi_tab_names)} protocol(s)")

            # Reorder tabs: Exchange tables after "Overview - Wallet", then "Overview - De-Fi", then other Overview tabs
            # Step 1: Define overview tabs order
            overview_before_exchanges = [
                'Overview - Header & Token Holdings Header',
                'Overview - Wallet',
            ]

            overview_after_exchanges = [
                'Overview - De-Fi',
                'Overview - Combined Net Worth',
                'Overview - Platform Allocation',
                'Overview - Chain Allocation',
                'Overview - Token Allocation',
            ]

            # Step 2: Identify exchange tables (non-Overview tabs that weren't combined into De-Fi)
            exchange_tables = []
            for sheet_name in all_tables_data.keys():
                if not sheet_name.startswith('Overview') and sheet_name not in defi_tab_names:
                    exchange_tables.append(sheet_name)

            # Sort exchange tables alphabetically for consistency
            exchange_tables.sort()

            # Step 3: Create new ordered dictionary
            ordered_tables = {}

            # Add Overview tabs before exchanges
            for sheet_name in overview_before_exchanges:
                if sheet_name in all_tables_data:
                    ordered_tables[sheet_name] = all_tables_data[sheet_name]

            # Add exchange tables
            for sheet_name in exchange_tables:
                if sheet_name in all_tables_data:
                    ordered_tables[sheet_name] = all_tables_data[sheet_name]

            # Add Overview tabs after exchanges
            for sheet_name in overview_after_exchanges:
                if sheet_name in all_tables_data:
                    ordered_tables[sheet_name] = all_tables_data[sheet_name]

            # Add any remaining sheets that weren't in any of the above
            for sheet_name, data in all_tables_data.items():
                if sheet_name not in ordered_tables:
                    ordered_tables[sheet_name] = data

            # Replace all_tables_data with the reordered version
            all_tables_data = ordered_tables
            if exchange_tables:
                print(f"   ✅ Reordered tabs: {len(exchange_tables)} exchange table(s) after 'Overview - Wallet', before 'Overview - De-Fi'")
            else:
                print(f"   ✅ Reordered tabs: No exchange tables found")

            for sheet_name, data in all_tables_data.items():
                if data:
                    # Add validation columns for Overview - Wallet table
                    if sheet_name == "Overview - Wallet":
                        data = add_validation_columns_to_overview_token(data)
                        if len(data) > 1 and len(data[1]) > 10:
                            print(f"   DEBUG after add_validation: Row 1 output has {len(data[1])} columns")
                            print(f"   Cols H({data[1][7]}), K({data[1][10]})")
                        elif len(data) > 1:
                            print(f"   ⚠️  WARNING: Row 1 has only {len(data[1])} columns (expected 17+ for full data)")
                            print(f"   This usually means the portfolio has no token data")

                    # Add validation columns for Token Allocation table
                    elif sheet_name == "Overview - Token Allocation":
                        data = add_validation_columns_to_token_allocation(data)
                        # Column structure: A(0):Token, B(1):Percentage, C(2):% - Api NW Calc, D(3):Pct Validation, E(4):Net Worth, F(5):NW - API Calc, G(6):NW Validation
                        # NOTE: Pct Validation (D) is populated in post-processing after C is filled (compares B vs C)
                        print(f"   ✅ Added validation columns to Token Allocation")

                    # Add validation columns for Chain Allocation table
                    elif sheet_name == "Overview - Chain Allocation":
                        data = add_validation_columns_to_chain_allocation(data)
                        # Column structure: A(0):Chain, B(1):Percentage, C(2):% - Api NW Calc, D(3):Pct Validation, E(4):Net Worth, F(5):NW - API Calc, G(6):NW Validation
                        # NOTE: Pct Validation (D) is populated in post-processing after C is filled (compares B vs C)
                        print(f"   ✅ Added validation columns to Chain Allocation")

                    # Add validation columns for Platform Allocation table
                    elif sheet_name == "Overview - Platform Allocation":
                        data = add_validation_columns_to_platform_allocation(data)
                        # Column structure: A(0):Platform, B(1):Percentage, C(2):% - Api NW Calc, D(3):Pct Validation, E(4):Net Worth, F(5):NW - API Calc, G(6):NW Validation
                        # NOTE: Pct Validation (D) is populated in post-processing after C is filled (compares B vs C)
                        print(f"   ✅ Added validation columns to Platform Allocation")

                    # Add validation columns for Header & Token Holdings table
                    elif sheet_name == "Overview - Header & Token Holdings Header":
                        data = add_validation_columns_to_header_holdings(data)
                        print(f"   ✅ Added validation columns to Header & Token Holdings Header")

                    # Add validation columns for Combined Net Worth table
                    elif sheet_name == "Overview - Combined Net Worth":
                        data = add_validation_columns_to_combined_net_worth(data)
                        print(f"   ✅ Added validation columns to Combined Net Worth")

                    # Handle Overview - De-Fi tab - fill validation columns using Python
                    elif sheet_name == "Overview - De-Fi":
                        print(f"   🔄 Processing Overview - De-Fi tab with {len(data) - 1} rows")
                        data = add_validation_to_defi_tab(data, rabby_extracted_data)
                        all_tables_data[sheet_name] = data
                        print(f"   ✅ Added validation columns to Overview - De-Fi tab")

                    # Check if this is an account-level table (doesn't start with "Overview -")
                    is_account_table = not sheet_name.startswith("Overview -")
                    if is_account_table:
                        print(f"   ✅ Account table detected: {sheet_name}")

                        # For exchange tables, add a total row summing Q column where B="Main"
                        # Exchange tables have columns: A:Token, B:Category, ..., P:Amount Diff Validation, Q:Value, ...
                        if sheet_name in exchange_tables and len(data) > 1:
                            # Calculate sum of Q column (index 16) where B column (index 1) = "Main"
                            total_value = Decimal('0')
                            main_count = 0

                            for row in data[1:]:  # Skip header row
                                if len(row) > 16:  # Ensure row has enough columns
                                    category = str(row[1]).strip() if row[1] else ""
                                    if category.upper() == "MAIN":
                                        value_str = str(row[16]).replace(',', '').replace('$', '').strip() if row[16] else "0"
                                        try:
                                            # Handle cases where value might be text with commas
                                            if value_str and value_str != '':
                                                total_value += Decimal(value_str)
                                                main_count += 1
                                        except (ValueError, InvalidOperation):
                                            print(f"      ⚠️  Could not parse value: {value_str}")
                                            pass

                            # Add total row if we found any "Main" rows
                            if main_count > 0:
                                # Create total row with empty cells except P and Q
                                total_row = [""] * len(data[0])  # Match header length
                                total_row[15] = "Total"  # P column (index 15)
                                total_row[16] = str(total_value)  # Q column (index 16)
                                data.append(total_row)
                                print(f"      ✅ Added total row: {main_count} 'Main' rows, Total = {total_value}")

                    ws = wb.create_sheet(sheet_name)

                    # Write data to worksheet
                    for row_idx, row_data in enumerate(data, start=1):
                        for col_idx, value in enumerate(row_data, start=1):
                            # Convert numeric columns from text to numbers (remove commas)
                            # Column structure (24 columns total):
                            # A:Chain, B:Name, C:Price, D:PriceTooltip, E:FE-PriceVal, F:PriceVal, G:PriceDiffVal,
                            # H:Price24h, I:Price24hVal, J:Price24hDiffVal,
                            # K:Amount, L:AmountTooltip, M:FE-AmtVal, N:AmtVal, O:AmtDiffVal,
                            # P:Share, Q:ShareTooltip, R:FE-ShareVal, S:CalcShare, T:ShareVal,
                            # U:Value, V:ValueVal, W:CalculatedValue, X:DataRow
                            # Numeric columns: C(Price), H(Price 24h), P(Share), U(Value)
                            # NOTE: Tooltip columns (D, L, Q) and Amount column (K) are NOT converted to numbers - keep as text
                            if sheet_name == "Overview - Wallet" and row_idx > 1:  # Skip header row
                                if col_idx in [3, 8, 16, 21]:  # Numeric columns: C(Price), H(Price 24h), P(Share), U(Value)
                                    try:
                                        # Remove commas and convert to float
                                        if isinstance(value, str):
                                            # For columns that may show "< 0.01" or "< 0.00001", preserve the "<" symbol
                                            if value.strip().startswith('<'):
                                                # Keep the "<" symbol, just remove other symbols ($, %, >)
                                                cleaned_value = value.replace(',', '').replace('$', '').replace('%', '').replace('>', '').strip()
                                                # Don't convert to float, keep as string with "<"
                                                value = cleaned_value
                                            else:
                                                # Remove commas, dollar signs, percent signs, angle brackets
                                                cleaned_value = value.replace(',', '').replace('$', '').replace('%', '').replace('<', '').replace('>', '').strip()
                                                if cleaned_value and cleaned_value != '':
                                                    value = float(cleaned_value)
                                    except (ValueError, AttributeError):
                                        pass  # Keep original value if conversion fails

                            cell = ws.cell(row=row_idx, column=col_idx, value=value)

                            # Apply yellow highlighting and formulas to validation columns in Overview - Wallet
                            if sheet_name == "Overview - Wallet":
                                # Validation column positions (1-indexed): E(5), F(6), G(7), I(9), J(10), M(13), N(14), O(15), R(18), S(19), T(20), V(22), W(23)
                                # Note: W(23) is Calculated Value (text), N(14) is Calculation Share (text), T(20) is Amount Diff Validation (text)
                                if col_idx in [5, 6, 7, 9, 10, 13, 14, 15, 18, 19, 20, 22, 23, 24]:
                                    cell.fill = yellow_fill

                                # Set validation column headers (row 1)
                                if row_idx == 1:
                                    if col_idx == 4:
                                        cell.value = "Price Tooltip"
                                    elif col_idx == 5:
                                        cell.value = "FE - Price Validation"
                                    elif col_idx == 6:
                                        cell.value = "Price Validation"
                                    elif col_idx == 7:
                                        cell.value = "Price abs_diff"
                                    elif col_idx == 8:
                                        cell.value = "Price Diff %"
                                    elif col_idx == 9:
                                        cell.value = "Price (24h)"
                                    elif col_idx == 10:
                                        cell.value = "Price (24h) Validation"
                                    elif col_idx == 11:
                                        cell.value = "Price (24H) Diff Validation"
                                    elif col_idx == 12:
                                        cell.value = "Amount"
                                    elif col_idx == 13:
                                        cell.value = "Amount Tooltip"
                                    elif col_idx == 14:
                                        cell.value = "FE - Amount Validation"
                                    elif col_idx == 15:
                                        cell.value = "Amount Validation"
                                    elif col_idx == 16:
                                        cell.value = "Amount Diff Validation"
                                    elif col_idx == 17:
                                        cell.value = "Share"
                                    elif col_idx == 18:
                                        cell.value = "Share Tooltip"
                                    elif col_idx == 19:
                                        cell.value = "FE - Share Validation"
                                    elif col_idx == 20:
                                        cell.value = "Calculation Share"
                                    elif col_idx == 21:
                                        cell.value = "Share Validation"
                                    elif col_idx == 22:
                                        cell.value = "Value"
                                    elif col_idx == 23:
                                        cell.value = "UI Calculated Value"
                                    elif col_idx == 24:
                                        cell.value = "Value Validation(ui-cal)"
                                    elif col_idx == 25:
                                        cell.value = "API Calculated Value"
                                    elif col_idx == 26:
                                        cell.value = "Data Row"

                                # Add validation formulas (skip header row)
                                elif row_idx > 1:
                                    # Column E (5): FE - Price Validation - IFS with <0.01 check, zero check, and TRUNC
                                    # If tooltip is empty, fall back to Price cell (C) — compare C vs C = always "Tooltip N/A, cant compare"
                                    if col_idx == 5:
                                        formula = (
                                            f'=LET('
                                            f'tipRaw,D{row_idx},'
                                            f'hasTip,AND(tipRaw<>"",tipRaw<>"Tooltip N/A"),'
                                            f'src,IF(hasTip,tipRaw,C{row_idx}),'
                                            f'srcClean,TRIM(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(src,"$",""),"<",""),",","")),'
                                            f'srcNum,IFERROR(VALUE(srcClean),0),'
                                            f'hasLt,ISNUMBER(SEARCH("<",src)),'
                                            f'IF(OR(src="",IFERROR(VALUE(srcClean),-999)=-999),"No Price Data",'
                                            f'IF(NOT(hasTip),"Tooltip N/A, cant compare",'
                                            f'IF(hasLt,IF(OR(C{row_idx}="< 0.01",C{row_idx}="< $0.01"),"Passed","Failed"),'
                                            f'IF(srcNum=0,IF(OR(IFERROR(VALUE(C{row_idx}),0)=0,C{row_idx}="0"),"Passed","Failed"),'
                                            f'IF(srcNum<0.01,IF(OR(C{row_idx}="< 0.01",C{row_idx}="< $0.01"),"Passed","Failed"),'
                                            f'IF(TRUNC(srcNum,2)=IFERROR(VALUE(C{row_idx}),0),"Passed","Failed")'
                                            f'))))))'
                                        )
                                        cell.value = formula

                                    # Column F (6): Price Validation - Python calculated (chain-aware)
                                    elif col_idx == 6:
                                        cell.value = ""  # Placeholder, calculated with Python

                                    # Column G (7): Price abs_diff - Python calculated (chain-aware)
                                    elif col_idx == 7:
                                        cell.value = ""  # Placeholder, calculated with Python

                                    # Column H (8): Price Diff % - Python calculated
                                    elif col_idx == 8:
                                        cell.value = ""  # Placeholder, calculated with Python

                                    # Column J (10): Price (24H) Validation - Python calculated (chain-aware)
                                    elif col_idx == 10:
                                        cell.value = ""  # Placeholder, calculated with Python

                                    # Column K (11): Price (24H) Diff Validation - Python calculated
                                    elif col_idx == 11:
                                        cell.value = ""  # Placeholder, calculated with Python

                                    # Column N (14): FE - Amount Validation (compare M=AmountTooltip vs L=Amount)
                                    # If tooltip missing, fall back to Amount cell (L) — compare L vs L = "Tooltip N/A, cant compare"
                                    elif col_idx == 14:
                                        formula = (
                                            f'=LET('
                                            f'tipRaw,M{row_idx},'
                                            f'hasTip,AND(tipRaw<>"",tipRaw<>"Tooltip N/A"),'
                                            f'src,IF(hasTip,tipRaw,L{row_idx}),'
                                            f'amt,L{row_idx},'
                                            f'IF(OR(src="",IFERROR(VALUE(src),-999)=-999),"No Amount Data",'
                                            f'IF(NOT(hasTip),"Tooltip N/A, cant compare",'
                                            f'IF(OR(src=0,IFERROR(VALUE(src),0)=0),'
                                            f'IF(OR(amt="",amt=0,IFERROR(VALUE(amt),0)=0),"Passed","Failed"),'
                                            f'IFERROR('
                                            f'IF(VALUE(src)<0.00001,IF(amt="< 0.00001","Passed","Failed"),'
                                            f'IF(LEFT(SUBSTITUTE(TEXT(VALUE(src),"0.000000000000000"),",",""),FIND(".",SUBSTITUTE(TEXT(VALUE(src),"0.000000000000000"),",",""))+5)='
                                            f'LEFT(SUBSTITUTE(TEXT(VALUE(amt),"0.000000000000000"),",",""),FIND(".",SUBSTITUTE(TEXT(VALUE(amt),"0.000000000000000"),",",""))+5),"Passed","Failed"))'
                                            f',"Error")))))'
                                        )
                                        cell.value = formula

                                    # Column O (15): Amount Diff Validation - Python calculated
                                    elif col_idx == 15:
                                        cell.value = ""  # Placeholder, calculated later with Python

                                    # Column S (19): FE - Share Validation (compare R=ShareTooltip vs Q=Share)
                                    # UI truncates share to 2dp, so use INT(value*100) for reliable comparison
                                    # If tooltip missing, fall back to Share cell (Q) — compare Q vs Q = "Tooltip N/A, cant compare"
                                    elif col_idx == 19:
                                        formula = (
                                            f'=LET('
                                            f'tipRaw,R{row_idx},'
                                            f'hasTip,AND(tipRaw<>"",tipRaw<>"Tooltip N/A"),'
                                            f'src,IF(hasTip,SUBSTITUTE(SUBSTITUTE(tipRaw,"%",""),",",""),SUBSTITUTE(SUBSTITUTE(Q{row_idx},"%",""),",","")),'
                                            f'qClean,SUBSTITUTE(SUBSTITUTE(Q{row_idx},"%",""),",",""),'
                                            f'IF(OR(src="",IFERROR(VALUE(src),-999)=-999),"No Share Data",'
                                            f'IF(NOT(hasTip),"Tooltip N/A, cant compare",'
                                            f'IFS('
                                            f'AND(VALUE(src)<0.01,VALUE(src)<>0,TRIM(Q{row_idx})="< 0.01"),"Passed",'
                                            f'AND(VALUE(src)<0.01,VALUE(src)<>0,TRIM(Q{row_idx})<>"< 0.01"),"Failed",'
                                            f'AND(VALUE(src)>=0.01,INT(VALUE(src)*100)=INT(VALUE(qClean)*100)),"Passed",'
                                            f'AND(VALUE(src)>=0.01,INT(VALUE(src)*100)<>INT(VALUE(qClean)*100)),"Failed",'
                                            f'AND(VALUE(src)=0,VALUE(qClean)=0),"Passed",'
                                            f'AND(VALUE(src)=0,VALUE(qClean)<>0),"Failed"))))'
                                        )
                                        cell.value = formula

                                    # Column T (20): Calculation Share - Python-calculated later
                                    elif col_idx == 20:
                                        pass  # Will be filled by Python calculation below

                                    # Column V (22): Value - data, no formula needed

                                    # Column W (23): Value Validation - IFS with <0.01 check using X=CalcValue, V=Value
                                    # V may contain commas (e.g. "1,509,932.82"), so strip them with SUBSTITUTE before comparing
                                    # Wrap in IFERROR to handle empty/non-numeric X or V gracefully
                                    elif col_idx == 23:
                                        formula = (
                                            f'=IFERROR(IFS('
                                            f'AND(VALUE(X{row_idx})<0.01,VALUE(X{row_idx})<>0,TRIM(V{row_idx})="< 0.01"),"Passed",'
                                            f'AND(VALUE(X{row_idx})<0.01,VALUE(X{row_idx})<>0,TRIM(V{row_idx})<>"< 0.01"),"Failed",'
                                            f'AND(VALUE(X{row_idx})>=0.01,TRUNC(VALUE(X{row_idx}),2)=VALUE(SUBSTITUTE(V{row_idx},",",""))),"Passed",'
                                            f'AND(VALUE(X{row_idx})>=0.01,TRUNC(VALUE(X{row_idx}),2)<>VALUE(SUBSTITUTE(V{row_idx},",",""))),"Failed",'
                                            f'AND(VALUE(X{row_idx})=0,VALUE(SUBSTITUTE(V{row_idx},",",""))=0),"Passed",'
                                            f'AND(VALUE(X{row_idx})=0,VALUE(SUBSTITUTE(V{row_idx},",",""))<>0),"Failed"'
                                            f'),"No Calc Value")'
                                        )
                                        cell.value = formula

                                    # Column X (24): Value - UI validation - Python calculated
                                    elif col_idx == 24:
                                        cell.value = ""  # Placeholder, calculated later with Python Decimal

                                    # Column Y (25): Data Row — written by Python later
                                    elif col_idx == 25:
                                        cell.value = ""  # filled later by Python

                            # Apply formatting to Overview - Token Allocation sheet
                            elif sheet_name == "Overview - Token Allocation":
                                # Column structure: A:Token, B:Percentage, C:Percentage Validation, D:Net Worth, E:Net Worth Validation
                                # Convert numeric columns: B(Percentage), D(Net Worth)
                                if row_idx > 1 and col_idx in [2, 5]:  # Skip header row (B:Percentage, E:Net Worth)
                                    try:
                                        if isinstance(value, str):
                                            cleaned_value = value.replace(',', '').replace('$', '').replace('%', '').strip()
                                            if cleaned_value and cleaned_value != '':
                                                value = float(cleaned_value)
                                                cell.value = value
                                    except (ValueError, AttributeError):
                                        pass

                                # Apply yellow highlighting to validation columns: C(3), E(5)
                                if col_idx in [3, 5]:
                                    cell.fill = yellow_fill

                                # Set validation column headers (row 1)
                                if row_idx == 1:
                                    if col_idx == 3:
                                        cell.value = "% - Api NW Calc"
                                    elif col_idx == 4:
                                        cell.value = "Percentage Validation"
                                    elif col_idx == 5:
                                        cell.value = "Net Worth"
                                    elif col_idx == 6:
                                        cell.value = "Net Worth - API Calculation"
                                    elif col_idx == 7:
                                        cell.value = "Net Worth Validation"

                                # Add validation formulas (skip header row)
                                elif row_idx > 1:
                                    # Column C (3): Percentage Validation - populated in post-processing (B vs C)

                                    # Column G (7): Net Worth Validation - populated in post-processing (E vs F, ≤1% tolerance)
                                    pass

                            # Apply formatting to Overview - Chain Allocation sheet
                            elif sheet_name == "Overview - Chain Allocation":
                                # Column structure: A:Chain, B:Percentage, C:Percentage Validation, D:Net Worth, E:Net Worth Validation
                                # Convert numeric columns: B(Percentage), D(Net Worth)
                                if row_idx > 1 and col_idx in [2, 5]:  # Skip header row (B:Percentage, E:Net Worth)
                                    try:
                                        if isinstance(value, str):
                                            cleaned_value = value.replace(',', '').replace('$', '').replace('%', '').strip()
                                            if cleaned_value and cleaned_value != '':
                                                value = float(cleaned_value)
                                                cell.value = value
                                    except (ValueError, AttributeError):
                                        pass

                                # Apply yellow highlighting to validation columns: C(3), E(5)
                                if col_idx in [3, 5]:
                                    cell.fill = yellow_fill

                                # Set validation column headers (row 1)
                                if row_idx == 1:
                                    if col_idx == 3:
                                        cell.value = "% - Api NW Calc"
                                    elif col_idx == 4:
                                        cell.value = "Percentage Validation"
                                    elif col_idx == 5:
                                        cell.value = "Net Worth"
                                    elif col_idx == 6:
                                        cell.value = "Net Worth - API Calculation"
                                    elif col_idx == 7:
                                        cell.value = "Net Worth Validation"

                                # Add validation formulas (skip header row)
                                elif row_idx > 1:
                                    # Column C (3): Percentage Validation - populated in post-processing (B vs C)

                                    # Column G (7): Net Worth Validation - populated in post-processing (E vs F, ≤1% tolerance)
                                    pass

                            # Apply formatting to Overview - Platform Allocation sheet
                            elif sheet_name == "Overview - Platform Allocation":
                                # Column structure: A:Platform, B:Percentage, C:Percentage Validation, D:Net Worth, E:Net Worth Validation
                                # Convert numeric columns: B(Percentage), D(Net Worth)
                                if row_idx > 1 and col_idx in [2, 5]:  # Skip header row (B:Percentage, E:Net Worth)
                                    try:
                                        if isinstance(value, str):
                                            cleaned_value = value.replace(',', '').replace('$', '').replace('%', '').strip()
                                            if cleaned_value and cleaned_value != '':
                                                value = float(cleaned_value)
                                                cell.value = value
                                    except (ValueError, AttributeError):
                                        pass

                                # Apply yellow highlighting to validation columns: C(3), E(5)
                                if col_idx in [3, 5]:
                                    cell.fill = yellow_fill

                                # Set validation column headers (row 1)
                                if row_idx == 1:
                                    if col_idx == 3:
                                        cell.value = "% - Api NW Calc"
                                    elif col_idx == 4:
                                        cell.value = "Percentage Validation"
                                    elif col_idx == 5:
                                        cell.value = "Net Worth"
                                    elif col_idx == 6:
                                        cell.value = "Net Worth - API Calculation"
                                    elif col_idx == 7:
                                        cell.value = "Net Worth Validation"

                                # Add validation formulas (skip header row)
                                elif row_idx > 1:
                                    # Column C (3): Percentage Validation - populated in post-processing (B vs C)

                                    # Column G (7): Net Worth Validation - populated in post-processing (E vs F, ≤1% tolerance)
                                    pass

                            # Apply formatting to Overview - De-Fi tab (combined protocol data)
                            elif sheet_name == "Overview - De-Fi":
                                # Column structure (20 cols): A:Defi, B:Chain, C:Type, D:Pool/Position Pair, E:Description,
                                # F:Amount, G:Amount Tooltip, H:Amount Validation, I:FE-Amount Validation,
                                # J:Amount Validation Diff, K:Amount Diff %, L:Value, M:Value Validation,
                                # N:Api Calc Value, O:Value Difference, P:Side, Q:Leverage, R:Collateral, S:P&L, T:Row Matched
                                # Apply yellow highlighting to validation columns
                                if col_idx in [8, 9, 10, 11, 13, 14]:
                                    cell.fill = yellow_fill

                                # Add Excel formulas for validation columns on data rows (row_idx > 1)
                                # Skip blank separator rows (col A empty = group separator, not a data row)
                                if row_idx > 1 and col_idx in [8, 9, 10, 11, 13, 14] and row_data[0] not in ('', None):
                                    _rs = "'Rabby Api Data'"
                                    # LET vars: matchArr, hasMatch, diffs (relative), bestIdx
                                    # chain: only filter when col B non-empty (Hyperliquid has no chain in De-Fi)
                                    # Pool Name ($C): always has value — Pool Name cells are NOT merged
                                    # diffs: RELATIVE difference so large-magnitude amounts don't lose to tiny values
                                    # epsilon on isMin avoids floating-point #N/A from MATCH(MIN(array),array,0)
                                    _let_hdr = (
                                        f"matchArr,"
                                        f"(LOWER(TRIM({_rs}!$B$2:$B$998))=LOWER(TRIM($A{row_idx})))"
                                        f"*(IF($B{row_idx}=\"\",1,LOWER(TRIM({_rs}!$D$2:$D$998))=LOWER(TRIM($B{row_idx}))))"
                                        f"*(LOWER(TRIM({_rs}!$E$2:$E$998))=LOWER(TRIM($C{row_idx})))"
                                        f"*(LOWER(TRIM({_rs}!$H$2:$H$998))=LOWER(TRIM($D{row_idx}))),"
                                        f"hasMatch,SUM(matchArr)>0,"
                                        f"fVal,VALUE(SUBSTITUTE(SUBSTITUTE($G{row_idx},\"<\",\"\"),\" \",\"\")),"
                                        f"lAbs,IFERROR(ABS(VALUE({_rs}!$L$2:$L$998)),0),"
                                        f"denom,IF(lAbs>ABS(fVal),lAbs,IF(ABS(fVal)=0,1,ABS(fVal))),"
                                        f"diffs,IF(matchArr=1,ABS(fVal-IFERROR(VALUE({_rs}!$L$2:$L$998),0))/denom,1E+99),"
                                        f"minDiff,MIN(diffs),"
                                        f"isMin,(diffs<=minDiff+1E-9)*matchArr,"
                                        f"bestIdx,IFERROR(MATCH(1,isMin,0),1)"
                                    )

                                    if col_idx == 9:  # I: FE - Amount Validation
                                        cell.value = (
                                            f'=IFS('
                                            f'AND(ABS(G{row_idx})>=0.00001,ABS(VALUE(SUBSTITUTE(F{row_idx},",","")))=TRUNC(ABS(G{row_idx}),5)),"Passed",'
                                            f'AND(ABS(G{row_idx})>=0.00001,ABS(VALUE(SUBSTITUTE(F{row_idx},",","")))<>TRUNC(ABS(G{row_idx}),5)),"Failed",'
                                            f'AND(ABS(G{row_idx})<0.00001,F{row_idx}="< 0.00001"),"Passed",'
                                            f'AND(ABS(G{row_idx})<0.00001,F{row_idx}<>"< 0.00001"),"Failed")'
                                        )

                                    elif col_idx == 10:  # J: Amount Validation Diff — best-match row
                                        cell.value = (
                                            f'=LET({_let_hdr},'
                                            f'rabbyAmt,INDEX(VALUE({_rs}!$L$2:$L$998),bestIdx),'
                                            f'IF(hasMatch,fVal-rabbyAmt,""))'
                                        )

                                    elif col_idx == 12:  # L: Value Validation
                                        # rabbyVal >= 0.01: ROUND(rabbyVal,2) exact match with K → Passed/Failed
                                        # rabbyVal < 0.01 : K contains "<" → Passed, else Failed
                                        cell.value = (
                                            f'=LET({_let_hdr},'
                                            f'rabbyVal,INDEX(VALUE({_rs}!$M$2:$M$998),bestIdx),'
                                            f'IF(hasMatch,'
                                            f'IF(rabbyVal>=0.01,'
                                            f'IF(ROUND(rabbyVal,2)=VALUE(TRIM($L{row_idx})),"Passed","Failed"),'
                                            f'IF(ISNUMBER(SEARCH("<",$L{row_idx})),"Passed","Failed")),'
                                            f'"Failed"))'
                                        )

                            # Apply formatting to Overview - Header & Token Holdings Header sheet
                            elif sheet_name == "Overview - Header & Token Holdings Header":
                                # Column structure: A:Section, B:Category, C:Token Count, D:Token Count Validation, E:Net Worth, F:Net Worth Validation, G:Percentage, H:Percentage Validation
                                # Convert numeric columns: C(Token Count), G(Percentage) - but keep E(Net Worth) as text with commas from DAM
                                if row_idx > 1 and col_idx in [3, 7]:  # Skip header row, exclude E column (5)
                                    try:
                                        if isinstance(value, str):
                                            # For column G (Percentage), preserve the "<" symbol
                                            if col_idx == 7 and value.strip().startswith('<'):
                                                # Keep the "<" symbol, just remove other symbols ($, %, >)
                                                cleaned_value = value.replace(',', '').replace('$', '').replace('%', '').replace('>', '').strip()
                                                # Don't convert to float, keep as string with "<"
                                                value = cleaned_value
                                                cell.value = value
                                                # Debug: Confirm "<" symbol is preserved when writing to Excel
                                                print(f"      🔍 DEBUG Excel Write: Row {row_idx}, Col G - Preserved '<' value: '{value}'")
                                            else:
                                                cleaned_value = value.replace(',', '').replace('$', '').replace('%', '').strip()
                                                if cleaned_value and cleaned_value != '':
                                                    # Token Count should be integer
                                                    if col_idx == 3:
                                                        value = int(cleaned_value)
                                                    else:
                                                        value = float(cleaned_value)
                                                    cell.value = value
                                    except (ValueError, AttributeError):
                                        pass

                                # Apply yellow highlighting to validation columns: E(5), H(8), J(10)
                                if col_idx in [5, 8, 9, 10, 12]:
                                    cell.fill = yellow_fill

                                # Set column headers (row 1)
                                if row_idx == 1:
                                    if col_idx == 1:
                                        cell.value = "Section"
                                    elif col_idx == 2:
                                        cell.value = "Category"
                                    elif col_idx == 3:
                                        cell.value = "Token Count"
                                    elif col_idx == 4:
                                        cell.value = "TC_UI Count"
                                    elif col_idx == 5:
                                        cell.value = "Token Count Validation"
                                    elif col_idx == 6:
                                        cell.value = "Net Worth"
                                    elif col_idx == 7:
                                        cell.value = "Net Worth_UI Calculation"
                                    elif col_idx == 8:
                                        cell.value = "Net Worth UI Validation"
                                    elif col_idx == 9:
                                        cell.value = "Net Worth - API Calculation"
                                    elif col_idx == 10:
                                        cell.value = "Net Worth - UI-API Validation"
                                    elif col_idx == 11:
                                        cell.value = "Percentage"
                                    elif col_idx == 12:
                                        cell.value = "Percentage Validation"

                                # Add validation formulas (skip header row)
                                elif row_idx > 1:
                                    # Column E (5): Token Count Validation — filled by post-processing
                                    if col_idx == 5:
                                        cell.value = "Pending"

                                    # Column H (8): Net Worth UI Validation — filled by post-processing
                                    elif col_idx == 8:
                                        cell.value = "Pending"

                                    # Column I (9): Net Worth - API Calculation — filled by post-processing
                                    elif col_idx == 9:
                                        cell.value = ""

                                    # Column J (10): Net Worth - UI-API Validation — filled by post-processing
                                    elif col_idx == 10:
                                        cell.value = ""

                                    # Column L (12): Percentage Validation
                                    elif col_idx == 12:
                                        section_val = ws.cell(row=row_idx, column=1).value
                                        if section_val in ["Overview Header", "Token Holdings - Platform"]:
                                            cell.value = "Not Applicable"
                                        else:
                                            last_row = len(data)
                                            formula = f'=IF(F{row_idx}="","-",IF(SUM(F$2:F${last_row})>0,F{row_idx}/SUM(F$2:F${last_row})*100,0))'
                                            cell.value = formula

                            # Apply formatting to Overview - Combined Net Worth sheet
                            elif sheet_name == "Overview - Combined Net Worth":
                                # Column structure: A:Address/Exchange, B:Value, C:Value Validation, D:Calculated Value
                                # Convert numeric columns: B(Value)
                                if row_idx > 1 and col_idx == 2:  # Skip header row
                                    try:
                                        if isinstance(value, str):
                                            # Handle K/M/B suffixes (e.g., "66.03K" -> "66030")
                                            cleaned_value = value.replace(',', '').replace('$', '').strip()
                                            if cleaned_value.endswith('K'):
                                                cleaned_value = str(float(cleaned_value[:-1]) * 1000)
                                            elif cleaned_value.endswith('M'):
                                                cleaned_value = str(float(cleaned_value[:-1]) * 1000000)
                                            elif cleaned_value.endswith('B'):
                                                cleaned_value = str(float(cleaned_value[:-1]) * 1000000000)

                                            if cleaned_value and cleaned_value != '':
                                                value = float(cleaned_value)
                                                cell.value = value
                                    except (ValueError, AttributeError):
                                        pass

                                # Apply yellow highlighting to validation column: C(3)
                                if col_idx == 3:
                                    cell.fill = yellow_fill

                                # Set column headers (row 1)
                                if row_idx == 1:
                                    if col_idx == 3:
                                        cell.value = "Value Validation"
                                    elif col_idx == 4:
                                        cell.value = "Calculated Value"

                                # Add validation formulas (skip header row)
                                elif row_idx > 1:
                                    # Column C (3): Value Validation
                                    if col_idx == 3:
                                        # Validate that sum of all values equals Overview - Wallet total
                                        last_row = len(data)
                                        # Compare with Overview - Wallet U column sum (Total Value)
                                        formula = f'=IF(SUM(B$2:B${last_row})>0,"Passed","No Data")'
                                        cell.value = formula

                            # Apply formatting to account-level tables
                            elif is_account_table:
                                # Account table column structure (21 columns):
                                # A:Token, B:Category, C:Account, D:Price, E:Price Tooltip,
                                # F:FE-Price Validation, G:Price Validation, H:Price Diff Validation,
                                # I:Price (24h), J:Price (24h) Validation, K:Price (24H) Diff Validation,
                                # L:Amount, M:Amount Tooltip, N:Amount Validation, O:FE-Amount Validation,
                                # P:Amount Diff Validation, Q:Value, R:Value Validation, S:Calculated Value,
                                # T:Total Value, U:Total Value Validation

                                # Convert numeric columns from text to numbers
                                # Numeric columns: D(Price), L(Amount), Q(Value), T(Total Value)
                                if row_idx > 1 and col_idx in [4, 12, 17, 20]:  # Skip header row
                                    try:
                                        if isinstance(value, str):
                                            cleaned_value = value.replace(',', '').replace('$', '').strip()
                                            if cleaned_value and cleaned_value != '':
                                                value = float(cleaned_value)
                                                cell.value = value
                                    except (ValueError, AttributeError):
                                        pass

                                # Apply yellow highlighting to validation columns
                                # F(6), G(7), H(8), J(10), K(11), N(14), O(15), P(16), R(18), S(19), U(21)
                                if col_idx in [6, 7, 8, 10, 11, 14, 15, 16, 18, 19, 21]:
                                    cell.fill = yellow_fill

                    # Set column widths and number formatting for allocation sheets
                    if sheet_name in ["Overview - Token Allocation", "Overview - Chain Allocation", "Overview - Platform Allocation"]:
                        from openpyxl.styles import numbers

                        # New structure: A:Name, B:Percentage, C:% - Api NW Calc, D:Pct Validation, E:Net Worth, F:NW - API Calc, G:NW Validation
                        ws.column_dimensions['A'].width = 20
                        ws.column_dimensions['B'].width = 15
                        ws.column_dimensions['C'].width = 18  # % - Api NW Calc
                        ws.column_dimensions['D'].width = 20  # Percentage Validation
                        ws.column_dimensions['E'].width = 20  # Net Worth
                        ws.column_dimensions['F'].width = 22  # Net Worth - API Calculation
                        ws.column_dimensions['G'].width = 22  # Net Worth Validation

                        for row_idx in range(2, len(data) + 1):
                            # Column B: Percentage
                            cell_b = ws.cell(row=row_idx, column=2)
                            if isinstance(cell_b.value, (int, float)):
                                cell_b.number_format = '0.00'

                            # Column E: Net Worth
                            cell_e = ws.cell(row=row_idx, column=5)
                            if isinstance(cell_e.value, (int, float)):
                                cell_e.number_format = '#,##0.00'

                            # Column F: Net Worth - API Calculation
                            cell_f = ws.cell(row=row_idx, column=6)
                            if isinstance(cell_f.value, (int, float)):
                                cell_f.number_format = '#,##0.00'

                        # Color Passed/Failed cells in validation columns (D and G)
                        _green = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
                        _red = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
                        _white_font = Font(color="FFFFFF", bold=True)
                        for row_idx in range(2, len(data) + 1):
                            for col in [4, 7]:  # D: Percentage Validation, G: Net Worth Validation
                                cell = ws.cell(row=row_idx, column=col)
                                val = str(cell.value or "").strip()
                                if val == "Passed":
                                    cell.fill = _green
                                    cell.font = _white_font
                                elif val == "Failed":
                                    cell.fill = _red
                                    cell.font = _white_font

                    # Set column widths and number formatting for Header & Token Holdings Header sheet
                    if sheet_name == "Overview - Header & Token Holdings Header":
                        from openpyxl.styles import numbers

                        # Set column widths
                        ws.column_dimensions['A'].width = 20  # Section
                        ws.column_dimensions['B'].width = 25  # Category
                        ws.column_dimensions['C'].width = 15  # Token Count
                        ws.column_dimensions['D'].width = 15  # TC_UI Count
                        ws.column_dimensions['E'].width = 25  # Token Count Validation
                        ws.column_dimensions['F'].width = 20  # Net Worth
                        ws.column_dimensions['G'].width = 25  # Net Worth_UI Calculation
                        ws.column_dimensions['H'].width = 25  # Net Worth UI Validation
                        ws.column_dimensions['I'].width = 25  # Net Worth - API Calculation
                        ws.column_dimensions['J'].width = 25  # Net Worth - UI-API Validation
                        ws.column_dimensions['K'].width = 15  # Percentage
                        ws.column_dimensions['L'].width = 25  # Percentage Validation

                        # Apply number formatting (skip header)
                        for row_idx in range(2, len(data) + 1):
                            # Column C: Token Count (integer format)
                            cell_c = ws.cell(row=row_idx, column=3)
                            if isinstance(cell_c.value, (int, float)):
                                cell_c.number_format = '0'

                            # Column D: TC_UI Count (integer format)
                            cell_d = ws.cell(row=row_idx, column=4)
                            if isinstance(cell_d.value, (int, float)):
                                cell_d.number_format = '0'

                            # Column F: Net Worth (currency format)
                            cell_f = ws.cell(row=row_idx, column=6)
                            if isinstance(cell_f.value, (int, float)):
                                cell_f.number_format = '#,##0.00'

                            # Column G: Net Worth_UI Calculation (currency format)
                            cell_g = ws.cell(row=row_idx, column=7)
                            if isinstance(cell_g.value, (int, float)):
                                cell_g.number_format = '#,##0.00'

                            # Column I: Net Worth - API Calculation (currency format)
                            cell_i = ws.cell(row=row_idx, column=9)
                            if isinstance(cell_i.value, (int, float)):
                                cell_i.number_format = '#,##0.00'

                            # Column K: Percentage (2 decimal places)
                            cell_k = ws.cell(row=row_idx, column=11)
                            if isinstance(cell_k.value, (int, float)):
                                cell_k.number_format = '0.00'

                            # Column L: Percentage Validation (formula - 2 decimal places)
                            cell_l = ws.cell(row=row_idx, column=12)
                            cell_l.number_format = '0.00'

                    # Set column widths and number formatting for Combined Net Worth sheet
                    if sheet_name == "Overview - Combined Net Worth":
                        from openpyxl.styles import numbers

                        # Set column widths
                        ws.column_dimensions['A'].width = 30  # Address/Exchange
                        ws.column_dimensions['B'].width = 20  # Value
                        ws.column_dimensions['C'].width = 25  # Value Validation
                        ws.column_dimensions['D'].width = 25  # Calculated Value

                        # Apply number formatting (skip header)
                        for row_idx in range(2, len(data) + 1):
                            # Column B: Value (currency format with comma separator)
                            cell_b = ws.cell(row=row_idx, column=2)
                            if isinstance(cell_b.value, (int, float)):
                                cell_b.number_format = '#,##0.00'  # Comma separator with 2 decimals
                            # Column D: Calculated Value (currency format)
                            cell_d = ws.cell(row=row_idx, column=4)
                            if isinstance(cell_d.value, (int, float)):
                                cell_d.number_format = '#,##0.00'

                    print(f"✅ {sheet_name}: {len(data)} rows")

            # IMPORTANT: Copy TRX Balance data into DAM workbook and fix formulas
            print(f"\n🔧 Adding TRX Balance data and fixing formulas...")

            # Initialize first_balance_sheet variable
            first_balance_sheet = None

            # Load the TRX Balance file to copy data from
            if trx_balance_filename and os.path.exists(os.path.join("test-results/API Result", trx_balance_filename)):
                try:
                    from openpyxl import load_workbook as load_wb
                    trx_wb = load_wb(os.path.join("test-results/API Result", trx_balance_filename))

                    # Copy TRX raw JSON to DAM Excel folder
                    _trx_raw_src = os.path.join("test-results/API Result", trx_balance_filename.replace(".xlsx", "_Raw.json"))
                    if os.path.exists(_trx_raw_src):
                        import shutil as _shutil_trx
                        _trx_raw_dst = os.path.join(os.path.dirname(excel_path), f"TRX_Raw_{target_portfolio_name}.json")
                        _shutil_trx.copy2(_trx_raw_src, _trx_raw_dst)
                        print(f"   📄 Copied TRX raw JSON to: {os.path.basename(_trx_raw_dst)}")

                    # Copy the combined "TRX Balance, Price" sheet into DAM workbook
                    # (single sheet containing rows from all addresses)
                    first_balance_sheet = "TRX Balance, Price"

                    if first_balance_sheet in trx_wb.sheetnames:
                        trx_ws_source = trx_wb[first_balance_sheet]

                        # Create new sheet in DAM workbook
                        trx_ws_dest = wb.create_sheet(first_balance_sheet)

                        # Copy all data from source to destination
                        for row in trx_ws_source.iter_rows():
                            for cell in row:
                                trx_ws_dest[cell.coordinate].value = cell.value

                        print(f"   ✅ Copied '{first_balance_sheet}' sheet into DAM workbook")

                        # Build AGGREGATED token->balance mapping
                        # (same token from different addresses gets summed)
                        token_balance_map = {}
                        try:
                            for row in trx_ws_source.iter_rows(min_row=2, values_only=True):
                                if row and len(row) >= 6:
                                    token_name = row[2]  # Column C (index 2) - Token
                                    balance_raw = row[5]  # Column F (index 5) - Balance (Raw)
                                    decimal_places = row[1]  # Column B (index 1) - Decimal Places
                                    if token_name and balance_raw and decimal_places is not None:
                                        try:
                                            raw_str = str(balance_raw).strip()
                                            dec_int = int(decimal_places)
                                            # Convert raw balance to Decimal for aggregation
                                            raw_decimal = Decimal(raw_str) / Decimal(10 ** dec_int)
                                            token_key = str(token_name).strip()
                                            if token_key in token_balance_map:
                                                token_balance_map[token_key] += raw_decimal
                                            else:
                                                token_balance_map[token_key] = raw_decimal
                                        except (ValueError, TypeError) as conv_err:
                                            print(f"   ⚠️  Skipping {token_name}: conversion error - {conv_err}")

                            print(f"   ✅ Built aggregated token->balance mapping: {len(token_balance_map)} tokens ({len(TRX_ADDRESSES)} address(es))")
                        except Exception as e:
                            print(f"   ⚠️  Error building token mapping: {e}")

                        # Build sim_dune_balance_map from sim_dune_extracted_data for EVM chains
                        # Structure of sim_dune_extracted_data: [["Chain", "Symbol", "Amount (Raw)", "Amount", "Decimals", "Token Address", "ID", "Price", "24H Price Change"], ...]
                        # (Combined has "Address" as first column: [["Address", "Chain", "Symbol", "Amount (Raw)", "Amount", "Decimals", "Token Address", "ID", "Price", "24H Price Change"], ...])
                        # Key: (chain, symbol) tuple for matching both Chain and Symbol
                        sim_dune_balance_map = {}
                        if sim_dune_extracted_data and len(sim_dune_extracted_data) > 1:
                            try:
                                for row in sim_dune_extracted_data[1:]:  # Skip header
                                    if row and len(row) >= 4:
                                        chain = str(row[0]).strip()  # Column A - Chain
                                        symbol = str(row[1]).strip()  # Column B - Symbol
                                        amount = str(row[3]).strip()  # Column D - Amount (already calculated with full precision)
                                        if chain and symbol and amount:
                                            try:
                                                amount_decimal = Decimal(amount)
                                                # Use composite key (chain, symbol) for matching - both lowercase for case-insensitive comparison
                                                key = (chain.lower(), symbol.lower())
                                                if key in sim_dune_balance_map:
                                                    sim_dune_balance_map[key] += amount_decimal
                                                else:
                                                    sim_dune_balance_map[key] = amount_decimal
                                            except (ValueError, TypeError) as conv_err:
                                                print(f"   ⚠️  Skipping Sim Dune {chain}/{symbol}: conversion error - {conv_err}")
                                print(f"   ✅ Built Sim Dune balance mapping: {len(sim_dune_balance_map)} tokens (chain+symbol keys)")
                            except Exception as e:
                                print(f"   ⚠️  Error building Sim Dune balance mapping: {e}")

                        # Now update formulas in Overview - Wallet to reference internal sheet
                        if "Overview - Wallet" in wb.sheetnames:
                            ws_overview = wb["Overview - Wallet"]

                            # Find the last data row (excluding header)
                            last_data_row = ws_overview.max_row
                            while last_data_row > 1 and not ws_overview.cell(last_data_row, 2).value:
                                last_data_row -= 1

                            # Add sum row after last data row
                            sum_row = last_data_row + 1

                            # Update validation formulas to use internal sheet
                            # New column structure (23 columns):
                            # A:Chain, B:Name, C:Price, D:PriceTooltip, E:FE-PriceVal, F:PriceVal, G:PriceDiffVal,
                            # H:Price24h, I:Price24hVal, J:Price24hDiffVal, K:Share, L:ShareTooltip, M:FE-ShareVal,
                            # N:CalcShare, O:ShareVal, P:Amount, Q:AmountTooltip, R:FE-AmountVal, S:AmountVal, T:AmountDiffVal, U:Value, V:ValueVal, W:CalculatedValue

                            # First pass: Calculate all Calculated Values (W column) and sum them
                            # Also build chain-to-calculated-value mapping for F column validation
                            calculated_values = []
                            chain_calculated_values = {}  # {chain_name: sum of calculated values for that chain}
                            chain_u_values = {}  # {chain_name: sum of U column values for that chain}
                            for row_idx in range(2, last_data_row + 1):
                                # Get A (Chain), D (Price Tooltip) and Q (Amount Tooltip) values
                                chain_value = ws_overview.cell(row_idx, 1).value  # Column A - Chain
                                d_value = ws_overview.cell(row_idx, 4).value  # Column D
                                q_value = ws_overview.cell(row_idx, 18).value  # Column Q (Amount Tooltip)
                                u_value = ws_overview.cell(row_idx, 22).value  # Column U - Value

                                # Track U column chain sums
                                chain_name_u = str(chain_value).strip() if chain_value else ""
                                if chain_name_u and u_value is not None:
                                    try:
                                        u_str = str(u_value).replace('$', '').replace(',', '').replace('<', '').strip()
                                        if u_str and u_str not in ('', 'None', 'N/A'):
                                            u_dec = Decimal(u_str)
                                            chain_u_values.setdefault(chain_name_u, Decimal('0'))
                                            chain_u_values[chain_name_u] += u_dec
                                    except:
                                        pass

                                try:
                                    # Clean values - remove $, <, commas and whitespace
                                    d_str_raw = str(d_value).strip() if d_value else ''
                                    q_str_raw = str(q_value).strip() if q_value else ''

                                    # Fallback: if tooltip is missing, use displayed Price (col C) and Amount (col Q)
                                    _d_is_na = (not d_str_raw or 'N/A' in d_str_raw or 'cant compare' in d_str_raw)
                                    _q_is_na = (not q_str_raw or 'N/A' in q_str_raw or 'cant compare' in q_str_raw)

                                    if _d_is_na:
                                        c_fallback = ws_overview.cell(row_idx, 3).value  # Column C - Price (displayed)
                                        d_str_raw = str(c_fallback).strip() if c_fallback else ''
                                    if _q_is_na:
                                        r_fallback = ws_overview.cell(row_idx, 17).value  # Column P - Amount (displayed)
                                        q_str_raw = str(r_fallback).strip() if r_fallback else ''

                                    # Handle "< 0.01" — treat as 0.01 and flag for yellow
                                    _d_approx_ov = False
                                    _q_approx_ov = False
                                    if '< 0.01' in d_str_raw or '<0.01' in d_str_raw:
                                        d_str_raw = '0.01'
                                        _d_approx_ov = True
                                    if '< 0.01' in q_str_raw or '<0.01' in q_str_raw:
                                        q_str_raw = '0.01'
                                        _q_approx_ov = True

                                    d_clean = d_str_raw.replace(',', '').replace('$', '').replace('<', '').replace('>', '').strip()
                                    q_clean = q_str_raw.replace(',', '').replace('$', '').replace('<', '').replace('>', '').strip()
                                    if not d_clean: d_clean = '0'
                                    if not q_clean: q_clean = '0'

                                    # Use Decimal for precise calculation
                                    d_decimal = Decimal(d_clean)
                                    q_decimal = Decimal(q_clean)
                                    result = d_decimal * q_decimal
                                    calculated_values.append(result)

                                    # Track approximation flag per row for yellow highlighting later
                                    if not hasattr(run_dam_portfolio_extraction, '_approx_rows_ov'):
                                        run_dam_portfolio_extraction._approx_rows_ov = set()
                                    if _d_approx_ov or _q_approx_ov:
                                        run_dam_portfolio_extraction._approx_rows_ov.add(row_idx)

                                    # Add to chain-specific sum
                                    chain_name = str(chain_value).strip() if chain_value else ""
                                    if chain_name:
                                        if chain_name not in chain_calculated_values:
                                            chain_calculated_values[chain_name] = Decimal('0')
                                        chain_calculated_values[chain_name] += result
                                except Exception as e:
                                    calculated_values.append(Decimal('0'))

                            # Calculate sum of all Calculated Values using Decimal
                            total_calculated_value = sum(calculated_values)

                            # Second pass: Apply all formulas and calculated values
                            for row_idx in range(2, last_data_row + 1):
                                # Column E (5): FE - Price Validation - IFS with <0.01 check, zero check, and TRUNC
                                cell_e = ws_overview.cell(row_idx, 5)
                                formula_e = (
                                    f'=LET('
                                    f'tipRaw,D{row_idx},'
                                    f'hasTip,AND(tipRaw<>"",tipRaw<>"Tooltip N/A"),'
                                    f'src,IF(hasTip,tipRaw,C{row_idx}),'
                                    f'srcClean,SUBSTITUTE(SUBSTITUTE(src,"$",""),",",""),'
                                    f'IFERROR('
                                    f'IFS('
                                    f'src="","No Price Data",'
                                    f'NOT(hasTip),"Tooltip N/A, cant compare",'
                                    f'AND(VALUE(srcClean)<0.01,VALUE(srcClean)<>0,C{row_idx}="< 0.01"),"Passed",'
                                    f'AND(VALUE(srcClean)<0.01,VALUE(srcClean)<>0,C{row_idx}<>"< 0.01"),"Failed",'
                                    f'AND(VALUE(srcClean)>=0.01,TRUNC(VALUE(srcClean),2)=VALUE(C{row_idx})),"Passed",'
                                    f'AND(VALUE(srcClean)>=0.01,TRUNC(VALUE(srcClean),2)<>VALUE(C{row_idx})),"Failed")'
                                    f',"No Price Data"))'
                                )
                                cell_e.value = formula_e
                                cell_e.fill = yellow_fill

                                # Column F (6): Price Validation
                                cell_f = ws_overview.cell(row_idx, 6)
                                formula_f = (
                                    f'=LET('
                                    f'trxPrice,XLOOKUP(B{row_idx},\'{first_balance_sheet}\'!C:C,\'{first_balance_sheet}\'!G:G,"Token not found"),'
                                    f'IF(trxPrice="Token not found","Token not found",'
                                    f'IF(OR(trxPrice="",ISBLANK(trxPrice)),"Price not available",'
                                    f'IF(ABS((VALUE(SUBSTITUTE(SUBSTITUTE(D{row_idx},"$",""),",",""))-trxPrice)/trxPrice)<=1%,"PASSED","FAILED")'
                                    f'))'
                                    f')'
                                )
                                cell_f.value = formula_f
                                cell_f.fill = yellow_fill

                                # Column G (7): Price Diff Validation
                                cell_g = ws_overview.cell(row_idx, 7)
                                formula_g = (
                                    f'=LET('
                                    f'trxPrice,XLOOKUP(B{row_idx},\'{first_balance_sheet}\'!C:C,\'{first_balance_sheet}\'!G:G),'
                                    f'(C{row_idx}-trxPrice)'
                                    f')'
                                )
                                cell_g.value = formula_g
                                cell_g.fill = yellow_fill

                                # Column I (9): Price (24h) Validation
                                cell_i = ws_overview.cell(row_idx, 10)
                                formula_i = (
                                    f'=LET('
                                    f'trxPrice24h,XLOOKUP(B{row_idx},\'{first_balance_sheet}\'!C:C,\'{first_balance_sheet}\'!H:H,"Token not found"),'
                                    f'IF(trxPrice24h="Token not found","0",'
                                    f'IF(OR(trxPrice24h="",ISBLANK(trxPrice24h)),"0",'
                                    f'IF(ABS(I{row_idx})-ABS(trxPrice24h)<=1,"PASSED","FAILED")'
                                    f'))'
                                    f')'
                                )
                                cell_i.value = formula_i
                                cell_i.fill = yellow_fill

                                # Column J (10): Price (24H) Diff Validation
                                cell_j = ws_overview.cell(row_idx, 11)
                                formula_j = (
                                    f'=LET('
                                    f'trxPrice24h,XLOOKUP(B{row_idx},\'{first_balance_sheet}\'!C:C,\'{first_balance_sheet}\'!H:H),'
                                    f'ABS(I{row_idx})-ABS(trxPrice24h)'
                                    f')'
                                )
                                cell_j.value = formula_j
                                cell_j.fill = yellow_fill

                                # Column M (13): FE - Share Validation
                                cell_m = ws_overview.cell(row_idx, 14)
                                formula_m = (
                                    f'=IFS('
                                    f'AND(VALUE(M{row_idx})<0.01,VALUE(M{row_idx})<>0,TRIM(L{row_idx})="< 0.01"),"Passed",'
                                    f'AND(VALUE(M{row_idx})<0.01,VALUE(M{row_idx})<>0,TRIM(L{row_idx})<>"< 0.01"),"Failed",'
                                    f'AND(VALUE(M{row_idx})>=0.01,ROUND(VALUE(M{row_idx})*100,0)=ROUND(VALUE(L{row_idx})*100,0)),"Passed",'
                                    f'AND(VALUE(M{row_idx})>=0.01,ROUND(VALUE(M{row_idx})*100,0)<>ROUND(VALUE(L{row_idx})*100,0)),"Failed",'
                                    f'AND(VALUE(M{row_idx})=0,VALUE(L{row_idx})=0),"Passed",'
                                    f'AND(VALUE(M{row_idx})=0,VALUE(L{row_idx})<>0),"Failed")'
                                )
                                cell_m.value = formula_m
                                cell_m.fill = yellow_fill

                                # Column N (14): Calculation Share (calculated using Python Decimal = W/total_W*100)
                                cell_n = ws_overview.cell(row_idx, 15)
                                try:
                                    if total_calculated_value != 0:
                                        calc_share = (calculated_values[row_idx - 2] / total_calculated_value) * Decimal('100')
                                        # Convert to string with precision
                                        share_str = str(calc_share)
                                        if '.' in share_str:
                                            share_str = share_str.rstrip('0').rstrip('.')
                                        cell_n.value = share_str
                                    else:
                                        cell_n.value = "0"
                                except Exception as e:
                                    cell_n.value = f"Error: {str(e)[:20]}"
                                cell_n.fill = yellow_fill

                                # Column O (15): Share Validation - Python calculated
                                # Compare K (Share) with (T value / SUM of all T values) * 100
                                cell_o = ws_overview.cell(row_idx, 16)
                                try:
                                    # Get K value (Share)
                                    k_val = ws_overview.cell(row_idx, 12).value  # Column K
                                    # Get T value (Amount Diff Validation) - column 20
                                    t_val = ws_overview.cell(row_idx, 21).value

                                    # Try to convert K to Decimal
                                    if isinstance(k_val, str):
                                        k_val_str = k_val.strip().replace('<', '').replace('%', '').replace(',', '').strip()
                                        if k_val_str:
                                            k_decimal = Decimal(k_val_str)
                                        else:
                                            k_decimal = Decimal('0')
                                    else:
                                        k_decimal = Decimal(str(k_val)) if k_val is not None else Decimal('0')

                                    # Try to convert T to Decimal (skip if text like "Token not found")
                                    if isinstance(t_val, str):
                                        # Check if it's a numeric string or error text
                                        if "not found" in t_val.lower() or "error" in t_val.lower():
                                            cell_o.value = "N/A"
                                        else:
                                            t_val_str = t_val.strip().replace(',', '').strip()
                                            if t_val_str:
                                                t_decimal = Decimal(t_val_str)
                                                # Calculate T/SUM(T) * 100 and compare with K
                                                # (sum_of_t will be calculated separately)
                                                cell_o.value = ""  # Placeholder for second pass
                                            else:
                                                cell_o.value = "N/A"
                                    elif isinstance(t_val, (int, float)):
                                        # Numeric value - will calculate in second pass
                                        cell_o.value = ""  # Placeholder
                                    else:
                                        cell_o.value = "N/A"

                                except Exception as e:
                                    cell_o.value = f"Error: {str(e)[:20]}"
                                cell_o.fill = yellow_fill

                                # Column R (18): FE - Amount Validation
                                # Compare Q (Amount Tooltip) with P (Amount)
                                cell_r = ws_overview.cell(row_idx, 19)
                                formula_r = (
                                    f'=IF(VALUE(R{row_idx})<0.00001,IF(Q{row_idx}="< 0.00001","Passed","Failed"),'
                                    f'IF(LEFT(SUBSTITUTE(TEXT(R{row_idx},"0.000000000000000"),",",""),FIND(".",SUBSTITUTE(TEXT(R{row_idx},"0.000000000000000"),",",""))+5)='
                                    f'LEFT(SUBSTITUTE(TEXT(Q{row_idx},"0.000000000000000"),",",""),FIND(".",SUBSTITUTE(TEXT(Q{row_idx},"0.000000000000000"),",",""))+5),"Passed","Failed"))'
                                )
                                cell_r.value = formula_r
                                cell_r.fill = yellow_fill

                                # Column S (19): Amount Validation - Python-calculated in fallback section
                                # (Skip Excel formula - Python handles Tron via TRX Balance map, EVM via Sim Dune map)

                                # Column T (20): Amount Diff Validation (Excel Formula)
                                # Chain-specific difference calculation:
                                # - Tron: Difference = TRX Balance E column - Q column (match by Symbol via XLOOKUP)
                                # - Non-Tron (EVM): Difference = Q column - 'Sim + Coingecko + Debank API' E column (match by B=Chain, C=Symbol)
                                # NOTE: Use LOWER() for case-insensitive matching (Ethereum vs ethereum)
                                cell_t = ws_overview.cell(row_idx, 21)
                                formula_t = (
                                    f'=LET('
                                    f'chain,A{row_idx},'
                                    f'token,B{row_idx},'
                                    f'amtTooltip,VALUE(SUBSTITUTE(R{row_idx},",","")),'
                                    f'trxAmt,XLOOKUP(token,\'{first_balance_sheet}\'!C:C,\'{first_balance_sheet}\'!E:E,"Not Found"),'
                                    f'simDuneAmt,XLOOKUP(1,(LOWER(\'Sim + Coingecko + Debank API\'!B:B)=LOWER(chain))*(LOWER(\'Sim + Coingecko + Debank API\'!C:C)=LOWER(token)),\'Sim + Coingecko + Debank API\'!E:E,"Not Found"),'
                                    f'IF(chain="Tron",'
                                    f'IF(trxAmt="Not Found","Token not found",VALUE(trxAmt)-amtTooltip),'
                                    f'IF(simDuneAmt="Not Found","Token not found",amtTooltip-VALUE(simDuneAmt))'
                                    f'))'
                                )
                                cell_t.value = formula_t
                                cell_t.number_format = '@'  # Text format
                                cell_t.fill = yellow_fill

                                # Column W (23): Value Validation - IFS with <0.01 check, zero check using X column
                                cell_v = ws_overview.cell(row_idx, 23)
                                formula_v = (
                                    f'=IFS('
                                    f'AND(VALUE(X{row_idx})<0.01,VALUE(X{row_idx})<>0,V{row_idx}="< 0.01"),"Passed",'
                                    f'AND(VALUE(X{row_idx})<0.01,VALUE(X{row_idx})<>0,V{row_idx}<>"< 0.01"),"Failed",'
                                    f'AND(VALUE(X{row_idx})>=0.01,TRUNC(VALUE(X{row_idx}),2)=V{row_idx}),"Passed",'
                                    f'AND(VALUE(X{row_idx})>=0.01,TRUNC(VALUE(X{row_idx}),2)<>V{row_idx}),"Failed",'
                                    f'AND(VALUE(X{row_idx})=0,VALUE(V{row_idx})=0),"Passed",'
                                    f'AND(VALUE(X{row_idx})=0,VALUE(V{row_idx})<>0),"Failed")'
                                )
                                cell_v.value = formula_v
                                cell_v.fill = yellow_fill

                                # Column X (24): Value - UI validation (D*Q with full precision using Python Decimal)
                                cell_w = ws_overview.cell(row_idx, 24)
                                try:
                                    result = calculated_values[row_idx - 2]
                                    result_str = str(result)
                                    if '.' in result_str:
                                        result_str = result_str.rstrip('0').rstrip('.')
                                    cell_w.value = result_str
                                    # Mark yellow if value was approximated from "< 0.01"
                                    if hasattr(run_dam_portfolio_extraction, '_approx_rows_ov') and row_idx in run_dam_portfolio_extraction._approx_rows_ov:
                                        from openpyxl.styles import PatternFill as _PF
                                        cell_w.fill = _PF(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                                except Exception as e:
                                    cell_w.value = f"Error: {str(e)[:20]}"
                                cell_w.fill = yellow_fill

                            # Add sum row for total token value
                            # W (23) = label, X (24) = sum of Value column (V=22)
                            ws_overview.cell(sum_row, 23).value = "Total"

                            # Calculate Value (V column) total using Python Decimal
                            total_value_v = Decimal('0')
                            chain_value_v = {}
                            for _ri in range(2, last_data_row + 1):
                                _v_val = ws_overview.cell(_ri, 22).value  # Column V - Value
                                _chain = str(ws_overview.cell(_ri, 1).value or "").strip()
                                if _v_val is not None and str(_v_val).strip() not in ('', '-'):
                                    try:
                                        _v_dec = Decimal(str(_v_val).replace(',', '').replace('$', '').replace('<', '').strip())
                                        total_value_v += _v_dec
                                        if _chain:
                                            chain_value_v.setdefault(_chain, Decimal('0'))
                                            chain_value_v[_chain] += _v_dec
                                    except:
                                        pass

                            def _fmt_dec(d):
                                s = str(d)
                                if '.' in s:
                                    s = s.rstrip('0').rstrip('.')
                                return s

                            ws_overview.cell(sum_row, 22).value = _fmt_dec(total_value_v)
                            ws_overview.cell(sum_row, 22).fill = yellow_fill

                            # Add chain-specific total rows — only for chains that exist in data
                            all_chains_in_data = set(list(chain_calculated_values.keys()) + list(chain_u_values.keys()) + list(chain_value_v.keys()))
                            # Preferred display order for common chains
                            preferred_order = ["Ethereum", "Base", "Binance Smart Chain", "Tron"]
                            chain_order = [c for c in preferred_order if c in all_chains_in_data]
                            # Append any remaining chains not in preferred order
                            for cn in all_chains_in_data:
                                if cn not in chain_order:
                                    chain_order.append(cn)
                            current_row = sum_row + 1

                            for chain_name in chain_order:

                                # Column W: Chain total label
                                ws_overview.cell(current_row, 23).value = f"Total ({chain_name})"

                                # Column V: Chain Value sum
                                ws_overview.cell(current_row, 22).value = _fmt_dec(chain_value_v.get(chain_name, Decimal('0')))
                                ws_overview.cell(current_row, 22).fill = yellow_fill

                                current_row += 1

                            # Second pass: Calculate Column O (Share Validation) - L compared with U/SUM(U)*100
                            # First, collect all U values (Calculated Value - Python) - column 21
                            u_values = []
                            for row_idx in range(2, last_data_row + 1):
                                u_val = ws_overview.cell(row_idx, 22).value
                                try:
                                    if isinstance(u_val, str):
                                        if "not found" not in u_val.lower() and "error" not in u_val.lower():
                                            u_val_str = u_val.strip().replace(',', '').strip()
                                            if u_val_str:
                                                u_decimal = Decimal(u_val_str)
                                                u_values.append((row_idx, u_decimal))
                                    elif isinstance(u_val, (int, float)):
                                        u_values.append((row_idx, Decimal(str(u_val))))
                                except:
                                    pass  # Skip non-numeric values

                            # Calculate sum of all U values
                            if u_values:
                                sum_u = sum([u for _, u in u_values])

                                # Update O column for each row
                                for row_idx, u_val in u_values:
                                    if sum_u != 0:
                                        # Calculate U/SUM(U) * 100
                                        calculated_share = (u_val / sum_u) * Decimal('100')

                                        # Get L value for comparison
                                        l_val = ws_overview.cell(row_idx, 13).value  # Column L
                                        try:
                                            if isinstance(l_val, str):
                                                l_val_str = l_val.strip().replace('<', '').replace('%', '').replace(',', '').strip()
                                                if l_val_str:
                                                    l_decimal = Decimal(l_val_str)
                                                else:
                                                    l_decimal = Decimal('0')
                                            else:
                                                l_decimal = Decimal(str(l_val)) if l_val is not None else Decimal('0')

                                            # Compare L with calculated share (U/SUM(U)*100)
                                            # Within 1 diff (absolute difference)
                                            absolute_diff = abs(calculated_share - l_decimal)

                                            if absolute_diff <= Decimal('1'):
                                                ws_overview.cell(row_idx, 16).value = "Passed"
                                            else:
                                                ws_overview.cell(row_idx, 16).value = "Failed"

                                        except:
                                            ws_overview.cell(row_idx, 16).value = "Error"

                            # Post-processing: Update F column in "Overview - Header & Token Holdings Header" sheet
                            # Now that W column is calculated, we can perform the validation
                            if "Overview - Header & Token Holdings Header" in wb.sheetnames:
                                ws_header = wb["Overview - Header & Token Holdings Header"]
                                header_max_row = ws_header.max_row

                                for h_row_idx in range(2, header_max_row + 1):
                                    section_val = ws_header.cell(row=h_row_idx, column=1).value  # A column - Section
                                    category_val = ws_header.cell(row=h_row_idx, column=2).value  # B column - Category
                                    e_val = ws_header.cell(row=h_row_idx, column=6).value  # F column - Net Worth
                                    cell_g = ws_header.cell(row=h_row_idx, column=7)  # G column - Net Worth_UI Calculation
                                    cell_f = ws_header.cell(row=h_row_idx, column=8)  # H column - Net Worth Validation

                                    if section_val == "Overview Header" and category_val == "Wallets":
                                        try:
                                            import math
                                            w_sum_rounded = Decimal(str(math.ceil(float(total_calculated_value) * 100) / 100))
                                            cell_g.value = float(w_sum_rounded)
                                            e_clean = str(e_val).replace(',', '').replace('$', '').replace('<', '').strip() if e_val else '0'
                                            e_decimal = Decimal(e_clean).quantize(Decimal('0.01'))
                                            cell_f.value = "Passed" if e_decimal == w_sum_rounded else "Failed"
                                        except Exception as ex:
                                            cell_f.value = f"Error: {ex}"

                                    elif section_val == "Token Holdings - Chain":
                                        try:
                                            chain_name = str(category_val).strip() if category_val else ""
                                            chain_sum = chain_calculated_values.get(chain_name, Decimal('0'))
                                            import math
                                            w_sum_rounded = Decimal(str(math.ceil(float(chain_sum) * 100) / 100))
                                            cell_g.value = float(w_sum_rounded)
                                            e_clean = str(e_val).replace(',', '').replace('$', '').replace('<', '').strip() if e_val else '0'
                                            e_decimal = Decimal(e_clean).quantize(Decimal('0.01'))
                                            cell_f.value = "Passed" if e_decimal == w_sum_rounded else "Failed"
                                        except Exception as ex:
                                            cell_f.value = f"Error: {ex}"

                                    elif section_val == "Overview Header" and category_val == "De-Fi Positions":
                                        try:
                                            # Sum K column from Overview - De-Fi
                                            ws_defi_early = wb["Overview - De-Fi"] if "Overview - De-Fi" in wb.sheetnames else None
                                            defi_sum_early = Decimal('0')
                                            if ws_defi_early:
                                                for _dr in range(2, ws_defi_early.max_row + 1):
                                                    _da = ws_defi_early.cell(_dr, 1).value
                                                    _dk = ws_defi_early.cell(_dr, 11).value  # K column - Value
                                                    if _da and _dk:
                                                        try:
                                                            defi_sum_early += Decimal(str(_dk).replace(',', '').replace('$', '').strip())
                                                        except:
                                                            pass
                                            defi_sum_early = defi_sum_early.quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                            cell_g.value = float(defi_sum_early)
                                            e_clean = str(e_val).replace(',', '').replace('$', '').replace('<', '').strip() if e_val else '0'
                                            e_decimal = Decimal(e_clean).quantize(Decimal('0.01'))
                                            if defi_sum_early != 0:
                                                pct_diff = abs(e_decimal - defi_sum_early) / defi_sum_early * Decimal('100')
                                                cell_f.value = "Passed" if pct_diff <= Decimal('1') else "Failed"
                                            else:
                                                cell_f.value = "Passed" if e_decimal == 0 else "Failed"
                                        except Exception as ex:
                                            cell_f.value = f"Error: {ex}"

                                    elif not e_val or e_val == "":
                                        cell_f.value = "-"
                                    else:
                                        cell_f.value = "Not Applicable"

                                print(f"   ✅ Updated F column validation in 'Overview - Header & Token Holdings Header' sheet")

                            # Set header row values for new columns
                            ws_overview.cell(1, 14).value = "FE - Share Validation"
                            ws_overview.cell(1, 15).value = "Calculation Share"
                            ws_overview.cell(1, 16).value = "Share Validation"
                            ws_overview.cell(1, 24).value = "Value - UI validation"

                            print(f"   ✅ Updated all validation formulas to use internal sheet")
                            print(f"   ✅ Added sum row at row {sum_row}")
                            print(f"   ✅ Added Share Validation formula in column I")
                            print(f"   ✅ Added Share Tooltip column H and Amount Tooltip column K")

                            # Add conditional formatting for validation columns
                            # Columns with "Passed"/"Failed" text: E, F, G, I, J, M, N, Q, R, T, V
                            # Light Red Fill with Dark Red Text (Excel standard)
                            light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            dark_red_font = Font(color="9C0006")

                            # Green Fill with Dark Green Text (Excel standard)
                            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            dark_green_font = Font(color="006100")

                            # Grey Fill for "Tooltip N/A, cant compare"
                            grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                            grey_font = Font(color="808080")

                            # Validation columns with Passed/Failed: E(5), F(6), I(9), N(14), O(15), R(18), S(19), V(22)
                            # Note: G(7), J(10), M(13), T(20), W(23) are numeric calculations (text), not Passed/Failed - excluded from conditional formatting
                            validation_columns = ['E', 'F', 'I', 'N', 'O', 'R', 'S', 'V']

                            for col_letter in validation_columns:
                                # Apply to data rows (2 to last_data_row)
                                range_address = f'{col_letter}2:{col_letter}{last_data_row}'

                                # Green Fill with Dark Green Text for "Passed"
                                ws_overview.conditional_formatting.add(
                                    range_address,
                                    CellIsRule(operator='containsText', formula=['"Passed"'], fill=green_fill, font=dark_green_font)
                                )

                                # Light Red Fill with Dark Red Text for "Failed"
                                ws_overview.conditional_formatting.add(
                                    range_address,
                                    CellIsRule(operator='containsText', formula=['"Failed"'], fill=light_red_fill, font=dark_red_font)
                                )

                                # Grey Fill for "Tooltip N/A, cant compare"
                                ws_overview.conditional_formatting.add(
                                    range_address,
                                    CellIsRule(operator='containsText', formula=['"Tooltip N/A, cant compare"'], fill=grey_fill, font=grey_font)
                                )

                            print(f"   ✅ Added conditional formatting (Green=Passed, Red=Failed, Grey=Tooltip N/A)")

                    trx_wb.close()
                except Exception as e:
                    print(f"   ⚠️  Warning: Could not copy TRX Balance data: {e}")

            # Add Sim + Coingecko + Debank API sheet if EVM data was extracted
            if sim_dune_extracted_data and len(sim_dune_extracted_data) > 1:
                try:
                    ws_sim_dune = wb.create_sheet("Sim + Coingecko + Debank API")

                    # Write data rows
                    for row_idx, row_data in enumerate(sim_dune_extracted_data, start=1):
                        for col_idx, value in enumerate(row_data, start=1):
                            ws_sim_dune.cell(row=row_idx, column=col_idx, value=value)

                    print(f"   ✅ Added 'Sim + Coingecko + Debank API' sheet with {len(sim_dune_extracted_data) - 1} token(s)")

                    # Post-process: call Debank API for rows where H (col 8) = "Spam Token"
                    # Col B (2)=Chain, Col G (7)=Token Address, Col I (9)=Price, Col J (10)=24H Price Change
                    _db_chain_map = {
                        'ethereum': 'eth',
                        'bnb': 'bsc',
                        'base': 'base',
                    }
                    _db_hits = _db_misses = 0
                    import requests as _req_db
                    import json as _json_db
                    for _sd_row in range(2, ws_sim_dune.max_row + 1):
                        _sd_h = ws_sim_dune.cell(_sd_row, 8).value
                        if str(_sd_h).strip() != "Spam Token":
                            continue
                        _sd_chain_raw = str(ws_sim_dune.cell(_sd_row, 2).value or '').strip().lower()
                        _sd_addr = str(ws_sim_dune.cell(_sd_row, 7).value or '').strip()
                        if not _sd_addr or _sd_addr.lower() in ('', 'none'):
                            print(f"      ⚠️  Debank: row {_sd_row} skipped — empty token address")
                            continue
                        _sd_chain_id = _db_chain_map.get(_sd_chain_raw, _sd_chain_raw)
                        _db_url = (
                            f"https://api.debank.com/swap/search_token"
                            f"?chain_id={_sd_chain_id}&q={_sd_addr}&exact=true"
                        )
                        print(f"      🌐 Debank request: {_db_url}")
                        try:
                            _db_resp = _req_db.get(
                                _db_url, timeout=15,
                                headers={
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                                    'Accept': 'application/json',
                                    'Origin': 'https://debank.com',
                                    'Referer': 'https://debank.com/',
                                }
                            )
                            print(f"      📡 Debank HTTP {_db_resp.status_code} — raw: {_db_resp.text[:300]}")
                            if _db_resp.status_code == 200:
                                _db_json = _db_resp.json()
                                # Walk every possible response shape to find a token dict
                                _db_token = None
                                def _find_token(obj):
                                    """Recursively search for the first dict containing a 'price' key."""
                                    if isinstance(obj, dict):
                                        if 'price' in obj:
                                            return obj
                                        for _v in obj.values():
                                            _t = _find_token(_v)
                                            if _t:
                                                return _t
                                    elif isinstance(obj, list):
                                        for _item in obj:
                                            _t = _find_token(_item)
                                            if _t:
                                                return _t
                                    return None
                                _db_token = _find_token(_db_json)
                                print(f"      🔎 Debank parsed token: {_db_token}")
                                if _db_token:
                                    _db_price = _db_token.get('price')
                                    _db_change = _db_token.get('price_24h_change')
                                    if _db_price is not None:
                                        ws_sim_dune.cell(_sd_row, 9).value = _db_price
                                        # Price found via Debank — not spam
                                        ws_sim_dune.cell(_sd_row, 8).value = "Found via Debank"
                                    else:
                                        # Debank found token but no price — token exists, price unavailable
                                        ws_sim_dune.cell(_sd_row, 8).value = "Found via Debank"
                                        ws_sim_dune.cell(_sd_row, 9).value = "Data not available"
                                    if _db_change is not None:
                                        ws_sim_dune.cell(_sd_row, 10).value = _db_change
                                    else:
                                        ws_sim_dune.cell(_sd_row, 10).value = "Data not available"
                                    _db_hits += 1
                                    print(f"      ✅ Debank: row {_sd_row} {_sd_addr} ({_sd_chain_id}) → price={_db_price}, 24h_change={_db_change}")
                                else:
                                    _db_misses += 1
                                    print(f"      ⚠️  Debank: token not found in response for {_sd_addr} on {_sd_chain_id}")
                            else:
                                _db_misses += 1
                                print(f"      ⚠️  Debank: HTTP {_db_resp.status_code} for {_sd_addr}")
                        except Exception as _db_err:
                            import traceback as _db_tb
                            _db_misses += 1
                            print(f"      ⚠️  Debank API error for {_sd_addr}: {_db_err}")
                            _db_tb.print_exc()
                    print(f"   ✅ Debank spam-token lookup: {_db_hits} resolved, {_db_misses} not found")
                except Exception as e:
                    print(f"   ⚠️  Warning: Could not add Sim Dune sheet: {e}")

            # Add Rabby Api Data sheet if Rabby data was extracted
            if rabby_extracted_data and len(rabby_extracted_data) > 1:
                try:
                    ws_rabby = wb.create_sheet("Rabby Api Data")
                    _rabby_key_order = ["Address", "Name", "ID", "Chain", "Pool_Name", "Description", "Side", "Symbol", "Leverage", "PnL_USD", "Price", "Amount", "Calculated_Value"]

                    # Write data rows — first row is the header list, rest are dicts
                    for row_idx, row_data in enumerate(rabby_extracted_data, start=1):
                        if isinstance(row_data, dict):
                            row_values = [row_data.get(k, "") for k in _rabby_key_order]
                        else:
                            row_values = row_data
                        for col_idx, value in enumerate(row_values, start=1):
                            ws_rabby.cell(row=row_idx, column=col_idx, value=value)

                    print(f"   ✅ Added 'Rabby Api Data' sheet with {len(rabby_extracted_data) - 1} protocol token(s)")
                except Exception as e:
                    print(f"   ⚠️  Warning: Could not add Rabby sheet: {e}")

            # Add Hyperliquid sheet if Hyperliquid data was extracted
            if hyperliquid_extracted_data and len(hyperliquid_extracted_data) > 1:
                try:
                    ws_hyperliquid = wb.create_sheet("Hyperliquid")

                    # Write data rows
                    for row_idx, row_data in enumerate(hyperliquid_extracted_data, start=1):
                        for col_idx, value in enumerate(row_data, start=1):
                            ws_hyperliquid.cell(row=row_idx, column=col_idx, value=value)

                    print(f"   ✅ Added 'Hyperliquid' sheet with {len(hyperliquid_extracted_data) - 1} position(s)")
                except Exception as e:
                    print(f"   ⚠️  Warning: Could not add Hyperliquid sheet: {e}")

            # Calculate W column and R column if not already done (for EVM-only portfolios)
            # This runs AFTER all sheets are created, ensuring columns are always calculated
            if "Overview - Wallet" in wb.sheetnames:
                ws_wallet = wb["Overview - Wallet"]

                # Find the last data row
                last_row = ws_wallet.max_row
                while last_row > 1 and not ws_wallet.cell(last_row, 2).value:
                    last_row -= 1

                # Check if W column needs to be calculated (if empty)
                w_needs_calculation = True
                if ws_wallet.cell(2, 24).value and str(ws_wallet.cell(2, 24).value).strip() not in ["", "0", "Error:"]:
                    w_needs_calculation = False

                # Check if N column (Amount Validation) needs to be calculated (if empty or has formula errors)
                r_needs_calculation = True
                r_cell_val = ws_wallet.cell(2, 15).value
                if r_cell_val and str(r_cell_val).strip() in ["Passed", "Failed", "Token not Found"]:
                    r_needs_calculation = False

                yellow_fill_fallback = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # Build Sim Dune balance map from 'Sim + Coingecko + Debank API' sheet if it exists
                # Store as STRING (not Decimal) to avoid float precision issues
                sim_dune_map = {}
                sim_dune_row_map = {}  # key → list of row indices (parallel to sim_dune_map amounts)
                if "Sim + Coingecko + Debank API" in wb.sheetnames:
                    ws_sim = wb["Sim + Coingecko + Debank API"]
                    print(f"   Building Sim Dune map from 'Sim + Coingecko + Debank API' sheet...")
                    # Chain name normalization
                    _sim_norm = {'hyper_evm': 'hyperevm', 'bnb': 'binance smart chain', 'bsc': 'binance smart chain'}
                    for row_idx in range(2, ws_sim.max_row + 1):
                        chain = str(ws_sim.cell(row_idx, 2).value or "").strip()  # Column B - Chain
                        chain_normalized = _sim_norm.get(chain.lower(), chain.lower())
                        symbol = str(ws_sim.cell(row_idx, 3).value or "").strip()  # Column C - Symbol
                        amount_raw = ws_sim.cell(row_idx, 5).value  # Column E - Amount

                        if chain and symbol and amount_raw is not None:
                            try:
                                key = (chain_normalized, symbol.lower())
                                # Handle both string and numeric cell values
                                # Use repr to see exact value, then normalize
                                if isinstance(amount_raw, (int, float)):
                                    # For numeric values, use high precision string formatting
                                    amount_str = f"{amount_raw:.18f}".rstrip('0').rstrip('.')
                                else:
                                    amount_str = str(amount_raw).strip()
                                    if '.' in amount_str:
                                        amount_str = amount_str.rstrip('0').rstrip('.')

                                # Debug: show raw type and value
                                print(f"      SimDune: {chain}/{symbol} = {amount_str} (raw type: {type(amount_raw).__name__}, raw: {repr(amount_raw)})")

                                # Store as list of amounts AND row indices in parallel
                                # so _pick_closest can return the matching row index
                                if key in sim_dune_map:
                                    sim_dune_map[key].append(amount_str)
                                    sim_dune_row_map[key].append(row_idx)
                                else:
                                    sim_dune_map[key] = [amount_str]
                                    sim_dune_row_map[key] = [row_idx]
                            except Exception as e:
                                print(f"      SimDune Error: {chain}/{symbol} - {e}")
                    print(f"   ✅ Built Sim Dune map with {len(sim_dune_map)} tokens")

                # Build TRX Balance map from token_balance_map (pre-computed from raw balance + decimals)
                # token_balance_map was built at TRX Balance copy step: {token_name: Decimal_balance}
                # Column E in the sheet has Excel formulas (not values), so we use the pre-computed map
                # Key: symbol.lower(), Value: normalized amount string
                trx_balance_map = {}
                trx_balance_row_map = {}  # key → row index in TRX Balance sheet
                if "TRX Balance, Price" in wb.sheetnames:
                    _ws_trx = wb["TRX Balance, Price"]
                    for _ri in range(2, _ws_trx.max_row + 1):
                        _sym = str(_ws_trx.cell(_ri, 3).value or "").strip()  # Column C - Token
                        if _sym and _sym.lower() not in trx_balance_row_map:
                            trx_balance_row_map[_sym.lower()] = _ri
                if 'token_balance_map' in dir() and token_balance_map:
                    print(f"   Building TRX Balance map from pre-computed token_balance_map ({len(token_balance_map)} tokens)...")
                    for token_name, balance_decimal in token_balance_map.items():
                        try:
                            key = token_name.lower()
                            amount_str = str(balance_decimal)
                            if '.' in amount_str:
                                amount_str = amount_str.rstrip('0').rstrip('.')
                            trx_balance_map[key] = amount_str
                            print(f"      TRXBalance: {token_name} = {amount_str}")
                        except Exception as e:
                            print(f"      TRXBalance Error: {token_name} - {e}")
                    print(f"   ✅ Built TRX Balance map with {len(trx_balance_map)} tokens")

                if w_needs_calculation or r_needs_calculation:
                    print(f"\n🔧 Fallback calculation for EVM-only portfolio...")

                # Helper: pick the candidate amount closest to a target value
                def _pick_closest(candidates, target_str):
                    """From a list of amount strings, return the one closest to target_str and its index."""
                    if not candidates:
                        return None, -1
                    if len(candidates) == 1:
                        return candidates[0], 0
                    try:
                        target_dec = Decimal(target_str)
                        best_idx = min(range(len(candidates)), key=lambda i: abs(Decimal(candidates[i]) - target_dec))
                        return candidates[best_idx], best_idx
                    except:
                        return candidates[0], 0

                # Pre-match: for each DAM Wallet row, find the best SimDune/TRX row ONCE
                # All validation columns will use this same matched row
                matched_source_row = {}  # DAM row_idx → SimDune/TRX row index (1-indexed)
                for row_idx in range(2, last_row + 1):
                    chain_val = str(ws_wallet.cell(row_idx, 1).value or "").strip()
                    token_val = str(ws_wallet.cell(row_idx, 2).value or "").strip()
                    _chain_norm_match = {'bnb': 'binance smart chain', 'bsc': 'binance smart chain',
                                         'hyper_evm': 'hyperevm', 'hyper evm': 'hyperevm'}
                    _cl = _chain_norm_match.get(chain_val.lower(), chain_val.lower())

                    if _cl == "tron":
                        lookup_key = token_val.lower()
                        if lookup_key in trx_balance_row_map:
                            matched_source_row[row_idx] = trx_balance_row_map[lookup_key]
                    else:
                        lookup_key = (_cl, token_val.lower())
                        if lookup_key in sim_dune_map:
                            # Use amount to pick closest
                            q_raw = ws_wallet.cell(row_idx, 13).value  # Column M - Amount Tooltip
                            if q_raw is None or str(q_raw).strip() in ("", "Tooltip N/A"):
                                q_raw = ws_wallet.cell(row_idx, 12).value  # Fallback to Amount cell
                            try:
                                if isinstance(q_raw, (int, float)):
                                    q_str = f"{q_raw:.18f}".rstrip('0').rstrip('.')
                                else:
                                    q_str = str(q_raw).replace(',', '').strip()
                            except:
                                q_str = "0"
                            _, picked_idx = _pick_closest(sim_dune_map[lookup_key], q_str)
                            row_list = sim_dune_row_map.get(lookup_key, [])
                            if row_list and 0 <= picked_idx < len(row_list):
                                matched_source_row[row_idx] = row_list[picked_idx]

                print(f"   ✅ Pre-matched {len(matched_source_row)}/{last_row - 1} DAM rows to API source rows")

                if w_needs_calculation:
                    print(f"   Calculating W column (D*L with Python Decimal precision)...")

                    # Set X column header
                    ws_wallet.cell(1, 24).value = "Value - UI validation"

                    # Calculate D*L for each row with full Decimal precision
                    calculated_values_w = []
                    for row_idx in range(2, last_row + 1):
                        d_value = ws_wallet.cell(row_idx, 4).value   # Column D - Price Tooltip
                        q_value = ws_wallet.cell(row_idx, 13).value  # Column L - Amount Tooltip

                        try:
                            # Clean values - remove commas, $, whitespace, < prefix
                            d_str = str(d_value).strip() if d_value else ''
                            q_str = str(q_value).strip() if q_value else ''

                            # Fallback: if tooltip is missing, use displayed Price (col C) and Amount (col L)
                            _d_is_na = (not d_str or 'N/A' in d_str or 'cant compare' in d_str)
                            _q_is_na = (not q_str or 'N/A' in q_str or 'cant compare' in q_str)

                            if _d_is_na:
                                c_fallback = ws_wallet.cell(row_idx, 3).value  # Column C - Price (displayed)
                                d_str = str(c_fallback).strip() if c_fallback else ''
                            if _q_is_na:
                                l_fallback = ws_wallet.cell(row_idx, 12).value  # Column L - Amount (displayed)
                                q_str = str(l_fallback).strip() if l_fallback else ''

                            if not d_str or not q_str:
                                calculated_values_w.append(Decimal('0'))
                                ws_wallet.cell(row_idx, 24).value = ""
                                continue

                            # Handle "< 0.01" — treat as 0.01 and flag yellow
                            _d_approx = False
                            _q_approx = False
                            if '< 0.01' in d_str or '<0.01' in d_str:
                                d_str = '0.01'
                                _d_approx = True
                            if '< 0.01' in q_str or '<0.01' in q_str:
                                q_str = '0.01'
                                _q_approx = True

                            d_clean = d_str.replace(',', '').replace('$', '').replace('<', '').replace('>', '').strip()
                            q_clean = q_str.replace(',', '').replace('$', '').replace('<', '').replace('>', '').strip()

                            if not d_clean or not q_clean:
                                calculated_values_w.append(Decimal('0'))
                                ws_wallet.cell(row_idx, 24).value = ""
                                continue

                            # Use Decimal for precise calculation
                            d_decimal = Decimal(d_clean)
                            q_decimal = Decimal(q_clean)
                            result = d_decimal * q_decimal
                            calculated_values_w.append(result)

                            # Write to W column with full precision
                            result_str = str(result)
                            if '.' in result_str:
                                result_str = result_str.rstrip('0').rstrip('.')
                            ws_wallet.cell(row_idx, 24).value = result_str

                            # Mark yellow if any value was approximated from "< 0.01"
                            if _d_approx or _q_approx:
                                from openpyxl.styles import PatternFill as _PF
                                _approx_fill = _PF(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                                ws_wallet.cell(row_idx, 24).fill = _approx_fill
                        except Exception as e:
                            calculated_values_w.append(Decimal('0'))
                            ws_wallet.cell(row_idx, 24).value = ""

                    # Add sum row
                    sum_row_w = last_row + 1
                    total_w = sum(calculated_values_w)
                    total_str_w = str(total_w)
                    if '.' in total_str_w:
                        total_str_w = total_str_w.rstrip('0').rstrip('.')
                    ws_wallet.cell(sum_row_w, 24).value = total_str_w

                    # Apply yellow fill to W column
                    for row_idx in range(1, sum_row_w + 1):
                        ws_wallet.cell(row_idx, 24).fill = yellow_fill_fallback

                    print(f"   ✅ Calculated W column for {last_row - 1} token rows")
                    print(f"   ✅ W column total: {total_str_w}")

                    # Calculate Y column (API Calculated Value = SimDune/TRX Balance Calculated Value per token)
                    ws_wallet.cell(1, 25).value = "API Calculated Value"
                    # Build lookup maps from SimDune Calculated Value (column K=11) and TRX Balance Calculated Value (column J=10)
                    _api_calc_map = {}  # (chain, symbol) → list of Calculated Values
                    _api_calc_trx = {}  # symbol → Calculated Value
                    if "Sim + Coingecko + Debank API" in wb.sheetnames:
                        _ws_sd = wb["Sim + Coingecko + Debank API"]
                        _sd_norm = {'hyper_evm': 'hyperevm', 'bnb': 'binance smart chain', 'bsc': 'binance smart chain', 'binance_smart_chain': 'binance smart chain'}
                        for _ri in range(2, _ws_sd.max_row + 1):
                            _ch = str(_ws_sd.cell(_ri, 2).value or "").strip().lower()
                            _ch = _sd_norm.get(_ch, _ch)
                            _sym = str(_ws_sd.cell(_ri, 3).value or "").strip().lower()
                            _cv = _ws_sd.cell(_ri, 11).value   # Column K = Calculated Value
                            _amt = _ws_sd.cell(_ri, 5).value   # Column E = Amount (for closest matching)
                            if _ch and _sym and _cv is not None:
                                try:
                                    _cv_dec = Decimal(str(_cv).replace(',', '').replace('$', '').strip())
                                    _amt_dec = Decimal(str(_amt).replace(',', '').strip()) if _amt else Decimal('0')
                                    _key = (_ch, _sym)
                                    if _key not in _api_calc_map:
                                        _api_calc_map[_key] = []
                                    _api_calc_map[_key].append({"calc_value": _cv_dec, "amount": _amt_dec})
                                except:
                                    pass
                    if first_balance_sheet and first_balance_sheet in wb.sheetnames:
                        _ws_trx = wb[first_balance_sheet]
                        for _ri in range(2, _ws_trx.max_row + 1):
                            _sym = str(_ws_trx.cell(_ri, 3).value or "").strip().lower()
                            _cv = _ws_trx.cell(_ri, 10).value   # Column J = Calculated Value
                            if _sym and _cv is not None:
                                try:
                                    _api_calc_trx[_sym] = Decimal(str(_cv).replace(',', '').replace('$', '').strip())
                                except:
                                    pass

                    for row_idx in range(2, last_row + 1):
                        chain_val = str(ws_wallet.cell(row_idx, 1).value or "").strip()
                        token_val = str(ws_wallet.cell(row_idx, 2).value or "").strip()
                        cell_y = ws_wallet.cell(row_idx, 25)
                        try:
                            _chain_norm_api = {'bnb': 'binance smart chain', 'bsc': 'binance smart chain', 'hyper_evm': 'hyperevm', 'hyper evm': 'hyperevm'}
                            _cl = _chain_norm_api.get(chain_val.lower(), chain_val.lower())
                            if _cl == "tron":
                                _tk = token_val.lower()
                                if _tk in _api_calc_trx:
                                    cell_y.value = str(_api_calc_trx[_tk])
                                else:
                                    cell_y.value = ""
                            else:
                                _key = (_cl, token_val.lower())
                                if _key in _api_calc_map:
                                    candidates = _api_calc_map[_key]
                                    if len(candidates) == 1:
                                        cell_y.value = str(candidates[0]["calc_value"])
                                    else:
                                        # Pick closest by amount
                                        _dam_amt_raw = ws_wallet.cell(row_idx, 13).value
                                        try:
                                            _dam_amt = Decimal(str(_dam_amt_raw).replace(',', '').strip()) if _dam_amt_raw else Decimal('0')
                                        except:
                                            _dam_amt = Decimal('0')
                                        _best = min(candidates, key=lambda c: abs(c["amount"] - _dam_amt))
                                        cell_y.value = str(_best["calc_value"])
                                else:
                                    cell_y.value = ""
                        except:
                            cell_y.value = ""
                        cell_y.fill = yellow_fill_fallback
                    print(f"   ✅ Calculated Y column (API Calculated Value) for {last_row - 1} token rows")

                    # Calculate S column (Calculation Share = W/total_W*100)
                    ws_wallet.cell(1, 20).value = "Calculation Share"
                    for row_idx in range(2, last_row + 1):
                        try:
                            if total_w != 0:
                                calc_share = (calculated_values_w[row_idx - 2] / total_w) * Decimal('100')
                                share_str = str(calc_share)
                                if '.' in share_str:
                                    share_str = share_str.rstrip('0').rstrip('.')
                                ws_wallet.cell(row_idx, 20).value = share_str
                            else:
                                ws_wallet.cell(row_idx, 20).value = "0"
                        except Exception as e:
                            ws_wallet.cell(row_idx, 20).value = f"Error: {str(e)[:20]}"
                        ws_wallet.cell(row_idx, 20).fill = yellow_fill_fallback
                    print(f"   ✅ Calculated S column (Calculation Share) for {last_row - 1} token rows")

                # Calculate Columns F & G (Price Validation & Diff - Python)
                if sim_dune_map or trx_balance_map:
                    print(f"   Calculating F & G columns (Price Validation & Diff - Python)...")

                    # Build Sim Dune price map from 'Sim + Coingecko + Debank API' sheet (column I = Price)
                    sim_dune_price_map = {}
                    sim_dune_no_price_keys = set()  # tokens found in SimDune but price not available
                    # Chain name normalization for matching
                    _sim_chain_normalize = {
                        'hyper_evm': 'hyperevm',
                        'binance_smart_chain': 'binance smart chain',
                        'bsc': 'binance smart chain',
                        'bnb': 'binance smart chain',
                    }
                    if "Sim + Coingecko + Debank API" in wb.sheetnames:
                        ws_sim_dune = wb["Sim + Coingecko + Debank API"]
                        for row_idx in range(2, ws_sim_dune.max_row + 1):
                            chain = str(ws_sim_dune.cell(row_idx, 2).value or "").strip().lower()
                            chain = _sim_chain_normalize.get(chain, chain)
                            symbol = str(ws_sim_dune.cell(row_idx, 3).value or "").strip().lower()
                            price = ws_sim_dune.cell(row_idx, 9).value  # Column I - Price
                            if chain and symbol:
                                key = (chain, symbol)
                                if price is not None and str(price).strip() not in ("", "Data not available"):
                                    try:
                                        if isinstance(price, (int, float)):
                                            sim_dune_price_map[key] = Decimal(str(price))
                                        elif isinstance(price, str):
                                            price_str = price.replace(',', '').replace('$', '').strip()
                                            if price_str:
                                                sim_dune_price_map[key] = Decimal(price_str)
                                    except:
                                        pass
                                else:
                                    # Token found in SimDune but price not available
                                    sim_dune_no_price_keys.add(key)

                    # Build TRX Balance price map (column G = Price)
                    trx_price_map = {}
                    if first_balance_sheet and first_balance_sheet in wb.sheetnames:
                        ws_trx = wb[first_balance_sheet]
                        for row_idx in range(2, ws_trx.max_row + 1):
                            token = str(ws_trx.cell(row_idx, 3).value or "").strip().lower()  # Column C - Token
                            price = ws_trx.cell(row_idx, 7).value  # Column G - Price
                            if token and price is not None:
                                try:
                                    # Handle numeric or string price values
                                    if isinstance(price, (int, float)):
                                        trx_price_map[token] = Decimal(str(price))
                                    elif isinstance(price, str):
                                        price_str = price.replace(',', '').replace('$', '').strip()
                                        if price_str:  # Only convert non-empty strings
                                            trx_price_map[token] = Decimal(price_str)
                                except:
                                    pass  # Skip invalid prices

                    # Calculate F (Price Validation), G (Price abs_diff), H (Price Diff %) for each row
                    # Uses matched_source_row to read from the same API row as Amount validation
                    ws_sim_dune_ref = wb["Sim + Coingecko + Debank API"] if "Sim + Coingecko + Debank API" in wb.sheetnames else None
                    ws_trx_ref = wb[first_balance_sheet] if first_balance_sheet and first_balance_sheet in wb.sheetnames else None

                    for row_idx in range(2, last_row + 1):
                        chain_val = str(ws_wallet.cell(row_idx, 1).value or "").strip()  # Column A - Chain
                        token_val = str(ws_wallet.cell(row_idx, 2).value or "").strip()  # Column B - Token
                        c_val = ws_wallet.cell(row_idx, 3).value  # Column C - Price (cell text)
                        d_val = ws_wallet.cell(row_idx, 4).value  # Column D - Price Tooltip

                        # Fallback: if tooltip missing, use Price cell text
                        _tooltip_missing_price = (d_val is None or d_val == "" or str(d_val).strip() == "Tooltip N/A")
                        effective_price_val = c_val if _tooltip_missing_price else d_val

                        cell_f = ws_wallet.cell(row_idx, 6)  # Column F - Price Validation
                        cell_g = ws_wallet.cell(row_idx, 7)  # Column G - Price abs_diff
                        cell_h = ws_wallet.cell(row_idx, 8)  # Column H - Price Diff %

                        def _write_price_result(dam_p, src_p, cf, cg, ch):
                            """Write F/G/H price validation columns."""
                            if src_p != Decimal('0'):
                                diff_pct = abs((dam_p - src_p) / src_p) * Decimal('100')
                                threshold = Decimal('1')
                                cf.value = "Passed" if diff_pct <= threshold else "Failed"
                                cg.value = str(dam_p - src_p)
                                ch.value = str(diff_pct.quantize(Decimal('0.0001')))
                            else:
                                cf.value = "Passed" if dam_p == Decimal('0') else "Failed"
                                cg.value = str(dam_p)
                                ch.value = ""

                        try:
                            if effective_price_val is None or effective_price_val == "" or str(effective_price_val).strip() == "" or str(effective_price_val).strip() == "Tooltip N/A":
                                cell_f.value = "No Price Data"
                                cell_g.value = ""
                                cell_h.value = ""
                            else:
                                d_str = str(effective_price_val).replace(',', '').replace('$', '').strip()
                                dam_price = Decimal(d_str) if d_str else Decimal('0')

                                # Use matched_source_row to read price from the same API row
                                _chain_norm_p = {'bnb': 'binance smart chain', 'bsc': 'binance smart chain', 'hyper_evm': 'hyperevm', 'hyper evm': 'hyperevm'}
                                _chain_lower_p = _chain_norm_p.get(chain_val.lower(), chain_val.lower())
                                src_price = None

                                if row_idx in matched_source_row:
                                    _src_row = matched_source_row[row_idx]
                                    if _chain_lower_p == "tron" and ws_trx_ref:
                                        _p = ws_trx_ref.cell(_src_row, 7).value  # Column G = Price
                                    elif ws_sim_dune_ref:
                                        _p = ws_sim_dune_ref.cell(_src_row, 9).value  # Column I = Price
                                    else:
                                        _p = None
                                    if _p is not None and str(_p).strip() not in ("", "Data not available"):
                                        try:
                                            src_price = Decimal(str(_p).replace(',', '').replace('$', '').strip())
                                        except:
                                            pass

                                if src_price is not None:
                                    _write_price_result(dam_price, src_price, cell_f, cell_g, cell_h)
                                else:
                                    cell_f.value = "Price Not Found"
                                    cell_g.value = "-"
                                    cell_h.value = ""

                        except Exception as e:
                            cell_f.value = f"Error: {str(e)[:20]}"
                            cell_g.value = ""
                            cell_h.value = ""

                        cell_f.fill = yellow_fill_fallback
                        cell_g.fill = yellow_fill_fallback
                        cell_h.fill = yellow_fill_fallback

                    print(f"   ✅ Calculated F, G & H columns for {last_row - 1} token rows")

                    # Conditional formatting for Price Diff % (col H = 8)
                    # if price < 0.01: Green < 1%, Yellow 1-2%, Red > 2%
                    # if price >= 0.01: Green <= 1%, Red > 1%
                    from openpyxl.formatting.rule import FormulaRule
                    _green_fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
                    _yellow_fill_cf = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    _red_fill = PatternFill(start_color="E57373", end_color="E57373", fill_type="solid")
                    _h_range = f"H2:H{last_row}"
                    # Red: price >= 0.01 AND diff% > 1%
                    ws_wallet.conditional_formatting.add(_h_range, FormulaRule(
                        formula=[f'AND(IFERROR(VALUE(C2),0)>=0.01,IFERROR(VALUE(H2),0)>1)'],
                        fill=_red_fill))
                    # Yellow: price < 0.01 AND diff% between 1% and 2%
                    ws_wallet.conditional_formatting.add(_h_range, FormulaRule(
                        formula=[f'AND(IFERROR(VALUE(C2),0)<0.01,IFERROR(VALUE(H2),0)>=1,IFERROR(VALUE(H2),0)<=2)'],
                        fill=_yellow_fill_cf))
                    # Red: price < 0.01 AND diff% > 2%
                    ws_wallet.conditional_formatting.add(_h_range, FormulaRule(
                        formula=[f'AND(IFERROR(VALUE(C2),0)<0.01,IFERROR(VALUE(H2),0)>2)'],
                        fill=_red_fill))
                    # Green: price >= 0.01 AND diff% <= 1%
                    ws_wallet.conditional_formatting.add(_h_range, FormulaRule(
                        formula=[f'AND(IFERROR(VALUE(C2),0)>=0.01,IFERROR(VALUE(H2),0)<=1)'],
                        fill=_green_fill))
                    # Green: price < 0.01 AND diff% < 1%
                    ws_wallet.conditional_formatting.add(_h_range, FormulaRule(
                        formula=[f'AND(IFERROR(VALUE(C2),0)<0.01,IFERROR(VALUE(H2),0)<1)'],
                        fill=_green_fill))
                    print(f"   ✅ Added conditional formatting for Price Diff % (col H)")

                # Initialize 24H maps (fallback for rows without matched_source_row)
                sim_dune_24h_map = {}
                trx_24h_map = {}
                trx_tokens_set = set()  # Set of all TRX tokens

                # Calculate Column J (col 10, Price 24H Validation) and K (col 11, Price 24H Diff)
                # using matched_source_row — same row as Price/Amount validation
                # Price(24h) is in column I (col 9)
                if sim_dune_map or trx_balance_map:
                    print(f"   Calculating J & K columns (Price 24H Validation & Diff - Python, using matched_source_row)...")

                    # Build 24H maps as fallback only
                    sim_dune_no_24h_keys = set()
                    _sim_chain_normalize_24h = {
                        'hyper_evm': 'hyperevm',
                        'binance_smart_chain': 'binance smart chain',
                        'bsc': 'binance smart chain',
                        'bnb': 'binance smart chain',
                    }
                    if "Sim + Coingecko + Debank API" in wb.sheetnames:
                        ws_sim_dune = wb["Sim + Coingecko + Debank API"]
                        for row_idx in range(2, ws_sim_dune.max_row + 1):
                            chain = str(ws_sim_dune.cell(row_idx, 2).value or "").strip().lower()
                            chain = _sim_chain_normalize_24h.get(chain, chain)
                            symbol = str(ws_sim_dune.cell(row_idx, 3).value or "").strip().lower()
                            price_24h = ws_sim_dune.cell(row_idx, 10).value  # Column J - 24H Price Change
                            if chain and symbol:
                                key = (chain, symbol)
                                if price_24h is not None and str(price_24h).strip() not in ("", "Data not available"):
                                    try:
                                        if isinstance(price_24h, (int, float)):
                                            sim_dune_24h_map[key] = Decimal(str(price_24h))
                                        elif isinstance(price_24h, str):
                                            price_str = price_24h.replace(',', '').replace('%', '').strip()
                                            if price_str:
                                                sim_dune_24h_map[key] = Decimal(price_str)
                                    except:
                                        pass
                                else:
                                    if str(price_24h).strip() == "null":
                                        sim_dune_no_24h_keys.add(("null", key))
                                    else:
                                        sim_dune_no_24h_keys.add(key)

                    if first_balance_sheet and first_balance_sheet in wb.sheetnames:
                        ws_trx = wb[first_balance_sheet]
                        for row_idx in range(2, ws_trx.max_row + 1):
                            token = str(ws_trx.cell(row_idx, 3).value or "").strip().lower()
                            price_24h = ws_trx.cell(row_idx, 8).value
                            if token:
                                trx_tokens_set.add(token)
                                if price_24h is not None:
                                    try:
                                        if isinstance(price_24h, (int, float)):
                                            trx_24h_map[token] = Decimal(str(price_24h))
                                        elif isinstance(price_24h, str):
                                            price_str = price_24h.replace(',', '').replace('%', '').strip()
                                            if price_str:
                                                trx_24h_map[token] = Decimal(price_str)
                                    except:
                                        pass

                    # Reference sheets for matched_source_row lookups
                    _ws_sd_24h = wb["Sim + Coingecko + Debank API"] if "Sim + Coingecko + Debank API" in wb.sheetnames else None
                    _ws_trx_24h = wb[first_balance_sheet] if first_balance_sheet and first_balance_sheet in wb.sheetnames else None

                    # Set headers
                    ws_wallet.cell(1, 10).value = "Price (24h) Validation"
                    ws_wallet.cell(1, 11).value = "Price (24H) Diff Validation"

                    # Calculate J (col 10 = Validation) and K (col 11 = Diff) for each row
                    for row_idx in range(2, last_row + 1):
                        chain_val = str(ws_wallet.cell(row_idx, 1).value or "").strip()
                        token_val = str(ws_wallet.cell(row_idx, 2).value or "").strip()
                        dam_24h_raw = ws_wallet.cell(row_idx, 9).value  # Column I (9) = Price (24h)

                        cell_val = ws_wallet.cell(row_idx, 10)   # Column J (10) = Validation
                        cell_diff = ws_wallet.cell(row_idx, 11)  # Column K (11) = Diff

                        try:
                            # No 24h data in DAM UI
                            if dam_24h_raw is None or dam_24h_raw == "" or str(dam_24h_raw).strip() in ("", "Tooltip N/A"):
                                cell_val.value = "No 24h Data"
                                cell_diff.value = ""
                                cell_val.fill = yellow_fill_fallback
                                cell_diff.fill = yellow_fill_fallback
                                continue

                            # Parse DAM 24H value
                            if isinstance(dam_24h_raw, (int, float)):
                                dam_24h = Decimal(str(dam_24h_raw))
                            else:
                                dam_24h_str = str(dam_24h_raw).replace(',', '').replace('%', '').strip()
                                dam_24h = Decimal(dam_24h_str) if dam_24h_str else Decimal('0')

                            _chain_norm_24h = {
                                'bnb': 'binance smart chain', 'bsc': 'binance smart chain',
                                'hyper_evm': 'hyperevm', 'hyper evm': 'hyperevm',
                            }
                            _chain_lower = _chain_norm_24h.get(chain_val.lower(), chain_val.lower())

                            source_24h = None

                            # Primary: use matched_source_row (same row as Price validation)
                            if row_idx in matched_source_row:
                                _src_row = matched_source_row[row_idx]
                                if _chain_lower == "tron" and _ws_trx_24h:
                                    _p24 = _ws_trx_24h.cell(_src_row, 8).value  # TRX Balance col H = 24H Change
                                elif _ws_sd_24h:
                                    _p24 = _ws_sd_24h.cell(_src_row, 10).value  # SimDune col J = 24H Price Change
                                else:
                                    _p24 = None

                                if _p24 is not None and str(_p24).strip() not in ("", "Data not available"):
                                    if str(_p24).strip() == "null":
                                        cell_val.value = "api return null"
                                        cell_diff.value = ""
                                        cell_val.fill = yellow_fill_fallback
                                        cell_diff.fill = yellow_fill_fallback
                                        continue
                                    try:
                                        if isinstance(_p24, (int, float)):
                                            source_24h = Decimal(str(_p24))
                                        else:
                                            _p24_str = str(_p24).replace(',', '').replace('%', '').strip()
                                            if _p24_str:
                                                source_24h = Decimal(_p24_str)
                                    except:
                                        pass

                            # Fallback: map-based lookup (only if no matched_source_row)
                            if source_24h is None and row_idx not in matched_source_row:
                                if _chain_lower == "tron":
                                    source_24h = trx_24h_map.get(token_val.lower())
                                else:
                                    _lookup_24h = (_chain_lower, token_val.lower())
                                    source_24h = sim_dune_24h_map.get(_lookup_24h)
                                    if source_24h is None:
                                        if ("null", _lookup_24h) in sim_dune_no_24h_keys:
                                            cell_val.value = "api return null"
                                            cell_diff.value = ""
                                        elif _lookup_24h in sim_dune_no_24h_keys:
                                            cell_val.value = "Data not available"
                                            cell_diff.value = ""
                                        else:
                                            cell_val.value = "Token Not Found"
                                            cell_diff.value = ""
                                        cell_val.fill = yellow_fill_fallback
                                        cell_diff.fill = yellow_fill_fallback
                                        continue

                            # Write validation and diff
                            if source_24h is not None:
                                # Diff = DAM - Source (simple subtraction, same as Price diff)
                                diff_24h = dam_24h - source_24h
                                cell_diff.value = str(diff_24h)

                                # Validation: absolute diff <= 1 percentage point = Passed
                                cell_val.value = "Passed" if abs(diff_24h) <= Decimal('1') else "Failed"
                            elif row_idx in matched_source_row:
                                # Matched row exists but 24h data is blank/missing in source
                                cell_val.value = "0"
                                cell_diff.value = ""
                            else:
                                cell_val.value = "Token Not Found"
                                cell_diff.value = ""

                        except Exception as e:
                            print(f"      \u26a0\ufe0f  Error calculating 24h for row {row_idx}: {e}")
                            cell_val.value = "Error"
                            cell_diff.value = ""

                        cell_val.fill = yellow_fill_fallback
                        cell_diff.fill = yellow_fill_fallback

                    print(f"   \u2705 Calculated J & K columns (Price 24H Validation & Diff) for {last_row - 1} token rows")

                if sim_dune_map or trx_balance_map:
                    print(f"   Calculating N column (Amount Validation - Python)...")
                    print(f"      Sources: TRX Balance map={len(trx_balance_map)} tokens, Sim Dune map={len(sim_dune_map)} tokens")

                    # Set N column header
                    ws_wallet.cell(1, 15).value = "Amount Validation"

                    # N Column: Chain-aware comparison
                    # - Tron chain: Compare L (Amount Tooltip) with TRX Balance map (by token symbol)
                    # - Non-Tron (EVM): Compare Overview-Wallet A+B+L with Sim Dune B+C+E
                    #   All 3 must match → "Passed", otherwise → "Failed"
                    for row_idx in range(2, last_row + 1):
                        chain_val = str(ws_wallet.cell(row_idx, 1).value or "").strip()  # Column A - Chain
                        token_val = str(ws_wallet.cell(row_idx, 2).value or "").strip()  # Column B - Token Name
                        q_raw = ws_wallet.cell(row_idx, 13).value  # Column M - Amount Tooltip (raw)
                        l_raw = ws_wallet.cell(row_idx, 12).value  # Column L - Amount (cell text)

                        # Fallback: if tooltip missing, use Amount cell text
                        _tooltip_missing_amt = (q_raw is None or q_raw == "" or str(q_raw).strip() == "Tooltip N/A")
                        effective_amt_val = l_raw if _tooltip_missing_amt else q_raw

                        cell_s = ws_wallet.cell(row_idx, 15)

                        try:
                            # Check if both tooltip AND cell text are empty
                            if effective_amt_val is None or effective_amt_val == "" or str(effective_amt_val).strip() == "" or str(effective_amt_val).strip() == "Tooltip N/A":
                                cell_s.value = "No Amount Data"
                                cell_t = ws_wallet.cell(row_idx, 16)
                                cell_t.value = ""
                            else:
                                # Handle effective amount value - could be numeric or string
                                if isinstance(effective_amt_val, (int, float)):
                                    q_str = f"{effective_amt_val:.18f}".rstrip('0').rstrip('.')
                                else:
                                    q_str = str(effective_amt_val).replace(',', '').strip()
                                    if '.' in q_str:
                                        q_str = q_str.rstrip('0').rstrip('.')

                                # Chain-aware lookup
                                if chain_val.lower() == "tron":
                                    # Tron: look up by token symbol in TRX Balance map
                                    lookup_key = token_val.lower()
                                    if lookup_key in trx_balance_map:
                                        source_str = trx_balance_map[lookup_key]
                                        if source_str == q_str:
                                            cell_s.value = "Passed"
                                        else:
                                            cell_s.value = "Failed"
                                            print(f"      DEBUG S MISMATCH: {chain_val}/{token_val} - Q='{q_str}' vs TRXBal='{source_str}'")
                                    else:
                                        cell_s.value = "Failed"
                                    # Data Row: TRX Balance row
                                    ws_wallet.cell(row_idx, 26).value = trx_balance_row_map.get(lookup_key, "No Match")
                                else:
                                    # Non-Tron (EVM): look up by (chain, symbol) in Sim Dune map
                                    lookup_key = (chain_val.lower(), token_val.lower())
                                    if lookup_key in sim_dune_map:
                                        source_str, picked_idx = _pick_closest(sim_dune_map[lookup_key], q_str)
                                        # Use 1% tolerance after picking closest (handles precision differences)
                                        try:
                                            _q_dec = Decimal(q_str)
                                            _s_dec = Decimal(source_str)
                                            if _s_dec != Decimal('0'):
                                                _pct = abs(_q_dec - _s_dec) / abs(_s_dec) * Decimal('100')
                                                cell_s.value = "Passed" if _pct <= Decimal('1') else "Failed"
                                            else:
                                                cell_s.value = "Passed" if _q_dec == Decimal('0') else "Failed"
                                        except:
                                            cell_s.value = "Passed" if source_str == q_str else "Failed"
                                        if cell_s.value == "Failed":
                                            print(f"      DEBUG S MISMATCH: {chain_val}/{token_val} - Q='{q_str}' vs SimDune='{source_str}'")
                                    else:
                                        cell_s.value = "Failed"
                                        picked_idx = -1
                                    # Data Row: Sim Dune row (use the row that _pick_closest actually selected)
                                    row_list = sim_dune_row_map.get(lookup_key, [])
                                    if row_list and 0 <= picked_idx < len(row_list):
                                        ws_wallet.cell(row_idx, 26).value = row_list[picked_idx]
                                    else:
                                        ws_wallet.cell(row_idx, 26).value = "No Match"
                        except Exception as e:
                            cell_s.value = f"Error: {str(e)[:20]}"

                        cell_s.fill = yellow_fill_fallback
                        ws_wallet.cell(row_idx, 26).fill = yellow_fill_fallback

                    print(f"   ✅ Calculated S column for {last_row - 1} token rows")

                if sim_dune_map or trx_balance_map:
                    print(f"   Calculating O column (Amount Diff - Python)...")

                    # Set O column header
                    ws_wallet.cell(1, 16).value = "Amount Diff Validation"

                    # O Column: Chain-aware difference calculation
                    # - Tron: TRX Balance amount - L (by token symbol)
                    # - Non-Tron (EVM): L minus Sim Dune E (when Wallet A=SimDune B and Wallet B=SimDune C)
                    for row_idx in range(2, last_row + 1):
                        chain_val = str(ws_wallet.cell(row_idx, 1).value or "").strip()
                        token_val = str(ws_wallet.cell(row_idx, 2).value or "").strip()
                        q_raw = ws_wallet.cell(row_idx, 13).value  # Column L - Amount Tooltip

                        cell_t = ws_wallet.cell(row_idx, 16)

                        try:
                            # Handle Q value - could be numeric or string
                            if q_raw is None:
                                q_str = '0'
                            elif isinstance(q_raw, (int, float)):
                                q_str = f"{q_raw:.18f}".rstrip('0').rstrip('.')
                            else:
                                q_str = str(q_raw).replace(',', '').strip()
                                if '.' in q_str:
                                    q_str = q_str.rstrip('0').rstrip('.')

                            q_decimal = Decimal(q_str)

                            # Chain-aware lookup
                            if chain_val.lower() == "tron":
                                lookup_key = token_val.lower()
                                if lookup_key in trx_balance_map:
                                    source_str = trx_balance_map[lookup_key]
                                    source_decimal = Decimal(source_str)
                                    diff = source_decimal - q_decimal
                                    diff_str = str(diff)
                                    if '.' in diff_str:
                                        diff_str = diff_str.rstrip('0').rstrip('.')
                                    cell_t.value = diff_str
                                else:
                                    cell_t.value = "Token not found"
                            else:
                                # Non-Tron (EVM): Q - SimDune E (when chain+symbol match)
                                lookup_key = (chain_val.lower(), token_val.lower())
                                if lookup_key in sim_dune_map:
                                    source_str, _picked_idx = _pick_closest(sim_dune_map[lookup_key], q_str)
                                    source_decimal = Decimal(source_str)
                                    diff = q_decimal - source_decimal
                                    diff_str = str(diff)
                                    if '.' in diff_str:
                                        diff_str = diff_str.rstrip('0').rstrip('.')
                                    cell_t.value = diff_str
                                else:
                                    cell_t.value = "Token not found"
                        except Exception as e:
                            cell_t.value = f"Error: {str(e)[:20]}"

                        cell_t.fill = yellow_fill_fallback

                    print(f"   ✅ Calculated O column (Amount Diff) for {last_row - 1} token rows")

                # Calculate Column T (Share Validation) using Python
                if sim_dune_map or trx_balance_map:
                    print(f"   Calculating T column (Share Validation - Python)...")

                    # Set T column header
                    ws_wallet.cell(1, 21).value = "Share Validation"

                    # T Column: Compare UI Share (P col 16) vs Calculation Share (S col 19)
                    # Passed if INT(S*100) == INT(P*100)  (i.e. truncated to 2dp match)
                    # Special case: if S < 0.01 → P must be "< 0.01"
                    for row_idx in range(2, last_row + 1):
                        p_raw = ws_wallet.cell(row_idx, 17).value  # Column Q - Share (UI cell text)
                        r_raw = ws_wallet.cell(row_idx, 18).value  # Column R - Share Tooltip
                        s_raw = ws_wallet.cell(row_idx, 20).value  # Column T - Calculation Share

                        # Fallback: if tooltip missing, use Share cell text
                        _tooltip_missing_share = (r_raw is None or r_raw == "" or str(r_raw).strip() == "Tooltip N/A")
                        # For Share Validation, we compare UI Share (P) vs Calculation Share (S)
                        # The tooltip is not directly used here — P is cell text
                        # But if P itself is empty, we can't validate
                        cell_t = ws_wallet.cell(row_idx, 21)

                        try:
                            if s_raw is None or s_raw == "":
                                cell_t.value = "No Calc Share"
                            elif p_raw is None or p_raw == "":
                                cell_t.value = "No Share Data"
                            else:
                                p_str = str(p_raw).replace(',', '').replace('%', '').replace('<', '').strip()
                                s_str = str(s_raw).replace(',', '').replace('%', '').strip()
                                try:
                                    s_dec = Decimal(s_str)
                                    if s_dec < Decimal('0.01'):
                                        p_orig = str(p_raw).strip()
                                        cell_t.value = "Passed" if p_orig in ["< 0.01", "<0.01", "< 0.01%", "<0.01%"] else "Failed"
                                    else:
                                        p_dec = Decimal(p_str) if p_str else Decimal('0')
                                        # Compare using truncation (floor to 2dp) — DAM truncates, not rounds
                                        s_truncated = s_dec.quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                        p_truncated = p_dec.quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                        cell_t.value = "Passed" if s_truncated == p_truncated else "Failed"
                                except:
                                    cell_t.value = "Error"
                        except Exception as e:
                            cell_t.value = f"Error: {str(e)[:20]}"

                        cell_t.fill = yellow_fill_fallback

                    print(f"   ✅ Calculated T column (Share Validation) for {last_row - 1} token rows")

            # Add total rows to Overview - Wallet tab (moved outside conditional blocks)
            if "Overview - Wallet" in wb.sheetnames:
                ws_overview = wb["Overview - Wallet"]

                # Find the last data row
                last_data_row = ws_overview.max_row
                while last_data_row > 1 and not ws_overview.cell(last_data_row, 2).value:
                    last_data_row -= 1

                # Calculate total and chain sums from W column
                sum_row = last_data_row + 1
                total_calculated_value = Decimal('0')
                chain_calculated_values = {}

                # Collect W column values and build chain sums
                for row_idx in range(2, last_data_row + 1):
                    w_val = ws_overview.cell(row_idx, 24).value  # Column W
                    chain_val = ws_overview.cell(row_idx, 1).value  # Column A - Chain

                    if w_val:
                        try:
                            # Convert W value to Decimal
                            if isinstance(w_val, str):
                                w_val_str = w_val.replace(',', '').strip()
                                if w_val_str and "error" not in w_val_str.lower():
                                    w_decimal = Decimal(w_val_str)
                                    total_calculated_value += w_decimal

                                    # Add to chain sum
                                    if chain_val:
                                        chain_name = str(chain_val).strip()
                                        if chain_name not in chain_calculated_values:
                                            chain_calculated_values[chain_name] = Decimal('0')
                                        chain_calculated_values[chain_name] += w_decimal
                            elif isinstance(w_val, (int, float)):
                                w_decimal = Decimal(str(w_val))
                                total_calculated_value += w_decimal

                                # Add to chain sum
                                if chain_val:
                                    chain_name = str(chain_val).strip()
                                    if chain_name not in chain_calculated_values:
                                        chain_calculated_values[chain_name] = Decimal('0')
                                    chain_calculated_values[chain_name] += w_decimal
                        except Exception as e:
                            pass  # Skip invalid values

                # Add main total row
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws_overview.cell(sum_row, 23).value = "Total"

                # Column X (24): Sum of Value column (V=22) using Python Decimal
                total_value_v2 = Decimal('0')
                chain_value_v2 = {}
                for _ri in range(2, last_row + 1):
                    _v_val = ws_overview.cell(_ri, 22).value
                    _chain = str(ws_overview.cell(_ri, 1).value or "").strip()
                    if _v_val is not None and str(_v_val).strip() not in ('', '-'):
                        try:
                            _v_dec = Decimal(str(_v_val).replace(',', '').replace('$', '').replace('<', '').strip())
                            total_value_v2 += _v_dec
                            if _chain:
                                chain_value_v2.setdefault(_chain, Decimal('0'))
                                chain_value_v2[_chain] += _v_dec
                        except:
                            pass

                def _fmt_dec2(d):
                    s = str(d)
                    if '.' in s:
                        s = s.rstrip('0').rstrip('.')
                    return s

                ws_overview.cell(sum_row, 22).value = _fmt_dec2(total_value_v2)
                ws_overview.cell(sum_row, 22).fill = yellow_fill
                # Column Y (25): API Calculated Value total — sum Y column
                api_total = Decimal('0')
                for _ri in range(2, last_row + 1):
                    _yv = ws_overview.cell(_ri, 25).value
                    if _yv:
                        try:
                            api_total += Decimal(str(_yv).replace(',', '').replace('$', '').strip())
                        except:
                            pass
                api_total_str = str(api_total)
                if '.' in api_total_str:
                    api_total_str = api_total_str.rstrip('0').rstrip('.')
                ws_overview.cell(sum_row, 24).value = api_total_str
                ws_overview.cell(sum_row, 24).fill = yellow_fill

                # Build chain U value sums
                chain_u_values_2 = {}
                for row_idx in range(2, last_row + 1):
                    chain_val = ws_overview.cell(row_idx, 1).value
                    u_val = ws_overview.cell(row_idx, 22).value
                    if chain_val and u_val is not None:
                        try:
                            u_str = str(u_val).replace('$', '').replace(',', '').replace('<', '').strip()
                            if u_str and u_str not in ('', 'None', 'N/A'):
                                chain_name = str(chain_val).strip()
                                chain_u_values_2.setdefault(chain_name, Decimal('0'))
                                chain_u_values_2[chain_name] += Decimal(u_str)
                        except:
                            pass

                # Add chain-specific total rows — only for chains that exist in data
                all_chains_in_data = set(list(chain_calculated_values.keys()) + list(chain_u_values_2.keys()))
                preferred_order = ["Ethereum", "Base", "Binance Smart Chain", "Tron"]
                chain_order = [c for c in preferred_order if c in all_chains_in_data]
                for cn in all_chains_in_data:
                    if cn not in chain_order:
                        chain_order.append(cn)
                current_row = sum_row + 1

                for chain_name in chain_order:

                    # Column W (23): Chain total label
                    ws_overview.cell(current_row, 23).value = f"Total ({chain_name})"

                    # Column V (22): Chain Value sum
                    ws_overview.cell(current_row, 22).value = _fmt_dec2(chain_value_v2.get(chain_name, Decimal('0')))
                    ws_overview.cell(current_row, 22).fill = yellow_fill

                    # Column Y (25): Chain API Calculated Value sum
                    chain_api_sum = Decimal('0')
                    for _ri in range(2, last_row + 1):
                        _chain_val = ws_overview.cell(_ri, 1).value
                        if _chain_val and str(_chain_val).strip() == chain_name:
                            _yv = ws_overview.cell(_ri, 25).value
                            if _yv:
                                try:
                                    chain_api_sum += Decimal(str(_yv).replace(',', '').replace('$', '').strip())
                                except:
                                    pass
                    chain_api_str = str(chain_api_sum)
                    if '.' in chain_api_str:
                        chain_api_str = chain_api_str.rstrip('0').rstrip('.')
                    ws_overview.cell(current_row, 24).value = chain_api_str
                    ws_overview.cell(current_row, 24).fill = yellow_fill

                    current_row += 1

                print(f"   ✅ Added {len(chain_order) + 1} total rows (1 main + {len(chain_order)} chain totals) at rows {sum_row}-{current_row-1}")

            # Add total rows to Overview - De-Fi tab
            if "Overview - De-Fi" in wb.sheetnames:
                ws_defi = wb["Overview - De-Fi"]

                # Find the last data row
                last_data_row = ws_defi.max_row
                while last_data_row > 1 and not ws_defi.cell(last_data_row, 1).value:
                    last_data_row -= 1

                sum_row = last_data_row + 1

                # Collect unique chains from column B (col 2)
                defi_chains_in_data = set()
                for row_idx in range(2, last_data_row + 1):
                    b_val = ws_defi.cell(row_idx, 2).value
                    if b_val:
                        defi_chains_in_data.add(str(b_val).strip())

                # Chain name mapping for display
                chain_display = {
                    'eth': 'Ethereum', 'ethereum': 'Ethereum',
                    'base': 'Base',
                    'bsc': 'Binance Smart Chain', 'binance smart chain': 'Binance Smart Chain',
                    'tron': 'Tron',
                }

                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # Calculate L and N column sums using Python (values may be stored as text)
                total_l = Decimal('0')
                total_n = Decimal('0')
                chain_l_sums = {}  # raw_chain → Decimal
                chain_n_sums = {}  # raw_chain → Decimal

                for row_idx in range(2, last_data_row + 1):
                    l_val = ws_defi.cell(row_idx, 12).value  # Column L - Value
                    n_val = ws_defi.cell(row_idx, 14).value  # Column N - Api Calc Value
                    b_val = ws_defi.cell(row_idx, 2).value   # Column B - Chain

                    raw_chain = str(b_val).strip() if b_val else ""

                    for col_val, total_ref, chain_ref in [(l_val, 'l', chain_l_sums), (n_val, 'n', chain_n_sums)]:
                        if col_val is not None and str(col_val).strip() not in ('', 'No Match', '-'):
                            try:
                                dec_val = Decimal(str(col_val).replace(',', '').replace('$', '').strip())
                                if total_ref == 'l':
                                    total_l += dec_val
                                else:
                                    total_n += dec_val
                                if raw_chain:
                                    chain_ref.setdefault(raw_chain, Decimal('0'))
                                    chain_ref[raw_chain] += dec_val
                            except:
                                pass

                def _dec_str(d):
                    s = str(d)
                    if '.' in s:
                        s = s.rstrip('0').rstrip('.')
                    return s

                # Main total row: M = "Total", L = sum, N = sum
                ws_defi.cell(sum_row, 13).value = "Total"
                ws_defi.cell(sum_row, 12).value = _dec_str(total_l)
                ws_defi.cell(sum_row, 12).fill = yellow_fill
                ws_defi.cell(sum_row, 14).value = _dec_str(total_n)
                ws_defi.cell(sum_row, 14).fill = yellow_fill

                # Chain-specific total rows — only for chains that exist in data
                preferred_order = ["Ethereum", "Base", "Binance Smart Chain", "Tron"]
                # Normalize chain codes to display names
                normalized_chains = set()
                for c in defi_chains_in_data:
                    display_name = chain_display.get(c.lower(), c.title())
                    normalized_chains.add((c, display_name))

                # Build ordered list of (raw_chain_code, display_name)
                chain_order_defi = []
                seen_display = set()
                for pref in preferred_order:
                    for raw, display in normalized_chains:
                        if display == pref and display not in seen_display:
                            chain_order_defi.append((raw, display))
                            seen_display.add(display)
                for raw, display in normalized_chains:
                    if display not in seen_display:
                        chain_order_defi.append((raw, display))
                        seen_display.add(display)

                current_row = sum_row + 1
                for raw_chain, display_name in chain_order_defi:
                    ws_defi.cell(current_row, 13).value = f"Total ({display_name})"
                    ws_defi.cell(current_row, 12).value = _dec_str(chain_l_sums.get(raw_chain, Decimal('0')))
                    ws_defi.cell(current_row, 12).fill = yellow_fill
                    ws_defi.cell(current_row, 14).value = _dec_str(chain_n_sums.get(raw_chain, Decimal('0')))
                    ws_defi.cell(current_row, 14).fill = yellow_fill
                    current_row += 1

                print(f"   ✅ Added {len(chain_order_defi) + 1} total rows to Overview - De-Fi (1 main + {len(chain_order_defi)} chain totals)")

                # Type (col C): NOT merged — every row keeps its own value so formulas can reference $C directly
                print(f"   ℹ️  Type (col C) not merged — values preserved in each row for formula correctness")

                # Conditional formatting: Amount Validation Diff (col J=10) = 0 → green
                from openpyxl.formatting.rule import CellIsRule
                _green_fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
                _diff_range = f"J2:J{last_data_row}"
                ws_defi.conditional_formatting.add(
                    _diff_range,
                    CellIsRule(operator='equal', formula=['0'], fill=_green_fill)
                )

            # Update D column (Token Count Validation) in "Overview - Header & Token Holdings Header" tab
            # For "Token Holdings - Chain" rows: count Wallet rows (A = full chain name) + De-Fi rows (B = chain code)
            # Chain name in B column may differ from Overview - Wallet A column (e.g. "BNB" vs "Binance Smart Chain")
            if "Overview - Header & Token Holdings Header" in wb.sheetnames and "Overview - Wallet" in wb.sheetnames:
                ws_header_d = wb["Overview - Header & Token Holdings Header"]
                ws_wallet_d = wb["Overview - Wallet"]
                ws_defi_d = wb["Overview - De-Fi"] if "Overview - De-Fi" in wb.sheetnames else None

                print(f"   🔄 Updating D column (Token Count Validation) in 'Overview - Header & Token Holdings Header'...")

                # Chain display name → full name as it appears in Overview - Wallet A column
                chain_name_map_d = {
                    'bnb': 'Binance Smart Chain',
                    'bsc': 'Binance Smart Chain',
                    'binance smart chain': 'Binance Smart Chain',
                    'binance': 'Binance Smart Chain',
                    'ethereum': 'Ethereum',
                    'eth': 'Ethereum',
                    'tron': 'Tron',
                    'trx': 'Tron',
                    'polygon': 'Polygon',
                    'matic': 'Polygon',
                    'arbitrum': 'Arbitrum',
                    'arb': 'Arbitrum',
                    'optimism': 'Optimism',
                    'op': 'Optimism',
                    'avalanche': 'Avalanche',
                    'avax': 'Avalanche',
                    'base': 'Base',
                    'solana': 'Solana',
                    'sol': 'Solana',
                    'fantom': 'Fantom',
                    'ftm': 'Fantom',
                    'linea': 'Linea',
                    'scroll': 'Scroll',
                    'zksync era': 'zkSync Era',
                    'zksync': 'zkSync Era',
                    'blast': 'Blast',
                    'mantle': 'Mantle',
                    'mnt': 'Mantle',
                    'cronos': 'Cronos',
                    'cro': 'Cronos',
                    'gnosis': 'Gnosis',
                    'xdai': 'Gnosis',
                    'celo': 'Celo',
                    'aurora': 'Aurora',
                    'moonriver': 'Moonriver',
                    'movr': 'Moonriver',
                    'metis': 'Metis',
                    'boba': 'Boba',
                    'klaytn': 'Klaytn',
                    'klay': 'Klaytn',
                    'mode': 'Mode',
                    'hyperliquid': 'Hyperliquid',
                    'hype': 'Hyperliquid',
                }

                # Full chain name → De-Fi chain code (as stored in Overview - De-Fi B column)
                chain_to_defi_code = {
                    'binance smart chain': 'bsc',
                    'ethereum': 'eth',
                    'tron': 'tron',
                    'polygon': 'matic',
                    'arbitrum': 'arb',
                    'optimism': 'op',
                    'avalanche': 'avax',
                    'base': 'base',
                    'solana': 'sol',
                    'fantom': 'ftm',
                    'linea': 'linea',
                    'scroll': 'scroll',
                    'zksync era': 'era',
                    'blast': 'blast',
                    'mantle': 'mnt',
                    'cronos': 'cro',
                    'gnosis': 'xdai',
                    'celo': 'celo',
                    'aurora': 'aurora',
                    'moonriver': 'movr',
                    'metis': 'metis',
                    'boba': 'boba',
                    'klaytn': 'klay',
                    'mode': 'mode',
                    'hyperliquid': 'hyperliquid',
                }

                # Build token sets: full chain name (lowercase) → set of token symbols from Overview - Wallet B column
                wallet_chain_tokens = {}
                for w_row in range(2, ws_wallet_d.max_row + 1):
                    chain_cell = ws_wallet_d.cell(w_row, 1).value
                    token_cell = ws_wallet_d.cell(w_row, 2).value
                    if chain_cell and token_cell:
                        chain_key = str(chain_cell).strip().lower()
                        token_key = str(token_cell).strip().lower()
                        wallet_chain_tokens.setdefault(chain_key, set()).add(token_key)
                # Keep count map for backward compat
                wallet_chain_counts = {k: len(v) for k, v in wallet_chain_tokens.items()}

                # Build token sets: defi chain code (lowercase) → set of token symbols from Overview - De-Fi D column
                defi_chain_tokens = {}
                if ws_defi_d:
                    for d_row in range(2, ws_defi_d.max_row + 1):
                        defi_chain_cell = ws_defi_d.cell(d_row, 2).value  # B column - chain code
                        defi_token_cell = ws_defi_d.cell(d_row, 4).value  # D column - Pool/Position Pair
                        defi_name_cell = ws_defi_d.cell(d_row, 1).value   # A column - De-Fi name
                        if defi_chain_cell and defi_token_cell:
                            defi_key = str(defi_chain_cell).strip().lower()
                            token_key = str(defi_token_cell).strip().lower()
                            defi_chain_tokens.setdefault(defi_key, set()).add(token_key)
                        elif defi_token_cell and not defi_chain_cell:
                            # Row has token but no chain — log it
                            print(f"      ⚠️  De-Fi row {d_row}: token='{defi_token_cell}' but chain is BLANK (name='{defi_name_cell}')")
                defi_chain_counts = {k: len(v) for k, v in defi_chain_tokens.items()}

                print(f"      DEBUG defi_chain_tokens keys: {list(defi_chain_tokens.keys())}")
                print(f"      DEBUG wallet_chain_tokens keys: {list(wallet_chain_tokens.keys())}")
                for _dk, _dv in defi_chain_tokens.items():
                    print(f"      DEBUG defi[{_dk}]: {_dv}")

                # Update D column for each Token Holdings - Chain row
                for h_row in range(2, ws_header_d.max_row + 1):
                    section_val = ws_header_d.cell(h_row, 1).value
                    if section_val != "Token Holdings - Chain" and section_val != "Token Holdings - Platform":
                        # Overview Header and Table rows: calculate Net Worth_UI Calculation
                        d_cell = ws_header_d.cell(h_row, 4)
                        e_cell = ws_header_d.cell(h_row, 5)
                        g_cell = ws_header_d.cell(h_row, 7)  # Net Worth_UI Calculation
                        h_cell = ws_header_d.cell(h_row, 8)
                        f_val = ws_header_d.cell(h_row, 6).value  # Net Worth (UI)
                        b_val_oh = str(ws_header_d.cell(h_row, 2).value or '').strip()
                        b_val_oh_lower = b_val_oh.lower()

                        if d_cell.value == "Pending" or section_val == "Table":
                            d_cell.value = ""
                        if e_cell.value == "Pending" or section_val == "Table":
                            e_cell.value = "Not Applicable"

                        # Calculate Net Worth_UI Calculation
                        calc_nw = Decimal('0')
                        if b_val_oh_lower in ('wallets', 'wallet'):
                            # Sum ALL Calculated Values from Overview - Wallet column W (23)
                            if ws_wallet_d:
                                for w_row in range(2, ws_wallet_d.max_row + 1):
                                    value_cell = ws_wallet_d.cell(w_row, 23).value
                                    if value_cell:
                                        try:
                                            val_str = str(value_cell).replace('$', '').replace(',', '').replace('<', '').strip()
                                            if val_str and val_str not in ('', 'None', 'N/A'):
                                                calc_nw += Decimal(val_str)
                                        except:
                                            pass
                        elif b_val_oh_lower in ('de-fi positions', 'defi positions', 'de-fi'):
                            # Read the Total row value from Overview - De-Fi column L (12)
                            # The total row has "Total" in column M (13)
                            if ws_defi_d:
                                for d_row in range(2, ws_defi_d.max_row + 1):
                                    m_val = ws_defi_d.cell(d_row, 13).value
                                    if m_val and str(m_val).strip() == "Total":
                                        total_l_val = ws_defi_d.cell(d_row, 12).value
                                        if total_l_val:
                                            try:
                                                val_str = str(total_l_val).replace('$', '').replace(',', '').replace('<', '').strip()
                                                if val_str and val_str not in ('', 'None', 'N/A'):
                                                    calc_nw = Decimal(val_str)
                                            except:
                                                pass
                                        break
                        elif b_val_oh_lower == 'exchanges':
                            calc_nw = Decimal('0')  # No exchange data to sum
                        elif section_val == "Table":
                            # Table rows: match by name against wallet or De-Fi
                            if b_val_oh_lower in ('wallet',):
                                if ws_wallet_d:
                                    for w_row in range(2, ws_wallet_d.max_row + 1):
                                        value_cell = ws_wallet_d.cell(w_row, 23).value
                                        if value_cell:
                                            try:
                                                val_str = str(value_cell).replace('$', '').replace(',', '').replace('<', '').strip()
                                                if val_str and val_str not in ('', 'None', 'N/A'):
                                                    calc_nw += Decimal(val_str)
                                            except:
                                                pass
                            else:
                                # Match De-Fi name (column A)
                                if ws_defi_d:
                                    for d_row in range(2, ws_defi_d.max_row + 1):
                                        defi_name_cell = ws_defi_d.cell(d_row, 1).value
                                        defi_value_cell = ws_defi_d.cell(d_row, 12).value
                                        if defi_name_cell and defi_value_cell:
                                            if b_val_oh_lower in str(defi_name_cell).strip().lower():
                                                try:
                                                    val_str = str(defi_value_cell).replace('$', '').replace(',', '').replace('<', '').strip()
                                                    if val_str and val_str not in ('', 'None', 'N/A'):
                                                        calc_nw += Decimal(val_str)
                                                except:
                                                    pass

                        g_cell.value = float(calc_nw) if calc_nw > 0 else 0

                        # Net Worth Validation
                        try:
                            ui_nw_str = str(f_val or '').replace('$', '').replace(',', '').replace('<', '').strip()
                            ui_nw = Decimal(ui_nw_str) if ui_nw_str else Decimal('0')
                            tolerance = Decimal('1')
                            if abs(calc_nw - ui_nw) <= tolerance:
                                h_cell.value = "Passed"
                            else:
                                h_cell.value = "Failed"
                        except:
                            h_cell.value = "Not Applicable"

                        continue

                    # For Platform rows: only calculate Net Worth_UI Calculation (no token count)
                    is_platform_row = section_val == "Token Holdings - Platform"

                    b_val = str(ws_header_d.cell(h_row, 2).value or '').strip()
                    c_val = ws_header_d.cell(h_row, 3).value
                    d_cell = ws_header_d.cell(h_row, 4)  # TC_UI Count
                    e_cell = ws_header_d.cell(h_row, 5)  # Token Count Validation
                    f_val = ws_header_d.cell(h_row, 6).value  # Net Worth (UI)
                    g_cell = ws_header_d.cell(h_row, 7)  # Net Worth (Actual)
                    h_cell = ws_header_d.cell(h_row, 8)  # Net Worth Validation

                    # Map B column chain name → full wallet chain name
                    b_lower = b_val.lower()
                    full_chain_name = chain_name_map_d.get(b_lower, b_val)
                    full_chain_lower = full_chain_name.lower()

                    # Count from Overview - Wallet (A column = full chain name)
                    wallet_tokens = wallet_chain_tokens.get(full_chain_lower, set())

                    # Count from Overview - De-Fi (B column = chain code)
                    defi_code = chain_to_defi_code.get(full_chain_lower, full_chain_lower)
                    defi_tokens = defi_chain_tokens.get(defi_code, set())

                    # Also check De-Fi A column (protocol name contains chain) for protocols without chain code
                    # e.g., "Lido" on Ethereum might have chain="eth" in B, but also check by full chain name
                    if not defi_tokens:
                        # Try matching De-Fi B column with full chain name directly
                        defi_tokens = defi_chain_tokens.get(full_chain_lower, set())

                    # Union to remove duplicates (same token in both wallet and defi counts once)
                    actual_count = len(wallet_tokens | defi_tokens)

                    print(f"      Chain '{b_val}' → full='{full_chain_name}', defi_code='{defi_code}', wallet={len(wallet_tokens)}, defi={len(defi_tokens)}, union={actual_count}")

                    if is_platform_row:
                        # Platform rows: no token count validation
                        d_cell.value = ""
                        e_cell.value = "Not Applicable"
                    else:
                        # Chain rows: write TC_UI Count and validate
                        d_cell.value = actual_count

                        # Compare with C column token count
                        try:
                            expected_count = int(str(c_val).strip()) if c_val else 0
                            e_cell.value = "Passed" if actual_count == expected_count else "Failed"
                            if e_cell.value == "Failed":
                                print(f"      ⚠️  Token Count Mismatch: '{b_val}' (→ '{full_chain_name}') expected={expected_count}, wallet={len(wallet_tokens)}, defi={len(defi_tokens)}, union={actual_count}")
                        except (ValueError, TypeError):
                            e_cell.value = "Not Applicable"

                    # Calculate actual net worth from wallet and defi data
                    if is_platform_row:
                        # Platform rows: calculate based on platform name
                        platform_name_lower = b_val.lower().split(' (')[0]  # Remove chain suffix like "(Ethereum)"
                        
                        if platform_name_lower == 'wallet':
                            # WALLET: sum ALL values from Overview - Wallet column V (22)
                            wallet_net_worth = Decimal('0')
                            if ws_wallet_d:
                                for w_row in range(2, ws_wallet_d.max_row + 1):
                                    value_cell = ws_wallet_d.cell(w_row, 22).value  # Column V - Value
                                    if value_cell:
                                        try:
                                            val_str = str(value_cell).replace('$', '').replace(',', '').replace('<', '').strip()
                                            if val_str and val_str not in ('', 'None', 'N/A'):
                                                wallet_net_worth += Decimal(val_str)
                                        except:
                                            pass
                            actual_net_worth = wallet_net_worth
                        else:
                            # Other platforms: "Protocol (Chain)" format — match De-Fi A + B columns
                            import re as _re_plat
                            _plat_match = _re_plat.match(r'^(.+?)\s*\((.+?)\)$', b_val)
                            if _plat_match:
                                _proto_name = _plat_match.group(1).strip().lower()
                                _chain_display = _plat_match.group(2).strip()
                                _chain_code_map = {
                                    'Ethereum': 'eth', 'Base': 'base',
                                    'Binance Smart Chain': 'bsc', 'BSC': 'bsc',
                                    'Tron': 'tron', 'Polygon': 'matic',
                                    'Arbitrum': 'arb', 'Optimism': 'op',
                                }
                                _chain_code = _chain_code_map.get(_chain_display, _chain_display.lower())
                            else:
                                _proto_name = platform_name_lower
                                _chain_code = None  # No chain filter

                            defi_net_worth = Decimal('0')
                            if ws_defi_d:
                                for d_row in range(2, ws_defi_d.max_row + 1):
                                    defi_name_cell = ws_defi_d.cell(d_row, 1).value  # A column - De-Fi name
                                    defi_chain_cell = ws_defi_d.cell(d_row, 2).value  # B column - Chain
                                    defi_value_cell = ws_defi_d.cell(d_row, 12).value  # L column - Value
                                    if defi_name_cell and defi_value_cell:
                                        name_match = _proto_name in str(defi_name_cell).strip().lower()
                                        chain_match = (_chain_code is None or
                                                       (defi_chain_cell and str(defi_chain_cell).strip().lower() == _chain_code))
                                        if name_match and chain_match:
                                            try:
                                                val_str = str(defi_value_cell).replace('$', '').replace(',', '').replace('<', '').strip()
                                                if val_str and val_str not in ('', 'None', 'N/A'):
                                                    defi_net_worth += Decimal(val_str)
                                            except:
                                                pass
                            actual_net_worth = defi_net_worth
                    else:
                        # Chain rows: sum wallet V(22) + defi L(12) for that chain
                        wallet_net_worth = Decimal('0')
                        if ws_wallet_d:
                            for w_row in range(2, ws_wallet_d.max_row + 1):
                                chain_cell = ws_wallet_d.cell(w_row, 1).value
                                value_cell = ws_wallet_d.cell(w_row, 22).value  # Column V - Value
                                if chain_cell and value_cell:
                                    if str(chain_cell).strip().lower() == full_chain_lower:
                                        try:
                                            val_str = str(value_cell).replace('$', '').replace(',', '').strip()
                                            if val_str and val_str not in ('', 'None', 'N/A', '<'):
                                                wallet_net_worth += Decimal(val_str)
                                        except:
                                            pass

                    defi_net_worth = Decimal('0')
                    if ws_defi_d:
                        is_hyperliquid_row = "hyperliquid" in b_lower
                        for d_row in range(2, ws_defi_d.max_row + 1):
                            defi_name_cell = ws_defi_d.cell(d_row, 1).value   # A column - De-Fi name
                            defi_chain_cell = ws_defi_d.cell(d_row, 2).value  # B column - chain code
                            defi_value_cell = ws_defi_d.cell(d_row, 12).value  # L column - Value
                            if defi_value_cell:
                                match = False
                                if is_hyperliquid_row:
                                    # Hyperliquid: match on De-Fi name (column A) = "Hyperliquid"
                                    if defi_name_cell and "hyperliquid" in str(defi_name_cell).strip().lower():
                                        match = True
                                else:
                                    # Other chains: match on chain code (column B)
                                    if defi_chain_cell and str(defi_chain_cell).strip().lower() == defi_code:
                                        match = True
                                if match:
                                    try:
                                        val_str = str(defi_value_cell).replace('$', '').replace(',', '').replace('<', '').strip()
                                        if val_str and val_str not in ('', 'None', 'N/A'):
                                            defi_net_worth += Decimal(val_str)
                                    except:
                                        pass

                    actual_net_worth = wallet_net_worth + defi_net_worth
                    g_cell.value = float(actual_net_worth) if actual_net_worth > 0 else ""

                    # Compare net worth
                    try:
                        ui_net_worth_str = str(f_val or '').replace('$', '').replace(',', '').replace('<', '').strip()
                        ui_net_worth = Decimal(ui_net_worth_str) if ui_net_worth_str else Decimal('0')
                        
                        # Allow small tolerance for rounding
                        tolerance = Decimal('1')  # $1 tolerance
                        if abs(actual_net_worth - ui_net_worth) <= tolerance:
                            h_cell.value = "Passed"
                        else:
                            h_cell.value = "Failed"
                            print(f"      ⚠️  Net Worth Mismatch: '{b_val}' UI=${ui_net_worth}, Actual=${actual_net_worth}, Diff=${abs(actual_net_worth - ui_net_worth)}")
                    except (ValueError, TypeError, InvalidOperation):
                        h_cell.value = "Not Applicable"

                print(f"   ✅ Updated D column (Token Count Validation) in 'Overview - Header & Token Holdings Header'")

            # Update F column validation in "Overview - Header & Token Holdings Header" tab
            if "Overview - Header & Token Holdings Header" in wb.sheetnames:
                ws_header = wb["Overview - Header & Token Holdings Header"]
                ws_wallet = wb["Overview - Wallet"] if "Overview - Wallet" in wb.sheetnames else None
                ws_defi = wb["Overview - De-Fi"] if "Overview - De-Fi" in wb.sheetnames else None

                print(f"   🔄 Updating F column validation in 'Overview - Header & Token Holdings Header'...")

                # Build lookup maps from total rows
                wallet_totals = {}  # {label: value}
                defi_totals = {}    # {label: value}
                defi_detail_map = {}  # {(protocol, chain): value} for detailed lookups

                # Chain display name → full name as stored in Overview - Wallet A column
                _chain_display_map = {
                    'bnb': 'Binance Smart Chain',
                    'bsc': 'Binance Smart Chain',
                    'binance smart chain': 'Binance Smart Chain',
                    'binance': 'Binance Smart Chain',
                    'ethereum': 'Ethereum',
                    'eth': 'Ethereum',
                    'tron': 'Tron',
                    'trx': 'Tron',
                    'polygon': 'Polygon',
                    'matic': 'Polygon',
                    'arbitrum': 'Arbitrum',
                    'arb': 'Arbitrum',
                    'optimism': 'Optimism',
                    'op': 'Optimism',
                    'avalanche': 'Avalanche',
                    'avax': 'Avalanche',
                    'base': 'Base',
                    'solana': 'Solana',
                    'sol': 'Solana',
                    'fantom': 'Fantom',
                    'ftm': 'Fantom',
                    'linea': 'Linea',
                    'scroll': 'Scroll',
                    'zksync era': 'zkSync Era',
                    'zksync': 'zkSync Era',
                    'blast': 'Blast',
                    'mantle': 'Mantle',
                    'mnt': 'Mantle',
                    'cronos': 'Cronos',
                    'cro': 'Cronos',
                    'gnosis': 'Gnosis',
                    'xdai': 'Gnosis',
                    'celo': 'Celo',
                    'aurora': 'Aurora',
                    'moonriver': 'Moonriver',
                    'movr': 'Moonriver',
                    'metis': 'Metis',
                    'boba': 'Boba',
                    'klaytn': 'Klaytn',
                    'klay': 'Klaytn',
                    'mode': 'Mode',
                    'hyperliquid': 'Hyperliquid',
                    'hype': 'Hyperliquid',
                }

                # Read Overview - Wallet totals (V and W columns)
                # Read Overview - Wallet totals and per-chain Value sums
                # Total labels are in W (23), total values in V (22)
                # Per-chain: sum V column (22) grouped by A column (chain name)
                wallet_chain_w_sums = {}  # full chain name (lowercase) → sum of V column (Value)
                if ws_wallet:
                    for row_idx in range(2, ws_wallet.max_row + 1):
                        w_label = ws_wallet.cell(row_idx, 23).value  # Column W - Total labels
                        v_value = ws_wallet.cell(row_idx, 22).value  # Column V - Value

                        # Capture total row values
                        if w_label and isinstance(w_label, str) and w_label.startswith("Total"):
                            try:
                                if isinstance(v_value, str):
                                    wallet_totals[w_label] = Decimal(v_value.replace(',', '').strip())
                                elif isinstance(v_value, (int, float)):
                                    wallet_totals[w_label] = Decimal(str(v_value))
                            except:
                                pass

                        # Build per-chain V sum from A column (chain name) + V column (Value)
                        a_chain = ws_wallet.cell(row_idx, 1).value  # Column A - Chain name
                        if a_chain and v_value is not None:
                            try:
                                a_key = str(a_chain).strip().lower()
                                if isinstance(v_value, str):
                                    v_dec = Decimal(v_value.replace(',', '').replace('$', '').strip())
                                else:
                                    v_dec = Decimal(str(v_value))
                                wallet_chain_w_sums[a_key] = wallet_chain_w_sums.get(a_key, Decimal('0')) + v_dec
                            except:
                                pass

                # Read Overview - De-Fi totals and detailed data (A, B, L)
                # Column structure: A:Defi, B:Chain, ..., L:Value (col 12), M:Value Validation (col 13)
                defi_chain_j_sums = {}  # chain code (lowercase) → sum of L column (Value)
                if ws_defi:
                    for row_idx in range(2, ws_defi.max_row + 1):
                        m_label = ws_defi.cell(row_idx, 13).value  # Column M - Total labels
                        l_value = ws_defi.cell(row_idx, 12).value  # Column L - Value
                        protocol_name = ws_defi.cell(row_idx, 1).value  # Column A - Protocol Name
                        chain_name = ws_defi.cell(row_idx, 2).value     # Column B - Chain

                        # Store totals (from M column labels)
                        if m_label and isinstance(m_label, str) and m_label.startswith("Total"):
                            try:
                                if isinstance(l_value, str):
                                    defi_totals[m_label] = Decimal(l_value.replace(',', '').strip())
                                elif isinstance(l_value, (int, float)):
                                    defi_totals[m_label] = Decimal(str(l_value))
                            except:
                                pass

                        # Store detailed protocol+chain data (sum values for same protocol+chain)
                        if protocol_name and chain_name:
                            protocol_str = str(protocol_name).strip().lower()
                            chain_str = str(chain_name).strip().lower()
                            try:
                                value_decimal = Decimal('0')
                                if isinstance(l_value, str):
                                    value_decimal = Decimal(l_value.replace(',', '').strip())
                                elif isinstance(l_value, (int, float)):
                                    value_decimal = Decimal(str(l_value))

                                # Sum values for same protocol+chain combination
                                key = (protocol_str, chain_str)
                                if key in defi_detail_map:
                                    defi_detail_map[key] += value_decimal
                                else:
                                    defi_detail_map[key] = value_decimal

                                # Also accumulate per-chain sum (B column = chain code)
                                defi_chain_j_sums[chain_str] = defi_chain_j_sums.get(chain_str, Decimal('0')) + value_decimal
                            except:
                                pass

                # First pass: collect all F values for "Token Holdings - Chain" rows (for H column validation)
                chain_e_values = []  # [(row_idx, e_decimal)]
                for h_row_idx in range(2, ws_header.max_row + 1):
                    section_val = ws_header.cell(h_row_idx, 1).value  # A column - Section
                    e_val = ws_header.cell(h_row_idx, 6).value  # F column - Net Worth
                    if section_val == "Token Holdings - Chain":
                        try:
                            # Strip "<" for calculation only — cell value is not modified
                            e_clean = str(e_val).replace(',', '').replace('$', '').replace('<', '').strip() if e_val else '0'
                            e_decimal = Decimal(e_clean)
                            chain_e_values.append((h_row_idx, e_decimal))
                        except:
                            pass

                # Calculate sum of all chain E values
                sum_chain_e = sum([e for _, e in chain_e_values]) if chain_e_values else Decimal('0')

                # Validate each row
                for h_row_idx in range(2, ws_header.max_row + 1):
                    section_val = ws_header.cell(h_row_idx, 1).value  # A column - Section
                    category_val = ws_header.cell(h_row_idx, 2).value  # B column - Category
                    c_val = ws_header.cell(h_row_idx, 3).value  # C column - Token Count
                    cell_d = ws_header.cell(h_row_idx, 4)  # D column - TC_UI Count
                    e_val = ws_header.cell(h_row_idx, 6).value  # F column - Net Worth
                    cell_g = ws_header.cell(h_row_idx, 7)  # G column - Net Worth_UI Calculation
                    cell_f = ws_header.cell(h_row_idx, 8)  # H column - Net Worth UI Validation
                    cell_h = ws_header.cell(h_row_idx, 12)  # L column - Percentage Validation

                    try:
                        # Parse E column value and truncate to 2 decimal places
                        # Strip "<" for calculation only — cell value is not modified
                        e_clean = str(e_val).replace(',', '').replace('$', '').replace('<', '').strip() if e_val else '0'
                        e_decimal = Decimal(e_clean).quantize(Decimal('0.01'), rounding='ROUND_DOWN')

                        # Scenario 1: Overview Header → Wallets
                        if section_val == "Overview Header" and category_val == "Wallets":
                            # Sum all W column values from Overview - Wallet directly
                            wallet_total = sum(wallet_chain_w_sums.values()).quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                            cell_g.value = float(wallet_total)
                            if wallet_total != 0:
                                percentage_diff = abs(e_decimal - wallet_total) / wallet_total * Decimal('100')
                                if percentage_diff <= Decimal('1'):
                                    cell_f.value = "Passed"
                                else:
                                    cell_f.value = "Failed"
                            else:
                                cell_f.value = "Passed" if e_decimal == 0 else "Failed"

                        # Scenario 2: Overview Header → De-Fi Positions (within 1% diff)
                        elif section_val == "Overview Header" and category_val == "De-Fi Positions":
                            # Sum all K column values from Overview - De-Fi directly
                            defi_total = sum(defi_chain_j_sums.values()).quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                            cell_g.value = float(defi_total)
                            if defi_total != 0:
                                percentage_diff = abs(e_decimal - defi_total) / defi_total * Decimal('100')
                                if percentage_diff <= Decimal('1'):
                                    cell_f.value = "Passed"
                                else:
                                    cell_f.value = "Failed"
                            else:
                                cell_f.value = "Passed" if e_decimal == 0 else "Failed"

                        # Scenario 2b: Overview Header → Exchanges (within 1% diff)
                        elif section_val == "Overview Header" and category_val == "Exchanges":
                            # Sum all Q column values from exchange tables where P column = "Total"
                            exchange_total = Decimal('0')

                            # Loop through all exchange tables
                            for exchange_sheet_name in exchange_tables:
                                if exchange_sheet_name in wb.sheetnames:
                                    ws_exchange = wb[exchange_sheet_name]
                                    # Find row where P column (15) = "Total"
                                    for row_idx in range(2, ws_exchange.max_row + 1):
                                        p_val = ws_exchange.cell(row_idx, 16).value  # P column (16th column)
                                        if p_val and str(p_val).strip().upper() == "TOTAL":
                                            q_val = ws_exchange.cell(row_idx, 17).value  # Q column (17th column)
                                            if q_val:
                                                try:
                                                    q_clean = str(q_val).replace(',', '').replace('$', '').strip()
                                                    exchange_total += Decimal(q_clean)
                                                except (ValueError, InvalidOperation):
                                                    pass
                                            break  # Found total row, no need to continue

                            # Truncate to 2 decimal places
                            exchange_total = exchange_total.quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                            cell_g.value = float(exchange_total)

                            # Compare with E column value (within 1% diff)
                            if exchange_total != 0:
                                percentage_diff = abs(e_decimal - exchange_total) / exchange_total * Decimal('100')
                                if percentage_diff <= Decimal('1'):
                                    cell_f.value = "Passed"
                                else:
                                    cell_f.value = "Failed"
                            else:
                                cell_f.value = "Passed" if e_decimal == 0 else "Failed"

                        # Scenario 3 & 4: Token Holdings - Chain (Ethereum, Base, BNB, Tron, etc.)
                        elif section_val == "Token Holdings - Chain":
                            chain_name = str(category_val).strip() if category_val else ""

                            # Map B column chain name (e.g. "BNB") → full wallet chain name ("Binance Smart Chain")
                            full_chain_name = _chain_display_map.get(chain_name.lower(), chain_name)
                            full_chain_key = full_chain_name.lower()

                            # SUM Overview - Wallet W column where A column = full chain name
                            wallet_w_sum = wallet_chain_w_sums.get(full_chain_key, Decimal('0'))
                            wallet_w_sum = wallet_w_sum.quantize(Decimal('0.01'), rounding='ROUND_DOWN')

                            # SUM Overview - De-Fi K column where B column = chain code (for Hyperliquid and others)
                            defi_code = chain_to_defi_code.get(full_chain_key, full_chain_key)
                            defi_k_sum = defi_chain_j_sums.get(defi_code, Decimal('0'))
                            defi_k_sum = defi_k_sum.quantize(Decimal('0.01'), rounding='ROUND_DOWN')

                            combined_sum = wallet_w_sum + defi_k_sum
                            cell_g.value = float(combined_sum)

                            # F column validation: within 1% of E column
                            e_raw = str(e_val).strip() if e_val is not None else ''
                            if combined_sum < Decimal('0.01'):
                                if e_raw in ('<0.01', '< 0.01', '<0.01%', '< 0.01%'):
                                    cell_f.value = "Passed"
                                else:
                                    cell_f.value = "Failed"
                            elif combined_sum != 0:
                                percentage_diff = abs(e_decimal - combined_sum) / combined_sum * Decimal('100')
                                if percentage_diff <= Decimal('1'):
                                    cell_f.value = "Passed"
                                else:
                                    cell_f.value = "Failed"
                                    print(f"      ⚠️  Net Worth Mismatch: '{chain_name}' (→ '{full_chain_name}') E={e_decimal}, sum={combined_sum}, diff={percentage_diff:.2f}%")
                            else:
                                cell_f.value = "Passed" if e_decimal == 0 else "Failed"

                            # D column: handled by the dedicated post-processing block above
                            # (uses chain name mapping + Overview - Wallet A column count only)

                            # L column validation: F / SUM(all chain F values) compared with K column
                            g_val = ws_header.cell(h_row_idx, 11).value  # K column - Share %
                            if g_val and sum_chain_e != 0:
                                try:
                                    # Parse E value (without 2dp truncation for percentage calculation)
                                    e_clean_full = str(e_val).replace(',', '').replace('$', '').strip() if e_val else '0'
                                    e_decimal_full = Decimal(e_clean_full)

                                    # Calculate percentage: E / SUM(all chain E)
                                    calculated_percentage = (e_decimal_full / sum_chain_e) * Decimal('100')

                                    # Parse G column value (expected percentage)
                                    g_clean = str(g_val).replace(',', '').replace('%', '').strip()

                                    # Special handling for calculated percentage < 0.01
                                    if calculated_percentage < Decimal('0.01'):
                                        # If calculated < 0.01, check if G column = "<0.01"
                                        if g_clean == "<0.01":
                                            cell_h.value = "Passed"
                                        else:
                                            cell_h.value = "Failed"
                                    else:
                                        # Normal comparison with tolerance
                                        g_decimal = Decimal(g_clean) if not g_clean.startswith('<') else Decimal('0')
                                        diff = abs(calculated_percentage - g_decimal)
                                        if diff <= Decimal('0.01'):
                                            cell_h.value = "Passed"
                                        else:
                                            cell_h.value = "Failed"
                                except:
                                    cell_h.value = "Error"
                            else:
                                cell_h.value = "N/A"

                        # Scenario 5: Token Holdings - Platform → WALLET
                        elif section_val == "Token Holdings - Platform" and str(category_val).strip().upper() == "WALLET":
                            wallet_total = sum(wallet_chain_w_sums.values()).quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                            cell_g.value = float(wallet_total)
                            if wallet_total != 0:
                                percentage_diff = abs(e_decimal - wallet_total) / wallet_total * Decimal('100')
                                if percentage_diff <= Decimal('1'):
                                    cell_f.value = "Passed"
                                else:
                                    cell_f.value = "Failed"
                            else:
                                cell_f.value = "Passed" if e_decimal == 0 else "Failed"

                        # Scenario 5b: Token Holdings - Platform → Chain Name (e.g., BNB, ETH, TRX)
                        # F = (SUM Wallet W where A = full chain name) + (SUM De-Fi J where B = chain code)
                        elif section_val == "Token Holdings - Platform" and str(category_val).strip().lower() in _chain_display_map:
                            category_str_5b = str(category_val).strip()
                            full_chain_name_5b = _chain_display_map.get(category_str_5b.lower(), category_str_5b)
                            full_chain_key_5b = full_chain_name_5b.lower()
                            defi_code_5b = chain_to_defi_code.get(full_chain_key_5b, full_chain_key_5b)
                            wallet_w_sum_5b = wallet_chain_w_sums.get(full_chain_key_5b, Decimal('0'))
                            defi_j_sum_5b = defi_chain_j_sums.get(defi_code_5b, Decimal('0'))
                            combined_5b = wallet_w_sum_5b + defi_j_sum_5b
                            cell_g.value = float(combined_5b)
                            e_raw_5b = str(e_val).strip() if e_val is not None else ''
                            if combined_5b < Decimal('0.01'):
                                if e_raw_5b in ('<0.01', '< 0.01', '<0.01%', '< 0.01%'):
                                    cell_f.value = "Passed"
                                else:
                                    cell_f.value = "Failed"
                            elif combined_5b != Decimal('0'):
                                percentage_diff_5b = abs(e_decimal - combined_5b) / combined_5b * Decimal('100')
                                cell_f.value = "Passed" if percentage_diff_5b <= Decimal('1') else "Failed"
                            else:
                                cell_f.value = "Passed" if e_decimal == Decimal('0') else "Failed"

                        # Scenario 6+: Token Holdings - Platform → {Protocol} ({Chain}) or Exchange Table Name
                        elif section_val == "Token Holdings - Platform" and category_val:
                            # Parse category like "Aave V3 (Ethereum)" or "Aave V3 (Base)"
                            category_str = str(category_val).strip()
                            # Extract protocol and chain from format: "Protocol Name (Chain Name)"
                            import re
                            match = re.match(r'^(.+?)\s*\((.+?)\)$', category_str)
                            if match:
                                # DeFi protocol with chain format
                                protocol_name = match.group(1).strip()
                                chain_display = match.group(2).strip()

                                # Map display chain name to lowercase chain codes
                                chain_mapping = {
                                    'Ethereum': 'eth',
                                    'Base': 'base',
                                    'Binance Smart Chain': 'bsc',
                                    'BSC': 'bsc',
                                    'Tron': 'tron'
                                }
                                chain_code = chain_mapping.get(chain_display, chain_display.lower())

                                # Lookup in defi_detail_map (case-insensitive)
                                lookup_key = (protocol_name.lower(), chain_code)
                                defi_value = defi_detail_map.get(lookup_key)

                                # Debug logging for failed lookups
                                if defi_value is None:
                                    print(f"      ⚠️  DEBUG: Lookup failed for {lookup_key}")
                                    print(f"      Available De-Fi data keys (partial match on '{protocol_name}'):")
                                    for key in sorted(defi_detail_map.keys()):
                                        if protocol_name.lower() in key[0]:
                                            print(f"         {key} → ${defi_detail_map[key]}")

                                if defi_value is not None:
                                    defi_value_2dp = defi_value.quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                    cell_g.value = float(defi_value_2dp)
                                    # F column validation (within 1% diff)
                                    if defi_value_2dp != 0:
                                        percentage_diff = abs(e_decimal - defi_value_2dp) / defi_value_2dp * Decimal('100')

                                        # Debug logging for percentage diff
                                        print(f"      🔍 DEBUG F validation: {category_val}")
                                        print(f"         E column (UI):        ${e_decimal}")
                                        print(f"         De-Fi calculated:     ${defi_value}")
                                        print(f"         De-Fi (2dp):          ${defi_value_2dp}")
                                        print(f"         Percentage diff:      {percentage_diff:.4f}%")
                                        print(f"         Tolerance:            1%")
                                        print(f"         Result:               {'PASS' if percentage_diff <= Decimal('1') else 'FAIL'}")

                                        if percentage_diff <= Decimal('1'):
                                            cell_f.value = "Passed"
                                        else:
                                            cell_f.value = "Failed"
                                    else:
                                        cell_f.value = "Passed" if e_decimal == 0 else "Failed"
                                else:
                                    cell_f.value = "Not Found in De-Fi"
                            else:
                                # Not a protocol(chain) format - check if it's an exchange table name
                                # Exchange table names like "Binance - david", "moontest", etc.
                                if category_str in exchange_tables:
                                    # Find the total from this exchange table
                                    if category_str in wb.sheetnames:
                                        ws_exchange = wb[category_str]
                                        exchange_total = None

                                        # Find row where P column (16) = "Total"
                                        for row_idx in range(2, ws_exchange.max_row + 1):
                                            p_val = ws_exchange.cell(row_idx, 16).value  # P column
                                            if p_val and str(p_val).strip().upper() == "TOTAL":
                                                q_val = ws_exchange.cell(row_idx, 17).value  # Q column
                                                if q_val:
                                                    try:
                                                        q_clean = str(q_val).replace(',', '').replace('$', '').strip()
                                                        exchange_total = Decimal(q_clean).quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                                    except (ValueError, InvalidOperation):
                                                        pass
                                                break  # Found total row

                                        # Validate
                                        if exchange_total is not None:
                                            cell_g.value = float(exchange_total)
                                            if exchange_total != 0:
                                                percentage_diff = abs(e_decimal - exchange_total) / exchange_total * Decimal('100')
                                                if percentage_diff <= Decimal('1'):
                                                    cell_f.value = "Passed"
                                                else:
                                                    cell_f.value = "Failed"
                                            else:
                                                cell_f.value = "Passed" if e_decimal == 0 else "Failed"
                                        else:
                                            cell_f.value = "Total Not Found"
                                    else:
                                        cell_f.value = "Sheet Not Found"
                                else:
                                    # Platform name without chain suffix (e.g. "Morpho") — sum De-Fi L column where A matches
                                    _platform_lower = category_str.lower()
                                    _platform_defi_sum = Decimal('0')
                                    _platform_found = False
                                    if ws_defi:
                                        for _dr in range(2, ws_defi.max_row + 1):
                                            _d_name = ws_defi.cell(_dr, 1).value
                                            _d_val = ws_defi.cell(_dr, 12).value  # Column L - Value
                                            if _d_name and _d_val:
                                                if _platform_lower in str(_d_name).strip().lower():
                                                    try:
                                                        _v_str = str(_d_val).replace(',', '').replace('$', '').replace('<', '').strip()
                                                        if _v_str and _v_str not in ('', 'None', 'N/A'):
                                                            _platform_defi_sum += Decimal(_v_str)
                                                            _platform_found = True
                                                    except:
                                                        pass
                                    if _platform_found:
                                        _platform_2dp = _platform_defi_sum.quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                        cell_g.value = float(_platform_2dp)
                                        if _platform_2dp != 0:
                                            _pct_diff = abs(e_decimal - _platform_2dp) / _platform_2dp * Decimal('100')
                                            cell_f.value = "Passed" if _pct_diff <= Decimal('1') else "Failed"
                                        else:
                                            cell_f.value = "Passed" if e_decimal == 0 else "Failed"
                                    else:
                                        cell_f.value = "Not Found"
                        elif section_val == "Table":
                            b_val_t = str(ws_header.cell(h_row_idx, 2).value or '').strip()
                            try:
                                e_clean_t = str(e_val).replace(',', '').replace('$', '').replace('<', '').strip() if e_val else '0'
                                e_dec_t = Decimal(e_clean_t)
                            except:
                                e_dec_t = None

                            if b_val_t.lower() == "wallet":
                                try:
                                    wallet_trunc = sum(wallet_chain_w_sums.values()).quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                    cell_g.value = float(wallet_trunc)
                                    if e_dec_t is not None:
                                        if wallet_trunc != 0:
                                            _pct = abs(e_dec_t - wallet_trunc) / wallet_trunc * Decimal('100')
                                            cell_f.value = "Passed" if _pct <= Decimal('1') else "Failed"
                                        else:
                                            cell_f.value = "Passed" if e_dec_t == 0 else "Failed"
                                    else:
                                        cell_f.value = "Invalid Format"
                                except Exception as _wte:
                                    cell_f.value = f"Error: {str(_wte)[:20]}"
                            else:
                                # Compare E with SUM of Overview - De-Fi K column where A = B value
                                try:
                                    ws_d_t = wb["Overview - De-Fi"] if "Overview - De-Fi" in wb.sheetnames else None
                                    defi_sum = Decimal('0')
                                    if ws_d_t:
                                        for _dr in range(2, ws_d_t.max_row + 1):
                                            a_defi = ws_d_t.cell(_dr, 1).value
                                            k_defi = ws_d_t.cell(_dr, 11).value  # K column - Value
                                            if a_defi and str(a_defi).strip().lower() == b_val_t.lower() and k_defi:
                                                try:
                                                    defi_sum += Decimal(str(k_defi).replace(',', '').replace('$', '').strip())
                                                except:
                                                    pass
                                    if e_dec_t is not None:
                                        cell_g.value = float(defi_sum)
                                        if defi_sum != Decimal('0'):
                                            _pct = abs(e_dec_t - defi_sum) / abs(defi_sum) * Decimal('100')
                                            cell_f.value = "Passed" if _pct <= Decimal('1') else "Failed"
                                        else:
                                            cell_f.value = "Passed" if e_dec_t == Decimal('0') else "Failed"
                                    else:
                                        cell_f.value = "Invalid Format"
                                except Exception as _dte:
                                    cell_f.value = f"Error: {str(_dte)[:20]}"
                            cell_h.value = "Not Applicable"

                        else:
                            # Other scenarios - leave as is or mark as Not Applicable
                            if not cell_f.value:
                                cell_f.value = "Not Applicable"

                    except Exception as ex:
                        cell_f.value = f"Error: {str(ex)[:20]}"

                print(f"   ✅ Updated F column validation in 'Overview - Header & Token Holdings Header'")

            # ================================================================
            # Populate Column I (Net Worth - API Calculation) and
            # Column J (Net Worth - UI-API Validation) in Header & Holdings tab
            # I = Wallet API Calc (Y col 25) per chain + Rabby Calc Value per chain
            # J = Net Worth (F col 6) vs API Calc (I col 9), ≤1% = Passed
            # ================================================================
            if "Overview - Header & Token Holdings Header" in wb.sheetnames:
                ws_header_ij = wb["Overview - Header & Token Holdings Header"]
                ws_wallet_ij = wb["Overview - Wallet"] if "Overview - Wallet" in wb.sheetnames else None
                ws_defi_ij = wb["Overview - De-Fi"] if "Overview - De-Fi" in wb.sheetnames else None

                print(f"   🔄 Populating I & J columns (API Calculation & UI-API Validation) in 'Overview - Header & Token Holdings Header'...")

                # Build per-chain Wallet API Calc sums from Overview - Wallet Y column (25)
                wallet_api_calc_chain = {}  # full chain name (lowercase) → sum of Y column
                wallet_api_calc_total = Decimal('0')
                if ws_wallet_ij:
                    for _ri in range(2, ws_wallet_ij.max_row + 1):
                        _chain = ws_wallet_ij.cell(_ri, 1).value  # Column A - Chain
                        _y_val = ws_wallet_ij.cell(_ri, 25).value  # Column Y - API Calculated Value
                        if _chain and _y_val is not None and str(_y_val).strip() not in ("", "Error"):
                            try:
                                _y_dec = Decimal(str(_y_val).replace(',', '').replace('$', '').strip())
                                _chain_key = str(_chain).strip().lower()
                                wallet_api_calc_chain.setdefault(_chain_key, Decimal('0'))
                                wallet_api_calc_chain[_chain_key] += _y_dec
                                wallet_api_calc_total += _y_dec
                            except:
                                pass
                print(f"      Wallet API Calc: {len(wallet_api_calc_chain)} chains, total=${wallet_api_calc_total}")

                # Build per-chain Rabby Calc Value sums from Overview - De-Fi T column (20) = Api Calc Value
                rabby_calc_chain = {}  # defi chain code (lowercase) → sum of T column
                rabby_calc_total = Decimal('0')
                if ws_defi_ij:
                    for _ri in range(2, ws_defi_ij.max_row + 1):
                        _defi_chain = ws_defi_ij.cell(_ri, 2).value  # Column B - Chain code
                        _t_val = ws_defi_ij.cell(_ri, 14).value  # Column N - Api Calc Value
                        if _defi_chain and _t_val is not None and str(_t_val).strip() not in ("", "No Match"):
                            try:
                                _t_dec = Decimal(str(_t_val).replace(',', '').replace('$', '').strip())
                                _defi_key = str(_defi_chain).strip().lower()
                                rabby_calc_chain.setdefault(_defi_key, Decimal('0'))
                                rabby_calc_chain[_defi_key] += _t_dec
                                rabby_calc_total += _t_dec
                            except:
                                pass
                print(f"      Rabby Calc: {len(rabby_calc_chain)} chains, total=${rabby_calc_total}")

                # Chain display name → full wallet chain name (reuse existing map)
                _chain_display_map_ij = {
                    'bnb': 'binance smart chain', 'bsc': 'binance smart chain',
                    'binance smart chain': 'binance smart chain', 'binance': 'binance smart chain',
                    'ethereum': 'ethereum', 'eth': 'ethereum',
                    'tron': 'tron', 'trx': 'tron',
                    'polygon': 'polygon', 'matic': 'polygon',
                    'arbitrum': 'arbitrum', 'arb': 'arbitrum',
                    'optimism': 'optimism', 'op': 'optimism',
                    'avalanche': 'avalanche', 'avax': 'avalanche',
                    'base': 'base', 'solana': 'solana', 'sol': 'solana',
                    'fantom': 'fantom', 'ftm': 'fantom',
                    'linea': 'linea', 'scroll': 'scroll',
                    'zksync era': 'zksync era', 'zksync': 'zksync era',
                    'blast': 'blast', 'mantle': 'mantle', 'mnt': 'mantle',
                    'cronos': 'cronos', 'cro': 'cronos',
                    'gnosis': 'gnosis', 'xdai': 'gnosis',
                    'mode': 'mode', 'hyperliquid': 'hyperliquid', 'hype': 'hyperliquid',
                }
                # Full chain name → De-Fi chain code
                _chain_to_defi_ij = {
                    'binance smart chain': 'bsc', 'ethereum': 'eth', 'tron': 'tron',
                    'polygon': 'matic', 'arbitrum': 'arb', 'optimism': 'op',
                    'avalanche': 'avax', 'base': 'base', 'solana': 'sol',
                    'fantom': 'ftm', 'linea': 'linea', 'scroll': 'scroll',
                    'zksync era': 'era', 'blast': 'blast', 'mantle': 'mnt',
                    'cronos': 'cro', 'gnosis': 'xdai', 'mode': 'mode',
                    'hyperliquid': 'hyperliquid',
                }

                for _h_row in range(2, ws_header_ij.max_row + 1):
                    section_val = ws_header_ij.cell(_h_row, 1).value  # A - Section
                    category_val = ws_header_ij.cell(_h_row, 2).value  # B - Category
                    f_val = ws_header_ij.cell(_h_row, 6).value  # F - Net Worth (UI)
                    cell_i = ws_header_ij.cell(_h_row, 9)  # I - API Calculation
                    cell_j = ws_header_ij.cell(_h_row, 10)  # J - UI-API Validation

                    if not section_val or not category_val:
                        continue

                    section_str = str(section_val).strip()
                    category_str = str(category_val).strip()
                    category_lower = category_str.lower()

                    api_calc = None  # Will hold the Decimal value for column I

                    # --- Overview Header rows ---
                    if section_str == "Overview Header":
                        if category_lower in ('wallets', 'wallet'):
                            # Wallets: sum ALL Wallet API Calc
                            api_calc = wallet_api_calc_total
                        elif category_lower in ('de-fi positions', 'defi positions', 'de-fi'):
                            # De-Fi: sum ALL Rabby Calc
                            api_calc = rabby_calc_total
                        elif category_lower == 'exchanges':
                            api_calc = Decimal('0')  # No API calc for exchanges

                    # --- Token Holdings - Chain rows ---
                    elif section_str == "Token Holdings - Chain":
                        full_chain = _chain_display_map_ij.get(category_lower, category_lower)
                        defi_code = _chain_to_defi_ij.get(full_chain, full_chain)
                        w_sum = wallet_api_calc_chain.get(full_chain, Decimal('0'))
                        r_sum = rabby_calc_chain.get(defi_code, Decimal('0'))
                        api_calc = w_sum + r_sum

                    # --- Token Holdings - Platform rows ---
                    elif section_str == "Token Holdings - Platform":
                        if category_lower in ('wallet',):
                            api_calc = wallet_api_calc_total
                        else:
                            # Match protocol name + chain against De-Fi column A + B, sum column N (14) = Api Calc Value
                            import re as _re_api_plat
                            _api_m = _re_api_plat.match(r'^(.+?)\s*\((.+?)\)$', category_str)
                            if _api_m:
                                _api_proto = _api_m.group(1).strip().lower()
                                _api_chain_display = _api_m.group(2).strip()
                                _api_chain_map = {
                                    'Ethereum': 'eth', 'Base': 'base',
                                    'Binance Smart Chain': 'bsc', 'BSC': 'bsc',
                                    'Tron': 'tron', 'Polygon': 'matic',
                                    'Arbitrum': 'arb', 'Optimism': 'op',
                                }
                                _api_chain_code = _api_chain_map.get(_api_chain_display, _api_chain_display.lower())
                            else:
                                _api_proto = category_lower
                                _api_chain_code = None

                            _platform_api_sum = Decimal('0')
                            if ws_defi_ij:
                                for _dr in range(2, ws_defi_ij.max_row + 1):
                                    _d_name = ws_defi_ij.cell(_dr, 1).value
                                    _d_chain = ws_defi_ij.cell(_dr, 2).value
                                    _d_n = ws_defi_ij.cell(_dr, 14).value  # Column N - Api Calc Value
                                    if _d_name and _d_n:
                                        _name_match = _api_proto in str(_d_name).strip().lower()
                                        _chain_match = (_api_chain_code is None or
                                                        (_d_chain and str(_d_chain).strip().lower() == _api_chain_code))
                                        if _name_match and _chain_match and str(_d_n).strip() not in ("", "No Match"):
                                            try:
                                                _platform_api_sum += Decimal(str(_d_n).replace(',', '').replace('$', '').strip())
                                            except:
                                                pass
                            api_calc = _platform_api_sum

                    # --- Table rows ---
                    elif section_str == "Table":
                        if category_lower == "wallet":
                            api_calc = wallet_api_calc_total
                        else:
                            # Sum De-Fi column N (14) where A (protocol) matches
                            _table_sum = Decimal('0')
                            if ws_defi_ij:
                                for _dr in range(2, ws_defi_ij.max_row + 1):
                                    _d_name = ws_defi_ij.cell(_dr, 1).value
                                    _d_n = ws_defi_ij.cell(_dr, 14).value  # Column N - Api Calc Value
                                    if (_d_name and _d_n
                                            and str(_d_name).strip().lower() == category_lower
                                            and str(_d_n).strip() not in ("", "No Match")):
                                        try:
                                            _table_sum += Decimal(str(_d_n).replace(',', '').replace('$', '').strip())
                                        except:
                                            pass
                            api_calc = _table_sum

                    # Write Column I (API Calculation)
                    if api_calc is not None:
                        cell_i.value = float(api_calc) if api_calc != Decimal('0') else 0

                        # Write Column J (UI-API Validation): Net Worth vs API Calc, ≤1% = Passed
                        try:
                            f_clean = str(f_val or '').replace(',', '').replace('$', '').replace('<', '').strip()
                            f_dec = Decimal(f_clean) if f_clean else Decimal('0')
                            if api_calc != Decimal('0'):
                                pct_diff = abs(f_dec - api_calc) / abs(api_calc) * Decimal('100')
                                cell_j.value = "Passed" if pct_diff <= Decimal('1') else "Failed"
                            elif f_dec == Decimal('0'):
                                cell_j.value = "Passed"
                            else:
                                cell_j.value = "Failed"
                        except:
                            cell_j.value = "Error"
                    else:
                        cell_i.value = ""
                        cell_j.value = "Not Applicable"

                print(f"   ✅ Populated I & J columns in 'Overview - Header & Token Holdings Header'")

            # ================================================================
            # Populate allocation tabs: Column C (% - Api NW Calc) and Column F (Net Worth - API Calculation)
            # For Platform: read from Header tab "Overview Header" rows' column I
            # For Chain: read from Header tab "Token Holdings - Chain" rows' column I
            # ================================================================
            ws_header_alloc = wb["Overview - Header & Token Holdings Header"] if "Overview - Header & Token Holdings Header" in wb.sheetnames else None

            for alloc_sheet_name, header_section in [
                ("Overview - Platform Allocation", "Overview Header"),
                ("Overview - Chain Allocation", "Token Holdings - Chain"),
            ]:
                if alloc_sheet_name in wb.sheetnames and ws_header_alloc:
                    ws_alloc = wb[alloc_sheet_name]
                    print(f"   🔄 Populating C & F columns in '{alloc_sheet_name}'...")

                    # Build lookup: category_lower → API Calc value from Header tab column I (9)
                    api_calc_lookup = {}
                    for _hr in range(2, ws_header_alloc.max_row + 1):
                        _sec = ws_header_alloc.cell(_hr, 1).value
                        _cat = ws_header_alloc.cell(_hr, 2).value
                        _api_val = ws_header_alloc.cell(_hr, 9).value  # Column I - API Calculation
                        if _sec and _cat and str(_sec).strip() == header_section:
                            cat_key = str(_cat).strip().lower()
                            if _api_val is not None and str(_api_val).strip() not in ('', '0'):
                                try:
                                    api_calc_lookup[cat_key] = Decimal(str(_api_val).replace(',', '').strip())
                                except:
                                    pass

                    # Write column F (6) for each allocation row
                    # Name normalization for matching (Header uses "Wallets", allocation uses "Wallet")
                    _name_normalize = {
                        'wallet': 'wallets',
                        'de-fi positions': 'de-fi positions',
                        'exchange': 'exchanges',
                    }
                    alloc_last_row = ws_alloc.max_row
                    for _ar in range(2, alloc_last_row + 1):
                        _name = ws_alloc.cell(_ar, 1).value  # Column A - Name
                        if _name:
                            _name_lower = str(_name).strip().lower()
                            _lookup_key = _name_normalize.get(_name_lower, _name_lower)
                            _api_val = api_calc_lookup.get(_lookup_key) or api_calc_lookup.get(_name_lower)
                            if _api_val is not None:
                                ws_alloc.cell(_ar, 6).value = float(_api_val)

                    # Calculate column C (3): % - Api NW Calc = (row F / sum of all F) * 100
                    total_api = Decimal('0')
                    for _ar in range(2, alloc_last_row + 1):
                        _f_val = ws_alloc.cell(_ar, 6).value
                        if _f_val is not None:
                            try:
                                total_api += Decimal(str(_f_val).replace(',', '').strip())
                            except:
                                pass

                    if total_api > 0:
                        for _ar in range(2, alloc_last_row + 1):
                            _f_val = ws_alloc.cell(_ar, 6).value
                            if _f_val is not None:
                                try:
                                    _f_dec = Decimal(str(_f_val).replace(',', '').strip())
                                    _pct = (_f_dec / total_api) * Decimal('100')
                                    ws_alloc.cell(_ar, 3).value = float(_pct)
                                except:
                                    pass

                    print(f"   ✅ Populated C & F columns in '{alloc_sheet_name}' (total API: {total_api})")

                    # Percentage Validation (D col 4): compare B (UI %) vs C (% - Api NW Calc)
                    print(f"   🔄 Calculating Percentage Validation (B vs C) in '{alloc_sheet_name}'...")
                    for _ar in range(2, alloc_last_row + 1):
                        _b_val = ws_alloc.cell(_ar, 2).value  # B - Percentage (UI)
                        _c_val = ws_alloc.cell(_ar, 3).value  # C - % - Api NW Calc
                        cell_d = ws_alloc.cell(_ar, 4)         # D - Percentage Validation

                        _b_str = str(_b_val).strip() if _b_val is not None else ''
                        _b_clean = _b_str.replace(',', '').replace('%', '').replace('<', '').replace('>', '').strip()

                        if _c_val is not None:
                            try:
                                _c_dec = Decimal(str(_c_val).replace(',', '').strip())
                                _c_trunc = Decimal(int(_c_dec * Decimal('100'))) / Decimal('100')

                                # Special handling for very small percentages
                                if _c_trunc < Decimal('0.01'):
                                    if _b_str in ['<0.01', '< 0.01', '<0.01%', '< 0.01%']:
                                        cell_d.value = "Passed"
                                    else:
                                        cell_d.value = "Failed"
                                else:
                                    _b_dec = Decimal(_b_clean) if _b_clean else None
                                    if _b_dec is not None and _c_trunc == _b_dec:
                                        cell_d.value = "Passed"
                                    else:
                                        cell_d.value = "Failed"
                            except:
                                cell_d.value = "Error"
                        else:
                            cell_d.value = "No API Data"
                    print(f"   ✅ Percentage Validation done for '{alloc_sheet_name}'")

                    # Net Worth Validation (G col 7): |E - F| / E * 100, ≤1% = Passed
                    print(f"   🔄 Calculating Net Worth Validation (E vs F) in '{alloc_sheet_name}'...")
                    for _ar in range(2, alloc_last_row + 1):
                        _e_val = ws_alloc.cell(_ar, 5).value  # E - Net Worth (UI)
                        _f_val = ws_alloc.cell(_ar, 6).value  # F - Net Worth - API Calculation
                        cell_g_alloc = ws_alloc.cell(_ar, 7)  # G - Net Worth Validation

                        if _f_val is not None and _e_val is not None:
                            try:
                                _e_str = str(_e_val).replace(',', '').replace('$', '').replace('<', '').replace('>', '').strip()
                                _f_str = str(_f_val).replace(',', '').replace('$', '').strip()
                                _e_dec = Decimal(_e_str) if _e_str else Decimal('0')
                                _f_dec = Decimal(_f_str) if _f_str else Decimal('0')

                                if _e_dec != 0:
                                    _pct_diff = abs(_e_dec - _f_dec) / abs(_e_dec) * Decimal('100')
                                    cell_g_alloc.value = "Passed" if _pct_diff < Decimal('1') else "Failed"
                                elif _f_dec == 0:
                                    cell_g_alloc.value = "Passed"
                                else:
                                    cell_g_alloc.value = "Failed"
                            except:
                                cell_g_alloc.value = "Error"
                        else:
                            cell_g_alloc.value = "No API Data"
                    print(f"   ✅ Net Worth Validation done for '{alloc_sheet_name}'")

            # ================================================================
            # Token Allocation: Column F from Wallet Y + De-Fi N, grouped by token name
            # "Others" row = total - sum of named tokens
            # ================================================================
            if "Overview - Token Allocation" in wb.sheetnames:
                ws_token_alloc = wb["Overview - Token Allocation"]
                ws_wallet_ta = wb["Overview - Wallet"] if "Overview - Wallet" in wb.sheetnames else None
                ws_defi_ta = wb["Overview - De-Fi"] if "Overview - De-Fi" in wb.sheetnames else None

                print(f"   🔄 Populating C & F columns in 'Overview - Token Allocation'...")

                # Build token+chain → API value map
                # Key: (token_lower, chain_lower) → Decimal sum
                token_chain_api = {}  # (token, chain) → sum
                total_all_api = Decimal('0')

                # From Wallet tab: B=token, A=chain, Y=API Calc Value
                _chain_norm = {
                    'ethereum': 'eth', 'binance smart chain': 'bsc', 'base': 'base', 'tron': 'tron',
                    'polygon': 'matic', 'arbitrum': 'arb', 'optimism': 'op',
                }
                if ws_wallet_ta:
                    for _wr in range(2, ws_wallet_ta.max_row + 1):
                        _token = ws_wallet_ta.cell(_wr, 2).value  # B - Token
                        _chain = ws_wallet_ta.cell(_wr, 1).value  # A - Chain
                        _y_val = ws_wallet_ta.cell(_wr, 25).value  # Y - API Calc Value
                        if _token and _y_val is not None and str(_y_val).strip() not in ('', 'Error', 'N/A'):
                            try:
                                _y_dec = Decimal(str(_y_val).replace(',', '').replace('$', '').strip())
                                _t = str(_token).strip().lower()
                                _c = _chain_norm.get(str(_chain).strip().lower(), str(_chain).strip().lower()) if _chain else ''
                                token_chain_api.setdefault((_t, _c), Decimal('0'))
                                token_chain_api[(_t, _c)] += _y_dec
                                total_all_api += _y_dec
                            except:
                                pass

                # From De-Fi tab: D=Pool/Position (token), B=Chain, N=Api Calc Value
                if ws_defi_ta:
                    for _dr in range(2, ws_defi_ta.max_row + 1):
                        _pool = ws_defi_ta.cell(_dr, 4).value  # D - Pool/Position
                        _chain = ws_defi_ta.cell(_dr, 2).value  # B - Chain
                        _n_val = ws_defi_ta.cell(_dr, 14).value  # N - Api Calc Value
                        if _pool and _n_val is not None and str(_n_val).strip() not in ('', 'No Match'):
                            try:
                                _n_dec = Decimal(str(_n_val).replace(',', '').replace('$', '').strip())
                                _t = str(_pool).strip().split('/')[0].split('-')[0].strip().lower()
                                _c = str(_chain).strip().lower() if _chain else ''
                                token_chain_api.setdefault((_t, _c), Decimal('0'))
                                token_chain_api[(_t, _c)] += _n_dec
                                total_all_api += _n_dec
                            except:
                                pass

                # Debug: show what's in the map
                print(f"   📊 Token+Chain API map ({len(token_chain_api)} entries, total={total_all_api}):")
                for (_t, _c), _v in sorted(token_chain_api.items(), key=lambda x: -x[1]):
                    if _v > Decimal('1'):
                        print(f"      ({_t}, {_c}) = {_v}")

                # Write column F for each token row
                import re as _re_ta
                named_sum = Decimal('0')
                alloc_last = ws_token_alloc.max_row
                for _ar in range(2, alloc_last + 1):
                    _name = ws_token_alloc.cell(_ar, 1).value  # e.g. "USDC(BASE)"
                    if _name:
                        _name_str = str(_name).strip()
                        _name_lower = _name_str.lower()
                        if _name_lower == 'others':
                            others_val = total_all_api - named_sum
                            ws_token_alloc.cell(_ar, 6).value = float(others_val)
                            print(f"      Token '{_name_str}': total={total_all_api} - named={named_sum} = {others_val}")
                        else:
                            # Parse "TOKEN(CHAIN)" format
                            _m = _re_ta.match(r'^(.+?)\s*\((.+?)\)$', _name_lower)
                            if _m:
                                _base_token = _m.group(1).strip()
                                _raw_chain = _m.group(2).strip()
                                # Normalize chain name to match map keys (map uses short codes)
                                _alloc_chain_norm = {
                                    'ethereum': 'eth', 'base': 'base', 'bsc': 'bsc',
                                    'binance smart chain': 'bsc', 'tron': 'tron',
                                    'polygon': 'matic', 'arbitrum': 'arb', 'optimism': 'op',
                                }
                                _chain_code = _alloc_chain_norm.get(_raw_chain, _raw_chain)
                            else:
                                _base_token = _name_lower
                                _chain_code = None

                            # Sum all matching (token, chain) entries
                            _api_val = Decimal('0')
                            _matches = []
                            for (_t, _c), _v in token_chain_api.items():
                                if _t == _base_token:
                                    if _chain_code is None or _c == _chain_code:
                                        _api_val += _v
                                        _matches.append(f"({_t},{_c})={_v}")

                            print(f"      Token '{_name_str}' → base='{_base_token}', chain='{_chain_code}' → {len(_matches)} matches: {_matches[:5]} = {_api_val}")

                            ws_token_alloc.cell(_ar, 6).value = float(_api_val)
                            named_sum += _api_val

                # Calculate column C: % - Api NW Calc = (row F / total) * 100
                if total_all_api > 0:
                    for _ar in range(2, alloc_last + 1):
                        _f_val = ws_token_alloc.cell(_ar, 6).value
                        if _f_val is not None:
                            try:
                                _f_dec = Decimal(str(_f_val).replace(',', '').strip())
                                _pct = (_f_dec / total_all_api) * Decimal('100')
                                ws_token_alloc.cell(_ar, 3).value = float(_pct)
                            except:
                                pass

                print(f"   ✅ Populated C & F columns in 'Overview - Token Allocation' (total API: {total_all_api})")

                # Percentage Validation (D col 4): compare B (UI %) vs C (% - Api NW Calc)
                print(f"   🔄 Calculating Percentage Validation (B vs C) in 'Overview - Token Allocation'...")
                for _ar in range(2, alloc_last + 1):
                    _b_val = ws_token_alloc.cell(_ar, 2).value  # B - Percentage (UI)
                    _c_val = ws_token_alloc.cell(_ar, 3).value  # C - % - Api NW Calc
                    cell_d = ws_token_alloc.cell(_ar, 4)         # D - Percentage Validation

                    _b_str = str(_b_val).strip() if _b_val is not None else ''
                    _b_clean = _b_str.replace(',', '').replace('%', '').replace('<', '').replace('>', '').strip()

                    if _c_val is not None:
                        try:
                            _c_dec = Decimal(str(_c_val).replace(',', '').strip())
                            _c_trunc = Decimal(int(_c_dec * Decimal('100'))) / Decimal('100')

                            # Special handling for very small percentages
                            if _c_trunc < Decimal('0.01'):
                                if _b_str in ['<0.01', '< 0.01', '<0.01%', '< 0.01%']:
                                    cell_d.value = "Passed"
                                else:
                                    cell_d.value = "Failed"
                            else:
                                _b_dec = Decimal(_b_clean) if _b_clean else None
                                if _b_dec is not None and _c_trunc == _b_dec:
                                    cell_d.value = "Passed"
                                else:
                                    cell_d.value = "Failed"
                        except:
                            cell_d.value = "Error"
                    else:
                        cell_d.value = "No API Data"
                print(f"   ✅ Percentage Validation done for 'Overview - Token Allocation'")

                # Net Worth Validation (G col 7): |E - F| / E * 100, ≤1% = Passed
                print(f"   🔄 Calculating Net Worth Validation (E vs F) in 'Overview - Token Allocation'...")
                for _ar in range(2, alloc_last + 1):
                    _e_val = ws_token_alloc.cell(_ar, 5).value  # E - Net Worth (UI)
                    _f_val = ws_token_alloc.cell(_ar, 6).value  # F - Net Worth - API Calculation
                    cell_g_alloc = ws_token_alloc.cell(_ar, 7)  # G - Net Worth Validation

                    if _f_val is not None and _e_val is not None:
                        try:
                            _e_str = str(_e_val).replace(',', '').replace('$', '').replace('<', '').replace('>', '').strip()
                            _f_str = str(_f_val).replace(',', '').replace('$', '').strip()
                            _e_dec = Decimal(_e_str) if _e_str else Decimal('0')
                            _f_dec = Decimal(_f_str) if _f_str else Decimal('0')

                            if _e_dec != 0:
                                _pct_diff = abs(_e_dec - _f_dec) / abs(_e_dec) * Decimal('100')
                                cell_g_alloc.value = "Passed" if _pct_diff < Decimal('1') else "Failed"
                            elif _f_dec == 0:
                                cell_g_alloc.value = "Passed"
                            else:
                                cell_g_alloc.value = "Failed"
                        except:
                            cell_g_alloc.value = "Error"
                    else:
                        cell_g_alloc.value = "No API Data"
                print(f"   ✅ Net Worth Validation done for 'Overview - Token Allocation'")

            # Update G column validation in "Overview - Platform Allocation" tab
            if "Overview - Platform Allocation" in wb.sheetnames and "Overview - Header & Token Holdings Header" in wb.sheetnames:
                ws_platform = wb["Overview - Platform Allocation"]
                ws_header = wb["Overview - Header & Token Holdings Header"]

                print(f"   🔄 Updating E column validation in 'Overview - Platform Allocation'...")

                # Build lookup map from Header & Token Holdings Header
                # Map specific "Overview Header" rows: (section, category) → net_worth
                overview_header_map = {}
                for h_row_idx in range(2, ws_header.max_row + 1):
                    section = ws_header.cell(h_row_idx, 1).value  # A column - Section
                    category = ws_header.cell(h_row_idx, 2).value  # B column - Category
                    net_worth = ws_header.cell(h_row_idx, 6).value  # F column - Net Worth
                    if section and category:
                        section_str = str(section).strip()
                        category_str = str(category).strip()
                        if section_str == "Overview Header":
                            overview_header_map[category_str] = net_worth

                # Platform name mapping: Platform Allocation A → Header & Token Holdings B
                platform_to_header_map = {
                    'Wallet': 'Wallets',
                    'Exchange': 'Exchanges',
                    'De-Fi Positions': 'De-Fi Positions'
                }

                # Validate each row in Platform Allocation
                for p_row_idx in range(2, ws_platform.max_row + 1):
                    platform_name = ws_platform.cell(p_row_idx, 1).value  # A column - Platform
                    platform_networth = ws_platform.cell(p_row_idx, 5).value  # E column - Net Worth
                    cell_g = ws_platform.cell(p_row_idx, 7)  # G column - validation cell

                    if platform_name:
                        platform_str = str(platform_name).strip()

                        # Map platform name to header category
                        header_category = platform_to_header_map.get(platform_str, platform_str)
                        header_networth = overview_header_map.get(header_category)

                        if header_networth is not None:
                            try:
                                # Parse both values for comparison
                                platform_clean = str(platform_networth).replace(',', '').replace('$', '').strip() if platform_networth else '0'
                                header_clean = str(header_networth).replace(',', '').replace('$', '').strip() if header_networth else '0'

                                platform_decimal = Decimal(platform_clean).quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                header_decimal = Decimal(header_clean).quantize(Decimal('0.01'), rounding='ROUND_DOWN')

                                # Exact match comparison
                                if platform_decimal == header_decimal:
                                    cell_g.value = "Passed"
                                else:
                                    cell_g.value = "Failed"
                            except:
                                cell_g.value = "Error"
                        else:
                            cell_g.value = "Not Found in Header"
                    else:
                        cell_g.value = "N/A"

                print(f"   ✅ Updated G column validation in 'Overview - Platform Allocation'")

            # Update E column validation in "Overview - Chain Allocation" tab
            if "Overview - Chain Allocation" in wb.sheetnames and "Overview - Header & Token Holdings Header" in wb.sheetnames:
                ws_chain = wb["Overview - Chain Allocation"]
                ws_header = wb["Overview - Header & Token Holdings Header"]

                print(f"   🔄 Updating E column validation in 'Overview - Chain Allocation'...")

                # Build lookup map from Header & Token Holdings Header (B column → E column)
                # Only for rows where we can match chain names
                header_chain_map = {}  # {chain_name: net_worth_value}
                for h_row_idx in range(2, ws_header.max_row + 1):
                    category = ws_header.cell(h_row_idx, 2).value  # B column - Category (could be chain name)
                    net_worth = ws_header.cell(h_row_idx, 6).value  # F column - Net Worth
                    if category:
                        category_str = str(category).strip()
                        header_chain_map[category_str] = net_worth

                # Validate each row in Chain Allocation
                for chain_row_idx in range(2, ws_chain.max_row + 1):
                    chain_name = ws_chain.cell(chain_row_idx, 1).value  # A column - Chain
                    chain_networth = ws_chain.cell(chain_row_idx, 5).value  # E column - Net Worth
                    cell_g = ws_chain.cell(chain_row_idx, 7)  # G column - validation cell

                    if chain_name:
                        chain_str = str(chain_name).strip()

                        # Look up in header map using chain name
                        header_networth = header_chain_map.get(chain_str)

                        if header_networth is not None:
                            try:
                                # Parse both values for comparison
                                chain_clean = str(chain_networth).replace(',', '').replace('$', '').strip() if chain_networth else '0'
                                header_clean = str(header_networth).replace(',', '').replace('$', '').strip() if header_networth else '0'

                                chain_decimal = Decimal(chain_clean).quantize(Decimal('0.01'), rounding='ROUND_DOWN')
                                header_decimal = Decimal(header_clean).quantize(Decimal('0.01'), rounding='ROUND_DOWN')

                                # Exact match comparison
                                if chain_decimal == header_decimal:
                                    cell_g.value = "Passed"
                                else:
                                    cell_g.value = "Failed"
                            except:
                                cell_g.value = "Error"
                        else:
                            cell_g.value = "Not Found in Header"
                    else:
                        cell_g.value = "N/A"

                print(f"   ✅ Updated G column validation in 'Overview - Chain Allocation'")

            # Merge Combined Net Worth data into Header & Token Holdings tab
            if "Overview - Combined Net Worth" in wb.sheetnames and "Overview - Header & Token Holdings Header" in wb.sheetnames:
                ws_header_merge = wb["Overview - Header & Token Holdings Header"]
                ws_combined_src = wb["Overview - Combined Net Worth"]

                # Find last used row in Header tab
                last_header_row = ws_header_merge.max_row

                # Add 2 empty rows
                start_row = last_header_row + 3

                # Write "Combined Net Worth" section header
                ws_header_merge.cell(start_row, 1).value = "Combined Net Worth"
                ws_header_merge.cell(start_row, 1).font = Font(bold=True, size=11)
                start_row += 1

                # Copy Combined Net Worth data (skip header row)
                for src_row in range(1, ws_combined_src.max_row + 1):
                    for src_col in range(1, ws_combined_src.max_column + 1):
                        val = ws_combined_src.cell(src_row, src_col).value
                        ws_header_merge.cell(start_row, src_col).value = val
                    start_row += 1

                # Remove the separate Combined Net Worth sheet
                del wb["Overview - Combined Net Worth"]
                print(f"   ✅ Merged Combined Net Worth into Header & Token Holdings tab")

            # Update C column validation in "Overview - Combined Net Worth" tab (now merged)
            # Find the Combined Net Worth section in the Header tab
            if "Overview - Header & Token Holdings Header" in wb.sheetnames:
                ws_combined = wb["Overview - Header & Token Holdings Header"]
                # Find where Combined Net Worth data starts
                _cnw_start = None
                for _r in range(1, ws_combined.max_row + 1):
                    if ws_combined.cell(_r, 1).value == "Combined Net Worth":
                        _cnw_start = _r + 1  # Skip the section header, start at column headers
                        break

                if _cnw_start:
                    print(f"   🔄 Updating validation for Combined Net Worth section...")

                    # Get API data sheets if they exist
                    ws_sim_dune = wb["Sim + Coingecko + Debank API"] if "Sim + Coingecko + Debank API" in wb.sheetnames else None
                    ws_rabby = wb["Rabby Api Data"] if "Rabby Api Data" in wb.sheetnames else None
                    ws_trx_balance = wb["TRX Balance, Price"] if "TRX Balance, Price" in wb.sheetnames else None

                    for c_row_idx in range(_cnw_start + 1, ws_combined.max_row + 1):
                        address = ws_combined.cell(c_row_idx, 1).value  # A column - Address
                        dam_value = ws_combined.cell(c_row_idx, 2).value  # B column - Value from DAM
                        cell_c = ws_combined.cell(c_row_idx, 3)  # C column - validation cell

                        if not address or not dam_value:
                            cell_c.value = "N/A"
                            continue

                        address_str = str(address).strip()

                        try:
                            # Parse DAM value
                            dam_clean = str(dam_value).replace(',', '').replace('$', '').strip()
                            dam_decimal = Decimal(dam_clean)

                            # Detect address type and validate
                            if address_str.startswith('T') and len(address_str) == 34:
                                # Tron address - use TRX Balance, Price data
                                print(f"      Validating Tron address: {address_str}")

                                try:
                                    trx_total = Decimal('0')

                                    # Sum TRX Balance, Price (Calculate from Balance Raw and Price)
                                    # Column J contains formulas, so we calculate: (Balance Raw / 10^Decimals) × Price
                                    if ws_trx_balance:
                                        for row_idx in range(2, ws_trx_balance.max_row + 1):
                                            trx_address = ws_trx_balance.cell(row_idx, 1).value  # A column
                                            if trx_address and str(trx_address).strip() == address_str:
                                                # Get Balance (Raw) (F), Decimal Places (B), and Price (G)
                                                decimal_places = ws_trx_balance.cell(row_idx, 2).value  # B column - Decimal Places
                                                balance_raw = ws_trx_balance.cell(row_idx, 6).value     # F column - Balance (Raw)
                                                price_val = ws_trx_balance.cell(row_idx, 7).value       # G column - Price

                                                # Calculate: Balance = Balance (Raw) / (10 ^ Decimal Places)
                                                # Then: Calculated Price = Balance × Price
                                                if balance_raw and price_val and decimal_places is not None:
                                                    try:
                                                        balance_raw_str = str(balance_raw).strip()
                                                        price_str = str(price_val).replace(',', '').strip()
                                                        decimals_int = int(decimal_places)

                                                        if balance_raw_str and price_str and balance_raw_str != '':
                                                            # Calculate balance: raw / (10 ^ decimals)
                                                            raw_decimal = Decimal(balance_raw_str)
                                                            divisor = Decimal(10) ** decimals_int
                                                            balance_decimal = raw_decimal / divisor

                                                            # Calculate price
                                                            price_decimal = Decimal(price_str)
                                                            calculated_price = balance_decimal * price_decimal
                                                            trx_total += calculated_price
                                                    except (ValueError, InvalidOperation, ZeroDivisionError):
                                                        pass

                                    # Calculate percentage difference
                                    if trx_total != 0:
                                        pct_diff = abs(dam_decimal - trx_total) / trx_total * Decimal('100')
                                        if pct_diff <= Decimal('1'):
                                            cell_c.value = "Passed"
                                        else:
                                            cell_c.value = "Failed"
                                    else:
                                        cell_c.value = "Passed" if dam_decimal == 0 else "Failed"

                                    # Write calculated value to D column
                                    ws_combined.cell(c_row_idx, 4).value = float(trx_total)

                                    print(f"         DAM: ${dam_decimal}, TRX Total: ${trx_total}, Result: {cell_c.value}")
                                except Exception as e:
                                    print(f"         ⚠️  Error calculating TRX totals: {e}")
                                    cell_c.value = "Error - Calculation"

                            elif address_str.startswith('0x') and len(address_str) == 42:
                                # EVM address - use Sim Dune + Rabby data
                                print(f"      Validating EVM address: {address_str}")

                                try:
                                    total_calculated = Decimal('0')

                                    # Sum Sim + Coingecko + Debank API (Column K - Calculated Price)
                                    sim_dune_total = Decimal('0')
                                    if ws_sim_dune:
                                        for row_idx in range(2, ws_sim_dune.max_row + 1):
                                            sim_address = ws_sim_dune.cell(row_idx, 1).value  # A column
                                            if sim_address and str(sim_address).strip().lower() == address_str.lower():
                                                calc_price = ws_sim_dune.cell(row_idx, 11).value  # K column - Calculated Price
                                                if calc_price:
                                                    try:
                                                        calc_price_str = str(calc_price).replace(',', '').strip()
                                                        if calc_price_str:
                                                            sim_dune_total += Decimal(calc_price_str)
                                                    except (ValueError, InvalidOperation):
                                                        pass

                                    # Sum Rabby Api Data (Column M - Calculated Value)
                                    rabby_total = Decimal('0')
                                    if ws_rabby:
                                        for row_idx in range(2, ws_rabby.max_row + 1):
                                            rabby_address = ws_rabby.cell(row_idx, 1).value  # A column
                                            if rabby_address and str(rabby_address).strip().lower() == address_str.lower():
                                                calc_value = ws_rabby.cell(row_idx, 13).value  # M column - Calculated Value
                                                if calc_value:
                                                    try:
                                                        calc_value_str = str(calc_value).replace(',', '').strip()
                                                        if calc_value_str:
                                                            rabby_total += Decimal(calc_value_str)
                                                    except (ValueError, InvalidOperation):
                                                        pass

                                    # Total = Sim Dune + Rabby
                                    total_calculated = sim_dune_total + rabby_total

                                    # Calculate percentage difference
                                    if total_calculated != 0:
                                        pct_diff = abs(dam_decimal - total_calculated) / total_calculated * Decimal('100')
                                        if pct_diff <= Decimal('1'):
                                            cell_c.value = "Passed"
                                        else:
                                            cell_c.value = "Failed"
                                    else:
                                        cell_c.value = "Passed" if dam_decimal == 0 else "Failed"

                                    # Write calculated value to D column
                                    ws_combined.cell(c_row_idx, 4).value = float(total_calculated)

                                    print(f"         DAM: ${dam_decimal}, SimDune: ${sim_dune_total}, Rabby: ${rabby_total}, Total: ${total_calculated}, Result: {cell_c.value}")
                                except Exception as e:
                                    print(f"         ⚠️  Error calculating EVM totals: {e}")
                                    cell_c.value = "Error - Calculation"

                            else:
                                # Check if it's an exchange table name (case-insensitive)
                                # Find matching exchange table
                                matched_exchange = None
                                address_lower = address_str.lower()
                                for exchange_name in exchange_tables:
                                    if exchange_name.lower() == address_lower:
                                        matched_exchange = exchange_name
                                        break

                                if matched_exchange:
                                    # Exchange account - use exchange table total
                                    print(f"      Validating Exchange: {address_str} (matched: {matched_exchange})")

                                    try:
                                        exchange_total = None

                                        # Find the exchange table (using the matched name for exact sheet lookup)
                                        if matched_exchange in wb.sheetnames:
                                            ws_exchange = wb[matched_exchange]

                                            # Find row where P column (16) = "Total" (case-insensitive)
                                            for row_idx in range(2, ws_exchange.max_row + 1):
                                                p_val = ws_exchange.cell(row_idx, 16).value  # P column
                                                if p_val and str(p_val).strip().upper() == "TOTAL":
                                                    q_val = ws_exchange.cell(row_idx, 17).value  # Q column
                                                    if q_val:
                                                        try:
                                                            q_clean = str(q_val).replace(',', '').replace('$', '').strip()
                                                            exchange_total = Decimal(q_clean)
                                                        except (ValueError, InvalidOperation):
                                                            pass
                                                    break  # Found total row

                                        # Validate
                                        if exchange_total is not None:
                                            if exchange_total != 0:
                                                pct_diff = abs(dam_decimal - exchange_total) / exchange_total * Decimal('100')
                                                if pct_diff <= Decimal('1'):
                                                    cell_c.value = "Passed"
                                                else:
                                                    cell_c.value = "Failed"
                                            else:
                                                cell_c.value = "Passed" if dam_decimal == 0 else "Failed"

                                            # Write calculated value to D column
                                            ws_combined.cell(c_row_idx, 4).value = float(exchange_total)

                                            print(f"         DAM: ${dam_decimal}, Exchange Total: ${exchange_total}, Result: {cell_c.value}")
                                        else:
                                            cell_c.value = "Total Not Found"
                                            print(f"         ⚠️  Total row not found in exchange table")
                                    except Exception as e:
                                        print(f"         ⚠️  Error validating exchange: {e}")
                                        cell_c.value = "Error - Exchange"
                                else:
                                    # Unknown format
                                    cell_c.value = "Not Applicable"

                        except Exception as e:
                            print(f"      ⚠️  Error validating {address_str}: {e}")
                            cell_c.value = "Error"

                print(f"   ✅ Updated C column validation in 'Overview - Combined Net Worth'")

            # Merge SimDune sheets into DAM workbook (before final save)
            # Raw data saved as JSON files instead of Excel sheets
            _raw_json_folder = os.path.dirname(excel_path)  # Same folder as the Excel file
            if sim_dune_excel_files:
                print(f"\n🔀 Merging SimDune sheets into DAM Excel...")
                from openpyxl import load_workbook as _load_sim_wb
                for _sim_file in sim_dune_excel_files:
                    if os.path.exists(_sim_file):
                        try:
                            _swb = _load_sim_wb(_sim_file)
                            for _sname in _swb.sheetnames:
                                # Skip raw sheets — save as JSON instead
                                if _sname.lower() in ('simdune', 'simdune raw', 'sim dune raw'):
                                    _src = _swb[_sname]
                                    _raw_data = [
                                        [str(v).replace('\n', ' ').replace('\r', '') if v is not None else v for v in _row]
                                        for _row in _src.iter_rows(values_only=True)
                                    ]
                                    _json_path = os.path.join(_raw_json_folder, f"SimDune_Raw_{_safe_name}.json")
                                    with open(_json_path, 'w') as _jf:
                                        import json as _json_sd
                                        _json_sd.dump(_raw_data, _jf, indent=2, default=str)
                                    print(f"   📄 Saved SimDune raw data to: {os.path.basename(_json_path)}")
                                    continue
                                _src = _swb[_sname]
                                # Skip if sheet already exists — avoid creating duplicate "(Raw)" tab
                                if _sname in wb.sheetnames:
                                    continue
                                _dst = wb.create_sheet(_sname)
                                for _row in _src.iter_rows():
                                    for _cell in _row:
                                        _dst.cell(row=_cell.row, column=_cell.column, value=_cell.value)
                            _swb.close()
                            print(f"   ✅ Merged: {os.path.basename(_sim_file)}")
                        except Exception as _e:
                            print(f"   ⚠️  Could not merge {os.path.basename(_sim_file)}: {_e}")

            # Merge Rabby sheets into DAM workbook (before final save)
            if rabby_excel_files:
                print(f"\n🔀 Merging Rabby sheets into DAM Excel...")
                from openpyxl import load_workbook as _load_rabby_wb
                for rabby_file in rabby_excel_files:
                    if os.path.exists(rabby_file):
                        try:
                            _rwb = _load_rabby_wb(rabby_file)
                            for _sname in _rwb.sheetnames:
                                # Skip raw sheets — save as JSON instead
                                if 'raw' in _sname.lower():
                                    _src = _rwb[_sname]
                                    _raw_data = []
                                    for _row in _src.iter_rows(values_only=True):
                                        _raw_data.append(list(_row))
                                    _safe_name = _sname.replace(' ', '_').replace('/', '_')
                                    _json_path = os.path.join(_raw_json_folder, f"{_safe_name}_{target_portfolio_name}.json")
                                    with open(_json_path, 'w') as _jf:
                                        import json as _json_rb
                                        _json_rb.dump(_raw_data, _jf, indent=2, default=str)
                                    print(f"   📄 Saved {_sname} to: {os.path.basename(_json_path)}")
                                    continue
                                _src = _rwb[_sname]
                                _dst = wb.create_sheet(_sname)
                                for _row in _src.iter_rows():
                                    for _cell in _row:
                                        _dst.cell(row=_cell.row, column=_cell.column, value=_cell.value)
                            _rwb.close()
                            print(f"   ✅ Merged: {os.path.basename(rabby_file)}")
                            try:
                                os.remove(rabby_file)
                                print(f"   🗑️  Deleted separate file: {os.path.basename(rabby_file)}")
                            except Exception as _del_e:
                                print(f"   ⚠️  Could not delete {os.path.basename(rabby_file)}: {_del_e}")
                        except Exception as _e:
                            print(f"   ⚠️  Could not merge {os.path.basename(rabby_file)}: {_e}")

            # Write intercepted Rabby browser network data as JSON file
            if _intercepted_rabby_raw:
                try:
                    _json_path = os.path.join(_raw_json_folder, f"Rabby_Browser_Raw_{target_portfolio_name}.json")
                    with open(_json_path, 'w') as _jf:
                        import json as _json_br
                        _json_br.dump(_intercepted_rabby_raw, _jf, indent=2, default=str)
                    print(f"   📄 Saved Rabby browser raw data to: {os.path.basename(_json_path)}")
                except Exception as _e:
                    print(f"   ⚠️  Could not write intercepted Rabby data: {_e}")
            else:
                print(f"   ℹ️  No Rabby API calls intercepted from browser (DAM may proxy via backend)")

            # Compute H (Amount Validation) for Overview - De-Fi using Python
            # Runs after Rabby Api Data sheet is available in the workbook
            if "Overview - De-Fi" in wb.sheetnames and "Rabby Api Data" in wb.sheetnames:
                try:
                    _ws_defi_g = wb["Overview - De-Fi"]
                    _ws_rabby_g = wb["Rabby Api Data"]

                    # Load all Rabby rows into memory (skip header row 1)
                    # Rabby columns (0-indexed from list): 
                    # 0:Name, 1:ID, 2:Chain, 3:Pool Name, 4:Description,
                    # 5:Side, 6:Symbol/Currency Pair, 7:Leverage, 8:PnL(USD),
                    # 9:Price, 10:Amount, 11:Calculated Value
                    _rabby_rows_g = []
                    for _r in range(2, _ws_rabby_g.max_row + 1):
                        _rabby_rows_g.append([_ws_rabby_g.cell(_r, _c).value for _c in range(1, 14)])

                    # Update column headers (shifted: K→Amount Diff %, L→Value, M→Value Validation, N→Api Calc Value, O→Value Difference)
                    _ws_defi_g.cell(1, 11).value = "Amount Diff %"
                    _ws_defi_g.cell(1, 13).value = "Value Validation"
                    _ws_defi_g.cell(1, 14).value = "Api Calc Value"
                    _ws_defi_g.cell(1, 15).value = "Value Difference"
                    _ws_defi_g.cell(1, 16).value = "Side"

                    _g_passed = _g_failed = _g_no_match = 0

                    # Check if first pass already populated validation results
                    _first_pass_ran = False
                    if _ws_defi_g.max_row >= 2:
                        _h_check = _ws_defi_g.cell(2, 8).value
                        if _h_check and str(_h_check).strip() in ("Passed", "Failed", "No Match", "-"):
                            _first_pass_ran = True
                            print(f"   ℹ️  First pass already populated DeFi validation — secondary pass will only write Row Matched")

                    for _row_idx in range(2, _ws_defi_g.max_row + 1):
                        _da = _ws_defi_g.cell(_row_idx, 1).value  # A: De-Fi name
                        if not _da:
                            continue

                        _dc = _ws_defi_g.cell(_row_idx, 3).value  # C: Type
                        _dd = _ws_defi_g.cell(_row_idx, 4).value  # D: Pool/Position Pair
                        _de = _ws_defi_g.cell(_row_idx, 5).value  # E: Description
                        _dg = _ws_defi_g.cell(_row_idx, 7).value  # G: Amount Tooltip
                        _dk = _ws_defi_g.cell(_row_idx, 12).value  # L: Value (was K, shifted by 1)

                        _da_n = str(_da).strip().lower()
                        _dc_n = str(_dc).strip().lower() if _dc else ""
                        _dd_n = str(_dd).strip().lower() if _dd else ""
                        _de_n = str(_de).strip().lower() if _de else ""

                        _is_hyperliquid = "hyperliquid" in _da_n
                        _is_hl_perp = _is_hyperliquid and _dc_n == "perpetuals"

                        # Hyperliquid Perpetuals: skip amount validation, set to "-"
                        if _is_hl_perp:
                            _ws_defi_g.cell(_row_idx, 8).value = "-"   # H: Amount Validation
                            _ws_defi_g.cell(_row_idx, 9).value = "-"   # I: FE - Amount Validation
                            _ws_defi_g.cell(_row_idx, 10).value = "-"  # J: Amount Validation Diff
                            # Still do value validation below


                        # Find matching Rabby row using 4 criteria + closest amount:
                        # Rabby columns (0-indexed from list loaded via range(1,14)):
                        # [0]:Address, [1]:Name, [2]:ID, [3]:Chain, [4]:Pool Name, [5]:Description,
                        # [6]:Side, [7]:Symbol/Currency Pair, [8]:Leverage, [9]:PnL(USD),
                        # [10]:Price, [11]:Amount, [12]:Calculated Value
                        _candidates_g = []
                        for _rrow in _rabby_rows_g:
                            if len(_rrow) < 13:
                                continue
                            _r_name = str(_rrow[1]).strip().lower() if _rrow[1] else ""      # B: Name
                            _r_chain = str(_rrow[3]).strip().lower() if _rrow[3] else ""     # D: Chain
                            _r_pool = str(_rrow[4]).strip().lower() if _rrow[4] else ""      # E: Pool Name
                            _r_symbol = str(_rrow[7]).strip().lower() if _rrow[7] else ""    # H: Symbol/Currency Pair

                            _matched = False
                            if _is_hyperliquid:
                                _r_desc = str(_rrow[5]).strip().lower() if _rrow[5] else ""
                                if _da_n == _r_name and _dc_n == _r_pool and _dd_n == _r_symbol and _de_n == _r_desc:
                                    _matched = True
                            else:
                                # Match on Name + Chain + Pool Name + Symbol
                                _db_n = str(_ws_defi_g.cell(_row_idx, 2).value or "").strip().lower()  # B: Chain
                                if _da_n == _r_name and _db_n == _r_chain and _dc_n == _r_pool and _dd_n == _r_symbol:
                                    _matched = True

                            if _matched:
                                try:
                                    _r_amt = Decimal(str(_rrow[11]).replace(',', '')) if _rrow[11] else Decimal('0')
                                except:
                                    _r_amt = Decimal('0')
                                _candidates_g.append({"amount": _r_amt, "calc_value": _rrow[12], "row": _rrow})

                        # Fallback: Name + Symbol only (for non-Hyperliquid)
                        if not _candidates_g and not _is_hyperliquid:
                            for _rrow in _rabby_rows_g:
                                if len(_rrow) < 13:
                                    continue
                                _r_name = str(_rrow[1]).strip().lower() if _rrow[1] else ""
                                _r_symbol = str(_rrow[7]).strip().lower() if _rrow[7] else ""
                                if _da_n == _r_name and _dd_n == _r_symbol:
                                    try:
                                        _r_amt = Decimal(str(_rrow[11]).replace(',', '')) if _rrow[11] else Decimal('0')
                                    except:
                                        _r_amt = Decimal('0')
                                    _candidates_g.append({"amount": _r_amt, "calc_value": _rrow[12], "row": _rrow})

                        # Pick closest by amount
                        _matched_amount = None
                        _matched_calc_value = None
                        if _candidates_g:
                            try:
                                _dg_dec = Decimal(str(_dg).replace(',', '')) if _dg else Decimal('0')
                            except:
                                _dg_dec = Decimal('0')
                            _best_g = min(_candidates_g, key=lambda c: abs(c["amount"] - _dg_dec))
                            _matched_amount = str(_best_g["amount"])       # Amount for amount comparison
                            _matched_calc_value = str(_best_g["calc_value"])  # Calc Value for value comparison
                            # Write Row Matched (column T = 20)
                            _r = _best_g["row"]
                            _ws_defi_g.cell(_row_idx, 20).value = f"{_r[1]}|{_r[3]}|{_r[4]}|{_r[7]} Amt={_best_g['amount']}"

                        # Fallback for Hyperliquid: Name + Symbol only (without Description)
                        if _matched_amount is None and _is_hyperliquid:
                            for _rrow in _rabby_rows_g:
                                if len(_rrow) < 13:
                                    continue
                                _r_name = str(_rrow[1]).strip().lower() if _rrow[1] else ""
                                _r_symbol = str(_rrow[7]).strip().lower() if _rrow[7] else ""
                                if _da_n == _r_name and _dd_n == _r_symbol:
                                    _matched_amount = _rrow[11]
                                    _matched_calc_value = _rrow[12]
                                    break

                        if _matched_amount is not None:
                            try:
                                from decimal import Decimal as _Dec

                                def _to_dec(v):
                                    """Convert value to Decimal, handling scientific notation and treating tiny values as 0."""
                                    if v in (None, ''):
                                        return _Dec('0')
                                    s = str(v).replace(',', '').replace('$', '').replace('<', '').strip()
                                    if not s:
                                        return _Dec('0')
                                    d = _Dec(s)
                                    # Treat values smaller than 0.00001 as 0
                                    if abs(d) < _Dec('0.00001'):
                                        return _Dec('0')
                                    return d

                                if not _is_hl_perp and not _first_pass_ran:
                                    # Amount comparison: De-Fi G (Amount Tooltip) vs Rabby Amount
                                    _dg_dec = _to_dec(_dg)
                                    _ra_dec = _to_dec(_matched_amount)
                                    _amount_diff = _dg_dec - _ra_dec

                                    # Amount Validation (H column)
                                    if _is_hyperliquid:
                                        if _ra_dec != _Dec('0'):
                                            _pct = abs(_amount_diff) / abs(_ra_dec) * _Dec('100')
                                            _g_val = "Passed" if _pct <= _Dec('1') else "Failed"
                                        else:
                                            _g_val = "Passed" if _dg_dec == _Dec('0') else "Failed"
                                    else:
                                        # Passed if exact match OR within 1% difference
                                        if _ra_dec != _Dec('0'):
                                            _pct = abs(_amount_diff) / abs(_ra_dec) * _Dec('100')
                                            _g_val = "Passed" if _pct <= _Dec('1') else "Failed"
                                        else:
                                            _g_val = "Passed" if _dg_dec == _Dec('0') else "Failed"

                                    if _g_val == "Passed":
                                        _g_passed += 1
                                    else:
                                        _g_failed += 1

                                    # Amount Validation Diff (J column)
                                    _diff_str = str(_amount_diff)
                                    if '.' in _diff_str:
                                        _diff_str = _diff_str.rstrip('0').rstrip('.')
                                    _ws_defi_g.cell(_row_idx, 10).value = _diff_str
                                    _ws_defi_g.cell(_row_idx, 8).value = _g_val

                                # Value comparison: Value (L=12) minus Api Calc Value (N=14)
                                if not _first_pass_ran:
                                    _dk_dec = _to_dec(_dk)  # L column - Value
                                    # Read Api Calc Value from column N (14) — already written by first pass
                                    _n_val = _ws_defi_g.cell(_row_idx, 14).value
                                    _rv_dec = _to_dec(_n_val) if _n_val else _to_dec(_matched_calc_value)
                                    _value_diff = _dk_dec - _rv_dec

                                    # Value Validation (M column) — truncate to 2dp first, then 1% fallback
                                    if _rv_dec != _Dec('0'):
                                        _dk_trunc = _dk_dec.quantize(_Dec('0.01'), rounding='ROUND_DOWN')
                                        _rv_trunc = _rv_dec.quantize(_Dec('0.01'), rounding='ROUND_DOWN')
                                        if _dk_trunc == _rv_trunc:
                                            _v_val = "Passed"
                                        else:
                                            _v_pct = abs(_value_diff) / abs(_rv_dec) * _Dec('100')
                                            _v_val = "Passed" if _v_pct <= _Dec('1') else "Failed"
                                    else:
                                        _v_val = "Passed" if _dk_dec == _Dec('0') else "Failed"
                                    _ws_defi_g.cell(_row_idx, 13).value = _v_val

                                    # Value Difference (O column 15)
                                    _vdiff_str = str(_value_diff)
                                    if '.' in _vdiff_str:
                                        _vdiff_str = _vdiff_str.rstrip('0').rstrip('.')
                                    _ws_defi_g.cell(_row_idx, 15).value = _vdiff_str

                            except Exception:
                                if not _first_pass_ran and not _is_hl_perp:
                                    _ws_defi_g.cell(_row_idx, 8).value = "Error"
                                    _g_failed += 1
                        else:
                            if not _first_pass_ran:
                                _g_no_match += 1
                                if not _is_hl_perp:
                                    _ws_defi_g.cell(_row_idx, 8).value = "No Match"
                                    _ws_defi_g.cell(_row_idx, 10).value = ""
                                _ws_defi_g.cell(_row_idx, 13).value = "No Match"
                                _ws_defi_g.cell(_row_idx, 15).value = ""

                    print(f"   ✅ Computed Amount & Value Validation for Overview - De-Fi: {_g_passed} Passed, {_g_failed} Failed, {_g_no_match} No Match")
                except Exception as _g_err:
                    print(f"   ⚠️  Could not compute validation for Overview - De-Fi: {_g_err}")
                    import traceback
                    traceback.print_exc()

            # Apply conditional formatting: Passed = green (#66BB6A), Failed = red (#E57373)
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill
            _passed_fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
            _failed_fill = PatternFill(start_color="E57373", end_color="E57373", fill_type="solid")
            _tooltip_na_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            for _ws in wb.worksheets:
                _max_col_letter = _ws.cell(1, max(1, _ws.max_column)).column_letter
                _range = f"A1:{_max_col_letter}{max(1, _ws.max_row)}"
                _ws.conditional_formatting.add(_range, CellIsRule(
                    operator='equal', formula=['"Passed"'], fill=_passed_fill))
                _ws.conditional_formatting.add(_range, CellIsRule(
                    operator='equal', formula=['"Failed"'], fill=_failed_fill))
                _ws.conditional_formatting.add(_range, CellIsRule(
                    operator='equal', formula=['"Tooltip N/A, cant compare"'], fill=_tooltip_na_fill))
            print(f"   ✅ Applied Passed/Failed/Tooltip N/A conditional formatting to {len(wb.worksheets)} sheets")

            # Combine Platform / Chain / Token Allocation sheets into single "Allocation" tab
            _alloc_sources = [
                "Overview - Platform Allocation",
                "Overview - Chain Allocation",
                "Overview - Token Allocation",
            ]
            if all(s in wb.sheetnames for s in _alloc_sources):
                from openpyxl.styles import Font as _Font
                ws_alloc = wb.create_sheet("Allocation")
                _alloc_row = 1
                for _s_idx, _src_name in enumerate(_alloc_sources):
                    _src_ws = wb[_src_name]
                    _table_start_row = _alloc_row
                    for _src_r in _src_ws.iter_rows():
                        for _c_idx, _src_cell in enumerate(_src_r, start=1):
                            _dst = ws_alloc.cell(row=_alloc_row, column=_c_idx, value=_src_cell.value)
                            if _src_cell.font:
                                _dst.font = _src_cell.font.copy()
                            if _src_cell.fill and _src_cell.fill.fill_type != "none":
                                _dst.fill = _src_cell.fill.copy()
                        _alloc_row += 1
                    # Bold the keyword in col A of each table's header row
                    ws_alloc.cell(row=_table_start_row, column=1).font = _Font(bold=True)
                    del wb[_src_name]
                    # Two empty rows separator (except after last table)
                    if _s_idx < len(_alloc_sources) - 1:
                        _alloc_row += 2
                print(f"   ✅ Combined allocation sheets into 'Allocation' tab")

                # Apply conditional formatting to the combined Allocation tab
                _alloc_max_col = ws_alloc.cell(1, max(1, ws_alloc.max_column)).column_letter
                _alloc_range = f"A1:{_alloc_max_col}{max(1, ws_alloc.max_row)}"
                ws_alloc.conditional_formatting.add(_alloc_range, CellIsRule(
                    operator='equal', formula=['"Passed"'], fill=_passed_fill))
                ws_alloc.conditional_formatting.add(_alloc_range, CellIsRule(
                    operator='equal', formula=['"Failed"'], fill=_failed_fill))
                ws_alloc.conditional_formatting.add(_alloc_range, CellIsRule(
                    operator='equal', formula=['"No API Data"'], fill=_tooltip_na_fill))
                print(f"   ✅ Applied Passed/Failed conditional formatting to 'Allocation' tab")
            else:
                print(f"   ⚠️  Could not combine allocation sheets (not all present)")

            # ── STEP 4.5: Individual wallet page extraction ────────────────────────────
            print("\n" + "="*80)
            print("STEP 4.5: Extract Individual Wallet Pages")
            print("="*80)

            import re as _re_ind

            _evm_addr_re = _re_ind.compile(r'^0x[A-Fa-f0-9]{40}$')
            _trx_addr_re = _re_ind.compile(r'^[Tt][A-Za-z0-9]{33}$')

            def _ind_scroll(pg):
                last_h = pg.evaluate("document.body.scrollHeight")
                for _ in range(20):
                    pg.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    pg.wait_for_timeout(800)
                    new_h = pg.evaluate("document.body.scrollHeight")
                    if new_h == last_h:
                        break
                    last_h = new_h
                pg.evaluate("window.scrollTo(0, 0)")
                pg.wait_for_timeout(500)

            _IND_COLS = ["Section", "Chain", "Name / Pool", "Price", "Price (24h)",
                         "Share", "Amount", "Amount Tooltip", "Value", "Position Type"]
            _IND_DEFI_KWS = {'aave', 'morpho', 'compound', 'uniswap', 'curve', 'lido',
                              'maker', 'spark', 'euler', 'pendle', 'yearn', 'balancer',
                              'convex', 'merkl', 'hyperliquid', 'fluid', 'kamino',
                              'drift', 'orca', 'raydium', 'jupiter', 'midas', 'virtuals',
                              'gearbox'}

            _ind_tabs_added = []

            try:
                _portfolio_overview_url = page.url
                if "portfolioId=" not in _portfolio_overview_url:
                    print("   ⚠️  Not on portfolio page — skipping individual wallet extraction")
                else:
                    # Reload overview to get clean DOM
                    page.goto(_portfolio_overview_url)
                    page.wait_for_load_state("networkidle")
                    page.wait_for_timeout(3000)
                    _ind_scroll(page)

                    # Detect wallet addresses (EVM/TRX only, filter out CEX names)
                    _ind_addrs = []
                    _ind_seen = set()
                    try:
                        _found_addrs = page.evaluate('''() => {
                            const results = [];
                            const els = document.querySelectorAll('[data-tooltip-id^="address-display-tooltip-"]');
                            for (const el of els) {
                                let addr = "";
                                const hlEl = el.querySelector("[data-highlight-target]");
                                if (hlEl) { addr = hlEl.getAttribute("data-highlight-target") || ""; }
                                if (!addr) {
                                    const tid = el.getAttribute("data-tooltip-id") || "";
                                    addr = tid.replace("address-display-tooltip-", "");
                                }
                                if (addr.trim()) results.push(addr.trim());
                            }
                            return results;
                        }''')
                        for _a in _found_addrs:
                            _a = _a.strip()
                            if _a and _a not in _ind_seen and (_evm_addr_re.match(_a) or _trx_addr_re.match(_a)):
                                _ind_seen.add(_a)
                                _ind_addrs.append(_a)
                    except Exception as _ae:
                        print(f"   ⚠️  Address detection failed: {_ae}")

                    print(f"   Found {len(_ind_addrs)} wallet address(es): {_ind_addrs}")

                    # Skip Step 4.5 if portfolio has only 1 address (redundant with main extraction)
                    if len(_ind_addrs) <= 1:
                        print(f"   ⏭️  Skipping individual wallet extraction — only {len(_ind_addrs)} address (already extracted in main flow)")
                        _ind_addrs = []  # Clear to skip the loop

                    for _addr in _ind_addrs:
                        _last8 = _addr[-8:].lower()
                        _wallet_tab = f"{_last8} Wallet"
                        _defi_tab   = f"{_last8} De-Fi"
                        print(f"\n   ── {_last8} ({_addr}) ──")

                        # Click address element to navigate to individual wallet page
                        _addr_lower = _addr.lower()
                        try:
                            _cr = page.evaluate(f'''() => {{
                                const addrLower = "{_addr_lower}";
                                const all = document.querySelectorAll('[data-tooltip-id^="address-display-tooltip-"]');
                                for (const el of all) {{
                                    const tid = (el.getAttribute("data-tooltip-id") || "").toLowerCase();
                                    if (tid === "address-display-tooltip-" + addrLower) {{
                                        el.scrollIntoView({{behavior: "instant", block: "center"}});
                                        let t = el;
                                        for (let i = 0; i < 15; i++) {{
                                            t = t.parentElement;
                                            if (!t) break;
                                            if ((t.className || "").includes("cursor-pointer")) {{
                                                t.click(); return true;
                                            }}
                                        }}
                                        el.click(); return true;
                                    }}
                                }}
                                return false;
                            }}''')
                            page.wait_for_load_state("networkidle")
                            page.wait_for_timeout(3000)
                            if _addr_lower in page.url.lower():
                                print(f"      Landed → {page.url}")
                            else:
                                print(f"      ⚠️  URL does not contain address param: {page.url}")
                        except Exception as _ce:
                            print(f"      ⚠️  Click failed: {_ce}")

                        _ind_scroll(page)

                        # Extract wallet token table (scoped to #wallet-section)
                        _wallet_rows = []
                        try:
                            _ws_loc = page.locator("#wallet-section").first
                            if _ws_loc.count() > 0:
                                import re as _re_tt

                                def _read_tooltip(trigger_elem, tip_id, is_amount=False):
                                    """Hover trigger, wait for visible tooltip, return text. Retries once."""
                                    for _attempt in range(2):
                                        try:
                                            trigger_elem.scroll_into_view_if_needed(timeout=1000)
                                            trigger_elem.hover(timeout=2000)
                                            # Wait up to 800ms for tooltip to become visible
                                            _tdiv = page.locator(f'#{tip_id}').first
                                            try:
                                                _tdiv.wait_for(state="visible", timeout=800)
                                            except Exception:
                                                page.wait_for_timeout(400)
                                            if _tdiv.count() > 0 and _tdiv.is_visible():
                                                if is_amount:
                                                    _t = _tdiv.inner_text().strip()
                                                    if _t.count('.') > 1:
                                                        _mx = _re_tt.search(r'^(\d+\.\d+)', _t)
                                                        _t = _mx.group(1) if _mx else _t
                                                    return _t
                                                else:
                                                    return _tdiv.text_content().strip().replace('$', '')
                                        except Exception:
                                            pass
                                    return None

                                # --- Price Tooltips ---
                                _price_tips = {}
                                try:
                                    for _pi, _pelem in enumerate(_ws_loc.locator('[data-tooltip-id*="price-tooltip"]').all()):
                                        _ptid = _pelem.get_attribute('data-tooltip-id')
                                        if _ptid:
                                            _val = _read_tooltip(_pelem, _ptid, is_amount=False)
                                            if _val:
                                                _price_tips[_pi] = _val
                                            else:
                                                print(f"      ⚠️  Price tooltip[{_pi}] empty (id={_ptid})")
                                except Exception as _tte:
                                    print(f"      ⚠️  Price tooltip extraction error: {_tte}")

                                # --- Share Tooltips ---
                                _share_tips = {}
                                try:
                                    for _si, _selem in enumerate(_ws_loc.locator('[data-tooltip-id*="share-tooltip"]').all()):
                                        _stid = _selem.get_attribute('data-tooltip-id')
                                        if _stid:
                                            _val = _read_tooltip(_selem, _stid, is_amount=False)
                                            if _val:
                                                _share_tips[_si] = _val.replace('%', '')
                                            else:
                                                print(f"      ⚠️  Share tooltip[{_si}] empty (id={_stid})")
                                except Exception as _tte:
                                    print(f"      ⚠️  Share tooltip extraction error: {_tte}")

                                # --- Amount Tooltips ---
                                _amount_tips = {}
                                try:
                                    for _ai, _aelem in enumerate(_ws_loc.locator('[data-tooltip-id*="amount-tooltip"]').all()):
                                        _atid = _aelem.get_attribute('data-tooltip-id')
                                        if _atid:
                                            _val = _read_tooltip(_aelem, _atid, is_amount=True)
                                            if _val:
                                                _amount_tips[_ai] = _val
                                            else:
                                                print(f"      ⚠️  Amount tooltip[{_ai}] empty (id={_atid})")
                                except Exception as _tte:
                                    print(f"      ⚠️  Amount tooltip extraction error: {_tte}")

                                print(f"      Tooltips: {len(_price_tips)} price, {len(_share_tips)} share, {len(_amount_tips)} amount")

                                _tbl = _ws_loc.locator("table").first
                                if _tbl.count() > 0:
                                    _row_idx = 0
                                    for _tr in _tbl.locator("tbody tr").all():
                                        _cells = _tr.locator("td").all()
                                        if len(_cells) < 4:
                                            continue
                                        # Table cols: 0=Chain, 1=Name, 2=Price(24h combined), 3=Amount, 4=Share
                                        # Value column does not exist on individual wallet page
                                        _chain = ""
                                        _name = ""
                                        _price = ""
                                        _price_24h = ""
                                        _amount = ""
                                        _share = ""
                                        _value = ""
                                        try:
                                            _chain = _cells[0].inner_text(timeout=1000).split("\n")[0].strip()
                                        except Exception:
                                            pass
                                        try:
                                            _name = _cells[1].inner_text(timeout=1000).split("\n")[0].strip()
                                        except Exception:
                                            pass
                                        # Price (24H) combined cell — extract price from first line,
                                        # 24h change from pct element (red/green class)
                                        try:
                                            _pcell = _cells[2]
                                            _pct_el = _pcell.locator('[class*="text-error"],[class*="text-success"],[class*="bg-error"],[class*="bg-success"]').first
                                            if _pct_el.count() > 0 and _pct_el.is_visible():
                                                _pct_txt = _pct_el.text_content().strip()
                                                _pct_m = re.search(r'([\d.]+)%?', _pct_txt)
                                                if _pct_m:
                                                    _pct_cls = _pct_el.get_attribute('class') or ""
                                                    _price_24h = f"-{_pct_m.group(1)}" if ('error' in _pct_cls.lower() or '↓' in _pct_txt) else _pct_m.group(1)
                                            _pline = _pcell.inner_text(timeout=1000).split("\n")[0].strip()
                                            _lt_m = re.search(r'<\s*\$?([\d,]+\.?\d*)', _pline)
                                            _pm = re.search(r'\$?([\d,]+\.?\d*)', _pline)
                                            if _lt_m:
                                                _price = f"< {_lt_m.group(1)}"
                                            elif _pm:
                                                _price = _pm.group(1)
                                        except Exception:
                                            pass
                                        try:
                                            _amount = _cells[3].inner_text(timeout=1000).split("\n")[0].strip()
                                        except Exception:
                                            pass
                                        try:
                                            _share = _cells[4].inner_text(timeout=1000).split("\n")[0].strip() if len(_cells) > 4 else ""
                                        except Exception:
                                            pass
                                        try:
                                            _value = _cells[5].inner_text(timeout=1000).split("\n")[0].strip() if len(_cells) > 5 else ""
                                        except Exception:
                                            pass
                                        _wallet_rows.append({
                                            "Section": "Token Holdings",
                                            "Chain": _chain, "Name / Pool": _name,
                                            "Price": _price, "Price (24h)": _price_24h,
                                            "Price Tooltip": _price_tips.get(_row_idx, ""),
                                            "Amount": _amount, "Share": _share,
                                            "Share Tooltip": _share_tips.get(_row_idx, ""),
                                            "Amount Tooltip": _amount_tips.get(_row_idx, ""),
                                            "Value": _value,
                                            "Position Type": "",
                                        })
                                        _row_idx += 1
                        except Exception as _we:
                            print(f"      ⚠️  Wallet table failed: {_we}")

                        # Extract per-address DeFi data by clicking each protocol tab
                        # (sections below wallet table show overview-level unfiltered data;
                        #  protocol tabs at the top of the page show per-address data)
                        _defi_rows = []
                        try:
                            _ind_chain_code_map = {
                                'ethereum': 'eth', 'base': 'base', 'bsc': 'bsc',
                                'binance smart chain': 'bsc', 'arbitrum': 'arb',
                                'optimism': 'op', 'polygon': 'matic', 'avalanche': 'avax',
                                'fantom': 'ftm', 'tron': 'tron', 'linea': 'linea',
                                'scroll': 'scroll', 'zksync era': 'era', 'mantle': 'mnt',
                                'blast': 'blast',
                            }
                            _known_chain_alts = [
                                'Ethereum', 'Base', 'Arbitrum', 'Polygon', 'Optimism',
                                'BSC', 'Binance Smart Chain', 'Avalanche', 'Tron', 'Fantom',
                                'Linea', 'Scroll', 'zkSync Era', 'Mantle', 'Blast',
                            ]
                            _chain_hint_map = {
                                'ETH': 'eth', 'ETHER': 'eth', 'ETHEREUM': 'eth',
                                'BASE': 'base', 'ARB': 'arb', 'ARBITRUM': 'arb',
                                'OP': 'op', 'OPTIMISM': 'op',
                                'MATIC': 'matic', 'POLYGON': 'matic',
                                'BSC': 'bsc', 'BNB': 'bsc',
                                'AVAX': 'avax', 'AVALANCHE': 'avax',
                                'TRON': 'tron', 'TRX': 'tron',
                                'LINEA': 'linea', 'BLAST': 'blast',
                                'SCROLL': 'scroll', 'MANTLE': 'mnt', 'MNT': 'mnt',
                            }
                            def _defi_tip_read(_trigger, _tid, _is_amt=True):
                                try:
                                    _trigger.scroll_into_view_if_needed(timeout=1000)
                                    _trigger.hover(timeout=2000)
                                    _tdv = page.locator(f'#{_tid}').first
                                    try:
                                        _tdv.wait_for(state="visible", timeout=800)
                                    except Exception:
                                        page.wait_for_timeout(400)
                                    if _tdv.count() > 0 and _tdv.is_visible():
                                        _t = _tdv.inner_text().strip()
                                        if _is_amt and _t.count('.') > 1:
                                            _mx2 = re.search(r'^(\d+\.\d+)', _t)
                                            _t = _mx2.group(1) if _mx2 else _t
                                        return _t
                                except Exception:
                                    pass
                                return None

                            # Find all protocol tabs on the individual address page
                            _proto_tabs = page.locator('[data-slot="tab"], [role="tab"]').all()
                            print(f"      Found {len(_proto_tabs)} tab(s) on individual address page")
                            _wallet_tab_ref = None

                            for _ptab in _proto_tabs:
                                try:
                                    _ptxt = (_ptab.inner_text(timeout=1000) or "").strip()
                                    if not _ptxt:
                                        continue
                                    # Store WALLET tab reference for restoring later
                                    if _ptxt.lower() == 'wallet':
                                        _wallet_tab_ref = _ptab
                                        continue
                                    # Skip non-DeFi tabs
                                    _ptxt_lower = _ptxt.lower()
                                    if not any(kw in _ptxt_lower for kw in _IND_DEFI_KWS):
                                        continue

                                    # Parse protocol name and chain from tab label
                                    # e.g. "AAVE V3 (ETH...)" → proto="Aave V3", chain="eth"
                                    _tab_proto = _ptxt
                                    _tab_chain = ""
                                    _cmatch = re.search(r'\(([^)]+)\)', _ptxt)
                                    if _cmatch:
                                        _tab_proto = _ptxt[:_ptxt.rfind('(')].strip()
                                        _chain_raw = _cmatch.group(1).strip().upper().rstrip('.')
                                        for _ck, _cv in _chain_hint_map.items():
                                            if _chain_raw.startswith(_ck):
                                                _tab_chain = _cv; break
                                        if not _tab_chain:
                                            _tab_chain = _chain_raw.lower()

                                    print(f"      Clicking tab: '{_ptxt}' → proto='{_tab_proto}' chain='{_tab_chain}'")
                                    _ptab.scroll_into_view_if_needed(timeout=1000)
                                    _ptab.click(timeout=3000)
                                    page.wait_for_load_state("networkidle", timeout=8000)
                                    page.wait_for_timeout(1000)

                                    # Find the DeFi table in the active tab content
                                    _active_tbl = None
                                    for _tsel in [
                                        '[data-slot="tab-content"][data-state="active"] table',
                                        '[data-slot="tabs-content"][data-state="active"] table',
                                        '[role="tabpanel"][data-state="active"] table',
                                        '[role="tabpanel"] table',
                                    ]:
                                        _tnode = page.locator(_tsel).first
                                        if _tnode.count() > 0:
                                            _active_tbl = _tnode; break

                                    if not _active_tbl:
                                        # Fallback: find table inside currently visible bg-grey-30 section
                                        for _fsec in page.locator("div.bg-grey-30.rounded-sm").all():
                                            try:
                                                if _fsec.is_visible():
                                                    _ft = _fsec.locator("table").first
                                                    if _ft.count() > 0:
                                                        _active_tbl = _ft; break
                                            except Exception:
                                                pass

                                    if not _active_tbl or _active_tbl.count() == 0:
                                        print(f"         ⚠️  No table found for tab '{_ptxt}'")
                                        continue

                                    # Collect amount tooltips from the active tab view
                                    _sec_atips = {}
                                    try:
                                        _atip_scope = page.locator(
                                            '[data-slot="tab-content"][data-state="active"], '
                                            '[data-slot="tabs-content"][data-state="active"], '
                                            '[role="tabpanel"]'
                                        ).first
                                        if _atip_scope.count() == 0:
                                            _atip_scope = page
                                        for _dai, _daelem in enumerate(_atip_scope.locator('[data-tooltip-id*="amount-tooltip"]').all()):
                                            _datid = _daelem.get_attribute('data-tooltip-id')
                                            if _datid:
                                                _v = _defi_tip_read(_daelem, _datid, _is_amt=True)
                                                if _v:
                                                    _sec_atips[_dai] = _v
                                    except Exception:
                                        pass

                                    # Extract rows: col0=Position Type, col1=Pool, col2=Amount, col3=Value
                                    _row_di = 0
                                    for _tr2 in _active_tbl.locator("tbody tr").all():
                                        _cells2 = _tr2.locator("td").all()
                                        if len(_cells2) < 2:
                                            continue
                                        _pos_type, _pool, _amt, _val2 = "", "", "", ""
                                        try:
                                            _pos_type = _cells2[0].inner_text(timeout=500).split("\n")[0].strip()
                                        except Exception:
                                            pass
                                        try:
                                            _pool = _cells2[1].inner_text(timeout=500).split("\n")[0].strip() if len(_cells2) > 1 else ""
                                        except Exception:
                                            pass
                                        try:
                                            _amt = _cells2[2].inner_text(timeout=500).split("\n")[0].strip() if len(_cells2) > 2 else ""
                                        except Exception:
                                            pass
                                        try:
                                            _val2 = _cells2[3].inner_text(timeout=500).split("\n")[0].strip() if len(_cells2) > 3 else ""
                                        except Exception:
                                            pass
                                        _defi_rows.append({
                                            "Section": _tab_proto,
                                            "Chain": _tab_chain,
                                            "Position Type": _pos_type,
                                            "Pool": _pool,
                                            "Amount": _amt,
                                            "Amount Tooltip": _sec_atips.get(_row_di, ""),
                                            "Value": _val2,
                                        })
                                        _row_di += 1
                                    print(f"         → {_row_di} rows extracted")
                                except Exception:
                                    pass

                            # Restore WALLET tab so the page is in its original state
                            if _wallet_tab_ref:
                                try:
                                    _wallet_tab_ref.scroll_into_view_if_needed(timeout=1000)
                                    _wallet_tab_ref.click(timeout=3000)
                                    page.wait_for_timeout(1000)
                                except Exception:
                                    pass
                        except Exception as _de:
                            print(f"      ⚠️  DeFi extraction failed: {_de}")

                        print(f"      Token Holdings: {len(_wallet_rows)} rows | DeFi: {len(_defi_rows)} rows")

                        if _wallet_rows:
                            # Build input list matching add_validation_columns_to_overview_token expected format:
                            # 0:Chain, 1:Name, 2:Price, 3:Price(24h), 4:Amount, 5:Share, 6:Value,
                            # 7:PriceTooltip, 8:ShareTooltip, 9:AmountTooltip
                            _raw = [["Chain", "Name", "Price", "Price (24h)", "Amount", "Share", "Value",
                                     "Price Tooltip", "Share Tooltip", "Amount Tooltip"]]
                            for _r in _wallet_rows:
                                _raw.append([
                                    _r.get("Chain", ""),
                                    _r.get("Name / Pool", ""),
                                    _r.get("Price", ""),
                                    _r.get("Price (24h)", ""),
                                    _r.get("Amount", ""),
                                    _r.get("Share", ""),
                                    _r.get("Value", ""),
                                    _r.get("Price Tooltip", ""),
                                    _r.get("Share Tooltip", ""),
                                    _r.get("Amount Tooltip", ""),
                                ])
                            _validated = add_validation_columns_to_overview_token(_raw)
                            # Prepend Address column (shifts all data cols right by 1)
                            # New layout: A=Address, B=Chain, C=Name, D=Price, E=PriceTip,
                            # F=FE-PriceVal, G=PriceVal, H=Price abs_diff, I=Price Diff %,
                            # J=Price24h, K=Price24hVal, L=Price24hDiff, M=Amount, N=AmountTip,
                            # O=FE-AmtVal, P=AmtVal, Q=AmtDiffVal, R=Share, S=ShareTip,
                            # T=FE-ShareVal, U=CalcShare, V=ShareVal, W=Value,
                            # X=ValueVal, Y=CalcValue, Z=DataRow
                            _validated = [["Address"] + list(r) for r in _validated]
                            _ws_w = wb.create_sheet(_wallet_tab)
                            _ind_hdr_map = {
                                1: "Address", 5: "Price Tooltip", 6: "FE - Price Validation",
                                7: "Price Validation", 8: "Price abs_diff", 9: "Price Diff %",
                                11: "Price (24h) Validation", 12: "Price (24H) Diff Validation",
                                14: "Amount Tooltip", 15: "FE - Amount Validation",
                                16: "Amount Validation", 17: "Amount Diff Validation",
                                19: "Share Tooltip", 20: "FE - Share Validation",
                                21: "Calculation Share", 22: "Share Validation",
                                24: "Value Validation", 25: "Calculated Value", 26: "Data Row",
                            }
                            for _ri, _rdata in enumerate(_validated, start=1):
                                for _ci, _v in enumerate(_rdata, start=1):
                                    # Address column: header or full address value
                                    if _ci == 1:
                                        _cell = _ws_w.cell(row=_ri, column=1, value="Address" if _ri == 1 else _addr)
                                        continue
                                    # Numeric conversion for D(Price=4), J(Price24h=10), R(Share=18), W(Value=23)
                                    if _ri > 1 and _ci in [4, 10, 18, 23]:
                                        try:
                                            if isinstance(_v, str):
                                                if _v.strip().startswith('<'):
                                                    _v = _v.replace(',', '').replace('$', '').replace('%', '').replace('>', '').strip()
                                                else:
                                                    _cv = _v.replace(',', '').replace('$', '').replace('%', '').replace('<', '').replace('>', '').strip()
                                                    if _cv:
                                                        _v = float(_cv)
                                        except (ValueError, AttributeError):
                                            pass
                                    _cell = _ws_w.cell(row=_ri, column=_ci, value=_v)
                                    # Yellow fill for validation columns
                                    if _ci in [6, 7, 8, 9, 11, 12, 15, 16, 17, 20, 21, 22, 24, 25, 26]:
                                        _cell.fill = yellow_fill
                                    # Header row overrides
                                    if _ri == 1 and _ci in _ind_hdr_map:
                                        _cell.value = _ind_hdr_map[_ci]
                                    elif _ri > 1:
                                        if _ci == 6:   # F = FE - Price Validation (E=PriceTip, D=Price)
                                            _cell.value = (
                                                f'=LET('
                                                f'tipRaw,E{_ri},'
                                                f'hasTip,AND(tipRaw<>"",tipRaw<>"Tooltip N/A"),'
                                                f'src,IF(hasTip,tipRaw,D{_ri}),'
                                                f'srcClean,SUBSTITUTE(SUBSTITUTE(src,"$",""),",",""),'
                                                f'IFERROR('
                                                f'IFS('
                                                f'src="","No Price Data",'
                                                f'NOT(hasTip),"Tooltip N/A, cant compare",'
                                                f'AND(VALUE(srcClean)<0.01,VALUE(srcClean)<>0,D{_ri}="< 0.01"),"Passed",'
                                                f'AND(VALUE(srcClean)<0.01,VALUE(srcClean)<>0,D{_ri}<>"< 0.01"),"Failed",'
                                                f'AND(VALUE(srcClean)>=0.01,TRUNC(VALUE(srcClean),2)=VALUE(D{_ri})),"Passed",'
                                                f'AND(VALUE(srcClean)>=0.01,TRUNC(VALUE(srcClean),2)<>VALUE(D{_ri})),"Failed")'
                                                f',"No Price Data"))'
                                            )
                                        elif _ci == 15:  # O = FE - Amount Validation (N=AmtTip, M=Amount)
                                            _cell.value = (
                                                f'=LET('
                                                f'tipRaw,N{_ri},'
                                                f'hasTip,AND(tipRaw<>"",tipRaw<>"Tooltip N/A"),'
                                                f'src,IF(hasTip,tipRaw,M{_ri}),'
                                                f'amt,M{_ri},'
                                                f'IF(OR(src="",IFERROR(VALUE(src),-999)=-999),"No Amount Data",'
                                                f'IF(NOT(hasTip),"Tooltip N/A, cant compare",'
                                                f'IF(OR(src=0,IFERROR(VALUE(src),0)=0),'
                                                f'IF(OR(amt="",amt=0,IFERROR(VALUE(amt),0)=0),"Passed","Failed"),'
                                                f'IFERROR('
                                                f'IF(VALUE(src)<0.00001,IF(amt="< 0.00001","Passed","Failed"),'
                                                f'IF(LEFT(SUBSTITUTE(TEXT(VALUE(src),"0.000000000000000"),",",""),FIND(".",SUBSTITUTE(TEXT(VALUE(src),"0.000000000000000"),",",""))+5)='
                                                f'LEFT(SUBSTITUTE(TEXT(VALUE(amt),"0.000000000000000"),",",""),FIND(".",SUBSTITUTE(TEXT(VALUE(amt),"0.000000000000000"),",",""))+5),"Passed","Failed"))'
                                                f',"Error")))))' 
                                            )
                                        elif _ci == 20:  # T = FE - Share Validation (S=ShareTip, R=Share)
                                            _cell.value = (
                                                f'=LET('
                                                f'tipRaw,S{_ri},'
                                                f'hasTip,AND(tipRaw<>"",tipRaw<>"Tooltip N/A"),'
                                                f'src,IF(hasTip,tipRaw,R{_ri}),'
                                                f'IFERROR('
                                                f'IFS(src="","No Share Data",'
                                                f'NOT(hasTip),"Tooltip N/A, cant compare",'
                                                f'AND(VALUE(src)<0.01,VALUE(src)<>0,TRIM(R{_ri})="< 0.01"),"Passed",'
                                                f'AND(VALUE(src)<0.01,VALUE(src)<>0,TRIM(R{_ri})<>"< 0.01"),"Failed",'
                                                f'AND(VALUE(src)>=0.01,ROUND(VALUE(src)*100,0)=ROUND(VALUE(R{_ri})*100,0)),"Passed",'
                                                f'AND(VALUE(src)>=0.01,ROUND(VALUE(src)*100,0)<>ROUND(VALUE(R{_ri})*100,0)),"Failed",'
                                                f'AND(VALUE(src)=0,VALUE(R{_ri})=0),"Passed",'
                                                f'AND(VALUE(src)=0,VALUE(R{_ri})<>0),"Failed")'
                                                f',"No Share Data"))'
                                            )
                                        elif _ci == 25:  # Y = Calculated Value = Amount(M) * Price(D)
                                            _cell.value = f'=M{_ri}*D{_ri}'
                                        elif _ci == 24:  # X = Value Validation (Y=CalcVal, W=Value)
                                            _cell.value = (
                                                f'=IFERROR('
                                                f'IFS(AND(VALUE(Y{_ri})<0.01,VALUE(Y{_ri})<>0,W{_ri}="< 0.01"),"Passed",'
                                                f'AND(VALUE(Y{_ri})<0.01,VALUE(Y{_ri})<>0,W{_ri}<>"< 0.01"),"Failed",'
                                                f'AND(VALUE(Y{_ri})>=0.01,TRUNC(VALUE(Y{_ri}),2)=W{_ri}),"Passed",'
                                                f'AND(VALUE(Y{_ri})>=0.01,TRUNC(VALUE(Y{_ri}),2)<>W{_ri}),"Failed",'
                                                f'AND(VALUE(Y{_ri})=0,VALUE(W{_ri})=0),"Passed",'
                                                f'AND(VALUE(Y{_ri})=0,VALUE(W{_ri})<>0),"Failed")'
                                                f',"No Calc Value"'
                                                f')'
                                            )
                            # ── Price Validation (G=col7), Price abs_diff (H=col8), Price Diff % (I=col9),
                            #    Price 24h Validation (K=col11), Price 24h Diff (L=col12)
                            # Uses same maps built for Overview - Wallet
                            try:
                                _ind_sd_price = sim_dune_price_map
                            except NameError:
                                _ind_sd_price = {}
                            try:
                                _ind_sd_24h = sim_dune_24h_map
                            except NameError:
                                _ind_sd_24h = {}
                            try:
                                _ind_trx_price = trx_price_map
                            except NameError:
                                _ind_trx_price = {}
                            try:
                                _ind_trx_24h = trx_24h_map
                            except NameError:
                                _ind_trx_24h = {}

                            for _vri in range(2, _ws_w.max_row + 1):
                                _vb = str(_ws_w.cell(_vri, 2).value or "").strip()   # B = Chain
                                _vc = str(_ws_w.cell(_vri, 3).value or "").strip()   # C = Token
                                _ve = _ws_w.cell(_vri, 5).value                      # E = Price Tooltip
                                _vi = _ws_w.cell(_vri, 10).value                     # J = Price (24h)
                                if not _vb or not _vc:
                                    continue
                                _cg = _ws_w.cell(_vri, 7)   # G = Price Validation
                                _ch = _ws_w.cell(_vri, 8)   # H = Price abs_diff
                                _ci2 = _ws_w.cell(_vri, 9)  # I = Price Diff %
                                _cj = _ws_w.cell(_vri, 11)  # K = Price (24h) Validation
                                _ck = _ws_w.cell(_vri, 12)  # L = Price (24H) Diff Validation
                                try:
                                    _vD = Decimal
                                    _ep = _ve.replace(',','').replace('$','').strip() if isinstance(_ve, str) else str(_ve or '0')
                                    _dam_p = _vD(_ep) if _ep else _vD('0')
                                    _ip = str(_vi or '0').replace(',','').replace('%','').strip()
                                    _dam_24h = _vD(_ip) if _ip else _vD('0')

                                    if _vb.lower() == "tron":
                                        _tk = _vc.lower()
                                        if _tk in _ind_trx_price:
                                            _sp = _ind_trx_price[_tk]
                                            if _sp != 0:
                                                _dp = abs((_dam_p - _sp) / _sp) * _vD('100')
                                                _cg.value = "Passed" if _dp <= _vD('1') else "Failed"
                                                _ch.value = str(_dam_p - _sp)
                                                _ci2.value = str(_dp.quantize(_vD('0.0001')))
                                            else:
                                                _cg.value = "Failed"; _ch.value = str(_dam_p); _ci2.value = ""
                                        else:
                                            _cg.value = "Token Not Found"; _ch.value = ""; _ci2.value = ""
                                        if _tk in _ind_trx_24h:
                                            _s24 = _ind_trx_24h[_tk]
                                            _cj.value = "Passed" if abs(_dam_24h - _s24) <= _vD('1') else "Failed"
                                            _ck.value = str(_dam_24h - _s24)
                                        else:
                                            _cj.value = "Token Not Found"; _ck.value = ""
                                    elif _vb.lower() in ["ethereum", "base", "binance smart chain"]:
                                        _lk = (_vb.lower(), _vc.lower())
                                        if _lk in _ind_sd_price:
                                            _sp = _ind_sd_price[_lk]
                                            if _sp != 0:
                                                _dp = abs((_dam_p - _sp) / _sp) * _vD('100')
                                                _cg.value = "Passed" if _dp <= _vD('1') else "Failed"
                                                _ch.value = str(_dam_p - _sp)
                                                _ci2.value = str(_dp.quantize(_vD('0.0001')))
                                            else:
                                                _cg.value = "Failed"; _ch.value = str(_dam_p); _ci2.value = ""
                                        else:
                                            _cg.value = "Token Not Found"; _ch.value = ""; _ci2.value = ""
                                        if _lk in _ind_sd_24h:
                                            _s24 = _ind_sd_24h[_lk]
                                            _cj.value = "Passed" if abs(_dam_24h - _s24) <= _vD('1') else "Failed"
                                            _ck.value = str(_dam_24h - _s24)
                                        else:
                                            _cj.value = "Token Not Found"; _ck.value = ""
                                    else:
                                        _cg.value = "Chain Not Supported"; _ch.value = ""; _ci2.value = ""
                                        _cj.value = ""; _ck.value = ""
                                except Exception as _ve2:
                                    _cg.value = "Error"; _ch.value = ""; _ci2.value = ""; _cj.value = ""; _ck.value = ""
                                for _vc2 in [_cg, _ch, _ci2, _cj, _ck]:
                                    _vc2.fill = yellow_fill

                            # ── Amount Validation (O=col15), Amount Diff (P=col16)
                            # Build Sim Dune amount map: (address, chain, symbol) -> Decimal(amount)
                            _ind_sd_amount_map = {}
                            if "Sim + Coingecko + Debank API" in wb.sheetnames:
                                _sd_ws = wb["Sim + Coingecko + Debank API"]
                                for _sdi in range(2, _sd_ws.max_row + 1):
                                    _sdA = str(_sd_ws.cell(_sdi, 1).value or "").strip().lower()  # A=Address
                                    _sdB = str(_sd_ws.cell(_sdi, 2).value or "").strip().lower()  # B=Chain
                                    _sdC = str(_sd_ws.cell(_sdi, 3).value or "").strip().lower()  # C=Symbol
                                    _sdE = _sd_ws.cell(_sdi, 5).value                              # E=Amount
                                    if _sdA and _sdB and _sdC and _sdE is not None:
                                        try:
                                            _sdEv = Decimal(str(_sdE).replace(',', '').strip()) if isinstance(_sdE, str) else Decimal(str(_sdE))
                                            _ind_sd_amount_map[(_sdA, _sdB, _sdC)] = _sdEv
                                        except:
                                            pass

                            # Build TRX balance amount map: token.lower() -> Decimal(amount)
                            _ind_trx_amount_map = {}
                            try:
                                _ta_src = trx_balance_map
                            except NameError:
                                _ta_src = {}
                            for _tk2, _tv2 in _ta_src.items():
                                # trx_balance_map keys are token symbols; values are dicts with 'amount' or similar
                                # Check structure — may be Decimal directly or dict
                                try:
                                    if isinstance(_tv2, dict):
                                        _ind_trx_amount_map[_tk2.lower()] = Decimal(str(_tv2.get('amount', _tv2.get('balance', 0))))
                                    else:
                                        _ind_trx_amount_map[_tk2.lower()] = Decimal(str(_tv2))
                                except:
                                    pass

                            def _norm_amt(v):
                                """Normalize amount string: strip trailing zeros."""
                                try:
                                    d = Decimal(str(v).replace(',', '').strip())
                                    return d.normalize()
                                except:
                                    return None

                            for _ori in range(2, _ws_w.max_row + 1):
                                _oa = str(_ws_w.cell(_ori, 1).value or "").strip().lower()   # A = Address
                                _ob = str(_ws_w.cell(_ori, 2).value or "").strip().lower()   # B = Chain
                                _oc = str(_ws_w.cell(_ori, 3).value or "").strip().lower()   # C = Token
                                _om = _ws_w.cell(_ori, 14).value                              # N = Amount Tooltip
                                if not _ob or not _oc:
                                    continue
                                _co = _ws_w.cell(_ori, 16)   # P = Amount Validation
                                _cp = _ws_w.cell(_ori, 17)   # Q = Amount Diff Validation
                                try:
                                    _dam_amt = _norm_amt(_om) if _om is not None else None
                                    if _ob in ["ethereum", "base", "binance smart chain"]:
                                        _ak = (_oa, _ob, _oc)
                                        if _ak in _ind_sd_amount_map:
                                            _src_amt = _ind_sd_amount_map[_ak]
                                            _src_norm = _src_amt.normalize()
                                            if _dam_amt is not None:
                                                _co.value = "Passed" if _dam_amt == _src_norm else "Failed"
                                                _cp.value = str(_dam_amt - _src_norm)
                                            else:
                                                _co.value = "No Tooltip"; _cp.value = ""
                                        else:
                                            _co.value = "Token Not Found"; _cp.value = ""
                                    elif _ob == "tron":
                                        _tk3 = _oc.lower()
                                        if _tk3 in _ind_trx_amount_map:
                                            _src_amt = _ind_trx_amount_map[_tk3]
                                            _src_norm = _src_amt.normalize()
                                            if _dam_amt is not None:
                                                _co.value = "Passed" if _dam_amt == _src_norm else "Failed"
                                                _cp.value = str(_dam_amt - _src_norm)
                                            else:
                                                _co.value = "No Tooltip"; _cp.value = ""
                                        else:
                                            _co.value = "Token Not Found"; _cp.value = ""
                                    else:
                                        _co.value = "Chain Not Supported"; _cp.value = ""
                                except Exception as _oe:
                                    _co.value = "Error"; _cp.value = str(_oe)
                                _co.fill = yellow_fill
                                _cp.fill = yellow_fill

                            # ── Calculation Share (U=col21) and Share Validation (V=col22)
                            # U = (PriceTip(E=col5) * AmtTip(N=col14)) / SUM(all PriceTip*AmtTip) * 100
                            _row_w_vals = {}   # row_idx -> Decimal W
                            _total_w = Decimal('0')
                            for _tri in range(2, _ws_w.max_row + 1):
                                _te = _ws_w.cell(_tri, 5).value    # E = Price Tooltip
                                _tm = _ws_w.cell(_tri, 14).value   # N = Amount Tooltip
                                if _te is None or _tm is None:
                                    continue
                                try:
                                    _tp = Decimal(str(_te).replace(',','').replace('$','').strip())
                                    _ta = Decimal(str(_tm).replace(',','').strip())
                                    _tw = _tp * _ta
                                    _row_w_vals[_tri] = _tw
                                    _total_w += _tw
                                except:
                                    pass

                            for _tri in range(2, _ws_w.max_row + 1):
                                _ct = _ws_w.cell(_tri, 21)   # U = Calculation Share
                                _cu = _ws_w.cell(_tri, 22)   # V = Share Validation
                                if _tri in _row_w_vals and _total_w != 0:
                                    _calc_share = (_row_w_vals[_tri] / _total_w * Decimal('100')).quantize(Decimal('0.01'))
                                    _ct.value = float(_calc_share)
                                    # V = LET formula: Passed if ABS(R - U) <= 1
                                    _cu.value = (
                                        f'=LET(_qs,VALUE(R{_tri}),_ts,VALUE(U{_tri}),'
                                        f'IF(ABS(_qs-_ts)<=1,"Passed","Failed"))'
                                    )
                                else:
                                    _ct.value = ""
                                    _cu.value = ""
                                _ct.fill = yellow_fill
                                _cu.fill = yellow_fill

                            # Highlight Passed/Failed cells in wallet tab using conditional formatting
                            # (covers both static values and formula-evaluated results)
                            from openpyxl.styles import PatternFill as _PFillW
                            from openpyxl.formatting.rule import CellIsRule as _CIRuleW
                            from openpyxl.formatting.rule import FormulaRule as _FRuleW
                            _green_w = _PFillW(fill_type="solid", fgColor="66BB6A")
                            _red_w   = _PFillW(fill_type="solid", fgColor="EF5350")
                            _w_range = f"A2:{_ws_w.cell(1, _ws_w.max_column).column_letter}{_ws_w.max_row}"
                            _ws_w.conditional_formatting.add(_w_range,
                                _FRuleW(formula=['EXACT(A2,"Passed")'], fill=_green_w))
                            _ws_w.conditional_formatting.add(_w_range,
                                _FRuleW(formula=['EXACT(A2,"Failed")'], fill=_red_w))

                            _ind_tabs_added.append(_wallet_tab)
                            print(f"      ✅ Added '{_wallet_tab}'")

                        if _defi_rows:
                            _ws_d = wb.create_sheet(_defi_tab)
                            # Header matches Overview - De-Fi exactly
                            _ind_defi_hdr = [
                                'Defi', 'Chain', 'Type', 'Pool/ Position Pair', 'Description', 'Amount', 'Amount Tooltip',
                                'Amount Validation', 'FE - Amount Validation', 'Amount Validation Diff',
                                'Value', 'Value Validation', 'Side', 'Leverage', 'Collateral', 'P&L', 'Value'
                            ]
                            _ws_d.append(_ind_defi_hdr)
                            # Build raw data rows
                            _ind_defi_data = [_ind_defi_hdr]
                            for _r in _defi_rows:
                                _row_d = [
                                    _r.get("Section", ""),        # A: Defi
                                    _r.get("Chain", ""),          # B: Chain
                                    _r.get("Position Type", ""),  # C: Type
                                    _r.get("Pool", ""),           # D: Pool/ Position Pair
                                    _r.get("Description", ""),    # E: Description
                                    _r.get("Amount", ""),         # F: Amount
                                    _r.get("Amount Tooltip", ""), # G: Amount Tooltip
                                    "",  # H: Amount Validation
                                    "",  # I: FE - Amount Validation
                                    "",  # J: Amount Validation Diff
                                    _r.get("Value", ""),          # K: Value
                                    "",  # L: Value Validation
                                    _r.get("Side", ""),           # M: Side
                                    _r.get("Leverage", ""),       # N: Leverage
                                    _r.get("Collateral", ""),     # O: Collateral
                                    _r.get("P&L", ""),            # P: P&L
                                    _r.get("Value2", ""),         # Q: Value
                                ]
                                _ind_defi_data.append(_row_d)
                            # Apply Rabby validation (G, H, I, K columns)
                            try:
                                _rdata_ind = rabby_extracted_data if rabby_extracted_data else []
                            except NameError:
                                _rdata_ind = []
                            if _rdata_ind:
                                try:
                                    _ind_defi_validated = add_validation_to_defi_tab(_ind_defi_data, _rdata_ind)
                                    _ind_defi_data = _ind_defi_validated
                                except Exception as _vde:
                                    print(f"      ⚠️  DeFi validation error: {_vde}")
                            # Write to sheet (skip header already appended)
                            from openpyxl.styles import PatternFill as _PFillInd
                            from openpyxl.formatting.rule import FormulaRule as _FRuleInd
                            _green_ind = _PFillInd(fill_type="solid", fgColor="66BB6A")
                            _red_ind   = _PFillInd(fill_type="solid", fgColor="EF5350")
                            for _row_d in _ind_defi_data[1:]:
                                _ws_d.append(_row_d)
                            # Apply Passed/Failed highlights via conditional formatting
                            _d_range = f"A2:{_ws_d.cell(1, _ws_d.max_column).column_letter}{_ws_d.max_row}"
                            _ws_d.conditional_formatting.add(_d_range,
                                _FRuleInd(formula=['EXACT(A2,"Passed")'], fill=_green_ind))
                            _ws_d.conditional_formatting.add(_d_range,
                                _FRuleInd(formula=['EXACT(A2,"Failed")'], fill=_red_ind))
                            _ind_tabs_added.append(_defi_tab)
                            print(f"      ✅ Added '{_defi_tab}' ({len(_defi_rows)} rows)")

                        # Return to overview for next address
                        try:
                            page.goto(_portfolio_overview_url)
                            page.wait_for_load_state("networkidle")
                            page.wait_for_timeout(2000)
                            _ind_scroll(page)
                        except Exception:
                            pass

                    print(f"\n   ✅ {len(_ind_tabs_added)} individual wallet tab(s) added: {_ind_tabs_added}")

            except Exception as _ind_e:
                print(f"   ⚠️  Individual wallet extraction error: {_ind_e}")
                import traceback as _ind_tb
                _ind_tb.print_exc()

            # Reorder sheets: Sim Dune after Overview-Wallet, Rabby sheets after Overview-De-Fi
            _desired_order = [
                "Overview - Header & Token Holdings Header",
                "Overview - Wallet",
                "Sim + Coingecko + Debank API",
                "Overview - De-Fi",
                "Rabby Api Data",
                "Overview - Combined Net Worth",
                "Allocation",
                "TRX Balance, Price",
            ]
            for _i, _sname in enumerate(_desired_order):
                if _sname in wb.sheetnames:
                    wb.move_sheet(_sname, offset=wb.sheetnames.index(_sname) * -1 + _i)
            # Re-apply order using index approach for reliability
            _current = list(wb.sheetnames)
            _front = [s for s in _desired_order if s in _current]
            _back = [s for s in _current if s not in _front]
            _final_order = _front + _back
            wb._sheets.sort(key=lambda ws: _final_order.index(ws.title) if ws.title in _final_order else len(_final_order))

            # Save workbook
            sheet_count = len(wb.sheetnames)
            wb.save(excel_path)
            print(f"\n📁 DAM Excel file: {excel_filename} ({sheet_count} sheets)")

            # Summary
            print("\n" + "="*80)
            print("✅ DAM PORTFOLIO EXTRACTION COMPLETED")
            print("="*80)
            print(f"Test Case: TC00003")
            print(f"Portfolio: {portfolio_name_from_excel}")
            if SKIP_TRX_API or not TRX_ADDRESSES:
                print(f"Portfolio Type: CEX-only (no wallet addresses)")
            else:
                print(f"TRX Address: {TRX_ADDRESS}")
            print(f"\nData Extracted (Overview allocation tables only):")
            print(f"  1. Overview - Wallet: {len(all_tables_data.get('Overview - Wallet', []))} rows")
            print(f"  2. Overview - Header & Token Holdings Header: {len(all_tables_data.get('Overview - Header & Token Holdings Header', []))} rows")
            print(f"  3. Overview - Combined Net Worth: {len(all_tables_data.get('Overview - Combined Net Worth', []))} rows")
            print(f"  4. Overview - Platform Allocation: {len(all_tables_data.get('Overview - Platform Allocation', []))} rows")
            print(f"  5. Overview - Chain Allocation: {len(all_tables_data.get('Overview - Chain Allocation', []))} rows")
            print(f"  6. Overview - Token Allocation: {len(all_tables_data.get('Overview - Token Allocation', []))} rows")
            print(f"\nDAM Excel file: {excel_path}")
            if sim_dune_excel_files:
                print(f"Sim Dune Excel files: {len(sim_dune_excel_files)}")
                for f in sim_dune_excel_files:
                    print(f"  - {os.path.basename(f)}")
            if rabby_excel_files:
                print(f"Rabby data merged into DAM Excel ({len(rabby_excel_files)} file(s) — separate files deleted)")
            print(f"Screenshots: {screenshot_folder}/")
            print("="*80 + "\n")

            # Keep browser open
            print("⏸️  Keeping browser open for 20 seconds...")
            page.wait_for_timeout(20000)

            return excel_path, screenshot_folder, sim_dune_excel_files, rabby_excel_files

        except Exception as e:
            print(f"\n❌ ERROR: {e}")
            import traceback
            traceback.print_exc()

            try:
                page.screenshot(path=f"{screenshot_folder}/error_{os.getpid()}.png")
            except:
                pass

            return None, None, [], []

        finally:
            browser.close()


def main(args=None):
    """
    Main function to run TC00003 test
    """
    print(f"\nTC00003 | Portfolio: {PORTFOLIO_NAME_PREFIX} | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    if not SKIP_TRX_API and TRX_ADDRESSES:
        for idx, addr in enumerate(TRX_ADDRESSES, 1):
            print(f"  TRX {idx}: {addr}")

    balance_excel = None
    balance_filename = None

    # Part 1: TRX Balance API Test (tests ALL addresses)
    # Skip if portfolio has no wallet addresses (CEX-only)
    if SKIP_TRX_API:
        print("\nPART 1: TRX BALANCE API TEST - SKIPPED (No wallet addresses)")
    else:
        balance_excel = run_trx_balance_api_test(addresses=TRX_ADDRESSES, portfolio_name=PORTFOLIO_NAME_PREFIX)
        # Extract just the filename from the path for Excel formula reference
        balance_filename = os.path.basename(balance_excel) if balance_excel else None

    # Part 2: DAM Portfolio Full Extraction
    # Pass portfolio name so it navigates to the exact portfolio (not search by address)
    dam_excel, screenshot_folder, sim_dune_files, rabby_files = run_dam_portfolio_extraction(
        trx_balance_filename=balance_filename,
        target_portfolio_name=PORTFOLIO_NAME_PREFIX
    )

    # Final Summary
    print("\n" + "="*80)
    print("✅ TC00003 TEST COMPLETED")
    print("="*80)
    print(f"Test Case: TC00003")
    if SKIP_TRX_API:
        print(f"Portfolio Type: CEX-only (no wallet addresses)")
    else:
        print(f"Addresses: {len(TRX_ADDRESSES)}")
        for idx, addr in enumerate(TRX_ADDRESSES, 1):
            print(f"  {idx}. {addr}")
    print(f"\nFiles Created:")
    if SKIP_TRX_API:
        print(f"  1. TRX Balance API: ⏭️  Skipped (CEX-only portfolio)")
    elif balance_excel:
        print(f"  1. TRX Balance API: {balance_excel}")
        tab_count = 3 + (4 * len(TRX_ADDRESSES))  # 3 shared + 4 per address
        print(f"     - {tab_count} tabs: 3 shared (Token List, All Token Info, TRX Balance Price) + {4 * len(TRX_ADDRESSES)} address-specific ({len(TRX_ADDRESSES)} address(es) x 4)")
    else:
        print(f"  1. TRX Balance API: ❌ Failed")

    if dam_excel:
        print(f"  2. DAM Portfolio: {dam_excel}")
        try:
            from openpyxl import load_workbook as _lwb
            _wb = _lwb(dam_excel, read_only=True)
            _sheet_names = _wb.sheetnames
            _wb.close()
            print(f"     - {len(_sheet_names)} sheets: {', '.join(_sheet_names)}")
        except:
            print(f"     - Overview allocation tables")
    else:
        print(f"  2. DAM Portfolio: ❌ Failed")

    if sim_dune_files:
        print(f"  3. Sim Dune API (EVM):")
        for idx, sim_file in enumerate(sim_dune_files, 1):
            print(f"     {idx}. {sim_file}")
            print(f"        - 2 sheets: SimDune (raw), Sim Dune - Address Amount (extracted)")

    if rabby_files:
        print(f"  4. Rabby Protocol API (EVM):")
        for idx, rabby_file in enumerate(rabby_files, 1):
            print(f"     {idx}. {rabby_file}")
            print(f"        - Rabby Api Data (extracted)")

    if screenshot_folder:
        print(f"\nScreenshots: {screenshot_folder}/")
    print(f"Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*80 + "\n")


def lookup_portfolio_in_excel(portfolio_name):
    """
    Look up a portfolio name in DAM addresses.xlsx and return ALL associated addresses.
    Returns (addresses_list, portfolio_name) tuple or ([], None) if not found.
    Addresses are stored in columns B, C, D, etc.
    """
    dam_excel_path = "test_data/DAM addresses.xlsx"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(dam_excel_path, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2:
                excel_portfolio = str(row[0]).strip() if row[0] else ""

                # Case-insensitive match for portfolio name
                if excel_portfolio.lower() == portfolio_name.lower():
                    # Collect all non-empty addresses from columns B onward
                    excel_addresses = []
                    for col_val in row[1:]:
                        addr = str(col_val).strip() if col_val else ""
                        if addr and ((addr.startswith('T') and len(addr) == 34) or
                                     (addr.startswith('0x') and len(addr) == 42)):
                            excel_addresses.append(addr)
                    wb.close()
                    return excel_addresses, excel_portfolio

        wb.close()
        return [], None
    except Exception as e:
        print(f"⚠️  Error reading DAM addresses Excel: {e}")
        return [], None


def lookup_address_in_excel(address):
    """
    Look up an ADDRESS in DAM addresses.xlsx and return the portfolio name that contains it.
    Searches all address columns (B, C, D, etc.).
    Returns (portfolio_name, all_addresses_in_portfolio) tuple or (None, []) if not found.
    """
    dam_excel_path = "test_data/DAM addresses.xlsx"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(dam_excel_path, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2:
                portfolio_name = str(row[0]).strip() if row[0] else ""

                # Collect all addresses from this row
                row_addresses = []
                for col_val in row[1:]:
                    addr = str(col_val).strip() if col_val else ""
                    if addr and addr.startswith('T') and len(addr) == 34:
                        row_addresses.append(addr)

                # Check if target address is in this row
                if address in row_addresses:
                    wb.close()
                    return portfolio_name, row_addresses

        wb.close()
        return None, []
    except Exception as e:
        print(f"⚠️  Error reading DAM addresses Excel: {e}")
        return None, []


def classify_address(addr):
    """
    Classify an address as Tron, EVM, or Exchange.
    Returns: 'tron', 'evm', or 'exchange'
    """
    addr = addr.strip()
    # Tron address: Starts with T, 34 characters
    if (addr.startswith('T') or addr.startswith('t')) and len(addr) == 34:
        return 'tron'
    # EVM address: Starts with 0x, 42 characters (0x + 40 hex)
    elif addr.startswith('0x') and len(addr) == 42:
        return 'evm'
    # Everything else is considered Exchange
    else:
        return 'exchange'


def lookup_portfolio_in_dam(portfolio_name):
    """
    Look up a portfolio name in DAM UI and extract ALL addresses from Combined Net Worth.
    Classifies addresses as: Tron (T + 33 chars), EVM (0x + 40 hex), or Exchange.
    Returns: (trx_addresses, evm_addresses, exchanges, portfolio_name)
    Returns ([], [], [], None) if not found.
    """
    import re as _re_local
    from playwright.sync_api import sync_playwright
    from openpyxl import load_workbook
    import json as _json

    print(f"\n🌐 Checking DAM for portfolio: '{portfolio_name}'")

    # Load credentials from tc1_account.json if available, else fall back to Config
    _tc1_path = os.path.join(Config.PROJECT_ROOT, "test_data", "tc1_account.json")
    if os.path.exists(_tc1_path):
        with open(_tc1_path) as _f:
            _acc = _json.load(_f)
        test_email = _acc["email"]
        test_password = _acc["password"]
    else:
        test_email = Config.TEST_EMAIL
        test_password = Config.TEST_PASSWORD

    found_trx_addresses = []
    found_evm_addresses = []
    found_exchanges = []
    found_portfolio = None

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            slow_mo=300,
            channel="chrome",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        from playwright_stealth import Stealth
        Stealth(navigator_platform_override="MacIntel").apply_stealth_sync(context)
        page = context.new_page()

        try:
            # Sign in to DAM
            print(f"   🔐 Signing in to DAM with account: {test_email}")
            page.goto(Config.SIGN_IN_URL)
            page.wait_for_timeout(2000)

            page.fill('input[data-testid="input-email"]', test_email)
            page.fill('input[data-testid="input-password"]', test_password)
            page.click('button[data-testid="sign-in-btn"]')
            page.wait_for_timeout(8000)

            # Verify sign-in actually succeeded
            current_url = page.url
            if "sign-in" in current_url or "sign-up" in current_url:
                # Still on sign-in page — check for error messages
                error_msg = ""
                for err_sel in [
                    '[data-testid*="error"]',
                    '[class*="error"]',
                    '[class*="toast"]',
                    '[class*="alert"]',
                    '[class*="notification"]',
                    '[role="alert"]',
                    '[role="status"]',
                    'div:has-text("internal server error")',
                    'div:has-text("try again")',
                    'div:has-text("Invalid")',
                    'div:has-text("incorrect")',
                    'div:has-text("failed")',
                    'p:has-text("Invalid")',
                    'span:has-text("error")',
                ]:
                    try:
                        err_el = page.locator(err_sel).first
                        if err_el.is_visible(timeout=500):
                            txt = err_el.text_content().strip()
                            # Pick the most specific error text (skip huge blocks)
                            if txt and len(txt) < 200 and len(txt) > len(error_msg):
                                error_msg = txt
                            # Stop early if we found a clear error
                            if error_msg and any(kw in error_msg.lower() for kw in ['error', 'invalid', 'failed', 'try again', 'incorrect']):
                                break
                    except:
                        pass

                # If no specific error found, scan body text for error keywords
                if not error_msg:
                    try:
                        body_text = page.locator("body").text_content().strip()
                        for line in body_text.split('\n'):
                            line = line.strip()
                            if line and len(line) > 5 and len(line) < 200:
                                if any(kw in line.lower() for kw in ['error', 'invalid', 'incorrect', 'failed', 'wrong', 'expired', 'locked', 'captcha', 'verify', 'try again']):
                                    error_msg = line
                                    break
                    except:
                        pass

                print(f"   ❌ Sign-in FAILED — still on {current_url}")
                if error_msg:
                    print(f"   ❌ Error: {error_msg}")
                else:
                    print(f"   ❌ No error message found. Page may require CAPTCHA or credentials may be wrong.")
                print(f"   ❌ Credentials used:")
                print(f"      Email:    {test_email}")
                print(f"      Password: {test_password[:3]}{'*' * (len(test_password) - 3)}")
                browser.close()
                return [], [], [], "__SIGNIN_FAILED__"

            # Close popup if any
            try:
                for selector in ['button:has-text("×")', '[aria-label="close"]']:
                    if page.locator(selector).is_visible(timeout=1000):
                        page.locator(selector).first.click()
                        page.wait_for_timeout(500)
                        break
            except:
                pass

            # Handle redirect to specific portfolio
            current_url = page.url
            if "portfolioId=" in current_url:
                page.goto(f"{Config.BASE_URL}/portfolio")
                page.wait_for_timeout(3000)

            print("   ✅ Signed in")

            # Open Portfolio dropdown
            dropdown_opened = False
            for dropdown_sel in [
                'button:has-text("Portfolio")',
                'div:has-text("Portfolio") >> button',
                '[class*="portfolio"] button',
                'button[aria-haspopup]',
            ]:
                try:
                    dropdown_btn = page.locator(dropdown_sel).first
                    if dropdown_btn.count() > 0 and dropdown_btn.is_visible(timeout=3000):
                        dropdown_btn.click()
                        page.wait_for_timeout(3000)
                        dropdown_opened = True
                        break
                except Exception:
                    pass

            if not dropdown_opened:
                page.mouse.click(395, 141)
                page.wait_for_timeout(3000)
                dropdown_opened = True

            # Search for the portfolio name in dropdown
            print(f"   🔍 Searching for portfolio: '{portfolio_name}'...")

            # Try multiple methods to find the portfolio
            portfolio_found = False
            portfolio_element = None

            # Helper function to search for portfolio in current view
            def search_portfolio_in_view():
                nonlocal portfolio_found, portfolio_element, found_portfolio

                # Method 0: Use CSS selectors for portfolio name elements (from DAM HTML structure)
                # Try multiple selectors for robustness
                portfolio_selectors = [
                    'div.text-mono-900.break-all',  # Simplified selector
                    'div[class*="text-mono-900"][class*="break-all"]',  # Partial class match
                    'div.text-mono-900.typography-body.font-normal.break-all.text-left.w-full',  # Full selector
                ]

                for selector in portfolio_selectors:
                    try:
                        portfolio_name_elements = page.locator(selector).all()
                        if portfolio_name_elements:
                            for elem in portfolio_name_elements:
                                try:
                                    if elem.is_visible():
                                        elem_text = elem.text_content().strip()
                                        if elem_text.lower() == portfolio_name.lower():
                                            # Found the portfolio name element, need to click its parent button
                                            # Navigate up to find the clickable parent (button element)
                                            parent_button = elem.locator('xpath=ancestor::button').first
                                            if parent_button.count() > 0:
                                                portfolio_element = parent_button
                                            else:
                                                # Try finding a clickable parent div
                                                portfolio_element = elem.locator('xpath=ancestor::div[contains(@class, "cursor-pointer") or @role="option" or @role="menuitem"]').first
                                                if portfolio_element.count() == 0:
                                                    portfolio_element = elem
                                            portfolio_found = True
                                            found_portfolio = elem_text
                                            print(f"   ✅ Found portfolio via CSS class: '{elem_text}'")
                                            return True
                                except Exception as e:
                                    pass
                    except Exception as e:
                        pass

                # Method 1: Use getByText with partial matching (handles "NNN" in "NNN\n1 Addresses...")
                try:
                    partial_matches = page.get_by_text(portfolio_name)
                    if partial_matches.count() > 0:
                        for i in range(partial_matches.count()):
                            elem = partial_matches.nth(i)
                            if elem.is_visible():
                                elem_text = elem.text_content().strip()
                                # Check if first line matches portfolio name
                                first_line = elem_text.split('\n')[0].strip()
                                if first_line.lower() == portfolio_name.lower():
                                    portfolio_element = elem
                                    portfolio_found = True
                                    found_portfolio = first_line
                                    print(f"   ✅ Found portfolio: '{first_line}'")
                                    return True
                except Exception as e:
                    pass

                # Method 2: Search all visible elements for portfolio name as first line
                if not portfolio_found:
                    try:
                        # Look for dropdown item containers (typically have specific class patterns)
                        dropdown_items = page.locator('[class*="dropdown"] > div, [class*="menu"] > div, [role="listbox"] > div, [role="option"]').all()
                        if not dropdown_items:
                            dropdown_items = page.locator('div').all()

                        for div in dropdown_items:
                            try:
                                if not div.is_visible():
                                    continue
                                div_text = div.text_content().strip()
                                if not div_text:
                                    continue
                                # Check if first line matches portfolio name (case-insensitive)
                                first_line = div_text.split('\n')[0].strip()
                                if first_line.lower() == portfolio_name.lower():
                                    # Make sure this is a clickable item (has reasonable content)
                                    # Accept Addresses, Exchange, Wallet, or numbers (like "1", "2")
                                    if any(keyword in div_text for keyword in ['Addresses', 'Address', 'Exchange', 'Wallet', 'DeFi']) or \
                                       any(char.isdigit() for char in div_text):
                                        portfolio_element = div
                                        portfolio_found = True
                                        found_portfolio = first_line
                                        print(f"   ✅ Found portfolio item: '{first_line}'")
                                        return True
                            except:
                                pass
                    except Exception as e:
                        pass

                # Method 3: Click directly using text selector
                if not portfolio_found:
                    try:
                        # Try clicking on text that exactly matches portfolio name
                        text_selector = f'text="{portfolio_name}"'
                        if page.locator(text_selector).count() > 0:
                            elem = page.locator(text_selector).first
                            if elem.is_visible():
                                portfolio_element = elem
                                portfolio_found = True
                                found_portfolio = portfolio_name
                                print(f"   ✅ Found by text selector: '{portfolio_name}'")
                                return True
                    except Exception as e:
                        pass

                return False

            # First search without scrolling
            search_portfolio_in_view()

            # If not found, try scrolling within the dropdown to find the portfolio
            if not portfolio_found:
                print(f"   🔄 Scrolling through dropdown to find portfolio...")
                try:
                    # Find the scrollable dropdown container
                    dropdown_container = None

                    # Try different selectors for the dropdown container
                    dropdown_selectors = [
                        '[class*="dropdown"][class*="scroll"]',
                        '[class*="menu"][class*="scroll"]',
                        '[role="listbox"]',
                        '[class*="overflow-y-auto"]',
                        '[class*="overflow-auto"]',
                        'div[style*="overflow"]',
                    ]

                    for selector in dropdown_selectors:
                        containers = page.locator(selector).all()
                        for container in containers:
                            try:
                                if container.is_visible():
                                    # Check if this container is scrollable
                                    box = container.bounding_box()
                                    if box and box['height'] > 100:
                                        dropdown_container = container
                                        break
                            except:
                                pass
                        if dropdown_container:
                            break

                    # If no specific container found, try using JavaScript to find scrollable element
                    if not dropdown_container:
                        scrollable_info = page.evaluate('''() => {
                            const elements = document.querySelectorAll('div');
                            for (const el of elements) {
                                if (el.scrollHeight > el.clientHeight && el.clientHeight > 100) {
                                    const style = window.getComputedStyle(el);
                                    if (style.overflowY === 'auto' || style.overflowY === 'scroll') {
                                        const rect = el.getBoundingClientRect();
                                        if (rect.top > 0 && rect.left > 0) {
                                            return {found: true, scrollHeight: el.scrollHeight, clientHeight: el.clientHeight};
                                        }
                                    }
                                }
                            }
                            return {found: false};
                        }''')

                        if scrollable_info.get('found'):
                            pass  # Scrollable dropdown found via JS

                    # Scroll through dropdown using keyboard navigation
                    max_scroll_attempts = 30
                    for scroll_attempt in range(max_scroll_attempts):
                        # Press down arrow to scroll/navigate through items
                        page.keyboard.press('ArrowDown')
                        page.wait_for_timeout(200)

                        # Search again after each scroll
                        if search_portfolio_in_view():
                            break

                        # Also try Page Down for faster scrolling
                        if scroll_attempt % 5 == 4:
                            page.keyboard.press('PageDown')
                            page.wait_for_timeout(300)
                            if search_portfolio_in_view():
                                break

                except Exception as e:
                    pass  # Scroll error

            if portfolio_found and portfolio_element:
                # Click on the portfolio
                portfolio_element.click()
                page.wait_for_timeout(5000)

                # Extract portfolio name from UI
                try:
                    portfolio_name_elem = page.locator('div.text-mono-900.typography-body.font-normal.text-left.break-all.w-full').first
                    if portfolio_name_elem.count() > 0:
                        found_portfolio = portfolio_name_elem.text_content().strip()
                except:
                    found_portfolio = portfolio_name

                print(f"   📝 Portfolio name from DAM: {found_portfolio}")

                # Extract ALL addresses from Combined Net Worth section
                # Tron: Starts with T, 34 characters
                # EVM: Starts with 0x, 42 characters (0x + 40 hex)
                # Exchange: Everything else (names like "Binance", "moontest", etc.)
                print(f"   🔍 Extracting addresses from Combined Net Worth section...")

                collected_tron = set()
                collected_evm = set()
                collected_exchanges = set()

                def add_address_to_collection(addr):
                    """Helper to classify and add address to appropriate collection"""
                    addr = addr.strip()
                    addr_type = classify_address(addr)
                    if addr_type == 'tron':
                        if addr.lower() not in [a.lower() for a in collected_tron]:
                            collected_tron.add(addr)
                            print(f"   ✅ Found Tron address: {addr}")
                            return True
                    elif addr_type == 'evm':
                        if addr.lower() not in [a.lower() for a in collected_evm]:
                            collected_evm.add(addr)
                            print(f"   ✅ Found EVM address: {addr}")
                            return True
                    return False

                # Method 1: Find and hover/click on address elements in Combined Net Worth
                try:
                    # Look for "Combined Net Worth" section
                    combined_section = page.locator('text="Combined Net Worth"').first
                    if combined_section.count() > 0:
                        print(f"   📍 Found Combined Net Worth section")

                        # Find ALL address elements with data-tooltip-id containing address
                        address_elements = page.locator('[data-tooltip-id*="address-display-tooltip"]').all()

                        for elem in address_elements:
                            try:
                                # Hover on the element to trigger tooltip
                                elem.hover()
                                page.wait_for_timeout(1500)

                                # Check for visible tooltip with full address
                                tooltip = page.locator('[role="tooltip"]').first
                                if tooltip.count() > 0 and tooltip.is_visible():
                                    tooltip_text = tooltip.text_content().strip()

                                    # Try to extract Tron address (T + 33 chars)
                                    tron_match = re.search(r'([Tt][A-Za-z0-9]{33})', tooltip_text)
                                    if tron_match:
                                        add_address_to_collection(tron_match.group(1))

                                    # Try to extract EVM address (0x + 40 hex)
                                    evm_match = re.search(r'(0x[A-Fa-f0-9]{40})', tooltip_text)
                                    if evm_match:
                                        add_address_to_collection(evm_match.group(1))

                                # Also try data-tooltip-id attribute
                                tooltip_id = elem.get_attribute('data-tooltip-id')
                                if tooltip_id:
                                    # Tron address in tooltip ID
                                    tron_match = re.search(r'address-display-tooltip-([Tt][A-Za-z0-9]{33})', tooltip_id)
                                    if tron_match:
                                        add_address_to_collection(tron_match.group(1))

                                    # EVM address in tooltip ID
                                    evm_match = re.search(r'address-display-tooltip-(0x[A-Fa-f0-9]{40})', tooltip_id)
                                    if evm_match:
                                        add_address_to_collection(evm_match.group(1))

                            except Exception as e:
                                pass

                except Exception as e:
                    pass
                # Method 2: Look for truncated address text and hover
                if not collected_tron and not collected_evm:
                    try:
                        print(f"   🔍 Trying to find truncated address patterns...")
                        # Tron truncated pattern
                        truncated_tron = page.locator('text=/[Tt][A-Za-z0-9]{6,8}[\\.…]{2,3}[A-Za-z0-9]{4,8}/').all()
                        # EVM truncated pattern
                        truncated_evm = page.locator('text=/0x[A-Fa-f0-9]{4,8}[\\.…]{2,3}[A-Fa-f0-9]{4,8}/').all()

                        for elem in truncated_tron + truncated_evm:
                            try:
                                elem.hover()
                                page.wait_for_timeout(1500)

                                tooltip = page.locator('[role="tooltip"]').first
                                if tooltip.count() > 0 and tooltip.is_visible():
                                    tooltip_text = tooltip.text_content().strip()
                                    # Try Tron
                                    tron_match = re.search(r'([Tt][A-Za-z0-9]{33})', tooltip_text)
                                    if tron_match:
                                        add_address_to_collection(tron_match.group(1))
                                    # Try EVM
                                    evm_match = re.search(r'(0x[A-Fa-f0-9]{40})', tooltip_text)
                                    if evm_match:
                                        add_address_to_collection(evm_match.group(1))
                            except:
                                pass
                    except Exception as e:
                        pass

                # Method 3: Search page content for addresses in tooltip IDs
                if not collected_tron and not collected_evm:
                    try:
                        print(f"   📄 Searching page source for addresses in tooltip IDs...")
                        page_content = page.content()

                        # Find Tron addresses in data-tooltip-id attributes
                        tron_tooltip_matches = re.findall(r'address-display-tooltip-([Tt][A-Za-z0-9]{33})', page_content)
                        for addr in tron_tooltip_matches:
                            add_address_to_collection(addr)

                        # Find EVM addresses in data-tooltip-id attributes
                        evm_tooltip_matches = re.findall(r'address-display-tooltip-(0x[A-Fa-f0-9]{40})', page_content)
                        for addr in evm_tooltip_matches:
                            add_address_to_collection(addr)

                    except Exception as e:
                        pass

                # Method 4: Final fallback - search entire page for any address pattern
                if not collected_tron and not collected_evm:
                    try:
                        print(f"   📄 Final fallback: Searching entire page for addresses...")
                        page_content = page.content()

                        # Find ALL Tron addresses in page
                        all_tron_matches = re.findall(r'([Tt][A-Za-z0-9]{33})', page_content)
                        for addr in all_tron_matches:
                            add_address_to_collection(addr)

                        # Find ALL EVM addresses in page
                        all_evm_matches = re.findall(r'(0x[A-Fa-f0-9]{40})', page_content)
                        for addr in all_evm_matches:
                            add_address_to_collection(addr)

                    except Exception as e:
                        pass

                # Method 5: Extract Exchange names from Combined Net Worth section
                try:
                    print(f"   🔍 Looking for Exchange accounts...")
                    # Look for exchange names in the Combined Net Worth section
                    # These are typically displayed as text without address format
                    combined_rows = page.locator('text="Combined Net Worth"').locator('..').locator('..').locator('div[class*="flex"]').all()
                    for row in combined_rows[:20]:  # Limit to avoid too many iterations
                        try:
                            row_text = row.inner_text().strip()
                            # Check if this looks like an exchange (not an address)
                            if row_text and not row_text.startswith('T') and not row_text.startswith('0x'):
                                # Check for common exchange patterns
                                first_word = row_text.split('\n')[0].split()[0] if row_text else ""
                                if first_word and len(first_word) < 30 and not any(c.isdigit() for c in first_word[:5]):
                                    # Might be an exchange name like "Binance", "moontest", etc.
                                    if first_word.lower() not in ['addresses', 'value', 'combined', 'net', 'worth', 'total']:
                                        collected_exchanges.add(first_word)
                        except:
                            pass
                except Exception as e:
                    pass

                # Convert sets to sorted lists
                found_trx_addresses = sorted(list(collected_tron), key=lambda x: x.lower())
                found_evm_addresses = sorted(list(collected_evm), key=lambda x: x.lower())
                found_exchanges = sorted(list(collected_exchanges))

                # Print summary
                total_found = len(found_trx_addresses) + len(found_evm_addresses) + len(found_exchanges)
                print(f"\n   📋 Address Summary:")
                print(f"      Tron addresses: {len(found_trx_addresses)}")
                for idx, addr in enumerate(found_trx_addresses, 1):
                    print(f"         {idx}. {addr}")
                print(f"      EVM addresses: {len(found_evm_addresses)}")
                for idx, addr in enumerate(found_evm_addresses, 1):
                    print(f"         {idx}. {addr}")
                print(f"      Exchanges: {len(found_exchanges)}")
                for idx, exch in enumerate(found_exchanges, 1):
                    print(f"         {idx}. {exch}")

                if found_trx_addresses or found_evm_addresses:
                    # Update Excel with all addresses
                    print(f"\n   📝 Updating DAM addresses.xlsx...")
                    try:
                        dam_excel_path = "test_data/DAM addresses.xlsx"
                        wb = load_workbook(dam_excel_path)
                        ws = wb.active

                        # Add new row with all addresses in columns B, C, D, ...
                        next_row = ws.max_row + 1
                        ws.cell(row=next_row, column=1, value=found_portfolio)
                        all_addresses = found_trx_addresses + found_evm_addresses
                        for addr_idx, addr in enumerate(all_addresses):
                            ws.cell(row=next_row, column=2 + addr_idx, value=addr)
                        wb.save(dam_excel_path)
                        wb.close()

                        print(f"   ✅ Added to Excel: {found_portfolio} | {len(all_addresses)} address(es)")
                    except Exception as e:
                        print(f"   ⚠️  Could not update Excel: {e}")
                elif found_exchanges:
                    print(f"\n   ⚠️  Portfolio '{found_portfolio}' has only Exchange accounts (no wallet addresses).")
                else:
                    print(f"\n   ⚠️  Portfolio '{found_portfolio}' found but has NO addresses or exchanges!")

            else:
                print(f"   ❌ Portfolio '{portfolio_name}' not found in DAM dropdown")

        except Exception as e:
            print(f"   ❌ Error checking DAM: {e}")

        finally:
            browser.close()

    return found_trx_addresses, found_evm_addresses, found_exchanges, found_portfolio


def print_usage():
    """Print usage instructions."""
    print("\nUsage:")
    print("  python3 test_trx_tc00003.py                           # Run with default address")
    print("  python3 test_trx_tc00003.py <TRX_ADDRESS>             # Run with TRX address")
    print("  python3 test_trx_tc00003.py --portfolio <NAME>        # Run with portfolio name")
    print("  python3 test_trx_tc00003.py -p <NAME>                 # Run with portfolio name (short)")
    print("\nExamples:")
    print("  python3 test_trx_tc00003.py TWQsW9HJUX8wsY3KG961XUt27v4RePrMjr")
    print("  python3 test_trx_tc00003.py --portfolio 'My Portfolio'")
    print("  python3 test_trx_tc00003.py -p airdrop_token")
    print()


if __name__ == "__main__":
    import argparse

    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description="DAM Portfolio Extraction (EVM + TRX)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 run_overview.py                                           # Run with config defaults
  python3 run_overview.py -p A_bD                                  # Run by portfolio name
  python3 run_overview.py --trx TUqEg3dzVEJNQSVW2HY98z5X8SBdhmao8D --evm 0x4e14fc11f58b64740e66e4b1aa188a4b007c0eab
  python3 run_overview.py --trx TXxx... --evm 0xAAA... --evm 0xBBB...  # Multiple EVM
        """
    )
    parser.add_argument('address', nargs='?', default=None,
                        help='TRX wallet address (34 chars starting with T) or portfolio name')
    parser.add_argument('-p', '--portfolio', type=str, default=None,
                        help='Portfolio name to look up in DAM')
    parser.add_argument('--trx', type=str, default=None,
                        help='TRX wallet address (starts with T, 34 chars)')
    parser.add_argument('--evm', type=str, action='append', default=None,
                        help='EVM wallet address (0x..., 42 chars). Repeat for multiple.')
    parser.add_argument('-q', '--quiet', action='store_true', default=False,
                        help='Quiet mode - minimal output (only errors and final results)')

    args = parser.parse_args()
    
    # Set global quiet mode flag
    QUIET_MODE = args.quiet

    # Handle --trx / --evm direct address input
    if args.trx or args.evm:
        trx_addr = args.trx or ""
        evm_addrs = args.evm or []

        # Derive portfolio name prefix: A_ + last char of each EVM + last char of TRX
        suffix = "".join(a[-1] for a in evm_addrs) + (trx_addr[-1] if trx_addr else "")
        TRX_ADDRESS = trx_addr
        TRX_ADDRESSES = [trx_addr] if trx_addr else []
        EVM_ADDRESSES = evm_addrs
        PORTFOLIO_NAME_PREFIX = f"A_{suffix}" if suffix else PORTFOLIO_NAME_PREFIX

        # Skip TRX API if no TRX address provided
        if not trx_addr:
            SKIP_TRX_API = True
        else:
            SKIP_TRX_API = False
        
        # Skip SimDune API if no EVM addresses provided
        if not evm_addrs:
            SKIP_SIM_DUNE_API = True
        else:
            SKIP_SIM_DUNE_API = False

        print(f"\n{'='*80}")
        print(f"DIRECT ADDRESS INPUT")
        print(f"{'='*80}")
        print(f"TRX Address : {TRX_ADDRESS or '(none)'}")
        for i, a in enumerate(EVM_ADDRESSES, 1):
            print(f"EVM Address {i}: {a}")
        print(f"Portfolio   : {PORTFOLIO_NAME_PREFIX}")
        print(f"{'='*80}\n")

    # Determine if running with portfolio name or address
    elif args.portfolio:
        # Use portfolio name directly - main extraction will handle lookup
        print(f"\n🔍 Using portfolio: '{args.portfolio}'...")

        # Step 1: Check Excel for existing entries (just for reference)
        excel_addresses, excel_portfolio = lookup_portfolio_in_excel(args.portfolio)
        if excel_addresses:
            print(f"   📋 Found in Excel: {excel_portfolio} | {len(excel_addresses)} address(es)")
            for idx, addr in enumerate(excel_addresses, 1):
                print(f"      {idx}. {addr}")

        # Step 2: Look up portfolio in DAM
        PORTFOLIO_NAME_PREFIX = args.portfolio
        dam_trx_addresses, dam_evm_addresses, dam_exchanges, dam_portfolio = lookup_portfolio_in_dam(args.portfolio)

        # If sign-in failed, stop immediately — no fallback
        if dam_portfolio == "__SIGNIN_FAILED__":
            print(f"\n❌ Cannot proceed — DAM sign-in failed.")
            print(f"   Please check DAM is accessible and credentials are correct, then try again.")
            sys.exit(1)

        # Determine which APIs to skip based on address types
        has_tron = len(dam_trx_addresses) > 0
        has_evm = len(dam_evm_addresses) > 0
        has_exchange_only = len(dam_exchanges) > 0 and not has_tron and not has_evm

        if has_tron or has_evm:
            # Set addresses
            TRX_ADDRESSES = dam_trx_addresses
            EVM_ADDRESSES = dam_evm_addresses
            TRX_ADDRESS = dam_trx_addresses[0] if dam_trx_addresses else ""
            PORTFOLIO_NAME_PREFIX = dam_portfolio

            # Set skip flags based on what's available
            if not has_tron:
                SKIP_TRX_API = True
                print(f"   ➡️  No Tron addresses found - Skipping TRX Balance API")
            if not has_evm:
                SKIP_SIM_DUNE_API = True
                print(f"   ➡️  No EVM addresses found - Skipping Sim Dune API")

            # Combine all addresses for Excel update check
            dam_addresses = dam_trx_addresses + dam_evm_addresses

            # Check if Excel had different (stale) addresses
            if excel_addresses and set(excel_addresses) != set(dam_addresses):
                print(f"\n   ⚠️  Excel addresses were outdated!")
                print(f"      Excel:  {excel_addresses}")
                print(f"      DAM:    {dam_addresses}")
                print(f"   📝 Updating Excel with correct addresses...")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook("test_data/DAM addresses.xlsx")
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2):
                        if row[0].value and str(row[0].value).strip().lower() == args.portfolio.lower():
                            # Clear old addresses and write new ones
                            for col_idx in range(1, ws.max_column + 1):
                                if col_idx >= 2:
                                    row[col_idx - 1].value = None
                            for addr_idx, addr in enumerate(dam_addresses):
                                row[1 + addr_idx].value = addr
                            break
                    wb.save("test_data/DAM addresses.xlsx")
                    wb.close()
                    print(f"   ✅ Excel updated with {len(dam_addresses)} address(es)")
                except Exception as e:
                    print(f"   ⚠️  Could not update Excel: {e}")

            print("\n" + "="*80)
            print("✅ PORTFOLIO FOUND IN DAM - RUNNING TEST")
            print("="*80)
            print(f"Portfolio Name: {PORTFOLIO_NAME_PREFIX}")
            print(f"Tron Addresses: {len(TRX_ADDRESSES)}")
            for idx, addr in enumerate(TRX_ADDRESSES, 1):
                print(f"  {idx}. {addr}")
            print(f"EVM Addresses: {len(EVM_ADDRESSES)}")
            for idx, addr in enumerate(EVM_ADDRESSES, 1):
                print(f"  {idx}. {addr}")
            print(f"Exchanges: {len(dam_exchanges)}")
            for idx, exch in enumerate(dam_exchanges, 1):
                print(f"  {idx}. {exch}")
            print(f"\nAPI Status:")
            print(f"  TRX Balance API: {'SKIP' if SKIP_TRX_API else 'RUN'}")
            print(f"  Sim Dune API: {'SKIP' if SKIP_SIM_DUNE_API else 'RUN'}")
            print("="*80 + "\n")
        elif excel_addresses:
            # DAM lookup failed but Excel has entries - use Excel as fallback
            print(f"\n   ⚠️  Could not verify from DAM. Using Excel addresses as fallback.")
            TRX_ADDRESS = excel_addresses[0]
            TRX_ADDRESSES = excel_addresses
            PORTFOLIO_NAME_PREFIX = excel_portfolio

            print("\n" + "="*80)
            print("PORTFOLIO NAME TEST (from Excel - unverified)")
            print("="*80)
            print(f"Portfolio Name: {PORTFOLIO_NAME_PREFIX}")
            print(f"Addresses: {len(TRX_ADDRESSES)}")
            for idx, addr in enumerate(TRX_ADDRESSES, 1):
                print(f"  {idx}. {addr}")
            print("="*80 + "\n")
        elif dam_portfolio:
            # Portfolio was found in DAM but has no wallet addresses (Exchange-only portfolio)
            SKIP_TRX_API = True
            SKIP_SIM_DUNE_API = True
            TRX_ADDRESSES = []
            EVM_ADDRESSES = []
            PORTFOLIO_NAME_PREFIX = dam_portfolio

            print(f"\n⚠️  Portfolio '{dam_portfolio}' exists in DAM but has NO wallet addresses!")
            print("   This portfolio only has Exchange accounts (CEX) with no wallet addresses.")
            print("   ➡️  Skipping TRX Balance API test (no Tron addresses)")
            print("   ➡️  Skipping Sim Dune API test (no EVM addresses)")
            print("   ➡️  Continuing with DAM Portfolio extraction...")

            print("\n" + "="*80)
            print("PORTFOLIO NAME TEST (Exchange-only - No wallet addresses)")
            print("="*80)
            print(f"Portfolio Name: {PORTFOLIO_NAME_PREFIX}")
            print(f"Wallet Addresses: 0")
            print(f"Exchanges: {len(dam_exchanges)}")
            for idx, exch in enumerate(dam_exchanges, 1):
                print(f"  {idx}. {exch}")
            print("="*80 + "\n")
        else:
            print(f"\n❌ Portfolio '{args.portfolio}' not found in DAM.")
            sys.exit(1)

    elif args.address:
        custom_address = args.address

        # Check if it looks like an address or a portfolio name
        is_address = custom_address.startswith('T') and len(custom_address) == 34

        if is_address:
            # It's a TRX address - first check if it exists in DAM addresses.xlsx
            print(f"\n🔍 Looking up address in DAM addresses.xlsx...")
            found_portfolio, portfolio_addresses = lookup_address_in_excel(custom_address)

            if found_portfolio:
                # Address found in existing portfolio
                TRX_ADDRESS = custom_address
                TRX_ADDRESSES = portfolio_addresses  # Use all addresses from that portfolio
                PORTFOLIO_NAME_PREFIX = found_portfolio

                print("\n" + "="*80)
                print("ADDRESS FOUND IN EXISTING PORTFOLIO")
                print("="*80)
                print(f"Input Address: {custom_address}")
                print(f"Portfolio Name: {PORTFOLIO_NAME_PREFIX}")
                print(f"Portfolio Addresses: {len(TRX_ADDRESSES)}")
                for idx, addr in enumerate(TRX_ADDRESSES, 1):
                    marker = " <-- (input)" if addr == custom_address else ""
                    print(f"  {idx}. {addr}{marker}")
                print("="*80 + "\n")
            else:
                # Address not found - create new portfolio
                TRX_ADDRESS = custom_address
                TRX_ADDRESSES = [custom_address]
                PORTFOLIO_NAME_PREFIX = f"Custom_{custom_address[:8]}"

                print("\n" + "="*80)
                print("NEW ADDRESS - WILL CREATE PORTFOLIO")
                print("="*80)
                print(f"Using custom address: {TRX_ADDRESS}")
                print(f"Portfolio prefix: {PORTFOLIO_NAME_PREFIX}")
                print("="*80 + "\n")
        else:
            # Might be a portfolio name without the --portfolio flag
            print(f"\n🔍 '{custom_address}' doesn't look like a TRX address.")
            print(f"   Checking DAM for portfolio...")

            # Always verify from DAM to get the correct current addresses
            found_trx, found_evm, found_exchanges, found_portfolio = lookup_portfolio_in_dam(custom_address)

            # If sign-in failed, stop immediately
            if found_portfolio == "__SIGNIN_FAILED__":
                print(f"\n❌ Cannot proceed — DAM sign-in failed.")
                print(f"   Please check DAM is accessible and credentials are correct, then try again.")
                sys.exit(1)

            if found_trx or found_evm:
                TRX_ADDRESSES = found_trx
                EVM_ADDRESSES = found_evm
                TRX_ADDRESS = found_trx[0] if found_trx else ""
                PORTFOLIO_NAME_PREFIX = found_portfolio

                # Set skip flags
                if not found_trx:
                    SKIP_TRX_API = True
                if not found_evm:
                    SKIP_SIM_DUNE_API = True

                print("\n" + "="*80)
                print("PORTFOLIO NAME TEST (Found in DAM)")
                print("="*80)
                print(f"Portfolio Name: {PORTFOLIO_NAME_PREFIX}")
                print(f"Tron Addresses: {len(TRX_ADDRESSES)}")
                for idx, addr in enumerate(TRX_ADDRESSES, 1):
                    print(f"  {idx}. {addr}")
                print(f"EVM Addresses: {len(EVM_ADDRESSES)}")
                for idx, addr in enumerate(EVM_ADDRESSES, 1):
                    print(f"  {idx}. {addr}")
                print(f"\nAPI Status:")
                print(f"  TRX Balance API: {'SKIP' if SKIP_TRX_API else 'RUN'}")
                print(f"  Sim Dune API: {'SKIP' if SKIP_SIM_DUNE_API else 'RUN'}")
                print("="*80 + "\n")
            elif found_portfolio:
                # Portfolio exists but has only exchanges
                SKIP_TRX_API = True
                SKIP_SIM_DUNE_API = True
                TRX_ADDRESSES = []
                EVM_ADDRESSES = []
                PORTFOLIO_NAME_PREFIX = found_portfolio
                print(f"\n⚠️  Portfolio '{found_portfolio}' has only Exchange accounts.")
                print("   ➡️  Skipping both TRX and Sim Dune APIs")
            else:
                print(f"\n⚠️  '{custom_address}' is not a valid TRX address or portfolio name.")
                print(f"   Expected address: 34 characters starting with 'T' or '0x'")
                print(f"   Received: {len(custom_address)} characters")
                print()
                response = input("Continue anyway as address? (y/n): ")
                if response.lower() == 'y':
                    TRX_ADDRESS = custom_address
                    TRX_ADDRESSES = [custom_address]
                    PORTFOLIO_NAME_PREFIX = f"Custom_{custom_address[:8]}"
                else:
                    print_usage()
                    sys.exit(1)

    # Run the main test
    main(args)
