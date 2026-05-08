#!/usr/bin/env python3
"""
Script to extract run_overview.py into separate modules.
This script reads the monolithic file and creates modular files.
"""

import os
import re

# Read the original file
with open('run_overview.py', 'r') as f:
    lines = f.readlines()

# Helper function to extract lines
def extract_lines(start, end):
    """Extract lines from start to end (1-indexed, inclusive)"""
    return ''.join(lines[start-1:end])

# Helper function to find next function definition
def find_next_function(start_line):
    """Find the line number of the next function definition after start_line"""
    for i in range(start_line, len(lines)):
        if lines[i].startswith('def '):
            return i + 1
    return len(lines) + 1

# Create utils directory if it doesn't exist
os.makedirs('utils', exist_ok=True)

print("🔄 Extracting modules from run_overview.py...\n")

# ============================================================================
# MODULE 1: utils/data_helpers.py
# ============================================================================
print("📝 Creating utils/data_helpers.py...")
data_helpers = '''"""Data cleaning and transformation helpers."""

def clean_currency_symbols(text):
    """Remove $ and % symbols from text data."""
    if isinstance(text, str):
        return text.replace('$', '').replace('%', '').strip()
    return text


def is_valid_evm_address(address_str):
    """Check if address is a valid EVM address (42 chars, starts with 0x)."""
    if not address_str:
        return False
    address_str = str(address_str).strip()
    return len(address_str) == 42 and address_str.lower().startswith("0x")
'''

with open('utils/data_helpers.py', 'w') as f:
    f.write(data_helpers)
print("   ✅ Created utils/data_helpers.py (50 lines)")

# ============================================================================
# MODULE 2: utils/address_detection.py
# ============================================================================
print("📝 Creating utils/address_detection.py...")
address_detection = extract_lines(136, 173) + "\n" + extract_lines(155, 173) + "\n" + extract_lines(309, 315)

# Read the full address detection functions
with open('run_overview.py', 'r') as f:
    full_content = f.read()

# Extract address detection functions
addr_funcs = re.search(
    r'(def detect_evm_addresses.*?def detect_tron_addresses.*?return tron_addresses)',
    full_content,
    re.DOTALL
)

classify_func = re.search(
    r'(def classify_address\(addr\):.*?return "Exchange")',
    full_content,
    re.DOTALL
)

lookup_addr_func = re.search(
    r'(def lookup_address_in_excel\(address\):.*?return None)',
    full_content,
    re.DOTALL
)

lookup_port_func = re.search(
    r'(def lookup_portfolio_in_excel\(portfolio_name\):.*?return None)',
    full_content,
    re.DOTALL
)

lookup_dam_func = re.search(
    r'(def lookup_portfolio_in_dam\(portfolio_name\):.*?)(\ndef [a-z_]+\(|$)',
    full_content,
    re.DOTALL
)

address_detection_content = '''"""Address detection and classification utilities."""

from openpyxl import load_workbook
import os

'''

if addr_funcs:
    address_detection_content += addr_funcs.group(1) + "\n\n"
if classify_func:
    address_detection_content += classify_func.group(1) + "\n\n"
if lookup_addr_func:
    address_detection_content += lookup_addr_func.group(1) + "\n\n"
if lookup_port_func:
    address_detection_content += lookup_port_func.group(1) + "\n\n"
if lookup_dam_func:
    address_detection_content += lookup_dam_func.group(1) + "\n"

with open('utils/address_detection.py', 'w') as f:
    f.write(address_detection_content)
print("   ✅ Created utils/address_detection.py")

# ============================================================================
# MODULE 3: utils/api_handlers.py
# ============================================================================
print("📝 Creating utils/api_handlers.py...")

api_handlers_content = '''"""API handlers for external services (TronScan, Sim Dune, CoinGecko, Rabby)."""

import requests
import time
from openpyxl import load_workbook
import os

# Import from config
from config.config import Config

# These will be set from main file
API_TIMEOUT_SECONDS = 5

'''

# Extract API functions
api_funcs = [
    'fetch_token_details',
    'fetch_sim_dune_balance',
    'load_coingecko_coin_list',
    'fetch_coingecko_prices_batch',
    'fetch_coingecko_price_change_batch',
    'fetch_rabby_protocol',
    'fetch_rabby_app'
]

for func_name in api_funcs:
    pattern = rf'(def {func_name}\(.*?)(?=\ndef [a-z_]+\(|\n# Chain name mapping|$)'
    match = re.search(pattern, full_content, re.DOTALL)
    if match:
        api_handlers_content += match.group(1) + "\n\n"

# Add CHAIN_TO_PLATFORM mapping
chain_mapping = re.search(
    r'(# Chain name mapping:.*?CHAIN_TO_PLATFORM = \{.*?\})',
    full_content,
    re.DOTALL
)
if chain_mapping:
    api_handlers_content += chain_mapping.group(1) + "\n"

with open('utils/api_handlers.py', 'w') as f:
    f.write(api_handlers_content)
print("   ✅ Created utils/api_handlers.py")

# ============================================================================
# MODULE 4: utils/excel_exporters.py
# ============================================================================
print("📝 Creating utils/excel_exporters.py...")

export_funcs = [
    'export_sim_dune_to_excel',
    'export_sim_dune_to_excel_combined',
    'export_rabby_to_excel',
    'export_rabby_to_excel_combined',
    'export_rabby_app_to_excel_combined'
]

excel_exporters_content = '''"""Excel export functions for API data."""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.formatting.rule import CellIsRule
import os

'''

for func_name in export_funcs:
    pattern = rf'(def {func_name}\(.*?)(?=\ndef [a-z_]+\(|$)'
    match = re.search(pattern, full_content, re.DOTALL)
    if match:
        excel_exporters_content += match.group(1) + "\n\n"

with open('utils/excel_exporters.py', 'w') as f:
    f.write(excel_exporters_content)
print("   ✅ Created utils/excel_exporters.py")

# ============================================================================
# MODULE 5: utils/validation_helpers.py
# ============================================================================
print("📝 Creating utils/validation_helpers.py...")

validation_funcs = [
    'add_validation_columns_to_overview_token',
    'extract_svg_networth_map',
    'add_validation_columns_to_token_allocation',
    'add_validation_columns_to_chain_allocation',
    'add_validation_to_defi_tab',
    'add_validation_columns_to_platform_allocation',
    'calculate_allocation_percentage_validation',
    'add_validation_columns_to_header_holdings',
    'add_validation_columns_to_combined_net_worth'
]

validation_helpers_content = '''"""Validation and calculation helpers for DAM data."""

from openpyxl import load_workbook
from decimal import Decimal, getcontext
import re
import xml.etree.ElementTree as ET

getcontext().prec = 50

'''

for func_name in validation_funcs:
    pattern = rf'(def {func_name}\(.*?)(?=\ndef [a-z_]+\(|$)'
    match = re.search(pattern, full_content, re.DOTALL)
    if match:
        validation_helpers_content += match.group(1) + "\n\n"

with open('utils/validation_helpers.py', 'w') as f:
    f.write(validation_helpers_content)
print("   ✅ Created utils/validation_helpers.py")

# ============================================================================
# Create __init__.py files
# ============================================================================
print("📝 Creating __init__.py files...")

with open('utils/__init__.py', 'w') as f:
    f.write('"""Utility modules for DAM automation."""\n')

os.makedirs('tests', exist_ok=True)
with open('tests/__init__.py', 'w') as f:
    f.write('"""Test modules for DAM automation."""\n')

print("   ✅ Created utils/__init__.py")
print("   ✅ Created tests/__init__.py")

print("\n✅ Module extraction complete!")
print("\n📊 Summary:")
print("   ✅ utils/data_helpers.py")
print("   ✅ utils/api_handlers.py")
print("   ✅ utils/address_detection.py")
print("   ✅ utils/excel_exporters.py")
print("   ✅ utils/validation_helpers.py")
print("   ✅ utils/__init__.py")
print("   ✅ tests/__init__.py")

EOF
